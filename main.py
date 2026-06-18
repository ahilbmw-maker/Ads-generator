import os
import json
import re
import asyncio
import shutil
import uuid
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from urllib.parse import urlparse
from pathlib import Path

import httpx
try:
    import openpyxl
except ImportError:
    openpyxl = None
from fastapi import FastAPI, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import anthropic

app = FastAPI()


@app.get("/healthz")
async def healthz():
    """Health check + FFmpeg readiness probe."""
    import subprocess
    status = {"ok": True}
    try:
        result = subprocess.run(["ffmpeg", "-version"], capture_output=True, timeout=5)
        status["ffmpeg"] = "ready" if result.returncode == 0 else "error"
    except Exception as e:
        status["ffmpeg"] = f"missing: {e}"
        status["ok"] = False
    return status
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.mount("/static", StaticFiles(directory="static"), name="static")
client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

# ═══ DOSTOPNI GATE (eno skupno geslo) ═══════════════════════════════════════
# Geslo se nastavi prek okoljske spr. APP_PASSWORD na Renderju.
# Piškotek je podpisan s HMAC (skrivnost APP_SECRET ali fallback) — ne da se ponarediti.
import hmac as _hmac, hashlib as _hashlib, time as _time, base64 as _b64
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import RedirectResponse, PlainTextResponse

APP_PASSWORD = os.environ.get("APP_PASSWORD", "siluxar2026")  # SPREMENI prek env na Renderju!
APP_SECRET = os.environ.get("APP_SECRET", APP_PASSWORD + "_slx_sign_v1")
AUTH_COOKIE = "slx_auth"
AUTH_TTL = 60 * 60 * 24 * 30  # seja velja 30 dni

# Poti, ki so DOSTOPNE BREZ prijave
_AUTH_EXEMPT_EXACT = {"/login", "/healthz", "/favicon.ico"}
_AUTH_EXEMPT_PREFIX = ("/static/", "/regen-img/")

def _auth_make_token():
    exp = str(int(_time.time()) + AUTH_TTL)
    sig = _hmac.new(APP_SECRET.encode(), exp.encode(), _hashlib.sha256).hexdigest()
    raw = f"{exp}:{sig}"
    return _b64.urlsafe_b64encode(raw.encode()).decode()

def _auth_check_token(token: str) -> bool:
    try:
        raw = _b64.urlsafe_b64decode(token.encode()).decode()
        exp_str, sig = raw.split(":", 1)
        expected = _hmac.new(APP_SECRET.encode(), exp_str.encode(), _hashlib.sha256).hexdigest()
        if not _hmac.compare_digest(sig, expected):
            return False
        return int(exp_str) > int(_time.time())
    except Exception:
        return False

class AuthGateMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        path = request.url.path
        if path in _AUTH_EXEMPT_EXACT or path.startswith(_AUTH_EXEMPT_PREFIX):
            return await call_next(request)
        token = request.cookies.get(AUTH_COOKIE, "")
        if _auth_check_token(token):
            return await call_next(request)
        # ni prijavljen
        accept = request.headers.get("accept", "")
        if "text/html" in accept:
            return RedirectResponse(url="/login", status_code=302)
        return PlainTextResponse("401 — prijava potrebna", status_code=401)

app.add_middleware(AuthGateMiddleware)

_LOGIN_HTML = """<!DOCTYPE html><html lang="sl"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Prijava — Suban AI</title>
<link rel="icon" type="image/png" href="/static/suban-s-icon.png">
<style>
  * { box-sizing: border-box; }
  body { margin:0; font-family:'DM Sans',-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
    background:#f4f5f7; color:#1e293b; display:flex; align-items:center; justify-content:center; min-height:100vh; padding:20px; }
  .box { background:#fff; border:1px solid #e2e8f0; border-radius:16px; padding:36px 30px; width:360px; max-width:92vw;
    box-shadow:0 10px 40px rgba(0,0,0,0.08); text-align:center; }
  .logo { font-size:34px; margin-bottom:8px; }
  h1 { font-size:20px; font-weight:700; margin:0 0 4px; }
  p.sub { font-size:13px; color:#64748b; margin:0 0 24px; }
  input { width:100%; padding:12px 14px; border:1px solid #e2e8f0; border-radius:10px; font-size:15px;
    font-family:inherit; margin-bottom:12px; }
  input:focus { outline:none; border-color:#3a6fff; }
  button { width:100%; padding:12px; background:#3a6fff; color:#fff; border:none; border-radius:10px;
    font-size:15px; font-weight:700; cursor:pointer; font-family:inherit; }
  .err { color:#dc2626; font-size:13px; min-height:18px; margin-top:10px; }
</style></head><body>
  <div class="box">
    <img src="/static/suban-logo.png" alt="Suban AI" style="width:160px;height:auto;margin:0 auto 8px;display:block">
    <p class="sub">Vpiši geslo za dostop</p>
    <form method="POST" action="/login">
      <input type="password" name="password" placeholder="Geslo" autofocus autocomplete="current-password">
      <button type="submit">Vstopi</button>
    </form>
    <div class="err">__ERR__</div>
  </div>
</body></html>"""

@app.get("/login", response_class=HTMLResponse)
async def login_page(err: str = ""):
    msg = "Napačno geslo." if err else ""
    return HTMLResponse(_LOGIN_HTML.replace("__ERR__", msg))

@app.post("/login")
async def login_submit(password: str = Form("")):
    if _hmac.compare_digest(password, APP_PASSWORD):
        resp = RedirectResponse(url="/", status_code=302)
        resp.set_cookie(AUTH_COOKIE, _auth_make_token(), max_age=AUTH_TTL,
                        httponly=True, samesite="lax")
        return resp
    return RedirectResponse(url="/login?err=1", status_code=302)

@app.get("/logout")
async def logout():
    resp = RedirectResponse(url="/login", status_code=302)
    resp.delete_cookie(AUTH_COOKIE)
    return resp
# ════════════════════════════════════════════════════════════════════════════

# Maksimum sočasnih FFmpeg procesov — Render Pro 2 CPU/4GB ne zdrži več kot 1 hkrati
# brez health check timeoutov. (1 CPU za FFmpeg, 1 CPU za /healthz + uvicorn)
FFMPEG_SEMAPHORE = asyncio.Semaphore(1)

TEMPLATE_PATH = "static/tiktok_template.xlsx"
EXPORTS_DIR = Path("exports")
EXPORTS_DIR.mkdir(exist_ok=True)
DATA_DIR = Path(os.environ.get("DATA_DIR", "/data"))
DATA_DIR.mkdir(exist_ok=True, parents=True)
TT_HISTORY_FILE = DATA_DIR / "tiktok_history.json"
META_HISTORY_FILE = DATA_DIR / "meta_history.json"
KREATIVE_HISTORY_FILE = DATA_DIR / "kreative_history.json"
FORECAST_ENTRIES_FILE = DATA_DIR / "forecast_entries.json"
FORECAST_DELETED_FILE = DATA_DIR / "forecast_deleted.json"
FORECAST_HISTORY_FILE = DATA_DIR / "forecast_history.json"
RVC_FILE = DATA_DIR / "rvc_maaarket.json"  # RVC Maaarket: zadnji vnos (prepiše se ob novem)
SPOROCANJE_FILE = DATA_DIR / "sporocanje_common.json"
SCALING_STATE_FILE = DATA_DIR / "scaling_state.json"  # Scaling Recommender: shranjeni rezultati + scale zgodovina

# ─── ECONT GEO (cities / streets / quarters lookup) ──────────────────────────
def _load_econt_geo():
    for p in [DATA_DIR / "econt_geo.json", Path("econt_geo.json"), Path("static/econt_geo.json")]:
        if p.exists():
            try:
                with open(p, encoding="utf-8") as f:
                    data = json.load(f)
                print(f"[econt_geo] Loaded from {p} — {len(data.get('city_by_id',{}))} cities, {len(data.get('streets_by_city',{}))} cities w/ streets")
                return data
            except Exception as e:
                print(f"[econt_geo] Failed to load {p}: {e}")
    print("[econt_geo] NOT FOUND — street validation disabled")
    return {}

ECONT_GEO = _load_econt_geo()


def econt_lookup_city(zip_code: str, city_name: str) -> dict | None:
    """Vrni city entry (id, name_bg, name_en, zip) glede na ZIP ali ime mesta."""
    if not ECONT_GEO:
        return None
    zip_map = ECONT_GEO.get("zip_to_city_id", {})
    city_map = ECONT_GEO.get("city_by_id", {})
    name_map = ECONT_GEO.get("name_to_city_ids", {})
    city_key = city_name.lower().strip()

    # 1. Poskusi ZIP + ime skupaj — poišči city ki ima ta ZIP IN ime se ujema
    if zip_code and city_name:
        # Najdi vse city ki imajo ta ZIP
        candidates = []
        for cid, cdata in city_map.items():
            if cdata.get("zip") == str(zip_code).strip():
                candidates.append((cid, cdata))
        if candidates:
            # Med kandidati poišči najboljše ujemanje po imenu
            for cid, cdata in candidates:
                en = (cdata.get("name_en") or "").lower()
                bg = (cdata.get("name_bg") or "").lower()
                if city_key == en or city_key == bg:
                    return {"id": int(cid), **cdata}
            # Delno ujemanje
            for cid, cdata in candidates:
                en = (cdata.get("name_en") or "").lower()
                bg = (cdata.get("name_bg") or "").lower()
                if city_key in en or en in city_key or city_key in bg or bg in city_key:
                    return {"id": int(cid), **cdata}
            # ZIP ustreza ampak ime ne — vrni prvega (fallback)
            cid, cdata = candidates[0]
            return {"id": int(cid), **cdata}

    # 2. Samo ZIP lookup
    if zip_code:
        cid = zip_map.get(str(zip_code).strip())
        if cid:
            return {"id": cid, **city_map.get(str(cid), {})}

    # 3. Samo ime lookup (lowercase, točno)
    cids = name_map.get(city_key)
    if cids:
        cid = cids[0]
        return {"id": cid, **city_map.get(str(cid), {})}

    # 4. Delno ujemanje po imenu
    for k, ids in name_map.items():
        if city_key and (city_key in k or k in city_key):
            cid = ids[0]
            return {"id": cid, **city_map.get(str(cid), {})}
    return None


def econt_get_streets_context(city_id, street_query: str, max_results: int = 12) -> str:
    """Vrni seznam ulic iz mesta kot kontekst za AI — fuzzy match na street_query."""
    if not ECONT_GEO or not city_id:
        return ""
    streets = ECONT_GEO.get("streets_by_city", {}).get(str(city_id), [])
    quarters = ECONT_GEO.get("quarters_by_city", {}).get(str(city_id), [])
    if not streets and not quarters:
        return ""
    # Fuzzy match — poišči ulice ki vsebujejo ključne besede iz poizvedbe
    q = re.sub(r'(ul\.|bul\.|zh\.k\.|kv\.|bl\.|vh\.|et\.|ap\.|\d+)', '', street_query.lower()).strip()
    words = [w for w in q.split() if len(w) > 2]
    scored = []
    for s in streets:
        en = s[1].lower() if s[1] else ""
        bg = s[0].lower() if s[0] else ""
        score = sum(1 for w in words if w in en or w in bg)
        if score > 0:
            scored.append((score, s))
    scored.sort(key=lambda x: -x[0])
    top_streets = [s for _, s in scored[:max_results]]
    # Če ni zadetkov — vrni prvih N ulic (za osnovno validacijo)
    if not top_streets:
        top_streets = streets[:8]
    lines = []
    if quarters:
        q_sample = quarters[:6]
        lines.append("Četrti: " + ", ".join(f"{q[1]}" for q in q_sample if q[1]))
    lines.append("Ulice (vzorec):")
    for s in top_streets:
        lines.append(f"  {s[1]} / {s[0]}")
    return "\n".join(lines)

# ─── BRAND DOMAIN MAPS ───────────────────────────────────────────────────────

BRAND_DOMAINS = {
    "maaarket": {"sl":"www.maaarket.si","hr":"www.maaarket.hr","rs":"www.maaarket.rs","hu":"www.maaarket.hu","cz":"www.maaarket.cz","sk":"www.maaarket.sk","pl":"www.maaarket.pl","gr":"www.maaarket.gr","ro":"www.maaarket.ro","bg":"www.maaarket.bg"},
    "fluxigo":  {"sl":"www.fluxigo.si","hr":"www.fluxigo.hr","rs":"www.fluxigo.rs","hu":"www.fluxigo.hu","cz":"www.fluxigo.cz","sk":"www.fluxigo.sk","pl":"www.fluxigo.pl","gr":"www.fluxigo.gr","ro":"www.fluxigo.ro","bg":"www.fluxigo.bg"},
    "easyzo":   {"sl":"www.easyzo.si","hr":"www.easyzo.hr","rs":"www.easyzo.rs","hu":"www.easyzo.hu","cz":"www.easyzo.cz","sk":"www.easyzo.sk","pl":"www.easyzo.pl","gr":"www.easyzo.gr","ro":"www.easyzo.ro","bg":"www.easyzo.bg"},
    "zipply":   {"sl":"www.zipply.si","hr":"www.zipply.hr","rs":"www.zipply.rs","hu":"www.zipply.hu","cz":"www.zipply.cz","sk":"www.zipply.sk","pl":"www.zipply.pl","gr":"www.zipply.gr","ro":"www.zipply.ro","bg":"www.zipply.bg"},
    "thundershop": {"sl":"www.thundershop.si","hr":"www.thundershop.hr","rs":"www.thundershop.rs","hu":"www.thundershop.hu","cz":"www.thundershop.cz","sk":"www.thundershop.sk","gr":"www.thundershop.gr","ro":"www.thundershop.ro","bg":"www.thundershop.bg"},
    "colibrishop": {"sl":"www.colibrishop.si","hr":"www.colibrishop.hr","rs":"www.colibrishop.rs","cz":"www.colibrishop.cz","sk":"www.colibrishop.sk","gr":"www.colibrishop.gr","ro":"www.colibrishop.ro","bg":"www.colibrishop.bg"},
}

MAAARKET_FEEDS = {
    "sl":"https://api.maaarket.si/storage/exports/sl/google.xml",
    "hr":"https://api.maaarket.hr/storage/exports/hr/google.xml",
    "rs":"https://api.maaarket.rs/storage/exports/sr/google.xml",
    "hu":"https://api.maaarket.hu/storage/exports/hu/google.xml",
    "pl":"https://api.maaarket.pl/storage/exports/pl/google.xml",
    "cz":"https://api.maaarket.cz/storage/exports/cs/google.xml",
    "sk":"https://api.maaarket.sk/storage/exports/sk/google.xml",
    "gr":"https://api.maaarket.gr/storage/exports/el/google.xml",
    "bg":"https://api.maaarket.bg/storage/exports/bg/google.xml",
    "ro":"https://api.maaarket.ro/storage/exports/ro/google.xml",
}

G = "http://base.google.com/ns/1.0"
feed_by_lang: dict = {}
slug_to_id: dict = {}
sl_image_index: dict = {}   # SLO slike: { "mpn_upper": img, "slug_lower": img, "title_lower": img } za hitri lookup
last_fetch: Optional[datetime] = None
CACHE_TTL_HOURS = 168  # 7 dni — slike/SKU se redko spreminjajo, ni potrebe po pogostem osveževanju
_feed_lock: Optional["asyncio.Lock"] = None  # prepreči sočasne prenose feeda


def _get_feed_lock() -> "asyncio.Lock":
    global _feed_lock
    if _feed_lock is None:
        _feed_lock = asyncio.Lock()
    return _feed_lock


def is_cache_stale():
    return last_fetch is None or datetime.now() - last_fetch > timedelta(hours=CACHE_TTL_HOURS)


def extract_slug(url: str) -> Optional[str]:
    path = urlparse(url).path.rstrip('/')
    parts = [p for p in path.split('/') if p]
    return parts[-1].lower() if parts else None


def _extract_brand_sku(brand: str, image_url: str) -> Optional[str]:
    """Za Ikonka/Amio izlušči zanesljiv SKU iz image_link URL-ja.
    Ikonka: '...cache/ikonka{SKU}{ostalo}-{hash}.jpeg' → SKU takoj za 'ikonka'
    Amio:   '...cache/amio{XXXX}-{SKU}-{ostalo}-{hash}.jpeg' → SKU za prvim '-'
    Vrne SKU (upper) ali None."""
    import re as _re
    if not brand or not image_url:
        return None
    b = brand.strip().lower()
    # zajemi del datoteke za '/cache/'
    m = _re.search(r'/cache/(.+)$', image_url)
    fname = (m.group(1) if m else image_url).lower()

    if b == "ikonka":
        # ikonka{SKU}... — SKU je alfanumerični niz takoj za 'ikonka' do prvega '-' ali do hash
        mm = _re.search(r'ikonka([a-z]{0,3}\d[a-z0-9]*)', fname)
        if mm:
            raw = mm.group(1)
            # SKU je tipično tip 'kx7975' (črke+številke); odreži morebitno podvojitev/hash rep
            # vzemi vzorec črk + številk na začetku
            m2 = _re.match(r'([a-z]+\d+)', raw)
            return (m2.group(1) if m2 else raw).upper()

    if b == "amio":
        # amio{nekaj}-{SKU}-... → SKU za prvim '-'
        mm = _re.search(r'amio[^-]*-([a-z0-9]+)-', fname)
        if mm:
            return mm.group(1).upper()

    return None


def _sl_slugify(s: str) -> str:
    """Pretvori naziv v slug enako kot Maaarket (č→c, š→s, ž→z, presledki/ločila→vezaji)."""
    import re as _re
    if not s:
        return ""
    s = s.lower()
    for a, b in (('č','c'),('š','s'),('ž','z'),('ć','c'),('đ','d'),('á','a'),('é','e'),('í','i'),('ó','o'),('ú','u')):
        s = s.replace(a, b)
    s = _re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s


_FEED_CATEGORIES = [
    "a-mobile", "adapterji-za-zarnice", "akumulator", "avto-antene",
    "avto-moto", "avto-ventilatorji", "bluetooth-zvocniki", "centralno-zaklepanje",
    "cevi-in-crpalke", "ciscenje-sesanje-in-likanje", "ciscenje-vozila", "dekorativni-program",
    "dnevne-luci", "dodatki-za-avto", "dodatki-za-mobilni-telefon", "dodatki-za-tv-in-racunalnik",
    "dom-in-vrt", "druge-igrace", "drzala-telefona", "dusilci",
    "dvigalke", "figure-zivali", "fm-oddajniki", "gospodinjski-aparati",
    "gospodinjstvo", "hisni-tekstil", "hlace", "hlajenje-in-ogrevanje",
    "igrace", "igre-na-prostem", "jakne-brezrokavniki", "kampiranje-piknik-plaza",
    "kolesarjenje", "kompleti", "kopalke", "kostumi-in-dodatki",
    "kozmetika", "kozmetika-in-nega", "kreativne-igrace", "krila",
    "kuhinjski-pribor", "licenje", "moda-za-nosecnice", "modni-dodatki",
    "napihljive-blazine-in-obroci", "nega-las", "nega-obraza", "nega-telesa",
    "oblacila", "oblacila-in-obutev", "obleke", "obutev", "obutev-in-dodatki", "okrasni-pokrovi-koles",
    "opozorilne-luci", "organizatorji", "organizatorji-za-avto", "orodja-in-aparati",
    "osebna-nega", "ostalo", "osvezilci-zraka-za-avto", "otroska-kolesa",
    "otroska-moda", "pajkice", "pametna-hisa", "pametne-ure",
    "parkirni-senzorji", "plisaste-igrace", "pohistvo", "pohodnistvo",
    "pokrivala-za-avto", "polnilci-za-akumulatorje", "polnjenje-naprav", "popravilo-pnevmatike",
    "posode-za-hrano-in-vodo", "potovanja", "poucne-igrace", "povodci",
    "prestavne-rocice", "prevleke-sedezev-in-otroski-sedezi", "prevleke-volana", "pripomocki-za-kopalnico",
    "pripomocki-za-kuhinjo", "products", "racunalnistvo-in-telefonija", "rc-modeli",
    "rocne-ure", "rocno-orodje", "slike", "slusalke",
    "spodnje-perilo", "sport", "sport-in-prosti-cas", "sportna-oprema",
    "sportna-vadba", "sprostitev-in-dobro-pocutje", "svetilke-in-luci", "tehnika-in-orodje",
    "topi-majice", "transport-ljubljenckov", "transportni-trakovi-in-elasticne-vrvi", "ure",
    "usb-avto-polnilci", "usb-kabli", "ustvarjanje", "vse-za-vrt",
    "vticnice-in-varovalke", "xenon-kiti", "za-otroke", "za-potovanje",
    "za-zivali", "zabava-za-ljubljence", "zabavna-elektronika", "zari",
    "zarnice-za-avto", "zdravje-in-lepota", "zenski-kombinezoni",
]

# Dodatne kategorije, samodejno odkrite iz feeda (varovalo za morebitne nove).
_FEED_CATEGORIES_DYNAMIC: set = set()


def _all_feed_categories():
    """Združen seznam (ročne + dinamično odkrite), daljše najprej,
    da 'zdravje-in-lepota' preveri pred 'zdravje' in 'za-zivali' pred 'za'."""
    cats = set(_FEED_CATEGORIES) | _FEED_CATEGORIES_DYNAMIC
    return sorted(cats, key=len, reverse=True)

def _sku_in_image_url(sku: str, joined_urls: str, strict: bool = False) -> bool:
    """Maaarket image URL je oblike: .../cache/{kategorija}{SKU}{naziv-slike}-{hash}.ext
    SKU pride TAKOJ za kategorijo (npr. 'dom-in-vrt'+'wave-1'+'izdelek-brez-naslova...').
    Najprej poskusi {kategorija}{SKU} ujemanje (najzanesljiveje). Sicer fallback na
    pojavitev SKU z mejo (za SKU ne sme slediti dodatna številka, da silux38≠silux380).
    Podpira SKU z vezaji (wave-1, swc-09-l).
    strict=True: samo kategorija+SKU ujemanje (brez ohlapnega fallbacka) — za base-SKU trim."""
    import re as _re
    sl = sku.lower().strip()
    if len(sl) < 2:
        return False
    u = joined_urls.lower()

    # 1) Najmočnejše: kategorija neposredno pred SKU (z opcijskim '2020' vmes, ki ga feed včasih vrine)
    for cat in _all_feed_categories():
        for sep in ("", "2020"):
            needle = cat + sep + sl
            pos = u.find(needle)
            while pos != -1:
                after = u[pos + len(needle):]
                if strict:
                    # strict (base-SKU trim): za osnovo mora slediti vezaj ali enojni indeks,
                    # NE poljubna črka (da 'silu'+'x...' ne velja za zadetek)
                    if after == '' or after[0] in '-_':
                        return True
                    if after[0].isdigit() and (len(after) == 1 or not after[1].isdigit()):
                        return True
                else:
                    if after == '' or not after[0].isdigit():
                        return True
                    if len(after) == 1 or not after[1].isdigit():
                        return True
                pos = u.find(needle, pos + 1)

    # 2) Fallback: SKU kjerkoli, z mejo da mu ne sledi nadaljnja številka in
    #    da pred njim ni alfanumerik (da ne ujamemo sredine daljše kode)
    if strict:
        return False
    esc = _re.escape(sl)
    pat = r'(?<![a-z0-9])' + esc + r'(?![0-9])'
    return _re.search(pat, u) is not None


def _norm_sku(s: str) -> str:
    """Normalizira SKU za ujemanje: lowercase + podčrtaj→vezaj.
    Zaloga ima včasih 'ELIPACK_black', feed v URL 'elipack-black' — oba postaneta 'elipack-black'."""
    return s.lower().strip().replace('_', '-')


def _extract_skus_from_image_url(image_url: str) -> list:
    """Izlušči kandidat-SKU(je) iz Maaarket image URL-ja za predizračunan indeks.
    SKU se v URL-ju skoraj vedno pojavi obdan z vezaji: '...-silux100-...', '...-maaa61-...',
    '...-pma-520-...' (SKU z vezajem = dva zaporedna tokena). Filename razbijemo po vezajih
    in kot kandidate vzamemo posamezne tokene IN zlepljene pare/trojke (za SKU z vezaji).
    Vrne seznam normaliziranih kandidatov (lowercase). Hash (zadnji token) izpustimo."""
    import re as _re
    fname = image_url.rsplit('/', 1)[-1].lower()
    fname = _re.sub(r'\.(jpe?g|png|webp|gif)$', '', fname)
    toks = [t for t in fname.split('-') if t]
    if not toks:
        return []
    # zadnji token je hash (dolg hex) — izpusti
    if len(toks) > 1 and _re.fullmatch(r'[0-9a-f]{6,}', toks[-1]):
        toks = toks[:-1]
    out = []
    n = len(toks)
    # PRVI del filename-a: {kategorija}{SKU}{začetek naziva}. Odstrani znano kategorijo z
    # ZAČETKA celotnega filename-a (ne razbitega), da ohranimo SKU z vezaji (wave-1).
    # Nato dodaj prefikse preostanka — pravi SKU se ujame na enem prefiksu.
    for cat in _all_feed_categories():
        matched_cat = False
        for sep in ("2020", ""):
            pref = cat + sep
            if fname.startswith(pref) and len(fname) > len(pref):
                rest = fname[len(pref):]
                for L in range(3, min(len(rest), 16) + 1):
                    cand = rest[:L].rstrip('-')
                    if len(cand) >= 3:
                        out.append(cand)
                matched_cat = True
                break
        if matched_cat:
            break
    for i, t in enumerate(toks):
        if i == 0:
            continue
        if len(t) >= 2:
            out.append(t)
        # SKU z vezajem: zlepi 2 ali 3 zaporedne tokene (pma-520, swc-09-l)
        if i + 1 < n:
            two = t + '-' + toks[i+1]
            if len(two) >= 4:
                out.append(two)
        if i + 2 < n:
            three = t + '-' + toks[i+1] + '-' + toks[i+2]
            if len(three) >= 6:
                out.append(three)
    return out


def parse_feed(xml_content: str) -> dict:
    """Parse Google Shopping XML feed.
    Vrne: {g_id: {url, path, title, description, price, mpn, brand, product_type, image, availability}}
    """
    products = {}
    try:
        root = ET.fromstring(xml_content)
        channel = root.find('channel')
        if channel is None:
            return products
        for item in channel.findall('item'):
            gid_el = item.find(f'{{{G}}}id')
            link_el = item.find(f'{{{G}}}link')
            if gid_el is None or not gid_el.text or link_el is None or not link_el.text:
                continue
            g_id = gid_el.text.strip()
            url = link_el.text.strip()
            path = urlparse(url).path

            def _get(tag):
                el = item.find(f'{{{G}}}{tag}')
                return el.text.strip() if el is not None and el.text else ""

            # Standard RSS tags (brez namespace)
            def _get_rss(tag):
                el = item.find(tag)
                return el.text.strip() if el is not None and el.text else ""

            # vse slike (glavna + dodatne) — za SKU match v URL-ju
            all_imgs = []
            main_img = _get('image_link')
            if main_img:
                all_imgs.append(main_img)
            for ael in item.findall(f'{{{G}}}additional_image_link'):
                if ael.text and ael.text.strip():
                    all_imgs.append(ael.text.strip())

            products[g_id] = {
                "url": url,
                "path": path,
                "title": _get_rss('title') or _get('title'),
                "description": _get_rss('description') or _get('description'),
                "price": _get('price'),
                "sale_price": _get('sale_price'),
                "mpn": _get('mpn'),  # običajno SKU
                "brand": _get('brand'),
                "product_type": _get('product_type'),
                "google_category": _get('google_product_category'),
                "image": main_img,
                "all_images": all_imgs,
                "availability": _get('availability'),
            }
    except ET.ParseError:
        pass
    return products


async def fetch_all_feeds():
    global feed_by_lang, slug_to_id, last_fetch
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Fetching XML feeds...")
    async with httpx.AsyncClient(timeout=30.0) as hc:
        tasks = {lang: hc.get(url) for lang, url in MAAARKET_FEEDS.items()}
        new_cache = {}
        for lang, task in tasks.items():
            try:
                resp = await task
                new_cache[lang] = parse_feed(resp.text) if resp.status_code == 200 else {}
                print(f"  ✓ {lang}: {len(new_cache.get(lang,{}))} products")
            except Exception as e:
                new_cache[lang] = {}
                print(f"  ✗ {lang}: {e}")
    feed_by_lang = new_cache
    new_slug_to_id = {}
    for lang, lang_feed in feed_by_lang.items():
        for g_id, data in lang_feed.items():
            slug = extract_slug(data["url"])
            if slug and slug not in new_slug_to_id:
                new_slug_to_id[slug] = g_id
    slug_to_id = new_slug_to_id

    # SLO slik-indeks (zgradi 1x ob osvežitvi) — za hitri SKU→slika lookup
    global sl_image_index
    new_img_idx = {
        "slug": {},       # feed slug (= sluggificiran naziv) → slika
        "title_slug": {}, # sluggificiran naziv → slika (za naziv match)
        "img_corpus": [], # [(slika, "vse image poti zlepljene lowercase")] — legacy/fallback
        "mpn": {},        # če bi feed kdaj imel mpn
        "sku_exact": {},  # zanesljiv SKU (Ikonka/Amio iz image URL) UPPER → slika
        "sku_url": {},    # PREDIZRAČUNAN: SKU (lowercase, iz image URL) → slika, O(1) lookup
    }
    sl_feed = feed_by_lang.get("sl", {})
    for g_id, prod in sl_feed.items():
        img = prod.get("image", "")
        if not img:
            continue
        # Ikonka/Amio: zanesljiv SKU iz image URL
        bsku = _extract_brand_sku(prod.get("brand", ""), img)
        if bsku:
            new_img_idx["sku_exact"].setdefault(bsku, img)
        # slug izdelka
        slug = (extract_slug(prod.get("url", "")) or "").lower()
        if slug:
            new_img_idx["slug"].setdefault(slug, img)
        # sluggificiran naziv
        tslug = _sl_slugify(prod.get("title", ""))
        if tslug:
            new_img_idx["title_slug"].setdefault(tslug, img)
        # mpn (če obstaja)
        mpn = (prod.get("mpn") or "").strip().upper()
        if mpn:
            new_img_idx["mpn"].setdefault(mpn, img)
        # korpus vseh image poti (za SKU-v-URL match) — legacy fallback
        all_imgs = prod.get("all_images") or [img]
        joined = " ".join(all_imgs).lower()
        new_img_idx["img_corpus"].append((img, joined))
        # PREDIZRAČUNAN sku_url indeks: izlušči SKU iz vsake image poti → O(1) lookup
        for one_img in all_imgs:
            for sk in _extract_skus_from_image_url(one_img):
                new_img_idx["sku_url"].setdefault(sk, img)
    sl_image_index = new_img_idx

    last_fetch = datetime.now()
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Done. Slug index: {len(slug_to_id)}, slike: {len(sl_image_index.get('img_corpus',[]))} izdelkov")
    # shrani na disk, da preživi deploy/restart
    _save_feed_cache_to_disk()


FEED_CACHE_FILE = DATA_DIR / "feed_cache.json"
# Verzija formata indeksa. Dvigni ob spremembi ekstrakcijske logike (sku_url, kategorije ipd.),
# da se star/pokvarjen disk cache samodejno zavrže in zgradi znova s popravljeno kodo.
CACHE_FORMAT_VERSION = 3


def _save_feed_cache_to_disk():
    """Shrani feed cache + indekse na persistent disk (/data), da preživijo deploy."""
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        payload = {
            "format_version": CACHE_FORMAT_VERSION,
            "saved_at": (last_fetch or datetime.now()).isoformat(),
            "feed_by_lang": feed_by_lang,
            "slug_to_id": slug_to_id,
            "sl_image_index": sl_image_index,
        }
        tmp = FEED_CACHE_FILE.with_suffix(".tmp")
        tmp.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        tmp.replace(FEED_CACHE_FILE)
        print(f"[feed cache] shranjen na disk ({FEED_CACHE_FILE})")
    except Exception as e:
        print(f"[feed cache] shranjevanje spodletelo: {e}")


def _load_feed_cache_from_disk() -> bool:
    """Naloži feed cache z diska, če obstaja, je svež (<TTL) in pravilne format verzije.
    Vrne True ob uspehu."""
    global feed_by_lang, slug_to_id, sl_image_index, last_fetch
    try:
        if not FEED_CACHE_FILE.exists():
            return False
        payload = json.loads(FEED_CACHE_FILE.read_text(encoding="utf-8"))
        # zavrzi star format (npr. pokvarjen sku_url indeks iz prejšnje verzije ekstrakcije)
        if payload.get("format_version") != CACHE_FORMAT_VERSION:
            print(f"[feed cache] format verzija se ne ujema (disk={payload.get('format_version')}, koda={CACHE_FORMAT_VERSION}) — bo zgrajen znova")
            return False
        saved_at = datetime.fromisoformat(payload["saved_at"])
        if datetime.now() - saved_at > timedelta(hours=CACHE_TTL_HOURS):
            print("[feed cache] disk cache zastarel, bo osvežen")
            return False
        feed_by_lang = payload.get("feed_by_lang", {})
        slug_to_id = payload.get("slug_to_id", {})
        sl_image_index = payload.get("sl_image_index", {})
        # img_corpus se v JSON serializira kot seznam seznamov [img, joined] → pretvori nazaj v tuple
        if "img_corpus" in sl_image_index:
            sl_image_index["img_corpus"] = [tuple(x) for x in sl_image_index["img_corpus"]]
        last_fetch = saved_at
        print(f"[feed cache] naložen z diska ({len(feed_by_lang.get('sl',{}))} SLO izdelkov, star {(datetime.now()-saved_at).seconds//60} min)")
        return True
    except Exception as e:
        print(f"[feed cache] nalaganje spodletelo: {e}")
        return False


async def ensure_cache_fresh():
    if not is_cache_stale():
        return
    # poskusi z diska brez locka (hitro)
    if _load_feed_cache_from_disk():
        return
    # potrebujemo prenos — a samo EN naenkrat, ostale zahteve počakajo nanj
    lock = _get_feed_lock()
    async with lock:
        # ponovni pregled: med čakanjem je morda nekdo drug že napolnil cache
        if not is_cache_stale():
            return
        if _load_feed_cache_from_disk():
            return
        await fetch_all_feeds()


# ════════════════════════════════════════════════════════════════════
#  SKENIRANJE / INVENTURA — mapiranje črtne kode (kod_kreskowy) → koda (kod)
#  Skener vrne EAN/črtno kodo; mi jo prevedemo v dobaviteljevo kodo (npr. KX3116_2).
#  Zasnova je razširljiva: dodatne dobavitelje dodaš v BARCODE_FEEDS.
# ════════════════════════════════════════════════════════════════════
BARCODE_FEEDS = {
    # ime_dobavitelja: XML URL
    "ikonka": "https://api.ikonka.eu/d7121f05b75c448860ed5467ac1ba3caf2f307c2.xml?variant=b&lang=pl&currency=PLN",
    "amio": "https://amio.pl/xml?id=107&hash=c8c74346b9948675a93f531e4212318a9598e07dd97e48742e3fba5b0b4a68e1",
}
BARCODE_CACHE_FILE = DATA_DIR / "barcode_cache.json"
BARCODE_CACHE_VERSION = 1
BARCODE_TTL_HOURS = 168  # 7 dni

barcode_index: dict = {}          # kod_kreskowy (normaliziran) → {"kod":..., "supplier":..., "name":...}
barcode_last_fetch: Optional[datetime] = None
_barcode_lock: Optional[asyncio.Lock] = None


def _get_barcode_lock() -> asyncio.Lock:
    global _barcode_lock
    if _barcode_lock is None:
        _barcode_lock = asyncio.Lock()
    return _barcode_lock


def _norm_barcode(s: str) -> str:
    """Normaliziraj črtno kodo: samo števke (skenerji včasih dodajo presledke/CR)."""
    import re as _re
    return _re.sub(r"\D", "", (s or "").strip())


def _parse_barcode_feed(xml_text: str, supplier: str) -> dict:
    """Parsira dobaviteljev XML in vrne {kod_kreskowy → {kod, supplier, name}}.
    Robusten na različne sheme: išče <kod> in <kod_kreskowy> v vsakem <produkt>/<product>/<offer>."""
    import re as _re
    out = {}
    try:
        root = ET.fromstring(xml_text)
    except Exception as e:
        print(f"[barcode] {supplier}: XML parse napaka: {e}")
        return out

    # Poišči vse elemente, ki vsebujejo otroka s SKU + črtno kodo (ne glede na shemo/namespace)
    # Ikonka: <kod> (SKU) + <kod_kreskowy> (EAN);  Amio: <PN> (SKU) + <EAN>
    def _localname(tag: str) -> str:
        return tag.rsplit("}", 1)[-1].lower() if tag else ""

    SKU_TAGS = ("kod", "pn", "sku", "symbol", "index")          # ime polja za SKU
    BARCODE_TAGS = ("kod_kreskowy", "ean", "ean13", "barcode", "gtin")  # črtna koda
    NAME_TAGS = ("nazwa", "name", "naziv", "title", "nazev")

    for el in root.iter():
        kod = None
        barcode = None
        name = None
        for child in list(el):
            ln = _localname(child.tag)
            txt = (child.text or "").strip()
            if not txt:
                continue
            if kod is None and ln in SKU_TAGS:
                kod = txt
            elif barcode is None and ln in BARCODE_TAGS:
                barcode = txt
            elif name is None and ln in NAME_TAGS:
                name = txt
        if kod and barcode:
            nb = _norm_barcode(barcode)
            if nb:
                # prvi zmaga (ne povozi); če bi bil isti barcode pri dveh kodah, obdrži prvega
                out.setdefault(nb, {"kod": kod, "supplier": supplier, "name": name or ""})
    print(f"[barcode] {supplier}: {len(out)} barkod")
    return out


def _save_barcode_cache():
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        payload = {
            "format_version": BARCODE_CACHE_VERSION,
            "saved_at": (barcode_last_fetch or datetime.now()).isoformat(),
            "barcode_index": barcode_index,
        }
        tmp = BARCODE_CACHE_FILE.with_suffix(".tmp")
        tmp.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        tmp.replace(BARCODE_CACHE_FILE)
        print(f"[barcode cache] shranjen ({len(barcode_index)} barkod)")
    except Exception as e:
        print(f"[barcode cache] shranjevanje spodletelo: {e}")


def _load_barcode_cache() -> bool:
    global barcode_index, barcode_last_fetch
    try:
        if not BARCODE_CACHE_FILE.exists():
            return False
        payload = json.loads(BARCODE_CACHE_FILE.read_text(encoding="utf-8"))
        if payload.get("format_version") != BARCODE_CACHE_VERSION:
            return False
        saved_at = datetime.fromisoformat(payload["saved_at"])
        if datetime.now() - saved_at > timedelta(hours=BARCODE_TTL_HOURS):
            return False
        barcode_index = payload.get("barcode_index", {})
        barcode_last_fetch = saved_at
        print(f"[barcode cache] naložen ({len(barcode_index)} barkod, star {(datetime.now()-saved_at).seconds//60} min)")
        return True
    except Exception as e:
        print(f"[barcode cache] nalaganje spodletelo: {e}")
        return False


async def fetch_barcode_feeds():
    """Potegni vse dobaviteljeve XML in zgradi barcode_index (kod_kreskowy → kod)."""
    global barcode_index, barcode_last_fetch
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Fetching barcode feeds...")
    new_idx = {}
    async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as hc:
        for supplier, url in BARCODE_FEEDS.items():
            try:
                resp = await hc.get(url)
                if resp.status_code == 200:
                    new_idx.update(_parse_barcode_feed(resp.text, supplier))
                else:
                    print(f"[barcode] {supplier}: HTTP {resp.status_code}")
            except Exception as e:
                print(f"[barcode] {supplier}: {e}")
    if new_idx:
        barcode_index = new_idx
        barcode_last_fetch = datetime.now()
        _save_barcode_cache()
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Barcode index: {len(barcode_index)} kod")


def _barcode_is_stale() -> bool:
    if not barcode_index:
        return True
    if barcode_last_fetch is None:
        return True
    return datetime.now() - barcode_last_fetch > timedelta(hours=BARCODE_TTL_HOURS)


async def ensure_barcode_fresh():
    if not _barcode_is_stale():
        return
    if _load_barcode_cache():
        return
    lock = _get_barcode_lock()
    async with lock:
        if not _barcode_is_stale():
            return
        if _load_barcode_cache():
            return
        await fetch_barcode_feeds()


def detect_brand(url: str) -> Optional[str]:
    if not url:
        return None
    domain = urlparse(url).netloc.lower().replace("www.", "")
    for brand, lang_map in BRAND_DOMAINS.items():
        for d in lang_map.values():
            if domain == d.replace("www.", ""):
                return brand
    return None


def find_product_urls(source_url: Optional[str]) -> dict:
    if not source_url:
        return {}
    brand = detect_brand(source_url) or "maaarket"
    slug = extract_slug(source_url)
    if not slug:
        return {}
    g_id = slug_to_id.get(slug)
    if not g_id:
        return {}
    target_domains = BRAND_DOMAINS.get(brand, BRAND_DOMAINS["maaarket"])
    result = {}
    for lang, products in feed_by_lang.items():
        if lang not in target_domains or g_id not in products:
            continue
        result[lang] = f"https://{target_domains[lang]}{products[g_id]['path']}"
    return result


# ─── STARTUP ─────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup_event():
    # Startup mora biti HITER, da Render health check (/healthz) takoj uspe.
    # Vse počasne operacije (FFmpeg download, feed prenos) gredo v ozadje.

    # Feed cache: disk je hiter (preživi deploy), zato ga poskusimo takoj (ms).
    # Če diska ni / je zastarel, NE blokiramo zagona — prenos gre v ozadje.
    loaded = _load_feed_cache_from_disk()
    if not loaded:
        asyncio.create_task(fetch_all_feeds())

    # FFmpeg warm-up v ozadju (lahko traja do 30s ob prvem zagonu) — ne sme blokirati startupa
    asyncio.create_task(_ffmpeg_warmup())

    asyncio.create_task(periodic_refresh())
    asyncio.create_task(_daily_cashflow_sync())
    asyncio.create_task(_email_polling_loop())
    asyncio.create_task(_forecast2_scheduler_loop())
    asyncio.create_task(_zaloga_scheduler_loop())
    asyncio.create_task(_hsplus_daily_scheduler())
    asyncio.create_task(_hsplus_daily_scheduler())
    asyncio.create_task(_regen_worker_loop())


async def _ffmpeg_warmup():
    """FFmpeg priprava v ozadju, da prvi /merge-video-audio ne čaka in startup ni blokiran."""
    try:
        import static_ffmpeg
        static_ffmpeg.add_paths()
        import subprocess
        # teci v thread executorju, da ne blokira event loopa
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(
            None,
            lambda: subprocess.run(["ffmpeg", "-version"], capture_output=True, timeout=60)
        )
        if result.returncode == 0:
            print(f"[startup] FFmpeg ready: {result.stdout.decode()[:60]}")
        else:
            print(f"[startup] FFmpeg warning: returncode={result.returncode}")
    except Exception as e:
        print(f"[startup] FFmpeg warm-up failed: {e}")


async def periodic_refresh():
    while True:
        # počakaj do poteka TTL (7 dni) glede na zadnji prenos
        if last_fetch:
            age = (datetime.now() - last_fetch).total_seconds()
            wait = max(60, CACHE_TTL_HOURS * 3600 - age)
        else:
            wait = CACHE_TTL_HOURS * 3600
        await asyncio.sleep(wait)
        await fetch_all_feeds()


# ─── EMAIL IMAP POLLING ──────────────────────────────────────────────────────

EMAIL_CONFIG_FILE = DATA_DIR / "email_config.json"
EMAIL_LOG_FILE    = DATA_DIR / "email_log.json"
EMAIL_POLL_INTERVAL = 300  # 5 minut


def _load_email_config() -> dict:
    if EMAIL_CONFIG_FILE.exists():
        try:
            return json.loads(EMAIL_CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_email_config(cfg: dict):
    # Nikoli ne shranjujemo gesla v plaintext če je v env
    EMAIL_CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_email_log() -> list:
    if EMAIL_LOG_FILE.exists():
        try:
            return json.loads(EMAIL_LOG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return []


def _append_email_log(entry: dict):
    log = _load_email_log()
    log.insert(0, entry)  # newest first
    if len(log) > 500:
        log = log[:500]
    EMAIL_LOG_FILE.write_text(json.dumps(log, ensure_ascii=False), encoding="utf-8")


def _email_get_password(cfg: dict) -> str:
    """Geslo iz env vara ima prioriteto pred shranjenim."""
    return os.environ.get("EMAIL_PASSWORD", "") or cfg.get("password", "")


async def _email_polling_loop():
    """Background loop — preveri inbox vsake EMAIL_POLL_INTERVAL sekund."""
    await asyncio.sleep(30)  # počakaj na startup
    while True:
        try:
            cfg = _load_email_config()
            if cfg.get("enabled") and cfg.get("imap_host") and cfg.get("email"):
                password = _email_get_password(cfg)
                if password:
                    count = await asyncio.get_event_loop().run_in_executor(
                        None, _process_inbox, cfg, password
                    )
                    if count > 0:
                        print(f"[email] Processed {count} new invoice(s)")
        except Exception as e:
            print(f"[email] Polling error: {e}")
        await asyncio.sleep(EMAIL_POLL_INTERVAL)


def _process_inbox(cfg: dict, password: str) -> int:
    """Sinhrona IMAP obdelava — teče v executor threadu."""
    import imaplib
    import email as email_lib
    from email.header import decode_header as _dh
    from datetime import datetime as _dt

    host = cfg["imap_host"]
    port = int(cfg.get("imap_port", 993))
    username = cfg["email"]
    processed_folder = cfg.get("processed_folder", "Processed")
    count = 0

    try:
        if port == 993:
            mail = imaplib.IMAP4_SSL(host, port)
        else:
            mail = imaplib.IMAP4(host, port)
            mail.starttls()
        mail.login(username, password)
    except Exception as e:
        _append_email_log({
            "ts": _dt.now().isoformat(), "type": "error",
            "message": f"IMAP login failed: {e}", "from": "", "subject": ""
        })
        return 0

    try:
        mail.select("INBOX")
        # Išči neprebrane emaile z attachmenti
        _, msg_nums = mail.search(None, "UNSEEN")
        if not msg_nums[0]:
            mail.logout()
            return 0

        for num in msg_nums[0].split():
            try:
                _, msg_data = mail.fetch(num, "(RFC822)")
                raw = msg_data[0][1]
                msg = email_lib.message_from_bytes(raw)

                # Decode subject
                subj_parts = _dh(msg.get("Subject", ""))
                subject = "".join(
                    p.decode(enc or "utf-8") if isinstance(p, bytes) else p
                    for p, enc in subj_parts
                )
                sender = msg.get("From", "")

                # Najdi attachmente
                attachments = []
                for part in msg.walk():
                    cd = part.get("Content-Disposition", "")
                    if "attachment" not in cd and "inline" not in cd:
                        continue
                    filename = part.get_filename()
                    if not filename:
                        continue
                    # Decode filename
                    fn_parts = _dh(filename)
                    filename = "".join(
                        p.decode(enc or "utf-8") if isinstance(p, bytes) else p
                        for p, enc in fn_parts
                    )
                    content = part.get_payload(decode=True)
                    if content:
                        attachments.append((filename, content))

                if not attachments:
                    # Ni attachmentov — označi kot prebran brez obdelave
                    mail.store(num, "+FLAGS", "\\Seen")
                    continue

                # Obdelaj vsak attachment
                results = []
                for filename, content in attachments:
                    result = _process_email_attachment(filename, content, subject, sender)
                    results.append(result)

                # Log
                statuses = [r["status"] for r in results]
                _append_email_log({
                    "ts": _dt.now().isoformat(),
                    "type": "processed",
                    "from": sender,
                    "subject": subject,
                    "attachments": [r["filename"] for r in results],
                    "suppliers": [r.get("supplier", "?") for r in results],
                    "statuses": statuses,
                    "items": [r.get("item_count", 0) for r in results],
                    "record_ids": [r.get("record_id", "") for r in results],
                })

                # Označi kot prebran
                mail.store(num, "+FLAGS", "\\Seen")

                # Prestavi v Processed folder (ustvari če ne obstaja)
                try:
                    mail.create(processed_folder)
                except Exception:
                    pass
                try:
                    mail.copy(num, processed_folder)
                    mail.store(num, "+FLAGS", "\\Deleted")
                    mail.expunge()
                except Exception:
                    pass  # Folder move ni critical

                count += 1

            except Exception as e:
                _append_email_log({
                    "ts": _dt.now().isoformat(), "type": "error",
                    "message": f"Email processing error: {e}", "from": "", "subject": ""
                })

    finally:
        try:
            mail.logout()
        except Exception:
            pass

    return count


def _process_email_attachment(filename: str, content: bytes, subject: str, sender: str) -> dict:
    """Obdela en attachment — zaznaj dobavitelja in parsiraj."""
    from datetime import datetime as _dt

    ts = _dt.now().strftime("%Y%m%d_%H%M%S")
    supplier = _detect_supplier(filename, content)
    result = {
        "filename": filename,
        "supplier": supplier,
        "status": "ok",
        "record_id": "",
        "item_count": 0,
    }

    try:
        if supplier == "amio_xml":
            parsed = _parse_amio_xml(content)
            supplier = "amio"
        elif supplier == "amio":
            parsed = _parse_amio_pdf_bytes(content)
        elif supplier == "motoprofil":
            parsed = _parse_motoprofil_csv(content, filename)
        elif supplier == "intercars":
            parsed = _parse_intercars_xml(content)
        elif supplier == "abakus":
            parsed = _parse_abakus_xlsx(content)
        elif supplier == "ikonka":
            parsed = _parse_ikonka_pdf(content)
        else:
            # Neznan dobavitelj — shrani attachment za ročno obdelavo
            unknown_dir = PREVZEMI_DIR / f"_email_unknown_{ts}"
            unknown_dir.mkdir(parents=True, exist_ok=True)
            (unknown_dir / _safe_filename(filename)).write_bytes(content)
            (unknown_dir / "meta.json").write_text(json.dumps({
                "source": "email", "from": sender, "subject": subject,
                "filename": filename, "supplier": "unknown", "ts": ts
            }, ensure_ascii=False), encoding="utf-8")
            result["status"] = "unknown_supplier"
            return result

        if not parsed or not parsed.get("items"):
            result["status"] = "no_items"
            return result

        # Shrani enako kot ročni upload
        # Normaliziraj — odstrani presledke iz invoice_number (npr. 'FA 162383' → 'FA_162383')
        if parsed.get('invoice_number'):
            parsed['invoice_number'] = _normalize_invoice_number(parsed['invoice_number'])
        invoice_num_safe = _safe_filename(parsed.get('invoice_number', 'unknown'))
        record_id = f"{ts}_{invoice_num_safe}"
        target_dir = PREVZEMI_DIR / record_id
        target_dir.mkdir(parents=True, exist_ok=True)

        (target_dir / f"source_{_safe_filename(filename)}").write_bytes(content)
        (target_dir / "parsed.json").write_text(
            json.dumps(parsed, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        # Override supplier če pattern invoice številke ali vendor_name kaže drugače
        supplier = _override_supplier_by_invoice(
            supplier,
            parsed.get("invoice_number", ""),
            parsed.get("vendor_name", "")
        )
        (target_dir / "meta.json").write_text(json.dumps({
            "record_id": record_id,
            "source": "email",
            "original_filename": filename,
            "from": sender,
            "subject": subject,
            "created_ts": ts,
            "supplier": supplier,
            "supplier_name": SUPPLIER_NAMES.get(supplier, supplier),
            "vendor_id": SUPPLIER_VENDOR_IDS.get(supplier, ""),
            "invoice_number": parsed.get("invoice_number", ""),
            "invoice_date": parsed.get("invoice_date", ""),
            "vendor_name": parsed.get("vendor_name", ""),
            "item_count": len(parsed.get("items", [])),
        }, ensure_ascii=False, indent=2), encoding="utf-8")

        result["record_id"] = record_id
        result["item_count"] = len(parsed.get("items", []))

    except Exception as e:
        result["status"] = f"error: {e}"

    return result


def _parse_amio_pdf_bytes(content: bytes) -> dict:
    """AMiO PDF parse z Claude — enako kot /prevzemi-parse-pdf endpoint."""
    import base64
    pdf_b64 = base64.standard_b64encode(content).decode("utf-8")
    # Reuse obstoječe Claude prompt logike — pokliči parse endpoint interno
    # (Tukaj dupliciramo ključno logiko da se izognemo HTTP loop)
    msg = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=20000,
        messages=[{
            "role": "user",
            "content": [
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                {"type": "text", "text": """Parse this supplier invoice PDF. Return ONLY valid JSON.

JSON format:
{
  "invoice_number": "...",
  "invoice_date": "YYYY-MM-DD",
  "vendor_name": "...",
  "currency": "EUR",
  "items": [
    {"product_number": "...", "product_name": "...", "qty": 1, "unit_price": 0.0, "value": 0.0}
  ]
}

Rules:
- Net price only (brez DDV)
- Skip header/footer rows
- product_number: kataloška številka
"""}
            ]
        }]
    )
    import re as _re
    raw = msg.content[0].text.strip()
    raw = _re.sub(r'^```(?:json)?\s*', '', raw)
    raw = _re.sub(r'\s*```$', '', raw)
    parsed = json.loads(raw)
    for it in parsed.get("items", []):
        try:
            it["value"] = round(float(it.get("qty", 1)) * float(it.get("unit_price", 0)), 2)
        except Exception:
            it["value"] = 0
    parsed.setdefault("currency", "EUR")
    return parsed


# ─── EMAIL API ENDPOINTS ─────────────────────────────────────────────────────

class EmailConfigRequest(BaseModel):
    imap_host: str
    imap_port: int = 993
    email: str
    password: str = ""
    processed_folder: str = "Processed"
    enabled: bool = True


@app.post("/email-config")
async def email_config_save(req: EmailConfigRequest):
    """Shrani IMAP konfiguracija."""
    try:
        cfg = {
            "imap_host": req.imap_host.strip(),
            "imap_port": req.imap_port,
            "email": req.email.strip(),
            "processed_folder": req.processed_folder.strip() or "Processed",
            "enabled": req.enabled,
        }
        # Geslo shrani samo če ni v env varu
        env_pw = os.environ.get("EMAIL_PASSWORD", "")
        if req.password and not env_pw:
            cfg["password"] = req.password
        elif env_pw:
            cfg["password"] = ""  # Ni treba shraniti — je v env
        _save_email_config(cfg)
        return {"ok": True, "config": {k: v for k, v in cfg.items() if k != "password"}}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/email-config")
async def email_config_get():
    """Vrne trenutno konfiguracijo (brez gesla)."""
    cfg = _load_email_config()
    has_password = bool(_email_get_password(cfg))
    return {
        "ok": True,
        "config": {k: v for k, v in cfg.items() if k != "password"},
        "has_password": has_password,
        "password_source": "env_var" if os.environ.get("EMAIL_PASSWORD") else ("saved" if cfg.get("password") else "none"),
        "poll_interval_sec": EMAIL_POLL_INTERVAL,
    }


@app.post("/email-check-now")
async def email_check_now():
    """Ročni trigger — takoj preveri inbox."""
    try:
        cfg = _load_email_config()
        if not cfg.get("enabled"):
            return {"ok": False, "error": "Email polling ni omogočen"}
        if not cfg.get("imap_host") or not cfg.get("email"):
            return {"ok": False, "error": "Email ni konfiguriran"}
        password = _email_get_password(cfg)
        if not password:
            return {"ok": False, "error": "Geslo manjka — nastavi EMAIL_PASSWORD env var ali vnesi v formi"}

        count = await asyncio.get_event_loop().run_in_executor(
            None, _process_inbox, cfg, password
        )
        return {"ok": True, "processed": count, "message": f"Najdenih {count} novih računov"}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.post("/email-test-connection")
async def email_test_connection():
    """Testira IMAP povezavo brez obdelave emailov."""
    import imaplib
    try:
        cfg = _load_email_config()
        password = _email_get_password(cfg)
        if not cfg.get("imap_host") or not cfg.get("email") or not password:
            return {"ok": False, "error": "Konfiguracija ni popolna"}

        host = cfg["imap_host"]
        port = int(cfg.get("imap_port", 993))

        if port == 993:
            mail = imaplib.IMAP4_SSL(host, port)
        else:
            mail = imaplib.IMAP4(host, port)
            mail.starttls()

        mail.login(cfg["email"], password)
        _, folders = mail.list()
        mail.select("INBOX")
        _, unseen = mail.search(None, "UNSEEN")
        unseen_count = len(unseen[0].split()) if unseen[0] else 0
        mail.logout()

        return {
            "ok": True,
            "message": "Povezava uspešna",
            "unseen_emails": unseen_count,
            "folders": [f.decode() for f in (folders or [])[:10]],
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/email-log")
async def email_log_get(limit: int = 50):
    """Vrne zadnjih N log zapisov."""
    log = _load_email_log()
    return {"ok": True, "log": log[:limit], "total": len(log)}


@app.delete("/email-log")
async def email_log_clear():
    """Zbriše email log."""
    EMAIL_LOG_FILE.write_text("[]", encoding="utf-8")
    return {"ok": True}


# ─── KNJIGOVODSTVO EMAIL BATCH PROCESSING ────────────────────────────────────

STORAGE_KNJ_DIR = DATA_DIR / "storage" / "knjigovodstvo"
STORAGE_KNJ_DIR.mkdir(parents=True, exist_ok=True)


# ─── VRAČILA — SCANNER ───────────────────────────────────────────────────────

VRACILA_DIR = DATA_DIR / "vracila"
VRACILA_DIR.mkdir(parents=True, exist_ok=True)
VRACILA_CURRENT = VRACILA_DIR / "current.json"
VRACILA_ARCHIVE_DIR = VRACILA_DIR / "archive"
VRACILA_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)


class VracilaSaveRequest(BaseModel):
    items: list  # [{id, comment, ts}, ...]


@app.get("/vracila-current")
async def vracila_current_get():
    """Vrne aktivno sejo (current.json)."""
    try:
        if VRACILA_CURRENT.exists():
            data = json.loads(VRACILA_CURRENT.read_text(encoding="utf-8"))
            return {"ok": True, "items": data.get("items", []), "started_at": data.get("started_at")}
        return {"ok": True, "items": [], "started_at": None}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/vracila-save")
async def vracila_save(req: VracilaSaveRequest):
    """Shrani trenutno sejo — kliče se na vsako spremembo."""
    try:
        from datetime import datetime as _dt
        existing_started = None
        if VRACILA_CURRENT.exists():
            try:
                old = json.loads(VRACILA_CURRENT.read_text(encoding="utf-8"))
                existing_started = old.get("started_at")
            except Exception:
                pass
        data = {
            "started_at": existing_started or _dt.now().isoformat(),
            "updated_at": _dt.now().isoformat(),
            "items": req.items or [],
        }
        VRACILA_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "count": len(data["items"])}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.post("/vracila-archive")
async def vracila_archive():
    """Arhivira aktivno sejo in resetira."""
    try:
        from datetime import datetime as _dt
        if not VRACILA_CURRENT.exists():
            return {"ok": False, "error": "Ni aktivne seje"}
        data = json.loads(VRACILA_CURRENT.read_text(encoding="utf-8"))
        items = data.get("items", [])
        if not items:
            VRACILA_CURRENT.unlink()
            return {"ok": True, "message": "Seja je bila prazna, izbrisana"}
        ts = _dt.now().strftime("%Y%m%d_%H%M%S")
        archive_name = f"{ts}_{len(items)}items.json"
        archive_path = VRACILA_ARCHIVE_DIR / archive_name
        data["archived_at"] = _dt.now().isoformat()
        archive_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        VRACILA_CURRENT.unlink()
        return {"ok": True, "archived": archive_name, "items": len(items)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/vracila-history")
async def vracila_history():
    """Seznam arhiviranih sej."""
    try:
        from datetime import datetime as _dt
        items = []
        for f in sorted(VRACILA_ARCHIVE_DIR.iterdir(), reverse=True):
            if not f.is_file() or not f.name.endswith(".json"):
                continue
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                items.append({
                    "filename": f.name,
                    "started_at": data.get("started_at"),
                    "archived_at": data.get("archived_at"),
                    "count": len(data.get("items", [])),
                    "size_kb": round(f.stat().st_size / 1024, 1),
                })
            except Exception:
                continue
        return {"ok": True, "items": items}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/vracila-history/{filename}")
async def vracila_history_detail(filename: str):
    """Podrobnosti enega arhiva."""
    try:
        safe = filename.replace("/", "").replace("..", "")
        f = VRACILA_ARCHIVE_DIR / safe
        if not f.exists():
            return {"ok": False, "error": "Ne obstaja"}
        data = json.loads(f.read_text(encoding="utf-8"))
        return {"ok": True, "data": data}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ─── ZALOGA / NABIRANJE (picking list za skladišče) ──────────────────────────
ZALOGA_DIR = DATA_DIR / "zaloga"
ZALOGA_DIR.mkdir(parents=True, exist_ok=True)
ZALOGA_CURRENT = ZALOGA_DIR / "current.json"  # legacy SLO (nazaj-združljivost)
ZALOGA_ARCHIVE_DIR = ZALOGA_DIR / "archive"
ZALOGA_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

# Dodatne (backup) lokacije za SKU — neodvisno od siluxarja (preživi sync).
# Oblika: { "SKU": ["P2-B", "R4-A"], ... }  (primarna pozicija ostane v CSV/siluxar)
ZALOGA_EXTRA_POS = ZALOGA_DIR / "extra_positions.json"

def _zaloga_load_extra_pos() -> dict:
    """Naloži dodatne pozicije (sku → seznam). Vrne {} če ne obstaja."""
    try:
        if ZALOGA_EXTRA_POS.exists():
            d = json.loads(ZALOGA_EXTRA_POS.read_text(encoding="utf-8"))
            return d if isinstance(d, dict) else {}
    except Exception:
        pass
    return {}

def _zaloga_save_extra_pos(d: dict):
    """Atomično zapiši dodatne pozicije."""
    try:
        tmp = ZALOGA_EXTRA_POS.with_suffix(".tmp")
        tmp.write_text(json.dumps(d, ensure_ascii=False), encoding="utf-8")
        os.replace(tmp, ZALOGA_EXTRA_POS)
    except Exception as e:
        print(f"[extra_pos] save err: {e}")


def _zaloga_market(m: str) -> str:
    """Normalizira market kodo. Privzeto 'slo'."""
    m = (m or "slo").strip().lower()
    return m if m in ("slo", "rs") else "slo"


def _zaloga_current_path(market: str) -> Path:
    """Pot do aktivne seje za trg. SLO uporablja legacy current.json."""
    market = _zaloga_market(market)
    if market == "slo":
        return ZALOGA_CURRENT  # legacy pot ostane
    return ZALOGA_DIR / f"current_{market}.json"


def _zaloga_atomic_write(path: Path, data: dict):
    """Atomično zapiši sejo: najprej v .tmp, nato os.replace (atomično na istem disku).
    Tako prekinjen zapis (restart/deploy sredi pisanja) NE pokvari obstoječe datoteke."""
    import os as _os
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    _os.replace(str(tmp), str(path))  # atomično


def _zaloga_expected_totals(sess: dict):
    """Skupne PRIČAKOVANE postavke in kose v seji (police + čakajoče).
    To je 'obseg dela' — število nabranih raste, ta dva seštevka pa NE smeta padati."""
    items = sess.get("items") or []
    cak = sess.get("cakajoce") or []
    total_items = len(items) + len(cak)
    total_qty = 0
    for it in items:
        try:
            total_qty += int(it.get("qty", 0) or 0)
        except (ValueError, TypeError):
            pass
    for c in cak:
        try:
            total_qty += int(c.get("qty", 0) or 0)
        except (ValueError, TypeError):
            pass
    return total_items, total_qty


def _zaloga_update_peak(sess: dict):
    """Posodobi baseline (peak) skupnih postavk + kosov. Peak je največja doslej videna
    vrednost; če dejansko stanje pade POD peak, je prišlo do nepričakovanega izbrisa/napake.
    Vrne dict z zaznavo padca za opozorilo na frontendu."""
    cur_items, cur_qty = _zaloga_expected_totals(sess)
    peak_items = int(sess.get("peak_items", 0) or 0)
    peak_qty = int(sess.get("peak_qty", 0) or 0)
    # peak lahko le raste (legitimno: uvoz dobavnice doda postavke/kose)
    if cur_items > peak_items:
        peak_items = cur_items
    if cur_qty > peak_qty:
        peak_qty = cur_qty
    sess["peak_items"] = peak_items
    sess["peak_qty"] = peak_qty
    return {
        "cur_items": cur_items, "cur_qty": cur_qty,
        "peak_items": peak_items, "peak_qty": peak_qty,
        "dropped_items": max(0, peak_items - cur_items),
        "dropped_qty": max(0, peak_qty - cur_qty),
    }


def _zaloga_archive_dir(market: str) -> Path:
    """Arhiv mapa za trg. SLO uporablja legacy archive/."""
    market = _zaloga_market(market)
    if market == "slo":
        return ZALOGA_ARCHIVE_DIR  # legacy pot ostane
    d = ZALOGA_DIR / f"archive_{market}"
    d.mkdir(parents=True, exist_ok=True)
    return d


# ── HS PLUS trajna shramba (oznaka velja 5 koledarskih dni, čez seje/arhive) ──
HSPLUS_DAYS = 5   # dan uvoza = dan 1 → velja še +4 dni

def _hsplus_store_path(market: str) -> Path:
    return ZALOGA_DIR / f"hsplus_active_{_zaloga_market(market)}.json"

def _hsplus_load_active(market: str) -> dict:
    """Vrne {sku: expires_isodate} samo za SKU-je, ki ŠE NISO potekli (počisti potekle)."""
    from datetime import date as _date
    p = _hsplus_store_path(market)
    if not p.exists():
        return {}
    try:
        store = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}
    today = _date.today().isoformat()
    active = {sku: exp for sku, exp in store.items() if str(exp) >= today}
    if len(active) != len(store):   # počisti potekle z diska
        _zaloga_atomic_write(p, active)
    return active

def _hsplus_save_active(market: str, sku_list):
    """Dodaj/osveži SKU-je z rokom = danes + (HSPLUS_DAYS-1) dni (dan uvoza = dan 1)."""
    from datetime import date as _date, timedelta as _td
    active = _hsplus_load_active(market)
    expires = (_date.today() + _td(days=HSPLUS_DAYS - 1)).isoformat()
    for sku in sku_list:
        active[str(sku)] = expires   # najnovejši uvoz podaljša rok
    _zaloga_atomic_write(_hsplus_store_path(market), active)

def _hsplus_mark_items(market: str, items: list):
    """Označi postavke v seji s HS PLUS, če je SKU v aktivni shrambi (brez qty).
    Vrne število označenih."""
    active = _hsplus_load_active(market)
    if not active:
        return 0
    n = 0
    for it in items:
        if str(it.get("sku", "")).strip() in active:
            it["hsplus"] = True   # samo oznaka, brez hsplus_qty (prenesena oznaka)
            n += 1
    return n


def _zaloga_group(poz: str, sku: str = "") -> str:
    """Razvrsti pozicijo v skupino (zavihek). 01-1C→Polica 01, P13-C→P13, Ni podatka/Paleta/Pod Mizo ostanejo.
    Če ni podatka o poziciji, razvrsti po predponi SKU: KX*→Ikonka, 0*→Amio."""
    import re as _re
    poz = (poz or "").strip()
    if not poz or poz.lower() == "ni podatka":
        # brez pozicije → poskusi razvrstiti po predponi SKU
        s = (sku or "").strip()
        if s.upper().startswith("KX"):
            return "Ikonka"
        if s.startswith("0"):
            return "Amio"
        return "Ni podatka"
    if poz in ("Paleta", "Pod Mizo"):
        return poz
    m = _re.match(r'^(\d{2})-', poz)
    if m:
        return "Polica " + m.group(1)
    m = _re.match(r'^(P\d+)-', poz)
    if m:
        return m.group(1)
    return poz


@app.post("/zaloga-upload")
async def zaloga_upload(file: UploadFile = File(...), market: str = "slo"):
    """Naloži CSV za nabiranje. Parsira ID, SKU, Naziv, Količina, Pozicija SL → skupine.
    Ustvari novo aktivno sejo za izbrani trg. Obstoječa se prepiše."""
    try:
        import csv as _csv, io as _io
        from datetime import datetime as _dt
        raw = await file.read()
        text = raw.decode("utf-8-sig", errors="replace")
        reader = _csv.DictReader(_io.StringIO(text))

        items = []
        for i, row in enumerate(reader):
            # Sprejmi različne variante imen stolpcev
            def col(*names):
                for n in names:
                    if n in row and row[n] is not None:
                        return str(row[n]).strip()
                return ""
            sku = col("SKU", "sku")
            if not sku:
                continue
            poz = col("Pozicija SL", "Pozicija", "pozicija SL", "pozicija")
            qty_raw = col("Količina", "Kolicina", "kolicina", "qty")
            try:
                qty = int(float(qty_raw)) if qty_raw else 0
            except ValueError:
                qty = 0
            # Nizka zaloga = če Opomba stolpec vsebuje "NIZKA ZALOGA" (NE iz količine)
            opomba = col("Opomba", "opomba", "Note")
            is_low = "NIZKA" in opomba.upper()
            items.append({
                "idx": i,
                "id": col("ID naročila", "ID", "id"),
                "sku": sku,
                "naziv": col("Naziv", "naziv", "Name"),
                "qty": qty,            # potrebna količina
                "poz": poz or "Ni podatka",
                "group": _zaloga_group(poz, sku),
                "status": "",         # "" | "ok" | "ni"
                "picked": qty,        # koliko dejansko nabrано (privzeto = potrebna)
                "low": is_low,        # nizka zaloga tag — iz Opomba stolpca
                "box": "",            # RS: glavni box (zaklenjen ob Shrani)
                "locked": False,      # RS: zaklenjeno (box dodeljen)
                "opomba": "",         # RS: dodatni box (prosto, neodvisno od glavnega)
            })

        # HS PLUS: prenesi aktivne oznake (5-dnevno okno) na nove postavke po SKU
        _hsplus_mark_items(market, items)

        data = {
            "started_at": _dt.now().isoformat(),
            "updated_at": _dt.now().isoformat(),
            "filename": file.filename,
            "market": _zaloga_market(market),
            "items": items,
            "extra_boxes": {},   # RS: dodatni boxi (viški) — { "99": [{sku, naziv, kos}], ... }
            "pick_started_at": None,   # časovnica nabiranja — nova seja vedno začne pri 0
            "pick_finished_at": None,
        }
        # BASELINE (varovalka): pričakovane skupne postavke + kosi.
        # Ne sme padati; če dejansko stanje pade pod peak → opozorilo (možen izbris/napaka).
        _pi, _pq = _zaloga_expected_totals(data)
        data["peak_items"] = _pi
        data["peak_qty"] = _pq
        _zaloga_atomic_write(_zaloga_current_path(market), data)
        # Statistika skupin
        groups = {}
        for it in items:
            groups[it["group"]] = groups.get(it["group"], 0) + 1
        return {"ok": True, "count": len(items), "groups": groups}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


def _zaloga_import_group(sku: str) -> str:
    """Razvrsti uvoženo postavko po predponi SKU: KX*→Ikonka, številka→Amio, drugo→neznano."""
    s = (sku or "").strip().upper()
    if s.startswith("KX"):
        return "Ikonka"
    if s and s[0].isdigit():
        return "Amio"
    return "neznano"


@app.post("/zaloga-import-ikonka")
async def zaloga_import_ikonka(file: UploadFile = File(...), market: str = "rs"):
    """Uvoz dobavnice (CSV/XLS) — DODA k obstoječi RS seji.
    Format: ločilo ';', BOM, količine kot '1,000'. Stolpci (prvi 3):
      NazivMaterial_Ang = SKU, NazivMaterial = naziv, Kolicina = količina.
    Razvrstitev: KX*→Ikonka, 0*/št→Amio, drugo→neznano.
    Količina = 1 → na polico; količina > 1 → v čakajoče (razdelitev v bokse)."""
    try:
        import csv as _csv, io as _io
        from datetime import datetime as _dt
        raw = await file.read()
        fname = (file.filename or "").lower()

        parsed_rows = []  # (sku, naziv, kol)

        if fname.endswith(".xls") or fname.endswith(".xlsx"):
            # Excel: preberi prek openpyxl (xlsx) ali xlrd (xls)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                header = next(rows_iter, None)
                for r in rows_iter:
                    if not r or len(r) < 3:
                        continue
                    parsed_rows.append((str(r[0] or "").strip(), str(r[1] or "").strip(), r[2]))
            except Exception as ex:
                return {"ok": False, "error": f"Excel branje ni uspelo: {ex}. Shrani kot CSV."}
        else:
            # CSV: ločilo ';', BOM odstrani z utf-8-sig
            text = raw.decode("utf-8-sig", errors="replace")
            # zaznaj ločilo (privzeto ';', sicer ',')
            delim = ";" if text.count(";") >= text.count(",") else ","
            reader = _csv.reader(_io.StringIO(text), delimiter=delim)
            rowlist = list(reader)
            # preskoči glavo (prva vrstica)
            for r in rowlist[1:]:
                if not r or len(r) < 3:
                    continue
                parsed_rows.append((r[0].strip(), r[1].strip(), r[2]))

        def _parse_kol(v):
            """'1,000' / '3,000' / 1.0 / '2' → int."""
            if v is None:
                return 0
            s = str(v).strip().replace(",", ".")
            try:
                return int(float(s))
            except ValueError:
                return 0

        # ZDRUŽI podvojene SKU: dobavnica ima nezdružene postavke (isti SKU v več vrsticah).
        # Seštejemo količine PRED razporejanjem, sicer bi se napačno razvrstilo polica/čakajoče
        # (npr. SKU 3× po 1 → mora biti kol=3 v čakajoče, ne 3 ločene postavke na polici).
        _merged = {}        # sku -> {"naziv": prvi naziv, "kol": vsota}
        _order = []         # ohrani vrstni red prvega pojava
        _dup_count = 0
        for sku, naziv, kol_raw in parsed_rows:
            sku = (sku or "").strip()
            if not sku:
                continue
            k = _parse_kol(kol_raw)
            if sku not in _merged:
                _merged[sku] = {"naziv": (naziv or "").strip(), "kol": 0}
                _order.append(sku)
            else:
                _dup_count += 1
                # če prvi naziv prazen, vzemi tega
                if not _merged[sku]["naziv"] and naziv:
                    _merged[sku]["naziv"] = naziv.strip()
            _merged[sku]["kol"] += k
        parsed_rows = [(s, _merged[s]["naziv"], _merged[s]["kol"]) for s in _order]

        # naloži obstoječo sejo (ali ustvari novo) — DODAMO k njej
        mk = _zaloga_market(market)
        path = _zaloga_current_path(market)
        if path.exists():
            sess = json.loads(path.read_text(encoding="utf-8"))
        else:
            sess = {
                "started_at": _dt.now().isoformat(),
                "filename": file.filename, "market": mk,
                "items": [], "extra_boxes": {}, "cakajoce": [], "packing_boxes": {},
                "pick_started_at": None, "pick_finished_at": None,
            }
        sess.setdefault("items", [])
        sess.setdefault("cakajoce", [])
        sess.setdefault("packing_boxes", {})

        # naslednji prosti idx (čez items + cakajoce, da so unikatni)
        max_idx = -1
        for it in sess["items"]:
            max_idx = max(max_idx, int(it.get("idx", -1)))
        for c in sess["cakajoce"]:
            max_idx = max(max_idx, int(c.get("idx", -1)))
        next_idx = max_idx + 1

        added_police = 0
        added_cakajoce = 0
        groups_added = {}

        for sku, naziv, kol_raw in parsed_rows:
            if not sku:
                continue
            kol = _parse_kol(kol_raw)
            if kol <= 0:
                continue
            grp = _zaloga_import_group(sku)
            if kol == 1:
                # na polico
                sess["items"].append({
                    "idx": next_idx, "id": "", "sku": sku, "naziv": naziv,
                    "qty": 1, "poz": "Ni podatka", "group": grp,
                    "status": "", "picked": 1, "low": False,
                    "box": "", "locked": False, "opomba": "",
                })
                added_police += 1
                groups_added[grp] = groups_added.get(grp, 0) + 1
            else:
                # količina > 1 → čakajoče (razdelitev v bokse)
                sess["cakajoce"].append({
                    "idx": next_idx, "sku": sku, "naziv": naziv,
                    "qty": kol, "poz": grp, "assigned": 0, "done": False,
                })
                added_cakajoce += 1
            next_idx += 1

        sess["updated_at"] = _dt.now().isoformat()
        # HS PLUS: prenesi aktivne oznake (5-dnevno okno) na vse postavke seje po SKU
        _hsplus_mark_items(market, sess.get("items", []))
        # uvoz doda nove postavke → nabiranje se začne na novo: resetiraj časovnico,
        # da šteje od PRVE nabrane postavke (0:00:01), ne od starega ostanka
        sess["started_at"] = _dt.now().isoformat()
        sess["pick_started_at"] = None
        sess["pick_finished_at"] = None
        # BASELINE: uvoz dobavnice doda postavke/kose → peak legitimno naraste
        _zaloga_update_peak(sess)
        _zaloga_atomic_write(path, sess)
        return {"ok": True, "added_police": added_police, "added_cakajoce": added_cakajoce,
                "groups": groups_added, "total": added_police + added_cakajoce,
                "merged_dups": _dup_count, "unique_skus": len(_order)}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/zaloga-current")
async def zaloga_current_get(market: str = "slo"):
    """Vrne aktivno sejo nabiranja za trg (vsi nabiralci berejo isto)."""
    try:
        path = _zaloga_current_path(market)
        if path.exists():
            data = json.loads(path.read_text(encoding="utf-8"))
            # SAMO-POPRAVILO STARIH SEJ: časovnica mora izhajati iz dejanskih znamk
            # nabiranja (picked_at), ne iz starega pick_started_at (ki je bil čas uploada).
            its = data.get("items", [])
            stamps = [it.get("picked_at") for it in its if it.get("picked_at")]
            if stamps:
                real_start = min(stamps)
            else:
                # seja še nima nobene prave znamke → časovnica naj kaže 0:00:00,
                # tudi če stari pick_started_at obstaja (bil je čas uploada)
                real_start = None
            if data.get("pick_started_at") != real_start:
                data["pick_started_at"] = real_start
                if real_start is None:
                    data["pick_finished_at"] = None
                try:
                    _zaloga_atomic_write(path, data)
                except Exception:
                    pass
            # ── VAROVALKA: skupne postavke + kosi ne smejo padati pod baseline (peak) ──
            cur_items, cur_qty = _zaloga_expected_totals(data)
            has_peak = ("peak_items" in data) or ("peak_qty" in data)
            if not has_peak:
                # stara seja brez baseline → inicializiraj na trenutno stanje (brez alarma)
                data["peak_items"] = cur_items
                data["peak_qty"] = cur_qty
                try:
                    _zaloga_atomic_write(path, data)
                except Exception:
                    pass
            peak_items = int(data.get("peak_items", cur_items) or 0)
            peak_qty = int(data.get("peak_qty", cur_qty) or 0)
            dropped_items = max(0, peak_items - cur_items)
            dropped_qty = max(0, peak_qty - cur_qty)
            data["integrity"] = {
                "cur_items": cur_items, "cur_qty": cur_qty,
                "peak_items": peak_items, "peak_qty": peak_qty,
                "dropped_items": dropped_items, "dropped_qty": dropped_qty,
                "ok": (dropped_items == 0 and dropped_qty == 0),
            }
            # SKLADIŠČE: ponovno označi postavke po sejnem seznamu SKU (preživi osvežitev/sync)
            if data.get("skladisce_skus"):
                _skladisce_mark_items(data)
            return {"ok": True, **data}
        return {"ok": True, "items": [], "started_at": None}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/zaloga-hsplus-upload")
async def zaloga_hsplus_upload(file: UploadFile = File(...), market: str = "slo"):
    """Uvoz 'HS PLUS' seznama (XLSX/CSV s stolpcema sku, stock).
    Označi OBSTOJEČE postavke v seji (ujemanje po SKU) z vizualno značko HS PLUS.
    Pomeni: izdelek danes pride k nam (v sistemu je, fizično še ne) — nabiralci naj
    ga NE označijo kot 'ni zaloge'. Samo oznaka, nič drugega."""
    try:
        import io as _io
        from datetime import datetime as _dt
        raw = await file.read()
        fname = (file.filename or "").lower()

        hs_skus = {}
        if fname.endswith(".xlsx") or fname.endswith(".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                header = next(rows_iter, None)
                # poišči indeks stolpca sku + stock
                sku_i, stock_i = 0, 1
                if header:
                    for i, h in enumerate(header):
                        hl = str(h or "").strip().lower()
                        if hl == "sku": sku_i = i
                        elif hl in ("stock", "kolicina", "količina", "qty"): stock_i = i
                for r in rows_iter:
                    if not r or len(r) <= sku_i:
                        continue
                    s = str(r[sku_i] or "").strip()
                    if s:
                        q = 0
                        if stock_i is not None and len(r) > stock_i:
                            try: q = int(float(str(r[stock_i]).replace(",", ".")))
                            except (ValueError, TypeError): q = 0
                        hs_skus[s] = q
            except Exception as ex:
                return {"ok": False, "error": f"Excel branje ni uspelo: {ex}"}
        else:
            import csv as _csv
            text = raw.decode("utf-8-sig", errors="replace")
            delim = ";" if text.count(";") >= text.count(",") else ","
            reader = _csv.DictReader(_io.StringIO(text), delimiter=delim)
            for row in reader:
                sval, qval = "", 0
                for k, v in row.items():
                    kl = str(k or "").strip().lower()
                    if kl == "sku":
                        sval = str(v or "").strip()
                    elif kl in ("stock", "kolicina", "količina", "qty"):
                        try: qval = int(float(str(v or "0").replace(",", ".")))
                        except (ValueError, TypeError): qval = 0
                if sval:
                    hs_skus[sval] = qval

        if not hs_skus:
            return {"ok": False, "error": "V datoteki ni najdenih SKU-jev (stolpec 'sku')"}

        path = _zaloga_current_path(market)
        if not path.exists():
            return {"ok": False, "error": "Ni aktivne seje — najprej naloži seznam za nabiranje"}
        sess = json.loads(path.read_text(encoding="utf-8"))
        items = sess.get("items", [])

        matched = 0
        for it in items:
            sku = str(it.get("sku", "")).strip()
            if sku in hs_skus:
                it["hsplus"] = True
                it["hsplus_qty"] = hs_skus[sku]   # količina ki danes pride
                matched += 1
            # ne brišemo obstoječih oznak, ki niso v tem seznamu — pusti kot so

        # trajna shramba: oznaka velja 5 koledarskih dni (čez seje/arhive)
        _hsplus_save_active(market, list(hs_skus.keys()))

        sess["updated_at"] = _dt.now().isoformat()
        _zaloga_atomic_write(path, sess)
        return {"ok": True, "hs_count": len(hs_skus), "matched": matched,
                "unmatched": len(hs_skus) - matched}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/zaloga-debug")
async def zaloga_debug():
    """Diagnostika: pokaže vsebino /data/zaloga (ali current seje obstajajo, velikost, čas,
    koliko postavk). Za ugotavljanje, ali so podatki še na disku."""
    try:
        from datetime import datetime as _dtd
        info = {"ZALOGA_DIR": str(ZALOGA_DIR), "exists": ZALOGA_DIR.exists(), "files": []}
        if ZALOGA_DIR.exists():
            for f in sorted(ZALOGA_DIR.iterdir()):
                entry = {"name": f.name, "is_dir": f.is_dir()}
                if f.is_file():
                    entry["size_kb"] = round(f.stat().st_size / 1024, 1)
                    entry["modified"] = _dtd.fromtimestamp(f.stat().st_mtime).isoformat()
                    # če je current seja, preštej postavke
                    if f.name.startswith("current"):
                        try:
                            d = json.loads(f.read_text(encoding="utf-8"))
                            entry["items"] = len(d.get("items", []))
                            entry["cakajoce"] = len(d.get("cakajoce", []))
                            entry["filename"] = d.get("filename", "")
                            entry["started_at"] = d.get("started_at", "")
                            entry["pick_started_at"] = d.get("pick_started_at", "")
                        except Exception as ex:
                            entry["read_error"] = str(ex)
                info["files"].append(entry)
        # arhiv
        arch = ZALOGA_DIR / "archive"
        if arch.exists():
            info["archive_files"] = []
            for f in sorted(arch.iterdir()):
                if f.is_file():
                    info["archive_files"].append({
                        "name": f.name, "size_kb": round(f.stat().st_size / 1024, 1),
                        "modified": _dtd.fromtimestamp(f.stat().st_mtime).isoformat(),
                    })
                elif f.is_dir():
                    for ff in sorted(f.iterdir()):
                        info["archive_files"].append({
                            "name": f"{f.name}/{ff.name}", "size_kb": round(ff.stat().st_size / 1024, 1),
                            "modified": _dtd.fromtimestamp(ff.stat().st_mtime).isoformat(),
                        })
        return info
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}


def _skladisce_mark_items(sess: dict):
    """Označi postavke seje s 'skladisce', če je njihov SKU na sejnem seznamu.
    Seznam je skupen za vse postavke trenutne seje (SLO in RS berеta svojo sejo)."""
    skus = set(str(s).strip().upper() for s in sess.get("skladisce_skus", []) if str(s).strip())
    n = 0
    for it in sess.get("items", []):
        if str(it.get("sku", "")).strip().upper() in skus:
            it["skladisce"] = True
            n += 1
        elif it.get("skladisce"):
            it["skladisce"] = False  # umaknjen s seznama → odstrani oznako
    return n


def _parse_sku_list(raw: str):
    """Razčleni prilepljen seznam SKU (vrstice, vejice, podpičja, presledki, tabi)."""
    import re as _re
    if not raw:
        return []
    parts = _re.split(r"[\s,;]+", str(raw).strip())
    seen = []
    for p in parts:
        p = p.strip().upper()
        if p and p not in seen:
            seen.append(p)
    return seen


@app.post("/zaloga-skladisce")
async def zaloga_skladisce(data: dict):
    """Doda/odstrani SKU-je za tag 'Skladišče' v TRENUTNI seji (skupno SLO+RS prek seje).
    action: 'add' (dopolni seznam) | 'clear' (počisti vse) | 'remove' (odstrani en SKU).
    Vrne posodobljen seznam + število označenih postavk."""
    try:
        path = _zaloga_current_path(data.get("market", "slo"))
        if not path.exists():
            return {"ok": False, "error": "Ni aktivne seje"}
        sess = json.loads(path.read_text(encoding="utf-8"))
        cur = list(sess.get("skladisce_skus", []))
        action = data.get("action", "add")

        if action == "clear":
            cur = []
        elif action == "remove":
            rem = str(data.get("sku", "")).strip().upper()
            cur = [s for s in cur if str(s).strip().upper() != rem]
        else:  # add — dopolni seznam (brez dvojnikov)
            new = _parse_sku_list(data.get("raw", ""))
            existing = set(str(s).strip().upper() for s in cur)
            for s in new:
                if s not in existing:
                    cur.append(s)
                    existing.add(s)

        sess["skladisce_skus"] = cur
        n = _skladisce_mark_items(sess)
        from datetime import datetime as _dt
        sess["updated_at"] = _dt.now().isoformat()
        _zaloga_atomic_write(path, sess)
        # koliko jih ni najdenih v seji (za info)
        sess_skus = set(str(it.get("sku", "")).strip().upper() for it in sess.get("items", []))
        ni_najdenih = [s for s in cur if s not in sess_skus]
        return {"ok": True, "skladisce_skus": cur, "marked": n, "ni_najdenih": ni_najdenih}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _parse_iso_utc(s):
    """Pretvori ISO žig v UTC timestamp (sekunde). Brez cone → razumi kot UTC (Render)."""
    from datetime import datetime as _dt2
    if not s:
        return 0.0
    try:
        import re as _re2
        has_tz = bool(_re2.search(r'[zZ]$|[+-]\d{2}:?\d{2}$', s))
        dt = _dt2.fromisoformat(s.replace('Z', '+00:00')) if has_tz else _dt2.fromisoformat(s).replace(tzinfo=timezone.utc)
        return dt.timestamp()
    except Exception:
        return 0.0

def _dt_now_utc_ts():
    from datetime import datetime as _dt2
    return _dt2.now(timezone.utc).timestamp()


@app.post("/zaloga-update-item")
async def zaloga_update_item(data: dict):
    """Posodobi eno postavko (status/picked) — kliče se LIVE ob vsaki spremembi.
    Več nabiralcev hkrati: zaklepamo z atomarnim read-modify-write na idx."""
    try:
        from datetime import datetime as _dt
        path = _zaloga_current_path(data.get("market", "slo"))
        if not path.exists():
            return {"ok": False, "error": "Ni aktivne seje"}
        idx = data.get("idx")
        if idx is None:
            return {"ok": False, "error": "Manjka idx"}
        sess = json.loads(path.read_text(encoding="utf-8"))
        now_iso = _dt.now(timezone.utc).isoformat()
        found = False
        for it in sess.get("items", []):
            if it.get("idx") == idx:
                if "status" in data:
                    prev_status = it.get("status")
                    it["status"] = data["status"]
                    # ČASOVNICA: zabeleži DEJANSKI trenutek nabiranja (samo ob prvem prehodu
                    # iz nepotrjeno → ok/ni). To je edini vir resnice za začetek časovnice —
                    # nikoli čas uploada ali predhodno naložen status.
                    if data["status"] in ("ok", "ni") and prev_status not in ("ok", "ni") \
                            and not it.get("picked_at"):
                        it["picked_at"] = now_iso
                    # preklic statusa → pobriši pick znamko (da časovnica spet pravilno reagira)
                    elif data["status"] not in ("ok", "ni"):
                        it["picked_at"] = None
                if "picked" in data:
                    try:
                        it["picked"] = max(0, int(data["picked"]))
                    except (ValueError, TypeError):
                        pass
                if "opomba" in data:
                    it["opomba"] = str(data["opomba"]).strip()
                found = True
                break
        if not found:
            return {"ok": False, "error": "Postavka ne obstaja"}
        sess["updated_at"] = now_iso

        # ── ČASOVNICA NABIRANJA ──
        # Začetek = DEJANSKI čas najzgodnejše nabrane postavke (it["picked_at"]).
        # Konec = ko so VSE obdelane (100%). Nikoli se ne veže na čas uploada.
        items_all = sess.get("items", [])
        total = len(items_all)
        picked_ats = [it.get("picked_at") for it in items_all if it.get("picked_at")]
        obdelanih = sum(1 for it in items_all if it.get("status") in ("ok", "ni"))
        # vir resnice: najzgodnejša dejanska znamka nabiranja (ali None, če še nič)
        sess["pick_started_at"] = min(picked_ats) if picked_ats else None

        # ── PAVZA: če je bila štoparica pavzirana in nabiralec doda/spremeni postavko,
        # pavzo prekinemo — pretečeni pavzni čas pretvorimo v trajni offset, čas teče naprej. ──
        if sess.get("pick_paused_at") and data.get("status") in ("ok", "ni"):
            try:
                paused_ms = _parse_iso_utc(sess["pick_paused_at"])
                now_ms = _dt.now(timezone.utc).timestamp()
                gap = max(0, now_ms - paused_ms)
                sess["pick_pause_offset_s"] = round(sess.get("pick_pause_offset_s", 0) + gap, 3)
            except Exception:
                pass
            sess["pick_paused_at"] = None
        # konec: vse obdelano → zabeleži končni čas (samo prvič)
        if total > 0 and obdelanih >= total and sess.get("pick_started_at"):
            if not sess.get("pick_finished_at"):
                sess["pick_finished_at"] = now_iso
        else:
            # padlo pod 100% (npr. preklic statusa) → časovnica spet teče
            if sess.get("pick_finished_at"):
                sess["pick_finished_at"] = None

        _zaloga_atomic_write(path, sess)
        return {"ok": True,
                "pick_started_at": sess.get("pick_started_at"),
                "pick_finished_at": sess.get("pick_finished_at"),
                "pick_paused_at": sess.get("pick_paused_at"),
                "pick_pause_offset_s": sess.get("pick_pause_offset_s", 0)}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.post("/zaloga-lock-box")
async def zaloga_lock_box(data: dict):
    """RS: zakleni vse obkljukane (status=ok) IN še ne zaklenjene postavke v polici → dodeli box.
    Ali odkleni eno postavko (unlock=idx)."""
    try:
        from datetime import datetime as _dt
        path = _zaloga_current_path(data.get("market", "rs"))
        if not path.exists():
            return {"ok": False, "error": "Ni aktivne seje"}
        sess = json.loads(path.read_text(encoding="utf-8"))

        # Odklep ene postavke
        if "unlock_idx" in data:
            uidx = data["unlock_idx"]
            for it in sess.get("items", []):
                if it.get("idx") == uidx:
                    it["locked"] = False
                    it["box"] = ""
                    break
            sess["updated_at"] = _dt.now().isoformat()
            _zaloga_atomic_write(path, sess)
            return {"ok": True, "unlocked": uidx}

        # Zaklep obkljukanih — globalno (vse police) ali samo ena polica
        is_global = bool(data.get("global"))
        group = data.get("group")
        box = str(data.get("box", "")).strip()
        if not box:
            return {"ok": False, "error": "Manjka št. boxa"}
        if not is_global and not group:
            return {"ok": False, "error": "Manjka group"}
        locked_count = 0
        for it in sess.get("items", []):
            if it.get("status") == "ok" and not it.get("locked"):
                if is_global or it.get("group") == group:
                    it["box"] = box
                    it["locked"] = True
                    locked_count += 1
        if locked_count == 0:
            return {"ok": False, "error": "Ni obkljukanih (in še odklenjenih) postavk"}
        sess["updated_at"] = _dt.now().isoformat()
        _zaloga_atomic_write(path, sess)
        return {"ok": True, "locked": locked_count, "box": box}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.post("/zaloga-extra-box")
async def zaloga_extra_box(data: dict):
    """RS dodatni boxi (viški). Akcije:
    - add: doda postavko (sku, kos) v box (ustvari box če ne obstaja)
    - remove_item: odstrani postavko iz boxa
    - delete_box: izbriše cel box
    Vrne posodobljen extra_boxes."""
    try:
        from datetime import datetime as _dt
        path = _zaloga_current_path(data.get("market", "rs"))
        if not path.exists():
            return {"ok": False, "error": "Ni aktivne seje"}
        sess = json.loads(path.read_text(encoding="utf-8"))
        boxes = sess.get("extra_boxes") or {}
        action = data.get("action", "add")

        if action == "add":
            box = str(data.get("box", "")).strip()
            sku = str(data.get("sku", "")).strip()
            naziv = str(data.get("naziv", "")).strip()
            idx = data.get("idx")
            try:
                kos = max(1, int(data.get("kos", 1)))
            except (ValueError, TypeError):
                kos = 1
            if not box:
                return {"ok": False, "error": "Manjka št. boxa"}
            if not sku:
                return {"ok": False, "error": "Manjka SKU"}
            lst = boxes.get(box, [])
            # če isti sku že v boxu, prištej kose
            merged = False
            for entry in lst:
                if entry.get("sku") == sku:
                    entry["kos"] = entry.get("kos", 0) + kos
                    merged = True
                    break
            if not merged:
                lst.append({"idx": idx, "sku": sku, "naziv": naziv, "kos": kos})
            boxes[box] = lst

            # Označi postavko nabrano SAMO če je cel kos (kos >= qty) v box
            if idx is not None:
                for it in sess.get("items", []):
                    if it.get("idx") == idx:
                        try:
                            need = int(it.get("qty", 0) or 0)
                        except (ValueError, TypeError):
                            need = 0
                        # seštej vse kose tega sku po vseh extra boxih
                        total_in_boxes = 0
                        for bx_items in boxes.values():
                            for e in bx_items:
                                if e.get("sku") == it.get("sku"):
                                    total_in_boxes += e.get("kos", 0)
                        if need and total_in_boxes >= need:
                            it["status"] = "ok"
                        break

        elif action == "remove_item":
            box = str(data.get("box", "")).strip()
            sku = str(data.get("sku", "")).strip()
            if box in boxes:
                boxes[box] = [e for e in boxes[box] if e.get("sku") != sku]
                if not boxes[box]:
                    del boxes[box]

        elif action == "delete_box":
            box = str(data.get("box", "")).strip()
            if box in boxes:
                del boxes[box]

        elif action == "rename_box":
            old = str(data.get("box", "")).strip()
            new = str(data.get("new_box", "")).strip()
            if not new:
                return {"ok": False, "error": "Manjka nova št. boxa"}
            if old not in boxes:
                return {"ok": False, "error": "Box ne obstaja"}
            if new == old:
                pass  # nič za narediti
            elif new in boxes:
                # nova št. že obstaja → zlij postavke (isti sku združi kose)
                target = boxes[new]
                for entry in boxes[old]:
                    merged = False
                    for t in target:
                        if t.get("sku") == entry.get("sku"):
                            t["kos"] = t.get("kos", 0) + entry.get("kos", 0)
                            merged = True
                            break
                    if not merged:
                        target.append(entry)
                del boxes[old]
            else:
                boxes[new] = boxes.pop(old)

        sess["extra_boxes"] = boxes
        sess["updated_at"] = _dt.now().isoformat()
        _zaloga_atomic_write(path, sess)
        return {"ok": True, "extra_boxes": boxes}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.post("/zaloga-cakajoce")
async def zaloga_cakajoce(data: dict):
    """RS 'Čakajoče' — velike postavke za razdelitev v več packing boxov (carinska lista).
    Akcije:
    - transfer: prenese postavko iz police v čakajoče (izgine iz police)
    - return: vrne postavko nazaj v polico (iz čakajočih)
    - assign: dodeli N kosov postavke v packing box (box = posoda za več izdelkov)
    - remove_assign: odstrani dodelitev (sku iz packing boxa)
    - delete_pbox: izbriše cel packing box (vse dodelitve)
    Struktura v seji:
      cakajoce: [{idx, sku, naziv, qty, poz}]
      packing_boxes: { "20": [{sku, naziv, kos}], "21": [...] }
    Postavka šteje kot nabrana (status ok), ko je vseh qty kosov dodeljenih v bokse."""
    try:
        from datetime import datetime as _dt
        path = _zaloga_current_path(data.get("market", "rs"))
        if not path.exists():
            return {"ok": False, "error": "Ni aktivne seje"}
        sess = json.loads(path.read_text(encoding="utf-8"))
        cakajoce = sess.get("cakajoce") or []
        pboxes = sess.get("packing_boxes") or {}
        action = data.get("action", "")

        def _assigned_for(sku):
            """Koliko kosov tega SKU je že dodeljenih v packing bokse."""
            tot = 0
            for items in pboxes.values():
                for e in items:
                    if e.get("sku") == sku:
                        tot += e.get("kos", 0)
            return tot

        def _sync_item_status(idx, sku):
            """Postavka v cakajoce: nabrana (ok) ko so vsi kosi dodeljeni."""
            for c in cakajoce:
                if c.get("idx") == idx:
                    need = int(c.get("qty", 0) or 0)
                    c["assigned"] = _assigned_for(sku)
                    c["done"] = bool(need and c["assigned"] >= need)
                    return c
            return None

        if action == "transfer":
            idx = data.get("idx")
            if idx is None:
                return {"ok": False, "error": "Manjka idx"}
            # poišči postavko v items, jo prestavi v cakajoce in odstrani iz police
            moved = None
            for it in sess.get("items", []):
                if it.get("idx") == idx:
                    moved = {
                        "idx": it.get("idx"), "sku": it.get("sku", ""),
                        "naziv": it.get("naziv", ""), "qty": int(it.get("qty", 0) or 0),
                        "poz": it.get("poz", ""), "assigned": 0, "done": False,
                    }
                    break
            if not moved:
                return {"ok": False, "error": "Postavka ni najdena"}
            # ni dvojnikov
            if not any(c.get("idx") == idx for c in cakajoce):
                cakajoce.append(moved)
            # odstrani iz police (items)
            sess["items"] = [it for it in sess.get("items", []) if it.get("idx") != idx]

        elif action == "return":
            idx = data.get("idx")
            entry = next((c for c in cakajoce if c.get("idx") == idx), None)
            if not entry:
                return {"ok": False, "error": "Ni v čakajočih"}
            # vrni v polico kot todo postavko — group MORA biti nastavljen (sicer pristane v "undefined")
            _poz = entry.get("poz", "")
            _sku = entry.get("sku", "")
            sess.setdefault("items", []).append({
                "idx": entry["idx"], "sku": _sku, "naziv": entry["naziv"],
                "qty": entry["qty"], "picked": 0, "status": "", "poz": _poz,
                "group": _zaloga_group(_poz, _sku),   # ← izračunaj polico nazaj iz pozicije/SKU
                "locked": False, "box": "", "low": False, "opomba": "",
            })
            cakajoce = [c for c in cakajoce if c.get("idx") != idx]
            # počisti dodelitve tega sku iz packing boxov
            for b in list(pboxes.keys()):
                pboxes[b] = [e for e in pboxes[b] if e.get("sku") != entry["sku"]]
                if not pboxes[b]:
                    del pboxes[b]

        elif action == "assign":
            idx = data.get("idx")
            box = str(data.get("box", "")).strip()
            sku = str(data.get("sku", "")).strip()
            naziv = str(data.get("naziv", "")).strip()
            try:
                kos = max(1, int(data.get("kos", 1)))
            except (ValueError, TypeError):
                kos = 1
            if not box or not sku:
                return {"ok": False, "error": "Manjka box ali SKU"}
            lst = pboxes.get(box, [])
            merged = False
            for e in lst:
                if e.get("sku") == sku:
                    e["kos"] = e.get("kos", 0) + kos
                    merged = True
                    break
            if not merged:
                lst.append({"sku": sku, "naziv": naziv, "kos": kos})
            pboxes[box] = lst
            sess["packing_boxes"] = pboxes
            _sync_item_status(idx, sku)

        elif action == "assign_bulk":
            # Razdeli X kosov v VEČ boxov naenkrat: po kos_per_box v vsak zaporedni box,
            # od start_box naprej. Ostanek (če ni deljivo) ostane nerazdeljen.
            idx = data.get("idx")
            sku = str(data.get("sku", "")).strip()
            naziv = str(data.get("naziv", "")).strip()
            try:
                kos_per_box = max(1, int(data.get("kos_per_box", 1)))
            except (ValueError, TypeError):
                return {"ok": False, "error": "Neveljavna količina na box"}
            try:
                start_box = int(str(data.get("start_box", "")).strip())
            except (ValueError, TypeError):
                return {"ok": False, "error": "Neveljavna začetna št. boxa"}
            if not sku:
                return {"ok": False, "error": "Manjka SKU"}
            # koliko kosov je še nerazdeljenih za to postavko
            entry = next((c for c in cakajoce if c.get("idx") == idx), None)
            if not entry:
                return {"ok": False, "error": "Ni v čakajočih"}
            need = int(entry.get("qty", 0) or 0)
            already = _assigned_for(sku)
            remaining = need - already
            if remaining <= 0:
                return {"ok": False, "error": "Ni več kosov za razdeliti"}
            n_boxes = remaining // kos_per_box   # cele škatle; ostanek ostane
            if n_boxes <= 0:
                return {"ok": False, "error": f"Premalo kosov ({remaining}) za en box po {kos_per_box}"}
            box_num = start_box
            created = []
            for _ in range(n_boxes):
                b = str(box_num)
                lst = pboxes.get(b, [])
                merged = False
                for e in lst:
                    if e.get("sku") == sku:
                        e["kos"] = e.get("kos", 0) + kos_per_box
                        merged = True
                        break
                if not merged:
                    lst.append({"sku": sku, "naziv": naziv, "kos": kos_per_box})
                pboxes[b] = lst
                created.append(b)
                box_num += 1
            sess["packing_boxes"] = pboxes
            _sync_item_status(idx, sku)
            ostanek = remaining - (n_boxes * kos_per_box)
            sess["cakajoce"] = cakajoce
            sess["packing_boxes"] = pboxes
            sess["updated_at"] = _dt.now().isoformat()
            _zaloga_atomic_write(path, sess)
            return {"ok": True, "cakajoce": cakajoce, "packing_boxes": pboxes,
                    "created_boxes": created, "ostanek": ostanek}

        elif action == "remove_assign":
            box = str(data.get("box", "")).strip()
            sku = str(data.get("sku", "")).strip()
            idx = data.get("idx")
            if box in pboxes:
                pboxes[box] = [e for e in pboxes[box] if e.get("sku") != sku]
                if not pboxes[box]:
                    del pboxes[box]
            sess["packing_boxes"] = pboxes
            if idx is not None:
                _sync_item_status(idx, sku)

        elif action == "close_missing":
            # Zaključi čakajočo z manjkom: razdeljeni kosi ostanejo v boxih,
            # preostanek (qty - assigned) se šteje kot manjko. Postavka postane done.
            idx = data.get("idx")
            sku = str(data.get("sku", "")).strip()
            entry = next((c for c in cakajoce if c.get("idx") == idx), None)
            if not entry:
                return {"ok": False, "error": "Ni v čakajočih"}
            need = int(entry.get("qty", 0) or 0)
            assigned = _assigned_for(sku)
            entry["assigned"] = assigned
            entry["done"] = True
            entry["closed_missing"] = max(0, need - assigned)

        elif action == "delete_pbox":
            box = str(data.get("box", "")).strip()
            if box in pboxes:
                del pboxes[box]
            sess["packing_boxes"] = pboxes
            # osveži vse statuse
            for c in cakajoce:
                _sync_item_status(c.get("idx"), c.get("sku"))
        else:
            return {"ok": False, "error": "Neznana akcija"}

        sess["cakajoce"] = cakajoce
        sess["packing_boxes"] = pboxes
        sess["updated_at"] = _dt.now().isoformat()
        _zaloga_atomic_write(path, sess)
        return {"ok": True, "cakajoce": cakajoce, "packing_boxes": pboxes}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.post("/zaloga-archive")
async def zaloga_archive(data: dict = None):
    """Arhivira aktivno sejo nabiranja za trg in resetira."""
    try:
        from datetime import datetime as _dt
        market = _zaloga_market((data or {}).get("market", "slo"))
        path = _zaloga_current_path(market)
        adir = _zaloga_archive_dir(market)
        if not path.exists():
            return {"ok": False, "error": "Ni aktivne seje"}
        sess = json.loads(path.read_text(encoding="utf-8"))
        items = sess.get("items", [])
        if not items:
            path.unlink()
            return {"ok": True, "message": "Prazna seja izbrisana"}
        # RS: opozori če so nabrane (ok) postavke brez dodeljenega boxa — razen če force
        if market == "rs" and not (data or {}).get("force"):
            no_box = sum(1 for it in items if it.get("status") == "ok" and not it.get("box"))
            if no_box > 0:
                return {"ok": False, "warn_no_box": no_box,
                        "error": f"{no_box} nabranih postavk nima dodeljenega boxa"}
        ts = _dt.now().strftime("%Y%m%d_%H%M%S")
        archive_name = f"{ts}_{len(items)}items.json"
        sess["archived_at"] = _dt.now().isoformat()
        (adir / archive_name).write_text(json.dumps(sess, ensure_ascii=False), encoding="utf-8")
        path.unlink()
        return {"ok": True, "archived": archive_name, "items": len(items)}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/zaloga-history")
async def zaloga_history(market: str = "slo"):
    """Seznam arhiviranih sej nabiranja za trg (povzetki)."""
    try:
        adir = _zaloga_archive_dir(market)
        out = []
        for f in sorted(adir.glob("*.json"), reverse=True):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                items = data.get("items", [])
                total = len(items)
                ok = sum(1 for it in items if it.get("status") == "ok")
                ni = sum(1 for it in items if it.get("status") == "ni")
                qty_need = sum(int(it.get("qty", 0) or 0) for it in items)
                qty_picked = sum(int(it.get("picked", 0) or 0) for it in items if it.get("status") == "ok")
                # trajanje nabiranja (sekunde) iz časovnice
                pick_secs = None
                ps = data.get("pick_started_at")
                pf = data.get("pick_finished_at") or data.get("archived_at")
                if ps and pf:
                    try:
                        from datetime import datetime as _dt2
                        pick_secs = max(0, int((_dt2.fromisoformat(pf) - _dt2.fromisoformat(ps)).total_seconds()))
                    except Exception:
                        pick_secs = None
                out.append({
                    "filename": f.name,
                    "archived_at": data.get("archived_at"),
                    "total": total, "ok": ok, "ni": ni,
                    "qty_need": qty_need, "qty_picked": qty_picked,
                    "pct": round((ok + ni) / total * 100) if total else 0,
                    "pick_started_at": data.get("pick_started_at"),
                    "pick_finished_at": data.get("pick_finished_at"),
                    "pick_secs": pick_secs,
                })
            except Exception:
                continue
        return {"ok": True, "sessions": out}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/zaloga-history/{filename}")
async def zaloga_history_detail(filename: str, market: str = "slo"):
    """Polna vsebina ene arhivirane seje."""
    try:
        if "/" in filename or "\\" in filename or ".." in filename:
            return {"ok": False, "error": "neveljavno ime"}
        f = _zaloga_archive_dir(market) / filename
        if not f.exists():
            return {"ok": False, "error": "ni najdeno"}
        data = json.loads(f.read_text(encoding="utf-8"))
        return {"ok": True, "filename": filename, **data}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.delete("/zaloga-history/{filename}")
async def zaloga_history_delete(filename: str, market: str = "slo"):
    """Izbriše arhivirano sejo."""
    try:
        if "/" in filename or "\\" in filename or ".." in filename:
            return {"ok": False, "error": "neveljavno ime"}
        f = _zaloga_archive_dir(market) / filename
        if not f.exists():
            return {"ok": False, "error": "ni najdeno"}
        f.unlink()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/zaloga", response_class=HTMLResponse)
def zaloga_page():
    return FileResponse("static/zaloga.html")


# ═══ OBVESTILA (in-app "kaj je novega") + /admin ═══
NOTICES_FILE = DATA_DIR / "notices.json"

def _notices_load():
    if NOTICES_FILE.exists():
        try:
            return json.loads(NOTICES_FILE.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []

def _notices_write(items):
    NOTICES_FILE.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")

@app.get("/notices")
async def notices_get(scope: str = "all"):
    """Vrne objavljena obvestila. scope: 'zaloga' | 'index' | 'all'.
    Vrne obvestila, ki ciljajo na ta scope (ali 'both')."""
    items = _notices_load()
    if scope and scope != "all":
        items = [n for n in items if n.get("scope") in (scope, "both")]
    # najnovejša najprej
    items = sorted(items, key=lambda n: n.get("created_at", ""), reverse=True)
    return {"ok": True, "notices": items}

@app.post("/notices")
async def notices_save(data: dict):
    """Admin: dodaj ali izbriši obvestilo. action: 'add' | 'remove'."""
    try:
        items = _notices_load()
        action = data.get("action", "add")
        if action == "remove":
            nid = (data.get("id") or "").strip()
            items = [n for n in items if n.get("id") != nid]
            _notices_write(items)
            return {"ok": True, "notices": items}
        # add
        n = data.get("notice") or {}
        title = (n.get("title") or "").strip()
        body = (n.get("body") or "").strip()
        if not title and not body:
            return {"ok": False, "error": "Manjka naslov ali besedilo."}
        scope = n.get("scope") or "zaloga"
        if scope not in ("zaloga", "index", "both"):
            scope = "zaloga"
        icon = (n.get("icon") or "📢").strip()[:4] or "📢"
        nid = "n" + str(int(datetime.now().timestamp())) + "".join(c for c in title[:8] if c.isalnum())
        items.append({
            "id": nid, "icon": icon, "title": title, "body": body,
            "scope": scope, "created_at": _lj_now().isoformat(),
            "date": _lj_now().strftime("%d.%m.%Y %H:%M"),
        })
        _notices_write(items)
        return {"ok": True, "id": nid, "notices": items}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}

@app.get("/admin", response_class=HTMLResponse)
def admin_page():
    return FileResponse("static/admin.html")


@app.get("/admin-pick-timer-status")
async def admin_pick_timer_status(market: str = "slo"):
    """Vrne stanje štoparice nabiranja (za admin nadzor)."""
    try:
        path = _zaloga_current_path(market)
        if not path.exists():
            return {"ok": False, "error": "Ni aktivne seje"}
        sess = json.loads(path.read_text(encoding="utf-8"))
        start = sess.get("pick_started_at")
        paused = sess.get("pick_paused_at")
        offset = sess.get("pick_pause_offset_s", 0)
        # izračunaj trenutni prikazani čas
        elapsed = 0
        if start:
            start_ms = _parse_iso_utc(start)
            ref_ms = _parse_iso_utc(paused) if paused else _dt_now_utc_ts()
            elapsed = max(0, ref_ms - start_ms - offset)
        return {"ok": True, "market": market, "pick_started_at": start,
                "pick_finished_at": sess.get("pick_finished_at"),
                "pick_paused_at": paused, "pick_pause_offset_s": offset,
                "elapsed_s": round(elapsed), "is_paused": bool(paused)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/admin-pick-timer-control")
async def admin_pick_timer_control(data: dict):
    """Admin nadzor štoparice nabiranja:
      action=pause   → zamrzni štoparico na trenutnem času
      action=resume  → nadaljuj (pavzni čas → offset)
      action=set     → nastavi prikazani čas na seconds=N (in pavziraj)
    Ko nabiralec doda postavko, se pavza samodejno prekine (v zaloga-update-item)."""
    try:
        from datetime import datetime as _dt
        market = data.get("market", "slo")
        action = data.get("action", "")
        path = _zaloga_current_path(market)
        if not path.exists():
            return {"ok": False, "error": "Ni aktivne seje"}
        sess = json.loads(path.read_text(encoding="utf-8"))
        now_iso = _dt.now(timezone.utc).isoformat()
        now_ts = _dt_now_utc_ts()

        if action == "pause":
            if not sess.get("pick_paused_at"):
                sess["pick_paused_at"] = now_iso
        elif action == "resume":
            if sess.get("pick_paused_at"):
                gap = max(0, now_ts - _parse_iso_utc(sess["pick_paused_at"]))
                sess["pick_pause_offset_s"] = round(sess.get("pick_pause_offset_s", 0) + gap, 3)
                sess["pick_paused_at"] = None
        elif action == "set":
            try:
                target = max(0, float(data.get("seconds", 0)))
            except (ValueError, TypeError):
                return {"ok": False, "error": "Neveljaven 'seconds'"}
            start = sess.get("pick_started_at")
            if not start:
                # ni začetka — postavi start tako, da bo prikaz = target in pavziraj
                sess["pick_started_at"] = _dt.fromtimestamp(now_ts - target, tz=timezone.utc).isoformat()
                sess["pick_pause_offset_s"] = 0
            else:
                # nastavi offset tako, da prikaz ob TEM trenutku = target, nato pavziraj
                start_ms = _parse_iso_utc(start)
                sess["pick_pause_offset_s"] = round(max(0, (now_ts - start_ms) - target), 3)
            sess["pick_paused_at"] = now_iso  # pavziraj na nastavljenem času
            sess["pick_finished_at"] = None
        else:
            return {"ok": False, "error": "Neznana akcija (pause/resume/set)"}

        _zaloga_atomic_write(path, sess)
        # izračunaj prikaz za odgovor
        start = sess.get("pick_started_at")
        elapsed = 0
        if start:
            ref = _parse_iso_utc(sess["pick_paused_at"]) if sess.get("pick_paused_at") else now_ts
            elapsed = max(0, ref - _parse_iso_utc(start) - sess.get("pick_pause_offset_s", 0))
        return {"ok": True, "action": action, "is_paused": bool(sess.get("pick_paused_at")),
                "elapsed_s": round(elapsed), "pick_pause_offset_s": sess.get("pick_pause_offset_s", 0)}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


# ═══ RVC Maaarket ═══
def _rvc_num(s):
    """Parsira evropski zapis '4.440,69 €' → 4440.69. Vrne None če ni številka."""
    s = (s or "").replace("€", "").replace("\xa0", "").strip()
    if not s:
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _rvc_parse(text: str) -> dict:
    """Parsira prilepljen RVC izpis. Ignorira stolpca Facebook in Google.
    Vrne {markets:[{trg,narocila,postnina,rvc_nar,skupaj}], total:{...}}."""
    import re as _re
    chunks = _re.split(r'(?=Maaarket\.|SKUPAJ)', text or "", flags=_re.IGNORECASE)
    markets = []
    total = None
    for ch in chunks:
        ch = ch.strip()
        if not ch or ch.lower().startswith("trgovina"):
            continue
        toks = [t.strip() for t in _re.split(r'[\t\n]+', ch) if t.strip()]
        if not toks:
            continue
        name = toks[0]
        nums = [_rvc_num(t) for t in toks[1:]]
        is_total = name.upper().startswith("SKUPAJ")
        if is_total:
            # naročila, FB(0), Google(0), poštnina, RVC/nar, skupaj
            orders = nums[0] if len(nums) > 0 else None
            postnina = nums[3] if len(nums) > 3 else None
            rvc_nar = nums[4] if len(nums) > 4 else None
            skupaj = nums[5] if len(nums) > 5 else None
        else:
            # FB+Google sta prazni vrstici (split ju je odstranil) →
            # naročila, poštnina, [RVC/nar], skupaj
            orders = nums[0] if len(nums) > 0 else None
            rest = nums[1:]
            skupaj = rest[-1] if rest else None
            postnina = rest[0] if len(rest) >= 1 else None
            rvc_nar = rest[-2] if len(rest) >= 3 else None
        # če RVC/nar ni podan, izračunaj
        if rvc_nar is None and skupaj and orders:
            rvc_nar = round(skupaj / orders, 2)
        # RVC bruto = skupaj / naročila (brez odbitka poštnine)
        rvc_bruto = round(skupaj / orders, 2) if (skupaj and orders) else None
        row = {
            "trg": name,
            "narocila": int(orders) if orders is not None else 0,
            "postnina": postnina,
            "rvc_nar": rvc_nar,
            "rvc_bruto": rvc_bruto,
            "skupaj": skupaj,
        }
        if is_total:
            total = row
        else:
            markets.append(row)
    return {"markets": markets, "total": total}


@app.post("/rvc-save")
async def rvc_save(data: dict):
    """Parsira in shrani RVC vnos. Trenutni postane 'previous' za primerjavo."""
    try:
        from datetime import datetime as _dt
        text = (data or {}).get("text", "")
        parsed = _rvc_parse(text)
        if not parsed["markets"] and not parsed["total"]:
            return {"ok": False, "error": "ni veljavnih vrstic — preveri format"}
        # preberi obstoječega (postane previous)
        prev = None
        if RVC_FILE.exists():
            try:
                old = json.loads(RVC_FILE.read_text(encoding="utf-8"))
                prev = {
                    "updated_at": old.get("updated_at"),
                    "markets": old.get("markets", []),
                    "total": old.get("total"),
                }
            except Exception:
                prev = None
        out = {
            "updated_at": _dt.now().isoformat(),
            "markets": parsed["markets"],
            "total": parsed["total"],
            "previous": prev,
        }
        RVC_FILE.parent.mkdir(parents=True, exist_ok=True)
        RVC_FILE.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, **out}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/rvc-current")
async def rvc_current():
    """Vrne zadnji shranjeni RVC vnos (z 'previous' za primerjavo)."""
    try:
        if not RVC_FILE.exists():
            return {"ok": True, "markets": [], "total": None, "previous": None, "updated_at": None}
        data = json.loads(RVC_FILE.read_text(encoding="utf-8"))
        return {"ok": True, **data}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/rvc-clear")
async def rvc_clear():
    """Počisti RVC tabelo (za nov dan)."""
    try:
        if RVC_FILE.exists():
            RVC_FILE.unlink()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.delete("/vracila-history/{filename}")
async def vracila_history_delete(filename: str):
    """Briše arhiv."""
    try:
        safe = filename.replace("/", "").replace("..", "")
        f = VRACILA_ARCHIVE_DIR / safe
        if not f.exists():
            return {"ok": False, "error": "Ne obstaja"}
        f.unlink()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/knjigovodstvo-process-emails")
async def knjigovodstvo_process_emails():
    """Sproži ročno obdelavo neprebranih emailov v parcels@.
    Zbere XML/PDF attachmente, razdeli na DU/NU, zazipira vse, shrani v storage."""
    try:
        cfg = _load_email_config()
        if not cfg.get("imap_host") or not cfg.get("email"):
            return {"ok": False, "error": "Email ni konfiguriran (Knjigovodstvo → Email avtomatizacija)"}
        password = _email_get_password(cfg)
        if not password:
            return {"ok": False, "error": "Geslo manjka"}

        # Background obdelava (lahko traja par 10s)
        result = await asyncio.get_event_loop().run_in_executor(
            None, _process_emails_to_batch_zip, cfg, password
        )
        return result
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


def _process_emails_to_batch_zip(cfg: dict, password: str) -> dict:
    """Sinhrona obdelava — teče v executorju.
    1. IMAP fetch UNSEEN
    2. Zberi vse XML + PDF attachmente
    3. Razdeli XML-je na DU + NU (z dedup po številki računa)
    4. Generiraj CSV-je
    5. Zapakiraj vse + PDF-je v 1 ZIP
    6. Shrani v storage
    7. Označi emaile kot prebrane
    """
    import imaplib
    import email as email_lib
    from email.header import decode_header as _dh
    import io
    import zipfile
    import csv
    from datetime import datetime as _dt

    host = cfg["imap_host"]
    port = int(cfg.get("imap_port", 993))
    username = cfg["email"]

    # Naloži zgodovino že obdelanih faktur (cross-batch dedup)
    HISTORY_FILE = STORAGE_KNJ_DIR / "_processed_invoices.json"
    processed_history = {}  # {invoice_key: {batch, ts, email_from}}
    if HISTORY_FILE.exists():
        try:
            processed_history = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        except Exception:
            processed_history = {}

    try:
        if port == 993:
            mail = imaplib.IMAP4_SSL(host, port)
        else:
            mail = imaplib.IMAP4(host, port)
            mail.starttls()
        mail.login(username, password)
    except Exception as e:
        return {"ok": False, "error": f"IMAP login failed: {e}"}

    debit_rows = []
    credit_rows = []
    pdf_files = []  # (filename_in_zip, content_bytes)
    xml_files = []  # (filename, content_str)
    processed_email_nums = []
    email_meta = []  # za log

    # Dedup tracking
    seen_invoices_this_batch = set()       # invoice_key v tem batchu
    skipped_dup_in_batch = []              # [{invoice_num, type, from}]
    skipped_dup_in_history = []            # [{invoice_num, type, from, original_batch}]
    new_invoices_added = []                # [{invoice_num, type}]

    try:
        mail.select("INBOX")
        _, msg_nums = mail.search(None, "UNSEEN")
        if not msg_nums[0]:
            mail.logout()
            return {"ok": True, "message": "Ni novih emailov", "emails": 0, "zip_filename": None}

        for num in msg_nums[0].split():
            try:
                _, msg_data = mail.fetch(num, "(RFC822)")
                raw = msg_data[0][1]
                msg = email_lib.message_from_bytes(raw)

                subj_parts = _dh(msg.get("Subject", ""))
                subject = "".join(p.decode(enc or "utf-8") if isinstance(p, bytes) else p for p, enc in subj_parts)
                sender = msg.get("From", "")
                date_hdr = msg.get("Date", "")

                email_attachments = []
                for part in msg.walk():
                    cd = part.get("Content-Disposition", "")
                    if "attachment" not in cd and "inline" not in cd:
                        continue
                    filename = part.get_filename()
                    if not filename:
                        continue
                    fn_parts = _dh(filename)
                    filename = "".join(p.decode(enc or "utf-8") if isinstance(p, bytes) else p for p, enc in fn_parts)
                    content = part.get_payload(decode=True)
                    if content:
                        email_attachments.append((filename, content))

                if not email_attachments:
                    continue  # Ne procesiraj emailov brez attachmentov, pusti UNSEEN

                # Klasifikacija attachmentov
                ext_count = {"xml": 0, "pdf": 0, "other": 0, "xml_dup_batch": 0, "xml_dup_history": 0}
                for fn, content in email_attachments:
                    fn_up = fn.upper()
                    if fn_up.endswith(".XML"):
                        # Polcar XML — UTF-16 z BOM
                        try:
                            xml_text = content.decode('utf-16')
                        except UnicodeDecodeError:
                            try:
                                xml_text = content.decode('utf-8-sig')
                            except UnicodeDecodeError:
                                xml_text = content.decode('utf-8', errors='ignore')

                        # Parsiraj glede na ime ali heuristiko
                        parsed_d = []
                        parsed_c = []
                        invoice_type = None  # "DU" | "NU"
                        if '_DU_' in fn_up or '_DU.' in fn_up or 'DU.XML' in fn_up:
                            parsed_d = _parse_polcar_debit(xml_text)
                            invoice_type = "DU"
                        elif '_NU_' in fn_up or '_NU.' in fn_up or 'NU.XML' in fn_up:
                            parsed_c = _parse_polcar_credit(xml_text)
                            invoice_type = "NU"
                        else:
                            d = _parse_polcar_debit(xml_text)
                            c = _parse_polcar_credit(xml_text)
                            if d and not c:
                                parsed_d = d; invoice_type = "DU"
                            elif c and not d:
                                parsed_c = c; invoice_type = "NU"
                            elif len(d) > len(c):
                                parsed_d = d; invoice_type = "DU"
                            else:
                                parsed_c = c; invoice_type = "NU"

                        # === DEDUP CHECK ===
                        # Vzami številko fakture iz prve vrstice (vse vrstice istega XML imajo isti Številka)
                        invoice_num = ""
                        if parsed_d:
                            invoice_num = parsed_d[0].get('Številka', '').strip()
                        elif parsed_c:
                            invoice_num = parsed_c[0].get('Številka', '').strip()

                        if invoice_num and invoice_type:
                            invoice_key = f"{invoice_type}:{invoice_num}"

                            # 1. preveri ali že v tem batchu
                            if invoice_key in seen_invoices_this_batch:
                                skipped_dup_in_batch.append({
                                    "invoice_num": invoice_num, "type": invoice_type,
                                    "from": sender, "filename": fn
                                })
                                xml_files.append((f"DUPLICATE_BATCH__{fn}", content))
                                ext_count["xml_dup_batch"] += 1
                                continue

                            # 2. preveri zgodovino
                            if invoice_key in processed_history:
                                hist = processed_history[invoice_key]
                                skipped_dup_in_history.append({
                                    "invoice_num": invoice_num, "type": invoice_type,
                                    "from": sender, "filename": fn,
                                    "original_batch": hist.get("batch", "?"),
                                    "original_date": hist.get("ts", "?"),
                                })
                                xml_files.append((f"DUPLICATE_HISTORY__{fn}", content))
                                ext_count["xml_dup_history"] += 1
                                continue

                            # Ni dvojnik — dodaj v batch
                            seen_invoices_this_batch.add(invoice_key)
                            new_invoices_added.append({
                                "invoice_num": invoice_num, "type": invoice_type
                            })

                        # Doda vrstice
                        if parsed_d:
                            debit_rows.extend(parsed_d)
                        if parsed_c:
                            credit_rows.extend(parsed_c)

                        xml_files.append((fn, content))
                        ext_count["xml"] += 1
                    elif fn_up.endswith(".PDF"):
                        pdf_files.append((fn, content))
                        ext_count["pdf"] += 1
                    else:
                        # Vse ostalo dodaj v ZIP raw
                        pdf_files.append((fn, content))
                        ext_count["other"] += 1

                processed_email_nums.append(num)
                email_meta.append({
                    "from": sender, "subject": subject, "date": date_hdr,
                    "attachments": ext_count
                })

            except Exception as e:
                print(f"[knjigovodstvo-batch] Error parsing email {num}: {e}")
                continue

        if not processed_email_nums:
            mail.logout()
            return {"ok": True, "message": "Ni novih emailov z veljavnimi attachmenti", "emails": 0, "zip_filename": None}

        # === GENERIRAJ CSV-je ===
        def _make_csv(rows: list, kind: str) -> str:
            if not rows:
                return ""
            buf = io.StringIO()
            writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()), delimiter=';')
            writer.writeheader()
            writer.writerows(rows)
            return buf.getvalue()

        debit_csv = _make_csv(debit_rows, "DU")
        credit_csv = _make_csv(credit_rows, "NU")

        # === ZAPAKIRAJ V ZIP ===
        ts = _dt.now().strftime("%Y-%m-%d_%H%M%S")
        zip_name = f"{ts}_batch.zip"
        zip_path = STORAGE_KNJ_DIR / zip_name

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            if debit_csv:
                zf.writestr("DU_polcar.csv", debit_csv.encode("utf-8-sig"))
            if credit_csv:
                zf.writestr("NU_polcar.csv", credit_csv.encode("utf-8-sig"))

            # PDF-je v podmapo
            for fn, content in pdf_files:
                safe_fn = _safe_filename(fn)
                if not safe_fn.lower().endswith(".pdf") and fn.lower().endswith(".pdf"):
                    safe_fn += ".pdf"
                zf.writestr(f"PDFs/{safe_fn}", content)

            # Originalni XML-ji (backup)
            for fn, content in xml_files:
                safe_fn = _safe_filename(fn)
                if not safe_fn.lower().endswith(".xml"):
                    safe_fn += ".xml"
                zf.writestr(f"XMLs/{safe_fn}", content)

            # Manifest z dedup poročilom
            total_dups = len(skipped_dup_in_batch) + len(skipped_dup_in_history)
            manifest = {
                "created_at": _dt.now().isoformat(),
                "emails_processed": len(processed_email_nums),
                "debit_rows": len(debit_rows),
                "credit_rows": len(credit_rows),
                "pdf_count": len([p for p in pdf_files if p[0].lower().endswith(".pdf")]),
                "xml_count": len(xml_files),
                "emails": email_meta,
                "deduplication": {
                    "unique_invoices_added": len(new_invoices_added),
                    "duplicates_in_batch": len(skipped_dup_in_batch),
                    "duplicates_in_history": len(skipped_dup_in_history),
                    "total_duplicates_skipped": total_dups,
                    "new_invoices_list": new_invoices_added,
                    "duplicates_batch_detail": skipped_dup_in_batch,
                    "duplicates_history_detail": skipped_dup_in_history,
                },
            }
            zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

            # Dedup report kot ločen TXT (čitljivo za uporabnika)
            report_lines = [
                "═══ DEDUPLICATION REPORT ═══",
                f"Datum: {_lj_now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Batch ZIP: {zip_name}",
                "",
                f"📥 Obdelanih emailov: {len(processed_email_nums)}",
                f"✓ Unikatnih novih faktur: {len(new_invoices_added)}",
                f"⚠️  Duplikati v tem batchu: {len(skipped_dup_in_batch)}",
                f"♻️  Duplikati iz zgodovine: {len(skipped_dup_in_history)}",
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━",
                f"📊 Skupaj duplikatov preskočenih: {total_dups}",
                "",
            ]
            if skipped_dup_in_batch:
                report_lines.append("⚠️  DVOJNIKI V TEM BATCHU (isti račun prejel 2× v istem prevzemu):")
                for d in skipped_dup_in_batch:
                    report_lines.append(f"   • {d['type']} {d['invoice_num']}  ←  {d.get('from', '?')[:50]}")
                report_lines.append("")
            if skipped_dup_in_history:
                report_lines.append("♻️  DVOJNIKI IZ ZGODOVINE (že obdelan v prejšnjem batchu):")
                for d in skipped_dup_in_history:
                    report_lines.append(f"   • {d['type']} {d['invoice_num']}  ←  {d.get('from', '?')[:50]}")
                    report_lines.append(f"       (originalni batch: {d.get('original_batch', '?')})")
                report_lines.append("")
            if new_invoices_added:
                report_lines.append("✓ NOVE UNIKATNE FAKTURE V TEM BATCHU:")
                for inv in new_invoices_added:
                    report_lines.append(f"   • {inv['type']} {inv['invoice_num']}")
            zf.writestr("_DEDUP_REPORT.txt", "\n".join(report_lines).encode("utf-8"))

        # === Označi kot prebrane ===
        for num in processed_email_nums:
            try:
                mail.store(num, "+FLAGS", "\\Seen")
            except Exception:
                pass
        mail.logout()

        # === POSODOBI ZGODOVINO OBDELANIH FAKTUR ===
        for inv in new_invoices_added:
            key = f"{inv['type']}:{inv['invoice_num']}"
            processed_history[key] = {
                "batch": zip_name,
                "ts": _dt.now().isoformat(),
                "type": inv['type'],
                "invoice_num": inv['invoice_num'],
            }
        try:
            HISTORY_FILE.write_text(json.dumps(processed_history, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            print(f"[knj-batch] Cannot write history: {e}")

        # Log
        total_dups = len(skipped_dup_in_batch) + len(skipped_dup_in_history)
        _append_email_log({
            "ts": _dt.now().isoformat(),
            "type": "batch",
            "from": f"{len(processed_email_nums)} emailov",
            "subject": f"Batch ZIP: {zip_name}",
            "batch_zip": zip_name,
            "debit_rows": len(debit_rows),
            "credit_rows": len(credit_rows),
            "pdf_count": len(pdf_files),
            "duplicates_skipped": total_dups,
        })

        return {
            "ok": True,
            "message": f"Obdelano {len(processed_email_nums)} emailov",
            "emails": len(processed_email_nums),
            "debit_rows": len(debit_rows),
            "credit_rows": len(credit_rows),
            "pdf_count": len(pdf_files),
            "zip_filename": zip_name,
            "zip_size_kb": round(zip_path.stat().st_size / 1024, 1),
            "dedup": {
                "unique_invoices": len(new_invoices_added),
                "duplicates_in_batch": len(skipped_dup_in_batch),
                "duplicates_in_history": len(skipped_dup_in_history),
                "total_skipped": total_dups,
                "batch_duplicates_list": skipped_dup_in_batch[:20],   # max 20 za UI
                "history_duplicates_list": skipped_dup_in_history[:20],
            },
        }

    except Exception as e:
        try: mail.logout()
        except: pass
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/storage-knjigovodstvo")
async def storage_knj_list():
    """Vrne seznam vseh ZIP-ov v knjigovodstvo storage."""
    try:
        items = []
        for f in sorted(STORAGE_KNJ_DIR.iterdir(), reverse=True):
            if not f.is_file() or not f.name.endswith(".zip"):
                continue
            stat = f.stat()
            # Read manifest from inside zip for details
            import zipfile
            manifest = {}
            try:
                with zipfile.ZipFile(f, "r") as zf:
                    if "manifest.json" in zf.namelist():
                        manifest = json.loads(zf.read("manifest.json").decode("utf-8"))
            except Exception:
                pass
            from datetime import datetime as _dts
            items.append({
                "filename": f.name,
                "size_kb": round(stat.st_size / 1024, 1),
                "modified": _dts.fromtimestamp(stat.st_mtime).isoformat(),
                "manifest": manifest,
            })
        return {"ok": True, "items": items}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/storage-knjigovodstvo/{filename}")
async def storage_knj_download(filename: str):
    """Download ZIP."""
    from fastapi.responses import FileResponse
    safe = filename.replace("/", "").replace("..", "")
    f = STORAGE_KNJ_DIR / safe
    if not f.exists():
        return {"ok": False, "error": "Ne obstaja"}
    return FileResponse(str(f), filename=safe, media_type="application/zip")


@app.delete("/storage-knjigovodstvo/{filename}")
async def storage_knj_delete(filename: str):
    """Briše ZIP."""
    safe = filename.replace("/", "").replace("..", "")
    f = STORAGE_KNJ_DIR / safe
    if not f.exists():
        return {"ok": False, "error": "Ne obstaja"}
    try:
        f.unlink()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}



# ─── PROMPT BUILDERS ─────────────────────────────────────────────────────────

def build_meta_prompt(user_msg: str, pt_count: int, hl_count: int) -> str:
    pt_ph = ", ".join([f'"PT {i+1}"' for i in range(pt_count)])
    hl_ph = ", ".join([f'"HL {i+1}"' for i in range(hl_count)])
    return f"""{user_msg}

Ustvari Meta oglase za FB/Instagram v 10 jezikih.

Primary Text ({pt_count}x na jezik): 2-3 vrstice, 2-3 emoji-ji, prodajni ton, brez cen, vsak DRUGAČEN.
Headline ({hl_count}x na jezik): MAX 5 BESED, 1 emoji na začetku, brez cen, vsak DRUGAČEN.
EMOJI PRAVILO: Uporabljaj SAMO te emoji-je ki so zagotovo podprti na vseh napravah:
✅ ⭐ 🔥 💪 🎯 👍 ❤️ 💥 🚀 ✨ 💡 🎁 💰 👌 🙌 😍 💎 🏆 ⚡ 🌟 👏 💫 🛒 📦 🔑 💯 😊 🤩 🌿 🌱 🍃 🌸 🌻 🌞 🍀 🎉 🎊 🛍️ 💚 💙 🧡 💜 🤍 🖤
NE uporabljaj: redkih, novejših ali manj znanih emoji-jev ki se lahko prikažejo kot □

Jeziki — STROGO upoštevaj pisavo:
• SL = slovenščina (izvirnik)
• HR = HRVAŠČINA, SAMO latinica (npr. "Čista voda iz pipe!"). NE cirilica. NE srbohrvaški. HRVATSKI jezik z hrvaškim besediščem.
• RS = srbščina v SAMO LATINICI (npr. "Čista voda iz pipe!"). NE cirilica. Latinski srbski jezik.
• HU = madžarščina
• CZ = češčina
• SK = slovaščina
• PL = poljščina
• GR = grščina (grška pisava)
• RO = romunščina (latinica)
• BG = bolgarščina (SAMO cirilica)

KRITIČNO: HR mora biti V LATINICI in hrvaški jezik (Čista, hrvatski, voda iz pipe). Če napišeš cirilico za HR, je NAPAKA.

Vrni SAMO JSON brez markdown:
{{
  "product": "ime",
  "sl": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "hr": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "rs": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "hu": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "cz": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "sk": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "pl": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "gr": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "ro": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}},
  "bg": {{"pt": [{pt_ph}], "hl": [{hl_ph}]}}
}}"""


def build_tiktok_prompt(user_msg: str) -> str:
    return f"""{user_msg}

Ustvari TikTok oglasne tekste v 10 jezikih.

Pravila:
- Točno 4 variante na jezik
- Vsaka varianta MAX 80 znakov (strogo!)
- Brez emojiev, brez # in {{}}
- Kratek, direkten, akcijski stil
- Vsaka varianta drugačen pristop (korist, socialni dokaz, nujnost, radovednost)
- Brez "kakovost/dostava/zaloga" strukture — bodi kreativen

KRITIČNO PREPOVEDANO:
- NE omenjaj cen ali EUR (kot "6,99 EUR", "samo X EUR", itd.)
- NE omenjaj popustov ali procentov (kot "minus 62%", "do -50%", itd.)
- NE izmišljaj specifičnih številk (kot "100 funkcij", "375 kupcev") — uporabljaj generične izraze ("tisoči kupcev", "številne funkcije")
- NE omenjaj "danes/jutri" v povezavi z akcijo

Jeziki — STROGO upoštevaj pisavo:
• SL = slovenščina (izvirnik)
• HR = HRVAŠČINA, SAMO latinica (npr. "Čista voda iz pipe!"). NE cirilica. NE srbohrvaški. HRVATSKI jezik z hrvaškim besediščem.
• RS = srbščina v SAMO LATINICI (npr. "Čista voda iz pipe!"). NE cirilica. Latinski srbski jezik.
• HU = madžarščina
• CZ = češčina
• SK = slovaščina
• PL = poljščina
• GR = grščina (grška pisava)
• RO = romunščina (latinica)
• BG = bolgarščina (SAMO cirilica)

KRITIČNO: HR mora biti V LATINICI in hrvaški jezik (Čista, hrvatski, voda iz pipe). Če napišeš cirilico za HR, je NAPAKA.

KRITIČNO VAŽNO — FORMAT:
- Vsak jezik mora imeti TOČNO 4 oklepaje: [var1],[var2],[var3],[var4]
- Vejica ZNOTRAJ variante je PREPOVEDANA — če bi napisal "Brez kuhinje, brez stresa" → napiši "Brez kuhinje in brez stresa"
- Vejica se sme pojaviti SAMO med oklepaji kot separator: ],[
- Vrni IZKLJUČNO in SAMO JSON — nobenih uvodnih besed, nobenih razlag, nobenih komentarjev, nobenih markdown backticks. Prva in zadnja stvar v odgovoru mora biti {{ in }}. Nič drugega.
{{
  "product": "ime",
  "sl": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "hr": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "rs": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "hu": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "cz": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "sk": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "pl": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "gr": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "ro": "[tekst1],[tekst2],[tekst3],[tekst4]",
  "bg": "[tekst1],[tekst2],[tekst3],[tekst4]"
}}\n"""


# ─── GENERATE HELPERS ────────────────────────────────────────────────────────

def parse_json_response(text: str) -> Optional[dict]:
    text = re.sub(r"```json\s*", "", text)
    text = re.sub(r"```\s*", "", text).strip()
    # Try direct parse first
    match = re.search(r'\{[\s\S]*\}', text)
    if not match:
        return None
    json_str = match.group()
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        # Try to fix truncated JSON by finding last complete key-value
        try:
            # Remove trailing incomplete content after last complete string value
            fixed = re.sub(r',\s*"[^"]*":\s*"[^"]*$', '', json_str)
            fixed = re.sub(r',\s*"[^"]*":\s*$', '', fixed)
            if not fixed.endswith('}'):
                fixed = fixed.rstrip(',\n\r\t ') + '}'
            return json.loads(fixed)
        except Exception:
            return None


async def call_claude(prompt: str, model: str, tools=None, max_tokens: int = 4000) -> str:
    loop = asyncio.get_event_loop()
    kwargs = {"model": model, "max_tokens": max_tokens, "messages": [{"role": "user", "content": prompt}]}
    if tools:
        kwargs["tools"] = tools

    for attempt in range(4):
        try:
            msg = await loop.run_in_executor(None, lambda: client.messages.create(**kwargs))
            return "".join(b.text for b in msg.content if hasattr(b, "text"))
        except anthropic.RateLimitError:
            if attempt < 3:
                wait = (attempt + 1) * 20
                print(f"  Rate limit (429), waiting {wait}s...")
                await asyncio.sleep(wait)
            else:
                raise
        except Exception as e:
            # 529 Overloaded = Anthropic preobremenjen → poskusi znova z backoff
            is_529 = '529' in str(e) or 'overloaded' in str(e).lower()
            if is_529 and attempt < 3:
                wait = (attempt + 1) * 4   # 4s, 8s, 12s
                print(f"  Overloaded (529), waiting {wait}s... (poskus {attempt+1}/4)")
                await asyncio.sleep(wait)
            else:
                raise
    return ""


async def generate_meta_sl_only(user_msg: str, mode: str, source_url: Optional[str],
                                pt_count: int, hl_count: int) -> dict:
    """Generira samo SL tekste brez prevajanja — za streaming mode."""
    product_urls = find_product_urls(source_url)
    tools = [{"type": "web_search_20250305", "name": "web_search"}] if mode == "url" else []
    pt_ph = ", ".join([f'"PT {i+1}"' for i in range(pt_count)])
    hl_ph = ", ".join([f'"HL {i+1}"' for i in range(hl_count)])
    sl_prompt = f"""{user_msg}

Ustvari Meta oglase SAMO v slovenščini.
Primary Text ({pt_count}x): 2-3 vrstice, 2-3 emoji-ji, prodajni ton, brez cen, vsak DRUGAČEN.
Headline ({hl_count}x): MAX 5 BESED, 1 emoji na začetku, brez cen, vsak DRUGAČEN.
EMOJI PRAVILO: Uporabljaj SAMO te emoji-je ki so zagotovo podprti na vseh napravah:
✅ ⭐ 🔥 💪 🎯 👍 ❤️ 💥 🚀 ✨ 💡 🎁 💰 👌 🙌 😍 💎 🏆 ⚡ 🌟 👏 💫 🛒 📦 🔑 💯 😊 🤩 🌿 🌱 🍃 🌸 🌻 🌞 🍀 🎉 🎊 🛍️ 💚 💙 🧡 💜 🤍 🖤
NE uporabljaj: redkih, novejših ali manj znanih emoji-jev ki se lahko prikažejo kot □
Vrni SAMO JSON: {{"product": "ime", "pt": [{pt_ph}], "hl": [{hl_ph}]}}"""

    sl_text = await call_claude(sl_prompt, "claude-sonnet-4-6", tools if tools else None, 1500)
    sl_data = parse_json_response(sl_text)
    if not sl_data:
        return {"error": "Napaka pri generiranju SL tekstov."}
    return {
        "product": sl_data.get("product", "Izdelek"),
        "sl": {"pt": sl_data.get("pt", []), "hl": sl_data.get("hl", [])},
        "product_urls": product_urls,
    }


async def generate_meta_one(user_msg: str, mode: str, source_url: Optional[str],
                            pt_count: int, hl_count: int, qmode: str) -> dict:
    product_urls = find_product_urls(source_url)
    tools = [{"type": "web_search_20250305", "name": "web_search"}] if mode == "url" else []

    if qmode == "fast":
        pt_ph = ", ".join([f'"PT {i+1}"' for i in range(pt_count)])
        hl_ph = ", ".join([f'"HL {i+1}"' for i in range(hl_count)])
        sl_prompt = f"""{user_msg}

Ustvari Meta oglase SAMO v slovenščini.
Primary Text ({pt_count}x): 2-3 vrstice, 2-3 emoji-ji, prodajni ton, brez cen, vsak DRUGAČEN.
Headline ({hl_count}x): MAX 5 BESED, 1 emoji na začetku, brez cen, vsak DRUGAČEN.
EMOJI PRAVILO: Uporabljaj SAMO te emoji-je ki so zagotovo podprti na vseh napravah:
✅ ⭐ 🔥 💪 🎯 👍 ❤️ 💥 🚀 ✨ 💡 🎁 💰 👌 🙌 😍 💎 🏆 ⚡ 🌟 👏 💫 🛒 📦 🔑 💯 😊 🤩 🌿 🌱 🍃 🌸 🌻 🌞 🍀 🎉 🎊 🛍️ 💚 💙 🧡 💜 🤍 🖤 🟢 🔵 🟡
NE uporabljaj: redkih, novejših ali manj znanih emoji-jev ki se lahko prikažejo kot □
Vrni SAMO JSON: {{"product": "ime", "pt": [{pt_ph}], "hl": [{hl_ph}]}}"""

        sl_text = await call_claude(sl_prompt, "claude-sonnet-4-6", tools if tools else None, 1500)
        sl_data = parse_json_response(sl_text)
        if not sl_data:
            print(f"SL parse failed. Raw response: {sl_text[:500]}")
            return {"error": "Napaka pri generiranju SL tekstov."}

        sl_pts = sl_data.get("pt", [])
        sl_hls = sl_data.get("hl", [])
        trans_prompt = f"""Prevedi Meta oglase iz slovenščine v 9 jezikov. Ohrani ŠTEVILO in POZICIJO emoji-jev točno kot v originalu.

Primary Texts: {json.dumps(sl_pts, ensure_ascii=False)}
Headlines: {json.dumps(sl_hls, ensure_ascii=False)}

PRAVILA PREVAJANJA PO JEZIKIH:

HR (hrvaščina, latinica): Natural marketing ton. Pazi na "č/ć/š/ž/đ".

RS (srbščina, SAMO LATINICA — NIKOLI cirilica!): Natural marketing ton. Pazi na "č/ć/š/ž/đ".

HU (madžarščina): KRITIČNO - to ni indoevropski jezik, NE prevajaj dobesedno. Uporabljaj aglutinacijo (končnice) pravilno. CTA kot "Naroči zdaj" = "Rendeld meg most". Pazi da stavek ni predolg (madžarski stavki so lahko za 20-30% daljši).

CZ (češčina): Natural. Pazi na sklonjenje samostalnikov (akuzativ po glagolih).

SK (slovaščina): Podobno češčini. Natural.

PL (poljščina): Pazi na sklone (7 padežev). CTA "Naroči" = "Zamów". Uporabi neformalni ti-vi.

GR (grščina, grška pisava!): KRITIČNO - kompleksna slovnica s skloni (nominativ/akuzativ). CTA "Naroči zdaj" = "Παράγγειλε τώρα". Izogibaj se predolgih stavkov. Pazi na spol samostalnikov. Uporabljaj natural marketing grščino, ne dobesedni prevod.

RO (romunščina, latinica): Natural. Pazi na "ă/â/î/ș/ț". CTA "Naroči" = "Comandă".

BG (bolgarščina, SAMO CIRILICA — NIKOLI latinica!): Natural. CTA "Naroči" = "Поръчай".

SPLOŠNA PRAVILA:
- Ohrani prodajni/energičen ton
- Prevodi morajo zveneti kot da jih je pisal materni govorec, ne robot
- Ohrani ŠTEVILO emoji-jev (če ima SL 3 emoji, mora imeti tudi prevod 3)
- Headlines: ohrani MAX 5 besed tudi v prevodu
- Ne prevajaj blagovnih znamk, imen izdelkov, če so v originalu

Vrni SAMO JSON:
{{"hr":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"rs":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"hu":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"cz":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"sk":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"pl":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"gr":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"ro":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}},"bg":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}}}}"""

        trans_text = await call_claude(trans_prompt, "claude-haiku-4-5-20251001", None, 5000)
        trans_data = parse_json_response(trans_text)
        if not trans_data:
            return {"error": "Napaka pri prevajanju."}

        result = {"product": sl_data.get("product","Izdelek"), "sl": {"pt": sl_pts, "hl": sl_hls}, "product_urls": product_urls}
        result.update(trans_data)
        return result
    else:
        prompt = build_meta_prompt(user_msg, pt_count, hl_count)
        text = await call_claude(prompt, "claude-sonnet-4-6", tools if tools else None)
        data = parse_json_response(text)
        if not data:
            return {"error": "Claude ni vrnil veljavnega JSON."}
        data["product_urls"] = product_urls
        return data


async def generate_tiktok_one(user_msg: str, mode: str, source_url: Optional[str]) -> dict:
    product_urls = find_product_urls(source_url)
    tools = [{"type": "web_search_20250305", "name": "web_search"}] if mode == "url" else []
    prompt = build_tiktok_prompt(user_msg)
    text = await call_claude(prompt, "claude-sonnet-4-6", tools if tools else None)
    data = parse_json_response(text)
    if not data:
        print(f"TikTok JSON parse failed. Raw (first 800 chars): {text[:800]}")
        return {"error": "Claude ni vrnil veljavnega JSON."}
    data["product_urls"] = product_urls
    return data


# ─── TIKTOK XLSX BUILDER ─────────────────────────────────────────────────────

COUNTRY_TO_LANG = {
    "BG": "bg", "CZ": "cz", "GR": "gr", "SK": "sk", "RS": "rs",
    "RO": "ro", "HU": "hu", "HR": "hr", "PL": "pl", "SLO": "sl"
}


def normalize_tiktok_text(raw: str) -> str:
    """Zagotovi format [var1],[var2],[var3],[var4] ne glede na vhodni format."""
    raw = raw.strip()
    # Pravilni AI format: [var1],[var2],... — pusti kot je
    if re.match(r'^\[.*\],\[.*\]', raw):
        return raw
    # Format ],[  brez prefiksa/sufiksa oklepaja
    if '],[' in raw:
        parts = re.split(r'\]\s*,\s*\[', raw)
        parts = [p.strip().strip('[]') for p in parts if p.strip().strip('[]')]
        return ','.join(f'[{p}]' for p in parts)
    # Brez oklepajev — AI pozabil dodati []. Splittaj po vejici.
    # Varno ker pravilo narekuje max 80 znakov / varianto in vsebinska vejica ni pričakovana.
    if '[' not in raw:
        parts = [p.strip() for p in raw.split(',') if p.strip()]
        return ','.join(f'[{p}]' for p in parts)
    # En sam oklepaj npr. [t1,t2,t3,t4] — pusti kot je, TikTok bo rešil
    return raw


def build_tiktok_xlsx(sku: str, brand: str, video_names: str,
                      texts_by_lang: dict, urls_by_lang: dict,
                      skip_rs: bool = False) -> str:
    if not Path(TEMPLATE_PATH).exists():
        raise FileNotFoundError("TikTok template not found. Upload tiktok_template.xlsx to static/")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb['Ads']
    headers = [cell.value for cell in ws[1]]

    col_campaign = headers.index('Campaign Name') + 1
    col_bc_id    = headers.index('Business Center ID of the identity') + 1
    col_video    = headers.index('Video Name') + 1
    col_text     = headers.index('Text') + 1
    col_url      = headers.index('Web URL') + 1
    col_ag       = headers.index('Ad Group Name') + 1

    base_bc_raw = ws.cell(row=2, column=col_bc_id).value or ''
    single_id = base_bc_raw.split(',')[0].strip()
    videos = [v.strip() for v in re.findall(r'\[([^\]]+)\]', video_names)]
    new_bc_id = ','.join([single_id] * len(videos)) if videos else single_id
    today = _lj_now().strftime('%-d_%-m_%Y')
    new_campaign = f'[{brand}] Smart+ {sku} - {today}'

    # Zberemo vrstice za brisanje (ne brišemo med iteracijo!)
    rows_to_delete = []
    for row in ws.iter_rows(min_row=2):
        r = row[0].row
        country = ws.cell(row=r, column=col_ag).value
        if not country:
            continue
        lang = COUNTRY_TO_LANG.get(country)
        if skip_rs and lang == 'rs':
            rows_to_delete.append(r)
            continue
        ws.cell(row=r, column=col_campaign).value = new_campaign
        ws.cell(row=r, column=col_bc_id).value = new_bc_id
        ws.cell(row=r, column=col_video).value = video_names
        if lang and lang in texts_by_lang:
            ws.cell(row=r, column=col_text).value = normalize_tiktok_text(texts_by_lang[lang])
        url = (urls_by_lang.get(lang) if lang else None) or next(iter(urls_by_lang.values()), '')
        if url:
            ws.cell(row=r, column=col_url).value = url

    # Zbriši RS vrstice po iteraciji (v obratnem vrstnem redu da ne zamešamo indeksov)
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)

    out_path = str(EXPORTS_DIR / f"tiktok_{sku}_{uuid.uuid4().hex[:8]}.xlsx")
    wb.save(out_path)
    return out_path



def build_master_xlsx(skus: list) -> str:
    """Združi več SKU-jev v en master XLS — vsak SKU doda svoje vrstice (koliko jih je v template-u)."""
    if not Path(TEMPLATE_PATH).exists():
        raise FileNotFoundError("TikTok template not found.")

    # Load template to get headers and template rows
    wb_tmpl = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_tmpl = wb_tmpl['Ads']
    headers = [cell.value for cell in ws_tmpl[1]]

    col_campaign = headers.index('Campaign Name') + 1
    col_bc_id    = headers.index('Business Center ID of the identity') + 1
    col_video    = headers.index('Video Name') + 1
    col_text     = headers.index('Text') + 1
    col_url      = headers.index('Web URL') + 1
    col_ag       = headers.index('Ad Group Name') + 1

    base_bc_raw = ws_tmpl.cell(row=2, column=col_bc_id).value or ''
    single_id = base_bc_raw.split(',')[0].strip()

    # Collect template rows (row 2 onwards) as base structure
    tmpl_rows = []
    for row in ws_tmpl.iter_rows(min_row=2, values_only=False):
        country = row[col_ag - 1].value
        if not country:
            continue
        tmpl_rows.append({
            'country': country,
            'row_data': [cell.value for cell in row],
            'row_num': row[0].row,
        })

    # Create new workbook
    wb_out = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_out = wb_out['Ads']

    # Clear all data rows
    for row in ws_out.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    today = _lj_now().strftime('%-d_%-m_%Y')
    out_row = 2

    for sku_entry in skus:
        sku = sku_entry.get('sku', '')
        brand = sku_entry.get('brand', '')
        video_names = sku_entry.get('videos', '')
        texts_by_lang = sku_entry.get('texts', {})
        urls_by_lang = sku_entry.get('urls', {})
        fallback_url = sku_entry.get('url') or sku_entry.get('source_url') or ''
        print(f"[master] SKU={sku} url={fallback_url!r} urls={urls_by_lang}")
        # Fallback: če urls prazen, uporabi url za vse jezike
        if not urls_by_lang and fallback_url:
            all_langs = list(set(list(COUNTRY_TO_LANG.values()) + ['sl','hr','rs','hu','cz','sk','pl','gr','ro','bg']))
            urls_by_lang = {lang: fallback_url for lang in all_langs}
        print(f"[master] SKU={sku} urls_by_lang keys={list(urls_by_lang.keys())[:5]}")

        videos = [v.strip() for v in re.findall(r'\[([^\]]+)\]', video_names)]
        new_bc_id = ','.join([single_id] * len(videos)) if videos else single_id
        new_campaign = f'[{brand}] Smart+ {sku} - {today}'

        for tmpl_row in tmpl_rows:
            country = tmpl_row['country']
            lang = COUNTRY_TO_LANG.get(country)

            # Copy template row values
            orig_row = ws_tmpl[tmpl_row['row_num']]
            for col_idx, orig_cell in enumerate(orig_row, 1):
                ws_out.cell(row=out_row, column=col_idx).value = orig_cell.value

            # Fill in our data
            ws_out.cell(row=out_row, column=col_campaign).value = new_campaign
            ws_out.cell(row=out_row, column=col_bc_id).value = new_bc_id
            ws_out.cell(row=out_row, column=col_video).value = video_names
            if lang and lang in texts_by_lang:
                ws_out.cell(row=out_row, column=col_text).value = normalize_tiktok_text(texts_by_lang[lang])
            # URL: lang-specifičen ali fallback
            url = (urls_by_lang.get(lang) if lang else None) or fallback_url or next(iter(urls_by_lang.values()), '')
            if url:
                ws_out.cell(row=out_row, column=col_url).value = url

            out_row += 1

    out_path = str(EXPORTS_DIR / f"master_{uuid.uuid4().hex[:8]}.xlsx")
    wb_out.save(out_path)
    return out_path


class MasterXlsxRequest(BaseModel):
    skus: List[dict]  # [{sku, brand, url, videos, texts, urls}]


@app.post("/build-master-xlsx")
async def build_master_xlsx_endpoint(req: MasterXlsxRequest):
    """Sestavi master XLS iz že shranjenih tekstov — brez API klicev."""
    skus_data = []
    for entry in req.skus:
        if not entry.get('videos'):
            continue
        skus_data.append({
            'sku':    entry.get('sku', ''),
            'brand':  entry.get('brand', ''),
            'videos': entry.get('videos', ''),
            'texts':  entry.get('texts', {}),
            'urls':   entry.get('urls', {}),
        })

    if not skus_data:
        return {"error": "Ni veljavnih SKU-jev (manjkajo video imena)."}

    try:
        path = build_master_xlsx(skus_data)
        return {"status": "ok", "file": path}
    except FileNotFoundError as e:
        return {"error": str(e)}


@app.post("/generate-master-xlsx")
async def generate_master_xlsx(req: MasterXlsxRequest):
    await ensure_cache_fresh()

    skus_data = []
    for entry in req.skus:
        url = entry.get('url', '')
        sku = entry.get('sku', '')
        brand = entry.get('brand', '')
        videos = entry.get('videos', '')

        if not videos:
            continue  # preskoči SKU brez video imen

        # Generate TikTok texts
        user_msg = f"Preberi to stran in ustvari TikTok oglase: {url}"
        data = await generate_tiktok_one(user_msg, "url", url)
        if "error" in data:
            continue

        texts_by_lang = {lang: data[lang] for lang in ["sl","hr","rs","hu","cz","sk","pl","gr","ro","bg"] if lang in data}
        urls_by_lang = data.get("product_urls", {})

        skus_data.append({
            'sku': sku, 'brand': brand, 'videos': videos,
            'texts': texts_by_lang, 'urls': urls_by_lang
        })

        if len(skus_data) < len(req.skus):
            await asyncio.sleep(15)  # rate limit

    if not skus_data:
        return {"error": "Ni veljavnih SKU-jev za generiranje."}

    try:
        path = build_master_xlsx(skus_data)
        return {"status": "ok", "file": path}
    except FileNotFoundError as e:
        return {"error": str(e)}


    input: str
    mode: str
    pt_count: int = 1
    hl_count: int = 1
    source_url: Optional[str] = None
    qmode: str = "sonnet"


class MultiAdRequest(BaseModel):
    products: List[dict]
    pt_count: int = 1
    hl_count: int = 1
    qmode: str = "sonnet"


class TikTokRequest(BaseModel):
    source_url: str
    sku: str
    brand: str
    video_names: str
    skip_rs: bool = False
    product_data: Optional[dict] = None  # maaarket API podatki iz frontenda


# ─── ROUTES ──────────────────────────────────────────────────────────────────

@app.get("/tiktok-history")
async def get_tiktok_history():
    if TT_HISTORY_FILE.exists():
        try:
            return json.loads(TT_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/tiktok-history")
async def save_tiktok_history(data: dict):
    try:
        history = data.get("history", [])
        TT_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(history)}
    except Exception as e:
        return {"error": str(e)}

@app.get("/meta-history")
async def get_meta_history():
    if META_HISTORY_FILE.exists():
        try:
            return json.loads(META_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/meta-history")
async def save_meta_history(data: dict):
    try:
        history = data.get("history", [])
        META_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(history)}
    except Exception as e:
        return {"error": str(e)}

@app.get("/sporocanje-common")
async def sporocanje_get_common(brand: str = "maaarket"):
    """Vrne seznam pogostih odgovorov za določen brand."""
    file = DATA_DIR / f"sporocanje_{brand}.json"
    if file.exists():
        try:
            data = json.loads(file.read_text(encoding="utf-8"))
            return {"ok": True, "answers": data.get("answers", [])}
        except:
            pass
    return {"ok": True, "answers": []}

@app.post("/sporocanje-save")
async def sporocanje_save(data: dict):
    """Shrani odgovor v pogosto bazo za določen brand."""
    brand = data.get("brand", "maaarket")
    file = DATA_DIR / f"sporocanje_{brand}.json"
    existing = {"answers": []}
    if file.exists():
        try:
            existing = json.loads(file.read_text(encoding="utf-8"))
        except:
            pass
    answers = existing.get("answers", [])
    reply_sl = (data.get("reply_sl") or "").strip().lower()
    found = False
    for a in answers:
        if (a.get("reply_sl") or "").strip().lower() == reply_sl:
            a["count"] = a.get("count", 1) + 1
            a["ts"] = data.get("ts", a.get("ts"))
            found = True
            break
    if not found:
        answers.insert(0, {
            "question": data.get("question", ""),
            "reply_sl": data.get("reply_sl", ""),
            "reply_translated": data.get("reply_translated", ""),
            "reply_hr": data.get("reply_hr", ""),
            "lang": data.get("lang", ""),
            "brand": brand,
            "count": 1,
            "ts": data.get("ts", 0),
        })
    answers.sort(key=lambda x: -x.get("count", 1))
    answers = answers[:200]
    existing["answers"] = answers
    file.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "total": len(answers)}

@app.post("/ai-proxy")
async def ai_proxy(data: dict):
    """Proxy za AI klice iz frontenda (za Sporočanje)."""
    prompt = data.get("prompt", "")
    max_tokens = min(int(data.get("max_tokens", 500)), 800)
    model = data.get("model", "claude-haiku-4-5-20251001")
    if not prompt:
        return {"content": [{"text": ""}]}
    msg = client.messages.create(
        model=model,
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}]
    )
    return {"content": [{"type": "text", "text": msg.content[0].text}]}

@app.post("/kayako-kb-search")
async def kayako_kb_search(data: dict):
    """Poišče top N relevantnih Q&A parov iz KB za dano vprašanje."""
    brand   = data.get("brand", "maaarket")
    query   = data.get("query", "").lower().strip()
    top_n   = min(int(data.get("top_n", 8)), 20)
    if not query:
        return {"ok": True, "results": []}
    kb_file = KB_FILES.get(brand)
    if not kb_file or not kb_file.exists():
        return {"ok": True, "results": []}
    try:
        kb = json.loads(kb_file.read_text(encoding="utf-8"))
        pairs = kb.get("qa_pairs", [])
    except:
        return {"ok": True, "results": []}
    # Keyword scoring
    words = [w for w in query.split() if len(w) > 3]
    scored = []
    for p in pairs:
        hay = (p.get("subject","") + " " + p.get("question","") + " " + p.get("answer","")).lower()
        score = sum(1 for w in words if w in hay)
        if score > 0:
            scored.append((score, p))
    scored.sort(key=lambda x: -x[0])
    results = [p for _, p in scored[:top_n]]
    return {"ok": True, "results": results, "total_kb": len(pairs)}


@app.get("/forecast-entries")
async def get_forecast_entries():
    """Vrne entries samo če so za danes."""
    from datetime import datetime
    try:
        import pytz
        lj = pytz.timezone("Europe/Ljubljana")
        today = datetime.now(lj).strftime("%Y-%m-%d")
    except:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    if not FORECAST_ENTRIES_FILE.exists():
        return {}
    try:
        data = json.loads(FORECAST_ENTRIES_FILE.read_text(encoding="utf-8"))
        if data.get("date") == today:
            return data
        # Stari podatki — vrni prazen
        print(f"[forecast] entries are from {data.get('date')}, today is {today} — returning empty")
        return {}
    except:
        return {}

@app.post("/forecast-entries")
async def save_forecast_entries(data: dict):
    """Združi poslane entries z obstoječimi po času (multi-user safe).
    Spoštuje deleted-list: vnose ki so bili izbrisani po datumu deletion ne sprejme."""
    try:
        from datetime import datetime
        try:
            import pytz
            lj = pytz.timezone("Europe/Ljubljana")
            today = datetime.now(lj).strftime("%Y-%m-%d")
        except:
            today = datetime.utcnow().strftime("%Y-%m-%d")

        new_entries = data.get("entries", [])
        new_date = data.get("date", today)
        replace_mode = data.get("replace", False)

        existing = {}
        if FORECAST_ENTRIES_FILE.exists():
            try:
                existing = json.loads(FORECAST_ENTRIES_FILE.read_text(encoding="utf-8"))
            except:
                existing = {}

        if existing.get("date") != today:
            existing = {"date": today, "entries": []}

        if new_date != today:
            print(f"[forecast] ignored save with non-today date {new_date}")
            return {"status": "ok", "note": "date mismatch"}

        # Naloži deleted-list za danes
        deleted = {}
        if FORECAST_DELETED_FILE.exists():
            try:
                all_deleted = json.loads(FORECAST_DELETED_FILE.read_text(encoding="utf-8"))
                deleted = all_deleted.get(today, {})  # {label: timestamp_ms}
            except:
                deleted = {}

        # REPLACE mode — zamenja namesto združi
        if replace_mode:
            # Filtriraj tiste ki so bili brisani
            filtered = [e for e in new_entries if e.get("label","") not in deleted]
            result = {"date": today, "entries": sorted(filtered, key=lambda e: e.get("label",""))}
            FORECAST_ENTRIES_FILE.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"[forecast] REPLACE save: {len(filtered)} entries")
            return {"status": "ok", "replaced": len(filtered)}

        # MERGE mode
        merged = {}
        for e in existing.get("entries", []):
            lbl = e.get("label","")
            if lbl not in deleted:
                merged[lbl] = e
        for e in new_entries:
            lbl = e.get("label","")
            entry_date = e.get("_date", "")
            # Zavrni entry če _date obstaja in ni danes
            if entry_date and entry_date != today:
                print(f"[forecast] rejected entry '{lbl}' with _date={entry_date} (today={today})")
                continue
            if lbl not in deleted:
                merged[lbl] = e
        sorted_entries = sorted(merged.values(), key=lambda e: e.get("label",""))

        result = {"date": today, "entries": sorted_entries}
        FORECAST_ENTRIES_FILE.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[forecast] MERGE save: total {len(sorted_entries)} (filtered out {len(deleted)} deleted)")
        return {"status": "ok", "merged": len(sorted_entries)}
    except Exception as e:
        return {"error": str(e)}

@app.post("/forecast-entry-delete")
async def delete_forecast_entry(data: dict):
    """Izbriši vnos iz danes — doda na deleted-list ki traja čez seje."""
    try:
        from datetime import datetime
        try:
            import pytz
            lj = pytz.timezone("Europe/Ljubljana")
            today = datetime.now(lj).strftime("%Y-%m-%d")
        except:
            today = datetime.utcnow().strftime("%Y-%m-%d")

        label = data.get("label", "")
        if not label:
            return {"error": "missing label"}

        # Naloži deleted-list
        all_deleted = {}
        if FORECAST_DELETED_FILE.exists():
            try:
                all_deleted = json.loads(FORECAST_DELETED_FILE.read_text(encoding="utf-8"))
            except:
                all_deleted = {}

        # Dodaj v deleted za danes
        if today not in all_deleted:
            all_deleted[today] = {}
        import time
        all_deleted[today][label] = int(time.time() * 1000)

        # Pobriši stare dni (>7 dni) — sprosti prostor
        cutoff_keys = sorted(all_deleted.keys())
        if len(cutoff_keys) > 7:
            for k in cutoff_keys[:-7]:
                del all_deleted[k]

        FORECAST_DELETED_FILE.write_text(json.dumps(all_deleted, ensure_ascii=False, indent=2), encoding="utf-8")

        # Takoj odstrani iz entries fajla
        if FORECAST_ENTRIES_FILE.exists():
            try:
                existing = json.loads(FORECAST_ENTRIES_FILE.read_text(encoding="utf-8"))
                if existing.get("date") == today:
                    existing["entries"] = [e for e in existing.get("entries",[]) if e.get("label","") != label]
                    FORECAST_ENTRIES_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
            except:
                pass

        print(f"[forecast] DELETED entry '{label}' for {today}")
        return {"status": "ok", "deleted": label}
    except Exception as e:
        return {"error": str(e)}

@app.get("/forecast-history")
async def get_forecast_history():
    if FORECAST_HISTORY_FILE.exists():
        try:
            hist = json.loads(FORECAST_HISTORY_FILE.read_text(encoding="utf-8"))
            # Normaliziraj stare sl-SI ključe → ISO format
            changed = False
            new_hist = {}
            for key, val in hist.items():
                # Prepoznaj sl-SI format: "8. 5. 2026"
                if '.' in key and len(key.split('.')) == 3:
                    try:
                        parts = [p.strip().strip('.') for p in key.split('.') if p.strip()]
                        if len(parts) == 3:
                            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                            iso_key = f"{y}-{m:02d}-{d:02d}"
                            if iso_key not in hist:
                                new_hist[iso_key] = val
                                changed = True
                            else:
                                new_hist[key] = val  # ISO verzija že obstaja
                        else:
                            new_hist[key] = val
                    except:
                        new_hist[key] = val
                else:
                    new_hist[key] = val
            if changed:
                FORECAST_HISTORY_FILE.write_text(json.dumps(new_hist, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"[forecast] normalized {sum(1 for k in new_hist if '-' in k)} history keys to ISO format")
                return new_hist
            return hist
        except:
            return {}
    return {}

@app.post("/forecast-history")
async def save_forecast_history(data: dict):
    try:
        FORECAST_HISTORY_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok"}
    except Exception as e:
        return {"error": str(e)}

@app.get("/")
def root():
    return FileResponse("static/index.html")


@app.get("/cache-status")
async def cache_status():
    return {"last_fetch": last_fetch.isoformat() if last_fetch else None,
            "stale": is_cache_stale(),
            "products_per_lang": {lang: len(p) for lang, p in feed_by_lang.items()},
            "slug_index_size": len(slug_to_id)}


@app.post("/refresh-cache")
async def refresh_cache():
    await fetch_all_feeds()
    return {"status": "ok", "last_fetch": last_fetch.isoformat()}


@app.post("/generate")
async def generate(req: AdRequest):
    await ensure_cache_fresh()
    user_msg = f"Preberi to stran in ustvari Meta oglase: {req.input}" if req.mode == "url" else f"Na podlagi tega opisa ustvari Meta oglase:\n\n{req.input}"
    return await generate_meta_one(user_msg, req.mode, req.source_url, req.pt_count, req.hl_count, req.qmode)


@app.post("/generate-multi")
async def generate_multi(req: MultiAdRequest):
    await ensure_cache_fresh()
    results = []
    for i, p in enumerate(req.products):
        url = p.get("url", "").strip()
        mode = p.get("mode", "url")
        if not url:
            results.append({"error": "Prazen URL"})
            continue
        user_msg = f"Preberi to stran in ustvari Meta oglase: {url}" if mode == "url" else f"Na podlagi tega opisa:\n\n{url}"
        result = await generate_meta_one(user_msg, mode, url if mode == "url" else None,
                                         req.pt_count, req.hl_count, req.qmode)
        results.append(result)
        if i < len(req.products) - 1:
            await asyncio.sleep(15)
    return {"results": results}


@app.post("/generate-multi-stream")
async def generate_multi_stream(req: MultiAdRequest):
    """SSE streaming endpoint — pošilja rezultate batch po batch."""
    await ensure_cache_fresh()

    async def event_stream():
        for i, p in enumerate(req.products):
            url = p.get("url", "").strip()
            mode = p.get("mode", "url")
            if not url:
                yield f"data: {json.dumps({'type': 'result', 'index': i, 'data': {'error': 'Prazen URL'}})}\n\n"
                continue

            # Notify frontend that this product is being processed
            yield f"data: {json.dumps({'type': 'loading', 'index': i, 'url': url})}\n\n"

            user_msg = f"Preberi to stran in ustvari Meta oglase: {url}" if mode == "url" else f"Na podlagi tega opisa:\n\n{url}"

            if req.qmode == "fast":
                # Step 1: SL generation
                yield f"data: {json.dumps({'type': 'progress', 'index': i, 'step': 'sl'})}\n\n"
                result = await generate_meta_sl_only(user_msg, mode, url if mode == "url" else None,
                                                      req.pt_count, req.hl_count)
                if "error" in result:
                    yield f"data: {json.dumps({'type': 'result', 'index': i, 'data': result})}\n\n"
                    if i < len(req.products) - 1:
                        await asyncio.sleep(15)
                    continue

                # Send SL immediately
                yield f"data: {json.dumps({'type': 'partial', 'index': i, 'langs': ['sl'], 'data': result})}\n\n"

                sl_pts = result["sl"]["pt"]
                sl_hls = result["sl"]["hl"]
                product_urls = result.get("product_urls", {})
                product_name = result.get("product", "Izdelek")

                # Step 2: Parallel translation in 4 batches of 2-3 langs
                pt_ph = ", ".join([f'"PT {i2+1}"' for i2 in range(req.pt_count)])
                hl_ph = ", ".join([f'"HL {i2+1}"' for i2 in range(req.hl_count)])

                lang_batches = [
                    ["hr", "rs"],
                    ["hu", "cz"],
                    ["sk", "pl"],
                    ["gr", "ro", "bg"],
                ]

                lang_info = {
                    "hr": "HR (hrvaščina, latinica)",
                    "rs": "RS (srbščina, SAMO latinica!)",
                    "hu": "HU (madžarščina - aglutinacijski jezik, ne prevajaj dobesedno)",
                    "cz": "CZ (češčina)",
                    "sk": "SK (slovaščina)",
                    "pl": "PL (poljščina)",
                    "gr": "GR (grščina, grška pisava!)",
                    "ro": "RO (romunščina, latinica)",
                    "bg": "BG (bolgarščina, SAMO cirilica!)",
                }

                full_result = {
                    "product": product_name,
                    "product_urls": product_urls,
                    "sl": {"pt": sl_pts, "hl": sl_hls},
                }

                # GLOBAL flag — če Haiku enkrat zavrne s 529, preskakuj Haiku za vse naslednje batch-e
                # (ker je trenutno preobremenjen — Anthropic outage)
                skip_haiku_for_rest = False

                for batch in lang_batches:
                    yield f"data: {json.dumps({'type': 'progress', 'index': i, 'step': 'translating', 'langs': batch})}\n\n"

                    batch_json_keys = ", ".join([
                        f'"{lang}":{{"pt":[{pt_ph}],"hl":[{hl_ph}]}}'
                        for lang in batch
                    ])
                    batch_lang_lines = "\n".join([f"- {lang_info[lang]}" for lang in batch])

                    batch_prompt = f"""Prevedi Meta oglase iz slovenščine v naslednje jezike. Ohrani ŠTEVILO in POZICIJO emoji-jev točno kot v originalu.

Primary Texts: {json.dumps(sl_pts, ensure_ascii=False)}
Headlines: {json.dumps(sl_hls, ensure_ascii=False)}

Prevedi SAMO v te jezike:
{batch_lang_lines}

SPLOŠNA PRAVILA:
- Ohrani prodajni/energičen ton
- Prevodi morajo zveneti kot materni govorec
- Ohrani ŠTEVILO emoji-jev
- Headlines: MAX 5 besed
- Ne prevajaj imen izdelkov/blagovnih znamk

Vrni SAMO JSON: {{{batch_json_keys}}}"""

                    # Retry do 3x — pri 529 overload se model preklopi na Sonnet
                    # Če smo že kdaj v tem requestu videli 529, takoj uporabi Sonnet
                    batch_data = None
                    last_error_was_529 = False
                    for attempt in range(3):
                        # Izberi model:
                        # - če je globalni flag prižgan (Haiku že failal nekje) → Sonnet
                        # - če je trenutni batch že imel 529 → Sonnet
                        # - sicer Haiku za prva 2 poskusa, Sonnet za 3.
                        if skip_haiku_for_rest or last_error_was_529:
                            model = "claude-sonnet-4-6"
                        else:
                            model = "claude-haiku-4-5-20251001" if attempt < 2 else "claude-sonnet-4-6"
                        try:
                            batch_text = await call_claude(batch_prompt, model, None, 4000)
                            batch_data = parse_json_response(batch_text)
                            if batch_data:
                                missing = [lang for lang in batch if lang not in batch_data]
                                if missing:
                                    print(f"[meta-stream] batch {batch} ({model}) manjkajo jeziki: {missing}, retry {attempt+1}/3")
                                    batch_data = None
                                else:
                                    print(f"[meta-stream] batch {batch} ({model}) USPELO v {attempt+1}. poskusu")
                                    break
                            else:
                                print(f"[meta-stream] batch {batch} ({model}) JSON parse fail, retry {attempt+1}/3, response[:200]: {batch_text[:200]}")
                        except Exception as e:
                            err_str = str(e)
                            is_529 = '529' in err_str or 'overloaded' in err_str.lower()
                            last_error_was_529 = is_529
                            # Če je Haiku in 529, prižgi globalni flag — ne probaj več Haiku v naslednjih batch-ih
                            if is_529 and 'haiku' in model.lower():
                                if not skip_haiku_for_rest:
                                    print(f"[meta-stream] 🚨 Haiku overloaded — preklopi na Sonnet za VSE naslednje batch-e")
                                skip_haiku_for_rest = True
                            print(f"[meta-stream] batch {batch} ({model}) error attempt {attempt+1}/3 [{'OVERLOADED' if is_529 else 'OTHER'}]: {type(e).__name__}: {err_str[:200]}")
                        # Pavza pred retry-jem — daljša pri 529
                        if attempt < 2:
                            wait_s = 8 if last_error_was_529 else 2
                            await asyncio.sleep(wait_s)

                    if batch_data:
                        full_result.update(batch_data)
                        yield f"data: {json.dumps({'type': 'partial', 'index': i, 'langs': batch, 'data': full_result})}\n\n"
                    else:
                        print(f"[meta-stream] batch {batch} FAILED po 3 poskusih — jeziki manjkajo v rezultatu")
                        yield f"data: {json.dumps({'type': 'batch_error', 'index': i, 'langs': batch, 'error': f'Prevod za {batch} ni uspel'})}\n\n"

                yield f"data: {json.dumps({'type': 'result', 'index': i, 'data': full_result})}\n\n"

            else:
                # Kreativni način — en klic, pošlji ko konča
                result = await generate_meta_one(user_msg, mode, url if mode == "url" else None,
                                                  req.pt_count, req.hl_count, req.qmode)
                yield f"data: {json.dumps({'type': 'result', 'index': i, 'data': result})}\n\n"

            if i < len(req.products) - 1:
                await asyncio.sleep(15)

        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.post("/generate-tiktok")
async def generate_tiktok(req: TikTokRequest):
    await ensure_cache_fresh()

    # Če imamo API podatke iz frontenda → uporabimo jih (brez web_search)
    if req.product_data and req.product_data.get("title"):
        pd = req.product_data

        # Zberi features opise
        features_text = ""
        for f in (pd.get("features") or []):
            t = f.get("title", "")
            c = f.get("content", "")
            if t or c:
                features_text += f"\n- {t}: {c}"

        # Reviews
        reviews_text = ""
        for r in (pd.get("comments") or [])[:3]:
            reviews_text += f'\n- "{r.get("content","")}" — {r.get("name","")}'

        sales = pd.get("sales_count", "")
        rating = ""
        try:
            rating = pd.get("rating", {}).get("total", {}).get("average", "")
        except:
            pass

        user_msg = f"""Napiši TikTok oglase za ta izdelek. Piši kot izkušen copywriter — ne naštevaj specifikacij, piši zgodbe in čustva.

IME: {pd.get('title', '')}
OPIS: {pd.get('lead', '') or pd.get('short_desc', '')}
PODROBNOSTI: {pd.get('content', '')[:800]}
{f"PREDNOSTI:{features_text}" if features_text else ""}
{f"PRODANO: {sales}x" if sales else ""}
{f"OCENE kupcev:{reviews_text}" if reviews_text else ""}
{f"POVPREČNA OCENA: {rating}/5" if rating else ""}

Ustvari TikTok oglase za ta izdelek."""

        print(f"[tiktok] API podatki: {pd.get('title','')[:50]} | {sales} prodaj | {rating}★")
        data = await generate_tiktok_one(user_msg, "text", req.source_url)
    else:
        # Fallback: stari način z web_search
        user_msg = f"Preberi to stran in ustvari TikTok oglase: {req.source_url}"
        print(f"[tiktok] Fallback web_search za: {req.source_url}")
        data = await generate_tiktok_one(user_msg, "url", req.source_url)

    if "error" in data:
        return data

    product_urls = data.get("product_urls", {})
    texts_by_lang = {lang: data[lang] for lang in ["sl","hr","rs","hu","cz","sk","pl","gr","ro","bg"] if lang in data}

    try:
        xlsx_path = build_tiktok_xlsx(
            sku=req.sku,
            brand=req.brand,
            video_names=req.video_names,
            texts_by_lang=texts_by_lang,
            urls_by_lang=product_urls,
            skip_rs=req.skip_rs
        )
        return {"status": "ok", "file": xlsx_path, "data": data}
    except FileNotFoundError as e:
        return {"error": str(e)}


@app.post("/extract-videos")
async def extract_videos(data: dict):
    """Extract video filenames from a base64 screenshot using Claude vision."""
    image_b64 = data.get("image")
    media_type = data.get("media_type", "image/png")
    if not image_b64:
        return {"error": "Ni slike."}
    loop = asyncio.get_event_loop()

    _prompt = "Extract all video filenames from this screenshot. Return ONLY a JSON array of filenames, nothing else. Example: [\"VIDEO (1).mp4\", \"VIDEO (2).mp4\"]. Extract exactly as written, preserve spaces and capitalization."

    def _call(model):
        return client.messages.create(
            model=model,
            max_tokens=500,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_b64}},
                {"type": "text", "text": _prompt}
            ]}]
        )

    # Retry do 3x — pri 529 overload preklopi na Sonnet
    msg = None
    last_err = None
    for attempt in range(3):
        # Prva 2 poskusa Haiku, 3. Sonnet (ali takoj Sonnet po 529)
        model = "claude-haiku-4-5-20251001" if attempt == 0 else "claude-sonnet-4-6"
        try:
            msg = await loop.run_in_executor(None, lambda m=model: _call(m))
            break
        except Exception as e:
            last_err = e
            is_529 = '529' in str(e) or 'overloaded' in str(e).lower()
            print(f"[extract-videos] ({model}) attempt {attempt+1}/3 error [{'OVERLOADED' if is_529 else 'OTHER'}]: {type(e).__name__}")
            if attempt < 2:
                await asyncio.sleep(2)

    if msg is None:
        return {"error": "Anthropic je trenutno preobremenjen (529). Poskusi znova čez nekaj sekund."}

    text = msg.content[0].text.strip()
    text = re.sub(r"```json\s*", "", text)
    text = re.sub(r"```\s*", "", text).strip()
    try:
        filenames = json.loads(text)
        formatted = ",".join([f"[{f}]" for f in filenames])
        return {"filenames": filenames, "formatted": formatted}
    except json.JSONDecodeError:
        return {"error": "Ni uspelo prebrati imen. Poskusi znova z jasnejšo sliko."}


@app.get("/download/{filename}")
def download_file(filename: str):
    path = EXPORTS_DIR / filename
    if not path.exists():
        return {"error": "File not found"}
    return FileResponse(str(path), filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/kreative-history")
async def get_kreative_history():
    if KREATIVE_HISTORY_FILE.exists():
        try:
            return json.loads(KREATIVE_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/kreative-history")
async def save_kreative_history(data: dict):
    try:
        history = data.get("history", [])
        KREATIVE_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(history)}
    except Exception as e:
        return {"error": str(e)}


KREATIVE_QUEUE_FILE = DATA_DIR / "kreative_queue.json"
KREATIVE_QUEUE_IMG_DIR = DATA_DIR / "kreative_queue_imgs"
KREATIVE_QUEUE_IMG_DIR.mkdir(exist_ok=True, parents=True)

def _qimg_safe_id(jid):
    return "".join(c for c in str(jid) if c.isalnum() or c in "-_")[:64]

@app.get("/kreative-queue")
async def get_kreative_queue():
    """Obstojnost čakalne vrste — opisi opravil (brez velikih slik) za preživetje F5."""
    if KREATIVE_QUEUE_FILE.exists():
        try:
            return json.loads(KREATIVE_QUEUE_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/kreative-queue")
async def save_kreative_queue(data: dict):
    try:
        queue = data.get("queue", [])
        KREATIVE_QUEUE_FILE.write_text(json.dumps(queue, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(queue)}
    except Exception as e:
        return {"error": str(e)}

@app.post("/kreative-queue-images")
async def save_kreative_queue_images(data: dict):
    """Shrani slike (results) enega opravila vrste v svojo datoteko (preživi F5)."""
    try:
        jid = _qimg_safe_id(data.get("id") or "")
        results = data.get("results") or []
        if not jid:
            return JSONResponse({"error": "Manjka ID"}, status_code=400)
        (KREATIVE_QUEUE_IMG_DIR / f"{jid}.json").write_text(
            json.dumps({"results": results}, ensure_ascii=False), encoding="utf-8")
        count = sum(len(r.get("images") or []) for r in results)
        return {"ok": True, "id": jid, "count": count}
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/kreative-queue-images/{jid}")
async def get_kreative_queue_images(jid: str):
    """Naloži slike enega opravila vrste (ob kliku / po F5)."""
    sid = _qimg_safe_id(jid)
    f = KREATIVE_QUEUE_IMG_DIR / f"{sid}.json"
    if not f.exists():
        return {"results": []}
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except Exception:
        return {"results": []}

@app.delete("/kreative-queue-images/{jid}")
async def del_kreative_queue_images(jid: str):
    sid = _qimg_safe_id(jid)
    try:
        f = KREATIVE_QUEUE_IMG_DIR / f"{sid}.json"
        if f.exists():
            f.unlink()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


# ── Shranjena opravila iz vrste (vsako svoja datoteka s slikami) ──
KREATIVE_SAVED_DIR = DATA_DIR / "kreative_saved"
KREATIVE_SAVED_DIR.mkdir(exist_ok=True, parents=True)
KREATIVE_SAVED_INDEX = KREATIVE_SAVED_DIR / "_index.json"

def _ksaved_load_index():
    if KREATIVE_SAVED_INDEX.exists():
        try:
            return json.loads(KREATIVE_SAVED_INDEX.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []

def _ksaved_write_index(idx):
    KREATIVE_SAVED_INDEX.write_text(json.dumps(idx, ensure_ascii=False, indent=2), encoding="utf-8")

def _ksaved_safe_id(jid):
    # dovoli le varne znake za ime datoteke
    return "".join(c for c in str(jid) if c.isalnum() or c in "-_")[:64]

@app.get("/kreative-saved")
async def kreative_saved_list():
    """Vrne LAHKI seznam shranjenih opravil (brez vseh slik — samo opis + 1 sličica)."""
    return _ksaved_load_index()

@app.get("/kreative-saved/{jid}")
async def kreative_saved_get(jid: str):
    """Naloži celo shranjeno opravilo s slikami (ob kliku)."""
    sid = _ksaved_safe_id(jid)
    f = KREATIVE_SAVED_DIR / f"{sid}.json"
    if not f.exists():
        return JSONResponse({"error": "Ni najdeno"}, status_code=404)
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/kreative-saved")
async def kreative_saved_add(data: dict):
    """Shrani opravilo (z rezultati/slikami) v svojo datoteko + dopiše v index."""
    try:
        job = data.get("job") or {}
        jid = _ksaved_safe_id(job.get("id") or str(int(datetime.now().timestamp())))
        if not jid:
            return JSONResponse({"error": "Neveljaven ID"}, status_code=400)
        # cela datoteka s slikami
        (KREATIVE_SAVED_DIR / f"{jid}.json").write_text(json.dumps(job, ensure_ascii=False), encoding="utf-8")
        # lahek vnos v index (brez vseh slik — samo 1 sličica)
        results = job.get("results") or []
        thumb = ""
        for r in results:
            imgs = r.get("images") or []
            if imgs:
                thumb = imgs[0]; break
        result_count = sum(len(r.get("images") or []) for r in results)
        idx = _ksaved_load_index()
        idx = [e for e in idx if e.get("id") != jid]  # dedup
        idx.insert(0, {
            "id": jid,
            "name": job.get("name") or "Izdelek",
            "date": job.get("date") or _lj_now().strftime("%d.%m.%Y %H:%M"),
            "saved_at": _lj_now().isoformat(),
            "result_count": result_count,
            "aCount": len(job.get("aOpts") or []),
            "bCount": len(job.get("bOpts") or []),
            "thumb": thumb,
        })
        _ksaved_write_index(idx)
        return {"ok": True, "id": jid, "count": len(idx)}
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)

@app.delete("/kreative-saved/{jid}")
async def kreative_saved_delete(jid: str):
    sid = _ksaved_safe_id(jid)
    try:
        f = KREATIVE_SAVED_DIR / f"{sid}.json"
        if f.exists():
            f.unlink()
        idx = [e for e in _ksaved_load_index() if e.get("id") != sid]
        _ksaved_write_index(idx)
        return {"ok": True, "count": len(idx)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── KREATIVE ENDPOINTS ───────────────────────────────────────────────────────

@app.post("/analyze-product-kreative")
async def analyze_product_kreative(data: dict):
    """Prebere stran izdelka in generira A/B/C opcije za kreative."""
    url = data.get("url", "").strip()
    if not url:
        return {"error": "Manjka URL."}

    # Fetch page
    try:
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as hc:
            resp = await hc.get(url, headers={"User-Agent": "Mozilla/5.0"})
            html = resp.text if resp.status_code == 200 else ""
    except Exception as e:
        return {"error": f"Ne morem prebrati strani: {e}"}


    # Build prompt for Claude to analyze product
    tools = [{"type": "web_search_20250305", "name": "web_search"}]
    analysis_prompt = f"""Preberi to spletno stran izdelka: {url}

Na podlagi opisa izdelka generiraj strukturirane podatke za kreative FB oglasov.

Vrni SAMO JSON (brez markdown) v tej obliki:
{{
  "name": "IME_IZDELKA_CAPS",
  "aOptions": [
    {{"label": "MAIN VERSION", "text": "benefit1, benefit2, benefit3"}},
    {{"label": "HIGH-CONVERT", "text": "benefit1, benefit2, benefit3"}},
    {{"label": "PROBLEM-SOLUTION", "text": "benefit1, benefit2, benefit3"}},
    {{"label": "ULTRA SIMPLE", "text": "benefit1, benefit2"}},
    {{"label": "SOCIAL PROOF", "text": "benefit1, benefit2, benefit3"}}
  ],
  "bOptions": [
    {{"label": "BEST", "text": "kratko ime vibeа ozadja"}},
    {{"label": "SECOND OPTION", "text": "kratko ime vibeа ozadja"}},
    {{"label": "SCROLL STOPPER", "text": "kratko ime vibeа ozadja"}},
    {{"label": "BONUS", "text": "kratko ime vibeа ozadja"}},
    {{"label": "LIFESTYLE", "text": "kratko ime vibeа ozadja"}}
  ]
}}

PRAVILA:
- name: samo ime blagovne znamke/modela z CAPS
- aOptions: PRODAJNI UDARCI (ne lastnosti!), max 2-3 besede vsak, v angleščini
- bOptions: KONCEPT/VIBE specifičen za TA izdelek, max 3 besede, vedno akcija ali moment
- VSE mora biti v angleščini
- Vrni točno 5 aOptions in 5 bOptions

PRIMERI (uči se iz teh vzorcev):

WEEDZAP (weed removal tool):
aOptions: "Pull the Root, Stop the Weed, No Chemicals" | "Weeds Gone, Root and All, One Tool" | "Weeds Keep Coming Back, Pull the Root, Done for Good" | "Twist, Pull, Gone"
bOptions: "root pull satisfying moment" | "no chemicals angle" | "before/after garden" | "no back pain angle"

ASHIRAFLUX (smokeless fire pit):
aOptions: "Real Flame, No Smoke, Anywhere You Are" | "No Fireplace? No Problem, Instant Cozy, Zero Smoke" | "Missing That Fire Feeling, Warm Vibes Instantly" | "Fill, Light, Relax"
bOptions: "evening atmosphere shot" | "smoke vs no smoke" | "first light moment" | "indoor safe angle"

STAXA (steel organizer shelf):
aOptions: "Double Your Space, Zero Clutter, Instant Order" | "Messy Counter, One Shelf, Problem Solved" | "No Room, Stack It Up, Space Created" | "Stack, Store, Done"
bOptions: "before/after counter" | "multi-room tour" | "satisfying load test" | "steel vs plastic"

SIZZELA (electric frying pan):
aOptions: "No Stove Needed, Cook Anywhere, Instant Heat" | "One Pan, Any Meal, Zero Hassle" | "No Stove, No Smoke, No Problem" | "Plug & Sizzle, Done"
bOptions: "sizzle sound moment" | "speed cooking demo" | "steam lid reveal" | "no stove freedom"

PLANTDRILL (garden auger):
aOptions: "Drill, Plant, Done, No Digging" | "One Bit, Perfect Holes, Zero Effort" | "Back-Breaking Digging, One Drill Bit, Done" | "Drill, Drop, Grow"
bOptions: "speed demo" | "planting demo" | "satisfying drill moment"

SMARTFITNESS (EMS stimulator):
aOptions: "Train Anywhere, No Gym, Real Results" | "On the Couch, Still Training, Zero Effort" | "No Time to Work Out, Wear It, Feel It Work" | "Stick, Activate, Tone"
bOptions: "lifestyle demo" | "before/after body" | "reaction moment"

SOWSYNC (seed spacing tool):
aOptions: "Plant Smart, Perfect Spacing, Grow Better" | "Even Rows, No Waste, Strong Growth" | "Messy Planting, Seed Tool, Better Yield" | "Place, Press, Plant"
bOptions: "before/after planting" | "planting demo" | "grid effect"

Sedaj generiraj za izdelek na tej strani po ISTEM vzorcu:"""

    text = await call_claude(analysis_prompt, "claude-sonnet-4-6", tools, 800)
    result = parse_json_response(text)

    if not result:
        return {"error": "Ni uspelo analizirati izdelka. Poskusi znova."}

    return result


@app.post("/fetch-product-images")
async def fetch_product_images(data: dict):
    """Pobere slike izdelka s podane URL strani."""
    url = data.get("url", "").strip()
    if not url:
        return {"error": "Manjka URL."}
    try:
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as hc:
            resp = await hc.get(url, headers={"User-Agent": "Mozilla/5.0"})
            if resp.status_code != 200:
                return {"error": f"Stran ni dostopna ({resp.status_code})."}
            html = resp.text

        # Extract img src attributes
        import re as _re
        srcs = _re.findall(r'<img[^>]+src=["\']([^"\']+)["\']', html, _re.IGNORECASE)

        # Filter: only product-like images (skip icons, logos, tiny tracking pixels)
        from urllib.parse import urljoin
        base = url
        images = []
        seen = set()
        for src in srcs:
            if not src or src.startswith("data:"):
                continue
            full = urljoin(base, src)
            # Skip obviously non-product images
            skip_words = ["logo", "icon", "favicon", "sprite", "banner", "flag",
                         "payment", "badge", "star", "rating", "arrow", "tracking"]
            if any(w in full.lower() for w in skip_words):
                continue
            # Only jpg/png/webp
            if not any(ext in full.lower() for ext in [".jpg", ".jpeg", ".png", ".webp"]):
                continue
            if full not in seen:
                seen.add(full)
                images.append(full)
            if len(images) >= 20:
                break

        return {"images": images, "count": len(images)}
    except Exception as e:
        return {"error": str(e)}


@app.post("/generate-kreative")
async def generate_kreative(data: dict):
    """Generira kreative z Google Gemini (Nano Banana 2) API."""
    import base64, struct, zlib

    product_name = data.get("productName", "")
    a_options = data.get("aOptions", [])
    b_options = data.get("bOptions", [])
    count = data.get("count", 4)
    ref_images = data.get("images", [])  # base64 data URLs

    if not product_name or not a_options or not b_options:
        return {"error": "Manjkajo podatki (ime izdelka, A ali B opcije)."}

    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    if not gemini_key:
        return {"error": "GEMINI_API_KEY ni nastavljen v Render environment variables."}

    # Globalni izbrani model (fallback, če B-opcija nima svojega): 'flash'|'pro'|'image2'
    _model_choice = (data.get("model") or "flash").lower()

    # ── GEMINI FILE URI CACHE ──────────────────────────────────────────────────
    # Gemini Files API vrne URI ki velja 48h — cachiramo na disk da ne uploadamo vsakič
    import hashlib, json as _json, time as _time
    GEMINI_CACHE_FILE = DATA_DIR / "gemini_file_cache.json"

    def _load_cache():
        try:
            if GEMINI_CACHE_FILE.exists():
                return _json.loads(GEMINI_CACHE_FILE.read_text(encoding="utf-8"))
        except: pass
        return {}

    def _save_cache(cache):
        try: GEMINI_CACHE_FILE.write_text(_json.dumps(cache, ensure_ascii=False), encoding="utf-8")
        except: pass

    def _img_hash(b64_data):
        # Hash samo prvih 2000 znakov (dovolj za unikatnost, hitro)
        return hashlib.md5(b64_data[:2000].encode()).hexdigest()

    gemini_cache = _load_cache()
    now_ts = _time.time()
    cache_updated = False

    # Build combinations — vsak combo dobi PROCESOR iz svoje B-opcije (vibe/ozadje)
    # b.model: 'flash' (Nano Banana 2) | 'pro' (Nano Banana Pro) | 'image2' (GPT Image 2)
    # POMEMBNO: pri image2 ozadju omejimo GPT Image 2 na max 1 sliko SKUPAJ za ta vibe
    # (čez vse A-tekste), ostale slike tega vibe-a naredi NB2 (hitreje/ceneje).
    # image2 je konsistenten (malo variira), zato 1 slika zadošča.
    IMAGE2_CAP = 1
    combos = []
    for b in b_options:
        b_model = (b.get("model") or _model_choice or "flash").lower()
        # koliko slik skupaj ima ta vibe: št. A-tekstov × count
        prompts_for_b = []
        for a in a_options:
            prompt = (
                f"From these reference images create a new FB ad creative. "
                f"Try a more intense '{b.get('text', '')}' background. "
                f"Do not include any text/words on the image except the device name in capital letters '{product_name}' — place it where it fits best or makes sense. "
                f"If possible (if you recognize any suitable English naming styles), you can also create a logo from the name. "
                f"Highlight (can be through icons or text in English) that it is: {a.get('text', '')}. "
                f"Keep all text and icons well within the image borders — nothing should be cut off at the edges. Square 1:1 format."
            )
            prompts_for_b.append((a, prompt))

        if b_model == "image2":
            # razdeli IMAGE2_CAP image2 slik čez kombinacije, ostalo NB2
            image2_left = IMAGE2_CAP
            for a, prompt in prompts_for_b:
                n_img2 = min(count, image2_left)
                n_nb2 = count - n_img2
                image2_left -= n_img2
                if n_img2 > 0:
                    combos.append({"combo": f"{a.get('label','A')} × {b.get('label','B')}", "prompt": prompt, "model": "image2", "n_images": n_img2})
                if n_nb2 > 0:
                    combos.append({"combo": f"{a.get('label','A')} × {b.get('label','B')}", "prompt": prompt, "model": "flash", "n_images": n_nb2})
        else:
            for a, prompt in prompts_for_b:
                combos.append({"combo": f"{a.get('label','A')} × {b.get('label','B')}", "prompt": prompt, "model": b_model, "n_images": count})

    # Prepare reference image parts for Gemini
    # Upload referenčne slike enkrat na Gemini Files API — z 48h cache
    file_uris = []
    if ref_images:
        async with httpx.AsyncClient(timeout=60.0) as hc:
            for img_data in ref_images:  # vse referenčne slike, cache poskrbi za optimizacijo
                try:
                    if "," in img_data:
                        header, b64 = img_data.split(",", 1)
                        mime = header.split(":")[1].split(";")[0]
                    else:
                        b64 = img_data
                        mime = "image/jpeg"

                    # Preveri cache — če URI še velja (< 46h star), preskoči upload
                    img_key = _img_hash(b64)
                    cached = gemini_cache.get(img_key)
                    if cached and (now_ts - cached.get("ts", 0)) < 46 * 3600:
                        # Cache hit — ne uploadamo znova
                        file_uris.append({"fileData": {"mimeType": cached["mime"], "fileUri": cached["uri"]}})
                        continue

                    # Cache miss — uploadaj
                    img_bytes = __import__("base64").b64decode(b64)
                    upload_url = f"https://generativelanguage.googleapis.com/upload/v1beta/files?key={gemini_key}"
                    resp = await hc.post(
                        upload_url,
                        content=img_bytes,
                        headers={"Content-Type": mime, "X-Goog-Upload-Content-Type": mime,
                                 "X-Goog-Upload-Protocol": "raw"}
                    )
                    if resp.status_code == 200:
                        uri = resp.json().get("file", {}).get("uri", "")
                        if uri:
                            file_uris.append({"fileData": {"mimeType": mime, "fileUri": uri}})
                            # Shrani v cache
                            gemini_cache[img_key] = {"uri": uri, "mime": mime, "ts": now_ts}
                            cache_updated = True
                except Exception:
                    continue

        if cache_updated:
            _save_cache(gemini_cache)

    # Fallback: če Files API ne dela, uporabi inline za prvo sliko
    if file_uris:
        image_parts = file_uris
    elif ref_images:
        try:
            img_data = ref_images[0]
            if "," in img_data:
                header, b64 = img_data.split(",", 1)
                mime = header.split(":")[1].split(";")[0]
            else:
                b64 = img_data; mime = "image/jpeg"
            image_parts = [{"inline_data": {"mime_type": mime, "data": b64}}]
        except Exception:
            image_parts = []
    else:
        image_parts = []

    # Za GPT Image 2 (images/edits) potrebujemo SUROVE bajte referenčne slike (multipart)
    openai_key = os.environ.get("OPENAI_API_KEY", "")
    _ref_raw = None  # (bytes, mime, filename)
    if ref_images:
        try:
            _img0 = ref_images[0]
            if "," in _img0:
                _hdr, _b64 = _img0.split(",", 1)
                _mime = _hdr.split(":")[1].split(";")[0]
            else:
                _b64 = _img0; _mime = "image/jpeg"
            _ext = "png" if "png" in _mime else "jpg"
            _ref_raw = (__import__("base64").b64decode(_b64), _mime, f"ref.{_ext}")
        except Exception:
            _ref_raw = None

    async def generate_one_gemini(combo_prompt, model_key):
        model_id = "gemini-3-pro-image" if model_key == "pro" else "gemini-3.1-flash-image-preview"
        api_url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_id}:generateContent?key={gemini_key}"
        parts = list(image_parts) + [{"text": combo_prompt}]
        payload = {"contents": [{"parts": parts}], "generationConfig": {"responseModalities": ["IMAGE", "TEXT"]}}
        try:
            async with httpx.AsyncClient(timeout=120.0) as hc:
                resp = await hc.post(api_url, json=payload, headers={"Content-Type": "application/json"})
                result = resp.json()
            if resp.status_code != 200:
                return None, result.get("error", {}).get("message", str(result))
            for candidate in result.get("candidates", []):
                for part in candidate.get("content", {}).get("parts", []):
                    if "inlineData" in part:
                        inline = part["inlineData"]
                        return f"data:{inline.get('mimeType','image/png')};base64,{inline.get('data','')}", None
            return None, "Gemini ni vrnil slike: " + str(result)[:150]
        except Exception as e:
            return None, str(e)

    async def generate_one_image2(combo_prompt):
        """GPT Image 2 prek OpenAI images/edits (multipart). Referenčna slika obvezna."""
        if not openai_key:
            return None, "OPENAI_API_KEY ni nastavljen."
        if not _ref_raw:
            return None, "GPT Image 2 potrebuje referenčno sliko."
        try:
            files = {"image[]": (_ref_raw[2], _ref_raw[0], _ref_raw[1])}
            form = {"model": "gpt-image-2", "prompt": combo_prompt, "size": "1024x1024", "n": "1", "output_format": "jpeg"}
            async with httpx.AsyncClient(timeout=180.0) as hc:
                resp = await hc.post(
                    "https://api.openai.com/v1/images/edits",
                    headers={"Authorization": f"Bearer {openai_key}"},
                    data=form, files=files,
                )
                result = resp.json()
            if resp.status_code != 200:
                return None, result.get("error", {}).get("message", str(result))[:200]
            data_arr = result.get("data", [])
            if data_arr and data_arr[0].get("b64_json"):
                return f"data:image/jpeg;base64,{data_arr[0]['b64_json']}", None
            return None, "GPT Image 2 ni vrnil slike: " + str(result)[:150]
        except Exception as e:
            return None, str(e)

    async def generate_one_image(combo_prompt, model_key, idx):
        if model_key == "image2":
            return await generate_one_image2(combo_prompt)
        return await generate_one_gemini(combo_prompt, model_key)

    async def generate_combo(combo):
        """Generate slike za eno kombinacijo vzporedno — model in št. slik iz comboja."""
        mk = combo.get("model", "flash")
        n = combo.get("n_images", count)
        tasks = [generate_one_image(combo["prompt"], mk, i) for i in range(n)]
        img_results = await asyncio.gather(*tasks)
        imgs = [img for img, err in img_results if img]
        errors = [err for img, err in img_results if err]
        if not imgs:
            return {"combo": combo["combo"], "model": mk, "images": [], "error": errors[0] if errors else "Ni slike"}
        return {"combo": combo["combo"], "model": mk, "images": imgs}

    # Vse kombinacije + vse slike vzporedno
    results = await asyncio.gather(*[generate_combo(combo) for combo in combos])
    return {"results": list(results)}


# ─── ASANA ENDPOINTS ──────────────────────────────────────────────────────────

ASANA_API = "https://app.asana.com/api/1.0"

def asana_headers():
    token = os.environ.get("ASANA_API_KEY", "")
    return {"Authorization": f"Bearer {token}", "Accept": "application/json"}

@app.get("/asana-search")
async def asana_search(q: str = ""):
    """Išče Asana taske po imenu."""
    if not q:
        return {"tasks": []}
    token = os.environ.get("ASANA_API_KEY", "")
    if not token:
        return {"error": "ASANA_API_KEY ni nastavljen."}
    try:
        async with httpx.AsyncClient(timeout=15.0) as hc:
            # Get workspaces first
            ws_resp = await hc.get(f"{ASANA_API}/workspaces", headers=asana_headers())
            workspaces = ws_resp.json().get("data", [])
            if not workspaces:
                return {"error": "Ni najdenih workspace-ov."}
            ws_gid = workspaces[0]["gid"]

            # Search tasks
            params = {"workspace": ws_gid, "text": q, "resource_type": "task",
                      "opt_fields": "gid,name,projects.name"}
            search_resp = await hc.get(f"{ASANA_API}/workspaces/{ws_gid}/tasks/search",
                                        params=params, headers=asana_headers())
            tasks_raw = search_resp.json().get("data", [])

            tasks = []
            for t in tasks_raw[:10]:
                projects = t.get("projects", [])
                proj_name = projects[0]["name"] if projects else ""
                tasks.append({"gid": t["gid"], "name": t["name"], "project": proj_name})

            return {"tasks": tasks}
    except Exception as e:
        return {"error": str(e)}


@app.post("/asana-attach")
async def asana_attach(data: dict):
    """Priloži slike (base64 data URLs) na Asana task."""
    task_id = data.get("task_id", "")
    image_urls = data.get("image_urls", [])

    if not task_id or not image_urls:
        return {"error": "Manjka task_id ali slike."}

    token = os.environ.get("ASANA_API_KEY", "")
    if not token:
        return {"error": "ASANA_API_KEY ni nastavljen."}

    attached = 0
    errors = []

    async with httpx.AsyncClient(timeout=60.0) as hc:
        for i, img_url in enumerate(image_urls):
            try:
                # Decode base64 data URL
                if img_url.startswith("data:"):
                    header, b64data = img_url.split(",", 1)
                    mime = header.split(":")[1].split(";")[0]
                    ext = mime.split("/")[1] if "/" in mime else "png"
                    img_bytes = __import__("base64").b64decode(b64data)
                else:
                    # Regular URL — fetch it
                    img_resp = await hc.get(img_url)
                    img_bytes = img_resp.content
                    mime = img_resp.headers.get("content-type", "image/png")
                    ext = "png"

                filename = f"kreativa_{i+1}.{ext}"

                # Upload to Asana as attachment
                files = {"file": (filename, img_bytes, mime)}
                attach_resp = await hc.post(
                    f"{ASANA_API}/tasks/{task_id}/attachments",
                    headers={"Authorization": f"Bearer {token}"},
                    files=files
                )

                if attach_resp.status_code in (200, 201):
                    attached += 1
                else:
                    errors.append(f"Slika {i+1}: {attach_resp.text[:100]}")

            except Exception as e:
                errors.append(f"Slika {i+1}: {str(e)}")

    return {"attached": attached, "errors": errors, "total": len(image_urls)}


# ─── LOKALIZACIJA ENDPOINT ───────────────────────────────────────────────────

LANG_NAMES = {
    "HR": "Croatian", "RS": "Serbian (Latin script)",
    "HU": "Hungarian", "CZ": "Czech", "SK": "Slovak",
    "PL": "Polish", "RO": "Romanian", "BG": "Bulgarian",
    "GR": "Greek", "SL": "Slovenian"
}

# Per-jezik opozorila za image prevod (Gemini meša sorodne jezike).
# Poudarek na pogostih pasteh + nekaj ključnih pravilnih besed.
LANG_TRANSLATE_HINTS = {
    "SL": ("This is SLOVENIAN, NOT Czech or Slovak. "
           "Common words: done/finished='GOTOVO' (never 'hotovo' which is Czech), "
           "blow='PIHNI', cover='POKRIJ/PREKRIJ', now='ZDAJ', new='NOVO', "
           "buy='KUPI', order='NAROČI'. Use Slovenian diacritics č, š, ž correctly."),
    "HR": ("This is CROATIAN, NOT Serbian or Slovenian. "
           "Use Croatian forms: 'tko' not 'ko', 'što' not 'šta'. "
           "done='GOTOVO', buy='KUPI', now='SADA', new='NOVO'. Diacritics: č, ć, š, ž, đ."),
    "RS": ("This is SERBIAN in LATIN script (not Cyrillic, not Croatian). "
           "Use Serbian forms: 'šta' not 'što'. done='GOTOVO', buy='KUPI', now='SADA'."),
    "SK": ("This is SLOVAK, NOT Czech or Slovenian. "
           "done='HOTOVO', buy='KÚPIŤ', now='TERAZ'. Slovak-specific letters: ľ, ô, ŕ."),
    "CZ": ("This is CZECH, NOT Slovak or Slovenian. "
           "done='HOTOVO', buy='KOUPIT', now='TEĎ/NYNÍ'. Czech-specific: ř, ě, ů."),
    "PL": ("This is POLISH. done='GOTOWE', buy='KUP', now='TERAZ'. Letters: ą, ę, ł, ń, ó, ś, ź, ż."),
    "HU": ("This is HUNGARIAN (not a Slavic language). done='KÉSZ', buy='VÁSÁROLJ', now='MOST'. Letters: á, é, í, ó, ö, ő, ú, ü, ű."),
    "RO": ("This is ROMANIAN. done='GATA', buy='CUMPĂRĂ', now='ACUM'. Letters: ă, â, î, ș, ț."),
    "BG": ("This is BULGARIAN in CYRILLIC script. done='ГОТОВО', buy='КУПИ', now='СЕГА'. Use Cyrillic only."),
    "GR": ("This is GREEK in Greek script. done='ΕΤΟΙΜΟ', buy='ΑΓΟΡΑΣΕ', now='ΤΩΡΑ'. Use Greek alphabet only."),
}

@app.post("/localize-kreativa")
async def localize_kreativa(data: dict):
    """Prevede tekst na hero kreativu v izbrane jezike z Gemini."""
    image_data = data.get("image", "") or (data.get("images", [None])[0] or "")
    images_data = data.get("images", [])
    if not images_data and image_data:
        images_data = [image_data]
    languages = data.get("languages", [])
    asana_task_id = data.get("asana_task_id")
    sku = data.get("sku", "SKU").strip().upper()
    brand = data.get("brand", "").strip()

    if not images_data or not languages:
        return {"error": "Manjka slika ali jeziki."}

    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    if not gemini_key:
        return {"error": "GEMINI_API_KEY ni nastavljen."}

    # Decode base64 image
    try:
        if "," in image_data:
            header, b64 = image_data.split(",", 1)
            mime = header.split(":")[1].split(";")[0]
        else:
            b64 = image_data
            mime = "image/jpeg"
    except Exception as e:
        return {"error": f"Napaka pri dekodiranju slike: {e}"}

    async def translate_one(lang_code, img_b64, img_mime, img_idx):
        lang_name = LANG_NAMES.get(lang_code, lang_code)
        api_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3.1-flash-image-preview:generateContent?key={gemini_key}"
        brand_note = f" Do NOT translate these brand/product names (keep exactly as-is): {brand}." if brand else ""
        # Per-jezik opozorila — preprečijo mešanje sorodnih jezikov (npr. SL "hotovo" je češko)
        lang_hint = LANG_TRANSLATE_HINTS.get(lang_code, "")
        prompt = (
            f"Edit this image to translate all visible text into {lang_name}. "
            f"Keep EVERYTHING exactly the same — same people, same background, same layout, same design, same product, same icons, same colors, same fonts. "
            f"ONLY translate the text that is NOT a brand name or logo."
            f"{brand_note} "
            f"Do NOT translate or modify any brand names, logos, or product names. "
            f"Keep all text in the same position, same style, same size. "
            f"CRITICAL TRANSLATION QUALITY: Use ONLY correct, native {lang_name}. "
            f"Do NOT use words from other similar/neighboring languages. "
            f"Translate the MEANING correctly, not phonetically. "
            f"{lang_hint}"
        )
        payload = {
            "contents": [{"parts": [
                {"inline_data": {"mime_type": img_mime, "data": img_b64}},
                {"text": prompt}
            ]}],
            "generationConfig": {"responseModalities": ["IMAGE", "TEXT"]}
        }
        try:
            async with httpx.AsyncClient(timeout=120.0) as hc:
                resp = await hc.post(api_url, json=payload, headers={"Content-Type": "application/json"})
                result = resp.json()
            if resp.status_code != 200:
                return {"lang": lang_code, "lang_name": lang_name, "url": None,
                        "error": result.get("error", {}).get("message", str(result))[:200]}
            for candidate in result.get("candidates", []):
                for part in candidate.get("content", {}).get("parts", []):
                    if "inlineData" in part:
                        out_mime = part["inlineData"].get("mimeType", "image/png")
                        out_b64 = part["inlineData"].get("data", "")
                        img_url = f"data:{out_mime};base64,{out_b64}"
                        filename = f"{sku}_{lang_code}_v{img_idx}.png"
                        asana_ok = False
                        if asana_task_id:
                            try:
                                img_bytes = __import__("base64").b64decode(out_b64)
                                token = os.environ.get("ASANA_API_KEY", "")
                                async with httpx.AsyncClient(timeout=30.0) as hc2:
                                    attach_resp = await hc2.post(
                                        f"{ASANA_API}/tasks/{asana_task_id}/attachments",
                                        headers={"Authorization": f"Bearer {token}"},
                                        files={"file": (filename, img_bytes, out_mime)}
                                    )
                                asana_ok = attach_resp.status_code in (200, 201)
                            except Exception:
                                pass
                        return {"lang": lang_code, "lang_name": lang_name, "url": img_url, "filename": filename, "asana_ok": asana_ok}
            return {"lang": lang_code, "lang_name": lang_name, "url": None, "error": "Ni slike"}
        except Exception as e:
            return {"lang": lang_code, "lang_name": lang_name, "url": None, "error": str(e)}

    # Pripravi vse kombinacije slika × jezik vzporedno
    tasks = []
    for img_idx, img_data_item in enumerate(images_data, 1):
        try:
            if "," in img_data_item:
                header, b64 = img_data_item.split(",", 1)
                mime = header.split(":")[1].split(";")[0]
            else:
                b64 = img_data_item; mime = "image/jpeg"
        except Exception:
            continue
        for lang_code in languages:
            tasks.append(translate_one(lang_code, b64, mime, img_idx))
    results = await asyncio.gather(*tasks)
    return {"results": list(results)}


# ─── NAROČILNICE HISTORY ─────────────────────────────────────────────────────

NAROCILNICE_HISTORY_FILE = DATA_DIR / "narocilnice_history.json"

@app.get("/narocilnice-history")
async def get_narocilnice_history():
    if NAROCILNICE_HISTORY_FILE.exists():
        try:
            return json.loads(NAROCILNICE_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/narocilnice-history")
async def save_narocilnice_history(data: dict):
    try:
        history = []
        if NAROCILNICE_HISTORY_FILE.exists():
            try:
                history = json.loads(NAROCILNICE_HISTORY_FILE.read_text(encoding="utf-8"))
            except:
                history = []
        
        csv_text = data.get("csv", "")
        date = data.get("date", "")
        
        # Count negative rows
        rows = 0
        for line in csv_text.split('\n')[1:]:
            if line.strip():
                rows += 1
        
        history.append({"csv": csv_text, "date": date, "rows": rows})
        # Keep last 50
        if len(history) > 50:
            history = history[-50:]
        
        NAROCILNICE_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(history)}
    except Exception as e:
        return {"error": str(e)}


# ─── NAROČILNICE SKU LOOKUP ───────────────────────────────────────────────────

@app.post("/narocilnice-lookup")
async def narocilnice_lookup(data: dict):
    """Poišče URL izdelka v Maaarket XML feedu po SKU ali nazivu."""
    skus = data.get("skus", [])  # list of {sku, naziv}
    
    await ensure_cache_fresh()
    
    results = {}
    sl_feed = feed_by_lang.get("sl", {})
    
    for item in skus:
        sku = item.get("sku", "").strip().upper()
        naziv = item.get("naziv", "").strip().lower()
        
        found_url = None
        
        # Search by SKU in title/id
        for g_id, prod in sl_feed.items():
            prod_title = prod.get("title", "").strip()
            prod_url = prod.get("url", "")
            prod_slug = extract_slug(prod_url) or ""
            
            # Match SKU in title or slug
            if (sku and (sku.lower() in prod_title.lower() or sku.lower() in prod_slug.lower())):
                found_url = prod_url
                break
        
        # Fallback: search by naziv words (first 3 significant words)
        if not found_url and naziv:
            words = [w for w in naziv.split() if len(w) > 3][:3]
            if words:
                best_score = 0
                for g_id, prod in sl_feed.items():
                    prod_title = prod.get("title", "").strip().lower()
                    score = sum(1 for w in words if w in prod_title)
                    if score > best_score and score >= 2:
                        best_score = score
                        found_url = prod.get("url", "")
        
        if found_url:
            results[sku] = found_url
    
    return {"urls": results}


@app.post("/zaloga-sku-images")
async def zaloga_sku_images(data: dict):
    """Vrne slike izdelkov po SKU iz SLO feed-a (za preview v nabiranju).
    Feed nima mpn; SKU je vgrajen v image_link URL, slug = sluggificiran naziv.
    Strategija: SKU v image URL → sluggificiran naziv == feed slug.
    Vhod: {skus: [...], naziv_map: {SKU: naziv}}. Izhod: {images: {SKU: url}}."""
    skus = data.get("skus", [])
    naziv_map = data.get("naziv_map", {})
    if not skus:
        return {"ok": True, "images": {}}
    try:
        idx = sl_image_index or {}
        # Če cache ni pripravljen, NE čakaj na prenos (sicer 17-39s blokada → Render restart).
        # Sproži osvežitev v ozadju in takoj vrni prazno; slike pridejo ob naslednjem klicu.
        if not idx.get("sku_url") and not idx.get("slug"):
            if is_cache_stale():
                asyncio.create_task(ensure_cache_fresh())
            return {"ok": True, "images": {}, "note": "feed se nalaga, poskusi ponovno", "feed_size": 0}
        slug_idx = idx.get("slug", {})
        tslug_idx = idx.get("title_slug", {})
        mpn_idx = idx.get("mpn", {})
        sku_exact = idx.get("sku_exact", {})
        sku_url = idx.get("sku_url", {})  # PREDIZRAČUNAN SKU→slika (O(1))

        out = {}
        matched_by = {"brand_sku": 0, "image_url": 0, "base_sku": 0, "naziv_slug": 0, "slug": 0, "mpn": 0, "none": 0}
        unresolved = []

        # Faza 1: O(1) lookup po predizračunanih indeksih (brand SKU, mpn, sku_url).
        # Nobene linearne preiskave korpusa — zato hitro tudi za 200 SKU × 9000 izdelkov.
        for raw in skus:
            sku = str(raw).strip()
            su = sku.upper()
            sl = sku.lower()
            snorm = _norm_sku(sku)  # podčrtaj→vezaj (ELIPACK_black → elipack-black)
            found = None
            if su in sku_exact:
                found = sku_exact[su]; matched_by["brand_sku"] += 1
            elif su in mpn_idx:
                found = mpn_idx[su]; matched_by["mpn"] += 1
            elif sl in sku_url:
                found = sku_url[sl]; matched_by["image_url"] += 1
            elif snorm in sku_url:
                found = sku_url[snorm]; matched_by["image_url"] += 1
            if found:
                out[sku] = found
            else:
                unresolved.append(sku)

        # Faza 1b: osnovni SKU (child/variant) — če točen SKU ni našel slike,
        # postopno krajšaj z desne (Maaa61black → ... → Maaa61) in preveri v O(1) sku_url.
        # Min dolžina 4, da ne ujamemo prekratke/skupne osnove.
        if unresolved:
            still = []
            for sku in unresolved:
                sl = _norm_sku(sku)  # podčrtaj→vezaj tudi pri base trim
                found = None
                base = sl
                while len(base) > 4:
                    base = base[:-1]
                    if base in sku_url:
                        found = sku_url[base]; break
                if not found and len(sl) > 4 and sl[:4] in sku_url:
                    found = sku_url[sl[:4]]
                if found:
                    out[sku] = found; matched_by["base_sku"] += 1
                else:
                    still.append(sku)
            unresolved = still

        # Faza 2: sluggificiran naziv == feed slug (ali title_slug)
        if unresolved:
            still = []
            for sku in unresolved:
                naziv = naziv_map.get(sku) or ""
                nslug = _sl_slugify(naziv)
                found = None
                if nslug:
                    # točen slug
                    if nslug in slug_idx:
                        found = slug_idx[nslug]
                    elif nslug in tslug_idx:
                        found = tslug_idx[nslug]
                    else:
                        # prefix match (slug se lahko konča z dodatki kot -premium)
                        for s, img in slug_idx.items():
                            if s.startswith(nslug) or nslug.startswith(s):
                                if abs(len(s) - len(nslug)) <= 15:
                                    found = img; break
                if found:
                    out[sku] = found; matched_by["naziv_slug"] += 1
                else:
                    still.append(sku)
            unresolved = still

        for sku in unresolved:
            matched_by["none"] += 1

        resp = {"ok": True, "images": out, "feed_size": len(sku_url),
                "matched": len(out), "total": len(skus), "match_stats": matched_by}
        if data.get("debug"):
            resp["unresolved"] = unresolved[:50]
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e), "images": {}}


@app.post("/zaloga-refresh-feed")
async def zaloga_refresh_feed():
    """Ročna prisilna osvežitev feed cache-a (obide 7-dnevni TTL).
    Uporabno ko dodaš nov izdelek in želiš slike takoj, brez čakanja na tedensko osvežitev.
    Uporablja isti lock kot samodejni prenos, da se ne prekrivata."""
    lock = _get_feed_lock()
    if lock.locked():
        return {"ok": False, "note": "osvežitev že poteka, počakaj trenutek"}
    try:
        async with lock:
            await fetch_all_feeds()
        idx = sl_image_index or {}
        return {
            "ok": True,
            "note": "feed osvežen",
            "feed_size": len(idx.get("img_corpus", [])),
            "sku_exact": len(idx.get("sku_exact", {})),
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


# ── SKENIRANJE / INVENTURA endpointi ──
@app.post("/skeniranje-lookup")
async def skeniranje_lookup(data: dict):
    """Preslikaj eno ali več črtnih kod v dobaviteljevo kodo.
    Vhod: {barcodes: ["5903039773509", ...]}  ali  {barcode: "..."}.
    Izhod: {ok, results: {barcode: {kod, supplier, name} | null}, ready}."""
    await ensure_barcode_fresh()
    raw = data.get("barcodes")
    if raw is None and data.get("barcode") is not None:
        raw = [data.get("barcode")]
    raw = raw or []
    results = {}
    for b in raw:
        nb = _norm_barcode(str(b))
        hit = barcode_index.get(nb)
        results[str(b)] = hit if hit else None
    return {"ok": True, "ready": bool(barcode_index), "count": len(barcode_index), "results": results}


@app.post("/skeniranje-refresh")
async def skeniranje_refresh():
    """Ročna osvežitev barkoda-indeksa (potegne dobaviteljeve XML znova)."""
    lock = _get_barcode_lock()
    if lock.locked():
        return {"ok": False, "note": "osvežitev že poteka"}
    try:
        async with lock:
            await fetch_barcode_feeds()
        return {"ok": True, "note": "barkode osvežene", "count": len(barcode_index)}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/skeniranje-status")
async def skeniranje_status():
    """Diagnostika: koliko barkod je naloženih + primer."""
    await ensure_barcode_fresh()
    sample = dict(list(barcode_index.items())[:3])
    return {"ok": True, "count": len(barcode_index),
            "last_fetch": barcode_last_fetch.isoformat() if barcode_last_fetch else None,
            "sample": sample}


@app.get("/zaloga-sku-debug")
async def zaloga_sku_debug(skus: str = "", nazivi: str = ""):
    """Hitri pregled: /zaloga-sku-debug?skus=silux38,KX7907&nazivi=Naziv1|Naziv2
    Pove za vsak SKU kako (in če) se najde slika."""
    await ensure_cache_fresh()
    sl_feed = feed_by_lang.get("sl", {})
    sku_list = [s.strip() for s in skus.split(",") if s.strip()]
    naziv_list = [n.strip() for n in nazivi.split("|")] if nazivi else []
    out = {"feed_size": len(sl_feed), "results": {}}
    if not sku_list:
        return {"napotek": "uporabi ?skus=SKU1,SKU2 (in po želji &nazivi=Naziv1|Naziv2)", **out}
    for i, sku in enumerate(sku_list):
        naziv = naziv_list[i] if i < len(naziv_list) else ""
        nslug = _sl_slugify(naziv)
        su = sku.upper()
        hits = []
        for g_id, prod in sl_feed.items():
            img = prod.get("image", "")
            all_imgs = prod.get("all_images") or ([img] if img else [])
            joined = " ".join(all_imgs).lower()
            slug = (extract_slug(prod.get("url", "")) or "").lower()
            bsku = _extract_brand_sku(prod.get("brand", ""), img)
            via = None
            if bsku and bsku == su:
                via = f"brand_sku({prod.get('brand')})"
            elif _sku_in_image_url(sku, joined):
                via = "image_url"
            elif nslug and (slug == nslug or _sl_slugify(prod.get("title","")) == nslug):
                via = "naziv_slug"
            if via:
                hits.append({"g_id": g_id, "brand": prod.get("brand",""),
                             "title": (prod.get("title") or "")[:50],
                             "ima_sliko": bool(img), "kje": via, "img": img[:85]})
        out["results"][sku] = {
            "naziv": naziv, "naziv_slug": nslug,
            "najden": len(hits) > 0, "st_zadetkov": len(hits), "zadetki": hits[:3]
        }
    return out


@app.post("/narocilnice-history-set")
async def set_narocilnice_history(data: dict):
    """Nastavi celotno zgodovino (za brisanje)."""
    try:
        history = data.get("history", [])
        NAROCILNICE_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(history)}
    except Exception as e:
        return {"error": str(e)}



# ─── FORECAST EOD (End-of-day končna naročila) ───────────────────────────────

FORECAST_EOD_FILE = DATA_DIR / "forecast_eod.json"

@app.get("/forecast-eod")
async def get_forecast_eod():
    if FORECAST_EOD_FILE.exists():
        try:
            return json.loads(FORECAST_EOD_FILE.read_text(encoding="utf-8"))
        except:
            pass
    return {}

@app.post("/forecast-eod")
async def save_forecast_eod(data: dict):
    try:
        FORECAST_EOD_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok", "count": len(data)}
    except Exception as e:
        return {"error": str(e)}


# ─── KARANTENA PDF PARSER ─────────────────────────────────────────────────────

from fastapi import UploadFile, File, Form
import io
import re as _re

@app.post("/parse-karantena-pdf")
async def parse_karantena_pdf(file: UploadFile = File(...)):
    """Parsira PDF karantene in vrne strukturirane podatke."""
    try:
        import pdfplumber
        content = await file.read()
        rows = []
        
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            full_text = ""
            for page in pdf.pages:
                # Poskusi extract table najprej
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            if not row or not any(row):
                                continue
                            # Preskoči header vrstice
                            if row[0] and str(row[0]).strip().lower() in ('product_id', 'id', '#'):
                                continue
                            # Normalizacija vrstice
                            cells = [str(c).strip() if c else '' for c in row]
                            if len(cells) >= 3:
                                product_id = cells[0] if cells[0] else ''
                                sku = cells[1] if len(cells) > 1 else ''
                                title = cells[2] if len(cells) > 2 else ''
                                stock = cells[3] if len(cells) > 3 else '0'
                                stock_actual = cells[4] if len(cells) > 4 else '0'
                                position = cells[5] if len(cells) > 5 else ''
                                
                                # Preskoči header
                                if sku.lower() in ('product_sku', 'sku', ''):
                                    continue
                                
                                try: stock = int(float(stock))
                                except: stock = 0
                                try: stock_actual = int(float(stock_actual))
                                except: stock_actual = 0
                                
                                rows.append({
                                    'product_id': product_id,
                                    'sku': sku,
                                    'title': title,
                                    'stock': stock,
                                    'stock_actual': stock_actual,
                                    'position': position,
                                })
                else:
                    full_text += (page.extract_text() or "") + "\n"
            
            # Fallback: text parsing če ni tabel
            if not rows and full_text:
                lines = full_text.split('\n')
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    parts = line.split()
                    if len(parts) >= 3 and parts[0].isdigit():
                        product_id = parts[0]
                        sku = parts[1]
                        # Preskoči header
                        if sku.lower() in ('product_sku', 'sku'):
                            continue
                        # Poišči zadnji del ki je številka (stock)
                        stock = 0
                        stock_actual = 0
                        position = ''
                        title_parts = []
                        for i, p in enumerate(parts[2:], 2):
                            if _re.match(r'^\d+$', p):
                                stock = int(p)
                                if i+1 < len(parts) and _re.match(r'^\d+$', parts[i+1]):
                                    stock_actual = int(parts[i+1])
                                    if i+2 < len(parts):
                                        position = parts[i+2]
                                break
                            else:
                                title_parts.append(p)
                        title = ' '.join(title_parts)
                        rows.append({
                            'product_id': product_id,
                            'sku': sku,
                            'title': title,
                            'stock': stock,
                            'stock_actual': stock_actual,
                            'position': position,
                        })
        
        if not rows:
            return {"error": "Ni podatkov v PDF-u."}
        
        print(f"[karantena] Parsed {len(rows)} rows from PDF")
        return {"rows": rows, "count": len(rows)}
        
    except ImportError:
        return {"error": "pdfplumber ni nameščen. Dodaj ga v requirements.txt."}
    except Exception as e:
        import traceback
        print(f"[karantena] Error: {e}\n{traceback.format_exc()[-500:]}")
        return {"error": str(e)}


# ─── KARANTENA HISTORY ────────────────────────────────────────────────────────

KARANTENA_HISTORY_FILE = DATA_DIR / "karantena_history.json"

@app.get("/karantena-history")
async def get_karantena_history():
    if KARANTENA_HISTORY_FILE.exists():
        try:
            return json.loads(KARANTENA_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/karantena-history")
async def save_karantena_history(data: dict):
    try:
        history = []
        if KARANTENA_HISTORY_FILE.exists():
            try:
                history = json.loads(KARANTENA_HISTORY_FILE.read_text(encoding="utf-8"))
            except:
                history = []
        history.append({
            "rows": data.get("rows", []),
            "filename": data.get("filename", ""),
            "date": data.get("date", "")
        })
        if len(history) > 30:
            history = history[-30:]
        KARANTENA_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok"}
    except Exception as e:
        return {"error": str(e)}

@app.post("/karantena-history-set")
async def set_karantena_history(data: dict):
    try:
        history = data.get("history", [])
        KARANTENA_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok"}
    except Exception as e:
        return {"error": str(e)}


# ─── VIDEO ADS — SCRIPT GENERATION ───────────────────────────────────────────

@app.post("/generate-video-scripts")
async def generate_video_scripts(data: dict):
    input_text = data.get("input", "").strip()
    duration = data.get("duration", 15)
    durations = data.get("durations", [])  # seznam dolžin videov (npr. [18,7,19,8])
    only_en = bool(data.get("only_en", False))  # generiraj SAMO angleščino (univerzalni video)
    if not input_text:
        return {"error": "Manjka vnos."}

    mode = "url" if input_text.startswith("http") else "text"
    tools = [{"type": "web_search_20250305", "name": "web_search"}] if mode == "url" else []

    def _words_for(dur):
        # ElevenLabs govori ~2.6 besede/s. Da govor ZAPOLNI video (ne pusti tišine),
        # ciljamo na ~95% dolžine z 0.5s bufferjem na koncu.
        return max(10, min(150, int((dur - 0.5) * 2.6)))

    # Unikatne dolžine (da ne generiramo isto skripto večkrat)
    unique_durs = sorted(set(int(d) for d in durations if d and d > 0)) if durations else [int(duration)]

    def _build_prompt_en(dur, words):
        return f"""{'Read this page and' if mode == 'url' else 'Based on this description,'} create a voice-over script for a video ad in ENGLISH only.

{'Page: ' + input_text if mode == 'url' else 'Description: ' + input_text}

Rules:
- Target ~{words} words — the speech must FILL a {dur}s video, not too short, not too long
- Natural spoken style, like a friend talking
- Focus on ONE main product benefit, add detail to fill the time
- No prices, no "click", no "order"
- End with a strong statement (not a call to action)
- The script must take ABOUT {dur}s when read aloud (not more!)

Return ONLY JSON without markdown:
{{"product": "product name", "en": "..."}}"""

    def _build_prompt(dur, words):
        if only_en:
            return _build_prompt_en(dur, words)
        return f"""{'Preberi to stran in' if mode == 'url' else 'Na podlagi tega opisa'} ustvari voice over skripte za video oglas v 10 jezikih.

{'Stran: ' + input_text if mode == 'url' else 'Opis: ' + input_text}

Pravila:
- Ciljaj na ~{words} besed na jezik — govor mora ZAPOLNITI {dur}s video, ne prekratek ne predolg
- Naravni govorni slog, kot da govori prijatelj
- Poudarek na eni glavni koristi izdelka, dodaj podrobnosti da zapolniš čas
- Brez cen, brez "klikni", brez "naroči"
- Konec z močno izjavo (ne pozivom k akciji)
- KLJUČNO — tempo govora se razlikuje po jezikih, prilagodi število besed:
  · Hitrejši jeziki (manj besed za isti čas): madžarščina, poljščina, češčina, slovaščina — uporabi ~15% MANJ besed
  · Daljše besede (počasneje): grščina, romunščina, bolgarščina — uporabi ~10% MANJ besed
  · Slovenščina, hrvaščina, srbščina — standardno {words} besed
- Cilj: vsak jezik naj traja PRIBLIŽNO {dur}s ko se prebere naglas (ne več!)
- SL: slovenščina, HR: hrvaščina (latinica), RS: srbščina (SAMO latinica), HU: madžarščina, CZ: češčina, SK: slovaščina, PL: poljščina, GR: grščina (grška pisava), RO: romunščina, BG: bolgarščina (SAMO cirilica)

Vrni SAMO JSON brez markdown:
{{"product": "ime izdelka", "sl": "...", "hr": "...", "rs": "...", "hu": "...", "cz": "...", "sk": "...", "pl": "...", "gr": "...", "ro": "...", "bg": "..."}}"""

    try:
        # Generiraj skripto za VSAKO unikatno dolžino (paralelno)
        async def _gen_for_dur(dur):
            words = _words_for(dur)
            text = await call_claude(_build_prompt(dur, words), "claude-sonnet-4-6", tools if tools else None, 10000)
            parsed = parse_json_response(text)
            return dur, parsed

        results = await asyncio.gather(*[_gen_for_dur(d) for d in unique_durs])

        # scripts_by_dur: {18: {sl:..., hr:...}, 7: {...}}
        scripts_by_dur = {}
        product = ""
        for dur, parsed in results:
            if not parsed:
                continue
            if not product:
                product = parsed.get("product", "")
            scripts_by_dur[str(dur)] = {k: v for k, v in parsed.items() if k != "product"}

        if not scripts_by_dur:
            return {"error": "Napaka pri generiranju skript."}

        # Za kompatibilnost: "scripts" = skripta za najkrajšo dolžino (privzeto)
        first_dur = str(unique_durs[0]) if unique_durs else None
        default_scripts = scripts_by_dur.get(first_dur, {})

        return {
            "scripts": default_scripts,         # privzeto (najkrajši) — kompatibilnost
            "scripts_by_dur": scripts_by_dur,   # NOVO: skripte po dolžini
            "product": product,
        }
    except Exception as e:
        return {"error": str(e)}


# ─── VIDEO ADS — ELEVENLABS AUDIO ────────────────────────────────────────────

ELEVENLABS_VOICES = {
    "sl": "bu5eKETbFKC8G702EAU4",  # Liam — Energetic, Social Media Creator (v3)
    "hr": "FXFcxnjikw0naYO1PPrU",  # Adnan — Energetic, Educational
    "rs": "eWKPI657Btpf4xbqX4x6",
    "hu": "M336tBVZHWWiWb4R54ui",
    "cz": "uYFJyGaibp4N2VwYQshk",
    "sk": "2ST3sI2j7fz4A5oXjnbA",
    "pl": "H5xTcsAIeS5RAykjz57a",
    "gr": "n0vzWypeCK1NlWPVwhOc",
    "ro": "xHIzJ4zBhlGcvJscsdON",
    "bg": "pVnrL6sighQX7hVz89cp",
    "en": "21m00Tcm4TlvDq8ikWAM",  # Rachel — privzeti angleški ženski glas
}

# Angleški glasovi na izbiro (uradni ElevenLabs prednastavljeni ID-ji) — ženski + moški
ELEVENLABS_EN_VOICES = {
    # ženski
    "rachel":  {"id": "21m00Tcm4TlvDq8ikWAM", "name": "♀ Rachel — čist, umirjen (naracija/oglas)", "gender": "f"},
    "bella":   {"id": "EXAVITQu4vr4xnSDxMaL", "name": "♀ Bella — mehek, topel ženski", "gender": "f"},
    "jessica": {"id": "cgSgspJ2msm6clMCkdW9", "name": "♀ Jessica — mlad, energičen (social/oglas)", "gender": "f"},
    # moški
    "adam":    {"id": "pNInz6obpgDQGcFmaJgB", "name": "♂ Adam — globok, avtoritativen (oglas)", "gender": "m"},
    "antoni":  {"id": "ErXwobaYiN019PkySvjV", "name": "♂ Antoni — topel, prijazen (vsestranski)", "gender": "m"},
    "josh":    {"id": "TxGEqnHWrfWFTfGW9XjX", "name": "♂ Josh — mlad, energičen (UGC/social)", "gender": "m"},
}

# Per-language model — SLO gre na novi v3 (z audio tagi), ostali ohranijo v2
ELEVENLABS_MODELS = {
    "sl": "eleven_v3",
    # vsi ostali jeziki: eleven_multilingual_v2 (default)
}

@app.get("/vads-en-voices")
async def vads_en_voices():
    """Seznam angleških glasov za izbiro (ženski + moški)."""
    return {"voices": [{"key": k, "name": v["name"], "gender": v["gender"]} for k, v in ELEVENLABS_EN_VOICES.items()]}

def _parse_words(alignment: dict):
    """Iz ElevenLabs alignment podatkov izloci seznam (beseda, start, end)."""
    chars = alignment.get("characters", [])
    starts = alignment.get("character_start_times_seconds", [])
    ends = alignment.get("character_end_times_seconds", [])
    words = []
    cur_word, cur_start, cur_end = "", None, None
    for i, ch in enumerate(chars):
        if i >= len(starts):
            break
        t_start = starts[i]
        t_end = ends[i] if i < len(ends) else t_start + 0.1
        if ch in (' ', '\n', '\t'):
            if cur_word:
                words.append((cur_word, cur_start, cur_end))
            cur_word, cur_start, cur_end = "", None, None
        else:
            if cur_start is None:
                cur_start = t_start
            cur_end = t_end
            cur_word += ch
    if cur_word:
        words.append((cur_word, cur_start, cur_end))
    return words


def build_srt(alignment: dict) -> str:
    """Generira SRT iz ElevenLabs alignment (za download)."""
    words = _parse_words(alignment)
    if not words:
        return ""

    def fmt(s):
        h, m = int(s // 3600), int((s % 3600) // 60)
        sec, ms = int(s % 60), int((s - int(s)) * 1000)
        return f"{h:02d}:{m:02d}:{sec:02d},{ms:03d}"

    lines, i, idx = [], 0, 1
    while i < len(words):
        grp = [words[i]]
        i += 1
        while i < len(words) and len(grp) < 5 and (words[i][1] - grp[0][1]) < 3.0:
            grp.append(words[i]); i += 1
        lines.append(f"{idx}\n{fmt(grp[0][1])} --> {fmt(grp[-1][2])}\n{' '.join(w[0] for w in grp)}\n")
        idx += 1
    return "\n".join(lines)


def get_subtitle_style_for_format(width: int, height: int) -> dict:
    """Vrne optimalne subtitle nastavitve glede na video format.
    MarginV optimiran za Facebook (Feed + Reels) — podnapisi nad UI cono."""
    if width == 0 or height == 0:
        # Default: assume 9:16
        return {"fontsize": 64, "marginv": 230, "outline": 5, "max_words": 5, "playresx": 1080, "playresy": 1920}

    ratio = width / height

    # MarginV kot % višine — podnapisi so vedno vidni ne glede na crop
    # Za pokončne videe damo podnapise višje (večji %) da ne padejo v FB UI safe-zone
    if ratio < 0.4:
        # Ultra ozki pokončni videi (6:19, 9:21 ipd.)
        # FB ne reže tako agresivno kot TikTok, zato 20% (prej 35%)
        marginv = max(int(height * 0.20), 100)
        return {"fontsize": max(int(height * 0.038), 48), "marginv": marginv, "outline": 5, "max_words": 3, "playresx": width, "playresy": height}
    elif ratio < 0.7:
        # 9:16 vertical (FB Reels + FB Feed prikaz vertikalnega) — 0.5625
        # 16% (prej 12%) — varneje za FB Feed kjer ima vec UI elementov spodaj
        marginv = max(int(height * 0.16), 170)
        return {"fontsize": max(int(height * 0.038), 56), "marginv": marginv, "outline": 5, "max_words": 5, "playresx": width, "playresy": height}
    elif ratio < 1.2:
        # 1:1 square in 4:5 portrait (FB Feed) — najpogostejši FB ad format
        # 12% (prej 7%) — FB feed UI (ime, opis, like/komentar gumbi) prekriva ~10-13% spodnjega dela
        marginv = max(int(height * 0.12), 90)
        return {"fontsize": max(int(height * 0.045), 44), "marginv": marginv, "outline": 4, "max_words": 6, "playresx": width, "playresy": height}
    elif ratio < 1.6:
        # 4:3
        # 10% (prej 7%)
        marginv = max(int(height * 0.10), 70)
        return {"fontsize": max(int(height * 0.05), 40), "marginv": marginv, "outline": 4, "max_words": 6, "playresx": width, "playresy": height}
    else:
        # 16:9 horizontal
        # 8% (prej 6%)
        marginv = max(int(height * 0.08), 50)
        return {"fontsize": max(int(height * 0.055), 36), "marginv": marginv, "outline": 3, "max_words": 7, "playresx": width, "playresy": height}

def build_ass(alignment: dict, video_width: int = 1080, video_height: int = 1920) -> str:
    """Generira ASS karaoke podnapise — Stil B (bela+rumena, debela obroba), prilagojen formatu."""
    words = _parse_words(alignment)
    if not words:
        return ""

    style = get_subtitle_style_for_format(video_width, video_height)
    fontsize = style["fontsize"]
    marginv = style["marginv"]
    outline = style["outline"]
    max_words = style["max_words"]
    playresx = style["playresx"]
    playresy = style["playresy"]

    # ASS header — prilagojen video formatu
    header = f"""[Script Info]
ScriptType: v4.00+
PlayResX: {playresx}
PlayResY: {playresy}

[V4+ Styles]
Format: Name,Fontname,Fontsize,PrimaryColour,SecondaryColour,OutlineColour,BackColour,Bold,Italic,Underline,StrikeOut,ScaleX,ScaleY,Spacing,Angle,BorderStyle,Outline,Shadow,Alignment,MarginL,MarginR,MarginV,Encoding
Style: Default,Arial,{fontsize},&H00FFFFFF,&H00FFD600,&H00000000,&H00000000,-1,0,0,0,100,100,0,0,1,{outline},0,2,30,30,{marginv},1

[Events]
Format: Layer,Start,End,Style,Name,MarginL,MarginR,MarginV,Effect,Text
"""

    def fmt(s):
        h, m = int(s // 3600), int((s % 3600) // 60)
        sec, cs = int(s % 60), int((s - int(s)) * 100)
        return f"{h}:{m:02d}:{sec:02d}.{cs:02d}"

    # Grupiraj v vrstice (max max_words besed)
    lines_out = []
    i = 0
    while i < len(words):
        grp = [words[i]]; i += 1
        while i < len(words) and len(grp) < max_words and (words[i][1] - grp[0][1]) < 2.5:
            grp.append(words[i]); i += 1

        t_in = grp[0][1]
        t_out = grp[-1][2] + 0.05

        # Karaoke: vsaka beseda dobi {\kXX} tag (čas v centisekundah)
        parts = []
        for wi, (word, wstart, wend) in enumerate(grp):
            # Čas do naslednje besede ali konec
            if wi + 1 < len(grp):
                dur_cs = int((grp[wi+1][1] - wstart) * 100)
            else:
                dur_cs = int((wend - wstart) * 100) + 5
            dur_cs = max(dur_cs, 5)
            parts.append("{" + chr(92) + "k" + str(dur_cs) + "}" + word)

        karaoke_text = " ".join(parts)
        lines_out.append(f"Dialogue: 0,{fmt(t_in)},{fmt(t_out)},Default,,0,0,0,,{karaoke_text}")

    return header + "\n".join(lines_out)


@app.post("/generate-audio")
async def generate_audio(data: dict):
    text = data.get("text", "").strip()
    lang = data.get("lang", "sl")
    target_dur = float(data.get("target_duration", 0) or 0)  # prava dolžina TEGA videa (0 = brez)
    if not text:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Manjka tekst."}, status_code=400)

    api_key = os.environ.get("ELEVENLABS_API_KEY", "")
    if not api_key:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "ELEVENLABS_API_KEY ni nastavljen."}, status_code=400)

    voice_id = ELEVENLABS_VOICES.get(lang, "bu5eKETbFKC8G702EAU4")
    # za angleščino: če je izbran specifičen glas, ga uporabi
    if lang == "en":
        sel = (data.get("en_voice") or "rachel").lower()
        if sel in ELEVENLABS_EN_VOICES:
            voice_id = ELEVENLABS_EN_VOICES[sel]["id"]

    def _audio_dur(alignment: dict) -> float:
        ends = alignment.get("character_end_times_seconds", [])
        return float(ends[-1]) if ends else 0.0

    def _shorten(txt: str, ratio: float) -> str:
        """Skrajša tekst na ~ratio dolžine, ohrani cele stavke (ne reže sredi besede)."""
        import re as _re
        ratio = max(0.4, min(0.97, ratio))  # nikoli pod 40% (varovalka proti agresivnemu rezanju)
        words_list = txt.split()
        target_words = max(6, int(len(words_list) * ratio))
        if target_words >= len(words_list):
            return txt
        truncated = " ".join(words_list[:target_words])
        # Nazaj do zadnjega celega stavka
        m = list(_re.finditer(r'[.!?]', truncated))
        if m and m[-1].end() > len(truncated) * 0.55:
            return truncated[:m[-1].end()].strip()
        return truncated.strip().rstrip(',;:') + "."

    async def _call(txt: str, speed: float = 1.0):
        vs = {"stability": 0.5, "similarity_boost": 0.75}
        # speed je podprt v eleven_multilingual_v2 (0.7–1.2). Za v3 ga izpustimo.
        model = ELEVENLABS_MODELS.get(lang, "eleven_multilingual_v2")
        if model != "eleven_v3" and abs(speed - 1.0) > 0.001:
            vs["speed"] = round(max(0.7, min(1.2, speed)), 3)
        async with httpx.AsyncClient(timeout=60.0) as hc:
            return await hc.post(
                f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}/with-timestamps",
                headers={"xi-api-key": api_key, "Content-Type": "application/json"},
                json={"text": txt, "model_id": model, "voice_settings": vs}
            )

    try:
        cur_text = text
        cur_speed = 1.0
        result = None
        shortened = False
        sped_up = False
        tol = 0.3              # govor sme segati do target+0.3s
        MAX_SPEED = 1.15       # naravna meja hitrosti (nad tem zveni hitelo)

        for attempt in range(4):
            resp = await _call(cur_text, cur_speed)
            if resp.status_code != 200:
                from fastapi.responses import JSONResponse
                return JSONResponse({"error": f"ElevenLabs napaka {resp.status_code}: {resp.text[:300]}"}, status_code=400)
            result = resp.json()
            actual = _audio_dur(result.get("alignment", {}))

            # Prilega se → konec
            if target_dur <= 0 or actual <= target_dur + tol:
                print(f"[generate-audio] {lang} #{attempt+1}: {actual:.1f}s (cilj {target_dur:.1f}s, speed {cur_speed:.2f}) ✓")
                break

            over_ratio = actual / target_dur  # koliko presega (npr. 1.20 = 20% predolg)

            # 1. KORAK: poskusi s hitrostjo (do MAX_SPEED) — ohrani vso vsebino
            #    Dejanska potrebna hitrost = trenutna × over_ratio
            needed_speed = cur_speed * over_ratio
            if needed_speed <= MAX_SPEED + 0.001:
                cur_speed = round(needed_speed, 3)
                sped_up = True
                print(f"[generate-audio] {lang} #{attempt+1}: {actual:.1f}s > {target_dur:.1f}s → hitrost {cur_speed:.2f}× (ohranim vsebino)")
                continue

            # 2. KORAK: hitrost ni dovolj → nastavi MAX_SPEED + reži ostanek
            cur_speed = MAX_SPEED
            sped_up = True
            # Po pospešitvi na MAX bo govor ~actual/MAX_SPEED. Koliko še reči?
            after_speed = (actual / over_ratio) * (cur_speed)  # ocena pri novi hitrosti
            # cilj: target. ratio rezanja glede na current text
            ratio = (target_dur / actual) * cur_speed * 0.96
            new_text = _shorten(cur_text, ratio)
            print(f"[generate-audio] {lang} #{attempt+1}: {actual:.1f}s > {target_dur:.1f}s → max hitrost {cur_speed:.2f}× + rez ({ratio:.0%})")
            if new_text == cur_text:
                break
            cur_text = new_text
            shortened = True

        import base64
        audio_b64 = result.get("audio_base64", "")
        alignment = result.get("alignment", {})
        srt = build_srt(alignment)
        ass = build_ass(alignment)

        return {
            "audio_base64": audio_b64,
            "srt": srt,
            "ass": ass,
            "alignment": alignment,
            "lang": lang,
            "duration": _audio_dur(alignment),
            "final_text": cur_text if shortened else text,
            "shortened": shortened,
            "speed": cur_speed,
            "sped_up": sped_up,
        }

    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── MERGE VIDEO + AUDIO ──────────────────────────────────────────────────────

# ─── VIDEO SESSION CACHE — single upload za batch merge ───────────────────────

VIDEO_SESSION_DIR = Path("/tmp/video_sessions")
VIDEO_SESSION_DIR.mkdir(exist_ok=True, parents=True)

@app.post("/upload-video-session")
async def upload_video_session(video: UploadFile = File(...)):
    """Naloži video enkrat, vrni session_id za večkratno uporabo."""
    try:
        import uuid as _u
        session_id = _u.uuid4().hex[:16]
        session_path = VIDEO_SESSION_DIR / f"{session_id}.mp4"
        content_bytes = await video.read()
        session_path.write_bytes(content_bytes)
        # Avtomatsko počisti starejše od 30 min
        try:
            now = datetime.now().timestamp()
            for f in VIDEO_SESSION_DIR.glob("*.mp4"):
                if now - f.stat().st_mtime > 1800:
                    f.unlink()
        except: pass
        print(f"[video-session] uploaded {session_id} ({len(content_bytes)} bytes)")
        return {"session_id": session_id, "size": len(content_bytes)}
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/cleanup-video-session")
async def cleanup_video_session(data: dict):
    """Počisti video session ko ni več potreben."""
    session_ids = data.get("session_ids", [])
    deleted = 0
    for sid in session_ids:
        try:
            p = VIDEO_SESSION_DIR / f"{sid}.mp4"
            if p.exists():
                p.unlink()
                deleted += 1
        except: pass
    return {"deleted": deleted}


@app.post("/merge-video-audio-session")
async def merge_video_audio_session(
    session_id: str = Form(...),
    audio: UploadFile = File(...),
    lang: str = Form("sl"),
    srt: UploadFile = File(None),
):
    """Merge z video iz cached session (faster — video že na strežniku)."""
    import subprocess, tempfile
    from fastapi.responses import JSONResponse as JR
    try:
        import static_ffmpeg
        static_ffmpeg.add_paths()
    except Exception:
        pass

    session_video_path = VIDEO_SESSION_DIR / f"{session_id}.mp4"
    if not session_video_path.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": f"Video session {session_id} not found."}, status_code=404)

    try:
        with tempfile.TemporaryDirectory() as tmp:
            audio_path = f"{tmp}/audio.mp3"
            srt_path = f"{tmp}/subs.srt"
            ass_path = f"{tmp}/subs.ass"
            output_path = f"{tmp}/output_{lang}.mp4"
            video_path = str(session_video_path)

            with open(audio_path, "wb") as f:
                f.write(await audio.read())

            has_srt = False
            ass_content_orig = None
            if srt:
                srt_content = await srt.read()
                if srt_content.strip():
                    with open(srt_path, "wb") as f:
                        f.write(srt_content)
                    if srt_content.startswith(b'[Script Info]'):
                        ass_content_orig = srt_content
                        has_srt = 'ass'
                    else:
                        has_srt = 'srt'

            # Detect video dimensions
            video_width, video_height = 0, 0
            try:
                probe_cmd = ["ffprobe", "-v", "error", "-select_streams", "v:0",
                             "-show_entries", "stream=width,height", "-of", "csv=p=0", video_path]
                probe_result = subprocess.run(probe_cmd, capture_output=True, timeout=10)
                if probe_result.returncode == 0:
                    dims = probe_result.stdout.decode().strip().split(',')
                    if len(dims) == 2:
                        video_width, video_height = int(dims[0]), int(dims[1])
            except: pass

            # Adapt ASS če imamo
            if has_srt == 'ass' and ass_content_orig and video_width > 0 and video_height > 0:
                style = get_subtitle_style_for_format(video_width, video_height)
                ass_text = ass_content_orig.decode('utf-8', errors='replace')
                ass_text = re.sub(r'PlayResX:\s*\d+', f'PlayResX: {style["playresx"]}', ass_text)
                ass_text = re.sub(r'PlayResY:\s*\d+', f'PlayResY: {style["playresy"]}', ass_text)
                new_style = f'Style: Default,Arial,{style["fontsize"]},&H00FFFFFF,&H00FFD600,&H00000000,&H00000000,-1,0,0,0,100,100,0,0,1,{style["outline"]},0,2,30,30,{style["marginv"]},1'
                ass_text = re.sub(r'Style:\s*Default,[^\n]+', new_style, ass_text)
                with open(ass_path, "wb") as f:
                    f.write(ass_text.encode('utf-8'))
            elif has_srt == 'ass' and ass_content_orig:
                with open(ass_path, "wb") as f:
                    f.write(ass_content_orig)

            if has_srt:
                sub_file = ass_path if has_srt == 'ass' else srt_path
                if has_srt == 'ass':
                    vf = f"ass={sub_file}"
                else:
                    s = get_subtitle_style_for_format(video_width, video_height)
                    vf = f"subtitles={sub_file}:force_style='FontName=Arial,FontSize={s['fontsize']},PrimaryColour=&H00FFFFFF,OutlineColour=&H00000000,Outline={s['outline']},Bold=1,Alignment=2,MarginV={s['marginv']}'"
                cmd = ["ffmpeg", "-y", "-threads", "1", "-i", video_path, "-i", audio_path, "-vf", vf,
                       "-map", "0:v:0", "-map", "1:a:0", "-c:v", "libx264", "-preset", "veryfast", "-c:a", "aac",
                       "-shortest", output_path]
            else:
                cmd = ["ffmpeg", "-y", "-threads", "1", "-i", video_path, "-i", audio_path,
                       "-map", "0:v:0", "-map", "1:a:0", "-c:v", "copy", "-c:a", "aac",
                       "-shortest", output_path]

            # FFmpeg v executorju + nice priority
            import os as _os
            if _os.name == 'posix':
                cmd = ['nice', '-n', '10'] + cmd

            # Semafor: samo 1 FFmpeg naenkrat na server
            async with FFMPEG_SEMAPHORE:
                loop = asyncio.get_event_loop()
                result = await loop.run_in_executor(
                    None,
                    lambda: subprocess.run(cmd, capture_output=True, timeout=180)
                )
            if result.returncode != 0:
                err_msg = result.stderr.decode(errors='replace')[-400:]
                return JR({"error": f"FFmpeg napaka: {err_msg}"}, status_code=500)

            with open(output_path, "rb") as f:
                video_bytes = f.read()

            return StreamingResponse(
                iter([video_bytes]),
                media_type="video/mp4",
                headers={"Content-Disposition": f"attachment; filename=video_{lang}.mp4"}
            )
    except subprocess.TimeoutExpired:
        return JR({"error": "FFmpeg timeout."}, status_code=500)
    except Exception as e:
        return JR({"error": str(e)}, status_code=500)


# ─── MERGE VIDEO + AUDIO (full upload — fallback) ─────────────────────────────

@app.post("/merge-video-audio")
async def merge_video_audio(
    video: UploadFile = File(...),
    audio: UploadFile = File(...),
    lang: str = "sl",
    srt: UploadFile = File(None),
    emojis_json: str = Form(None),
):
    """Spoji video + audio (+ opcijsko SRT podnapisi + emoji overlay) z FFmpeg."""
    import subprocess, tempfile
    from fastapi.responses import JSONResponse as JR
    # Uporabi static-ffmpeg da dobimo ffmpeg binarko brez root
    try:
        import static_ffmpeg
        static_ffmpeg.add_paths()
    except Exception:
        pass
    try:
        with tempfile.TemporaryDirectory() as tmp:
            video_path = f"{tmp}/input.mp4"
            audio_path = f"{tmp}/audio.mp3"
            srt_path = f"{tmp}/subs.srt"
            ass_path = f"{tmp}/subs.ass"
            output_path = f"{tmp}/output_{lang}.mp4"

            with open(video_path, "wb") as f:
                f.write(await video.read())
            with open(audio_path, "wb") as f:
                f.write(await audio.read())

            has_srt = False
            ass_content_orig = None
            if srt:
                srt_content = await srt.read()
                if srt_content.strip():
                    with open(srt_path, "wb") as f:
                        f.write(srt_content)
                    # Poskusi parsati kot ASS (začne z [Script Info])
                    if srt_content.startswith(b'[Script Info]'):
                        ass_content_orig = srt_content
                        has_srt = 'ass'
                    else:
                        has_srt = 'srt'

            # Detect video dimensions z ffprobe za prilagoditev podnapisov
            video_width, video_height = 0, 0
            try:
                probe_cmd = [
                    "ffprobe", "-v", "error",
                    "-select_streams", "v:0",
                    "-show_entries", "stream=width,height",
                    "-of", "csv=p=0",
                    video_path
                ]
                probe_result = subprocess.run(probe_cmd, capture_output=True, timeout=10)
                if probe_result.returncode == 0:
                    dims = probe_result.stdout.decode().strip().split(',')
                    if len(dims) == 2:
                        video_width = int(dims[0])
                        video_height = int(dims[1])
                        print(f"[merge] Video {video_width}x{video_height} ratio={video_width/video_height:.2f}")
            except Exception as e:
                print(f"[merge] ffprobe failed: {e}")

            # Če imamo ASS — adaptiraj nastavitve glede na format
            if has_srt == 'ass' and ass_content_orig and video_width > 0 and video_height > 0:
                style = get_subtitle_style_for_format(video_width, video_height)
                # Prepiši Style: Default vrstico v ASS s pravimi parametri za format
                ass_text = ass_content_orig.decode('utf-8', errors='replace')
                # Prepiši PlayResX/PlayResY
                ass_text = re.sub(r'PlayResX:\s*\d+', f'PlayResX: {style["playresx"]}', ass_text)
                ass_text = re.sub(r'PlayResY:\s*\d+', f'PlayResY: {style["playresy"]}', ass_text)
                # Prepiši Style: Default vrstico
                new_style = f'Style: Default,Arial,{style["fontsize"]},&H00FFFFFF,&H00FFD600,&H00000000,&H00000000,-1,0,0,0,100,100,0,0,1,{style["outline"]},0,2,30,30,{style["marginv"]},1'
                ass_text = re.sub(r'Style:\s*Default,[^\n]+', new_style, ass_text)
                with open(ass_path, "wb") as f:
                    f.write(ass_text.encode('utf-8'))
                print(f"[merge] Adapted ASS for {video_width}x{video_height}: fontsize={style['fontsize']}, marginv={style['marginv']}")
            elif has_srt == 'ass' and ass_content_orig:
                # Ni dimensions — uporabi originalni ASS
                with open(ass_path, "wb") as f:
                    f.write(ass_content_orig)

            # Emoji overlay je bil odstranjen za performance (8GB/4CPU optimization).
            # Emojiji ostanejo le v podnapisih kot del besedila (če so v SRT/ASS).

            if has_srt:
                # ASS karaoke ali SRT fallback
                sub_file = ass_path if has_srt == 'ass' else srt_path
                if has_srt == 'ass':
                    sub_filter = f"ass={sub_file}"
                else:
                    s = get_subtitle_style_for_format(video_width, video_height)
                    sub_filter = f"subtitles={sub_file}:force_style='FontName=Arial,FontSize={s['fontsize']},PrimaryColour=&H00FFFFFF,OutlineColour=&H00000000,Outline={s['outline']},Bold=1,Alignment=2,MarginV={s['marginv']}'"

                vf = sub_filter
                cmd = [
                    "ffmpeg", "-y",
                    "-threads", "1",
                    "-i", video_path,
                    "-i", audio_path,
                    "-vf", vf,
                    "-map", "0:v:0",
                    "-map", "1:a:0",
                    "-c:v", "libx264",
                    "-preset", "veryfast",
                    "-c:a", "aac",
                    "-shortest",
                    output_path
                ]
            else:
                # Samo audio zamenjava, video brez rekodiranja
                cmd = [
                    "ffmpeg", "-y",
                    "-threads", "1",
                    "-i", video_path,
                    "-i", audio_path,
                    "-map", "0:v:0",
                    "-map", "1:a:0",
                    "-c:v", "copy",
                    "-c:a", "aac",
                    "-shortest",
                    output_path
                ]

            # Run FFmpeg v executorju da NE blokiramo async event loop-a
            # (preprečuje /healthz timeoute → Render restart)
            # nice=10 da imajo drugi procesi (npr. uvicorn /healthz) prednost
            import os as _os
            if _os.name == 'posix':
                cmd = ['nice', '-n', '10'] + cmd

            # Semafor: samo 1 FFmpeg naenkrat na server — preprečuje CPU saturation
            async with FFMPEG_SEMAPHORE:
                loop = asyncio.get_event_loop()
                result = await loop.run_in_executor(
                    None,
                    lambda: subprocess.run(cmd, capture_output=True, timeout=180)
                )

            if result.returncode != 0:
                err_msg = result.stderr.decode(errors='replace')[-400:]
                print(f"[merge] FFmpeg error: {err_msg}")
                return JR({"error": f"FFmpeg napaka: {err_msg}"}, status_code=500)

            with open(output_path, "rb") as f:
                video_bytes = f.read()

            return StreamingResponse(
                iter([video_bytes]),
                media_type="video/mp4",
                headers={"Content-Disposition": f"attachment; filename=video_{lang}.mp4"}
            )
    except subprocess.TimeoutExpired:
        return JR({"error": "FFmpeg timeout."}, status_code=500)
    except Exception as e:
        return JR({"error": str(e)}, status_code=500)


# ─── VIDEO ADS HISTORY ────────────────────────────────────────────────────────

VADS_HISTORY_FILE = DATA_DIR / "vads_history.json"

def vads_cleanup_old():
    """Zbriše vnose starejše od 7 dni."""
    if not VADS_HISTORY_FILE.exists():
        return
    try:
        history = json.loads(VADS_HISTORY_FILE.read_text(encoding="utf-8"))
        cutoff = datetime.now() - timedelta(days=7)
        history = [h for h in history if datetime.strptime(h.get("date","1.1.2000 00:00"), "%d.%m.%Y %H:%M") > cutoff]
        VADS_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
    except:
        pass

@app.get("/vads-history")
async def get_vads_history():
    vads_cleanup_old()
    if VADS_HISTORY_FILE.exists():
        try:
            return json.loads(VADS_HISTORY_FILE.read_text(encoding="utf-8"))
        except:
            return []
    return []

@app.post("/vads-history")
async def save_vads_history(data: dict):
    try:
        history = []
        if VADS_HISTORY_FILE.exists():
            try:
                history = json.loads(VADS_HISTORY_FILE.read_text(encoding="utf-8"))
            except:
                history = []
        history.append({
            "input": data.get("input", ""),
            "product": data.get("product", ""),
            "scripts": data.get("scripts", {}),
            "date": data.get("date", "")
        })
        if len(history) > 50:
            history = history[-50:]
        VADS_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok"}
    except Exception as e:
        return {"error": str(e)}

@app.post("/vads-history-set")
async def set_vads_history(data: dict):
    try:
        history = data.get("history", [])
        VADS_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "ok"}
    except Exception as e:
        return {"error": str(e)}




# ─── ASANA ATTACH BINARY (za ZIP, video, audio) ───────────────────────────────

@app.post("/asana-attach-binary")
async def asana_attach_binary(
    task_id: str = Form(...),
    file: UploadFile = File(...),
    filename: str = Form(None)
):
    """Priloži binarni fajl (ZIP, MP4, MP3) na Asana task."""
    if not task_id:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Manjka task_id."}, status_code=400)

    token = os.environ.get("ASANA_API_KEY", "")
    if not token:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "ASANA_API_KEY ni nastavljen."}, status_code=400)

    try:
        content = await file.read()
        upload_filename = filename or file.filename or "attachment"
        mime = file.content_type or "application/octet-stream"

        async with httpx.AsyncClient(timeout=120.0) as hc:
            files = {"file": (upload_filename, content, mime)}
            attach_resp = await hc.post(
                f"{ASANA_API}/tasks/{task_id}/attachments",
                headers={"Authorization": f"Bearer {token}"},
                files=files
            )

        if attach_resp.status_code in (200, 201):
            return {"status": "ok", "filename": upload_filename, "size": len(content)}
        else:
            from fastapi.responses import JSONResponse
            return JSONResponse(
                {"error": f"Asana napaka {attach_resp.status_code}: {attach_resp.text[:200]}"},
                status_code=400
            )
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── ORODJA: Združevalnik SKU+količin ─────────────────────────────────────────

ORODJA_HISTORY_DIR = DATA_DIR / "orodja_history"
ORODJA_HISTORY_DIR.mkdir(exist_ok=True, parents=True)

ORODJA_HS_HISTORY_DIR = DATA_DIR / "orodja_hs_history"
ORODJA_HS_HISTORY_DIR.mkdir(exist_ok=True, parents=True)


def cleanup_orodja_history():
    """Briše datoteke starejše od 30 dni (CSV združevalnik)."""
    try:
        cutoff = datetime.now().timestamp() - (30 * 86400)
        for f in ORODJA_HISTORY_DIR.glob("*.xlsx"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[orodja] cleanup err: {e}")


def cleanup_orodja_hs_history():
    """Briše datoteke starejše od 90 dni (HS+ uvoz)."""
    try:
        cutoff = datetime.now().timestamp() - (90 * 86400)
        for f in ORODJA_HS_HISTORY_DIR.glob("*.xlsx"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[orodja-hs] cleanup err: {e}")


@app.post("/orodja-merge-skus")
async def orodja_merge_skus(file: UploadFile = File(...)):
    """Sprejme CSV z SKU+Količina, združi dvojnike, vrne XLSX."""
    try:
        content_bytes = await file.read()
        text = content_bytes.decode('utf-8-sig', errors='replace')

        # Parsanje CSV
        import csv
        from io import StringIO
        reader = csv.reader(StringIO(text))
        headers = next(reader, None)
        if not headers:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)

        # Najdi indekse SKU in Količina kolon
        h_lower = [h.lower().strip() for h in headers]
        try:
            sku_idx = next(i for i, h in enumerate(h_lower) if h in ('sku', 'sku.'))
        except StopIteration:
            sku_idx = 1  # default druga kolona
        try:
            qty_idx = next(i for i, h in enumerate(h_lower) if 'količin' in h or 'kolicin' in h or h == 'qty' or h == 'quantity')
        except StopIteration:
            qty_idx = 3  # default

        # Združi po SKU
        sku_totals = {}
        for row in reader:
            if len(row) <= max(sku_idx, qty_idx):
                continue
            sku = (row[sku_idx] or '').strip()
            if not sku:
                continue
            try:
                qty = int(float((row[qty_idx] or '0').strip().replace(',', '.')))
            except:
                qty = 0
            sku_totals[sku] = sku_totals.get(sku, 0) + qty

        if not sku_totals:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "Ne najdem SKU/količina kolon v CSV."}, status_code=400)

        # Generiraj XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Order"
        # Brez headerjev — samo SKU | količina
        for sku, qty in sorted(sku_totals.items()):
            ws.append([sku, qty])

        # Shrani v history
        cleanup_orodja_history()  # počisti stare najprej
        ts = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        out_filename = f"Order_{ts}.xlsx"
        out_path = ORODJA_HISTORY_DIR / out_filename
        wb.save(out_path)

        # Vrni datoteko
        from fastapi.responses import FileResponse
        return FileResponse(
            path=str(out_path),
            filename=out_filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"X-Skus-Total": str(len(sku_totals)), "X-Filename": out_filename}
        )
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/orodja-history")
async def orodja_history():
    """Vrne seznam datotek v orodja_history (urejen po datumu od najnovejše)."""
    cleanup_orodja_history()
    items = []
    try:
        for f in sorted(ORODJA_HISTORY_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
            stat = f.stat()
            items.append({
                "filename": f.name,
                "size": stat.st_size,
                "created": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
    except Exception as e:
        print(f"[orodja] history err: {e}")
    return {"items": items[:100]}


@app.get("/orodja-download/{filename}")
async def orodja_download(filename: str):
    """Prenesi XLSX iz history."""
    # Sanitize filename — preprečimo path traversal
    if '/' in filename or '\\' in filename or '..' in filename:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)

    f = ORODJA_HISTORY_DIR / filename
    if not f.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Datoteka ne obstaja."}, status_code=404)

    from fastapi.responses import FileResponse
    return FileResponse(
        path=str(f),
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.delete("/orodja-history/{filename}")
async def orodja_history_delete(filename: str):
    """Zbriši posamezno datoteko iz history."""
    if '/' in filename or '\\' in filename or '..' in filename:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)

    f = ORODJA_HISTORY_DIR / filename
    if f.exists():
        try:
            f.unlink()
            return {"status": "ok"}
        except Exception as e:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": str(e)}, status_code=500)
    return {"status": "not_found"}


# ─── ORODJA: Uvoz HS+ PDF predračun ───────────────────────────────────────────

@app.post("/orodja-import-hs-pdf")
async def orodja_import_hs_pdf(file: UploadFile = File(...)):
    """Sprejme HS+ PDF predračun, vrne JSON s SKU + količinami.
    Uporablja Claude Vision za branje image-based PDF."""
    try:
        content_bytes = await file.read()
        items = []

        # Najprej poskus tekst-extraction (če je tekstovni PDF)
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp.write(content_bytes)
            tmp_path = tmp.name

        try:
            try:
                with pdfplumber.open(tmp_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        if text.strip():
                            for line in text.split('\n'):
                                line = line.strip()
                                m = re.match(r'^(\d{12,14})\s+(.+?)\s+(\d+)\s+(?:KOS|kos|PCS|pcs)', line)
                                if m:
                                    opis = m.group(2).strip()
                                    tokens = [t.rstrip('.,;:') for t in opis.split()]
                                    upper_t = [t for t in tokens if t.isupper() and len(t) >= 3 and not t.isdigit()]
                                    sku = upper_t[-1] if upper_t else (tokens[-1] if tokens else opis)
                                    items.append({
                                        "ean": m.group(1),
                                        "opis": opis,
                                        "sku": sku,
                                        "kolicina": int(m.group(3)),
                                    })
            except: pass

            # Fallback: Claude Vision (PDF kot dokument)
            if not items:
                import base64
                pdf_b64 = base64.b64encode(content_bytes).decode('utf-8')

                prompt = """Preberi ta predračun in vrni VSA postavke v JSON formatu.
Za vsako postavko izloci:
- ean: 13-mestna številčna koda na začetku vrstice
- opis: celoten opis postavke
- sku: zadnja SVE-VELIKA-ČRKA beseda v opisu (npr. "HYDRASPRINK HYDRASPRINK" → SKU = "HYDRASPRINK"; "WHEELPLAY yellow WHEELPLAY" → SKU = "WHEELPLAY"; "TOPKNER 180x200 TOPKNER" → SKU = "TOPKNER")
- kolicina: število pred "KOS" oznako

Vrni IZKLJUČNO valid JSON v formatu:
{"items": [{"ean": "...", "opis": "...", "sku": "...", "kolicina": 350}, ...]}

Brez dodatnih komentarjev, samo JSON."""

                try:
                    client = anthropic.Anthropic()
                    response = client.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=8000,
                        messages=[{
                            "role": "user",
                            "content": [
                                {
                                    "type": "document",
                                    "source": {
                                        "type": "base64",
                                        "media_type": "application/pdf",
                                        "data": pdf_b64
                                    }
                                },
                                {"type": "text", "text": prompt}
                            ]
                        }]
                    )
                    text = "".join([b.text for b in response.content if hasattr(b, 'text')])
                    parsed = parse_json_response(text)
                    if parsed and 'items' in parsed:
                        for it in parsed['items']:
                            try:
                                qty = int(it.get('kolicina', 0))
                            except:
                                qty = 0
                            items.append({
                                "ean": str(it.get('ean', '')),
                                "opis": str(it.get('opis', '')),
                                "sku": str(it.get('sku', '')).strip(),
                                "kolicina": qty,
                            })
                except Exception as e:
                    print(f"[hs-pdf] Claude vision error: {e}")
        finally:
            try:
                os.unlink(tmp_path)
            except: pass

        if not items:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "Iz PDF-ja ne morem prebrati postavk."}, status_code=400)

        # Skupna količina iz PDF-ja
        pdf_total = sum(int(it.get('kolicina', 0) or 0) for it in items)
        return {"items": items, "total": len(items), "pdf_total": pdf_total}
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/orodja-export-hs-xlsx")
async def orodja_export_hs_xlsx(data: dict):
    """Sprejme [{sku, kolicina}], generira XLSX in shrani v history."""
    try:
        items = data.get("items", [])
        if not items:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "Ni postavk."}, status_code=400)

        # Generiraj XLSX z headerji sku|stock (HS+ format)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = datetime.now().strftime('%Y-%m-%d_%H-%M-%S-export')
        ws.append(["sku", "stock"])  # header
        for item in items:
            sku = (item.get("sku") or "").strip()
            qty = item.get("kolicina") or 0
            if not sku:
                continue
            try:
                qty_int = int(float(qty))
            except:
                qty_int = 0
            ws.append([sku, qty_int])

        cleanup_orodja_hs_history()
        ts = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        out_filename = f"HS_PLUS_{ts}.xlsx"
        out_path = ORODJA_HS_HISTORY_DIR / out_filename
        wb.save(out_path)

        from fastapi.responses import FileResponse
        return FileResponse(
            path=str(out_path),
            filename=out_filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"X-Filename": out_filename}
        )
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)



# ─── HS+ HISTORY ENDPOINTS ────────────────────────────────────────────────────

@app.get("/orodja-hs-history")
async def orodja_hs_history():
    """Vrne seznam HS+ datotek (90 dni)."""
    cleanup_orodja_hs_history()
    items = []
    try:
        for f in sorted(ORODJA_HS_HISTORY_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
            stat = f.stat()
            items.append({
                "filename": f.name,
                "size": stat.st_size,
                "created": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
    except Exception as e:
        print(f"[orodja-hs] history err: {e}")
    return {"items": items[:200]}


@app.get("/orodja-hs-download/{filename}")
async def orodja_hs_download(filename: str):
    if '/' in filename or '\\' in filename or '..' in filename:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)

    f = ORODJA_HS_HISTORY_DIR / filename
    if not f.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Datoteka ne obstaja."}, status_code=404)

    from fastapi.responses import FileResponse
    return FileResponse(
        path=str(f),
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.delete("/orodja-hs-history/{filename}")
async def orodja_hs_history_delete(filename: str):
    if '/' in filename or '\\' in filename or '..' in filename:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)

    f = ORODJA_HS_HISTORY_DIR / filename
    if f.exists():
        try:
            f.unlink()
            return {"status": "ok"}
        except Exception as e:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": str(e)}, status_code=500)
    return {"status": "not_found"}



# ─── ORODJA: Kontrola cen — Stock CSV upload + match s PDF predračun ────────

STOCK_CSV_FILE = DATA_DIR / "stock_inventory.csv"
SILUXAR_PUSH_LOG = DATA_DIR / "siluxar_push_log.json"   # zadnja pošiljanja pozicij (debug)
SILUXAR_DELETE_LOG = DATA_DIR / "siluxar_delete_log.json"   # zadnja brisanja alarmov (debug)
STOCK_BACKUP_DIR = DATA_DIR / "stock_backups"   # avtomatski backupi zaloge pred sync
STOCK_CSV_META = DATA_DIR / "stock_inventory_meta.json"


@app.post("/zaloga-upload-obrat")
async def zaloga_upload_obrat(file: UploadFile = File(...)):
    """Iz naloženega CSV potegne SAMO obrat/30d (stock30) + trajanje (stock_duration) po SKU
    in ju OSVEŽI v obstoječi zalogi. NE dotika zaloge, pozicije, cene, naziva — samo ti dve polji.
    SKU-jev, ki niso v obstoječi zalogi, NE dodaja."""
    try:
        import csv as _csv
        from io import StringIO as _SIO

        if not STOCK_CSV_FILE.exists():
            return JSONResponse({"error": "Najprej naloži/sinhroniziraj zalogo, šele nato obrat+trajanje."}, status_code=400)

        content_bytes = await file.read()
        text = content_bytes.decode('utf-8-sig', errors='replace')
        sep = ';' if (text.split('\n', 1)[0].count(';') > text.split('\n', 1)[0].count(',')) else ','
        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        new_rows = [r for r in reader]
        if not new_rows:
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)

        keys = list(new_rows[0].keys())
        def find_col(*cands):
            for c in cands:
                for k in keys:
                    if k.strip().lower() == c.lower():
                        return k
            return None
        sku_col = find_col('product_sku', 'sku')
        s30_col = find_col('stock30', 'stock_30', 'obrat30', 'obrat_30', 'obrat', 'obr30', 'obrat_30d')
        dur_col = find_col('stock_duration', 'trajanje_zaloge', 'trajanje', 'duration', 'dni_zaloge', 'days_of_stock', 'zaloga_dni')
        if not sku_col:
            return JSONResponse({"error": f"Ne najdem SKU stolpca. Najdeni: {keys}"}, status_code=400)
        if not s30_col and not dur_col:
            return JSONResponse({"error": f"Ne najdem stolpca za obrat (stock30) ne trajanje. Najdeni: {keys}"}, status_code=400)

        # zgradi mapo SKU -> {stock30, stock_duration} iz naloženega CSV
        incoming = {}
        for row in new_rows:
            sku = (row.get(sku_col) or '').strip()
            if not sku:
                continue
            rec = {}
            if s30_col:
                v = (row.get(s30_col) or '').strip()
                if v != '':
                    try: rec['stock30'] = str(int(float(v.replace(',', '.'))))
                    except: pass
            if dur_col:
                d = (row.get(dur_col) or '').strip()
                if d != '':
                    rec['stock_duration'] = d
            if rec:
                incoming[sku] = rec

        if not incoming:
            return JSONResponse({"error": "V CSV ni veljavnih vrednosti obrat/trajanje."}, status_code=400)

        # naloži OBSTOJEČO zalogo, posodobi SAMO ti dve polji
        old_text = STOCK_CSV_FILE.read_text(encoding='utf-8-sig', errors='replace')
        rd = _csv.DictReader(_SIO(old_text))
        existing_fields = list(rd.fieldnames or [])
        rows = list(rd)
        # zagotovi, da sta stolpca v CSV
        for col in ('stock30', 'stock_duration'):
            if col not in existing_fields:
                existing_fields.append(col)

        updated = 0
        matched_skus = set()
        for row in rows:
            sku = (row.get('product_sku') or row.get('sku') or '').strip()
            if sku and sku in incoming:
                rec = incoming[sku]
                if 'stock30' in rec:
                    row['stock30'] = rec['stock30']
                if 'stock_duration' in rec:
                    row['stock_duration'] = rec['stock_duration']
                matched_skus.add(sku)
                updated += 1

        # shrani nazaj — iste vrstice, samo posodobljeni polji
        out = _SIO()
        writer = _csv.DictWriter(out, fieldnames=existing_fields, extrasaction='ignore')
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        STOCK_CSV_FILE.write_text(out.getvalue(), encoding='utf-8')

        not_found = len(incoming) - len(matched_skus)
        return {"ok": True, "updated": updated, "csv_skus": len(incoming),
                "not_found": not_found,
                "fields": [c for c in ('obrat/30d' if s30_col else None, 'trajanje' if dur_col else None) if c]}
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/orodja-stock-upload")
async def orodja_stock_upload(file: UploadFile = File(...)):
    """Naloži CSV zaloge — merge po SKU (sešteje stock iz več skladišč/uploadov)."""
    try:
        import csv as _csv
        from io import StringIO as _SIO

        content_bytes = await file.read()
        text = content_bytes.decode('utf-8-sig', errors='replace')
        sep = ';' if (text.split('\n', 1)[0].count(';') > text.split('\n', 1)[0].count(',')) else ','

        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        new_rows = [r for r in reader]
        if not new_rows:
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)

        # Normalizacija: najdi SKU, stock, stock30, title stolpce
        sample = new_rows[0]
        keys = list(sample.keys())

        def find_col(*candidates):
            for c in candidates:
                for k in keys:
                    if k.strip().lower() == c.lower():
                        return k
            return None

        sku_col   = find_col('product_sku', 'sku')
        stock_col = find_col('stock', 'qty', 'quantity', 'kolicina', 'količina')
        s30_col   = find_col('stock30', 'stock_30', 'obrat30', 'obrat_30')
        title_col = find_col('title', 'naziv', 'name')
        price_col = find_col('price_netto', 'price', 'cena')
        pos_col   = find_col('position', 'pozicija', 'lokacija')
        note_col  = find_col('note', 'opomba', 'komentar')
        wh_col    = find_col('skladisce', 'skladišče', 'warehouse', 'store', 'source')

        if not sku_col:
            return JSONResponse({"error": f"Ne najdem SKU stolpca. Najdeni: {keys}"}, status_code=400)

        # KLJUČ = SKU + skladišče (LOČENO po skladiščih, NE sešteva)
        def _mk_key(sku, wh):
            return (sku or '').strip() + '|' + (wh or '').strip()
        merged = {}
        added = 0
        updated = 0
        for row in new_rows:
            if not sku_col: continue
            sku = (row.get(sku_col) or '').strip()
            if not sku: continue
            try:
                new_stock = int(float((row.get(stock_col) or '0').replace(',', '.'))) if stock_col else 0
            except: new_stock = 0
            try:
                new_s30 = int(float((row.get(s30_col) or '0').replace(',', '.'))) if s30_col else 0
            except: new_s30 = 0
            title = (row.get(title_col) or '').strip() if title_col else ''
            wh    = (row.get(wh_col) or '').strip() if wh_col else ''
            rk = _mk_key(sku, wh)

            if rk in merged:
                # ista SKU+skladišče kombinacija v istem CSV — prepiši (NE sešteva med skladišči)
                merged[rk]['stock'] = new_stock
                if new_s30: merged[rk]['stock30'] = new_s30
                if title and not merged[rk]['title']:
                    merged[rk]['title'] = title
                for fld, src_col in [('price', price_col), ('position', pos_col), ('note', note_col), ('product_id', 'product_id')]:
                    new_val = (row.get(src_col) or '').strip() if src_col else ''
                    if new_val and not merged[rk].get(fld):
                        merged[rk][fld] = new_val
                updated += 1
            else:
                merged[rk] = {
                    'product_id': (row.get('product_id') or '').strip(),
                    'product_sku': sku,
                    'title': title,
                    'stock': new_stock,
                    'stock30': new_s30,
                    'price': (row.get(price_col) or '').strip() if price_col else '',
                    'position': (row.get(pos_col) or '').strip() if pos_col else '',
                    'note': (row.get(note_col) or '').strip() if note_col else '',
                    'warehouse': wh,
                }
                added += 1

        # Shrani nazaj kot CSV z vsemi kolonami
        out = _SIO()
        fieldnames = ['product_id', 'product_sku', 'title', 'stock', 'stock30', 'price', 'position', 'note', 'warehouse']
        writer = _csv.DictWriter(out, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(merged.values())
        STOCK_CSV_FILE.write_text(out.getvalue(), encoding='utf-8')

        total = len(merged)
        meta = {
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "filename": file.filename,
            "rows": total,
            "rows_added": added,
            "rows_merged": updated,
        }
        # Ohrani seznam uploadov
        old_meta = {}
        if STOCK_CSV_META.exists():
            try: old_meta = json.loads(STOCK_CSV_META.read_text(encoding='utf-8'))
            except: pass
        uploads = old_meta.get('uploads', [])
        uploads.append({"filename": file.filename, "uploaded_at": meta["uploaded_at"], "added": added, "merged": updated})
        meta['uploads'] = uploads[-5:]  # zadnjih 5
        STOCK_CSV_META.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')

        return {
            "status": "ok",
            "rows": total,
            "rows_added": added,
            "rows_merged": updated,
            "uploaded_at": meta["uploaded_at"],
            "filename": file.filename,
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)



@app.get("/siluxar-diag")
async def siluxar_diag():
    """Diagnostika povezave do siluxar.si — pove TOČNO, kje pade (DNS / TCP / HTTP).
    Pomaga ločiti omrežno težavo (požarni zid/geo) od avtentikacije."""
    import socket, ssl as _ssl
    out = {"steps": []}

    # 0) DEJANSKI odhodni IP tega strežnika (s katerega Render kliče ven)
    #    To je IP, ki ga mora siluxar dovoliti — preverimo, ali je v dovoljenih range.
    try:
        async with httpx.AsyncClient(timeout=15) as cli:
            ipr = await cli.get("https://api.ipify.org?format=json")
        my_ip = ipr.json().get("ip", "?")
        out["outbound_ip"] = my_ip
        # preveri ali je v dovoljenih range 74.220.51.0/24 ali 74.220.59.0/24
        import ipaddress
        allowed = ipaddress.ip_address(my_ip) in ipaddress.ip_network("74.220.51.0/24") or \
                  ipaddress.ip_address(my_ip) in ipaddress.ip_network("74.220.59.0/24")
        out["outbound_ip_in_allowed_ranges"] = allowed
        out["outbound_ip_note"] = ("✓ IP JE v dovoljenih range." if allowed else
                                    "✗ IP NI v dovoljenih range 74.220.51.0/24 ali 74.220.59.0/24 — TO je vzrok! Pošlji ta IP administratorju.")
    except Exception as e:
        out["outbound_ip"] = f"napaka: {e}"

    out["steps"] = []
    host = "www.siluxar.si"

    # 1) DNS razrešitev
    try:
        ips = socket.gethostbyname_ex(host)
        out["steps"].append({"step": "DNS", "ok": True, "resolved": ips[2]})
    except Exception as e:
        out["steps"].append({"step": "DNS", "ok": False, "error": str(e)})
        out["verdict"] = "DNS ne razreši — težava z imenom domene."
        return out

    # 2) TCP povezava na 443
    try:
        sock = socket.create_connection((host, 443), timeout=15)
        sock.close()
        out["steps"].append({"step": "TCP:443", "ok": True})
    except Exception as e:
        out["steps"].append({"step": "TCP:443", "ok": False, "error": str(e)})
        out["verdict"] = "TCP povezava pade → požarni zid/geo-filter blokira na omrežni ravni (NE avtentikacija)."
        return out

    # 3) HTTPS testi z RAZLIČNIMI avtentikacijami — da vidimo, kaj strežnik sprejme
    key = os.environ.get("SILUXAR_STOCK_KEY", "")
    bu = os.environ.get("SILUXAR_BASIC_USER", "")
    bp = os.environ.get("SILUXAR_BASIC_PASS", "")
    _ba = httpx.BasicAuth(bu, bp) if (bu or bp) else None

    def _snippet(resp):
        """Skrajšano telo + ključne glave (kdo zavrača)."""
        body = (resp.text or "")[:400]
        wa = resp.headers.get("www-authenticate", "")
        srv = resp.headers.get("server", "")
        ct = resp.headers.get("content-type", "")
        clen = resp.headers.get("content-length", "")
        return {"status": resp.status_code, "body_len": len(resp.text or ""),
                "body_preview": body, "content_type": ct, "content_length": clen,
                "www_authenticate": wa, "server": srv}

    tests = {}
    # (a) brez avtentikacije
    try:
        async with httpx.AsyncClient(timeout=30) as cli:
            r = await cli.get(f"https://{host}/apistockexport")
        tests["a_brez_auth"] = _snippet(r)
    except Exception as e:
        tests["a_brez_auth"] = {"error": str(e)}
    # (b) samo Basic Auth
    if _ba:
        try:
            async with httpx.AsyncClient(timeout=30, auth=_ba) as cli:
                r = await cli.get(f"https://{host}/apistockexport")
            tests["b_basic_auth"] = _snippet(r)
        except Exception as e:
            tests["b_basic_auth"] = {"error": str(e)}
    # (c) Basic Auth + X-API-Key
    try:
        async with httpx.AsyncClient(timeout=30, auth=_ba) as cli:
            r = await cli.get(f"https://{host}/apistockexport", headers={"X-API-Key": key})
        tests["c_basic_in_xapikey"] = _snippet(r)
    except Exception as e:
        tests["c_basic_in_xapikey"] = {"error": str(e)}
    # (d) samo X-API-Key (brez Basic) — če bi admin umaknil Basic Auth
    try:
        async with httpx.AsyncClient(timeout=30) as cli:
            r = await cli.get(f"https://{host}/apistockexport", headers={"X-API-Key": key})
        tests["d_samo_xapikey"] = _snippet(r)
    except Exception as e:
        tests["d_samo_xapikey"] = {"error": str(e)}
    # (e) ključ v Authorization (če Marko pričakuje tam, brez Basic)
    try:
        async with httpx.AsyncClient(timeout=30) as cli:
            r = await cli.get(f"https://{host}/apistockexport", headers={"Authorization": key})
        tests["e_kljuc_v_authorization"] = _snippet(r)
    except Exception as e:
        tests["e_kljuc_v_authorization"] = {"error": str(e)}

    # (f) Basic Auth + razni parametri za izvoz (telo je bilo prazno pri golem GET)
    param_variants = [
        "?format=csv", "?format=json", "?type=csv", "?type=json",
        "?export=1", "?download=1", "?output=csv", "?all=1",
    ]
    param_tests = {}
    for pv in param_variants:
        try:
            async with httpx.AsyncClient(timeout=30, auth=_ba) as cli:
                r = await cli.get(f"https://{host}/apistockexport{pv}")
            param_tests[pv] = {"status": r.status_code, "body_len": len(r.text or ""),
                               "preview": (r.text or "")[:120]}
        except Exception as e:
            param_tests[pv] = {"error": str(e)}

    out["steps"].append({"step": "TCP:443", "ok": True})
    out["auth_tests"] = tests
    out["param_tests"] = param_tests
    out["razlaga"] = ("Poišči test s statusom 200 IN body_len > 0 = ta kombinacija vrne podatke. "
                      "Če so vsi body_len=0, API rabi drugačen klic — vprašaj Marka. "
                      "param_tests preizkusi pogoste parametre za izvoz.")
    out["verdict"] = "Omrežje+auth deluje (200). Preveri body_len: če 0, manjka parameter za izvoz."
    return out


@app.post("/zaloga-reset")
async def zaloga_reset():
    """Počisti celotno shranjeno zalogo (CSV). Po tem prva sinhronizacija zgradi bazo od začetka."""
    try:
        if STOCK_CSV_FILE.exists():
            STOCK_CSV_FILE.unlink()
        if STOCK_CSV_META.exists():
            STOCK_CSV_META.unlink()
        return {"ok": True, "message": "Zaloga počiščena. Klikni Sinhroniziraj za uvoz s siluxar."}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _make_stock_backup(razlog="sync"):
    """Naredi backup trenutne zaloge (CSV) pred prepisom. Obdrži zadnjih 30 backupov.
    Vrne ime backup datoteke ali None, če ni kaj backupirati."""
    try:
        if not STOCK_CSV_FILE.exists():
            return None
        STOCK_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        ts = _lj_now().strftime("%Y%m%d_%H%M%S")
        fname = f"stock_{ts}_{razlog}.csv"
        dest = STOCK_BACKUP_DIR / fname
        # kopiraj vsebino
        dest.write_text(STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace"), encoding="utf-8")
        # počisti stare — obdrži zadnjih 30
        backups = sorted(STOCK_BACKUP_DIR.glob("stock_*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
        for old in backups[30:]:
            try: old.unlink()
            except Exception: pass
        return fname
    except Exception as e:
        print(f"[backup] napaka: {e}")
        return None


async def _zaloga_sync_core():
    """JEDRO sinhronizacije zaloge s siluxar (apistockexport). MERGE po SKU+skladišče.
    Kliče ga endpoint /zaloga-sync-siluxar (gumb) IN scheduler (vsake 4h).
    PRED prepisom naredi avtomatski backup trenutne zaloge."""
    import csv as _csv
    from io import StringIO as _SIO
    key = os.environ.get("SILUXAR_STOCK_KEY", "")
    if not key:
        return {"ok": False, "error": "Manjka SILUXAR_STOCK_KEY (Render okoljska spremenljivka)."}
    # BACKUP pred sync (da lahko hitro restoramo, če gre kaj narobe)
    _backup_ime = _make_stock_backup("pred-sync")
    basic_user = os.environ.get("SILUXAR_BASIC_USER", "")
    basic_pass = os.environ.get("SILUXAR_BASIC_PASS", "")
    url = "https://www.siluxar.si/apistockexport"
    # Avtentikacija: Markov ključ gre v Authorization glavo (kot v programerjevem curl).
    # Če sta nastavljena Basic user/pass IN ni ključa, uporabi Basic Auth (fallback).
    headers = {}
    _auth = None
    if key:
        headers["Authorization"] = key
    elif basic_user or basic_pass:
        _auth = httpx.BasicAuth(basic_user, basic_pass)
    try:
        async with httpx.AsyncClient(timeout=90, auth=_auth) as cli:
            r = await cli.get(url, headers=headers)
    except Exception as e:
        return {"ok": False, "error": f"Napaka pri klicu siluxar.si: {e}"}
    if r.status_code != 200:
        return {"ok": False, "error": f"siluxar.si vrnil status {r.status_code}", "status": r.status_code}
    text = r.text or ""
    if not text.strip():
        return {"ok": False, "error": "siluxar.si ni vrnil podatkov."}

    # 2) razberi vrstice (JSON seznam ALI CSV)
    incoming = []  # list of dict
    ctype = r.headers.get("content-type", "")
    parsed_json = False
    if "json" in ctype or text.lstrip()[:1] in ("[", "{"):
        try:
            jd = json.loads(text)
            if isinstance(jd, dict):
                # morda {"data":[...]} ali {"items":[...]}
                for kk in ("data", "items", "rows", "products", "stock"):
                    if isinstance(jd.get(kk), list):
                        jd = jd[kk]; break
            if isinstance(jd, list):
                incoming = [x for x in jd if isinstance(x, dict)]
                parsed_json = True
        except Exception:
            parsed_json = False
    if not parsed_json:
        # CSV
        sep = ';' if (text.split('\n',1)[0].count(';') > text.split('\n',1)[0].count(',')) else ','
        incoming = list(_csv.DictReader(_SIO(text), delimiter=sep))
    if not incoming:
        return {"ok": False, "error": "Ni veljavnih vrstic v odgovoru siluxar."}

    # 3) normalizacija stolpcev (prožno)
    keys = list(incoming[0].keys())
    def find_col(*cands):
        for c in cands:
            for k in keys:
                if k.strip().lower() == c.lower():
                    return k
        return None
    sku_col   = find_col('product_sku','sku')
    stock_col = find_col('product_stock','stock','qty','quantity','kolicina','količina','zaloga')
    s30_col   = find_col('stock30','stock_30','obrat30','obrat_30')
    title_col = find_col('product_title','title','naziv','name')
    price_col = find_col('product_price_netto','price_netto','product_price','price','cena')
    pos_col   = find_col('position','pozicija','lokacija')
    note_col  = find_col('note','opomba','komentar')
    id_col    = find_col('product_id')   # interni product_id (točno ujemanje)
    sid_col   = find_col('id')           # ps.id = siluxar skrit ključ (sinhronizacija brisanja)
    # Trajanje zaloge + Skladišče (prožno — Markovo ime morda drugačno)
    dur_col   = find_col('trajanje_zaloge','trajanje','stock_duration','duration','dni_zaloge','days_of_stock','zaloga_dni')
    wh_col    = find_col('skladisce','skladišče','warehouse','store','trgovina','lokacija_skladisce','source')  # source=silux/silux2=skladišče
    if not sku_col:
        return {"ok": False, "error": f"Ne najdem SKU stolpca. Najdeni: {keys}"}

    # 4) naloži OBSTOJEČO zalogo (merge cilj). KLJUČ = SKU + skladišče (ločeno po skladiščih).
    def _mk_key(sku, wh):
        return (sku or '').strip() + '|' + (wh or '').strip()
    existing = {}
    if STOCK_CSV_FILE.exists():
        try:
            old_text = STOCK_CSV_FILE.read_text(encoding='utf-8')
            for row in _csv.DictReader(_SIO(old_text)):
                s = (row.get('product_sku') or '').strip()
                w = (row.get('warehouse') or '').strip()
                if s:
                    existing[_mk_key(s, w)] = dict(row)
        except Exception:
            existing = {}

    def _to_int(v):
        try: return int(float(str(v).replace(',', '.')))
        except: return 0
    def _to_float(v):
        try: return float(str(v).replace(',', '.'))
        except: return 0.0

    added = 0; updated = 0; external_count = 0
    seen_this_sync = set()   # ključi, ki smo jih ŽE postavili v TEJ sinhronizaciji (za seštevanje podvojenih)
    price_acc = {}           # ključ -> [vsota_vrednosti, vsota_kosov] za TEHTANO ceno podvojenih
    for row in incoming:
        sku = (row.get(sku_col) or '').strip()
        if not sku:
            continue
        sid   = (row.get(sid_col) or '').strip() if sid_col else ''   # ps.id (skrit ključ)
        # EXTERNAL: naši izdelki imajo ps.id; external nimajo (prazen ali 0).
        # External UVOZIMO z oznako is_external="1" (privzeto skriti na frontu, checkbox jih pokaže).
        is_ext = '1' if (not sid or sid in ('0', '0.0')) else ''
        if is_ext:
            external_count += 1
        new_stock = _to_int(row.get(stock_col)) if stock_col else 0
        new_s30   = _to_int(row.get(s30_col)) if s30_col else 0
        title = (row.get(title_col) or '').strip() if title_col else ''
        price = (row.get(price_col) or '').strip() if price_col else ''
        pos   = (row.get(pos_col) or '').strip() if pos_col else ''
        note  = (row.get(note_col) or '').strip() if note_col else ''
        pid   = (row.get(id_col) or '').strip() if id_col else ''
        dur   = (row.get(dur_col) or '').strip() if dur_col else ''   # trajanje zaloge
        wh    = (row.get(wh_col) or '').strip() if wh_col else ''     # skladišče
        rk = _mk_key(sku, wh)   # ključ: SKU + skladišče (ločeno po skladiščih)
        first_in_sync = rk not in seen_this_sync   # prvi zapis tega ključa v TEJ sinhronizaciji?
        seen_this_sync.add(rk)
        # TEHTANA cena: akumuliraj vrednost (kosi×cena) + kose za ta ključ v tej sinhronizaciji
        _pn = _to_float(price)
        if first_in_sync:
            price_acc[rk] = [new_stock * _pn, new_stock]
        else:
            price_acc[rk][0] += new_stock * _pn
            price_acc[rk][1] += new_stock
        # izračunaj tehtano ceno (vrednost / kosi); če kosov 0, obdrži zadnjo ne-prazno ceno
        _acc_val, _acc_qty = price_acc[rk]
        if _acc_qty > 0:
            tehtana_cena = "{:.2f}".format(_acc_val / _acc_qty)
        else:
            tehtana_cena = price   # vsi kosi 0 → pusti ceno zadnjega zapisa
        if rk in existing:
            e = existing[rk]
            # stock: prvič v tej sinhronizaciji POSTAVI, vse naslednje zapise istega SKU+skladišča PRIŠTEJ
            # (isti SKU ima v siluxar lahko več zapisov v istem skladišču — seštejemo jih)
            if first_in_sync:
                e['stock'] = str(new_stock)
            else:
                e['stock'] = str(_to_int(e.get('stock')) + new_stock)
            if s30_col: e['stock30'] = str(new_s30)
            if title: e['title'] = title
            # cena = TEHTANA (po kosih) — pravilno tudi pri podvojenih z različnimi cenami
            if tehtana_cena and _to_float(tehtana_cena) > 0:
                e['price'] = tehtana_cena
            elif price and first_in_sync:
                e['price'] = price
            if pos: e['position'] = pos
            if note: e['note'] = note
            if pid: e['product_id'] = pid
            if sid: e['siluxar_id'] = sid
            if dur: e['stock_duration'] = dur
            e['warehouse'] = wh
            e['is_external'] = is_ext
            updated += 1
        else:
            existing[rk] = {
                'product_id': pid, 'product_sku': sku, 'title': title,
                'stock': str(new_stock), 'stock30': str(new_s30),
                'price': (tehtana_cena if (tehtana_cena and _to_float(tehtana_cena) > 0) else price),
                'position': pos, 'note': note,
                'siluxar_id': sid, 'stock_duration': dur, 'warehouse': wh,
                'is_external': is_ext,
            }
            added += 1

    # 5) shrani nazaj — siluxar_id je skrit ključ. External (is_external="1")
    #    OBDRŽIMO v CSV, a so privzeto skriti na frontu (checkbox jih pokaže).
    out = _SIO()
    fieldnames = ['product_id','product_sku','title','stock','stock30','price','position','note','siluxar_id','stock_duration','warehouse','is_external']
    writer = _csv.DictWriter(out, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()
    for v in existing.values():
        writer.writerow({k: v.get(k, '') for k in fieldnames})
    STOCK_CSV_FILE.write_text(out.getvalue(), encoding='utf-8')

    # meta
    meta = {}
    if STOCK_CSV_META.exists():
        try: meta = json.loads(STOCK_CSV_META.read_text(encoding='utf-8'))
        except: meta = {}
    meta['last_siluxar_sync'] = datetime.now(timezone.utc).isoformat()
    meta['rows'] = len(existing)
    STOCK_CSV_META.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')

    nasi_count = len(existing) - external_count
    return {"ok": True, "total": len(existing), "added": added, "updated": updated,
            "external": external_count, "nasi": nasi_count,
            "synced_at": meta['last_siluxar_sync']}


@app.post("/zaloga-sync-siluxar")
async def zaloga_sync_siluxar():
    """Ročna sinhronizacija zaloge s siluxar (gumb). Scheduler kliče isto jedro vsake 4h."""
    return await _zaloga_sync_core()


async def _zaloga_scheduler_loop():
    """Notranji scheduler (always-on Render Pro): vsake 4 ure v ozadju potegne zalogo s siluxar.
    Teče znotraj web procesa → dostop do /data brez konflikta. Gumb ostane za ročni poteg."""
    await asyncio.sleep(120)   # počakaj, da se startup dokonča (in feed/zaloga naložita)
    INTERVAL = 4 * 60 * 60     # 4 ure
    while True:
        try:
            res = await _zaloga_sync_core()
            if res.get("ok"):
                print(f"[zaloga-cron] OK — total {res.get('total')}, "
                      f"added {res.get('added')}, updated {res.get('updated')}")
            else:
                print(f"[zaloga-cron] FAIL — {res.get('error')}")
            await asyncio.sleep(INTERVAL)
        except Exception as e:
            print(f"[zaloga-cron] Error: {e}")
            await asyncio.sleep(1800)   # 30 min pred ponovnim poskusom


@app.get("/zaloga-backups")
async def zaloga_backups():
    """Seznam backupov zaloge (najnovejši prvi) — za pregled arhiva in restore."""
    try:
        if not STOCK_BACKUP_DIR.exists():
            return {"ok": True, "backups": []}
        out = []
        for p in sorted(STOCK_BACKUP_DIR.glob("stock_*.csv"), key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                txt = p.read_text(encoding="utf-8-sig", errors="replace")
                vrstic = max(0, txt.count("\n") - 1)  # brez glave
            except Exception:
                vrstic = None
            stat = p.stat()
            # razberi čas + razlog iz imena: stock_YYYYMMDD_HHMMSS_razlog.csv
            ime = p.name
            cas_str = ""
            razlog = ""
            try:
                deli = ime.replace("stock_", "").replace(".csv", "").split("_")
                if len(deli) >= 2:
                    d, t = deli[0], deli[1]
                    cas_str = f"{d[:4]}-{d[4:6]}-{d[6:8]} {t[:2]}:{t[2:4]}:{t[4:6]}"
                if len(deli) >= 3:
                    razlog = "_".join(deli[2:])
            except Exception:
                pass
            out.append({
                "ime": ime,
                "cas": cas_str,
                "razlog": razlog,
                "vrstic": vrstic,
                "velikost_kb": round(stat.st_size / 1024, 1),
            })
        return {"ok": True, "stevilo": len(out), "backups": out}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/zaloga-backup-now")
async def zaloga_backup_now():
    """Ročno naredi backup trenutne zaloge (gumb)."""
    ime = _make_stock_backup("rocno")
    if ime:
        return {"ok": True, "ime": ime}
    return {"ok": False, "error": "Ni zaloge za backup (ali napaka)."}


@app.post("/zaloga-restore")
async def zaloga_restore(data: dict):
    """Obnovi zalogo iz izbranega backupa. PRED obnovo naredi backup trenutnega stanja (varnost)."""
    try:
        ime = (data.get("ime") or "").strip()
        if not ime or "/" in ime or "\\" in ime or not ime.startswith("stock_"):
            return {"ok": False, "error": "Neveljavno ime backupa."}
        src = STOCK_BACKUP_DIR / ime
        if not src.exists():
            return {"ok": False, "error": f"Backup '{ime}' ne obstaja."}
        # backup trenutnega stanja PRED obnovo (da lahko razveljaviš restore)
        _make_stock_backup("pred-restore")
        # obnovi
        STOCK_CSV_FILE.write_text(src.read_text(encoding="utf-8-sig", errors="replace"), encoding="utf-8")
        # posodobi meta
        try:
            meta = {}
            if STOCK_CSV_META.exists():
                meta = json.loads(STOCK_CSV_META.read_text(encoding="utf-8"))
            txt = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
            meta["rows"] = max(0, txt.count("\n") - 1)
            meta["restored_from"] = ime
            meta["restored_at"] = _lj_now().strftime("%Y-%m-%d %H:%M:%S")
            STOCK_CSV_META.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass
        return {"ok": True, "obnovljeno_iz": ime}
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "tb": traceback.format_exc()}


@app.get("/orodja-stock-status")
async def orodja_stock_status():
    """Vrne info o trenutno shranjeni zalogi."""
    if not STOCK_CSV_FILE.exists() or not STOCK_CSV_META.exists():
        return {"loaded": False}
    try:
        meta = json.loads(STOCK_CSV_META.read_text(encoding="utf-8"))
        return {"loaded": True, **meta}
    except Exception as e:
        return {"loaded": False, "error": str(e)}


@app.post("/orodja-stock-clear")
async def orodja_stock_clear():
    """Počisti zalogo (reset za nov upload)."""
    try:
        if STOCK_CSV_FILE.exists(): STOCK_CSV_FILE.unlink()
        if STOCK_CSV_META.exists(): STOCK_CSV_META.unlink()
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/maaarket-sku-gid")
async def maaarket_sku_gid(lang: str = "sl"):
    """Vrne mapiranje SKU → Google Shopping g:id iz maaarket feeda.
    Feed pogosto NIMA <g:mpn> — SKU je vgrajen v image_link URL. Zato gradimo
    mapo iz obeh virov: mpn (če obstaja) IN SKU-ji izluščeni iz slik.
    Uporablja price_checker za hitri link do maaarket admina."""
    try:
        feed = feed_by_lang.get(lang, {})
        if not feed:
            # feed še ni naložen — sproži osvežitev v ozadju, vrni prazno
            if is_cache_stale():
                asyncio.create_task(ensure_cache_fresh())
            return {"ok": True, "lang": lang, "count": 0, "map": {}, "note": "feed se nalaga, poskusi ponovno"}
        m = {}
        for g_id, data in feed.items():
            # 1) mpn (če feed ga ima)
            mpn = (data.get("mpn") or "").strip()
            if mpn:
                m.setdefault(mpn.upper(), g_id)
            # 2) SKU-ji iz slik (glavna + dodatne)
            for img in (data.get("all_images") or []):
                for sk in _extract_skus_from_image_url(img):
                    if sk:
                        m.setdefault(sk.upper(), g_id)
        return {"ok": True, "lang": lang, "count": len(m), "map": m}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e), "map": {}}


PRICE_CHECKER_CACHE = DATA_DIR / "price_checker_cache.json"

@app.get("/price-checker-cache")
async def price_checker_cache_get():
    """Vrne skupni cache urejevalnika cen (tabela + resolved) — ena resnica za vse uporabnike.
    Tako vsi brskalniki/naprave vidijo isto, ne kopičijo lokalnih starih podatkov."""
    try:
        if PRICE_CHECKER_CACHE.exists():
            data = json.loads(PRICE_CHECKER_CACHE.read_text(encoding="utf-8"))
            return {"ok": True, "rows": data.get("rows", []), "resolved": data.get("resolved", {}),
                    "updated_at": data.get("updated_at")}
        return {"ok": True, "rows": [], "resolved": {}, "updated_at": None}
    except Exception as e:
        return {"ok": False, "error": str(e), "rows": [], "resolved": {}}

@app.post("/price-checker-cache")
async def price_checker_cache_set(data: dict):
    """Shrani skupni cache urejevalnika cen. Pošlje frontend ob vsaki spremembi
    (poteg podatkov, brisanje, označevanje rešeno)."""
    try:
        from datetime import datetime as _dt
        rows = data.get("rows")
        resolved = data.get("resolved")
        # naloži obstoječe, da delna posodobitev ne pobriše drugega polja
        cur = {}
        if PRICE_CHECKER_CACHE.exists():
            try: cur = json.loads(PRICE_CHECKER_CACHE.read_text(encoding="utf-8"))
            except Exception: cur = {}
        out = {
            "rows": rows if rows is not None else cur.get("rows", []),
            "resolved": resolved if resolved is not None else cur.get("resolved", {}),
            "updated_at": _dt.now(timezone.utc).isoformat(),
        }
        tmp = PRICE_CHECKER_CACHE.with_suffix(".tmp")
        tmp.write_text(json.dumps(out, ensure_ascii=False), encoding="utf-8")
        os.replace(tmp, PRICE_CHECKER_CACHE)
        return {"ok": True, "count": len(out["rows"]), "updated_at": out["updated_at"]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/price-stock-fetch")
async def price_stock_fetch():
    """Potegne podatke o zalogi/cenah s siluxar.si API-ja (za urejevalnik cen).
    Ključ je v okoljski spr. SILUXAR_STOCK_KEY (NE v kodi). Vrne surovo besedilo,
    ki ga frontend (price_checker) spusti skozi obstoječi parseFile."""
    key = os.environ.get("SILUXAR_STOCK_KEY", "")
    if not key:
        return {"ok": False, "error": "Manjka SILUXAR_STOCK_KEY (nastavi v Render okoljskih spremenljivkah)."}
    basic_user = os.environ.get("SILUXAR_BASIC_USER", "")
    basic_pass = os.environ.get("SILUXAR_BASIC_PASS", "")
    url = "https://www.siluxar.si/apistockalertsexport"
    # Ključ v Authorization glavo (kot programerjev curl); Basic Auth samo kot fallback brez ključa.
    headers = {}
    _auth = None
    if key:
        headers["Authorization"] = key
    elif basic_user or basic_pass:
        _auth = httpx.BasicAuth(basic_user, basic_pass)
    try:
        async with httpx.AsyncClient(timeout=60, auth=_auth) as cli:
            r = await cli.get(url, headers=headers)
        if r.status_code != 200:
            return {"ok": False, "error": f"siluxar.si vrnil status {r.status_code}", "status": r.status_code}
        text = r.text or ""
        # poskusi razbrati ali je JSON (za morebitno kasnejšo pretvorbo)
        ctype = r.headers.get("content-type", "")
        return {
            "ok": True,
            "content_type": ctype,
            "raw": text,
            "length": len(text),
            "lines": text.count("\n") + 1 if text else 0,
        }
    except Exception as e:
        return {"ok": False, "error": f"Napaka pri klicu siluxar.si: {e}"}


@app.post("/orodja-price-check")
async def orodja_price_check(file: UploadFile = File(...)):
    """Sprejme HS+ PDF + match s shranjeno zalogo, vrne primerjavo cen."""
    if not STOCK_CSV_FILE.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Najprej naloži CSV zaloge."}, status_code=400)

    try:
        # 1. Preberi PDF s Claude Vision (isti pristop kot orodja-import-hs-pdf)
        content_bytes = await file.read()
        items = []  # [{ean, opis, sku, kolicina, cena_pdf, popust_pct}]
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp.write(content_bytes)
            tmp_path = tmp.name

        try:
            # Najprej poskus pdfplumber
            try:
                with pdfplumber.open(tmp_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        if text.strip():
                            for line in text.split('\n'):
                                line = line.strip()
                                # Format: ean opis kolicina KOS cena popust DDV znesek
                                m = re.match(r'^(\d{12,14})\s+(.+?)\s+(\d+)\s+(?:KOS|kos)\s+([\d.,]+)\s+(\d+)?', line)
                                if m:
                                    opis = m.group(2).strip()
                                    tokens = [t.rstrip('.,;:') for t in opis.split()]
                                    upper_t = [t for t in tokens if t.isupper() and len(t) >= 3 and not t.isdigit()]
                                    sku = upper_t[-1] if upper_t else (tokens[-1] if tokens else opis)
                                    try:
                                        cena = float(m.group(4).replace(',', '.'))
                                    except:
                                        cena = 0
                                    try:
                                        popust = float(m.group(5)) if m.group(5) else 0
                                    except:
                                        popust = 0
                                    items.append({
                                        "ean": m.group(1),
                                        "opis": opis,
                                        "sku": sku,
                                        "kolicina": int(m.group(3)),
                                        "cena_pdf": cena,
                                        "popust_pct": popust,
                                    })
            except: pass

            # Fallback: Claude Vision
            if not items:
                import base64
                pdf_b64 = base64.b64encode(content_bytes).decode('utf-8')
                prompt = """Preberi ta predračun in vrni VSA postavke v JSON formatu.
Za vsako postavko izloci:
- ean: 13-mestna številka koda na začetku vrstice
- opis: celoten opis postavke
- sku: zadnja SVE-VELIKA-ČRKA beseda v opisu (npr. "HYDRASPRINK HYDRASPRINK" → SKU = "HYDRASPRINK"; "WHEELPLAY yellow WHEELPLAY" → SKU = "WHEELPLAY"; "TOPKNER 180x200 TOPKNER" → SKU = "TOPKNER_180x200")
- kolicina: število pred "KOS" oznako
- cena_pdf: cena enote (po KOS, npr. "2,67")
- popust_pct: popust v % (število iz Popust % stolpca, lahko je prazno → 0)

POMEMBNO: Pri SKU-jih z dimenzijami (TOPKNER, WHEELPLAY) vključi tudi dimenzijo/barvo, ker so to različni izdelki:
- "TOPKNER 180x200 TOPKNER" → "TOPKNER_180x200"
- "TOPKNER 160x200 TOPKNER" → "TOPKNER_160x200"
- "WHEELPLAY yellow WHEELPLAY" → "WHEELPLAY_yellow"
- "PLANTUP (white) PLANTUP" → "PLANTUP_white"
- "BEEWAX WOOD POLISH BEEWAX" → "BEEWAX"

Vrni IZKLJUČNO valid JSON v formatu:
{"items": [{"ean": "...", "opis": "...", "sku": "...", "kolicina": 350, "cena_pdf": 2.67, "popust_pct": 5}, ...]}

Brez dodatnih komentarjev, samo JSON."""

                try:
                    client = anthropic.Anthropic()
                    response = client.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=8000,
                        messages=[{
                            "role": "user",
                            "content": [
                                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                                {"type": "text", "text": prompt}
                            ]
                        }]
                    )
                    text_resp = "".join([b.text for b in response.content if hasattr(b, 'text')])
                    parsed = parse_json_response(text_resp)
                    if parsed and 'items' in parsed:
                        for it in parsed['items']:
                            try: qty = int(it.get('kolicina', 0))
                            except: qty = 0
                            try: cena = float(str(it.get('cena_pdf', 0)).replace(',', '.'))
                            except: cena = 0
                            try: popust = float(str(it.get('popust_pct', 0)).replace(',', '.'))
                            except: popust = 0
                            items.append({
                                "ean": str(it.get('ean', '')),
                                "opis": str(it.get('opis', '')),
                                "sku": str(it.get('sku', '')).strip(),
                                "kolicina": qty,
                                "cena_pdf": cena,
                                "popust_pct": popust,
                            })
                except Exception as e:
                    print(f"[price-check] Claude error: {e}")
        finally:
            try: os.unlink(tmp_path)
            except: pass

        if not items:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "PDF parse fail."}, status_code=400)

        # 2. Naloži zalogo iz CSV
        import csv
        from io import StringIO
        stock_text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        # Detect separator
        first_line = stock_text.split('\n')[0]
        delim = ';' if first_line.count(';') > first_line.count(',') else ','
        reader = csv.DictReader(StringIO(stock_text), delimiter=delim)

        # Map SKU → cena (lower-case za case-insensitive match)
        stock_map = {}
        for row in reader:
            sku = (row.get('product_sku') or row.get('sku') or '').strip()
            if not sku:
                continue
            try:
                cena = float(str(row.get('price_netto') or row.get('price') or 0).replace(',', '.'))
            except:
                cena = 0
            title = (row.get('title') or '').strip()
            stock_map[sku.lower()] = {"sku": sku, "cena_zaloga": cena, "title": title}

        # Fuzzy match helper
        def find_stock_match(sku_pdf):
            """Najprej exact match, nato koren match (PLANTUP_white → PLANTUP), nato fuzzy."""
            sku_lower = sku_pdf.lower()
            # 1. Exact match
            if sku_lower in stock_map:
                return stock_map[sku_lower]

            # 2. Koren match — vzami osnovno besedo (PLANTUP_white → PLANTUP, COVERKA_2x3m → COVERKA)
            koren = re.split(r'[_\-\s]', sku_lower)[0]
            if not koren or len(koren) < 3:
                return None

            # Najdi vse SKU-je v zalogi ki začnejo s tem korenom
            candidates = [(k, v) for k, v in stock_map.items() if k.split('_')[0] == koren or k.split('-')[0] == koren or k == koren]

            if not candidates:
                return None

            # Če je samo eden, vrni ga
            if len(candidates) == 1:
                return candidates[0][1]

            # 3. Fuzzy match — pri več zadetkih izberi najbolj podobnega
            from difflib import SequenceMatcher
            best = None
            best_ratio = 0
            for k, v in candidates:
                ratio = SequenceMatcher(None, sku_lower, k).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best = v
            # Vrni samo če je dovolj podoben (>0.6)
            return best if best_ratio >= 0.6 else None

        # 3. Match in primerjava
        results = []
        for item in items:
            sku_pdf = item["sku"]
            stock = find_stock_match(sku_pdf)

            cena_pdf_neto = item["cena_pdf"] * (1 - item["popust_pct"] / 100)

            if stock:
                cena_zaloga = stock["cena_zaloga"]
                razlika = cena_pdf_neto - cena_zaloga
                razlika_pct = (razlika / cena_zaloga * 100) if cena_zaloga > 0 else 0
                if abs(razlika) < 0.001:
                    status = "match"
                elif razlika > 0:
                    status = "vecja"  # PDF cena VEČJA = SLABO za nas
                else:
                    status = "manjsa"  # PDF cena MANJŠA = DOBRO za nas
            else:
                cena_zaloga = None
                razlika = None
                razlika_pct = None
                status = "no_match"

            results.append({
                "sku": sku_pdf,
                "opis": item["opis"],
                "title_zaloga": stock["title"] if stock else None,
                "kolicina": item["kolicina"],
                "cena_pdf": item["cena_pdf"],
                "popust_pct": item["popust_pct"],
                "cena_pdf_neto": round(cena_pdf_neto, 4),
                "cena_zaloga": cena_zaloga,
                "razlika": round(razlika, 4) if razlika is not None else None,
                "razlika_pct": round(razlika_pct, 2) if razlika_pct is not None else None,
                "status": status,
            })

        return {"items": results, "total": len(results), "matched": sum(1 for r in results if r["status"] != "no_match")}
    except Exception as e:
        import traceback
        traceback.print_exc()
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


# ═══════════════════════════════════════════════════════════════════════════
# XSELL — AI-powered cross-sell suggestions
# Združi: zalogo (SKU, naziv, cena, marža) + XML feed (opis, kategorija, URL)
# ═══════════════════════════════════════════════════════════════════════════

XSELL_DIR = DATA_DIR / "xsell"
XSELL_DIR.mkdir(parents=True, exist_ok=True)
XSELL_CACHE_FILE = XSELL_DIR / "suggestions_cache.json"


def _load_stock_lookup() -> dict:
    """Vrne dict {SKU_uppercase: {sku, title, price, stock, ...}} iz CSV zaloge."""
    if not STOCK_CSV_FILE.exists():
        return {}
    try:
        import csv as _csv
        from io import StringIO as _SIO
        text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        first_line = text.split('\n', 1)[0]
        sep = ';' if first_line.count(';') > first_line.count(',') else ','
        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        lookup = {}
        for row in reader:
            sku = (row.get('product_sku') or row.get('sku') or '').strip()
            if not sku:
                continue
            lookup[sku.upper()] = {
                "sku": sku,
                "product_id": (row.get('product_id') or '').strip(),
                "title": (row.get('title') or '').strip(),
                "stock": int(float(row.get('stock') or 0) or 0),
                "stock30": int(float(row.get('stock30') or 0) or 0),
                "price": float(row.get('price_netto') or row.get('price') or 0),
                "position": (row.get('position') or '').strip(),
            }
        return lookup
    except Exception as e:
        print(f"[xsell] stock load error: {e}")
        return {}


def _find_feed_product(sku_or_url: str, lang: str = "sl") -> Optional[dict]:
    """Najdi izdelek v cached feed-u po SKU (MPN), product_id, slug ali URL."""
    feed = feed_by_lang.get(lang, {})
    if not feed:
        return None
    query = sku_or_url.strip()
    query_up = query.upper()

    # 1. URL ali pot — match po path
    if query.startswith('http') or '/' in query:
        path = urlparse(query).path if query.startswith('http') else query
        for g_id, p in feed.items():
            if p.get("path") == path or p.get("url") == query:
                return {"g_id": g_id, **p}
        # Po slug
        slug = path.rstrip('/').split('/')[-1].lower()
        for g_id, p in feed.items():
            p_slug = urlparse(p.get("url", "")).path.rstrip('/').split('/')[-1].lower()
            if slug and p_slug == slug:
                return {"g_id": g_id, **p}

    # 2. Direct ID match (g_id)
    if query in feed:
        return {"g_id": query, **feed[query]}
    if query_up in feed:
        return {"g_id": query_up, **feed[query_up]}

    # 3. MPN (SKU) match — both cases
    for g_id, p in feed.items():
        if (p.get("mpn") or "").upper() == query_up:
            return {"g_id": g_id, **p}

    # 4. SKU substring v slug (URL)
    query_lower = query.lower()
    for g_id, p in feed.items():
        slug = urlparse(p.get("url", "")).path.rstrip('/').split('/')[-1].lower()
        if query_lower and query_lower in slug:
            return {"g_id": g_id, **p}

    # 5. SKU substring v title (zadnja možnost)
    for g_id, p in feed.items():
        if query_up in (p.get("title") or "").upper():
            return {"g_id": g_id, **p}

    return None


@app.get("/xsell-debug")
async def xsell_debug(lang: str = "sl"):
    """Diagnostika: pokaže sample feed + stock data za debugging match-a."""
    stock_lookup = _load_stock_lookup()
    feed = feed_by_lang.get(lang, {})

    sample_stock = []
    for sku_up, st in list(stock_lookup.items())[:10]:
        sample_stock.append({
            "sku": st["sku"],
            "product_id": st.get("product_id"),
            "title": st.get("title", "")[:80],
        })

    sample_feed = []
    for g_id, p in list(feed.items())[:10]:
        slug = urlparse(p.get("url", "")).path.rstrip('/').split('/')[-1]
        sample_feed.append({
            "g_id": g_id,
            "mpn": p.get("mpn", ""),
            "title": (p.get("title") or "")[:80],
            "slug": slug,
            "has_description": bool(p.get("description")),
        })

    return {
        "ok": True,
        "stock_total": len(stock_lookup),
        "feed_total": len(feed),
        "sample_stock": sample_stock,
        "sample_feed": sample_feed,
    }


@app.get("/xsell-feed-status")
async def xsell_feed_status():
    """Vrne status XML feed-a."""
    return {
        "ok": True,
        "last_fetch": last_fetch.isoformat() if last_fetch else None,
        "languages": {lang: len(prods) for lang, prods in feed_by_lang.items()},
        "total_sl": len(feed_by_lang.get("sl", {})),
        "stock_loaded": STOCK_CSV_FILE.exists(),
    }


def _scrape_maaarket_crosssell(product_url: str) -> list[dict]:
    """Iz Maaarket strani izdelka pobere 'Kupci pogosto izberejo še' linke.
    Vrne seznam {url, title} za predlagane izdelke."""
    if not product_url or 'maaarket.' not in product_url:
        return []
    try:
        import urllib.request
        req = urllib.request.Request(product_url, headers={
            'User-Agent': 'Mozilla/5.0 (compatible; Maaarket-XSell-Bot/1.0)'
        })
        with urllib.request.urlopen(req, timeout=8) as resp:
            html = resp.read().decode('utf-8', errors='replace')
    except Exception as e:
        print(f"[xsell-scrape] Failed to fetch {product_url}: {e}")
        return []

    # Najdi "Kupci pogosto izberejo še" sekcijo
    # Vzorec: <h2|h3|div>...Kupci pogosto izberejo še...</...>
    # Pri njej so <a href="/izdelek/..."> linki
    results = []
    # Heuristika: poišči vse linke /izdelek/ v zadnji tretjini strani (cross-sell je običajno na dnu)
    section_start = html.lower().find('kupci pogosto')
    if section_start < 0:
        section_start = html.lower().find('cross-sell')
    if section_start < 0:
        # Fallback: zadnja tretjina strani
        section_start = len(html) * 2 // 3

    section_html = html[section_start:]
    # Najdi /izdelek/ linke
    link_pattern = re.compile(r'href="(/izdelek/[^"#?]+)"[^>]*(?:title="([^"]+)")?')
    seen_slugs = set()
    for match in link_pattern.finditer(section_html):
        slug_path = match.group(1)
        title = (match.group(2) or '').strip()
        slug = slug_path.rstrip('/').split('/')[-1]
        if slug in seen_slugs:
            continue
        seen_slugs.add(slug)
        # Sestavi full URL
        from urllib.parse import urljoin
        full_url = urljoin(product_url, slug_path)
        results.append({"url": full_url, "slug": slug, "title": title})
        if len(results) >= 5:
            break
    print(f"[xsell-scrape] Found {len(results)} cross-sell links from {product_url}")
    return results


def _extract_category_from_url(url: str) -> Optional[str]:
    """Iz Maaarket URL-ja izvleče kategorijo (če je v 'izdelki' poti).
    Žal pa /izdelek/SLUG nima poti — zato to delamo iz scraped breadcrumbs ali category page."""
    if not url:
        return None
    path = urlparse(url).path
    # Maaarket strukture: /izdelek/{slug} — kategorija ni v URL-u
    # Vendar lahko fetch-amo stran in poiščemo breadcrumb
    return None


def _fetch_maaarket_category(product_url: str) -> Optional[str]:
    """Z fetcha strani pobere kategorijo iz breadcrumb-a."""
    if not product_url or 'maaarket.' not in product_url:
        return None
    try:
        import urllib.request
        req = urllib.request.Request(product_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=8) as resp:
            html = resp.read().decode('utf-8', errors='replace')
        # Najdi breadcrumb: ponavadi je tukaj /izdelki/{category-slug}
        match = re.search(r'/izdelki/([a-z0-9-]+)', html)
        if match:
            cat_slug = match.group(1)
            print(f"[xsell-cat] Detected category slug: {cat_slug}")
            return cat_slug
    except Exception as e:
        print(f"[xsell-cat] Failed: {e}")
    return None


@app.post("/xsell-suggest")
async def xsell_suggest(data: dict):
    """Vrne 3 Xsell predloge za vhodni SKU ali URL.

    Vhod: {"input": "M260" or "https://maaarket.si/product/...", "lang": "sl"}
    Izhod: {"original": {...}, "suggestions": [{...}, {...}, {...}]}
    """
    try:
        query = (data.get("input") or "").strip()
        lang = (data.get("lang") or "sl").strip()
        if not query:
            return {"ok": False, "error": "Vnesi SKU ali URL"}

        # Cache lookup (7-dnevni)
        cache_key = f"{lang}:{query.upper()}"
        try:
            if XSELL_CACHE_FILE.exists():
                cache = json.loads(XSELL_CACHE_FILE.read_text(encoding="utf-8"))
                if cache_key in cache:
                    cached = cache[cache_key]
                    cached_at = datetime.fromisoformat(cached.get("cached_at", ""))
                    if datetime.now() - cached_at < timedelta(days=7):
                        return {"ok": True, "cached": True, **cached["result"]}
        except Exception:
            pass

        # 1. Najdi izdelek v feed-u (opis, kategorija)
        feed_product = _find_feed_product(query, lang)

        # 2. Najdi tudi v zalogi (cena, stock)
        stock_lookup = _load_stock_lookup()
        if not stock_lookup:
            return {"ok": False, "error": "Zaloga ni naložena. Najprej naloži CSV zaloge v Zaloga tab."}

        # Najdi originalni izdelek v zalogi
        original_sku = None
        if feed_product:
            mpn = (feed_product.get("mpn") or "").upper()
            if mpn and mpn in stock_lookup:
                original_sku = mpn
        if not original_sku:
            if query.upper() in stock_lookup:
                original_sku = query.upper()

        if not original_sku and not feed_product:
            return {"ok": False, "error": f"Izdelek '{query}' ni najden v zalogi ali XML feed-u"}

        # Sestavi originalni izdelek
        original = {}
        if original_sku and original_sku in stock_lookup:
            original.update(stock_lookup[original_sku])
        if feed_product:
            original.update({
                "title_full": feed_product.get("title", ""),
                "description": feed_product.get("description", ""),
                "url": feed_product.get("url", ""),
                "image": feed_product.get("image", ""),
                "product_type": feed_product.get("product_type", ""),
                "google_category": feed_product.get("google_category", ""),
                "brand": feed_product.get("brand", ""),
                "feed_price": feed_product.get("price", ""),
            })

        # 3. Sestavi katalog kandidatov (presek zaloge + feed)
        # Match po: MPN, product_id, slug v URL, ali ime izdelka (slov. ujemanje)
        candidates = []
        feed = feed_by_lang.get(lang, {})

        # Naredi multiple lookup index-ov za feed
        feed_by_mpn = {}      # SKU upper -> feed product
        feed_by_gid = {}      # g_id -> feed product (običajno product_id)
        feed_by_slug = {}     # last URL slug -> feed product
        feed_by_title_norm = {}  # normalized title -> feed product

        def _norm(s):
            return re.sub(r'[^a-z0-9]', '', (s or '').lower())

        for g_id, p in feed.items():
            feed_by_gid[g_id] = {"g_id": g_id, **p}
            mpn = (p.get("mpn") or "").upper().strip()
            if mpn:
                feed_by_mpn[mpn] = {"g_id": g_id, **p}
            url = p.get("url") or ""
            slug = urlparse(url).path.rstrip('/').split('/')[-1].lower() if url else ""
            if slug:
                feed_by_slug[slug] = {"g_id": g_id, **p}
            title_norm = _norm(p.get("title", ""))
            if title_norm:
                feed_by_title_norm[title_norm] = {"g_id": g_id, **p}

        # Pomožni iskalniki feed produktov po SKU/title
        def _find_feed_for_stock_item(stock):
            sku_up = stock["sku"].upper()
            pid = (stock.get("product_id") or "").strip()
            title = stock.get("title", "")
            title_norm = _norm(title)
            # 1. MPN match
            if sku_up in feed_by_mpn:
                return feed_by_mpn[sku_up]
            # 2. product_id == g_id
            if pid and pid in feed_by_gid:
                return feed_by_gid[pid]
            # 3. SKU == g_id (some feeds use SKU as id)
            if sku_up in feed_by_gid:
                return feed_by_gid[sku_up]
            # 4. Naziv match (normaliziran)
            if title_norm and title_norm in feed_by_title_norm:
                return feed_by_title_norm[title_norm]
            # 5. Substring v slug
            if sku_up:
                sku_lower = sku_up.lower()
                for slug, fp in feed_by_slug.items():
                    if sku_lower in slug:
                        return fp
            return None

        matched_in_feed = 0
        for sku_up, stock in stock_lookup.items():
            if sku_up == original_sku:
                continue
            # ⛔ stock filter ODSTRANJEN — vključi tudi izdelke s stock=0 (nova zaloga lahko pride takoj)
            fp = _find_feed_for_stock_item(stock)
            if not fp:
                continue
            matched_in_feed += 1
            candidates.append({
                "sku": stock["sku"],
                "title": stock["title"] or fp.get("title", ""),
                "price": stock["price"],
                "stock": stock["stock"],
                "stock30": stock["stock30"],
                "description": (fp.get("description") or "")[:300],
                "product_type": fp.get("product_type", ""),
                "url": fp.get("url", ""),
                "image": fp.get("image", ""),
            })

        print(f"[xsell] Stock items: {len(stock_lookup)}, matched in feed: {matched_in_feed}, candidates: {len(candidates)}")

        if len(candidates) < 3:
            return {
                "ok": False,
                "error": f"Premalo kandidatov ({len(candidates)}) za {len(stock_lookup)} izdelkov v zalogi. Verjetno se SKU-ji iz zaloge ne ujemajo z MPN/g_id v XML feed-u. Preveri feed format (admin debug).",
                "debug": {
                    "stock_items": len(stock_lookup),
                    "feed_items": len(feed),
                    "matched": matched_in_feed,
                    "sample_stock_sku": list(stock_lookup.keys())[:5],
                    "sample_feed_mpns": [feed_by_mpn[k]["mpn"] for k in list(feed_by_mpn.keys())[:5]],
                    "sample_feed_gids": list(feed_by_gid.keys())[:5],
                }
            }

        # === SCRAPE MAAARKET CROSS-SELL + KATEGORIJA ===
        # Če imamo URL originalnega izdelka, scrape "Kupci pogosto izberejo še" + breadcrumb kategorijo
        scraped_xsell_urls = []
        scraped_category_slug = None
        original_url = original.get("url") or (feed_product.get("url") if feed_product else "")
        if original_url:
            try:
                # Fetch enkrat, parse oboje
                import urllib.request
                req = urllib.request.Request(original_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=8) as resp:
                    page_html = resp.read().decode('utf-8', errors='replace')

                # 1. Najdi kategorijo iz breadcrumb-a
                cat_match = re.search(r'/izdelki/([a-z0-9-]+)', page_html)
                if cat_match:
                    scraped_category_slug = cat_match.group(1)
                    print(f"[xsell] Category slug: {scraped_category_slug}")

                # 2. Najdi cross-sell linke (zadnja tretjina strani)
                section_start = page_html.lower().find('kupci pogosto')
                if section_start < 0:
                    section_start = len(page_html) * 2 // 3
                section_html = page_html[section_start:]
                link_pattern = re.compile(r'href="(/izdelek/([^"#?/]+))"')
                seen_slugs = set()
                from urllib.parse import urljoin
                for match in link_pattern.finditer(section_html):
                    slug = match.group(2)
                    if slug in seen_slugs or slug == urlparse(original_url).path.rstrip('/').split('/')[-1]:
                        continue
                    seen_slugs.add(slug)
                    full_url = urljoin(original_url, match.group(1))
                    scraped_xsell_urls.append({"url": full_url, "slug": slug})
                    if len(scraped_xsell_urls) >= 5:
                        break
                print(f"[xsell] Scraped {len(scraped_xsell_urls)} cross-sell URLs")
            except Exception as e:
                print(f"[xsell] Scrape error: {e}")

        # Najdi scraped izdelke med kandidati (po slug match)
        scraped_candidates = []
        for sx in scraped_xsell_urls:
            for c in candidates:
                c_slug = urlparse(c.get("url", "")).path.rstrip('/').split('/')[-1].lower()
                if c_slug == sx["slug"].lower():
                    scraped_candidates.append(c)
                    break
        print(f"[xsell] Matched {len(scraped_candidates)} scraped to feed candidates")

        # === KATEGORIJA-AWARE PRE-FILTERING ===
        # Najprej probaj URL slug kategorije (zelo zanesljivo za Maaarket)
        # Sicer fallback na product_type iz feed-a
        original_ptype = (original.get("product_type") or "").strip()
        original_ptype_parts = [p.strip().lower() for p in re.split(r'\s*[>/]\s*', original_ptype) if p.strip()]
        original_top_cat = original_ptype_parts[0] if original_ptype_parts else ""
        original_sub_cat = original_ptype_parts[1] if len(original_ptype_parts) > 1 else ""
        print(f"[xsell] Original category: top='{original_top_cat}', sub='{original_sub_cat}', full='{original_ptype}'")

        # Cache za kategorijo izdelkov iz URL-jev (da ne fetchamo isti URL večkrat)
        # Vrnemo izdelke ki imajo isto URL kategorijo

        # Najprej probaj poiskati izdelke iste kategorije preko **fetcha kategorije strani**
        # ALI alternative: filter kandidatov po sorodnih besedah v slugu
        same_category_slugs = set()
        if scraped_category_slug:
            # Fetch kategorijo stran in pridobi vse SKU slug-e
            try:
                cat_url = f"https://www.maaarket.si/izdelki/{scraped_category_slug}"
                import urllib.request
                req = urllib.request.Request(cat_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=10) as resp:
                    cat_html = resp.read().decode('utf-8', errors='replace')
                # Najdi vse /izdelek/{slug} v kategoriji
                for match in re.finditer(r'/izdelek/([a-z0-9-]+)', cat_html):
                    same_category_slugs.add(match.group(1).lower())
                print(f"[xsell] Category '{scraped_category_slug}' has {len(same_category_slugs)} products")
            except Exception as e:
                print(f"[xsell] Category fetch failed: {e}")

        def _cat_distance(cand):
            """Vrne 0 (ista kategorija po URL slug-u), 1 (sosednja sub), 2 (samo top), 3 (drugo)."""
            # NAJZANESLJIVEJŠE: ujemanje po URL slug-u kategorije
            cand_slug = urlparse(cand.get("url", "")).path.rstrip('/').split('/')[-1].lower()
            if same_category_slugs and cand_slug in same_category_slugs:
                return 0
            # Fallback na product_type ujemanje
            cand_ptype = cand.get("product_type", "")
            cand_parts = [p.strip().lower() for p in re.split(r'\s*[>/]\s*', cand_ptype) if p.strip()]
            if not cand_parts or not original_top_cat:
                return 3
            cand_top = cand_parts[0]
            cand_sub = cand_parts[1] if len(cand_parts) > 1 else ""
            if cand_ptype.strip().lower() == original_ptype.strip().lower():
                return 0
            if cand_top == original_top_cat and original_sub_cat and cand_sub == original_sub_cat:
                return 0
            if cand_top == original_top_cat:
                return 1
            orig_words = set(re.findall(r'[a-zA-ZčšžćđČŠŽĆĐ]+', original_ptype.lower()))
            cand_words = set(re.findall(r'[a-zA-ZčšžćđČŠŽĆĐ]+', cand_ptype.lower()))
            common = orig_words & cand_words
            common = {w for w in common if len(w) > 3}
            if len(common) >= 2:
                return 2
            if len(common) >= 1:
                return 3
            return 3

        # Razvrsti kandidate v 3 nivoje
        same_cat = []
        adjacent_cat = []
        other = []
        for c in candidates:
            d = _cat_distance(c)
            if d == 0:
                same_cat.append(c)
            elif d <= 1:
                adjacent_cat.append(c)
            elif d <= 2:
                other.append(c)  # poseben tier
        # Sortiraj po podobnosti naziva (preprosti word overlap)
        orig_title_words = set(re.findall(r'[a-zA-ZčšžćđČŠŽĆĐ]{3,}', (original.get("title") or original.get("title_full") or "").lower()))
        def _title_score(c):
            cw = set(re.findall(r'[a-zA-ZčšžćđČŠŽĆĐ]{3,}', (c.get("title") or "").lower()))
            return len(cw & orig_title_words)
        same_cat.sort(key=_title_score, reverse=True)
        adjacent_cat.sort(key=_title_score, reverse=True)
        other.sort(key=_title_score, reverse=True)

        print(f"[xsell] Category distribution: same={len(same_cat)}, adjacent={len(adjacent_cat)}, other={len(other)}")

        # Vzami uravnoteženi nabor: 70% iz iste/sosednje kategorije, 30% iz drugih
        ranked = same_cat[:35] + adjacent_cat[:25] + other[:20]
        if len(ranked) < 30:
            ranked = (same_cat + adjacent_cat + other)[:60]
        candidates_for_ai = ranked[:80]

        # Pripravi text za Claude — z eksplicitno kategorijsko strukturo
        def _fmt_cand(c, marker=""):
            return f"{marker}SKU={c['sku']} | {c['title'][:80]} | {c['price']:.2f}€ | stock={c['stock']} | kategorija={c.get('product_type','')[:80]} | opis={c['description'][:140]}"

        # Razdeli v sekcije za prompt
        same_cat_for_prompt = [c for c in candidates_for_ai if _cat_distance(c) == 0]
        adjacent_for_prompt = [c for c in candidates_for_ai if _cat_distance(c) == 1]
        other_for_prompt = [c for c in candidates_for_ai if _cat_distance(c) >= 2]

        # Če imamo malo categorisanih kandidatov (feed verjetno nima product_type),
        # potem damo VSE kandidate v "ostalo" da AI sam odloča.
        if not original_ptype or (len(same_cat_for_prompt) + len(adjacent_for_prompt) < 5):
            sections = ["📋 VSI KANDIDATI (kategorijski podatki niso jasni, izberi po nazivu/opisu):\n"
                        + "\n".join(_fmt_cand(c) for c in candidates_for_ai[:70])]
        else:
            sections = []
            if same_cat_for_prompt:
                sections.append("🎯 ISTA KATEGORIJA — PRIORITETA (priporočam izbrati prvi 2 predloga iz tega seznama):\n" + "\n".join(_fmt_cand(c) for c in same_cat_for_prompt[:30]))
            if adjacent_for_prompt:
                sections.append("📂 SOSEDNJA KATEGORIJA:\n" + "\n".join(_fmt_cand(c) for c in adjacent_for_prompt[:25]))
            if other_for_prompt:
                sections.append("🔀 OSTALI (uporabi LE če nimaš dovolj iz zgornjih kategorij):\n" + "\n".join(_fmt_cand(c) for c in other_for_prompt[:15]))
        cand_text = "\n\n".join(sections)

        original_desc = (original.get("description") or "")[:500]
        prompt = f"""Si strokovnjak za cross-sell v e-commerce trgovini Maaarket.si.

ORIGINALNI IZDELEK:
SKU: {original.get('sku', '?')}
Naziv: {original.get('title_full') or original.get('title', '?')}
Cena: {original.get('price', 0):.2f}€
Kategorija: {original_ptype or '(neznana — sklepaj iz naziva in opisa)'}
Opis: {original_desc}

KATALOG MOŽNIH XSELL IZDELKOV:

{cand_text}

KRITIČNO: Izberi 3 izdelke ki so res POVEZANI z originalnim izdelkom po NAMENU UPORABE.
Npr: izdelek za nego obraza → drugi face/skin care izdelki (NE kuhinjski pripomočki).
Če v katalogu obstajajo izdelki iz iste kategorije/namena, jih daj prednost.
Če pa ne najdeš popolnoma sorodne kategorije, izberi izdelke ki so VSAJ delno povezani po uporabniku ali namenu uporabe.

OBVEZNO vrni 3 predloge — če nisi prepričan kateri so najbolj sorodni, raje izberi tiste z najbolj podobnim opisom/namenom uporabe.

Vrni 3 raznolike predloge — vsakega drugačnega tipa:
1. KOMPLEMENTAREN: dopolnilo/pribor (npr. nadomestne glave, polnilo, dodatek)
2. KATEGORIJSKO PODOBEN: alternativa istega namena
3. CENOVNO SMISELN: add-on po nižji ceni iz iste/sosednje kategorije

Za vsakega:
- SKU (točno iz seznama)
- type ("komplementaren" | "kategorijsko_podoben" | "cenovno_smiseln")
- reason: 1 stavek ZAKAJ se odlično dopolnjuje (max 100 znakov, slovenščina)

VRNI EXACT JSON, brez dodatnega teksta, OBVEZNO 3 predloge:
{{
  "suggestions": [
    {{"sku": "...", "type": "komplementaren", "reason": "..."}},
    {{"sku": "...", "type": "kategorijsko_podoben", "reason": "..."}},
    {{"sku": "...", "type": "cenovno_smiseln", "reason": "..."}}
  ]
}}"""

        ai_response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1200,
            messages=[{"role": "user", "content": prompt}],
        )
        ai_text = ai_response.content[0].text.strip()
        # Strip ```json fences
        ai_text = re.sub(r'^```(?:json)?\s*', '', ai_text)
        ai_text = re.sub(r'\s*```$', '', ai_text)
        print(f"[xsell] AI response (first 600 chars): {ai_text[:600]}")

        try:
            ai_data = json.loads(ai_text)
        except json.JSONDecodeError:
            # Fallback: probaj najti JSON v odgovoru
            m = re.search(r'\{[\s\S]*\}', ai_text)
            if m:
                ai_data = json.loads(m.group())
            else:
                return {"ok": False, "error": f"AI vrnil neveljaven JSON: {ai_text[:200]}"}

        # 5. Sestavi končni odgovor — dopolni z metadata
        cand_by_sku = {c["sku"].upper(): c for c in candidates}
        suggestions_out = []
        used_skus = set()

        # === PRIORITY 1: Vstavi 1 scraped Maaarket cross-sell predlog (top spot) ===
        if scraped_candidates:
            sc = scraped_candidates[0]
            suggestions_out.append({
                "sku": sc["sku"],
                "type": "maaarket_pick",
                "reason": "Maaarket priporočilo: 'Kupci pogosto izberejo še'",
                "title": sc["title"],
                "price": sc["price"],
                "stock": sc["stock"],
                "stock30": sc["stock30"],
                "url": sc.get("url", ""),
                "image": sc.get("image", ""),
            })
            used_skus.add(sc["sku"].upper())

        missing_skus = []
        # === PRIORITY 2: AI predlogi (2 dodatna) ===
        for s in ai_data.get("suggestions", [])[:3]:
            if len(suggestions_out) >= 3:
                break
            sku = (s.get("sku") or "").strip().upper()
            if sku in used_skus:
                continue
            cand = cand_by_sku.get(sku)
            if not cand:
                missing_skus.append(sku)
                continue
            suggestions_out.append({
                "sku": cand["sku"],
                "type": s.get("type", ""),
                "reason": s.get("reason", ""),
                "title": cand["title"],
                "price": cand["price"],
                "stock": cand["stock"],
                "stock30": cand["stock30"],
                "url": cand.get("url", ""),
                "image": cand.get("image", ""),
            })
            used_skus.add(sku)

        if missing_skus:
            print(f"[xsell] WARN: AI returned SKUs not in candidates: {missing_skus}")

        # Fallback: VEDNO dopolni do 3 predlogov, ne glede na vir
        # Prioritetni vrstni red: same_cat > adjacent > other > kakršenkoli kandidat
        if len(suggestions_out) < 3:
            type_order = ["komplementaren", "kategorijsko_podoben", "cenovno_smiseln"]
            existing_types = {s["type"] for s in suggestions_out if s["type"] in type_order}
            # Zgradi fallback pool po prioritetah
            fallback_pool = list(same_cat_for_prompt) + list(adjacent_for_prompt) + list(other_for_prompt)
            # Če je še vedno prazno (ker ni kategorij), uporabi VSE kandidate
            if not fallback_pool:
                fallback_pool = candidates
            print(f"[xsell] AI vrnil le {len(suggestions_out)}, dopolnjujem iz fallback pool-a ({len(fallback_pool)} kandidatov)")

            for cand in fallback_pool:
                if len(suggestions_out) >= 3:
                    break
                if cand["sku"].upper() in used_skus:
                    continue
                # Najdi naslednji manjkajoči tip
                next_type = next((t for t in type_order if t not in existing_types), "kategorijsko_podoben")
                existing_types.add(next_type)
                suggestions_out.append({
                    "sku": cand["sku"],
                    "type": next_type,
                    "reason": "Iz iste/sosednje kategorije (AI fallback)" if cand in same_cat_for_prompt or cand in adjacent_for_prompt else "Najboljši preostali kandidat (fallback)",
                    "title": cand["title"],
                    "price": cand["price"],
                    "stock": cand["stock"],
                    "stock30": cand["stock30"],
                    "url": cand.get("url", ""),
                    "image": cand.get("image", ""),
                })
                used_skus.add(cand["sku"].upper())

        print(f"[xsell] Final suggestions count: {len(suggestions_out)}")

        result = {
            "original": {
                "sku": original.get("sku") or original_sku or query,
                "title": original.get("title_full") or original.get("title", ""),
                "price": original.get("price", 0),
                "stock": original.get("stock", 0),
                "url": original.get("url", ""),
                "image": original.get("image", ""),
                "product_type": original.get("product_type", ""),
            },
            "suggestions": suggestions_out,
            "candidates_total": len(candidates),
        }

        # Cache za 7 dni
        try:
            cache = {}
            if XSELL_CACHE_FILE.exists():
                cache = json.loads(XSELL_CACHE_FILE.read_text(encoding="utf-8"))
            cache[cache_key] = {"cached_at": datetime.now().isoformat(), "result": result}
            # Cleanup starih (> 7 dni)
            now_ts = datetime.now()
            cache = {k: v for k, v in cache.items()
                     if (now_ts - datetime.fromisoformat(v.get("cached_at", "1970-01-01"))).days < 7}
            XSELL_CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            print(f"[xsell] cache save error: {e}")

        return {"ok": True, "cached": False, **result}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.post("/xsell-cache-clear")
async def xsell_cache_clear():
    """Pobriše Xsell cache (forsiraj fresh AI re-rank)."""
    try:
        if XSELL_CACHE_FILE.exists():
            XSELL_CACHE_FILE.unlink()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/orodja-stock-data")
async def orodja_stock_data():
    """Vrne celoten seznam zaloge iz shranjenega CSV."""
    if not STOCK_CSV_FILE.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Najprej naloži CSV zaloge."}, status_code=400)
    try:
        import csv as _csv
        from io import StringIO as _SIO
        text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        first_line = text.split('\n', 1)[0]
        sep = ';' if first_line.count(';') > first_line.count(',') else ','

        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        items = []
        def _calc_trajanje(stock_s, s30_s):
            """Trajanje zaloge v dnevih = zaloga / (obrat30 / 30).
            Primer: obrat 150 → 5/dan; zaloga 25 → 25/5 = 5 dni."""
            try:
                st = float(str(stock_s).replace(',', '.'))
                s30 = float(str(s30_s).replace(',', '.'))
            except Exception:
                return ''
            if s30 <= 0:
                return ''            # ni obrata (ni prodaje) → trajanje ni smiselno
            per_day = s30 / 30.0
            if per_day <= 0:
                return ''
            days = st / per_day
            return str(int(round(days)))  # cele dni
        for row in reader:
            sku = (row.get('product_sku') or row.get('sku') or '').strip()
            if not sku:
                continue
            _stock = (row.get('stock') or '0').strip()
            _s30 = (row.get('stock30') or '0').strip()
            items.append({
                "sku": sku,
                "product_sku": sku,
                "product_id": (row.get('product_id') or row.get('siluxar_id') or '').strip(),
                "siluxar_id": (row.get('siluxar_id') or '').strip(),
                "title": (row.get('title') or '').strip(),
                "stock": _stock,
                "stock30": _s30,
                "price": (row.get('price_netto') or row.get('price') or '0').strip(),
                "position": (row.get('position') or '').strip(),
                "note": (row.get('note') or '').strip(),
                "stock_duration": _calc_trajanje(_stock, _s30),
                "warehouse": (row.get('warehouse') or '').strip(),
                "is_external": (row.get('is_external') or '').strip() == '1',
            })

        meta = {}
        if STOCK_CSV_META.exists():
            try:
                meta = json.loads(STOCK_CSV_META.read_text(encoding="utf-8"))
            except: pass

        return {"items": items, "total": len(items), **meta}
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── ANALIZA: TikTok Ads upload ──────────────────────────────────────────────

TIKTOK_CREATIVE_FILE = DATA_DIR / "tiktok_creative_map.json"

@app.post("/tiktok-creative-upload")
async def tiktok_creative_upload(file: UploadFile = File(...)):
    """Naloži TikTok Ad level export — zgradi Video→SKU mapping."""
    try:
        import io, re as _re
        content = await file.read()
        fname = file.filename or ""

        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows_raw = list(ws.iter_rows(values_only=True))
            headers = [str(h).strip() if h else '' for h in rows_raw[0]]
            data_rows = [dict(zip(headers, r)) for r in rows_raw[1:] if any(v is not None for v in r)]
        else:
            import csv as _csv
            text = content.decode('utf-8-sig', errors='replace')
            sep = ';' if text.split('\n')[0].count(';') > text.split('\n')[0].count(',') else ','
            data_rows = list(_csv.DictReader(io.StringIO(text), delimiter=sep))

        def extract_sku(campaign):
            c = str(campaign or '')
            m = _re.search(r'SKU:\s*([A-Za-z0-9_]+)', c)
            if m: return smart_root(m.group(1)).upper()
            m = _re.search(r'Smart\+\s+([A-Za-z0-9_]+)', c, _re.I)
            if m: return smart_root(m.group(1)).upper()
            return ''

        # Zgradi mapping: {sku: [{video, cost, conversions, status}]}
        sku_map = {}
        skipped = 0
        for row in data_rows:
            video = str(row.get('Video') or '').strip()
            if not video or video == '-': skipped += 1; continue
            campaign = row.get('Campaign name') or row.get('Campaign Name') or ''
            sku = extract_sku(campaign)
            if not sku: skipped += 1; continue

            try: cost = float(str(row.get('Cost') or 0).replace(',', '.'))
            except: cost = 0
            try: conversions = int(float(str(row.get('Conversions') or 0).replace(',', '.')))
            except: conversions = 0
            status = str(row.get('Primary status') or '').strip().lower()

            if sku not in sku_map:
                sku_map[sku] = {}
            if video not in sku_map[sku]:
                sku_map[sku][video] = {'video': video, 'cost': 0, 'conversions': 0, 'status': status}
            sku_map[sku][video]['cost'] += cost
            sku_map[sku][video]['conversions'] += conversions
            if status == 'active': sku_map[sku][video]['status'] = 'active'

        # Pretvori v seznam in sortiraj po cost desc
        result = {}
        for sku, videos in sku_map.items():
            result[sku] = sorted(videos.values(), key=lambda x: x['cost'], reverse=True)

        TIKTOK_CREATIVE_FILE.write_text(
            json.dumps({'map': result, 'uploaded_at': __import__('datetime').datetime.now().isoformat(), 'filename': fname, 'skus': len(result)}, ensure_ascii=False),
            encoding='utf-8'
        )
        return {"ok": True, "skus": len(result), "videos": sum(len(v) for v in result.values()), "skipped": skipped}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/tiktok-creative-search")
async def tiktok_creative_search(sku: str = ""):
    """Poišči videe za SKU."""
    if not TIKTOK_CREATIVE_FILE.exists():
        return JSONResponse({"error": "Naloži Creative CSV najprej."}, status_code=400)
    try:
        data = json.loads(TIKTOK_CREATIVE_FILE.read_text(encoding='utf-8'))
        cmap = data.get('map', {})
        sku_up = sku.upper().strip()
        if not sku_up:
            return {"skus": list(cmap.keys()), "videos": []}
        # Išči po SKU ali korenu
        root = smart_root(sku_up)
        videos = cmap.get(sku_up) or cmap.get(root) or []
        # Fuzzy: če ni exact match, vrni SKU-je ki vsebujejo iskalni niz
        if not videos:
            matches = {k: v for k, v in cmap.items() if sku_up in k or root in k}
            return {"skus": list(matches.keys()), "videos": [], "fuzzy": True}
        return {"sku": sku_up, "videos": videos, "total": len(videos)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/tiktok-creative-info")
async def tiktok_creative_info():
    if not TIKTOK_CREATIVE_FILE.exists():
        return {"loaded": False}
    try:
        data = json.loads(TIKTOK_CREATIVE_FILE.read_text(encoding='utf-8'))
        return {"loaded": True, "skus": data.get('skus', 0), "uploaded_at": data.get('uploaded_at'), "filename": data.get('filename')}
    except:
        return {"loaded": False}


TIKTOK_ADS_FILE = DATA_DIR / "tiktok_ads_report.csv"
TIKTOK_ADS_META = DATA_DIR / "tiktok_ads_meta.json"

@app.post("/analiza-tiktok-upload")
async def analiza_tiktok_upload(file: UploadFile = File(...)):
    """Naloži TikTok XLSX/CSV Campaign Report."""
    try:
        content = await file.read()
        fname = file.filename or ""

        # Preberi XLSX ali CSV
        import io
        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows_raw = list(ws.iter_rows(values_only=True))
            if not rows_raw: return JSONResponse({"error": "Prazna datoteka."}, status_code=400)
            headers = [str(h).strip() if h else '' for h in rows_raw[0]]
            data_rows = [dict(zip(headers, row)) for row in rows_raw[1:] if any(r is not None for r in row)]
        else:
            import csv as _csv
            text = content.decode('utf-8-sig', errors='replace')
            sep = ';' if text.split('\n')[0].count(';') > text.split('\n')[0].count(',') else ','
            reader = _csv.DictReader(io.StringIO(text), delimiter=sep)
            data_rows = list(reader)

        # Kolone za TikTok
        def fcol(row, *candidates):
            for c in candidates:
                for k in row.keys():
                    if k.strip().lower() == c.lower():
                        return row[k]
            return None

        parsed = []
        for row in data_rows:
            campaign = str(fcol(row, 'Campaign name', 'campaign_name') or '').strip()
            if not campaign: continue
            status = str(fcol(row, 'Primary status', 'Status', 'status') or '').strip().lower()
            spend_raw = fcol(row, 'Cost', 'Spend', 'cost', 'spend')
            conv_raw = fcol(row, 'Conversions', 'conversions', 'Cost per conversion')
            cpc_raw = fcol(row, 'CPC (destination)', 'CPC')

            try: spend = float(str(spend_raw).replace(',', '.')) if spend_raw else 0
            except: spend = 0
            try: conversions = float(str(conv_raw).replace(',', '.')) if conv_raw else 0
            except: conversions = 0

            # Izvleci SKU iz campaign name — vzorec: [Maaarket] Smart+ SKU ali SKU: ABPULLER kjerkoli
            import re
            sku = ''
            m = re.search(r'Smart\+\s+([A-Z0-9_]+)', campaign, re.IGNORECASE)
            if m: sku = smart_root(m.group(1).strip())
            else:
                m = re.search(r'SKU:\s*([A-Z0-9_]+)', campaign, re.IGNORECASE)
                if m: sku = smart_root(m.group(1).strip())
            if not sku:
                # Splošno: vzemi zadnji ALL-CAPS blok
                words = campaign.split()
                for w in words:
                    cleaned = re.sub(r'[^A-Z0-9_]', '', w.upper())
                    if len(cleaned) >= 3 and cleaned not in {'ALL','SMART','ADS','TT','TIKTOK','NEW','OLD'}:
                        sku = smart_root(cleaned)
                        break

            parsed.append({
                'campaign': campaign, 'sku': sku.upper(),
                'status': 'active' if 'active' in status else 'inactive',
                'spend': spend, 'conversions': conversions,
                'currency': str(fcol(row, 'Currency', 'currency') or 'EUR').strip(),
                'create_time': str(fcol(row, 'Date Created', 'Campaign create time', 'Create time', 'create_time', 'Created time', 'Ad group create time') or '').strip(),
            })

        if not parsed:
            return JSONResponse({"error": "Ni veljavnih kampanj v datoteki."}, status_code=400)

        # Shrani — accumulate (merge po campaign imenu)
        import csv as _csv2
        existing = {}
        if TIKTOK_ADS_FILE.exists():
            t = TIKTOK_ADS_FILE.read_text(encoding='utf-8-sig', errors='replace')
            for row in _csv2.DictReader(io.StringIO(t)):
                existing[row['campaign']] = row

        for row in parsed:
            existing[row['campaign']] = row

        out = io.StringIO()
        fieldnames = ['campaign', 'sku', 'status', 'spend', 'conversions', 'currency', 'create_time']
        writer = _csv2.DictWriter(out, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(existing.values())
        TIKTOK_ADS_FILE.write_text(out.getvalue(), encoding='utf-8')

        # Meta
        meta = {}
        if TIKTOK_ADS_META.exists():
            try: meta = json.loads(TIKTOK_ADS_META.read_text(encoding='utf-8'))
            except: pass
        uploads = meta.get('uploads', [])
        uploads.append({'filename': fname, 'rows': len(parsed), 'uploaded_at': __import__('datetime').datetime.now().isoformat()})
        TIKTOK_ADS_META.write_text(json.dumps({'uploads': uploads}, ensure_ascii=False), encoding='utf-8')

        return {"ok": True, "rows": len(existing), "new": len(parsed)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/analiza-tiktok-data")
async def analiza_tiktok_data():
    """Vrne TikTok kampanje z zalogo."""
    if not TIKTOK_ADS_FILE.exists():
        return JSONResponse({"error": "Naloži TikTok poročilo."}, status_code=400)
    try:
        import csv as _csv3, io as _io3
        text = TIKTOK_ADS_FILE.read_text(encoding='utf-8-sig', errors='replace')
        rows = list(_csv3.DictReader(_io3.StringIO(text)))

        # Dodaj zalogo iz stock CSV
        stock_map = {}
        stock_root_map = {}  # root → seznam variant
        if STOCK_CSV_FILE.exists():
            st = STOCK_CSV_FILE.read_text(encoding='utf-8-sig', errors='replace')
            sep = ';' if st.split('\n')[0].count(';') > st.split('\n')[0].count(',') else ','
            for r in _csv3.DictReader(_io3.StringIO(st), delimiter=sep):
                sku = (r.get('product_sku') or r.get('sku') or '').strip().upper()
                if sku:
                    entry = {
                        'stock': int(float(r.get('stock', 0) or 0)),
                        'stock30': int(float(r.get('stock30', 0) or 0)),
                        'title': r.get('title', ''),
                    }
                    stock_map[sku] = entry
                    # Dodaj pod koren (BATHFLEX_white → BATHFLEX)
                    root = smart_root(sku).upper()
                    if root not in stock_root_map:
                        stock_root_map[root] = []
                    stock_root_map[root].append(entry)

        def get_stock(sku):
            # 1. Točen match
            if sku in stock_map: return stock_map[sku]
            # 2. Koren match (BATHFLEX → BATHFLEX_white + BATHFLEX_black sešteto)
            root = smart_root(sku).upper()
            variants = stock_root_map.get(sku) or stock_root_map.get(root)
            if variants:
                return {
                    'stock': sum(v['stock'] for v in variants),
                    'stock30': sum(v['stock30'] for v in variants),
                    'title': variants[0]['title'],
                }
            return {}

        items = []
        for r in rows:
            sku = (r.get('sku') or '').strip().upper()
            st_data = get_stock(sku)
            items.append({
                'campaign': r.get('campaign', ''),
                'sku': sku,
                'title': st_data.get('title', ''),
                'status': r.get('status', 'inactive'),
                'spend': float(r.get('spend', 0) or 0),
                'conversions': float(r.get('conversions', 0) or 0),
                'currency': r.get('currency', 'EUR'),
                'stock': st_data.get('stock', 0),
                'stock30': st_data.get('stock30', 0),
                'create_time': r.get('create_time', ''),
            })

        meta = {}
        if TIKTOK_ADS_META.exists():
            try: meta = json.loads(TIKTOK_ADS_META.read_text(encoding='utf-8'))
            except: pass

        return {"items": items, "total": len(items), **meta}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/analiza-tiktok-clear")
async def analiza_tiktok_clear():
    if TIKTOK_ADS_FILE.exists(): TIKTOK_ADS_FILE.unlink()
    if TIKTOK_ADS_META.exists(): TIKTOK_ADS_META.unlink()
    return {"ok": True}

# ─── TIKTOK KREATIVE ──────────────────────────────────────────────────────────
TIKTOK_KR_FILE = DATA_DIR / "tiktok_kreative.csv"
TIKTOK_CL_FILE = DATA_DIR / "tiktok_creative_library.csv"  # Creative Library resolucije

@app.post("/analiza-ttcreative-library-upload")
async def analiza_ttcreative_library_upload(file: UploadFile = File(...)):
    """Naloži TikTok Creative Library CSV z Video name + Video ID + resolucijo."""
    try:
        content = await file.read()
        fname = file.filename or ""
        import io, re as _re

        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows_raw = list(ws.iter_rows(values_only=True))
            headers = [str(h).strip() if h else '' for h in rows_raw[0]]
            data_rows = [dict(zip(headers, row)) for row in rows_raw[1:] if any(r is not None for r in row)]
        else:
            import csv as _csv
            text = content.decode('utf-8-sig', errors='replace')
            sep = '\t' if '\t' in text.split('\n')[0] else (';' if ';' in text.split('\n')[0] else ',')
            data_rows = list(_csv.DictReader(io.StringIO(text), delimiter=sep))

        def fcol(row, *keys):
            for k in keys:
                for rk in row.keys():
                    if rk.strip().lower() == k.lower(): return row[rk]
            return None

        def extract_dims_from_str(s):
            # Podpira: 576 * 1024, 720 X 1280, 720x1280
            m = _re.search(r'(\d{3,5})\s*[*×xX]\s*(\d{3,5})', str(s or ''))
            if m:
                w, h = int(m.group(1)), int(m.group(2))
                if 100 < w < 10000 and 100 < h < 10000:
                    return w, h
            return None, None

        def clean_video_name(v):
            v = str(v or '').strip()
            if not v or v == '-': return ''
            parts = v.split(' ')
            if len(parts) > 1 and _re.match(r'^[a-f0-9]{32}$', parts[0]):
                return ' '.join(parts[1:])
            return v

        parsed = {}
        for row in data_rows:
            # Podpira oba formata: 'Video'/'Creative Name'
            video_raw = fcol(row, 'Creative Name', 'Video', 'video', 'Name', 'Ad name', 'Creative name')
            video_id = str(fcol(row, 'Video ID', 'Video material ID', 'video_id', 'ID') or '').strip()
            res_col = fcol(row, 'Resolution', 'Video resolution', 'Dimension', 'Size')

            if not video_raw or str(video_raw).strip() in ('-', '', 'None'): continue
            video = clean_video_name(video_raw)
            if not video: continue

            # Resolucija: iz stolpca (npr. "576 * 1024") ali iz video imena
            w, h = extract_dims_from_str(res_col) if res_col else (None, None)
            if not w:
                w, h = extract_dims_from_str(video_raw)

            if video not in parsed:
                parsed[video] = {'video': video, 'video_id': video_id, 'w': w or '', 'h': h or ''}
            elif w and not parsed[video]['w']:
                parsed[video]['w'] = w
                parsed[video]['h'] = h

        if not parsed:
            return JSONResponse({"error": "Ni veljavnih videov v datoteki."}, status_code=400)

        # Shrani / merge
        import csv as _csv2
        existing = {}
        if TIKTOK_CL_FILE.exists():
            t = TIKTOK_CL_FILE.read_text(encoding='utf-8-sig', errors='replace')
            for r in _csv2.DictReader(io.StringIO(t)):
                existing[r['video']] = r
        for v, data in parsed.items():
            if v not in existing:
                existing[v] = data
            else:
                # Posodobi resolucijo če je manjka
                if data['w'] and not existing[v].get('w'):
                    existing[v]['w'] = data['w']
                    existing[v]['h'] = data['h']
                if data['video_id'] and not existing[v].get('video_id'):
                    existing[v]['video_id'] = data['video_id']

        out = io.StringIO()
        fn = ['video', 'video_id', 'w', 'h']
        writer = _csv2.DictWriter(out, fieldnames=fn, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(existing.values())
        TIKTOK_CL_FILE.write_text(out.getvalue(), encoding='utf-8')

        with_res = sum(1 for v in existing.values() if v.get('w'))
        return {"ok": True, "videos": len(existing), "with_resolution": with_res}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/analiza-ttcreative-library-data")
async def analiza_ttcreative_library_data():
    if not TIKTOK_CL_FILE.exists():
        return {"items": [], "total": 0}
    import csv as _csv3, io as _io3
    text = TIKTOK_CL_FILE.read_text(encoding='utf-8-sig', errors='replace')
    rows = list(_csv3.DictReader(_io3.StringIO(text)))
    return {"items": rows, "total": len(rows)}

@app.get("/analiza-ttkreative-search")
async def analiza_ttkreative_search(sku: str = ""):
    """Poišče vse videote za SKU — spoji Ads + Creative Library."""
    import csv as _csv4, io as _io4, re as _re4
    sku = sku.strip().upper()
    if not sku:
        return JSONResponse({"error": "Vpišite SKU."}, status_code=400)

    # 1. Ads data — video→SKU mapping
    ads_videos = {}  # video_name → {cost, conversions, status, campaign}
    if TIKTOK_KR_FILE.exists():
        text = TIKTOK_KR_FILE.read_text(encoding='utf-8-sig', errors='replace')
        for r in _csv4.DictReader(_io4.StringIO(text)):
            row_sku = (r.get('sku') or '').strip().upper()
            row_root = smart_root(row_sku).upper()
            search_root = smart_root(sku).upper()
            if row_sku == sku or row_root == search_root or row_sku.startswith(search_root):
                v = r.get('video', '').strip()
                if v and v not in ads_videos:
                    ads_videos[v] = {
                        'video': v,
                        'cost': float(r.get('cost', 0) or 0),
                        'conversions': float(r.get('conversions', 0) or 0),
                        'status': r.get('status', ''),
                        'w': r.get('w') or None,
                        'h': r.get('h') or None,
                    }
                elif v in ads_videos:
                    ads_videos[v]['cost'] += float(r.get('cost', 0) or 0)
                    ads_videos[v]['conversions'] += float(r.get('conversions', 0) or 0)

    # 2. Creative Library — resolucije + vsi videoti za ta SKU
    cl_map = {}  # video_name → {w, h, video_id}
    cl_sku_videos = []  # videoti iz CL ki so vezani na ta SKU (po imenu)
    search_root = smart_root(sku).upper()
    if TIKTOK_CL_FILE.exists():
        text = TIKTOK_CL_FILE.read_text(encoding='utf-8-sig', errors='replace')
        for r in _csv4.DictReader(_io4.StringIO(text)):
            v = r.get('video', '').strip()
            if v:
                cl_map[v] = r
                # Preveri ali video ime vsebuje SKU (npr. ABPULLER19.mp4, PILARAFIT (5).mp4)
                import re as _re5
                v_clean = _re5.sub(r'\.mp4$', '', v, flags=_re5.I)
                v_clean = _re5.sub(r'\s*\(\d+\)$', '', v_clean).strip()  # odstrani (5), (1)
                v_root = _re5.sub(r'\d+$', '', v_clean).upper().strip('_- ')
                if v_root == search_root or v_clean.upper() == search_root or v_clean.upper().startswith(search_root):
                    if v not in ads_videos:
                        cl_sku_videos.append(v)

    # 3. Spoji — Ads videoti + CL-only videoti
    results = []
    for v, data in ads_videos.items():
        cl = cl_map.get(v, {})
        w = data.get('w') or cl.get('w') or None
        h = data.get('h') or cl.get('h') or None
        try: w = int(w) if w else None
        except: w = None
        try: h = int(h) if h else None
        except: h = None
        results.append({
            'video': v,
            'video_id': cl.get('video_id', ''),
            'w': w, 'h': h,
            'cost': round(data['cost'], 2),
            'conversions': int(data['conversions']),
            'status': data['status'],
            'source': 'ads',
            'res_status': _res_status(w, h),
        })

    # Dodaj CL-only videote (brez spend, samo v Creative Library)
    for v in cl_sku_videos:
        cl = cl_map.get(v, {})
        try: w = int(cl.get('w')) if cl.get('w') else None
        except: w = None
        try: h = int(cl.get('h')) if cl.get('h') else None
        except: h = None
        results.append({
            'video': v,
            'video_id': cl.get('video_id', ''),
            'w': w, 'h': h,
            'cost': 0,
            'conversions': 0,
            'status': '',
            'source': 'library',
            'res_status': _res_status(w, h),
        })

    # Sortiraj po cost desc
    results.sort(key=lambda x: x['cost'], reverse=True)
    return {"sku": sku, "items": results, "total": len(results)}

def _res_status(w, h):
    if not w or not h: return "unknown"
    ratio = w / h
    if ratio < 0.7: min_w, min_h = 540, 960
    elif ratio < 1.3: min_w, min_h = 640, 640
    else: min_w, min_h = 960, 540
    if w < min_w or h < min_h: return "bad"
    if min(w, h) < 720: return "warn"
    return "ok"


@app.post("/analiza-ttkreative-upload")
async def analiza_ttkreative_upload(file: UploadFile = File(...)):
    """Naloži TikTok Ad level XLSX z Video + Campaign name stolpci."""
    try:
        content = await file.read()
        fname = file.filename or ""
        import io, re as _re

        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows_raw = list(ws.iter_rows(values_only=True))
            headers = [str(h).strip() if h else '' for h in rows_raw[0]]
            data_rows = [dict(zip(headers, row)) for row in rows_raw[1:] if any(r is not None for r in row)]
        else:
            import csv as _csv, io as _io
            text = content.decode('utf-8-sig', errors='replace')
            sep = ';' if text.split('\n')[0].count(';') > text.split('\n')[0].count(',') else ','
            data_rows = list(_csv.DictReader(_io.StringIO(text), delimiter=sep))

        def fcol(row, *keys):
            for k in keys:
                for rk in row.keys():
                    if rk.strip().lower() == k.lower(): return row[rk]
            return None

        def extract_dims(video):
            m = _re.search(r'(\d{3,4})\s*[xX×]\s*(\d{3,4})', str(video or ''))
            if m:
                w, h = int(m.group(1)), int(m.group(2))
                return w, h
            return None, None

        def clean_video_name(v):
            v = str(v or '').strip()
            if not v or v == '-': return ''
            parts = v.split(' ')
            if len(parts) > 1 and _re.match(r'^[a-f0-9]{32}$', parts[0]):
                return ' '.join(parts[1:])
            return v

        parsed = []
        for row in data_rows:
            campaign = str(fcol(row, 'Campaign name') or '').strip()
            video_raw = fcol(row, 'Video', 'video')
            if not video_raw or str(video_raw).strip() in ('-', '', 'None'): continue

            # SKU iz campaign name
            m = _re.search(r'Smart\+\s+([A-Z0-9_]+)', campaign, _re.I)
            if m: sku = smart_root(m.group(1)).upper()
            else:
                m = _re.search(r'SKU:\s*([A-Z0-9_]+)', campaign, _re.I)
                sku = smart_root(m.group(1)).upper() if m else ''
            if not sku: continue

            status = str(fcol(row, 'Primary status', 'Status') or '').strip()
            try: cost = float(str(fcol(row, 'Cost') or 0).replace(',', '.'))
            except: cost = 0
            try: conversions = float(str(fcol(row, 'Conversions') or 0).replace(',', '.'))
            except: conversions = 0

            # Razdeli po vejici — en oglas ima lahko več videov
            raw_str = str(video_raw).strip()
            video_parts = [v.strip() for v in raw_str.split(',') if v.strip() and v.strip() != '-']

            for vp in video_parts:
                video = clean_video_name(vp)
                if not video: continue
                w, h = extract_dims(vp)
                # Cost/conversions delimo enakomerno med videote
                n = len(video_parts)
                parsed.append({
                    'sku': sku, 'video': video,
                    'w': w or '', 'h': h or '',
                    'cost': round(cost / n, 4), 'conversions': round(conversions / n, 4),
                    'status': status, 'campaign': campaign,
                })

        if not parsed:
            return JSONResponse({"error": "Ni veljavnih videov v datoteki."}, status_code=400)

        # Dedup po sku+video, summiraj cost/conversions
        import csv as _csv2
        existing = {}
        if TIKTOK_KR_FILE.exists():
            t = TIKTOK_KR_FILE.read_text(encoding='utf-8-sig', errors='replace')
            for r in _csv2.DictReader(io.StringIO(t)):
                key = r['sku'] + '||' + r['video']
                existing[key] = r
        for r in parsed:
            key = r['sku'] + '||' + r['video']
            if key in existing:
                try: existing[key]['cost'] = float(existing[key]['cost']) + r['cost']
                except: existing[key]['cost'] = r['cost']
                try: existing[key]['conversions'] = float(existing[key]['conversions']) + r['conversions']
                except: existing[key]['conversions'] = r['conversions']
            else:
                existing[key] = r

        out = io.StringIO()
        fieldnames = ['sku', 'video', 'w', 'h', 'cost', 'conversions', 'status', 'campaign']
        writer = _csv2.DictWriter(out, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(existing.values())
        TIKTOK_KR_FILE.write_text(out.getvalue(), encoding='utf-8')

        skus = len(set(r['sku'] for r in existing.values()))
        return {"ok": True, "videos": len(existing), "skus": skus}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/analiza-ttkreative-data")
async def analiza_ttkreative_data():
    if not TIKTOK_KR_FILE.exists():
        return JSONResponse({"error": "Ni podatkov."}, status_code=400)
    try:
        import csv as _csv3, io as _io3
        text = TIKTOK_KR_FILE.read_text(encoding='utf-8-sig', errors='replace')
        rows = list(_csv3.DictReader(_io3.StringIO(text)))
        items = []
        for r in rows:
            try: w = int(r['w']) if r.get('w') else None
            except: w = None
            try: h = int(r['h']) if r.get('h') else None
            except: h = None
            try: cost = float(r.get('cost', 0) or 0)
            except: cost = 0
            try: conv = float(r.get('conversions', 0) or 0)
            except: conv = 0
            items.append({
                'sku': r.get('sku', ''), 'video': r.get('video', ''),
                'w': w, 'h': h, 'cost': cost, 'conversions': conv,
                'status': r.get('status', ''), 'campaign': r.get('campaign', ''),
            })
        return {"items": items, "total": len(items)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)



# ─── ANALIZA: Meta Ads CSV upload ──────────────────────────────────────────────

META_ADS_FILE = DATA_DIR / "meta_ads_report.csv"
META_ADS_META = DATA_DIR / "meta_ads_meta.json"


# Stop words - skupne besede ki niso SKU
SKU_STOPWORDS = {
    'STOP', 'BIDCAP', 'COSTCAP', 'BID', 'CPA', 'BC', 'EX', 'WP', 'WV', 'EU', 'OFF',
    'LOCAL', 'OUTLET', 'MAAARKET', 'MULTIPLE', 'MAAARKET.HR', 'MAAARKET.RS', 'MAAARKET.SK',
    'NI', 'ZALOGE', 'NOVO', 'CATALOG', 'CATALOGSALE', 'CATEGORY', 'INTERESTED', 'AUTO',
    'ADVANTAGE', 'PROSPECTING', 'REMARKETING', 'CAMPAIGN', 'LOOKALIKE', 'BROAD', 'AAA',
    'DABA', 'COLD', 'KATALOG', 'INFLATED', 'INTERESTS', 'ABO', 'ZIPPLY', 'EASYZO', 'SUBAN',
    'INTEREST', 'TOFU', 'BOFU', 'MOFU', 'ALL'
}


def _load_known_skus():
    """Naloži SKU-je iz CSV zaloge (case-insensitive set)."""
    if not STOCK_CSV_FILE.exists():
        return set()
    try:
        text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        first_line = text.split('\n', 1)[0]
        sep = ';' if first_line.count(';') > first_line.count(',') else ','
        import csv as _csv
        from io import StringIO as _SIO
        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        skus = set()
        for row in reader:
            sku = (row.get('product_sku') or row.get('sku') or '').strip()
            if sku:
                skus.add(sku.upper())
                # Dodaj tudi koren (PLANTUP_white -> PLANTUP, Maaa61lightBrown -> Maaa61)
                koren = smart_root(sku).upper()
                if koren and len(koren) >= 4:
                    skus.add(koren)
        return skus
    except Exception as e:
        print(f"[meta] _load_known_skus err: {e}")
        return set()


def smart_root(s: str) -> str:
    """Pridobi koren SKU-ja:
    1. Razdeli po _-/presledek (PLANTUP_white → PLANTUP, COVERKA_2x3m → COVERKA)
    2. Camel-case prehod (Maaa61lightBrown → Maaa61)
    3. Digit-pred-male-črke (M261red → M261, Maaa6red → Maaa6)
    """
    if not s:
        return s
    base = re.split(r'[_\-\s]', s)[0]
    cut = None
    # Camel-case: lower → Upper
    m = re.search(r'([a-z])([A-Z])', base)
    if m:
        idx = m.start() + 1
        while idx > 1 and base[idx-1].islower():
            idx -= 1
        if idx > 0:
            cut = idx
    # Digit followed by lowercase (M261red, Maaa6red)
    if cut is None:
        m2 = re.search(r'(\d)([a-z])', base)
        if m2:
            cut = m2.start() + 1
    return base[:cut] if cut is not None else base


def extract_skus_from_text(text: str, known_skus: set = None) -> list[str]:
    """Izvleče SKU tokene iz teksta. Če je known_skus dan, vrača samo te.
    
    Match strategija:
    1. Exact match (case-insensitive) v known_skus
    2. Koren match - PLANTUP_white → PLANTUP
    3. Mixed-case dovoljen če je v known_skus (npr. Maaa61, silux74)
    """
    if not text:
        return []
    tokens = []
    
    # Pripravi case-insensitive lookup
    known_upper = set()
    if known_skus is not None:
        known_upper = {s.upper() for s in known_skus}
    
    for raw in text.split():
        # Odstrani emoji in posebne znake na začetku/koncu
        cleaned = re.sub(r'^[^\w]+|[^\w]+$', '', raw)
        if not cleaned or len(cleaned) < 4:
            continue
        # Ne sme vsebovati piko/decimal (filtrira 9.0BID, BID8.5)
        if '.' in cleaned:
            continue
        # Mora vsebovati vsaj 1 črko
        if not any(c.isalpha() for c in cleaned):
            continue
        # Skip stopwords (case-insensitive)
        if cleaned.upper() in SKU_STOPWORDS:
            continue
        # Skip čiste številke + max 1 črka (90D, 7D, 30D, 1ER ipd.)
        if re.match(r'^\d+[A-Z]?$', cleaned, re.IGNORECASE):
            continue
        
        cleaned_upper = cleaned.upper()
        
        # Če imamo known_skus, preverjamo pripadnost
        if known_skus is not None:
            # Exact match (case-insensitive)
            if cleaned_upper in known_upper:
                tokens.append(cleaned)
                continue
            # Koren match (PLANTUP_white → PLANTUP, Maaa61lightBrown → Maaa61)
            koren = smart_root(cleaned)
            if koren.upper() in known_upper:
                tokens.append(cleaned)
                continue
            # Le UPPERCASE besede (>=4 znaki) ki niso v knownh — ignoriraj
        else:
            # Brez known_skus - sprejmemo le UPPERCASE besede (legacy)
            if cleaned == cleaned.upper():
                tokens.append(cleaned)
    return tokens


@app.post("/analiza-meta-upload")
async def analiza_meta_upload(file: UploadFile = File(...), account_name: str = Form("")):
    """Sprejme CSV iz FB Ads Manager export, DODA k obstoječim (accumulate po Campaign name unikatnosti).
    Če CSV nima 'Account name' stolpca (single-account izvoz), uporabi podani account_name."""
    try:
        import csv as _csv
        from io import StringIO as _SIO

        content_bytes = await file.read()
        new_text = content_bytes.decode('utf-8-sig', errors='replace')
        first_line = new_text.split('\n', 1)[0]
        sep = ';' if first_line.count(';') > first_line.count(',') else ','

        # Preberi nove vrstice
        new_reader = _csv.DictReader(_SIO(new_text), delimiter=sep)
        new_rows = [r for r in new_reader if r.get('Campaign name', '').strip()]
        if not new_rows:
            return JSONResponse({"error": "CSV nima veljavnih vrstic."}, status_code=400)

        # Če CSV nima Account name (ali je prazen) IN je podan account_name → napolni
        acc_override = (account_name or "").strip()
        if acc_override:
            for r in new_rows:
                if not (r.get('Account name') or '').strip():
                    r['Account name'] = acc_override

        headers = list(new_rows[0].keys())
        if 'Account name' not in headers:
            headers.append('Account name')

        # Preberi obstoječe vrstice (če obstajajo)
        existing_rows = []
        if META_ADS_FILE.exists():
            try:
                ex_text = META_ADS_FILE.read_text(encoding='utf-8-sig', errors='replace')
                ex_sep = ';' if ex_text.split('\n',1)[0].count(';') > ex_text.split('\n',1)[0].count(',') else ','
                ex_reader = _csv.DictReader(_SIO(ex_text), delimiter=ex_sep)
                existing_rows = [r for r in ex_reader if r.get('Campaign name', '').strip()]
            except: pass

        # Deduplikacija: ključ = Campaign name + Account name + Reporting starts
        def row_key(r):
            return (r.get('Campaign name','').strip(), r.get('Account name','').strip(), r.get('Reporting starts','').strip())

        existing_keys = {row_key(r) for r in existing_rows}
        added = [r for r in new_rows if row_key(r) not in existing_keys]
        merged = existing_rows + added

        # Shrani merged CSV
        out = _SIO()
        # Združi vse headerje (union)
        all_headers = list(dict.fromkeys(headers + [h for h in (existing_rows[0].keys() if existing_rows else []) if h not in headers]))
        writer = _csv.DictWriter(out, fieldnames=all_headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(merged)
        META_ADS_FILE.write_text(out.getvalue(), encoding='utf-8')

        # Zapiši meta (seznam vseh naloženih fileov)
        meta = {}
        if META_ADS_META.exists():
            try: meta = json.loads(META_ADS_META.read_text(encoding='utf-8'))
            except: pass
        uploads = meta.get('uploads', [])
        uploads.append({
            "filename": file.filename,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "rows_added": len(added),
            "rows_total": len(merged),
        })
        meta = {
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "filename": file.filename,
            "rows": len(merged),
            "rows_added": len(added),
            "uploads": uploads[-10:],  # ohrani zadnjih 10
        }
        META_ADS_META.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')

        # Zberi accounte iz merged podatkov
        accounts = sorted(set(r.get('Account name','').strip() for r in merged if r.get('Account name','').strip()))

        return {
            "status": "ok",
            "rows_added": len(added),
            "rows_total": len(merged),
            "uploaded_at": meta["uploaded_at"],
            "filename": file.filename,
            "accounts": accounts,
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/analiza-meta-debug")
async def analiza_meta_debug():
    """Diagnostika: kaj backend vidi v Meta podatkih (accounti, SKU match, vzorci)."""
    if not META_ADS_FILE.exists():
        return {"loaded": False, "msg": "META_ADS_FILE ne obstaja"}
    try:
        import csv as _csv
        from io import StringIO as _SIO
        text = META_ADS_FILE.read_text(encoding="utf-8-sig", errors="replace")
        sep = ';' if text.split('\n',1)[0].count(';') > text.split('\n',1)[0].count(',') else ','
        rows = [r for r in _csv.DictReader(_SIO(text), delimiter=sep) if r.get('Campaign name','').strip()]

        known = _load_known_skus()

        # Account distribucija (surovo)
        from collections import Counter
        acc_counter = Counter((r.get('Account name') or '—').strip() for r in rows)

        # Za vsak account: koliko kampanj dobi SKU match
        acc_match = {}
        samples = {}
        for r in rows:
            acc = (r.get('Account name') or '—').strip()
            name = r.get('Campaign name','').strip()
            skus = extract_skus_from_text(name, known if known else None)
            acc_match.setdefault(acc, {"total": 0, "matched": 0})
            acc_match[acc]["total"] += 1
            if skus:
                acc_match[acc]["matched"] += 1
            elif acc not in samples:
                # shrani primer kampanje brez matcha
                samples[acc] = {"campaign": name, "extracted_no_filter": extract_skus_from_text(name, None)}

        # Ali so headerji v redu?
        headers = list(rows[0].keys()) if rows else []
        has_account_col = 'Account name' in headers

        return {
            "loaded": True,
            "total_rows": len(rows),
            "headers": headers,
            "has_account_name_column": has_account_col,
            "known_skus_count": len(known),
            "accounts_raw": dict(acc_counter),
            "account_match": acc_match,
            "no_match_samples": samples,
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/analiza-meta-fix-account")
async def analiza_meta_fix_account(data: dict):
    """Popravi obstoječe vrstice brez Account name (—) → dodeli izbran account.
    Uporabno če si naložil single-account CSV brez da bi vpisal ime accounta."""
    new_account = (data.get("account_name") or "").strip()
    if not new_account:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Manjka account_name."}, status_code=400)
    if not META_ADS_FILE.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Ni naloženih Meta podatkov."}, status_code=400)
    try:
        import csv as _csv
        from io import StringIO as _SIO
        text = META_ADS_FILE.read_text(encoding="utf-8-sig", errors="replace")
        sep = ';' if text.split('\n',1)[0].count(';') > text.split('\n',1)[0].count(',') else ','
        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        rows = [r for r in reader if r.get('Campaign name','').strip()]

        headers = list(rows[0].keys()) if rows else []
        if 'Account name' not in headers:
            headers.append('Account name')

        fixed = 0
        for r in rows:
            if not (r.get('Account name') or '').strip():
                r['Account name'] = new_account
                fixed += 1

        out = _SIO()
        writer = _csv.DictWriter(out, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)
        META_ADS_FILE.write_text(out.getvalue(), encoding='utf-8')

        return {"status": "ok", "fixed_rows": fixed, "account": new_account, "total_rows": len(rows)}
    except Exception as e:
        import traceback; traceback.print_exc()
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/analiza-meta-rename-account")
async def analiza_meta_rename_account(data: dict):
    """Preimenuje account v vseh vrsticah: from_account → to_account.
    Uporabno za odpravo podvojenih accountov (npr. Colibrishop → Colibrishop_EU)."""
    from_acc = (data.get("from_account") or "").strip()
    to_acc = (data.get("to_account") or "").strip()
    if not from_acc or not to_acc:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Manjka from_account ali to_account."}, status_code=400)
    if not META_ADS_FILE.exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Ni naloženih Meta podatkov."}, status_code=400)
    try:
        import csv as _csv
        from io import StringIO as _SIO
        text = META_ADS_FILE.read_text(encoding="utf-8-sig", errors="replace")
        sep = ';' if text.split('\n',1)[0].count(';') > text.split('\n',1)[0].count(',') else ','
        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        rows = [r for r in reader if r.get('Campaign name','').strip()]
        headers = list(rows[0].keys()) if rows else []
        if 'Account name' not in headers:
            headers.append('Account name')

        renamed = 0
        for r in rows:
            if (r.get('Account name') or '').strip() == from_acc:
                r['Account name'] = to_acc
                renamed += 1

        out = _SIO()
        writer = _csv.DictWriter(out, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)
        META_ADS_FILE.write_text(out.getvalue(), encoding='utf-8')

        return {"status": "ok", "renamed_rows": renamed, "from": from_acc, "to": to_acc}
    except Exception as e:
        import traceback; traceback.print_exc()
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/analiza-meta-clear")
async def analiza_meta_clear():
    """Počisti vse naložene Meta Ads CSV podatke."""
    try:
        if META_ADS_FILE.exists(): META_ADS_FILE.unlink()
        if META_ADS_META.exists(): META_ADS_META.unlink()
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)



@app.get("/analiza-meta-data")
async def analiza_meta_data():
    """Vrne agregirane podatke iz Meta Ads CSV: SKU → metrike po accountu."""
    if not META_ADS_FILE.exists():
        return {"loaded": False}

    try:
        text = META_ADS_FILE.read_text(encoding="utf-8-sig", errors="replace")
        first_line = text.split('\n', 1)[0]
        sep = ';' if first_line.count(';') > first_line.count(',') else ','

        import csv as _csv
        from io import StringIO as _SIO
        reader = _csv.DictReader(_SIO(text), delimiter=sep)

        # Agregirajmo po SKU (in kasneje v JS po accountu)
        sku_data = {}  # sku → {accounts: {acc: {spend, purchases, etc}}, campaigns: [...]}

        def _f(v):
            try:
                return float(str(v or '0').replace(',', '.'))
            except:
                return 0.0

        def _i(v):
            try:
                return int(float(str(v or '0').replace(',', '.')))
            except:
                return 0

        # Naloži znane SKU-je iz zaloge (za boljši filter)
        known_skus = _load_known_skus()

        for row in reader:
            campaign_name = (row.get('Campaign name') or '').strip()
            if not campaign_name:
                continue

            account = (row.get('Account name') or '').strip() or '—'
            spend = _f(row.get('Amount spent (EUR)'))
            purchases = _i(row.get('Purchases'))
            cpa = _f(row.get('Cost per purchase'))
            cpc = _f(row.get('CPC (cost per link click)'))
            ctr = _f(row.get('CTR (link click-through rate)'))
            atc = _i(row.get('Adds to cart'))
            freq = _f(row.get('Frequency'))
            # Status SAMO iz FB Campaign Delivery kolone (edina resnica)
            # Ime kampanje (@STOP, ⛔, OFF) se IGNORIRA — to so naši interni oznaki
            delivery = (row.get('Campaign Delivery') or '').strip().lower()
            if delivery == 'inactive':
                is_stopped = True
            elif delivery == 'active':
                is_stopped = False
            else:
                # Stolpec manjka ali neznana vrednost → privzeto aktivna
                # (raje napačno aktivna kot napačno pavzirana — ker FB sam ne ve)
                is_stopped = False

            # Izvleci SKU-je iz imena (filtrirano po znanih SKU iz zaloge)
            skus = extract_skus_from_text(campaign_name, known_skus if known_skus else None)
            skus = list(dict.fromkeys(skus))

            # FALLBACK: če kampanja ne najde SKU v glavni zalogi (npr. account z ločeno
            # zalogo kot Colibrishop), izvleci kandidat brez whitelist filtra — da se
            # kampanja vseeno prikaže in account dobi kljukico.
            if not skus:
                fb = extract_skus_from_text(campaign_name, None)  # UPPERCASE kandidati
                if not fb:
                    # mixed-case brand (SensaTouch, Hairrevive) → prvi smiseln token
                    for raw in campaign_name.split():
                        cl = re.sub(r'^[^\w]+|[^\w]+$', '', raw)
                        if len(cl) >= 4 and any(c.isalpha() for c in cl) \
                           and cl.upper() not in SKU_STOPWORDS \
                           and not re.match(r'^\d+[A-Z]?$', cl, re.IGNORECASE) \
                           and '.' not in cl:
                            fb = [cl]; break
                skus = list(dict.fromkeys(fb))

            for sku in skus:
                if sku not in sku_data:
                    sku_data[sku] = {
                        "sku": sku,
                        "campaigns": [],
                        "accounts": {},
                        "total_spend": 0,
                        "total_purchases": 0,
                        "total_atc": 0,
                        "total_clicks_value": 0,  # za uteženo CPC
                        "campaign_count": 0,
                        "stopped_count": 0,
                    }
                d = sku_data[sku]
                d["campaigns"].append({
                    "name": campaign_name,
                    "account": account,
                    "spend": spend,
                    "purchases": purchases,
                    "cpa": cpa,
                    "atc": atc,
                    "stopped": is_stopped,
                })
                if account not in d["accounts"]:
                    d["accounts"][account] = {"spend": 0, "purchases": 0, "campaigns": 0, "active": 0, "paused": 0}
                d["accounts"][account]["spend"] += spend
                d["accounts"][account]["purchases"] += purchases
                d["accounts"][account]["campaigns"] += 1
                if is_stopped:
                    d["accounts"][account]["paused"] += 1
                else:
                    d["accounts"][account]["active"] += 1
                d["total_spend"] += spend
                d["total_purchases"] += purchases
                d["total_atc"] += atc
                d["campaign_count"] += 1
                if is_stopped:
                    d["stopped_count"] += 1

        # Pripravi flat seznam za frontend
        items = []
        for sku, d in sku_data.items():
            avg_cpa = (d["total_spend"] / d["total_purchases"]) if d["total_purchases"] > 0 else None
            # Izračunaj per-account status (active = vsaj 1 aktivna, paused = vse pavzirane)
            accounts_with_status = []
            for k, v in d["accounts"].items():
                acc = {"name": k, **v}
                if v.get("active", 0) > 0:
                    acc["status"] = "active"
                elif v.get("paused", 0) > 0:
                    acc["status"] = "paused"
                else:
                    acc["status"] = "none"
                accounts_with_status.append(acc)
            items.append({
                "sku": sku,
                "total_spend": round(d["total_spend"], 2),
                "total_purchases": d["total_purchases"],
                "total_atc": d["total_atc"],
                "avg_cpa": round(avg_cpa, 2) if avg_cpa is not None else None,
                "campaign_count": d["campaign_count"],
                "stopped_count": d["stopped_count"],
                "active_count": d["campaign_count"] - d["stopped_count"],
                "accounts": accounts_with_status,
            })

        # Zberi vse accounte dinamično iz podatkov.
        # POMEMBNO: ne samo iz sku_data (kampanje s prepoznanim SKU), ampak iz VSEH
        # vrstic CSV — sicer accounti katerih SKU-ji niso v glavni zalogi (npr. Colibrishop
        # z ločeno zalogo) ne dobijo kljukice, čeprav imajo kampanje.
        all_accounts_set = set()
        for d in sku_data.values():
            for acc_name in d["accounts"].keys():
                if acc_name and acc_name != '—':
                    all_accounts_set.add(acc_name)
        # Dodaj še accounte iz surovih vrstic (tudi brez SKU match)
        try:
            text2 = META_ADS_FILE.read_text(encoding="utf-8-sig", errors="replace")
            sep2 = ';' if text2.split('\n',1)[0].count(';') > text2.split('\n',1)[0].count(',') else ','
            import csv as _csv2
            from io import StringIO as _SIO2
            for r in _csv2.DictReader(_SIO2(text2), delimiter=sep2):
                an = (r.get('Account name') or '').strip()
                if an and an != '—':
                    all_accounts_set.add(an)
        except Exception as _e:
            print(f"[meta] all_accounts raw scan err: {_e}")
        all_accounts = sorted(all_accounts_set)

        meta = {}
        if META_ADS_META.exists():
            try:
                meta = json.loads(META_ADS_META.read_text(encoding="utf-8"))
            except: pass

        return {
            "loaded": True,
            "items": items,
            "total_skus": len(items),
            "accounts": all_accounts,  # dynamic lista za frontend checkboxe
            **{k: v for k, v in meta.items() if k != 'accounts'},
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── ANALIZA: Obrat 14 dni ─────────────────────────────────────────────────────

OBRAT14_FILE = DATA_DIR / "obrat_14dni.txt"
OBRAT14_META = DATA_DIR / "obrat_14dni_meta.json"

# Whitelist accounts za Obrat 14dni view
TARGET_ACCOUNTS = ['Maaarket X', 'Maaarket ALL', 'Maaarket ALL2', 'Maaarket ALL3 + RS', 'Zipply.', 'si_SUBAN_Maaarket SK', 'Maaarket PL/RO', 'Maaarket HR', 'si_Suban_Maaarket HR', 'Easyzo', 'Thundershop ALL HU', 'ThunderShop HR', 'ThunderShop RS', 'Colibrishop_EU']
# Ko dobiš nova accounta, dodaj ju sem IN v AD_ACCOUNTS_CONFIG v index.html


@app.post("/analiza-obrat14-upload")
async def analiza_obrat14_upload(file: UploadFile = File(...)):
    """Sprejme TXT/TSV iz top obratov, shrani."""
    try:
        content_bytes = await file.read()
        OBRAT14_FILE.write_bytes(content_bytes)

        text = content_bytes.decode('utf-8-sig', errors='replace')
        lines = [l.strip() for l in text.splitlines() if l.strip()]

        meta = {
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "filename": file.filename,
            "size": len(content_bytes),
            "rows": max(0, len(lines) - 1),
        }
        OBRAT14_META.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"status": "ok", "rows": meta["rows"], "uploaded_at": meta["uploaded_at"], "filename": file.filename}
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/analiza-obrat14-data")
async def analiza_obrat14_data():
    """Vrne obrat 14dni podatke + match z FB Ads (po accountu, samo whitelist)."""
    if not OBRAT14_FILE.exists():
        return {"loaded": False}

    try:
        # Naloži obrat14
        text = OBRAT14_FILE.read_text(encoding="utf-8-sig", errors="replace")
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if len(lines) < 2:
            return {"loaded": False, "error": "Prazna datoteka."}

        # Detect separator (tab ali ;)
        first = lines[0]
        if '\t' in first:
            sep = '\t'
        elif ';' in first:
            sep = ';'
        else:
            sep = ','

        items = []
        # Skip header
        for line in lines[1:]:
            parts = line.split(sep)
            if len(parts) < 3:
                continue
            sku = parts[0].strip()
            naziv = parts[1].strip()
            try:
                kolicina = int(float(parts[2].strip().replace(',', '.')))
            except:
                kolicina = 0
            if not sku:
                continue
            items.append({
                "sku": sku,
                "naziv": naziv,
                "kolicina": kolicina,
            })

        # Match z FB Ads — uporabljaj že shranjen META_ADS_FILE
        sku_ads_map = {}  # sku.upper() → {account: {spend, purchases}, ...}
        if META_ADS_FILE.exists():
            try:
                fb_text = META_ADS_FILE.read_text(encoding="utf-8-sig", errors="replace")
                fb_first = fb_text.split('\n', 1)[0]
                fb_sep = ';' if fb_first.count(';') > fb_first.count(',') else ','
                import csv as _csv
                from io import StringIO as _SIO
                reader = _csv.DictReader(_SIO(fb_text), delimiter=fb_sep)

                # Pripravi seznam znanih SKU iz obrat14 (za extract_skus filter)
                # Vključimo VSE variante — uppercase (za standardni match) + originalne + korene
                known_skus_obrat = set()
                for it in items:
                    s = it["sku"]
                    known_skus_obrat.add(s)
                    known_skus_obrat.add(s.upper())
                    # Koren (PLANTUP_white -> PLANTUP, Maaa61lightBrown -> Maaa61, COVERKA_2x3m -> COVERKA)
                    koren = smart_root(s)
                    if koren and len(koren) >= 4:
                        known_skus_obrat.add(koren)
                        known_skus_obrat.add(koren.upper())

                def _f(v):
                    try: return float(str(v or '0').replace(',', '.'))
                    except: return 0.0
                def _i(v):
                    try: return int(float(str(v or '0').replace(',', '.')))
                    except: return 0

                for row in reader:
                    cname = (row.get('Campaign name') or '').strip()
                    if not cname:
                        continue
                    account = (row.get('Account name') or '').strip() or '—'
                    spend = _f(row.get('Amount spent (EUR)'))
                    purchases = _i(row.get('Purchases'))
                    
                    # Status SAMO iz FB Campaign Delivery kolone (edina resnica)
                    # Ime kampanje (@STOP, ⛔, OFF) se IGNORIRA — to so naši interni oznaki
                    delivery = (row.get('Campaign Delivery') or '').strip().lower()
                    if delivery == 'inactive':
                        is_active_campaign = False
                    elif delivery == 'active':
                        is_active_campaign = True
                    else:
                        # Stolpec manjka ali neznana vrednost → privzeto aktivna
                        is_active_campaign = True

                    # Izvleci SKU iz imena
                    skus = extract_skus_from_text(cname, known_skus_obrat)
                    skus = list(dict.fromkeys(skus))

                    # Za vsak SKU token iz kampanje, najdi VSE matching izdelke v 14dni
                    for sku_token in skus:
                        sku_upper = sku_token.upper()
                        token_koren = smart_root(sku_upper)
                        target_skus = []
                        
                        # 1. Exact (case-insensitive)
                        for it in items:
                            if it["sku"].upper() == sku_upper:
                                target_skus = [it["sku"]]
                                break
                        
                        if not target_skus:
                            # 2. Koren match — vsi izdelki s tem korenom
                            candidates = []
                            for it in items:
                                item_koren = smart_root(it["sku"]).upper()
                                if item_koren == token_koren or item_koren == sku_upper or token_koren == it["sku"].upper():
                                    candidates.append(it["sku"])
                            
                            if len(candidates) == 0:
                                continue
                            elif len(candidates) == 1:
                                target_skus = candidates
                            else:
                                # 3. Več zadetkov: če je sku_token preprosto koren (npr. "Maaa61") → vse variante
                                if sku_upper == token_koren:
                                    target_skus = candidates
                                else:
                                    # Fuzzy match — najdi najbolj podobnega
                                    from difflib import SequenceMatcher
                                    best, best_ratio = None, 0
                                    for c in candidates:
                                        ratio = SequenceMatcher(None, sku_upper, c.upper()).ratio()
                                        if ratio > best_ratio:
                                            best_ratio = ratio
                                            best = c
                                    if best and best_ratio >= 0.6:
                                        target_skus = [best]
                                    else:
                                        target_skus = candidates  # fallback: vse

                        # Zabeleži za vse target SKU-je
                        for target_sku in target_skus:
                            if target_sku not in sku_ads_map:
                                sku_ads_map[target_sku] = {}
                            if account not in sku_ads_map[target_sku]:
                                sku_ads_map[target_sku][account] = {"spend": 0, "purchases": 0, "campaigns": 0, "active": 0, "paused": 0}
                            # Razdelimo spend/purchases enakomerno če je več target SKU
                            split = max(1, len(target_skus))
                            sku_ads_map[target_sku][account]["spend"] += spend / split
                            sku_ads_map[target_sku][account]["purchases"] += purchases / split
                            sku_ads_map[target_sku][account]["campaigns"] += 1
                            if is_active_campaign:
                                sku_ads_map[target_sku][account]["active"] += 1
                            else:
                                sku_ads_map[target_sku][account]["paused"] += 1
            except Exception as e:
                print(f"[obrat14] FB match err: {e}")

        # Agregiraj — vsakemu accountu dodaj status field (active/paused/none)
        for it in items:
            accounts_data = sku_ads_map.get(it["sku"], {})
            for acc_name, acc_data in accounts_data.items():
                if acc_data.get("active", 0) > 0:
                    acc_data["status"] = "active"
                elif acc_data.get("paused", 0) > 0:
                    acc_data["status"] = "paused"
                else:
                    acc_data["status"] = "none"
            it["accounts"] = accounts_data
            it["has_ads"] = len(accounts_data) > 0
            # Skupni status
            statuses = [a.get("status") for a in accounts_data.values()]
            if "active" in statuses:
                it["overall_status"] = "active"
            elif "paused" in statuses:
                it["overall_status"] = "paused"
            else:
                it["overall_status"] = "none"

        meta = {}
        if OBRAT14_META.exists():
            try:
                meta = json.loads(OBRAT14_META.read_text(encoding="utf-8"))
            except: pass

        return {
            "loaded": True,
            "items": items,
            "total": len(items),
            "target_accounts": TARGET_ACCOUNTS,
            **meta,
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)

# ═══════════════════════════════════════════════════════════════
# HS+ UVOZ — naročilnice z združevanjem SKU + history
# ═══════════════════════════════════════════════════════════════
HSUVOZ_DIR = DATA_DIR / "hsuvoz_history"
HSUVOZ_DIR.mkdir(exist_ok=True, parents=True)
HSUVOZ_CURRENT = DATA_DIR / "hsuvoz_current.json"


# ═══════════════════════════════════════════════════════════════
# HS+ KATALOG — živa povezava na HS-plus XML izvoz
# Kredenciali so v Render env: HSPLUS_XML_URL, HSPLUS_USER, HSPLUS_PASS
# ═══════════════════════════════════════════════════════════════
HSPLUS_CATALOG_CACHE = DATA_DIR / "hsplus_catalog_cache.json"
HSPLUS_CACHE_TTL = 3600  # 1 ura
HSPLUS_SNAPSHOT = DATA_DIR / "hsplus_stock_snapshot.json"   # zadnji posnetek zaloge (sku→stock) za diff
HSPLUS_DIFF = DATA_DIR / "hsplus_stock_diff.json"           # zadnja sprememba zaloge med uploadi

def _hsplus_clean_xml(xml_bytes):
    """Očisti pogoste napake v XML (surovi & in neveljavni kontrolni znaki),
    da ga parser prebere tudi če HS+ vrne ne-povsem-veljaven XML."""
    import re as _re
    if isinstance(xml_bytes, bytes):
        text = xml_bytes.decode("utf-8", errors="replace")
    else:
        text = xml_bytes
    # surovi & ki ni del entitete (&amp; &lt; &gt; &quot; &apos; &#123;) → &amp;
    text = _re.sub(r'&(?!(?:amp|lt|gt|quot|apos|#\d+|#x[0-9a-fA-F]+);)', '&amp;', text)
    # odstrani neveljavne XML kontrolne znake (razen \t \n \r)
    text = _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    return text

_HSPLUS_COLORS = {'blue','pink','black','white','red','green','yellow','grey','gray','beige','brown',
    'purple','orange','navy','gold','silver','cream','khaki','light','dark','bg','bl','wh','bk','gr','rd',
    'modra','roza','crna','bela','rdeca','zelena','rumena','siva','bez','rjava','short','long','kratke','dolge'}
import re as _re_hsplus
_HSPLUS_SIZE_RE = _re_hsplus.compile(r'^(XS|S|M|L|XL|XXL|XXXL|S/M|L/XL|M/L|XL/XXL|\d+/\d+|\d+(CM|ML|L|KG|G)?)$', _re_hsplus.I)

def _hsplus_root_key(name):
    """Koren naziva = prva beseda (tvoja logika: prvi naziv je koren, ostalo variacije).
    Preskoči vodilni promocijski prefiks tipa '(SP) ' / '(UP) '.
    Npr. 'TOPKNER 90x200' → 'TOPKNER', 'STRAPIES Beige 2XL' → 'STRAPIES',
    'BODY-FIT BL S/M' → 'BODY-FIT', '(SP) CORALCLOTH' → 'CORALCLOTH'."""
    if not name:
        return ""
    s = _re_hsplus.sub(r'^\([^)]*\)\s*', '', name).strip()  # odstrani vodilni (XX) prefiks
    m = _re_hsplus.split(r'[\s(]', s, 1)  # razdeli na prvem presledku ali oklepaju
    root = m[0].strip() if m and m[0].strip() else s
    return root

def _hsplus_parse_xml(xml_bytes):
    """Parsira HS+ catalog XML → seznam izdelkov. Odporen na manjše napake v XML."""
    import xml.etree.ElementTree as _ET
    try:
        root = _ET.fromstring(xml_bytes if isinstance(xml_bytes, (bytes, str)) else xml_bytes)
    except _ET.ParseError:
        # poskusi s čiščenjem
        cleaned = _hsplus_clean_xml(xml_bytes)
        root = _ET.fromstring(cleaned)
    out = []
    for p in root.findall(".//product"):
        imgs = [im.text for im in p.findall(".//image") if im.text]
        def _num(tag, cast, default=0):
            v = (p.findtext(tag) or "").strip()
            try:
                return cast(v) if v else default
            except (ValueError, TypeError):
                return default
        nm = (p.findtext("name") or "").strip()
        raw_sku = (p.findtext("sku") or "").strip()
        # HS+ ima polji obrnjeni: <sku> dejansko vsebuje EAN (samo številke),
        # prava koda izdelka (npr. BRONZIE) pa je v <name>.
        # Zato: SKU = name (koda), EAN = sku polje. Če name prazen, fallback na raw_sku.
        is_ean = bool(_re_hsplus.fullmatch(r'\d{8,14}', raw_sku))
        if is_ean and nm:
            sku_val = nm
            ean_val = raw_sku
        else:
            sku_val = raw_sku
            ean_val = ""
        out.append({
            "sku": sku_val,
            "ean": ean_val,
            "name": nm,
            "description": (p.findtext("description") or "").strip(),
            "category": (p.findtext("category") or "").strip(),
            "stock": _num("stock", int, 0),
            "price": _num("price", float, 0.0),
            "image": imgs[0] if imgs else "",
            "images": imgs,
            "root": _hsplus_root_key(nm),
        })
    return out

async def _hsplus_fetch_core(force=False, cache_only=False):
    """Potegne HS+ katalog (kliče ga endpoint + dnevni scheduler).
    Vrne dict z ok/products/fetched_at. force=True preskoči cache TTL.
    cache_only=True vrne samo iz cache (brez živega potega — za odpiranje taba)."""
    import json as _json, time as _time
    # cache
    if not force and HSPLUS_CATALOG_CACHE.exists():
        try:
            cached = _json.loads(HSPLUS_CATALOG_CACHE.read_text(encoding="utf-8"))
            if cache_only or (_time.time() - cached.get("fetched_ts", 0) < HSPLUS_CACHE_TTL):
                return {"ok": True, "cached": True, "fetched_at": cached.get("fetched_at"),
                        "count": len(cached.get("products", [])), "products": cached.get("products", [])}
        except Exception:
            pass
    if cache_only:
        # ni cache — ne sproži živega potega, vrni prazno (uporabnik naj naloži ročno)
        return {"ok": True, "cached": True, "fetched_at": None, "count": 0, "products": []}
    # živi poteg
    url = os.environ.get("HSPLUS_XML_URL", "https://hsb2b.hs-plus.com/catalog/export?format=xml")
    user = os.environ.get("HSPLUS_USER", "")
    pw = os.environ.get("HSPLUS_PASS", "")
    if not user or not pw:
        return {"ok": False, "error": "Manjkata HSPLUS_USER / HSPLUS_PASS (Render okoljski spremenljivki)."}
    try:
        import httpx as _httpx
        async with _httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            r = await client.get(url, auth=(user, pw))
        if r.status_code != 200:
            return {"ok": False, "error": f"HS+ vrnil status {r.status_code}", "status": r.status_code}
        products = _hsplus_parse_xml(r.content)
        fetched_at = datetime.now().strftime("%d.%m.%Y %H:%M")
        try:
            HSPLUS_CATALOG_CACHE.write_text(_json.dumps({
                "fetched_ts": _time.time(), "fetched_at": fetched_at, "products": products
            }, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass
        return {"ok": True, "cached": False, "fetched_at": fetched_at,
                "count": len(products), "products": products}
    except Exception as e:
        return {"ok": False, "error": f"Napaka pri povezavi na HS+: {e}"}


@app.get("/hsplus-catalog")
async def hsplus_catalog(refresh: str = "0", cache: str = "0"):
    """Vrne HS+ katalog. refresh=1 živi poteg; cache=1 samo iz cache (brez živega potega)."""
    return await _hsplus_fetch_core(force=(refresh == "1"), cache_only=(cache == "1"))

@app.get("/hsplus-debug")
async def hsplus_debug():
    """Diagnostika: pokaže kaj HS+ dejansko vrne (status, content-type, začetek vsebine).
    Pomaga ugotoviti, ali prijava deluje in ali je odgovor res XML."""
    url = os.environ.get("HSPLUS_XML_URL", "https://hsb2b.hs-plus.com/catalog/export?format=xml")
    user = os.environ.get("HSPLUS_USER", "")
    pw = os.environ.get("HSPLUS_PASS", "")
    if not user or not pw:
        return {"ok": False, "error": "Manjkata HSPLUS_USER / HSPLUS_PASS."}
    try:
        import httpx as _httpx
        async with _httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            r = await client.get(url, auth=(user, pw))
        body = r.text
        # poskusi najti vrstico 19 (kjer je bila napaka)
        lines = body.split("\n")
        around_19 = "\n".join(lines[15:25]) if len(lines) >= 20 else body[:800]
        # ali je videti kot XML?
        looks_xml = body.lstrip()[:50].startswith("<?xml") or "<product>" in body[:2000]
        looks_html = "<html" in body[:500].lower() or "<!doctype html" in body[:500].lower()
        return {
            "ok": True,
            "status": r.status_code,
            "content_type": r.headers.get("content-type", ""),
            "dolzina": len(body),
            "izgleda_kot_xml": looks_xml,
            "izgleda_kot_html_login": looks_html,
            "prvih_500_znakov": body[:500],
            "okolica_vrstice_19": around_19,
        }
    except Exception as e:
        return {"ok": False, "error": f"{e}"}

@app.get("/hsplus-catalog-stats")
async def hsplus_catalog_stats():
    """Hitri povzetek kataloga (kategorije, skupno) iz cache, brez velikega prenosa."""
    import json as _json
    if not HSPLUS_CATALOG_CACHE.exists():
        return {"ok": False, "error": "Katalog še ni potegnjen."}
    try:
        cached = _json.loads(HSPLUS_CATALOG_CACHE.read_text(encoding="utf-8"))
        products = cached.get("products", [])
        from collections import Counter as _Counter
        cats = _Counter(p.get("category") or "(brez)" for p in products)
        return {"ok": True, "fetched_at": cached.get("fetched_at"), "count": len(products),
                "categories": dict(sorted(cats.items())),
                "zero_stock": sum(1 for p in products if p.get("stock", 0) == 0)}
    except Exception as e:
        return {"ok": False, "error": str(e)}

def _hsplus_compute_diff(products):
    """Primerja novo zalogo z zadnjim snapshotom (sku→stock) in shrani diff.
    Vrne povzetek sprememb. Nato posodobi snapshot na trenutno stanje."""
    import json as _json, time as _time
    # naloži prejšnji snapshot
    prev = {}
    prev_at = None
    if HSPLUS_SNAPSHOT.exists():
        try:
            snap = _json.loads(HSPLUS_SNAPSHOT.read_text(encoding="utf-8"))
            prev = snap.get("stock", {})
            prev_at = snap.get("fetched_at")
        except Exception:
            prev = {}
    # trenutno stanje (sku → {stock, name, price, category, image})
    cur = {}
    for p in products:
        sku = p.get("sku")
        if sku:
            cur[sku] = p
    # VAROVALO: stari snapshot je morda še v EAN formatu (pred popravkom sku=koda).
    # Če se ključi sploh ne ujemajo z novimi (presek prazen), tretiraj kot prvi zagon
    # — sicer bi bil cel katalog lažno označen kot "novo".
    if prev and cur:
        overlap = set(prev.keys()) & set(cur.keys())
        if not overlap:
            prev = {}  # nezdružljiv star snapshot → brez lažnega diffa
    # izračunaj spremembe (samo če imamo prejšnji snapshot)
    changes = []
    if prev:
        for sku, p in cur.items():
            old = prev.get(sku)
            new_stock = p.get("stock", 0)
            new_price = p.get("price", 0)
            if old is None:
                # nov izdelek (prej ga ni bilo)
                changes.append({"sku": sku, "name": p.get("name",""), "category": p.get("category",""),
                                "image": p.get("image",""), "price": new_price,
                                "old": None, "new": new_stock, "delta": new_stock, "is_new": True,
                                "old_price": None, "new_price": new_price, "price_delta": 0})
            else:
                old_stock = old if isinstance(old, (int, float)) else (old.get("stock", 0) if isinstance(old, dict) else 0)
                old_price = (old.get("price", 0) if isinstance(old, dict) else 0)
                delta = new_stock - old_stock
                price_delta = round(new_price - old_price, 2)
                # zabeleži če se je spremenila zaloga ALI cena
                if delta != 0 or price_delta != 0:
                    changes.append({"sku": sku, "name": p.get("name",""), "category": p.get("category",""),
                                    "image": p.get("image",""), "price": new_price,
                                    "old": old_stock, "new": new_stock, "delta": delta, "is_new": False,
                                    "old_price": old_price, "new_price": new_price, "price_delta": price_delta})
        # izdelki ki so izginili (bili prej, zdaj jih ni)
        for sku, old in prev.items():
            if sku not in cur:
                old_stock = old if isinstance(old, (int, float)) else (old.get("stock", 0) if isinstance(old, dict) else 0)
                changes.append({"sku": sku, "name": (old.get("name","") if isinstance(old, dict) else ""),
                                "category": (old.get("category","") if isinstance(old, dict) else ""),
                                "image": "", "price": 0,
                                "old": old_stock, "new": None, "delta": -old_stock, "is_gone": True,
                                "old_price": (old.get("price",0) if isinstance(old, dict) else 0), "new_price": None, "price_delta": 0})
    diff_payload = {
        "computed_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "prev_at": prev_at,
        "has_prev": bool(prev),
        "changes": changes,
    }
    try:
        HSPLUS_DIFF.write_text(_json.dumps(diff_payload, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass
    # posodobi snapshot (shrani lahke podatke: sku → {stock,name,price,category})
    new_snap = {sku: {"stock": p.get("stock",0), "name": p.get("name",""),
                      "price": p.get("price",0), "category": p.get("category","")}
                for sku, p in cur.items()}
    try:
        HSPLUS_SNAPSHOT.write_text(_json.dumps({
            "fetched_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "fetched_ts": _time.time(), "stock": new_snap
        }, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass
    return diff_payload

@app.post("/hsplus-catalog-upload")
async def hsplus_catalog_upload(file: UploadFile = File(...)):
    """Ročni upload HS+ XML (potegnjen iz b2b logina). Parsira in shrani v isti
    cache, ki ga katalog bere — vsi filtri/kartice delujejo takoj.
    Hkrati izračuna spremembo zaloge glede na prejšnji upload (winner signal)."""
    import json as _json, time as _time
    try:
        raw = await file.read()
        products = _hsplus_parse_xml(raw)
        if not products:
            return {"ok": False, "error": "V XML ni najdenih izdelkov (preveri, da je pravi catalog/export XML)."}
        # izračunaj diff PRED prepisom snapshota
        diff = _hsplus_compute_diff(products)
        fetched_at = datetime.now().strftime("%d.%m.%Y %H:%M")
        HSPLUS_CATALOG_CACHE.write_text(_json.dumps({
            "fetched_ts": _time.time(), "fetched_at": fetched_at,
            "source": "manual_upload", "products": products
        }, ensure_ascii=False), encoding="utf-8")
        # povzetek diffa za odgovor
        n_changes = len(diff.get("changes", []))
        return {"ok": True, "count": len(products), "fetched_at": fetched_at,
                "diff_has_prev": diff.get("has_prev"), "diff_changes": n_changes}
    except Exception as e:
        return {"ok": False, "error": f"Napaka pri branju XML: {e}"}

@app.get("/hsplus-stock-diff")
async def hsplus_stock_diff():
    """Vrne zadnjo spremembo zaloge (winnerji = padec, polnila = porast)."""
    import json as _json
    if not HSPLUS_DIFF.exists():
        return {"ok": True, "has_prev": False, "changes": []}
    try:
        d = _json.loads(HSPLUS_DIFF.read_text(encoding="utf-8"))
        return {"ok": True, **d}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def hsuvoz_cleanup():
    """Briše JSON datoteke starejše od 30 dni."""
    try:
        cutoff = datetime.now().timestamp() - (30 * 86400)
        for f in HSUVOZ_DIR.glob("*.json"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[hsuvoz] cleanup err: {e}")


@app.post("/hsuvoz-upload")
async def hsuvoz_upload(file: UploadFile = File(...)):
    """Sprejme CSV naročilnic, združi količine po SKU, shrani v current + history."""
    import csv
    from io import StringIO
    try:
        content = (await file.read()).decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(StringIO(content))

        # Normalizacija headerjev
        rows = []
        for row in reader:
            norm = {k.strip().replace('\ufeff', ''): v.strip() for k, v in row.items()}
            rows.append(norm)

        if not rows:
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)

        # Najdi kolone (ID naročila, SKU, Naziv, Količina)
        sample = rows[0]
        keys = list(sample.keys())

        def find_col(candidates):
            for c in candidates:
                for k in keys:
                    if c.lower() in k.lower():
                        return k
            return None

        id_col  = find_col(["id naročila", "id narocila", "id"])
        sku_col = find_col(["sku"])
        naz_col = find_col(["naziv", "name"])
        qty_col = find_col(["količina", "kolicina", "qty", "quantity"])

        if not sku_col or not qty_col:
            return JSONResponse({"error": f"Ne najdem SKU/Količina kolon. Najdene: {keys}"}, status_code=400)

        # Združi po SKU
        sku_map = {}  # sku → {naziv, qty, orders: [id,...]}
        for row in rows:
            sku = (row.get(sku_col) or "").strip()
            if not sku:
                continue
            try:
                qty = int(float((row.get(qty_col) or "0").replace(",", ".")))
            except:
                qty = 0
            naziv = (row.get(naz_col) or "").strip() if naz_col else ""
            order_id = (row.get(id_col) or "").strip() if id_col else ""

            if sku not in sku_map:
                sku_map[sku] = {"sku": sku, "naziv": naziv, "qty": 0, "orders": [], "done": False}
            sku_map[sku]["qty"] += qty
            if order_id and order_id not in sku_map[sku]["orders"]:
                sku_map[sku]["orders"].append(order_id)
            # Ohrani naziv (vzemi prvega ki ni prazen)
            if not sku_map[sku]["naziv"] and naziv:
                sku_map[sku]["naziv"] = naziv

        items = sorted(sku_map.values(), key=lambda x: -x["qty"])

        # DEDUP: odstrani postavke, ki so že v "NAROČILO" (HSUVOZ_ORDER) —
        # da ni dvojnega dela z ročnim brisanjem že obdelanih SKU-jev.
        skipped = []
        try:
            if HSUVOZ_ORDER.exists():
                order_data = json.loads(HSUVOZ_ORDER.read_text(encoding="utf-8"))
                order_items = order_data.get("items", []) if isinstance(order_data, dict) else (order_data or [])
                order_skus = set((it.get("sku") or "").strip().upper() for it in order_items if it.get("sku"))
                if order_skus:
                    kept = []
                    for it in items:
                        if (it.get("sku") or "").strip().upper() in order_skus:
                            skipped.append(it["sku"])
                        else:
                            kept.append(it)
                    items = kept
        except Exception:
            pass

        # Naloži obstoječe done state iz currenta (ohrani done flag)
        if HSUVOZ_CURRENT.exists():
            try:
                existing = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
                done_map = {it["sku"]: it.get("done", False) for it in existing.get("items", [])}
                for it in items:
                    if it["sku"] in done_map:
                        it["done"] = done_map[it["sku"]]
            except:
                pass

        ts = datetime.now(timezone.utc).isoformat()
        payload = {
            "uploaded_at": ts,
            "filename": file.filename,
            "total_skus": len(items),
            "items": items,
        }

        # Shrani current
        HSUVOZ_CURRENT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        # Shrani v history
        hsuvoz_cleanup()
        hist_file = HSUVOZ_DIR / f"hsuvoz_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.json"
        hist_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True, "total_skus": len(items), "uploaded_at": ts, "filename": file.filename, "skipped_in_order": len(skipped), "skipped_skus": skipped, "total_in_csv": len(items) + len(skipped)}

    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-filter-ordered")
async def hsuvoz_filter_ordered(file: UploadFile = File(...)):
    """Naloži file ŽE NAROČENIH (stolpca: sku, stock). Odstrani postavke iz HS+ seznama,
    kjer naročeno (stock) >= potreba (kolicina). Če stock < kolicina, postavko pusti
    in doda 'se_potrebuje' = kolicina - stock. Če SKU ni v file, pusti celo."""
    try:
        if not HSUVOZ_CURRENT.exists():
            return JSONResponse({"error": "Najprej naloži HS+ seznam za naročit."}, status_code=400)

        raw = await file.read()
        fn = (file.filename or "").lower()

        # zgradi mapo sku.upper() -> stock
        ordered = {}
        if fn.endswith(".xlsx") or fn.endswith(".xlsm"):
            from openpyxl import load_workbook
            from io import BytesIO
            wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            header_seen = False
            sku_col, stock_col = 0, 1
            for r in ws.iter_rows(values_only=True):
                if r is None:
                    continue
                if not header_seen:
                    header_seen = True
                    # zaznaj stolpca iz glave (sku, stock) če obstajata
                    hdr = [str(c).strip().lower() if c is not None else "" for c in r]
                    if "sku" in hdr: sku_col = hdr.index("sku")
                    for cand in ("stock", "kolicina", "qty", "naroceno", "naroceno"):
                        if cand in hdr: stock_col = hdr.index(cand); break
                    # če glava ni besedilna (so že podatki), ne preskoči
                    if not any(isinstance(c, str) and c.strip() for c in r):
                        header_seen = True  # vseeno; spodaj obdela kot podatke
                    else:
                        continue
                try:
                    sku = str(r[sku_col]).strip() if len(r) > sku_col and r[sku_col] is not None else ""
                    stk = r[stock_col] if len(r) > stock_col else 0
                    stock = int(float(str(stk).replace(",", "."))) if stk not in (None, "") else 0
                except (ValueError, TypeError, IndexError):
                    continue
                if sku:
                    ordered[sku.upper()] = ordered.get(sku.upper(), 0) + stock
        else:
            # CSV
            text = raw.decode("utf-8-sig", errors="replace")
            first = text.split("\n", 1)[0]
            sep = ";" if first.count(";") > first.count(",") else ("\t" if "\t" in first else ",")
            import csv as _csv
            from io import StringIO as _SIO
            reader = _csv.reader(_SIO(text), delimiter=sep)
            rows = list(reader)
            start = 0
            if rows:
                h = [c.strip().lower() for c in rows[0]]
                sku_col, stock_col = 0, 1
                if "sku" in h: sku_col = h.index("sku")
                for cand in ("stock","kolicina","qty","naroceno","naroceno"):
                    if cand in h: stock_col = h.index(cand); break
                # preskoči glavo če je besedilna
                if any(not c.replace(".","").replace(",","").isdigit() and c for c in rows[0]):
                    start = 1
            for r in rows[start:]:
                if len(r) <= max(sku_col, stock_col):
                    continue
                sku = (r[sku_col] or "").strip()
                try:
                    stock = int(float((r[stock_col] or "0").strip().replace(",", "."))) if r[stock_col] else 0
                except (ValueError, TypeError):
                    stock = 0
                if sku:
                    ordered[sku.upper()] = ordered.get(sku.upper(), 0) + stock

        if not ordered:
            return JSONResponse({"error": "V datoteki ni najdenih SKU/stock podatkov."}, status_code=400)

        # obdelaj HS+ seznam
        data = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
        items = data.get("items", [])
        kept = []
        removed = []        # popolnoma pokrito → odstranjeno
        partial = []        # delno (stock < kolicina)
        def _to_int(v):
            try:
                return int(float(str(v).replace(",", ".")))
            except (ValueError, TypeError):
                return 0
        for it in items:
            sku_u = (it.get("sku") or "").strip().upper()
            # HS+ postavke uporabljajo 'qty' (potreba); fallback na 'kolicina'
            kolicina = _to_int(it.get("qty", it.get("kolicina", 0)))
            if sku_u in ordered:
                stock = _to_int(ordered[sku_u])
                if stock >= kolicina:
                    removed.append({"sku": it.get("sku"), "kolicina": kolicina, "stock": stock})
                    continue  # odstrani
                else:
                    # pusti, pokaži še potrebno
                    it["se_potrebuje"] = kolicina - stock
                    it["naroceno_stock"] = stock
                    partial.append({"sku": it.get("sku"), "kolicina": kolicina, "stock": stock, "se_potrebuje": kolicina - stock})
                    kept.append(it)
            else:
                # ni v file → pusti celo, počisti morebitne stare oznake
                it.pop("se_potrebuje", None)
                it.pop("naroceno_stock", None)
                kept.append(it)

        data["items"] = kept
        data["total_skus"] = len(kept)
        HSUVOZ_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

        return {
            "ok": True,
            "removed_count": len(removed),
            "partial_count": len(partial),
            "remaining": len(kept),
            "removed_skus": [r["sku"] for r in removed][:100],
            "partial": partial[:100],
            "ordered_total": len(ordered),
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/hsuvoz-data")
async def hsuvoz_data():
    """Vrne trenutne HS+ uvoz podatke."""
    try:
        if not HSUVOZ_CURRENT.exists():
            return {"loaded": False, "items": []}
        data = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
        return {"loaded": True, **data}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-set-done")
async def hsuvoz_set_done(sku: str = None, done: str = "1", request: Request = None):
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw:
                    body = json.loads(raw)
                    sku = body.get("sku")
                    done = str(body.get("done", True)).lower()
            except: pass
        done_bool = done not in ("0", "false", "False")
        if not HSUVOZ_CURRENT.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)
        data = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
        for it in data.get("items", []):
            if it["sku"] == sku:
                it["done"] = done_bool
                break
        HSUVOZ_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-edit-sku")
async def hsuvoz_edit_sku(old_sku: str = None, new_sku: str = None, source: str = "current", request: Request = None):
    """Preimenuje SKU v current ali order."""
    try:
        if (not old_sku or not new_sku) and request:
            try:
                raw = await request.body()
                if raw:
                    body = __import__("json").loads(raw)
                    old_sku = old_sku or body.get("old_sku")
                    new_sku = new_sku or (body.get("new_sku") or "").strip()
                    source = body.get("source", source)
            except: pass
        new_sku = (new_sku or "").strip()
        if not new_sku:
            return JSONResponse({"error": "Nov SKU je prazen."}, status_code=400)
        file = HSUVOZ_CURRENT if source == "current" else HSUVOZ_ORDER
        if not file.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)
        data = json.loads(file.read_text(encoding="utf-8"))
        for it in data.get("items", []):
            if it["sku"] == old_sku:
                it["sku"] = new_sku
                break
        file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/hsuvoz-history")
async def hsuvoz_history():
    """Vrne seznam zgodovinskih uploadov (30 dni)."""
    hsuvoz_cleanup()
    try:
        items = []
        for f in sorted(HSUVOZ_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                d = json.loads(f.read_text(encoding="utf-8"))
                items.append({
                    "filename": f.name,
                    "original_filename": d.get("filename", f.name),
                    "uploaded_at": d.get("uploaded_at", ""),
                    "total_skus": d.get("total_skus", 0),
                    "size": f.stat().st_size,
                })
            except:
                pass
        return {"items": items}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-load-history")
async def hsuvoz_load_history(filename: str = None, request: Request = None):
    """Naloži zgodovinski upload kot current."""
    try:
        if not filename and request:
            try:
                raw = await request.body()
                if raw: filename = __import__("json").loads(raw).get("filename")
            except: pass
        fname = filename
        hist_file = HSUVOZ_DIR / fname
        if not hist_file.exists():
            return JSONResponse({"error": "Datoteka ne obstaja."}, status_code=404)
        data = json.loads(hist_file.read_text(encoding="utf-8"))
        HSUVOZ_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "total_skus": data.get("total_skus", 0)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

# ═══════════════════════════════════════════════════════════════
# HS+ NAROČANJE — naročilo composer (ločen persistent state)
# ═══════════════════════════════════════════════════════════════
HSUVOZ_ORDER = DATA_DIR / "hsuvoz_order.json"

@app.post("/hsuvoz-move-to-order")
async def hsuvoz_move_to_order(sku: str = None, request: Request = None):
    """Premakne SKU iz 'za naročilo' v 'naročilo'."""
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw: sku = __import__("json").loads(raw).get("sku")
            except: pass
        if not sku or not HSUVOZ_CURRENT.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=400)

        # Poberi iz current
        current = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
        item = next((it for it in current.get("items", []) if it["sku"] == sku), None)
        if not item:
            return JSONResponse({"error": "SKU ne obstaja."}, status_code=404)

        # Dodaj v order (ali posodobi qty)
        order = {"items": []}
        if HSUVOZ_ORDER.exists():
            try: order = json.loads(HSUVOZ_ORDER.read_text(encoding="utf-8"))
            except: pass

        existing = next((it for it in order["items"] if it["sku"] == sku), None)
        if existing:
            existing["qty"] += item["qty"]
            existing["orders"] = list(set(existing.get("orders", []) + item.get("orders", [])))
        else:
            order["items"].append({**item, "done": False})

        HSUVOZ_ORDER.write_text(json.dumps(order, ensure_ascii=False, indent=2), encoding="utf-8")

        # Odstrani iz current
        current["items"] = [it for it in current["items"] if it["sku"] != sku]
        HSUVOZ_CURRENT.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-move-back")
async def hsuvoz_move_back(sku: str = None, request: Request = None):
    """Vrne SKU iz naročila nazaj v 'za naročilo'."""
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw: sku = __import__("json").loads(raw).get("sku")
            except: pass
        if not sku or not HSUVOZ_ORDER.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=400)

        order = json.loads(HSUVOZ_ORDER.read_text(encoding="utf-8"))
        item = next((it for it in order["items"] if it["sku"] == sku), None)
        if not item:
            return JSONResponse({"error": "SKU ne obstaja v naročilu."}, status_code=404)

        # Vrni v current
        current = {"items": []}
        if HSUVOZ_CURRENT.exists():
            try: current = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
            except: pass

        if not any(it["sku"] == sku for it in current["items"]):
            current["items"].append({**item, "done": False})
            HSUVOZ_CURRENT.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")

        order["items"] = [it for it in order["items"] if it["sku"] != sku]
        HSUVOZ_ORDER.write_text(json.dumps(order, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-delete-item")
async def hsuvoz_delete_item(sku: str = None, source: str = "current", request: Request = None):
    """Zbriše SKU iz seznama. Sprejme query param ali JSON body."""
    try:
        # Poskusi dobiti iz JSON body če query param ni podan
        if not sku and request:
            try:
                raw = await request.body()
                if raw:
                    body = json.loads(raw)
                    sku = body.get("sku") or (body.get("skus") or [None])[0]
                    source = body.get("source", source)
            except: pass

        if not sku:
            return JSONResponse({"error": "Manjka SKU."}, status_code=400)

        file = HSUVOZ_CURRENT if source == "current" else HSUVOZ_ORDER
        if not file.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)

        data = json.loads(file.read_text(encoding="utf-8"))
        before = len(data.get("items", []))
        data["items"] = [it for it in data.get("items", []) if str(it.get("sku","")).strip() != str(sku).strip()]
        file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "deleted": before - len(data["items"])}
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/hsuvoz-order-data")
async def hsuvoz_order_data():
    """Vrne seznam SKU-jev v naročilu."""
    try:
        if not HSUVOZ_ORDER.exists():
            return {"items": []}
        return json.loads(HSUVOZ_ORDER.read_text(encoding="utf-8"))
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/hsuvoz-order-clear")
async def hsuvoz_order_clear():
    """Počisti celotno naročilo."""
    try:
        HSUVOZ_ORDER.write_text(json.dumps({"items": []}, ensure_ascii=False), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/hsuvoz-current-clear")
async def hsuvoz_current_clear():
    """Počisti celoten seznam 'za naročilo'."""
    try:
        if HSUVOZ_CURRENT.exists():
            data = json.loads(HSUVOZ_CURRENT.read_text(encoding="utf-8"))
            data["items"] = []
            HSUVOZ_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── APTEL NAROČANJE ─────────────────────────────────────────────────────────
APTEL_DIR = DATA_DIR / "aptel_history"
APTEL_DIR.mkdir(exist_ok=True, parents=True)
APTEL_CURRENT = DATA_DIR / "aptel_current.json"


def aptel_cleanup():
    """Briše JSON datoteke starejše od 30 dni."""
    try:
        cutoff = datetime.now().timestamp() - (30 * 86400)
        for f in APTEL_DIR.glob("*.json"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[aptel] cleanup err: {e}")


@app.post("/aptel-upload")
async def aptel_upload(file: UploadFile = File(...)):
    """Sprejme CSV naročilnic, združi količine po SKU, shrani v current + history."""
    import csv
    from io import StringIO
    try:
        content = (await file.read()).decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(StringIO(content))

        # Normalizacija headerjev
        rows = []
        for row in reader:
            norm = {k.strip().replace('\ufeff', ''): v.strip() for k, v in row.items()}
            rows.append(norm)

        if not rows:
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)

        # Najdi kolone (ID naročila, SKU, Naziv, Količina)
        sample = rows[0]
        keys = list(sample.keys())

        def find_col(candidates):
            for c in candidates:
                for k in keys:
                    if c.lower() in k.lower():
                        return k
            return None

        id_col  = find_col(["id naročila", "id narocila", "id"])
        sku_col = find_col(["sku"])
        naz_col = find_col(["naziv", "name"])
        qty_col = find_col(["količina", "kolicina", "qty", "quantity"])

        if not sku_col or not qty_col:
            return JSONResponse({"error": f"Ne najdem SKU/Količina kolon. Najdene: {keys}"}, status_code=400)

        # Združi po SKU
        sku_map = {}  # sku → {naziv, qty, orders: [id,...]}
        for row in rows:
            sku = (row.get(sku_col) or "").strip()
            if not sku:
                continue
            try:
                qty = int(float((row.get(qty_col) or "0").replace(",", ".")))
            except:
                qty = 0
            naziv = (row.get(naz_col) or "").strip() if naz_col else ""
            order_id = (row.get(id_col) or "").strip() if id_col else ""

            if sku not in sku_map:
                sku_map[sku] = {"sku": sku, "naziv": naziv, "qty": 0, "orders": [], "done": False}
            sku_map[sku]["qty"] += qty
            if order_id and order_id not in sku_map[sku]["orders"]:
                sku_map[sku]["orders"].append(order_id)
            # Ohrani naziv (vzemi prvega ki ni prazen)
            if not sku_map[sku]["naziv"] and naziv:
                sku_map[sku]["naziv"] = naziv

        items = sorted(sku_map.values(), key=lambda x: -x["qty"])

        # Naloži obstoječe done state iz currenta (ohrani done flag)
        if APTEL_CURRENT.exists():
            try:
                existing = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
                done_map = {it["sku"]: it.get("done", False) for it in existing.get("items", [])}
                for it in items:
                    if it["sku"] in done_map:
                        it["done"] = done_map[it["sku"]]
            except:
                pass

        ts = datetime.now(timezone.utc).isoformat()
        payload = {
            "uploaded_at": ts,
            "filename": file.filename,
            "total_skus": len(items),
            "items": items,
        }

        # Shrani current
        APTEL_CURRENT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        # Shrani v history
        aptel_cleanup()
        hist_file = APTEL_DIR / f"aptel_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.json"
        hist_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True, "total_skus": len(items), "uploaded_at": ts, "filename": file.filename}

    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/aptel-data")
async def aptel_data():
    """Vrne trenutne Aptel uvoz podatke."""
    try:
        if not APTEL_CURRENT.exists():
            return {"loaded": False, "items": []}
        data = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
        return {"loaded": True, **data}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-set-done")
async def aptel_set_done(sku: str = None, done: str = "1", request: Request = None):
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw:
                    body = json.loads(raw)
                    sku = body.get("sku")
                    done = str(body.get("done", True)).lower()
            except: pass
        done_bool = done not in ("0", "false", "False")
        if not APTEL_CURRENT.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)
        data = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
        for it in data.get("items", []):
            if it["sku"] == sku:
                it["done"] = done_bool
                break
        APTEL_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-edit-sku")
async def aptel_edit_sku(old_sku: str = None, new_sku: str = None, source: str = "current", request: Request = None):
    """Preimenuje SKU v current ali order."""
    try:
        if (not old_sku or not new_sku) and request:
            try:
                raw = await request.body()
                if raw:
                    body = __import__("json").loads(raw)
                    old_sku = old_sku or body.get("old_sku")
                    new_sku = new_sku or (body.get("new_sku") or "").strip()
                    source = body.get("source", source)
            except: pass
        new_sku = (new_sku or "").strip()
        if not new_sku:
            return JSONResponse({"error": "Nov SKU je prazen."}, status_code=400)
        file = APTEL_CURRENT if source == "current" else APTEL_ORDER
        if not file.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)
        data = json.loads(file.read_text(encoding="utf-8"))
        for it in data.get("items", []):
            if it["sku"] == old_sku:
                it["sku"] = new_sku
                break
        file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/aptel-history")
async def aptel_history():
    """Vrne seznam zgodovinskih uploadov (30 dni)."""
    aptel_cleanup()
    try:
        items = []
        for f in sorted(APTEL_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                d = json.loads(f.read_text(encoding="utf-8"))
                items.append({
                    "filename": f.name,
                    "original_filename": d.get("filename", f.name),
                    "uploaded_at": d.get("uploaded_at", ""),
                    "total_skus": d.get("total_skus", 0),
                    "size": f.stat().st_size,
                })
            except:
                pass
        return {"items": items}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-load-history")
async def aptel_load_history(filename: str = None, request: Request = None):
    """Naloži zgodovinski upload kot current."""
    try:
        if not filename and request:
            try:
                raw = await request.body()
                if raw: filename = __import__("json").loads(raw).get("filename")
            except: pass
        fname = filename
        hist_file = APTEL_DIR / fname
        if not hist_file.exists():
            return JSONResponse({"error": "Datoteka ne obstaja."}, status_code=404)
        data = json.loads(hist_file.read_text(encoding="utf-8"))
        APTEL_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "total_skus": data.get("total_skus", 0)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

# ═══════════════════════════════════════════════════════════════
# Aptel NAROČANJE — naročilo composer (ločen persistent state)
# ═══════════════════════════════════════════════════════════════
APTEL_ORDER = DATA_DIR / "aptel_order.json"

@app.post("/aptel-move-to-order")
async def aptel_move_to_order(sku: str = None, request: Request = None):
    """Premakne SKU iz 'za naročilo' v 'naročilo'."""
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw: sku = __import__("json").loads(raw).get("sku")
            except: pass
        if not sku or not APTEL_CURRENT.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=400)

        # Poberi iz current
        current = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
        item = next((it for it in current.get("items", []) if it["sku"] == sku), None)
        if not item:
            return JSONResponse({"error": "SKU ne obstaja."}, status_code=404)

        # Dodaj v order (ali posodobi qty)
        order = {"items": []}
        if APTEL_ORDER.exists():
            try: order = json.loads(APTEL_ORDER.read_text(encoding="utf-8"))
            except: pass

        existing = next((it for it in order["items"] if it["sku"] == sku), None)
        if existing:
            existing["qty"] += item["qty"]
            existing["orders"] = list(set(existing.get("orders", []) + item.get("orders", [])))
        else:
            order["items"].append({**item, "done": False})

        APTEL_ORDER.write_text(json.dumps(order, ensure_ascii=False, indent=2), encoding="utf-8")

        # Odstrani iz current
        current["items"] = [it for it in current["items"] if it["sku"] != sku]
        APTEL_CURRENT.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-move-back")
async def aptel_move_back(sku: str = None, request: Request = None):
    """Vrne SKU iz naročila nazaj v 'za naročilo'."""
    try:
        if not sku and request:
            try:
                raw = await request.body()
                if raw: sku = __import__("json").loads(raw).get("sku")
            except: pass
        if not sku or not APTEL_ORDER.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=400)

        order = json.loads(APTEL_ORDER.read_text(encoding="utf-8"))
        item = next((it for it in order["items"] if it["sku"] == sku), None)
        if not item:
            return JSONResponse({"error": "SKU ne obstaja v naročilu."}, status_code=404)

        # Vrni v current
        current = {"items": []}
        if APTEL_CURRENT.exists():
            try: current = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
            except: pass

        if not any(it["sku"] == sku for it in current["items"]):
            current["items"].append({**item, "done": False})
            APTEL_CURRENT.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")

        order["items"] = [it for it in order["items"] if it["sku"] != sku]
        APTEL_ORDER.write_text(json.dumps(order, ensure_ascii=False, indent=2), encoding="utf-8")

        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-delete-item")
async def aptel_delete_item(sku: str = None, source: str = "current", request: Request = None):
    """Zbriše SKU iz seznama. Sprejme query param ali JSON body."""
    try:
        # Poskusi dobiti iz JSON body če query param ni podan
        if not sku and request:
            try:
                raw = await request.body()
                if raw:
                    body = json.loads(raw)
                    sku = body.get("sku") or (body.get("skus") or [None])[0]
                    source = body.get("source", source)
            except: pass

        if not sku:
            return JSONResponse({"error": "Manjka SKU."}, status_code=400)

        file = APTEL_CURRENT if source == "current" else APTEL_ORDER
        if not file.exists():
            return JSONResponse({"error": "Ni podatkov."}, status_code=404)

        data = json.loads(file.read_text(encoding="utf-8"))
        before = len(data.get("items", []))
        data["items"] = [it for it in data.get("items", []) if str(it.get("sku","")).strip() != str(sku).strip()]
        file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "deleted": before - len(data["items"])}
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/aptel-order-data")
async def aptel_order_data():
    """Vrne seznam SKU-jev v naročilu."""
    try:
        if not APTEL_ORDER.exists():
            return {"items": []}
        return json.loads(APTEL_ORDER.read_text(encoding="utf-8"))
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/aptel-order-clear")
async def aptel_order_clear():
    """Počisti celotno naročilo."""
    try:
        APTEL_ORDER.write_text(json.dumps({"items": []}, ensure_ascii=False), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/aptel-current-clear")
async def aptel_current_clear():
    """Počisti celoten seznam 'za naročilo'."""
    try:
        if APTEL_CURRENT.exists():
            data = json.loads(APTEL_CURRENT.read_text(encoding="utf-8"))
            data["items"] = []
            APTEL_CURRENT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

# ─── INVENTURA ────────────────────────────────────────────────────────────────

INVENTURA_DIR = DATA_DIR / "inventura"
INVENTURA_DIR.mkdir(exist_ok=True, parents=True)

INVENTURA_CURRENT = INVENTURA_DIR / "_current.json"

DEJAVU_REGULAR = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
DEJAVU_BOLD    = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"


@app.post("/zaloga-packing-pdf")
async def zaloga_packing_pdf(data: dict):
    """Carinska packing lista (PDF) iz RS packing_boxes.
    Za vsak box izpiše izdelke: naziv + koda + količina kosov.
    Vhod: {market}. Bere packing_boxes iz aktivne RS seje."""
    try:
        import io
        # Uporabi skupni helper — vsebuje srbske nazive po SKU (RS feed) + fallback
        pboxes, _err = _packing_collect(data.get("market", "rs"))
        if _err:
            return JSONResponse({"error": _err}, status_code=400)

        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        pdfmetrics.registerFont(TTFont("DejaVu", DEJAVU_REGULAR))
        pdfmetrics.registerFont(TTFont("DejaVu-Bold", DEJAVU_BOLD))

        datum = _lj_now().strftime("%d. %m. %Y  %H:%M")
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)

        s_title = ParagraphStyle("t", fontSize=15, fontName="DejaVu-Bold", spaceAfter=3)
        s_sub   = ParagraphStyle("s", fontSize=9, fontName="DejaVu", textColor=colors.HexColor("#64748b"), spaceAfter=14)
        s_box   = ParagraphStyle("b", fontSize=12, fontName="DejaVu-Bold", spaceBefore=10, spaceAfter=5, textColor=colors.HexColor("#1e293b"))
        s_hdr   = ParagraphStyle("h", fontSize=9, fontName="DejaVu-Bold", textColor=colors.white, alignment=0, leading=12)
        s_cell  = ParagraphStyle("c", fontSize=9, fontName="DejaVu", leading=12)
        s_code  = ParagraphStyle("k", fontSize=9, fontName="DejaVu-Bold", leading=12)
        s_qty   = ParagraphStyle("q", fontSize=9, fontName="DejaVu-Bold", alignment=1, leading=12)

        total_boxes = len(pboxes)
        total_pcs = sum(e.get("kos", 0) for items in pboxes.values() for e in items)

        story = [
            Paragraph("PACKING LISTA / SPECIFIKACIJA", s_title),
            Paragraph(f"Datum: {datum}  |  Št. boxov: {total_boxes}  |  Skupaj kosov: {total_pcs}", s_sub),
        ]

        # sortiraj bokse po številki (numerično če gre)
        def _boxkey(b):
            try: return (0, int(b))
            except (ValueError, TypeError): return (1, b)
        for box in sorted(pboxes.keys(), key=_boxkey):
            items = pboxes[box]
            box_pcs = sum(e.get("kos", 0) for e in items)
            story.append(Paragraph(f"📦 BOX {box}  ·  {len(items)} izdelkov  ·  {box_pcs} kosov", s_box))
            tdata = [[Paragraph("Naziv", s_hdr), Paragraph("Koda", s_hdr), Paragraph("Kosov", s_hdr)]]
            for e in items:
                tdata.append([
                    Paragraph(str(e.get("naziv") or ""), s_cell),
                    Paragraph(str(e.get("sku") or ""), s_code),
                    Paragraph(str(e.get("kos") or 0), s_qty),
                ])
            t = Table(tdata, colWidths=[11.0*cm, 4.5*cm, 2.0*cm], repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1e293b")),
                ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                ("TOPPADDING", (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                ("LEFTPADDING", (0,0), (-1,-1), 6),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
            ]))
            story.append(t)

        doc.build(story)
        buf.seek(0)
        fn = f"packing_lista_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.pdf"
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fn}"'})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


def _packing_sr_names():
    """Zgradi mapo SKU(MPN, velike črke) -> srbski naziv iz RS feeda.
    Če feed ni naložen ali SKU ni najden, vrne prazno mapo (fallback na original)."""
    out = {}
    try:
        rs_feed = feed_by_lang.get("rs", {})
        for _gid, prod in rs_feed.items():
            mpn = (prod.get("mpn") or "").strip().upper()
            title = (prod.get("title") or "").strip()
            if mpn and title:
                out.setdefault(mpn, title)
    except Exception:
        pass
    return out


def _packing_collect(market: str):
    """Zbere packing podatke (isto kot PDF): {box: [{sku,naziv,kos}]}.
    Naziv = srbski iz RS feeda (po SKU), fallback na originalni naziv.
    Vrne (pboxes_final, error_or_None)."""
    path = _zaloga_current_path(market)
    if not path.exists():
        return None, "Ni aktivne seje"
    sess = json.loads(path.read_text(encoding="utf-8"))
    sr_names = _packing_sr_names()
    def _sr(sku, fallback):
        return sr_names.get((sku or "").strip().upper(), fallback or "")
    combined = {}
    def _add(box, sku, naziv, kos):
        box = str(box).strip()
        if not box or not sku or kos <= 0:
            return
        naziv = _sr(sku, naziv)  # srbski naziv (fallback original)
        combined.setdefault(box, {})
        if sku in combined[box]:
            combined[box][sku]["kos"] += kos
            if naziv and not combined[box][sku].get("naziv"):
                combined[box][sku]["naziv"] = naziv
        else:
            combined[box][sku] = {"naziv": naziv or "", "kos": kos}
    for it in sess.get("items", []):
        box = str(it.get("box", "") or "").strip()
        if box and it.get("status") == "ok":
            try:
                kos = int(it.get("picked", 0) or 0)
            except (ValueError, TypeError):
                kos = 0
            if kos <= 0:
                try: kos = int(it.get("qty", 0) or 0)
                except (ValueError, TypeError): kos = 0
            _add(box, it.get("sku", ""), it.get("naziv", ""), kos)
    pboxes = sess.get("packing_boxes") or {}
    for box, items in pboxes.items():
        for e in items:
            _add(box, e.get("sku", ""), e.get("naziv", ""), e.get("kos", 0))
    if not combined:
        return None, "Ni boxov za izpis (ni zaklenjenih postavk ne razdeljenih)"
    pboxes_final = {}
    for box, skus in combined.items():
        pboxes_final[box] = [{"sku": sku, "naziv": v["naziv"], "kos": v["kos"]} for sku, v in skus.items()]
    return pboxes_final, None


def _packing_boxkey(b):
    try: return (0, int(b))
    except (ValueError, TypeError): return (1, b)


@app.post("/zaloga-packing-xlsx")
async def zaloga_packing_xlsx(data: dict):
    """Packing lista v XLSX — stolpec Box v vsaki vrstici + list Pregled."""
    try:
        import io
        pboxes, err = _packing_collect(data.get("market", "rs"))
        if err:
            return JSONResponse({"error": err}, status_code=400)

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        NAVY="1E293B"; GREY="64748B"; LIGHT="F1F5F9"; WHITE="FFFFFF"
        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin,right=thin,top=thin,bottom=thin)

        wb = Workbook()
        ws = wb.active; ws.title = "Packing lista"; ws.sheet_view.showGridLines = False

        total_boxes = len(pboxes)
        total_pcs = sum(e["kos"] for items in pboxes.values() for e in items)
        ws["A1"] = "PACKING LISTA / SPECIFIKACIJA"
        ws["A1"].font = Font(name="Arial", size=15, bold=True, color=NAVY)
        ws.merge_cells("A1:D1")
        ws["A2"] = f"Datum: {_lj_now().strftime('%d. %m. %Y  %H:%M')}   |   Št. boxov: {total_boxes}   |   Skupaj kosov: {total_pcs}"
        ws["A2"].font = Font(name="Arial", size=9, color=GREY)
        ws.merge_cells("A2:D2")

        # glava — Box je prvi stolpec, v vsaki vrstici
        hdrs = ["Box","Naziv","Koda","Kosov"]
        hr = 4
        for c,h in enumerate(hdrs,1):
            cell = ws.cell(row=hr,column=c,value=h)
            cell.font = Font(name="Arial",size=9,bold=True,color=WHITE)
            cell.fill = PatternFill("solid",start_color=NAVY)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c in (1,4) else "left", vertical="center")
        row = hr + 1
        for box in sorted(pboxes.keys(), key=_packing_boxkey):
            for e in pboxes[box]:
                ws.cell(row=row,column=1,value=box).alignment = Alignment(horizontal="center")
                ws.cell(row=row,column=1).font = Font(name="Arial",size=9,bold=True)
                ws.cell(row=row,column=2,value=e.get("naziv","")).font = Font(name="Arial",size=9)
                ws.cell(row=row,column=3,value=e.get("sku","")).font = Font(name="Arial",size=9,bold=True)
                qc = ws.cell(row=row,column=4,value=e.get("kos",0)); qc.font = Font(name="Arial",size=9,bold=True); qc.alignment = Alignment(horizontal="center")
                for c in range(1,5):
                    ws.cell(row=row,column=c).border = border
                row += 1
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 56
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 10

        # List Pregled
        ws2 = wb.create_sheet("Pregled"); ws2.sheet_view.showGridLines = False
        ws2["A1"] = "PREGLED BOXOV"; ws2["A1"].font = Font(name="Arial",size=13,bold=True,color=NAVY)
        ws2.merge_cells("A1:C1")
        for c,h in enumerate(["Box","Št. izdelkov","Kosov"],1):
            cell = ws2.cell(row=3,column=c,value=h)
            cell.font = Font(name="Arial",size=10,bold=True,color=WHITE)
            cell.fill = PatternFill("solid",start_color=NAVY); cell.border=border
            cell.alignment=Alignment(horizontal="center")
        r=4
        for box in sorted(pboxes.keys(), key=_packing_boxkey):
            items = pboxes[box]
            ws2.cell(row=r,column=1,value=box).font = Font(name="Arial",size=10,bold=True)
            ws2.cell(row=r,column=1).alignment = Alignment(horizontal="center")
            ws2.cell(row=r,column=2,value=len(items)).alignment = Alignment(horizontal="center")
            ws2.cell(row=r,column=3,value=sum(e["kos"] for e in items)).alignment = Alignment(horizontal="center")
            for c in range(1,4): ws2.cell(row=r,column=c).border=border
            r+=1
        ws2.cell(row=r,column=1,value="SKUPAJ").font=Font(name="Arial",size=10,bold=True)
        ws2.cell(row=r,column=1).fill=PatternFill("solid",start_color=LIGHT)
        ws2.cell(row=r,column=2,value=f"=SUM(B4:B{r-1})").alignment=Alignment(horizontal="center")
        ws2.cell(row=r,column=2).font=Font(bold=True)
        ws2.cell(row=r,column=3,value=f"=SUM(C4:C{r-1})").alignment=Alignment(horizontal="center")
        ws2.cell(row=r,column=3).font=Font(bold=True)
        for c in range(1,4): ws2.cell(row=r,column=c).border=border
        ws2.column_dimensions["A"].width=10; ws2.column_dimensions["B"].width=14; ws2.column_dimensions["C"].width=12

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fn = f"packing_lista_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fn}"'})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/zaloga-packing-csv")
async def zaloga_packing_csv(data: dict):
    """Packing lista v CSV (ločilo ';', BOM za Excel) — stolpec Box v vsaki vrstici."""
    try:
        import io, csv as _csv
        pboxes, err = _packing_collect(data.get("market", "rs"))
        if err:
            return JSONResponse({"error": err}, status_code=400)
        out = io.StringIO()
        w = _csv.writer(out, delimiter=";")
        w.writerow(["Box","Naziv","Koda","Kosov"])
        for box in sorted(pboxes.keys(), key=_packing_boxkey):
            for e in pboxes[box]:
                w.writerow([box, e.get("naziv",""), e.get("sku",""), e.get("kos",0)])
        # BOM za pravilne šumnike v Excelu
        body = "\ufeff" + out.getvalue()
        fn = f"packing_lista_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.csv"
        return StreamingResponse(io.BytesIO(body.encode("utf-8")), media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{fn}"'})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


def inventura_cleanup():
    """Zbriše PDF-je in JSON-e starejše od 30 dni (ne _current.json)."""
    try:
        cutoff = datetime.now().timestamp() - (30 * 86400)
        for f in list(INVENTURA_DIR.glob("*.pdf")) + list(INVENTURA_DIR.glob("*.json")):
            if f.name == "_current.json":
                continue
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[inventura] cleanup err: {e}")


@app.get("/inventura-lookup-sku")
async def inventura_lookup_sku(sku: str = ""):
    """Vrne naziv + pozicijo za dani SKU iz baze zaloge (za dodajanje v inventuro).
    Inventura vleče te podatke samodejno — delavec vpiše samo SKU."""
    sku = (sku or "").strip()
    if not sku:
        return {"ok": False, "error": "Manjka SKU"}
    if not STOCK_CSV_FILE.exists():
        return {"ok": False, "error": "Zaloga ni naložena", "naziv": "", "pozicija": ""}
    import csv as _csv
    from io import StringIO as _SIO
    try:
        text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        naziv, pozicija = "", ""
        sku_l = sku.lower()
        for row in _csv.DictReader(_SIO(text)):
            rsku = (row.get("product_sku") or row.get("sku") or "").strip()
            if rsku.lower() == sku_l:
                # vzemi prvi zapis tega SKU (silux ima pozicijo); če pozneje najdemo silux s pozicijo, raje to
                wh = (row.get("warehouse") or "").strip().lower()
                rpos = (row.get("position") or "").strip()
                rtitle = (row.get("title") or "").strip()
                if not naziv and rtitle:
                    naziv = rtitle
                # pozicijo vzemi iz silux (tam je prava), sicer prvo neprazno
                if rpos and (wh == "silux" or not pozicija):
                    pozicija = rpos
        if not naziv and not pozicija:
            return {"ok": True, "najden": False, "naziv": "", "pozicija": "",
                    "opozorilo": f"SKU '{sku}' ni v zalogi — dodan bo brez naziva/pozicije."}
        return {"ok": True, "najden": True, "naziv": naziv, "pozicija": pozicija}
    except Exception as e:
        return {"ok": False, "error": str(e), "naziv": "", "pozicija": ""}


@app.get("/inventura-current")
async def inventura_get_current():
    """Vrne trenutno aktivno inventuro z diska."""
    if not INVENTURA_CURRENT.exists():
        return {"ok": False, "items": []}
    try:
        data = json.loads(INVENTURA_CURRENT.read_text(encoding="utf-8"))
        return {"ok": True, **data}
    except Exception:
        return {"ok": False, "items": []}


@app.post("/inventura-save-current")
async def inventura_save_current(data: dict):
    """Shrani trenutno stanje inventure na disk."""
    try:
        INVENTURA_CURRENT.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/inventura-clear-current")
async def inventura_clear_current():
    """Pobriše trenutno aktivno inventuro z diska."""
    try:
        if INVENTURA_CURRENT.exists():
            INVENTURA_CURRENT.unlink()
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/inventura-upload")
async def inventura_upload(file: UploadFile = File(...)):
    """Sprejme CSV izvoz, združi po SKU, vrne seznam."""
    try:
        import csv
        from io import StringIO
        content = (await file.read()).decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(StringIO(content))
        rows = [{k.strip().replace('\ufeff', ''): (v or '').strip() for k, v in row.items()} for row in reader]
        if not rows:
            return JSONResponse({"error": "Prazen CSV."}, status_code=400)
        keys = list(rows[0].keys())

        def fc(*cands):
            for c in cands:
                for k in keys:
                    if c.lower() in k.lower(): return k
            return None

        sku_col = fc("sku", "SKU")
        naz_col = fc("naziv", "name", "Naziv")
        pos_col = fc("pozicija", "position", "Pozicija")

        if not sku_col:
            return JSONResponse({"error": f"Ne najdem SKU stolpca. Najdeni: {keys}"}, status_code=400)

        sku_map = {}
        for row in rows:
            sku = (row.get(sku_col) or "").strip()
            if not sku: continue
            naziv    = (row.get(naz_col) or "").strip() if naz_col else ""
            pozicija = (row.get(pos_col) or "").strip() if pos_col else ""
            if sku not in sku_map:
                sku_map[sku] = {"sku": sku, "naziv": naziv, "pozicija": pozicija, "komentar": "", "kolicina_dejansko": None}
            if not sku_map[sku]["pozicija"] and pozicija: sku_map[sku]["pozicija"] = pozicija
            if not sku_map[sku]["naziv"] and naziv: sku_map[sku]["naziv"] = naziv

        items = sorted(sku_map.values(), key=lambda x: (x["pozicija"] or "zzz", x["sku"]))
        return {"ok": True, "total_skus": len(items), "filename": file.filename, "items": items}
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/inventura-pdf")
async def inventura_pdf(data: dict):
    """Generira PDF inventurni list z DejaVu fontom."""
    try:
        items = data.get("items", [])
        title_text = data.get("title", "Inventurni list")
        datum = data.get("datum", _lj_now().strftime("%d. %m. %Y"))
        filename_hint = data.get("filename", f"inventura_{datetime.now().strftime('%Y-%m-%d_%H-%M')}")
        if not items:
            return JSONResponse({"error": "Ni postavk."}, status_code=400)

        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase.pdfmetrics import stringWidth

        pdfmetrics.registerFont(TTFont("DejaVu", DEJAVU_REGULAR))
        pdfmetrics.registerFont(TTFont("DejaVu-Bold", DEJAVU_BOLD))

        def trunc_ellipsis(s, font, font_size, max_width_pt):
            if not s: return ""
            for sep in [",", "(", " -"]:
                idx = s.find(sep)
                if 0 < idx < len(s):
                    candidate = s[:idx].strip()
                    if stringWidth(candidate, font, font_size) <= max_width_pt:
                        s = candidate; break
            if stringWidth(s, font, font_size) <= max_width_pt: return s
            while s and stringWidth(s + "...", font, font_size) > max_width_pt: s = s[:-1]
            return s.strip() + "..."

        ROW_H = 0.65 * cm
        HDR_H = 0.7  * cm

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)

        s_title = ParagraphStyle("t", fontSize=14, fontName="DejaVu-Bold", spaceAfter=4)
        s_sub   = ParagraphStyle("s", fontSize=9,  fontName="DejaVu", textColor=colors.HexColor("#64748b"), spaceAfter=12)
        s_num   = ParagraphStyle("n", fontSize=7.5, fontName="DejaVu", leading=10, alignment=1)
        s_cell  = ParagraphStyle("c", fontSize=7.5, fontName="DejaVu", leading=10)
        s_sku   = ParagraphStyle("k", fontSize=7.5, fontName="DejaVu-Bold", leading=10)
        s_kom   = ParagraphStyle("m", fontSize=7.5, fontName="DejaVu", textColor=colors.HexColor("#7c3aed"), leading=10)
        s_hdr   = ParagraphStyle("h", fontSize=7.5, fontName="DejaVu-Bold", textColor=colors.white, leading=10, alignment=1)

        story = [
            Paragraph(title_text, s_title),
            Paragraph(f"Datum: {datum}  |  Skupaj SKU-jev: {len(items)}", s_sub),
        ]

        col_widths = [0.7*cm, 4.2*cm, 5.8*cm, 2.0*cm, 3.5*cm, 1.8*cm]
        naziv_max_pt = 5.8 * 28.35 - 10
        kom_max_pt   = 3.5 * 28.35 - 10

        table_data = [[
            Paragraph("#", s_hdr), Paragraph("SKU", s_hdr), Paragraph("Naziv", s_hdr),
            Paragraph("Pozicija", s_hdr), Paragraph("Komentar", s_hdr), Paragraph("Fizično ✓", s_hdr),
        ]]
        for i, it in enumerate(items, 1):
            komentar_raw = str(it.get("komentar") or "").strip()
            naziv_short  = trunc_ellipsis(str(it.get("naziv") or ""), "DejaVu", 7.5, naziv_max_pt)
            komentar_short = trunc_ellipsis(komentar_raw, "DejaVu", 7.5, kom_max_pt) if komentar_raw else ""
            table_data.append([
                Paragraph(str(i), s_num),
                Paragraph(str(it.get("sku") or ""), s_sku),
                Paragraph(naziv_short, s_cell),
                Paragraph(str(it.get("pozicija") or "—"), s_cell),
                Paragraph(komentar_short, s_kom) if komentar_short else Paragraph("", s_cell),
                Paragraph("", s_cell),
            ])

        row_heights = [HDR_H] + [ROW_H] * (len(table_data) - 1)
        tbl = Table(table_data, colWidths=col_widths, rowHeights=row_heights, repeatRows=1)
        row_styles = [
            ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#1e293b")),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#e2e8f0")),
            ("LINEBELOW",     (0,0), (-1,0), 1.5, colors.HexColor("#1e293b")),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 5),
            ("RIGHTPADDING",  (0,0), (-1,-1), 5),
            ("ALIGN",         (0,0), (0,-1), "CENTER"),
            ("ALIGN",         (3,1), (3,-1), "CENTER"),
            ("BOX",           (5,1), (5,-1), 0.8, colors.HexColor("#94a3b8")),
            ("BACKGROUND",    (5,1), (5,-1), colors.HexColor("#f0fdf4")),
        ]
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                row_styles.append(("BACKGROUND", (0,i), (-1,i), colors.HexColor("#f8fafc")))
        tbl.setStyle(TableStyle(row_styles))
        story.append(tbl)
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph(
            "Navodilo: V stolpec 'Fizično ✓' vpišite dejansko stanje zaloge. "
            "Prazno = ni pregledano.  ✓ = potrjeno.  0 = ni na zalogi.",
            ParagraphStyle("f", fontSize=7, fontName="DejaVu", textColor=colors.HexColor("#94a3b8"))
        ))
        doc.build(story)

        inventura_cleanup()
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        base = filename_hint.replace(".pdf", "")
        save_name = f"{base}_{ts}.pdf"
        buf.seek(0)
        (INVENTURA_DIR / save_name).write_bytes(buf.read())
        (INVENTURA_DIR / save_name.replace(".pdf", ".json")).write_text(
            json.dumps({"filename": save_name, "datum": datum, "items": items}, ensure_ascii=False),
            encoding="utf-8"
        )

        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={save_name}"})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/inventura-history")
async def inventura_history():
    """Vrne seznam shranjenih inventurnih PDF-jev."""
    inventura_cleanup()
    items = []
    try:
        for f in sorted(INVENTURA_DIR.glob("*.pdf"), key=lambda x: x.stat().st_mtime, reverse=True):
            stat = f.stat()
            items.append({
                "filename": f.name,
                "size": stat.st_size,
                "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
    except Exception as e:
        print(f"[inventura] history err: {e}")
    return {"items": items[:50]}


@app.get("/inventura-history-download/{filename}")
async def inventura_history_download(filename: str):
    if "/" in filename or "\\" in filename or ".." in filename:
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)
    f = INVENTURA_DIR / filename
    if not f.exists():
        return JSONResponse({"error": "Datoteka ne obstaja."}, status_code=404)
    return FileResponse(str(f), filename=filename, media_type="application/pdf")


@app.get("/inventura-history-load/{filename}")
async def inventura_history_load(filename: str):
    if "/" in filename or "\\" in filename or ".." in filename:
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)
    json_name = filename.replace(".pdf", ".json")
    f = INVENTURA_DIR / json_name
    if not f.exists():
        return JSONResponse({"error": "Podatki niso na voljo (star zapis)."}, status_code=404)
    return json.loads(f.read_text(encoding="utf-8"))


# ─── ODPREMA / AI ADDRESS VALIDATION ─────────────────────────────────────────

GEOAPIFY_KEY = os.environ.get("GEOAPIFY_API_KEY", "")
GOOGLE_MAPS_KEY = os.environ.get("GOOGLE_MAPS_API_KEY", "")
ODPREMA_DIR = DATA_DIR / "odprema"
ODPREMA_DIR.mkdir(exist_ok=True, parents=True)


@app.post("/odprema-google-validate")
async def odprema_google_validate(data: dict):
    """Google Address Validation API za problematične naslove."""
    if not GOOGLE_MAPS_KEY:
        return JSONResponse({"error": "GOOGLE_MAPS_API_KEY ni nastavljen."}, status_code=500)

    street = (data.get("street") or "").strip()
    city = (data.get("city") or "").strip()
    zip_code = (data.get("zip") or "").strip()
    order = data.get("order", "")

    address_line = f"{street}, {city}, {zip_code}, Bulgaria"

    payload = {
        "address": {
            "addressLines": [address_line],
            "regionCode": "BG",
            "languageCode": "bg",  # cirilica - Econt sprejme oboje
        },
        "enableUspsCass": False,
    }

    try:
        async with httpx.AsyncClient(timeout=10.0) as hc:
            resp = await hc.post(
                f"https://addressvalidation.googleapis.com/v1:validateAddress?key={GOOGLE_MAPS_KEY}",
                json=payload,
            )
            resp.raise_for_status()
            result = resp.json()

        verdict = result.get("result", {}).get("verdict", {})
        address = result.get("result", {}).get("address", {})
        components = address.get("addressComponents", [])

        # Izvleci komponente
        def get_comp(comp_type):
            for c in components:
                if comp_type in c.get("componentType", ""):
                    return c.get("componentName", {}).get("text", "")
            return ""

        fix_street = ""
        route = get_comp("route")
        street_nr = get_comp("street_number")
        subpremise = get_comp("subpremise")
        if route:
            fix_street = route
            if street_nr:
                fix_street += f" {street_nr}"
            if subpremise:
                fix_street += f", ap. {subpremise}"

        fix_city = get_comp("locality") or get_comp("administrative_area_level_2") or city
        fix_zip = get_comp("postal_code") or zip_code

        # Mesto ohranimo v latinici (original iz naše baze, ne Google cirilica)
        fix_city = city  # vedno originalno latinično ime mesta
        formatted = address.get("formattedAddress", "")

        validation_granularity = verdict.get("validationGranularity", "")
        has_unconfirmed = verdict.get("hasUnconfirmedComponents", False)

        if validation_granularity in ("PREMISE", "SUB_PREMISE", "ROUTE"):
            status = "CONFIRMED"
        elif validation_granularity in ("BLOCK", "PREMISE_PROXIMITY"):
            status = "PARTIALLY_CONFIRMED"
        else:
            status = "NOT_CONFIRMED"

        return {
            "order": order,
            "status": status,
            "fix_street": fix_street or street,
            "fix_city": fix_city,
            "fix_zip": fix_zip,
            "formatted": formatted,
            "granularity": validation_granularity,
            "has_unconfirmed": has_unconfirmed,
            "note": f"Google: {validation_granularity}",
        }

    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


def odprema_cleanup():
    """Zbriše zapise starejše od 90 dni."""
    try:
        cutoff = datetime.now().timestamp() - (90 * 86400)
        for f in ODPREMA_DIR.glob("*.json"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
    except Exception as e:
        print(f"[odprema] cleanup err: {e}")


@app.post("/odprema-save")
async def odprema_save(data: dict):
    """Shrani batch podatke pošiljk na disk (90 dni)."""
    try:
        odprema_cleanup()
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"odprema_{ts}.json"
        payload = {
            "saved_at": datetime.now().isoformat(),
            "filename": data.get("filename", ""),
            "total": data.get("total", 0),
            "rows": data.get("rows", []),
            "validation": data.get("validation", {}),
        }
        (ODPREMA_DIR / filename).write_text(
            json.dumps(payload, ensure_ascii=False), encoding="utf-8"
        )
        return {"ok": True, "filename": filename}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/odprema-history")
async def odprema_history_list():
    """Vrne seznam shranjenih batch zapisov (90 dni)."""
    odprema_cleanup()
    items = []
    try:
        for f in sorted(ODPREMA_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                items.append({
                    "filename": f.name,
                    "saved_at": data.get("saved_at", ""),
                    "original_file": data.get("filename", ""),
                    "total": data.get("total", 0),
                    "size": f.stat().st_size,
                })
            except:
                pass
    except Exception as e:
        print(f"[odprema] history err: {e}")
    return {"items": items}


@app.get("/odprema-history-load/{filename}")
async def odprema_history_load(filename: str):
    """Naloži shranjeni batch zapis."""
    if "/" in filename or "\\" in filename or ".." in filename:
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)
    f = ODPREMA_DIR / filename
    if not f.exists():
        return JSONResponse({"error": "Ni najdeno."}, status_code=404)
    return json.loads(f.read_text(encoding="utf-8"))


@app.delete("/odprema-history-delete/{filename}")
async def odprema_history_delete(filename: str):
    """Zbriše shranjeni batch zapis."""
    if "/" in filename or "\\" in filename or ".." in filename:
        return JSONResponse({"error": "Neveljavno ime."}, status_code=400)
    f = ODPREMA_DIR / filename
    if f.exists():
        f.unlink()
    return {"ok": True}


@app.get("/odprema-test")
async def odprema_test():
    """Debug: preveri Geoapify API ključ in naredi en testni klic."""
    if not GEOAPIFY_KEY:
        return {"ok": False, "error": "GEOAPIFY_API_KEY ni nastavljen", "key_len": 0}
    try:
        async with httpx.AsyncClient(timeout=10.0) as hc:
            resp = await hc.get(
                "https://api.geoapify.com/v1/geocode/search",
                params={"text": "Sofia, Bulgaria", "limit": 1, "apiKey": GEOAPIFY_KEY, "format": "json"}
            )
            return {
                "ok": resp.status_code == 200,
                "status_code": resp.status_code,
                "key_prefix": GEOAPIFY_KEY[:8] + "...",
                "response_preview": resp.text[:200]
            }
    except Exception as e:
        return {"ok": False, "error": str(e), "key_prefix": GEOAPIFY_KEY[:8] + "..."}


@app.post("/odprema-validate")
async def odprema_validate(data: dict):
    """
    Normalizira BG naslove z Claude AI — razčleni, popravi format, standardizira za Econt.
    """
    addresses = data.get("addresses", [])
    if not addresses:
        return JSONResponse({"error": "Ni naslovov za validacijo."}, status_code=400)

    async def normalize_one(addr: dict) -> dict:
        order = addr.get("order", "")
        zip_code = (addr.get("zip") or "").strip()
        city = (addr.get("city") or "").strip()
        street = (addr.get("street") or "").strip()
        street_nr = (addr.get("streetNr") or "").strip()

        # ── Geo lookup — poišči mesto in ulice iz Econt baze ──────────────────
        city_entry = econt_lookup_city(zip_code, city)
        streets_context = ""
        city_info = ""
        if city_entry:
            cid = city_entry.get("id")
            name_bg = city_entry.get("name_bg", "")
            name_en = city_entry.get("name_en", "")
            city_zip = city_entry.get("zip", "")
            city_info = f"\nVERIFICIRANO MESTO (Econt baza): {name_en} / {name_bg}, ZIP={city_zip} (cityID={cid})"
            streets_context = econt_get_streets_context(cid, street)
            if streets_context:
                streets_context = f"\nULICE TEGA MESTA (Econt baza — predlagaj SAMO iz tega seznama):\n{streets_context}"
        else:
            city_info = "\n⚠ Mesto ni najdeno v Econt bazi — previdno z validacijo."

        prompt = f"""Si ekspert za bolgarske poštne naslove za Econt Express dostavo. Normaliziraj vhodni naslov v standardni format ZA ECONT VMESNIK (latinica).

VHODNI NASLOV:
- Ulica/naslov: {street}
- Hišna številka: {street_nr}
- Mesto: {city}
- ZIP: {zip_code}
{city_info}{streets_context}

BOLGARSKE OKRAJŠAVE:
- ul. = ulitsa (ulica), bul. = bulevard
- zh.k. / jk / kv. / zk. / z.k. = zhilishten kompleks (stanovanjska četrt)
- bl. = blok, vh. = vhod, et. = etazh (nadstropje), ap. = apartament
- s. / selo = vas, gr. / grad = mesto
- ofis ekont / ekont / paketomat = Econt pisarna

KRITIČNA PRAVILA (po prioriteti):

1. LATINICA OBVEZNA — vse vrni v latinici, tudi ulice (Econt vmesnik je v latinici)
   **IZJEMA: fix_city VEDNO v latinici** — nikoli ne prevajaj ali transliteriraj imena mesta v cirilico, tudi če Google ali drug vir vrne cirilico. Mesto ohrani točno tako kot je napisano v originalnem naslovu ali v standardni latinični obliki.

2. HIŠNA ŠTEVILKA — NIKOLI NE BRIŠI!
   Hišno številko VEDNO ohrani v fix_street — stranka jo je vpisala in ve kje stanuje.
   Odstrani SAMO če je očiten placeholder: "nn", "NN", "N/A", "n/a" — to so sistemske vrednosti brez pomena.
   VSE ostalo ohrani: številke, črke, duplikate (65 65), 0, 36b, 4A itd.
   Če je hišna številka v polju streetNr, jo dodaj na konec fix_street.
   Primer: ulica="Vasil Levski", streetNr="36b" → fix_street="ul. Vasil Levski 36b"

3. ULICA IN MESTO — POMEMBNO!
   Če imaš seznam ulic tega mesta (zgoraj), ga uporabi za validacijo in popravke pravopisnih napak.
   NIKOLI ne predlagaj ulice iz drugega mesta.
   AMPAK: Econt baza ulic ni popolna — manjkajo nove ulice, manjše vasi, četrti brez ulic.
   Če ulica ni v seznamu, jo VSEENO potrdi (status FIXED ali OK) če je naslov smiselno formatiran.
   Status UNCLEAR nastavi SAMO če naslov je res nerazumljiv ali manjka ulica/hišna številka — NE samo zato ker ulice ni v bazi.

4. ZIP EKSTRAKCIJA — če je ZIP v polju ulice ali mesta, ga prestavi v fix_zip. Popravi ZIP SAMO če je očitno napačen (vsebuje črke, ni 4 cifre, ali je ZIP drugega mesta). Če si negotov → pusti originalni ZIP.

5. SOFIJSKE ČETRTI IN ZIP — poznaj pravilne ZIP-e za sofijske četrti:
   zh.k. Lulin → 1343, zh.k. Mladost → 1750/1784, zh.k. Lyulin → 1343
   zh.k. Druzhba → 1582, zh.k. Nadezhda → 1220, zh.k. Ovcha Kupel → 1618
   zh.k. Bukston/Bakston → 1618, zh.k. Borovo → 1680, zh.k. Lozenets → 1164
   zh.k. Dianabad → 1172, zh.k. Manastirski livadi → 1404
   Če stranka piše ZIP 1000 za četrt → popravi na pravilen ZIP četrti!

6. ULICA = IME MESTA → ni ulice
   Če je ime ulice enako imenu mesta (npr. "ul. Kardzhali" v mestu Kardzhali) → fix_street=""
   Če je naslov samo ime mesta brez ulice (npr. "Cerven bryag 3") → fix_street="", status UNCLEAR

7. ULICA V VEČ ČETRTIH → opozori
   Če ulica obstaja v več četrtih Sofije → status UNCLEAR, note mora vsebovati "Quarter needed: X, Y, Z"
   Znani primeri: ul. Elin Pelin (Lozenets/Dragalevtsi/Pancharevo)

8. ECONT OFFICE V LATINICI
   Vse Econt office naslove vrni v latinici:
   "Ofis Ekont" → "Econt office [mesto]"
   Primer: "Ekont Bokar, zh.k. Manastirski livadi" → "Econt office Bokar, zh.k. Manastirski livadi"

9. IME MESTA V ULICI → odstrani
   "Sofia bul. Bulgaria 102" → fix_street="bul. Bulgaria 102", fix_city="Sofia"
   "Bansko Glazne 6" → fix_street="ul. Glazne 6", fix_city="Bansko"

10. PODVOJENI DELI NASLOVA → ohrani, ne briši
   Podvojena hišna številka (npr. "65 65", "9 9") je verjetno resnična — ohrani kot je.
   Odstrani duplikat SAMO pri okrajšavah formata (npr. "bl. 5 bl. 5" → "bl. 5").
   "Bl. 503 vh.A ap 65 et 11" → "bl. 503, vh. A, et. 11, ap. 65"

11. TIPKARSKE NAPAKE V IMENIH MEST:
   Vitosa → Vitosha, Blgaria → Bulgaria, Sofiq → Sofia
   Plovdic → Plovdiv, Kustendil → Kyustendil, Vraca → Vratsa
   Carevo → Tsarevo, Dupnica → Dupnitsa, Krdzali → Kardzhali
   Trstenik → Trastenik, Satovca → Satovcha

12. MALE VASI BREZ ULICE → status UNCLEAR z opombo
    Če je naslov samo ime vasi brez ulice in hišne številke → UNCLEAR, note="No street - recommend Econt office [najbližje mesto]"

PRIMERI (few-shot):
Input: ulica="Sofia bul.Vitosa 38", mesto="Sofia", ZIP="1000"
Output: {{"status":"FIXED","fix_street":"bul. Vitosha 38","fix_city":"Sofia","fix_zip":"1000","note":"Removed city from street, fixed Vitosa→Vitosha"}}

Input: ulica="Vasil Levski", streetNr="36b", mesto="Starcevo", ZIP="4987"
Output: {{"status":"FIXED","fix_street":"ul. Vasil Levski 36b","fix_city":"Presoka","fix_zip":"4987","note":"Added ul. prefix, kept house number 36b, corrected city name"}}

Input: ulica="Zk.Borovo bl.5 vh B ET.6 AP.34 34", mesto="Sofia", ZIP="1000"
Output: {{"status":"FIXED","fix_street":"zh.k. Borovo, bl. 5, vh. B, et. 6, ap. 34","fix_city":"Sofia","fix_zip":"1680","note":"Fixed format, removed duplicate 34, corrected ZIP for Borovo"}}

Input: ulica="Lulin5 bl540vhb 65etz11", mesto="Sofia", ZIP="1000"
Output: {{"status":"FIXED","fix_street":"zh.k. Lulin 5, bl. 540, vh. B, et. 11, ap. 65","fix_city":"Sofia","fix_zip":"1343","note":"Parsed compressed format, corrected ZIP for Lulin"}}

Input: ulica="ul. Elin Pelin 18", mesto="Sofia", ZIP="1000"
Output: {{"status":"UNCLEAR","fix_street":"ul. Elin Pelin 18","fix_city":"Sofia","fix_zip":"1000","note":"Quarter needed: Lozenets (1164), Dragalevtsi (1415), Pancharevo (1137)"}}

Input: ulica="Cerven brag 3", mesto="Cerven bryag", ZIP="5980"
Output: {{"status":"UNCLEAR","fix_street":"","fix_city":"Cherven Bryag","fix_zip":"5980","note":"No street provided - recommend Econt office Cherven Bryag, ul. Hristo Botev 2"}}

Input: ulica="Sofia Manastirski livadi Ekont Bokar", mesto="Sofia", ZIP="1000"
Output: {{"status":"ECONT_OFFICE","fix_street":"Econt office Bokar, zh.k. Manastirski livadi","fix_city":"Sofia","fix_zip":"1404","note":"Econt office Bokar in Manastirski livadi"}}

Vrni SAMO JSON, brez razlag:
{{"status": "OK|FIXED|ECONT_OFFICE|UNCLEAR", "fix_street": "...", "fix_city": "...", "fix_zip": "4-mestni ZIP", "note": "..."}}"""

        loop = asyncio.get_event_loop()
        try:
            msg = await loop.run_in_executor(None, lambda: client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=400,
                messages=[{"role": "user", "content": prompt}]
            ))
            text = msg.content[0].text.strip()
            text = re.sub(r'```json\s*', '', text)
            text = re.sub(r'```\s*', '', text).strip()
            result = json.loads(text)
            
            status_map = {
                "OK": "CONFIRMED",
                "FIXED": "PARTIALLY_CONFIRMED", 
                "ECONT_OFFICE": "ECONT_DEPO",
                "UNCLEAR": "NOT_CONFIRMED"
            }
            
            fix_street_raw = result.get("fix_street", street)

            # Ekstrahiraj hišno številko iz fix_street če je streetNr nn/N/A
            fix_nr_out = street_nr
            is_placeholder = street_nr.lower().strip() in ("nn", "n/a", "")
            if is_placeholder and fix_street_raw:
                if not re.search(r'econt\s*office', fix_street_raw, re.IGNORECASE):
                    nr_match = re.search(r'^(.*?)[\s,]+(\d+[A-Za-z]?(?:\s+\d+[A-Za-z]?)?)$', fix_street_raw.strip())
                    if nr_match:
                        fix_nr_out = nr_match.group(2).strip()
                        fix_street_raw = nr_match.group(1).strip().rstrip(',')

            # Fallback: če fix_nr še vedno nn, poišči številko v ORIGINALNEM naslovu
            # (AI jo je morda pozabil prenesti v fix_street)
            if fix_nr_out.lower().strip() in ("nn", "n/a", "") and street:
                if not re.search(r'econt\s*office', fix_street_raw, re.IGNORECASE):
                    nr_fallback = re.search(r'\b(\d+[A-Za-z]?)\s*$', street.strip())
                    if nr_fallback:
                        fix_nr_out = nr_fallback.group(1)

            return {
                "order": order,
                "status": status_map.get(result.get("status", "UNCLEAR"), "NOT_CONFIRMED"),
                "confidence": 1.0 if result.get("status") == "OK" else 0.7 if result.get("status") == "FIXED" else 0.3,
                "formatted": f"{fix_street_raw}, {result.get('fix_city','')}, {result.get('fix_zip','')}",
                "fix_street": fix_street_raw,
                "fix_nr": fix_nr_out,
                "fix_city": result.get("fix_city", city),
                "fix_zip": result.get("fix_zip", zip_code),
                "note": result.get("note", ""),
                "original": {"zip": zip_code, "city": city, "street": street, "streetNr": street_nr},
            }
        except Exception as e:
            print(f"[odprema] AI error for {order}: {e}")
            return {
                "order": order, "status": "ERROR", "confidence": 0,
                "formatted": "", "fix_street": street, "fix_nr": street_nr,
                "fix_city": city, "fix_zip": zip_code,
                "error": str(e),
                "original": {"zip": zip_code, "city": city, "street": street, "streetNr": street_nr},
            }

    # Vzporedno — max 5 hkrati (Claude rate limit)
    semaphore = asyncio.Semaphore(5)

    async def limited(addr):
        async with semaphore:
            return await normalize_one(addr)

    results = await asyncio.gather(*[limited(addr) for addr in addresses])
    return {"results": list(results), "total": len(results)}

# ─── ECONT OFFICES CACHE ──────────────────────────────────────────────────────

ECONT_OFFICES_CACHE: list = []
ECONT_OFFICES_LAST_FETCH: float = 0
ECONT_OFFICES_TTL = 86400  # 24h cache

ECONT_API_URL = "https://ee.econt.com/services"
ECONT_DEMO_URL = "https://demo.econt.com/ee/services"
ECONT_USER = os.environ.get("ECONT_USER", "iasp-dev")
ECONT_PASS = os.environ.get("ECONT_PASS", "1Asp-dev")


async def econt_fetch_offices() -> list:
    """Pobere vse BG Econt office-e z API-ja, cachira 24h."""
    global ECONT_OFFICES_CACHE, ECONT_OFFICES_LAST_FETCH
    import time

    now = time.time()
    if ECONT_OFFICES_CACHE and (now - ECONT_OFFICES_LAST_FETCH) < ECONT_OFFICES_TTL:
        return ECONT_OFFICES_CACHE

    # Poskusi produkcijo, potem demo
    for base_url in [ECONT_API_URL, ECONT_DEMO_URL]:
        try:
            async with httpx.AsyncClient(timeout=20.0) as hc:
                resp = await hc.post(
                    f"{base_url}/Nomenclatures/NomenclaturesService.getOffices.json",
                    json={"countryCode": "BG"},
                    auth=(ECONT_USER, ECONT_PASS)
                )
                if resp.status_code == 200:
                    data = resp.json()
                    offices = data.get("offices", [])
                    if offices:
                        ECONT_OFFICES_CACHE = offices
                        ECONT_OFFICES_LAST_FETCH = now
                        print(f"[econt] Naloženih {len(offices)} BG officeov iz {base_url}")
                        return offices
        except Exception as e:
            print(f"[econt] Napaka pri {base_url}: {e}")

    return ECONT_OFFICES_CACHE  # Vrni star cache če API ne dela


def econt_find_nearest_office(offices: list, zip_code: str, city_name: str) -> dict | None:
    """Poišče najbližji Econt office glede na ZIP kodo ali ime mesta."""
    if not offices:
        return None

    zip_clean = (zip_code or "").strip()
    city_lower = (city_name or "").lower().strip()

    # 1. Točno ujemanje ZIP
    for o in offices:
        addr = o.get("address") or {}
        city = addr.get("city") or {}
        office_zip = str(city.get("postCode") or "").strip()
        if office_zip and office_zip == zip_clean:
            return o

    # 2. Ujemanje mesta (case-insensitive)
    if city_lower:
        for o in offices:
            addr = o.get("address") or {}
            city = addr.get("city") or {}
            office_city = (city.get("name") or "").lower().strip()
            office_city_en = (city.get("nameEn") or "").lower().strip()
            if city_lower in office_city or city_lower in office_city_en:
                return o

    # 3. ZIP prefix match (prvih 2 cifri = regija)
    if len(zip_clean) >= 2:
        prefix = zip_clean[:2]
        for o in offices:
            addr = o.get("address") or {}
            city = addr.get("city") or {}
            office_zip = str(city.get("postCode") or "").strip()
            if office_zip.startswith(prefix):
                return o

    return None


def econt_office_to_address(office: dict) -> dict:
    """Pretvori Econt office objekt v naslovne polje."""
    addr = office.get("address") or {}
    city = addr.get("city") or {}
    return {
        "name": office.get("name") or office.get("nameEn") or "",
        "street": addr.get("street") or addr.get("fullAddress") or "",
        "num": addr.get("num") or "",
        "city": city.get("name") or city.get("nameEn") or "",
        "zip": str(city.get("postCode") or ""),
        "full": f"{office.get('name','')} — {addr.get('fullAddress') or addr.get('street','')} {addr.get('num','')}, {city.get('name','')}".strip(),
    }


@app.get("/econt-offices")
async def get_econt_offices():
    """Vrne seznam vseh BG Econt officeov (cachiran 24h)."""
    offices = await econt_fetch_offices()
    return {"ok": True, "total": len(offices), "offices": offices}


@app.get("/econt-valid-zips")
async def get_econt_valid_zips():
    """Vrne set vseh veljavnih ZIP kod iz Econt geo baze. ZIP ki NI tukaj = suspended."""
    if not ECONT_GEO:
        return {"ok": False, "error": "econt_geo.json not loaded", "zips": []}
    zips = list(ECONT_GEO.get("zip_to_city_id", {}).keys())
    return {"ok": True, "total": len(zips), "zips": zips}


@app.post("/econt-nearest-office")
async def get_nearest_office(data: dict):
    """Za dani ZIP/mesto vrne najbližji Econt office z naslovom."""
    zip_code = data.get("zip", "")
    city = data.get("city", "")

    offices = await econt_fetch_offices()
    if not offices:
        return JSONResponse({"error": "Econt API ni dosegljiv, ni officeov v cacheju."}, status_code=503)

    office = econt_find_nearest_office(offices, zip_code, city)
    if not office:
        return {"ok": False, "office": None, "message": "Ni najden office za ta ZIP/mesto"}

    return {"ok": True, "office": econt_office_to_address(office), "raw": office}


@app.post("/odprema-econt-check")
async def odprema_econt_check(data: dict):
    """
    Double-check naslova direktno pri Econt API — preveri ali mesto + ulica obstajata.
    Klic gre iz Render serverja (whitelist IP). 
    Input: {"zip": "1000", "city": "Sofia", "street": "bul. Vitosha 38"}
    """
    zip_code = (data.get("zip") or "").strip()
    city_name = (data.get("city") or "").strip()
    street = (data.get("street") or "").strip()
    order = data.get("order", "")

    result = {"order": order, "zip": zip_code, "city": city_name}

    async with httpx.AsyncClient(timeout=15.0) as hc:
        for base_url in [ECONT_API_URL, ECONT_DEMO_URL]:
            try:
                # 1. getCities — preveri ali mesto obstaja
                r = await hc.post(
                    f"{base_url}/Nomenclatures/NomenclaturesService.getCities.json",
                    json={"countryCode": "BG", "name": city_name},
                    auth=(ECONT_USER, ECONT_PASS)
                )
                if r.status_code != 200:
                    continue

                cities = r.json().get("cities", [])
                city_match = None
                for c in cities:
                    if str(c.get("postCode", "")) == zip_code or \
                       (c.get("nameEn", "").lower() == city_name.lower()) or \
                       (c.get("name", "").lower() == city_name.lower()):
                        city_match = c
                        break

                result["city_found"] = bool(city_match)
                result["city_econt"] = city_match.get("nameEn", "") if city_match else ""
                result["city_zip"] = str(city_match.get("postCode", "")) if city_match else ""

                if not city_match:
                    result["status"] = "CITY_NOT_FOUND"
                    result["note"] = f"Mesto '{city_name}' ni v Econt bazi"
                    return result

                # 2. getStreets — preveri ali ulica obstaja v tem mestu
                if street:
                    # Ekstrahiraj samo ime ulice brez številke
                    street_name = re.sub(r'\b\d+\b.*$', '', street).strip().rstrip(',').strip()
                    street_name = re.sub(r'^(ul\.|bul\.|zh\.k\.|kv\.)\s*', '', street_name, flags=re.IGNORECASE).strip()

                    r2 = await hc.post(
                        f"{base_url}/Nomenclatures/NomenclaturesService.getStreets.json",
                        json={"cityID": city_match.get("id"), "name": street_name},
                        auth=(ECONT_USER, ECONT_PASS)
                    )
                    if r2.status_code == 200:
                        streets = r2.json().get("streets", [])
                        result["street_found"] = bool(streets)
                        result["street_matches"] = [s.get("nameEn", s.get("name", "")) for s in streets[:3]]
                        if streets:
                            result["status"] = "OK"
                            result["note"] = f"Mesto in ulica potrjena pri Econt"
                        else:
                            result["status"] = "STREET_NOT_FOUND"
                            result["note"] = f"Ulica '{street_name}' ni v Econt bazi za {city_name}"
                    else:
                        result["street_found"] = None
                        result["status"] = "OK"
                        result["note"] = "Mesto potrjeno, ulica ni preverjena"
                else:
                    result["street_found"] = None
                    result["status"] = "OK"
                    result["note"] = "Mesto potrjeno (ni ulice za preverjanje)"

                return result

            except Exception as e:
                result["error"] = str(e)
                continue

    result["status"] = "API_UNAVAILABLE"
    result["note"] = "Econt API ni dosegljiv (IP whitelist?)"
    return result



async def odprema_resolve_suspended(data: dict):
    """
    Za seznam suspended naslovov poišče najbližji Econt office.
    Input: [{"order": "...", "zip": "...", "city": "..."}, ...]
    """
    addresses = data.get("addresses", [])
    if not addresses:
        return JSONResponse({"error": "Ni naslovov."}, status_code=400)

    offices = await econt_fetch_offices()

    results = []
    for addr in addresses:
        order = addr.get("order", "")
        zip_code = addr.get("zip", "")
        city = addr.get("city", "")

        office = econt_find_nearest_office(offices, zip_code, city)
        if office:
            office_addr = econt_office_to_address(office)
            results.append({
                "order": order,
                "found": True,
                "office_name": office_addr["name"],
                "office_street": office_addr["street"],
                "office_num": office_addr["num"],
                "office_city": office_addr["city"],
                "office_zip": office_addr["zip"],
                "office_full": office_addr["full"],
            })
        else:
            results.append({
                "order": order,
                "found": False,
                "office_name": "",
                "office_street": "",
                "office_num": "",
                "office_city": "",
                "office_zip": zip_code,
                "office_full": "Ni najden — preveri ročno",
            })

    return {"ok": True, "results": results, "offices_available": len(offices)}


# ─── EMAIL OBVESTILA ──────────────────────────────────────────────────────────

SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASS = os.environ.get("SMTP_PASS", "")
SMTP_FROM = os.environ.get("SMTP_FROM", "")


@app.post("/odprema-send-emails")
async def odprema_send_emails(data: dict):
    """
    Pošlje email obvestila strankam s suspended ZIP naslovi.
    Input: { "shipments": [{ "order": "...", "name": "...", "email": "...", "office": "...", "orig_city": "..." }, ...] }
    """
    if not all([SMTP_HOST, SMTP_USER, SMTP_PASS, SMTP_FROM]):
        return JSONResponse({"error": "SMTP ni konfiguriran."}, status_code=500)

    shipments = data.get("shipments", [])
    if not shipments:
        return JSONResponse({"error": "Ni pošiljk za obveščanje."}, status_code=400)

    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    sent = []
    failed = []

    for s in shipments:
        email = (s.get("email") or "").strip()
        if not email or "@" not in email:
            failed.append({"order": s.get("order"), "reason": "Ni e-mail naslova"})
            continue

        name = s.get("name", "")
        order = s.get("order", "")
        office = s.get("office", "najbližji Econt office")
        orig_city = s.get("orig_city", "")

        # Email v bolgarščini + angleščini
        subject = f"Вашата поръчка {order} - промяна на адрес за доставка / Your order {order} - delivery address change"

        html = f"""
<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px">
  <img src="https://maaarket.si/wp-content/uploads/2023/03/logo.png" style="height:40px;margin-bottom:20px" alt="Maaarket">
  
  <h2 style="color:#1a1a2e">Уважаеми {name},</h2>
  
  <p>Вашата поръчка с номер <strong>{order}</strong> не може да бъде доставена до посочения адрес 
  (<strong>{orig_city}</strong>), тъй като куриерската фирма <strong>Econt Express</strong> временно е 
  преустановила доставките до вашия район.</p>
  
  <div style="background:#fff3cd;border:1px solid #ffc107;border-radius:8px;padding:16px;margin:20px 0">
    <strong>📦 Вашата поръчка ще ви чака на:</strong><br><br>
    <span style="font-size:16px;color:#1a1a2e">{office}</span>
  </div>
  
  <p>Моля, посетете горепосочения офис с <strong>личен документ</strong> за получаване на поръчката.</p>
  
  <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
  
  <h3 style="color:#1a1a2e">Dear {name},</h3>
  
  <p>Your order <strong>{order}</strong> cannot be delivered to the specified address 
  (<strong>{orig_city}</strong>) as <strong>Econt Express</strong> has temporarily suspended 
  deliveries to your area.</p>
  
  <div style="background:#d1ecf1;border:1px solid #bee5eb;border-radius:8px;padding:16px;margin:20px 0">
    <strong>📦 Your order will be waiting at:</strong><br><br>
    <span style="font-size:16px;color:#1a1a2e">{office}</span>
  </div>
  
  <p>Please visit the above office with a <strong>valid ID</strong> to collect your order.</p>
  
  <p style="color:#666;font-size:12px;margin-top:30px">
    Maaarket.eu | За въпроси / For questions: <a href="mailto:info@maaarket.bg">info@maaarket.bg</a>
  </p>
</div>
"""

        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = SMTP_FROM
            msg["To"] = email
            msg["Reply-To"] = "info@maaarket.bg"
            msg.attach(MIMEText(html, "html", "utf-8"))

            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.sendmail(SMTP_FROM, email, msg.as_string())

            sent.append({"order": order, "email": email})
            print(f"[email] Sent to {email} for order {order}")

        except Exception as e:
            print(f"[email] Failed {email} for {order}: {e}")
            failed.append({"order": order, "email": email, "reason": str(e)})

    return {
        "ok": True,
        "sent": len(sent),
        "failed": len(failed),
        "sent_list": sent,
        "failed_list": failed,
    }


# ─── KAYAKO CLASSIC REST API — KB IMPORT ─────────────────────────────────────
import hashlib
import hmac
import base64
import random

KAYAKO_API_URL = os.environ.get("KAYAKO_API_URL", "https://support.silux.si/api/index.php")
KAYAKO_API_KEY = os.environ.get("KAYAKO_API_KEY", "")
KAYAKO_SECRET  = os.environ.get("KAYAKO_SECRET", "")

KAYAKO_DEPT = {
    "silux":    1,
    "maaarket": 27,
}

KB_FILES = {
    "silux":    DATA_DIR / "kb_silux.json",
    "maaarket": DATA_DIR / "kb_maaarket.json",
}

MACROS_FILE = Path("macros_maaarket.json")

def _kayako_build_url(path: str) -> str:
    """Zgradi popoln Kayako URL z auth - signature je quote_plus encoded."""
    from urllib.parse import quote_plus
    salt = str(random.randint(1000000000, 9999999999))
    raw_sig = hmac.new(
        key=KAYAKO_SECRET.encode("utf-8"),
        msg=salt.encode("utf-8"),
        digestmod=hashlib.sha256
    ).digest()
    signature = quote_plus(base64.b64encode(raw_sig).decode("utf-8"))
    return (
        f"{KAYAKO_API_URL}"
        f"?e={path}"
        f"&apikey={KAYAKO_API_KEY}"
        f"&salt={salt}"
        f"&signature={signature}"
    )

# backwards compat
def _kayako_url(path: str) -> str:
    return _kayako_build_url(path)

def _xml_text(el, tag: str, default: str = "") -> str:
    node = el.find(tag)
    return (node.text or default).strip() if node is not None and node.text else default

def _parse_ticket_xml(xml_text: str) -> list[dict]:
    """Razčleni XML odgovor za ticket/tickete → lista dict."""
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return []
    tickets = []
    for t in root.findall("ticket"):
        ticket_id  = t.get("id", "")
        subject    = _xml_text(t, "subject")
        dept_id    = _xml_text(t, "departmentid")
        status_id  = _xml_text(t, "statusid")
        created    = _xml_text(t, "creationtime")
        email      = _xml_text(t, "email")
        fullname   = _xml_text(t, "fullname")
        replies    = _xml_text(t, "totalreplies", "0")
        # posti (konverzacija)
        posts = []
        for p in t.findall(".//post"):
            creator   = p.get("creator", _xml_text(p, "creator", "2"))
            staff_id  = _xml_text(p, "staffid", "0")
            p_fullname  = _xml_text(p, "fullname")
            contents_node = p.find("contents")
            contents = ""
            if contents_node is not None:
                contents = (contents_node.text or "").strip()
            # creator=1 = staff, creator=2 = user/stranka
            role = "staff" if (staff_id != "0" or str(creator) == "1") else "customer"
            if contents:
                posts.append({"role": role, "name": p_fullname, "text": contents})
        tickets.append({
            "id": ticket_id,
            "subject": subject,
            "dept_id": dept_id,
            "status_id": status_id,
            "created": created,
            "email": email,
            "fullname": fullname,
            "replies": replies,
            "posts": posts,
        })
    return tickets

async def _fetch_tickets_batch(client_h: httpx.AsyncClient, dept_id: int, start: int, count: int = 50) -> list[dict]:
    """Potegne batch ticketov iz Kayako (samo header info, brez postov)."""
    path = f"/Tickets/Ticket/ListAll/{dept_id}/3/-1/-1/{count}/{start}/ticketid/DESC"
    url = _kayako_build_url(path)
    try:
        r = await client_h.get(url, timeout=30)
        print(f"[kayako] ListAll HTTP {r.status_code} | {url[:80]}")
        if r.status_code != 200:
            return []
        return _parse_ticket_xml(r.text)
    except Exception as e:
        print(f"[kayako] ListAll error: {e}")
        return []

async def _fetch_ticket_posts(client_h: httpx.AsyncClient, ticket_id: str) -> list[dict]:
    """Potegne posamezen ticket z vsemi posti."""
    path = f"/Tickets/Ticket/{ticket_id}"
    url = _kayako_build_url(path)
    try:
        r = await client_h.get(url, timeout=20)
        if r.status_code != 200:
            return []
        tickets = _parse_ticket_xml(r.text)
        return tickets[0]["posts"] if tickets else []
    except Exception as e:
        print(f"[kayako] Ticket {ticket_id} error: {e}")
        return []

def _tickets_to_kb(tickets_with_posts: list[dict]) -> dict:
    """
    Pretvori surove tickete v knowledge base format.
    Shrani samo tickete ki imajo vsaj 1 staff odgovor.
    Format: { "qa_pairs": [ {subject, question, answer, count:1} ] }
    """
    qa_pairs = []
    for t in tickets_with_posts:
        subject = t.get("subject", "")
        posts   = t.get("posts", [])
        if not posts:
            continue
        # Združi customer posti v eno vprašanje, staff posti v en odgovor
        customer_texts = [p["text"] for p in posts if p["role"] == "customer"]
        staff_texts    = [p["text"] for p in posts if p["role"] == "staff"]
        if not customer_texts or not staff_texts:
            continue
        question = " | ".join(customer_texts[:3])[:800]   # max 800 znakov
        answer   = staff_texts[-1][:800]                   # zadnji staff odgovor, max 800
        qa_pairs.append({
            "subject":  subject[:150],
            "question": question,
            "answer":   answer,
            "count":    1,
        })
    return {"qa_pairs": qa_pairs, "updated": datetime.now(timezone.utc).isoformat()}

# ─── MAKRI (KAYAKO MACROS) ────────────────────────────────────────────────────

@app.get("/macros")
async def get_macros():
    """Vrne vse makre iz /data/macros_maaarket.json"""
    if MACROS_FILE.exists():
        try:
            return json.loads(MACROS_FILE.read_text(encoding="utf-8"))
        except:
            pass
    return {"macros": [], "updated": ""}

@app.post("/macros-save")
async def save_macros(data: dict):
    """Shrani makre (za admin update)"""
    try:
        MACROS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "total": len(data.get("macros", []))}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/kayako-test")
async def kayako_test(brand: str = "silux"):
    """Test Kayako povezave — potegne samo 3 tickete da preverimo auth."""
    if not KAYAKO_API_KEY or not KAYAKO_SECRET:
        return {"ok": False, "error": "KAYAKO_API_KEY ali KAYAKO_SECRET nista nastavljena!"}
    dept_id = KAYAKO_DEPT.get(brand, 1)
    # Debug — pokaži točen URL ki ga kličemo
    path = f"/Tickets/Ticket/ListAll/{dept_id}/3/-1/-1/3/0/ticketid/DESC"
    debug_url = _kayako_build_url(path)
    print(f"[kayako] DEBUG URL: {debug_url}")
    async with httpx.AsyncClient() as h:
        tickets = await _fetch_tickets_batch(h, dept_id, start=0, count=3)
    if not tickets:
        return {
            "ok": False,
            "error": "Ni ticketov ali napaka pri povezavi",
            "debug_url_base": f"{KAYAKO_API_URL}?e={path}",
            "kayako_api_url_env": KAYAKO_API_URL,
        }
    return {
        "ok": True,
        "brand": brand,
        "dept_id": dept_id,
        "tickets_found": len(tickets),
        "sample": [{"id": t["id"], "subject": t["subject"], "posts_count": len(t["posts"])} for t in tickets],
    }

@app.get("/kayako-import")
async def kayako_import(
    brand: str = "maaarket",
    max_tickets: int = 500,
    batch_size: int = 50,
):
    """
    Importa tickete iz Kayako → shrani v /data/kb_{brand}.json
    Parametri:
      brand       = maaarket | silux
      max_tickets = koliko ticketov max (default 500)
      batch_size  = po koliko naenkrat (default 50, max 100)
    """
    if not KAYAKO_API_KEY or not KAYAKO_SECRET:
        return {"ok": False, "error": "KAYAKO_API_KEY ali KAYAKO_SECRET nista nastavljena v Render env vars!"}
    if brand not in KAYAKO_DEPT:
        return {"ok": False, "error": f"Neznan brand: {brand}"}

    dept_id    = KAYAKO_DEPT[brand]
    kb_file    = KB_FILES[brand]
    batch_size = min(batch_size, 100)

    # Naloži obstoječo KB če obstaja
    existing_kb = {"qa_pairs": [], "updated": ""}
    if kb_file.exists():
        try:
            existing_kb = json.loads(kb_file.read_text(encoding="utf-8"))
        except:
            pass
    existing_ids = {qa.get("ticket_id", "") for qa in existing_kb.get("qa_pairs", [])}

    all_tickets = []
    start = 0
    print(f"[kayako] Začenjam import za {brand} (dept={dept_id}, max={max_tickets})")

    async with httpx.AsyncClient() as h:
        # Faza 1: potegni liste ticketov
        while start < max_tickets:
            batch = await _fetch_tickets_batch(h, dept_id, start=start, count=batch_size)
            if not batch:
                break
            all_tickets.extend(batch)
            print(f"[kayako] ListAll: {start}–{start+len(batch)} ({len(batch)} ticketov)")
            if len(batch) < batch_size:
                break  # zadnja stran
            start += batch_size
            await asyncio.sleep(0.2)  # rate limit

        # Faza 2: za vsak ticket potegni posti (samo novi)
        new_count = 0
        for t in all_tickets:
            tid = t["id"]
            if tid in existing_ids:
                continue  # že imamo
            posts = await _fetch_ticket_posts(h, tid)
            t["posts"] = posts
            await asyncio.sleep(0.1)  # rate limit
            new_count += 1
            if new_count % 50 == 0:
                print(f"[kayako] Posti: {new_count}/{len(all_tickets)}")

    # Faza 3: pretvori v KB format
    new_qa = _tickets_to_kb([t for t in all_tickets if t["id"] not in existing_ids])

    # Dodaj ticket_id za deduplication
    for i, qa in enumerate(new_qa["qa_pairs"]):
        if i < len(all_tickets):
            qa["ticket_id"] = all_tickets[i]["id"]

    # Združi z obstoječim
    merged = existing_kb.get("qa_pairs", []) + new_qa["qa_pairs"]
    # Dedupliciraj po subject+question
    seen = set()
    deduped = []
    for qa in merged:
        key = qa["subject"] + "|" + qa["question"][:100]
        if key not in seen:
            seen.add(key)
            deduped.append(qa)

    final_kb = {
        "qa_pairs":   deduped,
        "updated":    datetime.now(timezone.utc).isoformat(),
        "brand":      brand,
        "total":      len(deduped),
    }
    kb_file.write_text(json.dumps(final_kb, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"[kayako] Import končan: {len(new_qa['qa_pairs'])} novih, {len(deduped)} skupaj")
    return {
        "ok":         True,
        "brand":      brand,
        "new":        len(new_qa["qa_pairs"]),
        "total":      len(deduped),
        "skipped":    len(all_tickets) - new_count,
        "kb_file":    str(kb_file),
    }

@app.get("/kayako-kb-stats")
async def kayako_kb_stats(brand: str = "maaarket"):
    """Vrne statistiko knowledge base za brand."""
    kb_file = KB_FILES.get(brand)
    if not kb_file or not kb_file.exists():
        return {"ok": True, "brand": brand, "total": 0, "updated": None, "sample": []}
    try:
        kb = json.loads(kb_file.read_text(encoding="utf-8"))
        pairs = kb.get("qa_pairs", [])
        return {
            "ok":      True,
            "brand":   brand,
            "total":   len(pairs),
            "updated": kb.get("updated"),
            "sample":  pairs[:5],
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ─── BETA ANALIZA — META ADS EXPORT HISTORY ──────────────────────────────────
BETA_DIR = DATA_DIR / "beta_exports"

@app.post("/beta-save-export")
async def beta_save_export(data: dict):
    """Shrani Meta Ads CSV export na disk z datumom."""
    try:
        BETA_DIR.mkdir(parents=True, exist_ok=True)
        from datetime import datetime
        try:
            import pytz
            lj = pytz.timezone("Europe/Ljubljana")
            now = datetime.now(lj)
        except Exception:
            now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        ts_str = now.strftime("%Y-%m-%d_%H-%M")

        # Meta info
        filename = f"{ts_str}_{data.get('filename','export').replace(' ','_')[:40]}.json"
        campaigns = data.get("campaigns", [])
        payload = {
            "filename": data.get("filename", ""),
            "date": date_str,
            "timestamp": ts_str,
            "campaigns": campaigns,
            "bc_adsets": data.get("bc_adsets", []),
            "total_spend": data.get("total_spend", 0),
            "total_purchases": data.get("total_purchases", 0),
        }
        out_path = BETA_DIR / filename
        out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        size_kb = out_path.stat().st_size / 1024
        print(f"[beta-save] OK: {filename} ({len(campaigns)} kampanj, {size_kb:.1f} KB) -> {out_path}")

        # Pobriši stare (> 30 dni)
        cutoff = now.timestamp() - 30*24*3600
        for f in BETA_DIR.glob("*.json"):
            if f.stat().st_mtime < cutoff:
                f.unlink()

        return {"ok": True, "saved": filename, "size_kb": round(size_kb, 1), "path": str(out_path)}
    except Exception as e:
        import traceback
        print(f"[beta-save] ERROR: {e}")
        traceback.print_exc()
        return {"ok": False, "error": str(e)}

@app.get("/beta-export-history")
async def beta_export_history():
    """Seznam shranjenih exportov (zadnjih 30 dni)."""
    try:
        BETA_DIR.mkdir(exist_ok=True)
        files = sorted(BETA_DIR.glob("*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
        history = []
        for f in files:
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                history.append({
                    "filename": f.name,
                    "original": data.get("filename",""),
                    "date": data.get("date",""),
                    "timestamp": data.get("timestamp",""),
                    "total_spend": data.get("total_spend",0),
                    "total_purchases": data.get("total_purchases",0),
                    "campaign_count": len(data.get("campaigns",[])),
                })
            except:
                pass
        return {"ok": True, "exports": history}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/beta-export-load/{filename}")
async def beta_export_load(filename: str):
    """Naloži specifičen export."""
    try:
        f = BETA_DIR / filename
        if not f.exists():
            return {"ok": False, "error": "Datoteka ne obstaja"}
        data = json.loads(f.read_text(encoding="utf-8"))
        return {"ok": True, **data}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/forecast-fix-today")
async def forecast_fix_today():
    """Ročno popravi entries za danes — združi vse ključe."""
    from datetime import datetime
    try:
        import pytz
        lj = pytz.timezone("Europe/Ljubljana")
        today = datetime.now(lj).strftime("%Y-%m-%d")
        d_now = datetime.now(lj)
    except:
        d_now = datetime.utcnow()
        today = d_now.strftime("%Y-%m-%d")

    slsi_key = f"{d_now.day}. {d_now.month}. {d_now.year}"

    if not FORECAST_HISTORY_FILE.exists():
        return {"ok": False, "error": "History file ne obstaja"}

    hist = json.loads(FORECAST_HISTORY_FILE.read_text(encoding="utf-8"))

    # Združi vnose iz VSEH možnih ključev za danes
    all_raw = {}
    keys_used = []
    for key in [today, slsi_key]:
        if hist.get(key):
            keys_used.append(key)
            for e in hist[key]:
                time_key = f"{e.get('h',0):02d}:{e.get('m',0):02d}"
                if time_key not in all_raw:
                    all_raw[time_key] = e

    if not all_raw:
        return {"ok": False, "error": "Ni vnosov za danes", "tried": [today, slsi_key], "available": list(hist.keys())[-5:]}

    # Sortiraj po času
    sorted_entries = sorted(all_raw.values(), key=lambda e: e.get('h',0)*60 + e.get('m',0))

    recovered = {
        "date": today,
        "entries": [
            {
                "label": str(e.get("h",0)).zfill(2) + ":" + str(e.get("m",0)).zfill(2),
                "dejanski": e.get("rev", 0),
                "dejanskiOrd": e.get("ord", 0),
                "napoved": None,
                "napovedOrd": None,
            }
            for e in sorted_entries
        ]
    }
    FORECAST_ENTRIES_FILE.write_text(json.dumps(recovered, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "keys_used": keys_used, "entries_recovered": len(sorted_entries), "entries": recovered["entries"]}

@app.get("/forecast-clear-today")
async def forecast_clear_today():
    """Počisti entries za danes (admin endpoint)."""
    try:
        if FORECAST_ENTRIES_FILE.exists():
            data = json.loads(FORECAST_ENTRIES_FILE.read_text(encoding="utf-8"))
            old_count = len(data.get("entries", []))
            FORECAST_ENTRIES_FILE.write_text(json.dumps({}, ensure_ascii=False, indent=2), encoding="utf-8")
            return {"ok": True, "cleared": old_count}
        return {"ok": True, "cleared": 0}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ─── SPAM FILTER ────────────────────────────────────────────────────────────
SPAM_DIR = DATA_DIR / "spam"
SPAM_DIR.mkdir(exist_ok=True, parents=True)
SPAM_CONFIRMED_FILE = SPAM_DIR / "confirmed.json"
SPAM_REJECTED_FILE = SPAM_DIR / "rejected.json"

def _spam_load_set(path: Path) -> set:
    if not path.exists(): return set()
    try:
        return set(json.loads(path.read_text(encoding="utf-8")))
    except: return set()

def _spam_save_set(path: Path, s: set):
    path.write_text(json.dumps(list(s), ensure_ascii=False, indent=2), encoding="utf-8")

SPAM_THRESHOLD = 80  # Spodnja meja za prikaz - samo >80% se vrne v ads.slx

@app.post("/spam-analyze")
async def spam_analyze():
    """Naloži zadnjih 100 ticketov iz vseh brand-ov, klasificira z AI in vrne SAMO tiste z score >80%."""
    if not KAYAKO_API_KEY or not KAYAKO_SECRET:
        return {"ok": False, "error": "Kayako API ni konfiguriran"}

    confirmed = _spam_load_set(SPAM_CONFIRMED_FILE)
    rejected  = _spam_load_set(SPAM_REJECTED_FILE)

    all_tickets = []
    async with httpx.AsyncClient() as client:
        # 100 odprtih ticketov iz vsakega branda = 200 skupaj
        for brand, dept_id in KAYAKO_DEPT.items():
            # Status filter: 1 = samo odprti (status_id=1)
            path = f"/Tickets/Ticket/ListAll/{dept_id}/1/-1/-1/100/0/ticketid/DESC"
            url = _kayako_build_url(path)
            try:
                r = await client.get(url, timeout=30)
                if r.status_code != 200:
                    print(f"[spam] ListAll {brand} HTTP {r.status_code}")
                    continue
                tickets = _parse_ticket_xml(r.text)
            except Exception as e:
                print(f"[spam] ListAll error {brand}: {e}")
                continue

            for t in tickets:
                tid = t.get("id", "")
                if tid in rejected or tid in confirmed:
                    continue  # že obdelano - preskoči
                t["brand"] = brand
                # Ne nalagamo postov — analiziramo samo naslov
                all_tickets.append(t)

    # AI klasifikacija
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"ok": False, "error": "ANTHROPIC_API_KEY ni nastavljen"}

    client_h = anthropic.Anthropic(api_key=api_key)
    classified = []
    skipped = 0

    # Hitri pre-filter: označi za AI samo angleške naslove
    # Ne-angleški znaki, šumniki, in tipične besede v drugih jezikih = preskoči
    import unicodedata as _ud

    def _is_english_only(text: str) -> bool:
        """True če je naslov verjetno angleški (brez šumnikov, brez ne-ASCII črk)."""
        if not text or len(text.strip()) < 3:
            return False
        # Skip prazni / "(no subject)" / "no subject"
        ts = text.strip().lower()
        if ts in ('(no subject)', 'no subject', '(brez naslova)', 'brez naslova', '-', '...'):
            return False
        # Vsebuje šumnike ali cirilico ali grško → ni angleški
        for ch in text:
            if ch in 'čšžćđČŠŽĆĐńąęłóśźżŃĄĘŁÓŚŹŻáéíóúýäöüÁÉÍÓÚÝÄÖÜőűŐŰâîăşţÂÎĂŞŢ':
                return False
            # Ne-latinski alfabet (cirilica, grščina, arabščina, kitajščina...)
            if ch.isalpha() and ord(ch) > 591:
                return False
        # Tipične ne-angleške besede (običajno na začetku naslova)
        text_lower = text.lower()
        non_english_keywords = [
            'naroč', 'pošilj', 'dostav', 'plač', 'računa', 'izdelek', 'vraač',
            'reklam', 'kupon', 'paket', 'cena', 'ponudb', 'poizvedb',
            'narudžb', 'isporuk', 'plać', 'račun', 'proizvod', 'povrat',
            'narudžba', 'poručio', 'paypal',
            'поръчк', 'доставк', 'плащане', 'продукт',
            'objedn', 'doruč', 'platb', 'výrobok', 'vrátenie',
            'rendel', 'szállí', 'fizetés', 'termék', 'visszaküld',
            'zamówi', 'dostawa', 'płatność', 'produkt', 'zwrot',
            'comand', 'livrare', 'plat', 'produs', 'retur',
            'παραγγελ', 'αποστολ', 'πληρωμ', 'προϊόν', 'επιστροφ'
        ]
        for kw in non_english_keywords:
            if kw in text_lower:
                return False
        return True

    for t in all_tickets:
        subject = t.get("subject", "")
        from_email = t.get("email", "")
        from_name = t.get("fullname", "")
        replies = int(t.get("replies", "0") or "0")

        # Skip ticket-i ki že imajo odgovore — verjetno legitimna komunikacija
        if replies > 0:
            rejected.add(t.get("id",""))
            skipped += 1
            continue

        # PRE-FILTER: samo angleški naslovi grejo skozi AI
        if not _is_english_only(subject):
            rejected.add(t.get("id",""))
            skipped += 1
            continue

        prompt = f"""Si strog klasifikator email spam-a. Analiziraj SAMO naslov ticket-a iz Kayako support sistema (e-trgovina).

POŠILJATELJ: {from_name} <{from_email}>
NASLOV: {subject}

Odgovori SAMO z JSON:
{{"score": <0-100, koliko verjetno je spam>, "reason": "<razlaga v slovenščini, max 1 stavek>"}}

Visok score (>80) = oglaševanje SEO/marketing/development storitev, B2B "we offer X services", "increase your traffic", phishing, prevare, "partnership opportunity", agencije ki ponujajo svoje storitve, generično angleško sporočilo brez konteksta naših izdelkov.
Srednji (50-80) = sumljivo a možno legitimno (npr. tuji kupec).
Nizek (<50) = pristno vprašanje o naročilih/izdelkih/dostavi (npr. order #12345, my order, refund request, where is my package)."""

        try:
            msg = client_h.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=200,
                messages=[{"role": "user", "content": prompt}]
            )
            txt = msg.content[0].text.strip() if msg.content else "{}"
            txt = re.sub(r'^```(?:json)?|```$', '', txt, flags=re.MULTILINE).strip()
            try:
                ai = json.loads(txt)
                score = int(ai.get("score", 0))
                reason = ai.get("reason", "")
            except:
                continue
        except Exception as e:
            print(f"[spam] AI error for {t.get('id','')}: {e}")
            continue

        # Samo nad threshold-om
        if score < SPAM_THRESHOLD:
            # Auto-zavrni tiste pod threshold-om da jih ne procesiramo več
            rejected.add(t.get("id",""))
            skipped += 1
            continue

        classified.append({
            "id": t.get("id", ""),
            "brand": t.get("brand", ""),
            "subject": subject,
            "from": f"{from_name} <{from_email}>" if from_email else from_name,
            "score": score,
            "reason": reason,
            "url": f"https://support.silux.si/staff/index.php?/Tickets/Ticket/View/{t.get('id','')}",
        })

    # Shrani auto-rejected da jih naslednji klic preskoči
    _spam_save_set(SPAM_REJECTED_FILE, rejected)

    return {"ok": True, "tickets": classified, "confirmed": list(confirmed), "stats": {"checked": len(all_tickets), "found": len(classified), "filtered_out": skipped}}

@app.post("/spam-confirm")
async def spam_confirm(data: dict):
    """Potrdi spam → premakni v Trash v Kayako, ali zavrni."""
    tid = str(data.get("id", ""))
    action = data.get("action", "")
    if not tid:
        return {"error": "missing id"}

    confirmed = _spam_load_set(SPAM_CONFIRMED_FILE)
    rejected  = _spam_load_set(SPAM_REJECTED_FILE)

    if action == "confirm":
        # Mark ticket as trash v Kayako (PUT z trash=1)
        try:
            salt = str(random.randint(1000000000, 9999999999))
            raw_sig = hmac.new(
                key=KAYAKO_SECRET.encode("utf-8"),
                msg=salt.encode("utf-8"),
                digestmod=hashlib.sha256
            ).digest()
            signature = base64.b64encode(raw_sig).decode("utf-8")
            
            url = f"{KAYAKO_API_URL}?e=/Tickets/Ticket/{tid}"
            form_data = {
                "apikey":    KAYAKO_API_KEY,
                "salt":      salt,
                "signature": signature,
                "trash":     "1",  # Premakne v Trash brez delete
            }
            async with httpx.AsyncClient() as client:
                r = await client.put(url, data=form_data, timeout=20)
                print(f"[spam] Trash ticket {tid}: HTTP {r.status_code} | response: {r.text[:200]}")
                if r.status_code not in (200, 204):
                    return {"ok": False, "error": f"Kayako error: {r.status_code} — {r.text[:100]}"}
        except Exception as e:
            print(f"[spam] Trash error: {e}")
            return {"ok": False, "error": str(e)}

        confirmed.add(tid)
        rejected.discard(tid)
        _spam_save_set(SPAM_CONFIRMED_FILE, confirmed)
        _spam_save_set(SPAM_REJECTED_FILE, rejected)
        return {"ok": True, "trashed": True}

    elif action == "reject":
        rejected.add(tid)
        confirmed.discard(tid)
    elif action == "unconfirm":
        confirmed.discard(tid)

    _spam_save_set(SPAM_CONFIRMED_FILE, confirmed)
    _spam_save_set(SPAM_REJECTED_FILE, rejected)
    return {"ok": True}

@app.post("/spam-clear-confirmed")
async def spam_clear_confirmed():
    """Počisti seznam potrjenih spam-ov."""
    _spam_save_set(SPAM_CONFIRMED_FILE, set())
    return {"ok": True}


# ─── KNJIGOVODSTVO (Polcar XML → CSV) ────────────────────────────────────────
KNJ_DIR = DATA_DIR / "knjigovodstvo"
KNJ_DIR.mkdir(exist_ok=True, parents=True)

def _parse_polcar_debit(content: str) -> list[dict]:
    """Polcar Debit Note (DU) — ROOT atributi + USL elementi."""
    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        print(f"[knj] debit parse error: {e}")
        return []
    rows = []
    for usl in root.findall('USL'):
        rows.append({
            'Tip':                'DEBIT NOTE',
            'Številka':           root.get('NumerRachunku', ''),
            'Datum prodaje':      root.get('DataSprzedazyText_Wartosc', ''),
            'Datum izdaje':       root.get('DataWystawieniaText_Wartosc', ''),
            'Rok plačila':        root.get('TerminPlatnosciText_Wartosc', ''),
            'Prodajalec':         root.get('WystawcaNazwa', ''),
            'Kupec':              root.get('OdbiorcaNazwa', '').strip(),
            'EU VAT':             root.get('OdbiorcaUEVAT', '') or root.get('OdbiorcaNIP',''),
            'Valuta':             root.get('Rachunek_Waluta', '').strip(),
            'Art. številka':      usl.get('NazwaUslugi', ''),
            'Opis':               usl.get('OpisUslugi', '') or usl.get('PowodWystawienia',''),
            'Količina':           '1',
            'Cena/kos':           usl.get('WartoscNetto', '').replace(',','.'),
            'Vrednost neto':      usl.get('WartoscNetto', '').replace(',','.'),
            'VAT %':              str(usl.get('StawkaVat', '0')),
            'Vrednost VAT':       usl.get('WartoscVat', '').replace(',','.'),
            'Vrednost bruto':     usl.get('WartoscBrutto', '').replace(',','.'),
            'Razlog':             usl.get('PowodWystawienia', ''),
            'Skupaj bruto':       root.get('Rachunek_WartoscBrutto', '').replace(',','.'),
        })
    if not rows:
        # Fallback — header only
        rows.append({
            'Tip':            'DEBIT NOTE',
            'Številka':       root.get('NumerRachunku', ''),
            'Datum prodaje':  root.get('DataSprzedazyText_Wartosc', ''),
            'Datum izdaje':   root.get('DataWystawieniaText_Wartosc', ''),
            'Rok plačila':    root.get('TerminPlatnosciText_Wartosc', ''),
            'Prodajalec':     root.get('WystawcaNazwa', ''),
            'Kupec':          root.get('OdbiorcaNazwa', '').strip(),
            'EU VAT':         root.get('OdbiorcaUEVAT', ''),
            'Valuta':         root.get('Rachunek_Waluta', '').strip(),
            'Art. številka':  '',
            'Opis':           '',
            'Količina':       '',
            'Cena/kos':       '',
            'Vrednost neto':  root.get('Rachunek_WartoscNetto','').replace(',','.'),
            'VAT %':          '',
            'Vrednost VAT':   root.get('Rachunek_WartoscVat','').replace(',','.'),
            'Vrednost bruto': root.get('Rachunek_WartoscBrutto','').replace(',','.'),
            'Razlog':         '',
            'Skupaj bruto':   root.get('Rachunek_WartoscBrutto','').replace(',','.'),
        })
    return rows

def _parse_polcar_credit(content: str) -> list[dict]:
    """Polcar Credit Note (NU) — ROOT atributi + POZ elementi."""
    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        print(f"[knj] credit parse error: {e}")
        return []
    rows = []
    for poz in root.findall('POZ'):
        razlog = ''
        for p in root.findall('Powody'):
            if p.get('Symbol') == poz.get('Symbol'):
                razlog = p.get('Powod', '').replace('(','| ').replace(')','').strip()
        rows.append({
            'Tip':            'CREDIT NOTE',
            'Številka':       root.get('NumerKorekty', ''),
            'Za dokument':    root.get('NumerRachunku', ''),
            'Datum prodaje':  root.get('Rachunek_DataSprzedazyText_Wartosc', ''),
            'Datum izdaje':   root.get('Korekta_DataWystawienia_Wartosc', ''),
            'Rok plačila':    root.get('TerminPlatnosciText_Wartosc', '') or root.get('TerminWyplatyText',''),
            'Prodajalec':     root.get('WystawcaNazwa', ''),
            'Kupec':          root.get('OdbiorcaNazwa', '').strip(),
            'EU VAT':         root.get('OdbiorcaUEVAT', '') or root.get('OdbiorcaNIP',''),
            'Valuta':         root.get('Korekta_Waluta', '').strip(),
            'Art. številka':  poz.get('Numer', ''),
            'Art. št. kupca': poz.get('NumerTowaruKlienta', ''),
            'Naziv artikla':  poz.get('Naziv', '') or poz.get('Nazwa',''),
            'Opis':           poz.get('Opis', ''),
            'Količina':       poz.get('Ilosc', '').split('[')[0].strip(),
            'Cena/kos':       poz.get('CenaJednostkowa', '').replace(',','.'),
            'Vrednost neto':  poz.get('WartoscNetto', '').replace(',','.'),
            'VAT %':          str(poz.get('StawkaVat', '0')),
            'Vrednost VAT':   poz.get('WartoscVat', '').replace(',','.'),
            'Vrednost bruto': poz.get('WartoscBrutto', '').replace(',','.'),
            'Razlog':         razlog,
            'Skupaj bruto':   root.get('Korekta_WartoscBrutto','').replace(',','.'),
        })
    return rows

@app.post("/knj-parse")
async def knj_parse(files: list[UploadFile] = File(...)):
    """Sprejme Polcar XML datoteke (DU/NU) in vrne razčlenjene vrstice."""
    debit_rows = []
    credit_rows = []
    parsed_count = 0

    for f in files:
        try:
            raw = await f.read()
            # Polcar XML so v UTF-16 z BOM (\xff\xfe)
            try:
                content = raw.decode('utf-16')
            except UnicodeDecodeError:
                try:
                    content = raw.decode('utf-8-sig')
                except UnicodeDecodeError:
                    content = raw.decode('utf-8', errors='ignore')

            fname = (f.filename or '').upper()
            if '_DU_' in fname or '_DU.' in fname or 'DU.XML' in fname:
                debit_rows.extend(_parse_polcar_debit(content))
            elif '_NU_' in fname or '_NU.' in fname or 'NU.XML' in fname:
                credit_rows.extend(_parse_polcar_credit(content))
            else:
                # Heuristika: poskusi oba parserja
                d = _parse_polcar_debit(content)
                c = _parse_polcar_credit(content)
                if d and not c: debit_rows.extend(d)
                elif c and not d: credit_rows.extend(c)
                elif len(d) > len(c): debit_rows.extend(d)
                else: credit_rows.extend(c)
            parsed_count += 1
        except Exception as e:
            print(f"[knj] error parsing {f.filename}: {e}")
            continue

    return {
        "ok": True,
        "parsed": parsed_count,
        "debit": debit_rows,
        "credit": credit_rows,
    }

@app.post("/knj-save")
async def knj_save(data: dict):
    """Shrani CSV podatke v zgodovino."""
    try:
        from datetime import datetime
        try:
            import pytz
            lj = pytz.timezone("Europe/Ljubljana")
            now = datetime.now(lj)
        except:
            now = datetime.utcnow()

        ts = now.strftime("%Y-%m-%d_%H-%M-%S")
        filename = data.get("filename", f"polcar_{ts}.csv")
        # Sanitiziraj filename
        safe = re.sub(r'[^\w\-_.]', '_', filename)
        save_path = KNJ_DIR / f"{ts}_{safe}.json"

        payload = {
            "filename": filename,
            "timestamp": now.strftime("%Y-%m-%d %H:%M"),
            "rows": data.get("rows", []),
            "columns": data.get("columns", []),
        }
        save_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "saved": save_path.name}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/knj-history")
async def knj_history():
    """Zgodovina shranjenih CSV-jev."""
    try:
        files = sorted(KNJ_DIR.glob("*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
        exports = []
        for f in files[:50]:
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                exports.append({
                    "filename": data.get("filename", f.name),
                    "timestamp": data.get("timestamp", ""),
                    "rows": len(data.get("rows", [])),
                    "internal_id": f.name,
                })
            except:
                pass
        return {"ok": True, "exports": exports}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/knj-load/{filename}")
async def knj_load(filename: str):
    """Naloži shranjen CSV iz zgodovine."""
    try:
        # Iščemo po imenu (filename je user-facing ime, ne internal)
        for f in KNJ_DIR.glob("*.json"):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                if data.get("filename") == filename:
                    return {"ok": True, **data}
            except:
                continue
        return {"ok": False, "error": "Ni najden"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ═══════════════════════════════════════════════════════════════════
# FORECAST 2.0 — Clean rewrite (paralelni sistem)
# ═══════════════════════════════════════════════════════════════════

FORECAST2_DIR = DATA_DIR / "forecast2"
FORECAST2_DIR.mkdir(exist_ok=True, parents=True)

# Struktura: /data/forecast2/YYYY-MM-DD.json
# Vsebina: {"entries":[{"time":"HH:MM","orders":N,"revenue":N,"_ts":timestamp}], "final":{"orders":N,"revenue":N}}
#
# Vsak dan = svoj fajl. Ne mešamo več datumov.

def _lj_now():
    """Aktualni čas v Ljubljani (Europe/Ljubljana, samodejno poletni/zimski čas)."""
    from datetime import datetime, timezone, timedelta
    # 1) zoneinfo (vgrajen v Python 3.9+, ne potrebuje pytz)
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("Europe/Ljubljana"))
    except Exception:
        pass
    # 2) pytz (če slučajno na voljo)
    try:
        import pytz
        return datetime.now(pytz.timezone("Europe/Ljubljana"))
    except Exception:
        pass
    # 3) Fallback: ročni odmik (Slovenija = UTC+1 zimi, UTC+2 poleti).
    #    Groba ocena DST: zadnja nedelja marca → zadnja nedelja oktobra = +2h, sicer +1h.
    now_utc = datetime.now(timezone.utc)
    year = now_utc.year
    # zadnja nedelja v marcu
    mar = datetime(year, 3, 31, tzinfo=timezone.utc)
    dst_start = mar - timedelta(days=(mar.weekday() + 1) % 7)
    # zadnja nedelja v oktobru
    oct_ = datetime(year, 10, 31, tzinfo=timezone.utc)
    dst_end = oct_ - timedelta(days=(oct_.weekday() + 1) % 7)
    offset = 2 if (dst_start <= now_utc < dst_end) else 1
    return now_utc + timedelta(hours=offset)

def _lj_today():
    return _lj_now().strftime("%Y-%m-%d")

# Sledenje stanja forecast cron-a (za diagnostiko manjkajočih urnih vnosov)
_FORECAST2_CRON_STATUS = {
    "started_at": None, "last_tick": None, "last_run": None,
    "last_result": None, "last_error": None, "tick_count": 0,
}

def _forecast2_path(date_iso: str) -> Path:
    """Pot do datoteke za določen datum."""
    # Sanity check format YYYY-MM-DD
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_iso):
        raise ValueError(f"Invalid date format: {date_iso}")
    return FORECAST2_DIR / f"{date_iso}.json"

def _forecast2_load_day(date_iso: str) -> dict:
    """Naloži en dan. Vrne {entries:[], final:{}} če ne obstaja."""
    p = _forecast2_path(date_iso)
    if not p.exists():
        return {"date": date_iso, "entries": [], "final": None}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[fc2] load error {date_iso}: {e}")
        return {"date": date_iso, "entries": [], "final": None}

def _forecast2_save_day(date_iso: str, data: dict):
    """Shrani en dan."""
    p = _forecast2_path(date_iso)
    p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

# ─── BACKUP STAREGA SISTEMA ────────────────────────────────────────

@app.get("/forecast-backup")
async def forecast_backup():
    """Vrne ves stari forecast state kot JSON za backup."""
    backup = {
        "timestamp": _lj_now().isoformat(),
        "entries_file": {},
        "history_file": {},
        "deleted_file": {},
    }
    try:
        if FORECAST_ENTRIES_FILE.exists():
            backup["entries_file"] = json.loads(FORECAST_ENTRIES_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        backup["entries_file_error"] = str(e)
    try:
        if FORECAST_HISTORY_FILE.exists():
            backup["history_file"] = json.loads(FORECAST_HISTORY_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        backup["history_file_error"] = str(e)
    try:
        if FORECAST_DELETED_FILE.exists():
            backup["deleted_file"] = json.loads(FORECAST_DELETED_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        backup["deleted_file_error"] = str(e)
    return backup

# ─── MIGRACIJA STAREGA SISTEMA → V NOV ───────────────────────────

@app.post("/forecast2-migrate")
async def forecast2_migrate():
    """Migracija starega forecast_history.json v nov format (en fajl per dan)."""
    migrated_days = 0
    skipped_days = 0
    errors = []

    if not FORECAST_HISTORY_FILE.exists():
        return {"ok": True, "message": "Stara zgodovina ne obstaja", "migrated": 0}

    try:
        hist = json.loads(FORECAST_HISTORY_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        return {"ok": False, "error": f"Cannot read history: {e}"}

    for key, val in hist.items():
        # Pretvori format datuma → ISO (YYYY-MM-DD)
        date_iso = None
        if re.match(r'^\d{4}-\d{2}-\d{2}$', key):
            date_iso = key
        elif '.' in key:
            parts = [p.strip().strip('.') for p in key.split('.') if p.strip()]
            if len(parts) == 3:
                try:
                    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                    date_iso = f"{y}-{m:02d}-{d:02d}"
                except:
                    pass

        if not date_iso:
            errors.append(f"Cannot parse date key: {key}")
            skipped_days += 1
            continue

        # Če ciljni dan že obstaja, ne prepisuj (varnost)
        target = _forecast2_path(date_iso)
        if target.exists():
            skipped_days += 1
            continue

        # Pretvori val v nov format
        # val je lahko različnih formatov — število, dict, list...
        day_data = {"date": date_iso, "entries": [], "final": None}

        if isinstance(val, dict):
            # Možen format: {"orders": N, "revenue": N} ali entries: [...]
            if "entries" in val and isinstance(val["entries"], list):
                day_data["entries"] = val["entries"]
            if "final" in val:
                day_data["final"] = val["final"]
            if "orders" in val or "revenue" in val:
                day_data["final"] = {
                    "orders": val.get("orders", 0),
                    "revenue": val.get("revenue", 0),
                }
        elif isinstance(val, (int, float)):
            day_data["final"] = {"orders": 0, "revenue": val}
        elif isinstance(val, list):
            day_data["entries"] = val

        try:
            _forecast2_save_day(date_iso, day_data)
            migrated_days += 1
        except Exception as e:
            errors.append(f"{date_iso}: {e}")

    return {
        "ok": True,
        "migrated": migrated_days,
        "skipped": skipped_days,
        "errors": errors[:20],
    }

# ─── ENDPOINTI ─────────────────────────────────────────────────────

@app.get("/forecast2-today")
async def forecast2_today():
    """Vrne entries za današnji dan."""
    today = _lj_today()
    data = _forecast2_load_day(today)
    return {"ok": True, "today": today, **data}

@app.get("/forecast2-stats")
async def forecast2_stats(year: int = 2026):
    """Seštevek naročil + prometa za celo leto + PROJEKCIJA leta (za domači števec).
    Projekcija uporablja isti model kot revenue_forecast.html:
    pretekli meseci = dejansko, tekoči = dnevno povprečje × dnevi, prihodnji = 2025 vzorec × YoY faktor."""
    try:
        # lanski mesečni neto (2025) — isti vir kot revenue_forecast.html
        REV_2025_NET = {0:567249, 1:443491, 2:443923, 3:393831, 4:390365, 5:526175,
                        6:528250, 7:404497, 8:407053, 9:508581, 10:724774, 11:701193}

        # zberi dejanske dni leta (final.orders/revenue)
        day_rows = []  # (date_iso, orders, revenue)
        prefix = f"{year}-"
        for f in FORECAST2_DIR.glob(f"{prefix}*.json"):
            try:
                day = json.loads(f.read_text(encoding="utf-8"))
            except Exception:
                continue
            final = day.get("final")
            if final and isinstance(final, dict):
                o = int(final.get("orders", 0) or 0)
                rv = float(final.get("revenue", 0) or 0)
                if o > 0:
                    day_rows.append((f.stem, o, rv))
        day_rows.sort(key=lambda x: x[0])

        total_orders = sum(r[1] for r in day_rows)
        total_revenue = sum(r[2] for r in day_rows)
        days_counted = len(day_rows)

        # === POVPREČJA zadnjih 7 in 22 dni (z dejanskimi končnimi podatki) ===
        def _avg_last(n):
            last = day_rows[-n:] if len(day_rows) >= 1 else []
            if not last:
                return 0, 0.0
            ao = sum(r[1] for r in last) / len(last)
            ar = sum(r[2] for r in last) / len(last)
            return round(ao), round(ar, 2)
        avg_7d_orders, avg_7d_revenue = _avg_last(7)
        avg_22d_orders, avg_22d_revenue = _avg_last(22)

        # === PROJEKCIJA (ista logika kot frontend) ===
        import calendar as _cal
        from datetime import date as _d
        now = _lj_now()
        cur_year = now.year
        proj_revenue = total_revenue
        proj_orders = total_orders
        if day_rows and cur_year == year:
            # meseci → orders, revenue, days
            months = {}
            for date_iso, o, rv in day_rows:
                m = int(date_iso[5:7]) - 1
                months.setdefault(m, {"orders":0, "revenue":0.0, "days":0})
                months[m]["orders"] += o
                months[m]["revenue"] += rv
                months[m]["days"] += 1
            cur_month = now.month - 1
            def _dim(m): return _cal.monthrange(year, m+1)[1]
            # zadnjih 30 dni povprečje
            last30 = day_rows[-30:]
            avg30_rev = sum(r[2] for r in last30) / max(1, len(last30))
            # YoY faktor (povprečje razmerij kompletnih 2026 mesecev vs 2025)
            yoy_ratios = []
            for m in range(cur_month):
                if m in months and months[m]["days"] >= _dim(m) - 1 and REV_2025_NET.get(m):
                    yoy_ratios.append(months[m]["revenue"] / REV_2025_NET[m])
            yoy = sum(yoy_ratios)/len(yoy_ratios) if yoy_ratios else 1.0
            # projekcija prometa
            proj_revenue = total_revenue
            for m in range(12):
                if m < cur_month:
                    pass  # že v total_revenue
                elif m == cur_month:
                    md = months.get(m, {"revenue":0.0, "days":0})
                    days_so_far = md["days"]
                    daily = (md["revenue"]/days_so_far) if days_so_far > 0 else avg30_rev
                    proj_revenue += daily * (_dim(m) - days_so_far)
                else:
                    proj_revenue += REV_2025_NET.get(m, 0) * yoy
            # projekcija naročil prek povprečnega AOV
            aov = total_revenue / max(1, total_orders)
            proj_orders = int(round(proj_revenue / aov)) if aov else total_orders

        # AOV (povprečna vrednost naročila) — vedno
        aov_value = (total_revenue / total_orders) if total_orders > 0 else 0
        # najboljši dan (največ naročil)
        best_day = None
        if day_rows:
            bd = max(day_rows, key=lambda r: r[1])
            # lep zapis datuma DD.MM.YYYY
            try:
                y, mo, dy = bd[0].split("-")
                best_date_fmt = f"{int(dy)}.{int(mo)}.{y}"
            except Exception:
                best_date_fmt = bd[0]
            best_day = {"date": bd[0], "date_fmt": best_date_fmt, "orders": bd[1], "revenue": round(bd[2], 2)}

        return {
            "ok": True, "year": year,
            "total_orders": total_orders,
            "total_revenue": round(total_revenue, 2),
            "days_counted": days_counted,
            "days_with_final": days_counted,
            "avg_7d_orders": avg_7d_orders,
            "avg_7d_revenue": avg_7d_revenue,
            "avg_22d_orders": avg_22d_orders,
            "avg_22d_revenue": avg_22d_revenue,
            "projection_orders": proj_orders,
            "projection_revenue": round(proj_revenue, 2),
            "aov": round(aov_value, 2),
            "best_day": best_day,
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}

@app.post("/forecast2-add-entry")
async def forecast2_add_entry(data: dict):
    """Doda nov entry za današnji dan (ali za določen datum)."""
    try:
        time_str = data.get("time", "")
        orders = int(data.get("orders", 0))
        revenue = float(data.get("revenue", 0))
        date_iso = data.get("date") or _lj_today()

        # Validacija
        if not re.match(r'^\d{1,2}:\d{2}$', time_str):
            return {"ok": False, "error": "Invalid time format (use HH:MM)"}

        # Normaliziraj čas
        h, m = time_str.split(":")
        time_norm = f"{int(h):02d}:{int(m):02d}"

        day = _forecast2_load_day(date_iso)

        # Če entry s tem časom že obstaja, ga zamenjaj (ne podvoji)
        day["entries"] = [e for e in day["entries"] if e.get("time") != time_norm]
        day["entries"].append({
            "time": time_norm,
            "orders": orders,
            "revenue": revenue,
            "_ts": _lj_now().isoformat(),
        })
        # Sortiraj po času
        day["entries"].sort(key=lambda e: e.get("time", ""))

        _forecast2_save_day(date_iso, day)
        print(f"[fc2] saved entry {date_iso} @ {time_norm}: orders={orders}, revenue={revenue}")
        return {"ok": True, "date": date_iso, "entries": day["entries"]}
    except Exception as e:
        print(f"[fc2] add-entry error: {e}")
        return {"ok": False, "error": str(e)}

@app.post("/forecast2-delete-entry")
async def forecast2_delete_entry(data: dict):
    """Briše entry po času."""
    try:
        time_str = data.get("time", "")
        date_iso = data.get("date") or _lj_today()

        day = _forecast2_load_day(date_iso)
        before = len(day["entries"])
        day["entries"] = [e for e in day["entries"] if e.get("time") != time_str]
        after = len(day["entries"])

        if before == after:
            return {"ok": False, "error": "Entry not found"}

        _forecast2_save_day(date_iso, day)
        return {"ok": True, "deleted": before - after}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/zaloga-debug-vsota")
async def zaloga_debug_vsota():
    """Diagnostika skupne vsote: sešteje vse kose v suban.ai CSV in v siluxar API,
    pokaže razliko + SKU-je, kjer se zaloga razlikuje. Razkrije, kje je manjko kosov."""
    import csv as _csv
    from io import StringIO as _SIO

    def _to_int(v):
        try: return int(float(str(v).replace(',', '.')))
        except: return 0

    # 1) suban.ai CSV — vsota + mapa (SKU+skladišče) -> stock
    csv_total = 0
    csv_rows = 0
    csv_map = {}  # key = sku|warehouse -> stock
    csv_skus = set()
    csv_vrednost = 0.0   # vsota stock × cena
    def _to_float(v):
        try: return float(str(v).replace(',', '.'))
        except: return 0.0
    if STOCK_CSV_FILE.exists():
        try:
            text = STOCK_CSV_FILE.read_text(encoding='utf-8-sig', errors='replace')
            for row in _csv.DictReader(_SIO(text)):
                sku = (row.get('product_sku') or '').strip()
                wh = (row.get('warehouse') or '').strip()
                st = _to_int(row.get('stock'))
                pr = _to_float(row.get('price') or row.get('price_netto'))
                if not sku:
                    continue
                csv_total += st
                csv_vrednost += st * pr
                csv_rows += 1
                csv_skus.add(sku)
                csv_map[sku + '|' + wh] = csv_map.get(sku + '|' + wh, 0) + st
        except Exception as e:
            return {"error": f"CSV branje: {e}"}

    # 2) siluxar API — vsota + mapa
    api_total = 0
    api_rows = 0
    api_vrednost = 0.0
    api_map = {}
    api_raw = {}
    api_skus = set()
    api_error = None
    prazni_sku_kosi = 0       # zapisi BREZ sku, a s stockom (suban.ai jih preskoči!)
    prazni_sku_vrstic = 0
    prazni_naziv_kosi = 0     # zapisi s sku, a BREZ naziva, a s stockom
    prazni_naziv_vrstic = 0
    prazni_primeri = []
    key = os.environ.get("SILUXAR_STOCK_KEY", "")
    basic_user = os.environ.get("SILUXAR_BASIC_USER", "")
    basic_pass = os.environ.get("SILUXAR_BASIC_PASS", "")
    headers = {}
    _auth = None
    if key: headers["Authorization"] = key
    elif basic_user or basic_pass: _auth = httpx.BasicAuth(basic_user, basic_pass)
    try:
        async with httpx.AsyncClient(timeout=90, auth=_auth) as cli:
            r = await cli.get("https://www.siluxar.si/apistockexport", headers=headers)
        jd = json.loads(r.text or "[]")
        if isinstance(jd, dict):
            for kk in ("data", "items", "rows", "products", "stock"):
                if isinstance(jd.get(kk), list):
                    jd = jd[kk]; break
        if isinstance(jd, list):
            for it in jd:
                if not isinstance(it, dict): continue
                sku = str(it.get('sku') or it.get('product_sku') or '').strip()
                wh = str(it.get('source') or '').strip()
                title = str(it.get('title') or '').strip()
                st = _to_int(it.get('stock'))
                if not sku:
                    # zapis BREZ SKU — suban.ai sync ga PRESKOČI, siluxar ga šteje
                    if st != 0:
                        prazni_sku_kosi += st
                        prazni_sku_vrstic += 1
                        if len(prazni_primeri) < 20:
                            prazni_primeri.append({"id": str(it.get('id') or ''), "sku": "(prazen)",
                                                   "title": title or "(prazen)", "stock": st, "source": wh})
                    continue
                if not title and st != 0:
                    # ima SKU, a prazen naziv (a s stockom) — informativno
                    prazni_naziv_kosi += st
                    prazni_naziv_vrstic += 1
                    if len(prazni_primeri) < 20:
                        prazni_primeri.append({"id": str(it.get('id') or ''), "sku": sku,
                                               "title": "(prazen)", "stock": st, "source": wh})
                api_total += st
                api_vrednost += st * _to_float(it.get('price_netto'))
                api_rows += 1
                api_skus.add(sku)
                _k = sku + '|' + wh
                api_map[_k] = api_map.get(_k, 0) + st
                api_raw.setdefault(_k, []).append(st)   # vsi posamezni zapisi za ta ključ
    except Exception as e:
        api_error = str(e)

    # SIMULACIJA SEŠTEVANJA: kateri ključi imajo VEČ zapisov + kakšen bi bil učinek
    podvojeni = []           # ključi z >1 zapisom
    sestevek_skupaj = 0      # koliko kosov bi PRIDOBILI s seštevanjem (vs trenutni suban.ai)
    sumljivi = []            # podvojeni, kjer sta DVA ne-ničelna zapisa (npr. 424+50) — pozor!
    for k, vals in api_raw.items():
        if len(vals) > 1:
            sku, _, wh = k.partition('|')
            vsota = sum(vals)
            trenutni = csv_map.get(k, 0)   # kar suban.ai zdaj ima
            pridobitev = vsota - trenutni
            nenicelni = [v for v in vals if v != 0]
            rec = {"sku": sku, "skladisce": wh, "zapisi": sorted(vals, reverse=True),
                   "sestevek": vsota, "suban_ai_zdaj": trenutni, "pridobitev": pridobitev}
            podvojeni.append(rec)
            sestevek_skupaj += pridobitev
            if len(nenicelni) > 1:   # dva ali več ne-ničelnih → seštevanje morda napačno
                sumljivi.append(rec)
    podvojeni.sort(key=lambda x: abs(x["pridobitev"]), reverse=True)

    # 3) razlike po ključu (SKU+skladišče)
    razlike = []
    vsi_kljuci = set(csv_map.keys()) | set(api_map.keys())
    for k in vsi_kljuci:
        cv = csv_map.get(k, 0)
        av = api_map.get(k, 0)
        if cv != av:
            sku, _, wh = k.partition('|')
            razlike.append({"sku": sku, "skladisce": wh, "suban_ai": cv, "siluxar": av, "razlika": av - cv})
    razlike.sort(key=lambda x: abs(x["razlika"]), reverse=True)

    # SKU-ji, ki so samo v enem viru
    samo_v_csv = sorted(csv_skus - api_skus)[:50]
    samo_v_api = sorted(api_skus - csv_skus)[:50]

    return {
        "suban_ai": {"vsota_kosov": csv_total, "vrstic": csv_rows, "unikatnih_sku": len(csv_skus),
                     "vrednost_eur": round(csv_vrednost, 2)},
        "siluxar_api": {"vsota_kosov": api_total, "vrstic": api_rows, "unikatnih_sku": len(api_skus),
                        "vrednost_eur": round(api_vrednost, 2)},
        "RAZLIKA_VREDNOSTI_eur": round(api_vrednost - csv_vrednost, 2),
        "siluxar_skupaj_z_praznimi": api_total + prazni_sku_kosi,
        "PRAZNI_SKU": {
            "kosov": prazni_sku_kosi, "vrstic": prazni_sku_vrstic,
            "opomba": "Zapisi BREZ sku — suban.ai jih PRESKOČI, siluxar pa morda šteje. VERJETEN izvor razlike."
        },
        "prazni_naziv": {"kosov": prazni_naziv_kosi, "vrstic": prazni_naziv_vrstic,
                         "opomba": "Imajo SKU a prazen naziv — suban.ai jih SHRANI (samo informativno)."},
        "prazni_primeri": prazni_primeri,
        "razlika_vsote": api_total - csv_total,
        "stevilo_razlik": len(razlike),
        "razlike_po_sku": razlike[:60],
        "SIMULACIJA_SESTEVANJA": {
            "kljucev_z_vec_zapisi": len(podvojeni),
            "kosov_pridobljenih_ce_sestejemo": sestevek_skupaj,
            "opomba": "Toliko kosov bi suban.ai PRIDOBIL, če bi seštevali podvojene SKU+skladišče. Naj bi ujemalo razliko.",
            "SUMLJIVI_dva_nenicelna": sumljivi,
            "sumljivih_opomba": "Tu sta DVA ne-ničelna zapisa (npr. 424+50) — seštevanje bi morda napačno podvojilo. Če prazno → seštevanje varno.",
            "podvojeni_primeri": podvojeni[:40],
        },
        "samo_v_suban_ai": samo_v_csv,
        "samo_v_siluxar": samo_v_api,
        "api_error": api_error,
        "namig": "Če PRAZNI_SKU.kosov == tvoja razlika → to je vzrok (zapisi brez SKU, ki jih suban.ai ne more shraniti).",
    }


@app.get("/zaloga-debug-sku")
async def zaloga_debug_sku(q: str = ""):
    """Diagnostika neskladja: za dani SKU ali ID pokaže (1) kaj je v shranjenem CSV
    in (2) kaj siluxar API vrne. Razkrije, zakaj se zaloga ne ujema.
    Uporaba: /zaloga-debug-sku?q=8488407  (ID ali SKU)"""
    import csv as _csv
    from io import StringIO as _SIO
    q = (q or "").strip()
    if not q:
        return {"error": "Podaj ?q=SKU ali ID"}
    ql = q.lower()

    # 1) iz shranjenega CSV (suban.ai)
    csv_matches = []
    if STOCK_CSV_FILE.exists():
        try:
            text = STOCK_CSV_FILE.read_text(encoding='utf-8-sig', errors='replace')
            for row in _csv.DictReader(_SIO(text)):
                sku = (row.get('product_sku') or '').strip()
                pid = (row.get('product_id') or '').strip()
                sid = (row.get('siluxar_id') or '').strip()
                if ql in (sku.lower(), pid.lower(), sid.lower()):
                    csv_matches.append({
                        "product_sku": sku, "product_id": pid, "siluxar_id": sid,
                        "stock": row.get('stock'), "stock30": row.get('stock30'),
                        "position": row.get('position'), "warehouse": row.get('warehouse'),
                        "is_external": row.get('is_external'),
                    })
        except Exception as e:
            csv_matches = [{"error": str(e)}]

    # 2) iz siluxar API (apistockexport) — poišči iste SKU/ID
    api_matches = []
    api_error = None
    api_total = 0
    key = os.environ.get("SILUXAR_STOCK_KEY", "")
    basic_user = os.environ.get("SILUXAR_BASIC_USER", "")
    basic_pass = os.environ.get("SILUXAR_BASIC_PASS", "")
    headers = {}
    _auth = None
    if key: headers["Authorization"] = key
    elif basic_user or basic_pass: _auth = httpx.BasicAuth(basic_user, basic_pass)
    try:
        async with httpx.AsyncClient(timeout=90, auth=_auth) as cli:
            r = await cli.get("https://www.siluxar.si/apistockexport", headers=headers)
        jd = json.loads(r.text or "[]")
        if isinstance(jd, dict):
            for kk in ("data", "items", "rows", "products", "stock"):
                if isinstance(jd.get(kk), list):
                    jd = jd[kk]; break
        api_total = 0
        if isinstance(jd, list):
            api_total = len(jd)
            for it in jd:
                if not isinstance(it, dict): continue
                sku = str(it.get('sku') or it.get('product_sku') or '').strip()
                iid = str(it.get('id') or '').strip()
                pid = str(it.get('product_id') or '').strip()   # NOVO polje v API
                if ql in (sku.lower(), iid.lower(), pid.lower()):
                    api_matches.append({
                        "id": iid, "product_id": pid, "sku": sku, "stock": it.get('stock'),
                        "position": it.get('position'), "source": it.get('source'),
                        "price_netto": it.get('price_netto'),
                    })
    except Exception as e:
        api_error = str(e)

    return {
        "query": q,
        "v_suban_csv": csv_matches or "NI v shranjeni zalogi",
        "csv_count": len(csv_matches),
        "v_siluxar_api": api_matches or "NI v siluxar API odgovoru",
        "api_count": len(api_matches),
        "api_total_pregledanih": api_total,
        "api_error": api_error,
        "namig": "Primerjaj SKU natančno (presledki, velike/male črke) + poglej, ali API vrne stock=0 ali več vrstic.",
    }


@app.get("/zaloga-sync-raw")
async def zaloga_sync_raw():
    """Diagnostika: pokaže SUROV odgovor apistockexport (da vidimo imena polj zaloge)."""
    key = os.environ.get("SILUXAR_STOCK_KEY", "")
    basic_user = os.environ.get("SILUXAR_BASIC_USER", "")
    basic_pass = os.environ.get("SILUXAR_BASIC_PASS", "")
    url = "https://www.siluxar.si/apistockexport"
    headers = {}
    _auth = None
    if key:
        headers["Authorization"] = key
    elif basic_user or basic_pass:
        _auth = httpx.BasicAuth(basic_user, basic_pass)
    try:
        async with httpx.AsyncClient(timeout=90, auth=_auth) as cli:
            r = await cli.get(url, headers=headers)
        text = r.text or ""
        # poskusi razbrati JSON in pokazati KLJUČE prve vrstice (imena polj)
        field_names = None
        source_counts = {}
        total_items = 0
        try:
            jd = json.loads(text)
            if isinstance(jd, dict):
                for kk in ("data", "items", "rows", "products", "stock"):
                    if isinstance(jd.get(kk), list):
                        jd = jd[kk]; break
            if isinstance(jd, list) and jd and isinstance(jd[0], dict):
                field_names = list(jd[0].keys())
                total_items = len(jd)
                # preštej vse 'source' vrednosti (silux / silux2 / ...) — celoten odgovor
                for it in jd:
                    sv = str(it.get("source", "") or "(prazen)").strip() or "(prazen)"
                    source_counts[sv] = source_counts.get(sv, 0) + 1
        except Exception:
            pass
        return {"status": r.status_code, "content_type": r.headers.get("content-type", ""),
                "field_names": field_names, "total_items": total_items,
                "source_values": source_counts, "raw": text[:2000]}
    except Exception as e:
        return {"error": str(e)}


@app.get("/forecast2-sum-raw")
async def forecast2_sum_raw():
    """Diagnostika: pokaže SUROV odgovor apisumexport (da vidimo imena polj)."""
    key = os.environ.get("SILUXAR_STOCK_KEY", "")
    basic_user = os.environ.get("SILUXAR_BASIC_USER", "")
    basic_pass = os.environ.get("SILUXAR_BASIC_PASS", "")
    url = "https://www.siluxar.si/apisumexport"
    headers = {}
    _auth = None
    if key:
        headers["Authorization"] = key
    elif basic_user or basic_pass:
        _auth = httpx.BasicAuth(basic_user, basic_pass)
    try:
        async with httpx.AsyncClient(timeout=60, auth=_auth) as cli:
            r = await cli.get(url, headers=headers)
        return {"status": r.status_code, "content_type": r.headers.get("content-type", ""),
                "raw": (r.text or "")[:2000]}
    except Exception as e:
        return {"error": str(e)}


async def _forecast2_fetch_core(force=False, min_gap_min=55):
    """Jedro: potegne apisumexport in zapiše vmesni vnos za danes.
    Kliče ga endpoint /forecast2-fetch-siluxar (force=True, gumb) IN scheduler (force=False).
    ZAŠČITA: če ni force in je od zadnjega siluxar-vnosa minilo manj kot min_gap_min minut,
    NE zapiše (prepreči podvajanje, če teče več scheduler instanc ali se kliče prepogosto)."""
    today = _lj_today()
    now = _lj_now()
    # — zaščita pred prepogostim zapisom (samo za scheduler, ne za ročni force) —
    if not force:
        try:
            day_check = _forecast2_load_day(today)
            siluxar_entries = [e for e in (day_check.get("entries") or [])
                               if e.get("_source") == "siluxar_apisumexport" and e.get("_ts")]
            if siluxar_entries:
                last_ts = max(e["_ts"] for e in siluxar_entries)
                from datetime import datetime as _dt
                last_dt = _dt.fromisoformat(last_ts)
                gap_min = (now - last_dt).total_seconds() / 60.0
                if gap_min < min_gap_min:
                    return {"ok": True, "skipped": True,
                            "reason": f"Zadnji siluxar-vnos pred {gap_min:.0f} min (< {min_gap_min} min) — preskočim.",
                            "date": today}
        except Exception:
            pass  # ob napaki vseeno nadaljuj (raje zapiši kot izgubi podatek)

    key = os.environ.get("SILUXAR_STOCK_KEY", "")
    basic_user = os.environ.get("SILUXAR_BASIC_USER", "")
    basic_pass = os.environ.get("SILUXAR_BASIC_PASS", "")
    if not key and not (basic_user or basic_pass):
        return {"ok": False, "error": "Manjka SILUXAR_STOCK_KEY (Render okoljska spremenljivka)."}
    url = "https://www.siluxar.si/apisumexport"
    headers = {}
    _auth = None
    if key:
        headers["Authorization"] = key
    elif basic_user or basic_pass:
        _auth = httpx.BasicAuth(basic_user, basic_pass)
    # 1) potegni
    try:
        async with httpx.AsyncClient(timeout=60, auth=_auth) as cli:
            r = await cli.get(url, headers=headers)
    except Exception as e:
        return {"ok": False, "error": f"Napaka pri klicu siluxar.si: {e}"}
    if r.status_code != 200:
        return {"ok": False, "error": f"siluxar.si vrnil status {r.status_code}", "status": r.status_code}
    text = (r.text or "").strip()
    if not text:
        return {"ok": False, "error": "siluxar.si ni vrnil podatkov."}

    # 2) razberi orders + revenue (JSON: objekt ali seznam z enim objektom)
    orders = None
    revenue = None
    try:
        jd = json.loads(text)
        if isinstance(jd, list) and jd:
            jd = jd[0]
        if isinstance(jd, dict):
            def _find(d, *names):
                for n in names:
                    for k in d.keys():
                        if k.strip().lower() == n.lower():
                            return d[k]
                return None
            ov = _find(jd, 'cnt', 'orders', 'order_count', 'count', 'st_narocil', 'narocila', 'num_orders')
            rv = _find(jd, 'sum', 'revenue', 'total', 'promet', 'amount', 'sum_eur', 'revenue_eur')
            if ov is not None:
                orders = int(float(str(ov).replace(',', '.')))
            if rv is not None:
                revenue = float(str(rv).replace(',', '.'))
    except Exception as e:
        return {"ok": False, "error": f"Ne morem razbrati odgovora (ni veljaven JSON): {e}", "raw_preview": text[:200]}

    if orders is None and revenue is None:
        return {"ok": False, "error": "V odgovoru ne najdem polj za naročila/promet.", "raw_preview": text[:200]}

    # 3) zapiši kot VMESNI VNOS (entry) za današnji dan — kot ročni vnos, ne final
    today = _lj_today()
    now = _lj_now()
    time_norm = now.strftime("%H:%M")
    day = _forecast2_load_day(today)
    if "entries" not in day or not isinstance(day.get("entries"), list):
        day["entries"] = []
    day["entries"] = [e for e in day["entries"] if e.get("time") != time_norm]
    day["entries"].append({
        "time": time_norm,
        "orders": orders or 0,
        "revenue": revenue or 0.0,
        "_ts": now.isoformat(),
        "_source": "siluxar_apisumexport",
    })
    day["entries"].sort(key=lambda e: e.get("time", ""))
    _forecast2_save_day(today, day)
    return {"ok": True, "date": today, "time": time_norm,
            "orders": orders or 0, "revenue": revenue or 0.0}


@app.post("/forecast2-fetch-siluxar")
async def forecast2_fetch_siluxar():
    """Potegne dnevno vsoto (št. naročil + promet €) s siluxar apisumexport in
    zapiše vmesni vnos za danes. ROČNI klic prek gumba → force=True (vedno zapiše)."""
    return await _forecast2_fetch_core(force=True)


async def _forecast2_scheduler_loop():
    """Notranji scheduler (always-on Render Pro): preveri vsakih 5 min in zapiše
    vmesni vnos NAJVEČ 1× na uro (ko v tekoči uri še ni siluxar-vnosa).
    Aktiven 6:00–23:00 po Ljubljani. Ta pristop je odporen na restarte: ritem ni
    vezan na minuto zagona, ampak na koledarsko uro — če en poskus preskoči
    (npr. svež ročni vnos), naslednji čez 5 min lahko zapiše."""
    global _FORECAST2_CRON_STATUS
    await asyncio.sleep(60)  # počakaj, da se startup dokonča
    CHECK_INTERVAL = 5 * 60  # preveri vsakih 5 min
    _FORECAST2_CRON_STATUS["started_at"] = _lj_now().isoformat()
    while True:
        try:
            now = _lj_now()
            hour = now.hour
            _FORECAST2_CRON_STATUS["last_tick"] = now.isoformat()
            _FORECAST2_CRON_STATUS["tick_count"] = _FORECAST2_CRON_STATUS.get("tick_count", 0) + 1
            if 6 <= hour < 23:
                # ali v TEKOČI koledarski uri že obstaja siluxar-vnos?
                already_this_hour = False
                try:
                    day_chk = _forecast2_load_day(_lj_today())
                    for e in (day_chk.get("entries") or []):
                        if e.get("_source") == "siluxar_apisumexport" and e.get("_ts"):
                            from datetime import datetime as _dt
                            ets = _dt.fromisoformat(e["_ts"])
                            if ets.hour == hour and ets.date() == now.date():
                                already_this_hour = True
                                break
                except Exception:
                    pass
                if already_this_hour:
                    _FORECAST2_CRON_STATUS["last_result"] = {"skipped": True, "reason": f"v uri {hour}h že obstaja vnos"}
                else:
                    # uporabi krajšo zaščito (10 min) — prepreči podvajanje, a ne požre urnega vnosa
                    res = await _forecast2_fetch_core(force=False, min_gap_min=10)
                    _FORECAST2_CRON_STATUS["last_run"] = now.isoformat()
                    _FORECAST2_CRON_STATUS["last_result"] = res
                    if res.get("skipped"):
                        print(f"[forecast2-cron] preskočeno — {res.get('reason')}")
                    elif res.get("ok"):
                        print(f"[forecast2-cron] OK {res.get('date')} {res.get('time')} — "
                              f"{res.get('orders')} naročil, {res.get('revenue')} €")
                    else:
                        print(f"[forecast2-cron] FAIL — {res.get('error')}")
            else:
                _FORECAST2_CRON_STATUS["last_result"] = {"skipped": True, "reason": f"izven ur ({hour}h)"}
            await asyncio.sleep(CHECK_INTERVAL)
        except Exception as e:
            _FORECAST2_CRON_STATUS["last_error"] = {"at": _lj_now().isoformat(), "error": str(e)}
            print(f"[forecast2-cron] Error: {e}")
            await asyncio.sleep(600)  # 10 min pred ponovnim poskusom


@app.get("/forecast2-cron-status")
async def forecast2_cron_status():
    """Diagnostika cron-a: kdaj se je nazadnje pognal, kaj je vrnil, kdaj naslednjič.
    Pomaga ugotoviti zakaj manjka kak urni vnos (restart / napaka / preskok)."""
    today = _lj_today()
    day = _forecast2_load_day(today)
    siluxar_entries = [{"time": e.get("time"), "orders": e.get("orders"), "_ts": e.get("_ts")}
                       for e in (day.get("entries") or [])
                       if e.get("_source") == "siluxar_apisumexport"]
    return {
        "ok": True,
        "zdaj_lj": _lj_now().isoformat(),
        "status": _FORECAST2_CRON_STATUS,
        "danes_siluxar_vnosi": siluxar_entries,
        "stevilo_danes": len(siluxar_entries),
    }


@app.post("/forecast2-set-final")
async def forecast2_set_final(data: dict):
    """Nastavi končna naročila/promet za določen dan."""
    try:
        date_iso = data.get("date", "")
        orders = int(data.get("orders", 0))
        revenue = float(data.get("revenue", 0))

        if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_iso):
            return {"ok": False, "error": "Invalid date format (YYYY-MM-DD)"}

        day = _forecast2_load_day(date_iso)
        day["final"] = {
            "orders": orders,
            "revenue": revenue,
            "_set_at": _lj_now().isoformat(),
        }
        _forecast2_save_day(date_iso, day)
        return {"ok": True, "date": date_iso, "final": day["final"]}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/forecast2-bulk-import")
async def forecast2_bulk_import(data: dict):
    """Bulk uvoz/posodobitev končnih dnevnih podatkov iz prilepljenega teksta.
    Sprejme različne formate (TAB/presledek ločeno):
      01.01.2026	980	€ 18.552,94
      2026-01-01  980  18552.94
    Prepiše obstoječe dni (posodobitev za storno/neprevzete pakete).
    """
    try:
        text = data.get("text", "") or ""
        if not text.strip():
            return {"ok": False, "error": "Prazen tekst."}

        imported = 0
        skipped = 0
        errors = []
        sample = []

        for raw_line in text.split('\n'):
            line = raw_line.strip()
            if not line:
                continue
            # Preskoči glave/seštevke
            low = line.lower()
            if low.startswith('dan') or low.startswith('skupaj') or low.startswith('total') or low.startswith('datum'):
                continue

            # Razdeli po tabih ALI 2+ presledkih
            parts = re.split(r'\t+|\s{2,}', line)
            if len(parts) < 3:
                # morda en presledek loči — poskusi splitati na 3 dele od konca
                parts = line.split()
                if len(parts) < 3:
                    skipped += 1
                    continue

            date_raw = parts[0].strip()

            # Datum: DD.MM.YYYY ali YYYY-MM-DD
            iso = None
            m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$', date_raw)
            if m:
                iso = f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
            elif re.match(r'^\d{4}-\d{2}-\d{2}$', date_raw):
                iso = date_raw
            if not iso:
                skipped += 1
                continue

            # Naročila — drugi stolpec, samo številke
            orders_raw = re.sub(r'[^\d]', '', parts[1])
            orders = int(orders_raw) if orders_raw else 0

            # Promet — zadnji stolpec (lahko vsebuje € in presledke)
            rev_raw = " ".join(parts[2:]).replace('€', '').strip()
            # Slovenski format: 18.552,94 (. tisočica, , decimalka)
            # Angleški: 18552.94
            if ',' in rev_raw and '.' in rev_raw:
                # oba — predpostavi . = tisočica, , = decimalka
                rev_clean = rev_raw.replace('.', '').replace(',', '.')
            elif ',' in rev_raw:
                # samo vejica = decimalka
                rev_clean = rev_raw.replace(',', '.')
            else:
                rev_clean = rev_raw
            rev_clean = re.sub(r'[^\d.]', '', rev_clean)
            try:
                revenue = float(rev_clean) if rev_clean else 0.0
            except ValueError:
                errors.append(f"{iso}: neveljaven promet '{rev_raw}'")
                skipped += 1
                continue

            # Shrani (prepiše obstoječ final)
            day = _forecast2_load_day(iso)
            day["final"] = {
                "orders": orders,
                "revenue": revenue,
                "_set_at": _lj_now().isoformat(),
                "_source": "bulk_import",
            }
            _forecast2_save_day(iso, day)
            imported += 1
            if len(sample) < 3:
                sample.append({"date": iso, "orders": orders, "revenue": revenue})

        return {
            "ok": True,
            "imported": imported,
            "skipped": skipped,
            "errors": errors[:10],
            "sample": sample,
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}

# ─── SCALING RECOMMENDER STATE ───────────────────────────────────────────────
def _scaling_state_load() -> dict:
    if SCALING_STATE_FILE.exists():
        try:
            return json.loads(SCALING_STATE_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"results": None, "saved_at": None, "scaled": {}}

def _scaling_state_save(state: dict):
    SCALING_STATE_FILE.write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")

@app.get("/scaling-state")
async def scaling_state_get():
    """Vrne shranjene scaling rezultate + scale zgodovino (skalirane kampanje)."""
    try:
        return {"ok": True, **_scaling_state_load()}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/scaling-state-save")
async def scaling_state_save(data: dict):
    """Shrani trenutne scaling rezultate (po kliku Izberi kandidate)."""
    try:
        state = _scaling_state_load()
        state["results"] = data.get("results")
        state["saved_at"] = _lj_now().isoformat()
        _scaling_state_save(state)
        return {"ok": True, "saved_at": state["saved_at"]}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/scaling-state-mark")
async def scaling_state_mark(data: dict):
    """Označi/odznači kampanjo kot skalirano. Ključ = SKU (obstojno čez sezone)."""
    try:
        sku = (data.get("sku") or "").strip()
        if not sku:
            return {"ok": False, "error": "Manjka sku."}
        scaled_flag = bool(data.get("scaled", True))
        state = _scaling_state_load()
        if "scaled" not in state or not isinstance(state["scaled"], dict):
            state["scaled"] = {}
        if scaled_flag:
            state["scaled"][sku] = {
                "scaled_at": _lj_now().isoformat(),
                "note": data.get("note", ""),
            }
        else:
            state["scaled"].pop(sku, None)
        _scaling_state_save(state)
        return {"ok": True, "sku": sku, "scaled": scaled_flag, "total_scaled": len(state["scaled"])}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/scaling-state-archive")
async def scaling_state_archive(data: dict):
    """Arhivira komplet seznam — počisti shranjene rezultate (gumb Izberi kandidate spet viden).
    Scale zgodovina (skalirane kampanje) OSTANE za prepoznavo v prihodnje."""
    try:
        state = _scaling_state_load()
        keep_scaled = state.get("scaled", {})
        if data.get("clear_scaled"):
            keep_scaled = {}  # počisti tudi zgodovino če izrecno zahtevano
        _scaling_state_save({"results": None, "saved_at": None, "scaled": keep_scaled})
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/forecast2-history")
async def forecast2_history(days: int = 60):
    """Vrne zgodovino zadnjih N dni — vsak dan z entries + final."""
    try:
        from datetime import timedelta
        today = _lj_now().date()
        result = {}
        for i in range(days):
            d = today - timedelta(days=i)
            iso = d.strftime("%Y-%m-%d")
            data = _forecast2_load_day(iso)
            if data["entries"] or data["final"]:
                result[iso] = data
        return {"ok": True, "days": result}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/forecast2-day/{date_iso}")
async def forecast2_day(date_iso: str):
    """Vrne podrobnosti za določen dan."""
    try:
        data = _forecast2_load_day(date_iso)
        return {"ok": True, **data}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/forecast2-stats")
async def forecast2_stats():
    """Statistika — koliko dni imamo final podatke."""
    try:
        from datetime import timedelta
        today = _lj_now().date()
        days_with_final = 0
        days_with_entries = 0
        finals = []
        for i in range(90):
            d = today - timedelta(days=i)
            iso = d.strftime("%Y-%m-%d")
            data = _forecast2_load_day(iso)
            if data["final"]:
                days_with_final += 1
                finals.append({
                    "date": iso,
                    "orders": data["final"].get("orders", 0),
                    "revenue": data["final"].get("revenue", 0),
                })
            if data["entries"]:
                days_with_entries += 1

        # Povprečje zadnjih 7 in 22 dni
        recent_7 = finals[:7]
        recent_22 = finals[:22]

        return {
            "ok": True,
            "today": today.strftime("%Y-%m-%d"),
            "days_with_final": days_with_final,
            "days_with_entries": days_with_entries,
            "avg_7d_orders": round(sum(f["orders"] for f in recent_7) / max(1, len(recent_7)), 1),
            "avg_7d_revenue": round(sum(f["revenue"] for f in recent_7) / max(1, len(recent_7)), 0),
            "avg_22d_orders": round(sum(f["orders"] for f in recent_22) / max(1, len(recent_22)), 1),
            "avg_22d_revenue": round(sum(f["revenue"] for f in recent_22) / max(1, len(recent_22)), 0),
            "finals_count": len(finals),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/forecast2-bulk-finals")
async def forecast2_bulk_finals(data: dict):
    """Bulk import končnih podatkov za več dni naenkrat.
    Body: {"finals": [{"date":"YYYY-MM-DD","orders":N,"revenue":N}, ...]}
    """
    try:
        finals = data.get("finals", [])
        saved = 0
        errors = []
        for item in finals:
            date_iso = item.get("date", "")
            if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_iso):
                errors.append(f"Invalid date: {date_iso}")
                continue
            try:
                orders = int(item.get("orders", 0))
                revenue = float(item.get("revenue", 0))
                day = _forecast2_load_day(date_iso)
                day["final"] = {
                    "orders": orders,
                    "revenue": revenue,
                    "_set_at": _lj_now().isoformat(),
                    "_bulk": True,
                }
                _forecast2_save_day(date_iso, day)
                saved += 1
            except Exception as e:
                errors.append(f"{date_iso}: {e}")

        return {"ok": True, "saved": saved, "errors": errors}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/forecast2-cleanup")
async def forecast2_cleanup():
    """Počisti phantom entries v vseh dnevnih fajlih."""
    try:
        cleaned_days = 0
        removed_entries = 0
        for fp in FORECAST2_DIR.glob("*.json"):
            try:
                data = json.loads(fp.read_text(encoding="utf-8"))
                entries = data.get("entries", [])
                valid = [e for e in entries if e and isinstance(e, dict) and e.get("time") and "orders" in e]
                if len(valid) != len(entries):
                    data["entries"] = valid
                    fp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
                    cleaned_days += 1
                    removed_entries += (len(entries) - len(valid))
            except Exception as e:
                print(f"[fc2 cleanup] error {fp}: {e}")
        return {"ok": True, "cleaned_days": cleaned_days, "removed_entries": removed_entries}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ═══════════════════════════════════════════════
# VIDEO OUTPUTS DISK STORAGE (ad localization workflow)
# ═══════════════════════════════════════════════
VIDEO_OUTPUTS_DIR = Path("/data/video_outputs")
VIDEO_OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)


def _safe_sku(name: str) -> str:
    """Sanitize SKU name for filesystem."""
    import re as _re
    s = _re.sub(r'[^A-Za-z0-9_\-]', '', name).upper()
    return s or 'UNKNOWN'


@app.post("/video-save-to-disk")
async def video_save_to_disk(
    video: UploadFile = File(...),
    sku_name: str = Form(...),
    country_code: str = Form(...),
    video_version: int = Form(1)
):
    """Shrani merged video v /data/video_outputs/{SKU}/video_{lang}_v{n}.mp4
    Returns: filename, download_url"""
    try:
        sku = _safe_sku(sku_name)
        lang = country_code.lower().strip()
        if not lang:
            return {"ok": False, "error": "Missing country_code"}

        sku_dir = VIDEO_OUTPUTS_DIR / sku
        sku_dir.mkdir(parents=True, exist_ok=True)

        filename = f"{sku}_{lang}_v{video_version}.mp4"
        target = sku_dir / filename

        content_bytes = await video.read()
        target.write_bytes(content_bytes)

        return {
            "ok": True,
            "filename": filename,
            "size": len(content_bytes),
            "download_url": f"/video-batch/{sku}/{filename}"
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/video-batch/{sku}")
async def video_batch_meta(sku: str):
    """List all videos for a SKU."""
    sku_safe = _safe_sku(sku)
    sku_dir = VIDEO_OUTPUTS_DIR / sku_safe
    if not sku_dir.exists():
        return {"ok": False, "error": "SKU not found"}

    videos = []
    for f in sorted(sku_dir.glob("*.mp4")):
        videos.append({
            "filename": f.name,
            "size": f.stat().st_size,
            "size_mb": round(f.stat().st_size / 1024 / 1024, 2),
            "url": f"/video-batch/{sku_safe}/{f.name}",
            "modified": f.stat().st_mtime
        })
    return {"ok": True, "sku": sku_safe, "videos": videos, "count": len(videos)}


@app.get("/video-batch/{sku}/zip-all")
async def video_batch_zip(sku: str):
    """Download all videos for SKU as ZIP."""
    import zipfile
    import io
    sku_safe = _safe_sku(sku)
    sku_dir = VIDEO_OUTPUTS_DIR / sku_safe
    if not sku_dir.exists():
        raise HTTPException(status_code=404, detail="SKU not found")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in sorted(sku_dir.glob("*.mp4")):
            zf.write(f, arcname=f.name)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([buf.read()]),
        media_type='application/zip',
        headers={'Content-Disposition': f'attachment; filename="{sku_safe}_videos.zip"'}
    )


@app.get("/video-batch/{sku}/country/{country_code}/zip")
async def video_batch_country_zip(sku: str, country_code: str):
    """Download all videos for one country (as ZIP)."""
    import zipfile
    import io
    sku_safe = _safe_sku(sku)
    sku_dir = VIDEO_OUTPUTS_DIR / sku_safe
    if not sku_dir.exists():
        raise HTTPException(status_code=404, detail="SKU not found")

    lang = country_code.lower()
    # Podpora za stari format (video_xx_v*.mp4) IN novi format (SKU_xx_v*.mp4)
    files_old = list(sku_dir.glob(f"video_{lang}_v*.mp4"))
    files_new = list(sku_dir.glob(f"{sku_safe}_{lang}_v*.mp4"))
    files = sorted(set(files_old + files_new))
    if not files:
        raise HTTPException(status_code=404, detail=f"No videos for country {country_code}")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in files:
            zf.write(f, arcname=f.name)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([buf.read()]),
        media_type='application/zip',
        headers={'Content-Disposition': f'attachment; filename="{sku_safe}_{lang}_videos.zip"'}
    )



@app.get("/video-batch/{sku}/{filename}")
async def video_batch_file(sku: str, filename: str, request: Request = None):
    """Serve individual MP4 file for download with HTTP Range support (fast partial downloads)."""
    sku_safe = _safe_sku(sku)
    # Sanitize filename
    if '/' in filename or '\\' in filename or '..' in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    target = VIDEO_OUTPUTS_DIR / sku_safe / filename
    if not target.exists():
        raise HTTPException(status_code=404, detail="File not found")

    file_size = target.stat().st_size
    range_header = request.headers.get('range') if request else None

    # Async streaming with bigger chunks (1MB instead of default 64KB)
    CHUNK_SIZE = 1024 * 1024  # 1MB

    if range_header:
        # Parse "bytes=START-END"
        try:
            range_str = range_header.replace('bytes=', '').strip()
            start_str, end_str = range_str.split('-')
            start = int(start_str) if start_str else 0
            end = int(end_str) if end_str else file_size - 1
            end = min(end, file_size - 1)
        except Exception:
            start, end = 0, file_size - 1

        async def range_iter():
            with open(target, 'rb') as f:
                f.seek(start)
                remaining = end - start + 1
                while remaining > 0:
                    chunk = f.read(min(CHUNK_SIZE, remaining))
                    if not chunk:
                        break
                    remaining -= len(chunk)
                    yield chunk

        headers = {
            'Content-Range': f'bytes {start}-{end}/{file_size}',
            'Accept-Ranges': 'bytes',
            'Content-Length': str(end - start + 1),
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Cache-Control': 'public, max-age=3600',
        }
        return StreamingResponse(
            range_iter(),
            status_code=206,
            media_type='video/mp4',
            headers=headers,
        )

    # Full download with streaming
    async def file_iter():
        with open(target, 'rb') as f:
            while True:
                chunk = f.read(CHUNK_SIZE)
                if not chunk:
                    break
                yield chunk

    headers = {
        'Content-Length': str(file_size),
        'Accept-Ranges': 'bytes',
        'Content-Disposition': f'attachment; filename="{filename}"',
        'Cache-Control': 'public, max-age=3600',
    }
    return StreamingResponse(
        file_iter(),
        media_type='video/mp4',
        headers=headers,
    )



@app.delete("/video-batch/{sku}")
async def video_batch_delete(sku: str):
    """Delete all videos for SKU."""
    import shutil
    sku_safe = _safe_sku(sku)
    sku_dir = VIDEO_OUTPUTS_DIR / sku_safe
    if not sku_dir.exists():
        return {"ok": False, "error": "SKU not found"}
    try:
        shutil.rmtree(sku_dir)
        return {"ok": True, "deleted": sku_safe}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/storage-list")
async def storage_list():
    """List all SKU folders in /data/video_outputs/ with stats."""
    try:
        skus = []
        total_size = 0
        total_videos = 0
        for sku_dir in sorted(VIDEO_OUTPUTS_DIR.iterdir()):
            if not sku_dir.is_dir():
                continue
            videos = []
            sku_size = 0
            for f in sorted(sku_dir.glob("*.mp4")):
                size = f.stat().st_size
                sku_size += size
                videos.append({
                    "filename": f.name,
                    "size": size,
                    "size_mb": round(size / 1024 / 1024, 2),
                    "modified": f.stat().st_mtime
                })
            if not videos:
                continue
            # Last modified time of SKU
            last_mtime = max(v['modified'] for v in videos)
            skus.append({
                "sku": sku_dir.name,
                "video_count": len(videos),
                "size_mb": round(sku_size / 1024 / 1024, 2),
                "modified": last_mtime,
                "videos": videos
            })
            total_size += sku_size
            total_videos += len(videos)

        # Sort by most recently modified first
        skus.sort(key=lambda s: -s['modified'])

        # Disk usage (Render disk = 5GB or whatever)
        import shutil as _shutil
        try:
            stat = _shutil.disk_usage(str(VIDEO_OUTPUTS_DIR))
            disk_total_mb = round(stat.total / 1024 / 1024, 1)
            disk_used_mb = round((stat.total - stat.free) / 1024 / 1024, 1)
            disk_free_mb = round(stat.free / 1024 / 1024, 1)
        except Exception:
            disk_total_mb = disk_used_mb = disk_free_mb = 0

        return {
            "ok": True,
            "skus": skus,
            "stats": {
                "sku_count": len(skus),
                "total_videos": total_videos,
                "total_size_mb": round(total_size / 1024 / 1024, 1),
                "disk_total_mb": disk_total_mb,
                "disk_used_mb": disk_used_mb,
                "disk_free_mb": disk_free_mb,
            }
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


class StorageBulkDeleteRequest(BaseModel):
    skus: list[str]


@app.post("/storage-bulk-delete")
async def storage_bulk_delete(req: StorageBulkDeleteRequest):
    """Delete multiple SKU folders at once."""
    import shutil
    deleted = []
    errors = []
    for sku in req.skus:
        sku_safe = _safe_sku(sku)
        sku_dir = VIDEO_OUTPUTS_DIR / sku_safe
        try:
            if sku_dir.exists():
                shutil.rmtree(sku_dir)
                deleted.append(sku_safe)
        except Exception as e:
            errors.append(f"{sku_safe}: {e}")
    return {"ok": True, "deleted": deleted, "errors": errors}


@app.get("/d/{sku}", response_class=HTMLResponse)
async def video_download_page(sku: str):
    """Public download page for SKU videos - what advertiser sees."""
    sku_safe = _safe_sku(sku)
    sku_dir = VIDEO_OUTPUTS_DIR / sku_safe
    if not sku_dir.exists():
        return HTMLResponse(content=f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Ni videov za {sku_safe}</title>
<style>body{{font-family:system-ui,sans-serif;background:#fafbfd;color:#0f172a;padding:3rem;text-align:center}}</style></head>
<body><h1>📁 Ni videov za SKU: {sku_safe}</h1>
<p>SKU mapa še ne obstaja ali so videi že bili izbrisani.</p></body></html>""", status_code=404)

    # Group videos by country
    flag_map = {
        'si': ('🇸🇮', 'Slovenija'), 'hr': ('🇭🇷', 'Hrvaška'), 'rs': ('🇷🇸', 'Srbija'),
        'hu': ('🇭🇺', 'Madžarska'), 'pl': ('🇵🇱', 'Poljska'), 'cz': ('🇨🇿', 'Češka'),
        'sk': ('🇸🇰', 'Slovaška'), 'gr': ('🇬🇷', 'Grčija'), 'bg': ('🇧🇬', 'Bolgarija'),
        'ro': ('🇷🇴', 'Romunija'), 'it': ('🇮🇹', 'Italija'), 'de': ('🇩🇪', 'Nemčija'),
        'at': ('🇦🇹', 'Avstrija'), 'ba': ('🇧🇦', 'BiH'), 'mk': ('🇲🇰', 'N. Makedonija')
    }

    import re as _re
    by_country = {}
    all_videos = []
    for f in sorted(sku_dir.glob("*.mp4")):
        size_mb = round(f.stat().st_size / 1024 / 1024, 2)
        url = f"/video-batch/{sku_safe}/{f.name}"
        # Podpora za stari format (video_xx_v<n>.mp4) IN novi format (SKU_xx_v<n>.mp4)
        # Vzami zadnje _xx_v<n>.mp4 (jezik + verzija) iz katerega koli prefiksa
        m = _re.search(r'_([a-z]{2})_v(\d+)\.mp4$', f.name)
        lang = m.group(1) if m else 'xx'
        version = int(m.group(2)) if m else 0
        video_info = {"filename": f.name, "size_mb": size_mb, "url": url, "version": version}
        all_videos.append({**video_info, "lang": lang})
        by_country.setdefault(lang, []).append(video_info)

    if not all_videos:
        return HTMLResponse(content=f"<h1>Ni videov za {sku_safe}</h1>", status_code=404)

    country_order = sorted(by_country.keys(), key=lambda l: flag_map.get(l, ('', l))[1])
    total_size_mb = round(sum(v['size_mb'] for v in all_videos), 1)
    total_count = len(all_videos)
    country_count = len(country_order)

    def country_data(lang):
        flag, name = flag_map.get(lang, ('🌐', lang.upper()))
        vids = by_country[lang]
        size = round(sum(v['size_mb'] for v in vids), 2)
        zip_url = f"/video-batch/{sku_safe}/country/{lang}/zip"
        return {"lang": lang, "flag": flag, "name": name, "videos": vids, "size_mb": size, "zip_url": zip_url, "count": len(vids)}

    countries = [country_data(l) for l in country_order]

    cards = ""
    for c in countries:
        vids_html = ""
        for v in c['videos']:
            vids_html += f'<a href="{v["url"]}" download="{v["filename"]}" class="vmini">v{v["version"]}</a>'
        cards += f"""
        <div class="ccard">
          <div class="cflag">{c['flag']}</div>
          <div class="cname">{c['name']}</div>
          <div class="cmeta">{c['count']} videi · {c['size_mb']} MB</div>
          <a href="{c['zip_url']}" class="cdown">⬇ Prenesi vse ({c['count']})</a>
          <div class="cvids">{vids_html}</div>
        </div>"""

    return HTMLResponse(content=f"""<!DOCTYPE html>
<html lang="sl"><head><meta charset="utf-8">
<title>📁 {sku_safe} — Video Ads</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  *{{box-sizing:border-box;margin:0;padding:0;-webkit-font-smoothing:antialiased}}
  body{{font-family:'DM Sans',-apple-system,system-ui,sans-serif;color:#0f172a;background:#fafbfd;min-height:100vh}}
  .container{{max-width:1100px;margin:0 auto;padding:2rem 1.25rem}}
  .pagehead{{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:1rem;padding:0 0 2rem;border-bottom:1px solid #e2e8f2;margin-bottom:2rem}}
  .pagehead h1{{font-size:28px;font-weight:800;letter-spacing:-0.03em}}
  .pagehead .sub{{font-size:13px;color:#64748b;margin-top:4px}}
  .zipall{{background:#7c3aed;color:white;padding:11px 20px;border-radius:9px;font-size:13px;font-weight:600;text-decoration:none;display:inline-flex;align-items:center;gap:6px;transition:all 0.15s}}
  .zipall:hover{{background:#6d28d9;transform:translateY(-1px)}}
  .grid-a{{display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:14px}}
  .ccard{{background:white;border:1px solid #e2e8f2;border-radius:14px;padding:1.5rem 1.25rem;text-align:center;transition:all 0.2s;display:flex;flex-direction:column;gap:8px}}
  .ccard:hover{{border-color:#4f6ef7;transform:translateY(-2px);box-shadow:0 8px 20px rgba(79,110,247,0.1)}}
  .cflag{{font-size:48px;line-height:1;margin-bottom:6px}}
  .cname{{font-size:16px;font-weight:700;color:#0f172a}}
  .cmeta{{font-size:11px;color:#8896b0;margin-bottom:8px}}
  .cdown{{background:#4f6ef7;color:white;padding:10px 14px;border-radius:8px;font-size:13px;font-weight:600;text-decoration:none;transition:background 0.15s}}
  .cdown:hover{{background:#3f5dd6}}
  .cvids{{display:flex;flex-wrap:wrap;gap:4px;justify-content:center;margin-top:8px}}
  .vmini{{font-size:10px;color:#64748b;text-decoration:none;border:1px solid #e2e8f2;border-radius:5px;padding:3px 7px;font-family:monospace;transition:all 0.1s}}
  .vmini:hover{{border-color:#4f6ef7;color:#4f6ef7;background:#eff3ff}}
  .footer{{text-align:center;margin-top:3rem;padding:1.5rem;font-size:11px;color:#94a3b8;border-top:1px solid #e2e8f2}}
</style></head>
<body>
<div class="container">
  <div class="pagehead">
    <div>
      <h1>📁 {sku_safe}</h1>
      <div class="sub">{country_count} držav · {total_count} videov · {total_size_mb} MB · Maaarket Ads</div>
    </div>
    <a href="/video-batch/{sku_safe}/zip-all" class="zipall">📦 Prenesi vse kot ZIP</a>
  </div>
  <div class="grid-a">{cards}</div>
  <div class="footer">ads.slxanalytics.org · Maaarket Video Ads</div>
</div>
</body></html>""")


# ═══════════════════════════════════════════════
# PREVZEMI — PDF invoice parser → XLS generator
# ═══════════════════════════════════════════════
PREVZEMI_DIR = Path("/data/prevzemi")
PREVZEMI_DIR.mkdir(parents=True, exist_ok=True)


def _clean_motoprofil_product_code(code: str) -> str:
    """Odstrani MotoProfil prefiks iz product_code-a.
    MotoProfil ima format '{PREFIKS} {SKU}' kjer je prefiks 2-4 črk (npr. ABS, AMT, WP, IZA, BLP).
    Primeri:
      'ABS SL1348'   → 'SL1348'
      'AMT 62-002'   → '62-002'
      'WP 038'       → '038'
      'IZA 35.025A'  → '35.025A'
      'BLP ADM55368' → 'ADM55368'
    Če format ne ustreza (ni prefiks + presledek), vrne original.
    """
    if not code:
        return code
    import re as _re
    s = str(code).strip()
    # Odreži samo če se začne z 2-4 VELIKIMI črkami, sledi presledek, nato SKU
    m = _re.match(r'^[A-Z]{2,4}\s+(\S.*)$', s)
    if m:
        return m.group(1).strip()
    return s


def _apply_motoprofil_prefix_cleanup(parsed: dict) -> dict:
    """Če je dobavitelj MotoProfil, očisti prefikse iz vseh product_number-jev."""
    if not parsed or not isinstance(parsed, dict):
        return parsed
    items = parsed.get("items", [])
    cleaned_count = 0
    for it in items:
        old_code = it.get("product_number", "")
        new_code = _clean_motoprofil_product_code(old_code)
        if new_code != old_code:
            it["product_number"] = new_code
            cleaned_count += 1
    if cleaned_count:
        print(f"[motoprofil] Očiščenih {cleaned_count} prefiksov iz product_number")
    return parsed


def _safe_filename(name: str) -> str:
    import re as _re
    s = _re.sub(r'[^A-Za-z0-9_\-]', '_', name)
    return s[:120] or 'prevzem'


def _normalize_invoice_number(num: str) -> str:
    """Normalizira invoice številko — odstrani presledke, ohrani strukturo.
    Primer: 'FA 162383/05/2026/U' → 'FA_162383/05/2026/U'
    """
    if not num:
        return num
    import re as _re
    s = str(num).strip()
    # Vsi presledki (in tabulatorji, newline) → _
    s = _re.sub(r'\s+', '_', s)
    return s


def _override_supplier_by_invoice(supplier: str, invoice_number: str, vendor_name: str = "") -> str:
    """Override supplier-ja na podlagi pattern-a invoice številke ali vendor name-a.
    Kliče se PRED shranjevanjem da popravi napake AI parser-ja.

    Pravila:
    - Če invoice začne s "FA " (s presledkom) ali "FA_" (po normalizaciji) → MotoProfil
    - Če vendor_name vsebuje "moto-profil" ali "motoprofil" → MotoProfil
    - Sicer obdrži original supplier
    """
    if not invoice_number and not vendor_name:
        return supplier
    import re as _re
    inv = str(invoice_number or "").strip()
    vname = (vendor_name or "").lower()

    # 1. Vendor name detection (najmočnejši signal)
    if "moto-profil" in vname or "motoprofil" in vname or "moto profil" in vname:
        return "motoprofil"
    if "amio" in vname or "knurowska" in vname or "suban" in vname:
        return "amio"
    if "abakus" in vname:
        return "abakus"
    if "inter cars" in vname or "intercars" in vname or "inter-cars" in vname:
        return "intercars"
    if "ikonka" in vname or vname.startswith("kik "):
        return "ikonka"

    # 2. Invoice pattern: MotoProfil ima "FA XXXX/YY/YYYY/U" (s presledkom ali _)
    # ali "FA_XXXX_YY_YYYY" iz CSV imena
    if _re.match(r'^FA[\s_]\d+[/_]', inv, _re.IGNORECASE):
        return "motoprofil"

    return supplier


def _repair_truncated_invoice_json(raw: str) -> Optional[dict]:
    """Poskusi rešiti odrezan invoice JSON (max_tokens prekoračen).
    Strategija: najdi zadnji KOMPLETEN item v items[] array-ju, odreži ostalo, zapri JSON.
    """
    if not raw:
        return None
    import re as _re

    # Odstrani markdown fences
    text = raw.strip()
    text = _re.sub(r'^```(?:json)?\s*', '', text)
    text = _re.sub(r'\s*```$', '', text)

    # Najdi začetek items array-ja
    items_match = _re.search(r'"items"\s*:\s*\[', text)
    if not items_match:
        return None

    # Pridobi header del (invoice_number, invoice_date, ipd. pred items)
    header_part = text[:items_match.end()]  # vključno z "items": [

    # Najdi vse komplette item objekte v items array-ju
    # Item je { ... } — najdemo balansirane oklepaje
    items_body = text[items_match.end():]
    complete_items = []
    depth = 0
    current = ""
    in_string = False
    escape = False
    for ch in items_body:
        if escape:
            current += ch
            escape = False
            continue
        if ch == '\\':
            current += ch
            escape = True
            continue
        if ch == '"':
            in_string = not in_string
            current += ch
            continue
        if in_string:
            current += ch
            continue
        # zunaj stringa
        if ch == '{':
            depth += 1
            current += ch
        elif ch == '}':
            depth -= 1
            current += ch
            if depth == 0:
                # Komplette item objekt
                try:
                    item_obj = json.loads(current.strip())
                    complete_items.append(item_obj)
                except Exception:
                    pass
                current = ""
        elif ch == ']' and depth == 0:
            # Konec items array-ja (cel JSON je dejansko OK)
            break
        else:
            if depth > 0:
                current += ch
            # zunaj objekta ignoriramo (presledki, vejice)

    if not complete_items:
        return None

    # Sestavi nov, veljaven JSON
    try:
        repaired = header_part + json.dumps(complete_items, ensure_ascii=False)[1:]  # [1:] da odstranimo začetni [
        repaired += "}"  # zapri root objekt
        result = json.loads(repaired)
        return result
    except Exception:
        # Fallback: sestavi ročno
        try:
            # Izvleci header polja
            inv_num = ""
            inv_date = ""
            m = _re.search(r'"invoice_number"\s*:\s*"([^"]*)"', header_part)
            if m: inv_num = m.group(1)
            m = _re.search(r'"invoice_date"\s*:\s*"([^"]*)"', header_part)
            if m: inv_date = m.group(1)
            m = _re.search(r'"vendor_name"\s*:\s*"([^"]*)"', header_part)
            vendor = m.group(1) if m else ""
            return {
                "invoice_number": inv_num,
                "invoice_date": inv_date,
                "vendor_name": vendor,
                "items": complete_items,
            }
        except Exception:
            return None


@app.post("/prevzemi-parse-pdf")
async def prevzemi_parse_pdf(file: UploadFile = File(...)):
    """Upload PDF invoice, parse with Claude API, save raw + parsed JSON to disk."""
    try:
        import base64
        pdf_bytes = await file.read()
        pdf_b64 = base64.standard_b64encode(pdf_bytes).decode('utf-8')

        # Save raw PDF temporarily for re-processing
        from datetime import datetime
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_name = _safe_filename(file.filename or 'invoice.pdf')

        # Use Claude to parse PDF — V EXECUTOR (da ne blokira event loop / healthz)
        _pdf_prompt = """Parse this supplier invoice PDF and extract structured data. Return ONLY a valid JSON object — no markdown, no explanation, no text before or after.

CRITICAL JSON RULES:
- Use ONLY straight double quotes ("), never curly quotes (" ")
- Escape double quotes inside string values with backslash (\\")
- Escape backslashes in strings (\\\\)
- No trailing commas after last item in arrays or objects
- All keys MUST be in double quotes
- No comments, no extra text

Schema:
{
  "invoice_number": "string (e.g. FV/KK/896/05/2026)",
  "invoice_date": "string in YYYY-MM-DD format",
  "vendor_name": "string",
  "vendor_tin": "string",
  "customer_name": "string",
  "total_value": "string (e.g. 1013.70 EUR)",
  "currency": "string (e.g. EUR)",
  "items": [
    {
      "lp": "string (item number from PDF, e.g. '1', '2')",
      "product_number": "string (P/N code)",
      "product_name": "string (full name, escape any quotes with backslash)",
      "ean": "string (EAN code, 13 digits)",
      "qty": "string (quantity)",
      "unit": "string (szt/kpl/para/set, copy as-is)",
      "vat": "string (e.g. '0%')",
      "unit_price": "string (subtotal price per unit, exactly as in PDF)",
      "value": "string (subtotal value, exactly as in PDF)"
    }
  ]
}

IMPORTANT:
- All numeric values as STRINGS (not numbers!) to preserve formatting
- Use dot as decimal separator (3.32 not 3,32)
- Copy product names EXACTLY as written (Polish/English mixed text is OK)
- Extract ALL items, do not skip any (typically 50-150 items)
- If field missing, use empty string ""
- Return ONLY the JSON object - NOTHING else"""

        loop = asyncio.get_event_loop()
        message = await loop.run_in_executor(None, lambda: client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=20000,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_b64
                        }
                    },
                    {
                        "type": "text",
                        "text": _pdf_prompt
                    }
                ]
            }]
        ))

        raw_text = message.content[0].text.strip()
        # Strip possible markdown
        if raw_text.startswith('```'):
            raw_text = raw_text.split('```', 2)
            raw_text = raw_text[1] if len(raw_text) > 1 else raw_text[0]
            if raw_text.startswith('json'):
                raw_text = raw_text[4:]
            raw_text = raw_text.strip()
        if raw_text.endswith('```'):
            raw_text = raw_text[:-3].strip()

        # Robust JSON parsing - try multiple strategies
        parsed = None
        last_err = None

        # Strategy 1: direct parse
        try:
            parsed = json.loads(raw_text)
        except json.JSONDecodeError as e:
            last_err = e

        # Strategy 2: fix common Claude JSON issues
        if parsed is None:
            try:
                fixed = raw_text
                # Replace curly/smart quotes with straight ones
                fixed = fixed.replace('\u201c', '"').replace('\u201d', '"')
                fixed = fixed.replace('\u2018', "'").replace('\u2019', "'")
                # Remove trailing commas before } or ]
                fixed = re.sub(r',(\s*[}\]])', r'\1', fixed)
                # Replace unescaped newlines inside strings (best-effort)
                # Pattern: unescaped " followed by content with newline, before next "
                parsed = json.loads(fixed)
            except json.JSONDecodeError as e:
                last_err = e

        # Strategy 3: use existing parse_json_response helper
        if parsed is None:
            try:
                parsed = parse_json_response(raw_text)
            except Exception as e:
                last_err = e

        # Strategy 4: repair TRUNCATED JSON (odgovor odrezan zaradi max_tokens)
        # Najpogostejši primer: velik račun z mnogo postavkami, JSON se odreže sredi items[]
        if parsed is None:
            try:
                parsed = _repair_truncated_invoice_json(raw_text)
                if parsed:
                    print(f"[prevzemi-pdf] Truncated JSON repaired — {len(parsed.get('items', []))} postavk rešenih")
            except Exception as e:
                last_err = e

        if parsed is None:
            # Last resort: save raw response to disk for debugging
            try:
                fail_dir = PREVZEMI_DIR / f"_failed_{ts}"
                fail_dir.mkdir(parents=True, exist_ok=True)
                (fail_dir / 'raw_response.txt').write_text(raw_text, encoding='utf-8')
                (fail_dir / 'source.pdf').write_bytes(pdf_bytes)
            except Exception:
                pass
            return {"ok": False, "error": f"JSON parse failed: {last_err}. Raw response saved for debugging."}

        # Save to disk
        # Normaliziraj — odstrani presledke iz invoice_number (npr. 'FA 162383' → 'FA_162383')
        if parsed.get('invoice_number'):
            parsed['invoice_number'] = _normalize_invoice_number(parsed['invoice_number'])
        invoice_num_safe = _safe_filename(parsed.get('invoice_number', 'unknown'))
        record_id = f"{ts}_{invoice_num_safe}"
        target_dir = PREVZEMI_DIR / record_id
        target_dir.mkdir(parents=True, exist_ok=True)

        # Save PDF
        (target_dir / 'source.pdf').write_bytes(pdf_bytes)
        # Save parsed JSON
        (target_dir / 'parsed.json').write_text(json.dumps(parsed, indent=2, ensure_ascii=False), encoding='utf-8')
        # Save metadata
        # Override supplier če je AI napačno določil (npr. AMiO PDF endpoint vrne hard "amio" tudi za MotoProfil)
        detected_supplier = _override_supplier_by_invoice(
            'amio',
            parsed.get('invoice_number', ''),
            parsed.get('vendor_name', '')
        )
        if detected_supplier != 'amio':
            print(f"[prevzemi-pdf] Supplier override: amio -> {detected_supplier} (invoice={parsed.get('invoice_number')}, vendor={parsed.get('vendor_name')})")
        meta = {
            'record_id': record_id,
            'original_filename': file.filename or 'invoice.pdf',
            'created_ts': ts,
            'supplier': detected_supplier,
            'supplier_name': SUPPLIER_NAMES.get(detected_supplier, detected_supplier),
            'vendor_id': SUPPLIER_VENDOR_IDS.get(detected_supplier, ''),
            'invoice_number': parsed.get('invoice_number', ''),
            'invoice_date': parsed.get('invoice_date', ''),
            'vendor_name': parsed.get('vendor_name', ''),
            'item_count': len(parsed.get('items', []))
        }
        (target_dir / 'meta.json').write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding='utf-8')

        return {
            "ok": True,
            "record_id": record_id,
            "parsed": parsed
        }

    except json.JSONDecodeError as e:
        import traceback
        traceback.print_exc()
        # Save raw to debug
        try:
            from datetime import datetime
            ts_fail = datetime.now().strftime('%Y%m%d_%H%M%S')
            fail_dir = PREVZEMI_DIR / f"_failed_{ts_fail}"
            fail_dir.mkdir(parents=True, exist_ok=True)
            if 'raw_text' in dir():
                (fail_dir / 'raw_response.txt').write_text(raw_text, encoding='utf-8')
        except Exception:
            pass
        return {"ok": False, "error": f"JSON parse error: {e}"}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/prevzemi-list")
async def prevzemi_list():
    """List all saved prevzem records (history)."""
    try:
        records = []
        for d in sorted(PREVZEMI_DIR.iterdir(), reverse=True):
            if not d.is_dir():
                continue
            meta_file = d / 'meta.json'
            if not meta_file.exists():
                continue
            try:
                meta = json.loads(meta_file.read_text(encoding='utf-8'))
                records.append(meta)
            except Exception:
                continue
        return {"ok": True, "records": records}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/prevzemi-record/{record_id}")
async def prevzemi_record_get(record_id: str):
    """Load saved parsed record."""
    try:
        rec_safe = _safe_filename(record_id)
        target = PREVZEMI_DIR / rec_safe / 'parsed.json'
        if not target.exists():
            raise HTTPException(status_code=404, detail="Record not found")
        parsed = json.loads(target.read_text(encoding='utf-8'))
        # Vklju4i tudi meta info (supplier, vendor_id) v parsed payload
        meta_path = PREVZEMI_DIR / rec_safe / 'meta.json'
        if meta_path.exists():
            try:
                meta = json.loads(meta_path.read_text(encoding='utf-8'))
                parsed['_supplier_key'] = meta.get('supplier', '')
                parsed['_supplier_name'] = meta.get('supplier_name', '')
                parsed['_vendor_id'] = meta.get('vendor_id', '')
            except Exception:
                pass
        return {"ok": True, "parsed": parsed, "record_id": rec_safe}
    except HTTPException:
        raise
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.delete("/prevzemi-record/{record_id}")
async def prevzemi_record_delete(record_id: str):
    """Delete prevzem record."""
    import shutil
    try:
        rec_safe = _safe_filename(record_id)
        target = PREVZEMI_DIR / rec_safe
        if not target.exists():
            return {"ok": False, "error": "Not found"}
        shutil.rmtree(target)
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


class PrevzemiXlsRequest(BaseModel):
    record_id: str
    selected_indices: list[int]  # which items to include (by original index)


@app.post("/prevzemi-generate-xls")
async def prevzemi_generate_xls(req: PrevzemiXlsRequest):
    """Generate Excel 97 (.xls) file with selected items in template format."""
    try:
        import xlwt
        import io

        rec_safe = _safe_filename(req.record_id)
        target = PREVZEMI_DIR / rec_safe / 'parsed.json'
        if not target.exists():
            return {"ok": False, "error": "Record not found"}
        parsed = json.loads(target.read_text(encoding='utf-8'))

        items = parsed.get('items', [])
        selected = [items[i] for i in req.selected_indices if 0 <= i < len(items)]
        if not selected:
            return {"ok": False, "error": "No items selected"}

        invoice_num = _normalize_invoice_number(parsed.get('invoice_number', ''))
        invoice_date_str = parsed.get('invoice_date', '')
        try:
            from datetime import datetime
            inv_dt = datetime.strptime(invoice_date_str, '%Y-%m-%d')
            date_excel = inv_dt.strftime('%Y-%m-%d')
        except Exception:
            date_excel = invoice_date_str

        # Excel 97 .xls workbook
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('GenRap')

        # Text format style — preserve leading zeros, no number truncation
        text_style = xlwt.easyxf('', num_format_str='@')
        header_style = xlwt.easyxf('font: bold on', num_format_str='@')

        # Headers (matching template)
        headers = ["Racun", "DatumPrejema", "Lp.", "P/N", "Nazwa towaru/usługi / Product name",
                   "Ilość/ Quantity", "J.m./ Unit", "VAT/ TAX", "Cena/ Price EUR", "Wartość/ Value EUR"]
        for col_idx, h in enumerate(headers):
            ws.write(0, col_idx, h, header_style)

        # Items — 2 rows per item (main + EAN)
        row = 1
        new_lp = 1
        for item in selected:
            # Round price to 2 decimals
            raw_price = str(item.get('unit_price', ''))
            try:
                price_val = float(raw_price.replace(',', '.').strip())
                price_str = f"{price_val:.2f}"
            except (ValueError, AttributeError):
                price_str = raw_price

            # Main row
            ws.write(row, 0, str(invoice_num), text_style)
            ws.write(row, 1, str(date_excel), text_style)
            ws.write(row, 2, str(new_lp), text_style)
            ws.write(row, 3, str(item.get('product_number', '')), text_style)
            ws.write(row, 4, str(item.get('product_name', '')), text_style)
            ws.write(row, 5, str(item.get('qty', '')), text_style)
            ws.write(row, 6, str(item.get('unit', '')), text_style)
            ws.write(row, 7, str(item.get('vat', '')), text_style)
            ws.write(row, 8, price_str, text_style)
            ws.write(row, 9, str(item.get('value', '')), text_style)

            # EAN row
            row += 1
            ws.write(row, 0, str(invoice_num), text_style)
            ws.write(row, 1, str(date_excel), text_style)
            ean = str(item.get('ean', ''))
            ws.write(row, 4, f"EAN: {ean}", text_style)

            row += 1
            new_lp += 1

        # Column widths (xlwt uses ~256 units per character)
        widths_chars = [22, 16, 6, 14, 80, 10, 10, 8, 14, 14]
        for i, w in enumerate(widths_chars):
            ws.col(i).width = w * 256

        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Save copy to disk
        (PREVZEMI_DIR / rec_safe / 'last_generated.xls').write_bytes(buf.getvalue())
        buf.seek(0)

        filename = f"{_safe_filename(invoice_num) or 'prevzem'}.xls"
        return StreamingResponse(
            iter([buf.read()]),
            media_type='application/vnd.ms-excel',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


class PrevzemiXmlRequest(BaseModel):
    record_id: str
    selected_indices: list[int]  # which items to include (by original index)


@app.post("/prevzemi-generate-xml")
async def prevzemi_generate_xml(req: PrevzemiXmlRequest):
    """Generate nextis.xmlservice XML file with selected items."""
    try:
        import xml.etree.ElementTree as _ET
        from xml.dom import minidom as _minidom
        import io

        rec_safe = _safe_filename(req.record_id)
        target = PREVZEMI_DIR / rec_safe / 'parsed.json'
        if not target.exists():
            return {"ok": False, "error": "Record not found"}
        parsed = json.loads(target.read_text(encoding='utf-8'))

        items = parsed.get('items', [])
        selected = [items[i] for i in req.selected_indices if 0 <= i < len(items)]
        if not selected:
            return {"ok": False, "error": "No items selected"}

        invoice_num = _normalize_invoice_number(parsed.get('invoice_number', ''))
        currency = parsed.get('currency', 'EUR')

        # Parse invoice date -> DD.MM.YYYY
        raw_date = parsed.get('invoice_date', '')
        try:
            from datetime import datetime as _dt
            date_str = _dt.strptime(raw_date, '%Y-%m-%d').strftime('%d.%m.%Y')
        except Exception:
            date_str = raw_date

        # Calculate totals from selected items
        def _to_float(s):
            try:
                return float(str(s).replace(',', '.').strip())
            except Exception:
                return 0.0

        total_value = sum(_to_float(it.get('value', 0)) for it in selected)
        total_str = f"{total_value:.2f}"

        # Get supplier vendor ID from meta.json
        meta_path = PREVZEMI_DIR / rec_safe / 'meta.json'
        supplier_key = None
        if meta_path.exists():
            try:
                meta = json.loads(meta_path.read_text(encoding='utf-8'))
                supplier_key = meta.get('supplier', '').lower()
            except Exception:
                pass
        vendor_id = SUPPLIER_VENDOR_IDS.get(supplier_key, '') if supplier_key else ''

        # Build XML
        root = _ET.Element("nextis.xmlservice")
        invoice_attrs = {
            "Number": invoice_num,
            "DateCreated": date_str,
            "Currency": currency,
            "Payed": "0.00",
            "ToPay": total_str,
            "Value": total_str,
            "ValueWithVAT": total_str,
            "SupplierID": vendor_id,  # vedno prisoten, "" če dobavitelj ni prepoznan
        }
        invoice_el = _ET.SubElement(root, "Invoice", **invoice_attrs)
        items_el = _ET.SubElement(invoice_el, "InvoiceItems")

        for item in selected:
            unit_price_raw = str(item.get('unit_price', ''))
            try:
                unit_price_f = _to_float(unit_price_raw)
                unit_price_str = f"{unit_price_f:.4f}".rstrip('0').rstrip('.')
                # ensure at least 2 decimal places
                if '.' not in unit_price_str:
                    unit_price_str += '.00'
                elif len(unit_price_str.split('.')[1]) < 2:
                    unit_price_str += '0'
            except Exception:
                unit_price_str = unit_price_raw

            total_item_raw = str(item.get('value', ''))
            try:
                total_item_f = _to_float(total_item_raw)
                total_item_str = f"{total_item_f:.2f}"
            except Exception:
                total_item_str = total_item_raw

            qty_raw = str(item.get('qty', '1'))
            try:
                qty_f = _to_float(qty_raw)
                qty_str = str(int(qty_f)) if qty_f == int(qty_f) else qty_raw
            except Exception:
                qty_str = qty_raw

            _ET.SubElement(items_el, "Item",
                ProductCode=str(item.get('product_number', '')),
                Name=str(item.get('product_name', '')),
                Quantity=qty_str,
                UnitPrice=unit_price_str,
                UnitPriceWithVAT=unit_price_str,
                TotalPrice=total_item_str,
                TotalPriceWithVAT=total_item_str,
                VatValue="0.00"
            )

        # Pretty print
        xml_raw = _ET.tostring(root, encoding='unicode')
        pretty = _minidom.parseString(xml_raw).toprettyxml(indent="  ")
        # Strip XML declaration (nextis doesn't need it)
        lines = pretty.split('\n')
        if lines[0].startswith('<?xml'):
            lines = lines[1:]
        xml_output = '\n'.join(lines).strip()

        # Save copy to disk
        xml_bytes = xml_output.encode('utf-8')
        (PREVZEMI_DIR / rec_safe / 'last_generated.xml').write_bytes(xml_bytes)

        filename = f"{_safe_filename(invoice_num) or 'prevzem'}.xml"
        return StreamingResponse(
            iter([xml_bytes]),
            media_type='application/xml',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


# ─── GENERATOR PRODUKTA — iz linkov (SLO + Aliexpress) ──────────────────────

class ProductGenRequest(BaseModel):
    slo_url: str
    ali_urls: list[str] = []
    manual_text: str = ""  # fallback če scraping odpove


def _scrape_url(url: str, timeout: int = 15) -> dict:
    """Scrape spletne strani in vrne text + slike (URL-je)."""
    import re as _re
    import httpx as _httpx
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "sl,en-US;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    }
    try:
        with _httpx.Client(timeout=timeout, follow_redirects=True, headers=headers) as client:
            resp = client.get(url)
            if resp.status_code != 200:
                return {"ok": False, "error": f"HTTP {resp.status_code}", "url": url}
            html = resp.text

        title_m = _re.search(r'<title[^>]*>(.*?)</title>', html, _re.IGNORECASE | _re.DOTALL)
        title = title_m.group(1).strip() if title_m else ""

        meta_m = _re.search(r'<meta[^>]+name=["\']description["\'][^>]+content=["\']([^"\']+)', html, _re.IGNORECASE)
        meta_desc = meta_m.group(1) if meta_m else ""

        og_m = _re.search(r'<meta[^>]+property=["\']og:description["\'][^>]+content=["\']([^"\']+)', html, _re.IGNORECASE)
        og_desc = og_m.group(1) if og_m else ""

        images = []
        og_img = _re.findall(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)', html, _re.IGNORECASE)
        images.extend(og_img[:3])
        img_tags = _re.findall(r'<img[^>]+src=["\']([^"\']+)["\']', html, _re.IGNORECASE)
        for src in img_tags:
            if any(x in src.lower() for x in ['logo', 'icon', 'favicon', 'sprite', 'placeholder']):
                continue
            if src.startswith('//'):
                src = 'https:' + src
            elif src.startswith('/'):
                from urllib.parse import urlparse as _urlp
                p = _urlp(url)
                src = f"{p.scheme}://{p.netloc}{src}"
            if src not in images and src.startswith('http'):
                images.append(src)
            if len(images) >= 8:
                break

        body_m = _re.search(r'<body[^>]*>(.*?)</body>', html, _re.IGNORECASE | _re.DOTALL)
        body = body_m.group(1) if body_m else html
        body = _re.sub(r'<script[^>]*>.*?</script>', '', body, flags=_re.IGNORECASE | _re.DOTALL)
        body = _re.sub(r'<style[^>]*>.*?</style>', '', body, flags=_re.IGNORECASE | _re.DOTALL)
        text = _re.sub(r'<[^>]+>', ' ', body)
        text = _re.sub(r'&nbsp;', ' ', text)
        text = _re.sub(r'&amp;', '&', text)
        text = _re.sub(r'&[a-z]+;', '', text)
        text = _re.sub(r'\s+', ' ', text).strip()
        text = text[:8000]

        return {
            "ok": True,
            "url": url,
            "title": title,
            "meta_description": meta_desc,
            "og_description": og_desc,
            "text": text,
            "images": images[:8],
        }
    except Exception as e:
        return {"ok": False, "error": str(e), "url": url}


@app.post("/product-generate")
async def product_generate(req: ProductGenRequest):
    """Generira Maaarket-style besedilo izdelka iz SLO + Aliexpress linkov."""
    try:
        scraped = []
        slo_data = _scrape_url(req.slo_url)
        scraped.append(("SLO", slo_data))
        for ali_url in req.ali_urls[:2]:
            ali_data = _scrape_url(ali_url)
            scraped.append(("Aliexpress", ali_data))

        context_parts = []
        all_images = []
        scraping_warnings = []

        for source, data in scraped:
            if data.get("ok"):
                context_parts.append(f"\n=== {source} VIR: {data['url']} ===")
                context_parts.append(f"Naslov: {data.get('title', '')}")
                if data.get("meta_description"):
                    context_parts.append(f"Meta opis: {data['meta_description']}")
                if data.get("og_description"):
                    context_parts.append(f"OG opis: {data['og_description']}")
                context_parts.append(f"Vsebina: {data.get('text', '')[:4000]}")
                all_images.extend(data.get("images", []))
            else:
                scraping_warnings.append(f"{source} ({data['url']}): {data.get('error', 'failed')}")

        if req.manual_text:
            context_parts.append(f"\n=== ROČNO VNESENO ===\n{req.manual_text[:4000]}")

        if not context_parts:
            return {"ok": False, "error": "Nobeden link ni bil uspešno scrape-an in ni ročnega teksta."}

        full_context = "\n".join(context_parts)

        prompt = f"""Generiraj besedilo izdelka v slovenščini, v Maaarket.si stilu.

VSEBINA IZ LINKOV:
{full_context}

ZAHTEVE:
- Stil: prodajni, prijateljski, zgovoren ampak ne agresiven
- Jezik: slovenščina (slovenske besede, ne anglicizmi)
- Emoji: uporabi zmerno, ne preveč
- Ime izdelka: izmisli si privlačno ime (kombinacija 2 angleških besed, npr. WarmStep, Vapurex, NailGloss, CloudComfort)
- Tehnične podatke vzemi PREDVSEM iz SLO vira (ker so v slovenščini in zanesljivi)
- Aliexpress vir uporabi za dodatne specifikacije in vizualne reference

DOLŽINA BESEDILA (POMEMBNO — opisi naj bodo bogati in informativni):
- "glavni_opis": vsak odstavek naj ima 4-6 povedi (NE 1-2). Opisuj problem, rešitev, koristi, scenarije uporabe, počutje uporabnika. Vsebina naj se bere kot mini članek, ne kot bullet list. Vsaj 2-3 odstavki.
- "izpostavitve.opis": vsak naj ima 3-4 povedi (NE samo 1-2). Razloži funkcijo, zakaj je pomembna, kako koristi uporabniku, in konkretno povej kako se to manifestira v praksi.
- "kratek_opis": ostane 1 stavek (max 25 besed) za hook na vrhu

FORMATIRANJE (POMEMBNO):
- Za "glavni_opis" uporabi HTML tage, NE markdown:
  - <p>...</p> za odstavke
  - <strong>...</strong> za bold (NE **zvezdice**)
  - <ul><li>...</li></ul> za bullet sezname
  - Bold uporabi za KLJUČNE besede/fraze znotraj stavkov (npr. <p>Naprava deluje <strong>brezžično</strong> in zdrži <strong>do 8 ur</strong>.</p>)
- "naslov", "kratek_opis", "izpostavitve" naslov in opis pišejo se kot PLAIN TEXT brez HTML tagov
- "tehnicne_lastnosti" in "v_paketu" so PLAIN TEXT array (vsak element vrstica)

VRNI EXACT JSON v tej obliki, brez dodatnega teksta:
{{
  "naslov": "Ime izdelka IZMIŠLJEN_IME 🔥 Glavna lastnost",
  "kratek_opis": "En stavek (max 25 besed) ki privablja in pove kaj izdelek dela.",
  "glavni_opis": "<p><strong>Naslov 1. odstavka</strong> 🏠</p><p>Vsebina prvega odstavka s 4-6 povedmi: opiši problem ki ga uporabnik ima, rešitev ki jo izdelek prinaša, glavne <strong>poudarke</strong>, konkretne scenarije uporabe in počutje uporabnika ko ga uporablja.</p><p><strong>Naslov 2. odstavka</strong> ⚡</p><p>Drugi odstavek z 4-6 povedmi: dodatne lastnosti, materiali ali tehnologija, primerjava z alternativami, in zakaj je ravno ta izdelek smiselna izbira.</p>",
  "izpostavitve": [
    {{"naslov": "🎯 Naslov 1", "opis": "3-4 povedi: razloži funkcijo izdelka, zakaj je ta lastnost pomembna, kako konkretno koristi uporabniku v vsakdanjem življenju, in kaj naredi razliko od konkurence."}},
    {{"naslov": "⚡ Naslov 2", "opis": "3-4 povedi: tehnični aspekt z razlago, prednost ki jo uporabnik občuti, primer realne uporabe, in kako prihrani čas ali denar."}},
    {{"naslov": "💪 Naslov 3", "opis": "3-4 povedi: trajnost ali kakovost gradnje, kaj to pomeni za uporabnika dolgoročno, jamstvo ali zanesljivost, in zakaj se splača investirati."}}
  ],
  "tehnicne_lastnosti": [
    "Material: ...",
    "Mere: ...",
    "Napajanje: ..."
  ],
  "v_paketu": [
    "1× Naprava ...",
    "1× ..."
  ]
}}"""

        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return {"ok": False, "error": "ANTHROPIC_API_KEY ni nastavljen."}

        async with httpx.AsyncClient(timeout=120.0) as hc:
            resp = await hc.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json"
                },
                json={
                    "model": "claude-sonnet-4-6",
                    "max_tokens": 6000,
                    "messages": [{"role": "user", "content": prompt}],
                }
            )

        if resp.status_code != 200:
            return {"ok": False, "error": f"Claude API: {resp.status_code} {resp.text[:300]}"}

        result = resp.json()
        ai_text = result["content"][0]["text"].strip()

        import re as _re
        ai_text = _re.sub(r'^```(?:json)?\s*', '', ai_text)
        ai_text = _re.sub(r'\s*```$', '', ai_text)

        try:
            parsed = json.loads(ai_text)
        except Exception as e:
            return {"ok": False, "error": f"JSON parse error: {e}", "raw": ai_text}

        return {
            "ok": True,
            "data": parsed,
            "images": all_images[:8],
            "warnings": scraping_warnings,
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


# ─── EMOJI OVERLAY ZA PODNAPISE ─────────────────────────────────────────────

TWEMOJI_DIR = DATA_DIR / "twemoji_cache"
TWEMOJI_DIR.mkdir(exist_ok=True, parents=True)
TWEMOJI_CDN = "https://cdn.jsdelivr.net/gh/twitter/twemoji@latest/assets/72x72"


def _emoji_to_codepoint(emoji: str) -> str:
    """Pretvori emoji znak v hex codepoint za Twemoji URL.
    Primer: '🔒' -> '1f512', '🇸🇮' (flag) -> '1f1f8-1f1ee'."""
    codes = []
    for ch in emoji:
        cp = ord(ch)
        # Filter ZWJ (zero-width joiner) in variation selectors (FE0F)
        if cp == 0xFE0F:
            continue
        codes.append(f"{cp:x}")
    return "-".join(codes)


def _get_twemoji_png(emoji: str) -> Path | None:
    """Vrne pot do PNG datoteke za emoji. Če manjka, prenesi iz CDN.
    Vrne None če download failed."""
    codepoint = _emoji_to_codepoint(emoji)
    if not codepoint:
        return None
    png_path = TWEMOJI_DIR / f"{codepoint}.png"
    if png_path.exists() and png_path.stat().st_size > 100:
        return png_path

    # Download iz Twemoji CDN
    url = f"{TWEMOJI_CDN}/{codepoint}.png"
    try:
        import httpx as _httpx
        with _httpx.Client(timeout=10.0, follow_redirects=True) as client:
            resp = client.get(url)
            if resp.status_code == 200 and len(resp.content) > 100:
                png_path.write_bytes(resp.content)
                return png_path
    except Exception as e:
        print(f"[twemoji] Download failed for {emoji} ({codepoint}): {e}")
    return None


class SuggestEmojisRequest(BaseModel):
    text: str
    alignment: dict  # ElevenLabs alignment object
    lang: str = "sl"


@app.post("/suggest-emojis")
async def suggest_emojis(req: SuggestEmojisRequest):
    """Claude analizira tekst in predlaga emoji za ~30% segmentov (mix style)."""
    try:
        # Razdeli tekst v stavke/segmente
        words = _parse_words(req.alignment)
        if not words:
            return {"ok": False, "error": "Ni alignment podatkov"}

        # Grupiraj v segmente (vrstice) — uporabimo isto logiko kot build_ass
        # Vsakih 4-6 besed = 1 segment
        segments = []
        i = 0
        while i < len(words):
            grp = [words[i]]; i += 1
            while i < len(words) and len(grp) < 5 and (words[i][1] - grp[0][1]) < 2.5:
                grp.append(words[i]); i += 1
            seg_text = " ".join(w[0] for w in grp)
            seg_start = grp[0][1]
            seg_end = grp[-1][2]
            segments.append({"text": seg_text, "start": seg_start, "end": seg_end})

        # Claude prompt — predlagaj emoji samo za KLJUČNE segmente (mix style)
        segments_json = json.dumps([{"idx": i, "text": s["text"]} for i, s in enumerate(segments)], ensure_ascii=False)

        prompt = f"""Imaš seznam tekstualnih segmentov iz video oglasa (jezik: {req.lang}).

Tvoja naloga: za nekatere segmente predlagaj 1 emoji ki vizualno poudari ključni koncept.

PRAVILA:
- Emoji predlagaj samo za KLJUČNE segmente (cca. 30-50% segmentov, NE vse)
- Preskoči segmente ki so:
  - Splošni opisi brez vizualnega konteksta ("ki vam pomaga", "in zato")
  - Drugi del stavka (povezovalni)
  - Manjši kot 2 besedi
- Predlagaj emoji za segmente kjer:
  - Glavna beseda ima jasen vizualni koncept (okna → 🪟, hitrost → ⚡, varnost → 🔒)
  - Klavzule glavnih koristi (lažje → 💪, doma → 🏠, naravno → 🌿)
  - Močne čustvene besede (super → ✨, novo → 🆕, brez skrbi → 😌)
- Uporabljaj POPULARNE emoji (✨ 🔥 💪 ⚡ 🎯 🪟 🔒 🏠 🌿 💧 🌡 🛡 ⏱ 📦 🎁 ❤️ 👶 🐾 🌟 💎 🍃 ☀️ 🌙 🛌 👍 🚀)
- Različni emoji za različne segmente (NE isti emoji vsakič)

SEGMENTI:
{segments_json}

VRNI EXACT JSON brez markdown:
{{"emojis": [{{"idx": 0, "emoji": "🪟"}}, {{"idx": 3, "emoji": "🔒"}}, ...]}}

Vrni samo segmente ki potrebujejo emoji, ostale preskoči."""

        text = await call_claude(prompt, "claude-sonnet-4-6", None, 2000)
        parsed = parse_json_response(text)
        if not parsed:
            return {"ok": False, "error": "JSON parse failed", "raw": text[:300]}

        emoji_map = {}  # segment_idx -> emoji
        for item in parsed.get("emojis", []):
            try:
                idx = int(item["idx"])
                emoji = item["emoji"]
                if 0 <= idx < len(segments):
                    emoji_map[idx] = emoji
            except Exception:
                continue

        # Pripravi rezultat
        result_segments = []
        for i, seg in enumerate(segments):
            result_segments.append({
                "idx": i,
                "text": seg["text"],
                "start": seg["start"],
                "end": seg["end"],
                "emoji": emoji_map.get(i, None),
            })

        return {"ok": True, "segments": result_segments}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


# ─── CASHFLOW ANALIZA (iz Excel uploada) ────────────────────────────────────

CASHFLOW_FILE = DATA_DIR / "cashflow.xlsx"


@app.post("/cashflow-upload")
async def cashflow_upload(file: UploadFile = File(...)):
    """Naloži in shrani CashFlow.xlsx. Vrne število uspešno parsanih vrstic."""
    try:
        content = await file.read()
        if len(content) > 50 * 1024 * 1024:  # 50MB max
            return {"ok": False, "error": "File too large (max 50MB)"}
        CASHFLOW_FILE.write_bytes(content)

        # Test parse
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        if "Cash Flow" not in wb.sheetnames:
            return {"ok": False, "error": f"Tab 'Cash Flow' ne obstaja. Tabi: {wb.sheetnames}"}
        ws = wb["Cash Flow"]
        n_rows = 0
        for r in range(2, ws.max_row + 1):
            d = ws.cell(row=r, column=1).value
            if hasattr(d, 'year'):
                n_rows += 1
        return {"ok": True, "rows": n_rows, "size_kb": round(len(content) / 1024, 1)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/cashflow-data")
async def cashflow_data():
    """Vrne strukturirane podatke iz CashFlow.xlsx (samo tab 'Cash Flow')."""
    try:
        if not CASHFLOW_FILE.exists():
            return {"ok": False, "error": "no_file", "message": "CashFlow.xlsx še ni naložen."}

        import openpyxl
        from datetime import datetime as _dt
        wb = openpyxl.load_workbook(CASHFLOW_FILE, data_only=True)
        if "Cash Flow" not in wb.sheetnames:
            return {"ok": False, "error": "Tab 'Cash Flow' ne obstaja."}
        ws = wb["Cash Flow"]

        rows = []
        for r in range(2, ws.max_row + 1):
            d = ws.cell(row=r, column=1).value
            if not hasattr(d, 'year'):
                continue
            # Skip rows where everything is None/empty
            cf_all = ws.cell(row=r, column=15).value
            suma_promet = ws.cell(row=r, column=21).value
            if cf_all is None and suma_promet is None:
                continue

            def num(c):
                v = ws.cell(row=r, column=c).value
                if v is None or v == '':
                    return None
                try:
                    return float(v)
                except Exception:
                    return None

            rows.append({
                "date": d.strftime('%Y-%m-%d'),
                "paypal": num(3),
                "wise": num(4),
                "intesa_hr": num(5),
                "unicredit": num(6),
                "ibkr": num(7),
                "intesa_slo": num(8),
                "racun_dh": num(9),
                "suban": num(10),
                "srbija_rsd": num(11),
                "intesa_rs": num(12),
                "silux_b": num(13),
                "cf_all": num(15),
                "cf_w_loan": num(25),
                "zaloga": num(27),
                "st_kosov": num(28),
                "posojila": num(29),
            })

        return {"ok": True, "rows": rows, "count": len(rows)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


# ─── GOOGLE SHEETS SYNC za CashFlow ─────────────────────────────────────────

CASHFLOW_SHEET_CONFIG = DATA_DIR / "cashflow_sheet_config.json"


def _load_cf_sheet_config():
    if CASHFLOW_SHEET_CONFIG.exists():
        try:
            return json.loads(CASHFLOW_SHEET_CONFIG.read_text(encoding='utf-8'))
        except Exception:
            return {}
    return {}


def _save_cf_sheet_config(cfg: dict):
    CASHFLOW_SHEET_CONFIG.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding='utf-8')


def _get_sheets_service():
    """Inicializira Google Sheets API client s service account."""
    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    if not sa_json:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON env var ni nastavljen")

    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    sa_info = json.loads(sa_json)
    creds = service_account.Credentials.from_service_account_info(
        sa_info,
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
    )
    service = build('sheets', 'v4', credentials=creds, cache_discovery=False)
    return service, sa_info.get("client_email", "unknown")


def _sync_cashflow_from_sheets(sheet_id: str = None, tab_name: str = "Cash Flow"):
    """Prenese podatke iz Google Sheets in shrani kot Excel v /data/cashflow.xlsx."""
    if not sheet_id:
        cfg = _load_cf_sheet_config()
        sheet_id = os.environ.get("CASHFLOW_SHEET_ID") or cfg.get("sheet_id")
    if not sheet_id:
        return {"ok": False, "error": "Sheet ID ni nastavljen"}

    try:
        service, sa_email = _get_sheets_service()

        # Najprej dobimo seznam vseh tabov da najdemo pravega
        try:
            meta = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
            all_tabs = [s["properties"]["title"] for s in meta.get("sheets", [])]
        except Exception as e:
            return {"ok": False, "error": f"Cannot read sheet metadata: {e}. Check sharing permissions."}

        # Avto-resolve tab name (case insensitive + brez presledkov)
        target_tab = None
        tab_lower = tab_name.lower().replace(" ", "")
        for t in all_tabs:
            t_norm = t.lower().replace(" ", "")
            if t_norm == tab_lower or t_norm == "cashflow":
                target_tab = t
                break

        if not target_tab:
            return {"ok": False, "error": f"Tab '{tab_name}' ne obstaja. Najdeni tabi: {', '.join(all_tabs)}"}

        # Vzami vse podatke iz tab-a
        result = service.spreadsheets().values().get(
            spreadsheetId=sheet_id,
            range=f"'{target_tab}'",
            valueRenderOption="UNFORMATTED_VALUE",
            dateTimeRenderOption="FORMATTED_STRING"
        ).execute()

        values = result.get("values", [])
        if not values:
            return {"ok": False, "error": f"Tab '{target_tab}' je prazen"}

        # Konvertiraj v Excel format in shrani na disk
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cash Flow"

        from datetime import datetime as _dt
        for row_idx, row in enumerate(values, start=1):
            for col_idx, cell in enumerate(row, start=1):
                # Try to parse date strings in column 1
                if col_idx == 1 and row_idx > 1 and isinstance(cell, str):
                    parsed_date = None
                    # Najprej počistimo presledke za slovenski format "1. 4. 2019"
                    cell_clean = cell.replace(" ", "")
                    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                        try:
                            parsed_date = _dt.strptime(cell_clean, fmt)
                            break
                        except Exception:
                            continue
                    if parsed_date:
                        ws.cell(row=row_idx, column=col_idx).value = parsed_date
                        continue
                ws.cell(row=row_idx, column=col_idx).value = cell

        wb.save(CASHFLOW_FILE)

        # Save config
        cfg = _load_cf_sheet_config()
        cfg["sheet_id"] = sheet_id
        cfg["tab_name"] = tab_name
        cfg["last_sync"] = _dt.now().isoformat()
        cfg["last_sync_rows"] = len(values)
        cfg["service_account_email"] = sa_email
        _save_cf_sheet_config(cfg)

        # Count data rows (with dates)
        n_data = sum(1 for r in range(2, ws.max_row + 1) if hasattr(ws.cell(row=r, column=1).value, 'year'))

        return {
            "ok": True,
            "rows": n_data,
            "raw_rows": len(values),
            "sheet_id": sheet_id,
            "tab_name": tab_name,
            "last_sync": cfg["last_sync"],
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


class CashFlowSyncRequest(BaseModel):
    sheet_id: str = ""
    tab_name: str = "Cash Flow"


@app.post("/cashflow-sync-sheets")
async def cashflow_sync_sheets(req: CashFlowSyncRequest):
    """Manual sync iz Google Sheets."""
    return _sync_cashflow_from_sheets(
        sheet_id=req.sheet_id or None,
        tab_name=req.tab_name or "Cash Flow"
    )


@app.get("/cashflow-sheets-status")
async def cashflow_sheets_status():
    """Vrne trenutno konfiguracijo Sheets sync-a."""
    cfg = _load_cf_sheet_config()
    env_id = os.environ.get("CASHFLOW_SHEET_ID", "")
    return {
        "ok": True,
        "configured_id": cfg.get("sheet_id") or env_id,
        "from_env": bool(env_id),
        "tab_name": cfg.get("tab_name", "Cash Flow"),
        "last_sync": cfg.get("last_sync"),
        "last_sync_rows": cfg.get("last_sync_rows"),
        "service_account_email": cfg.get("service_account_email"),
        "auto_refresh": "daily at 06:00",
    }


# Auto-refresh ob 6:00 zjutraj
async def _hsplus_daily_scheduler():
    """Poteg HS+ kataloga 1× na dan ob 5:00 (Render UTC; prilagodi po potrebi).
    Tiho v ozadju — zjutraj imaš svež katalog. Ob napaki počaka 1h in poskusi znova."""
    while True:
        try:
            from datetime import datetime as _dt, timedelta as _td
            now = _dt.now()
            next_run = now.replace(hour=5, minute=0, second=0, microsecond=0)
            if next_run <= now:
                next_run += _td(days=1)
            sleep_seconds = (next_run - now).total_seconds()
            print(f"[hsplus-daily] Next fetch at {next_run.isoformat()} (in {sleep_seconds:.0f}s)")
            await asyncio.sleep(sleep_seconds)
            result = await _hsplus_fetch_core(force=True)
            if result.get("ok"):
                print(f"[hsplus-daily] OK — {result.get('count')} izdelkov ({result.get('fetched_at')})")
            else:
                print(f"[hsplus-daily] FAIL — {result.get('error')}")
        except Exception as e:
            print(f"[hsplus-daily] Error: {e}")
            await asyncio.sleep(3600)


async def _daily_cashflow_sync():
    while True:
        try:
            from datetime import datetime as _dt, timedelta as _td
            now = _dt.now()
            # Naslednji 6:00
            next_run = now.replace(hour=6, minute=0, second=0, microsecond=0)
            if next_run <= now:
                next_run += _td(days=1)
            sleep_seconds = (next_run - now).total_seconds()
            print(f"[cashflow-sync] Next sync at {next_run.isoformat()} (in {sleep_seconds:.0f}s)")
            await asyncio.sleep(sleep_seconds)

            # Sync
            cfg = _load_cf_sheet_config()
            sheet_id = os.environ.get("CASHFLOW_SHEET_ID") or cfg.get("sheet_id")
            if sheet_id:
                tab = cfg.get("tab_name", "Cash Flow")
                result = _sync_cashflow_from_sheets(sheet_id, tab)
                if result.get("ok"):
                    print(f"[cashflow-sync] OK — {result['rows']} rows synced")
                else:
                    print(f"[cashflow-sync] FAIL — {result.get('error')}")
            else:
                print("[cashflow-sync] No sheet_id configured, skipping")
        except Exception as e:
            print(f"[cashflow-sync] Error: {e}")
            await asyncio.sleep(3600)  # počakaj 1h pred ponovnim poskusom


# ─── MULTI-SUPPLIER PREVZEMI PARSER ──────────────────────────────────────────

def _detect_supplier(filename: str, content: bytes) -> str:
    """Auto-detect dobavitelja iz imena fajla in vsebine."""
    fn = (filename or "").lower()

    # Filename heuristics najprej
    if fn.startswith("fa_") and fn.endswith(".csv"):
        return "motoprofil"
    if fn.startswith("shipment-") and fn.endswith(".xml"):
        return "intercars"
    if fn.startswith("spm_wdt") and (fn.endswith(".xlsx") or fn.endswith(".xls")):
        return "abakus"
    if fn.startswith("language_eng") and fn.endswith(".pdf"):
        return "ikonka"
    # AMiO ima drugačno ime
    if "fv_kk" in fn or fn.startswith("fv-kk") or fn.startswith("fv_kk"):
        return "amio"

    # Vsebinske heuristike
    try:
        head = content[:5000].decode("utf-8", errors="ignore").lower()
    except Exception:
        head = ""

    # AMiO XML format (<invoice><document_number>FV/KK/...) — NOV, brez Claude
    if "<invoice>" in head and ("fv/kk" in head or "amio" in head or "<product_number>" in head):
        return "amio_xml"

    if "prefiks" in head and "indeks" in head:
        return "motoprofil"
    if "despatchadvice" in head or "eslog" in head or "sisshl" in head:
        return "intercars"
    if "abakus sp" in head or "spm/wdt" in head:
        return "abakus"
    if "kik sp" in head or "kik sp\xf3\u0142ka" in head or "faktura vat ue" in head:
        return "ikonka"
    if "amio" in head or "kn\xf3wska" in head or "knurowska" in head:
        return "amio"

    return "unknown"


def _parse_amio_xml(content: bytes) -> dict:
    """Parse AMiO strukturiran XML (<invoice>...). Brez Claude — direktno iz XML.
    Format: <invoice><document_number>, <items><item><product_number>...
    """
    import xml.etree.ElementTree as ET
    from datetime import datetime

    root = ET.fromstring(content)

    def _txt(parent, tag, default=""):
        el = parent.find(tag)
        return el.text.strip() if el is not None and el.text else default

    # Header
    invoice_number = _txt(root, "document_number")
    issue_date = _txt(root, "issue_date")  # 20-05-2026 (DD-MM-YYYY)
    # Pretvori v YYYY-MM-DD
    invoice_date = issue_date
    if issue_date and "-" in issue_date:
        parts = issue_date.split("-")
        if len(parts) == 3 and len(parts[0]) == 2:  # DD-MM-YYYY
            invoice_date = f"{parts[2]}-{parts[1]}-{parts[0]}"

    vendor_el = root.find("vendor")
    vendor_name = _txt(vendor_el, "name", "AMiO") if vendor_el is not None else "AMiO"
    vendor_tin = _txt(vendor_el, "tin") if vendor_el is not None else ""

    customer_el = root.find("customer")
    customer_name = _txt(customer_el, "name") if customer_el is not None else ""

    totals_el = root.find("totals")
    total_value = ""
    currency = "EUR"
    if totals_el is not None:
        tv = _txt(totals_el, "total_value_eur")
        if tv:
            total_value = f"{tv} EUR"

    # Items
    items = []
    for it in root.findall(".//item"):
        qty_str = _txt(it, "quantity", "1").replace(",", ".")
        price_str = _txt(it, "unit_price_eur", "0").replace(",", ".")
        value_str = _txt(it, "value_eur", "0").replace(",", ".")
        try:
            qty = float(qty_str)
        except ValueError:
            qty = 1
        try:
            unit_price = float(price_str)
        except ValueError:
            unit_price = 0.0
        try:
            value = float(value_str)
        except ValueError:
            value = round(qty * unit_price, 2)

        items.append({
            "lp": _txt(it, "no"),
            "product_number": _txt(it, "product_number"),
            "product_name": _txt(it, "product_name"),
            "ean": _txt(it, "ean"),
            "qty": int(qty) if qty == int(qty) else qty,
            "unit": _txt(it, "uom"),
            "vat": _txt(it, "vat", "0%"),
            "unit_price": unit_price,
            "value": value,
        })

    return {
        "invoice_number": invoice_number,
        "invoice_date": invoice_date or datetime.now().strftime("%Y-%m-%d"),
        "vendor_name": vendor_name,
        "vendor_tin": vendor_tin,
        "customer_name": customer_name,
        "total_value": total_value,
        "currency": currency,
        "items": items,
    }


def _parse_motoprofil_csv(content: bytes, filename: str) -> dict:
    """Parse MotoProfil CSV. Invoice num + date iz imena fajla."""
    import csv, io, re
    from datetime import datetime

    # Invoice number iz imena: FA_155887_05_2026 → FA/155887/05/2026
    base = re.sub(r"\.[^.]+$", "", filename or "")
    parts = re.match(r"FA_(\d+)_(\d+)_(\d+)", base, re.IGNORECASE)
    if parts:
        invoice_number = f"FA/{parts.group(1)}/{parts.group(2)}/{parts.group(3)}"
        # date approximation: 1. v mesecu (ne moremo brez konkretnih podatkov)
        try:
            invoice_date = f"{parts.group(3)}-{parts.group(2).zfill(2)}-01"
        except Exception:
            invoice_date = datetime.now().strftime("%Y-%m-%d")
    else:
        invoice_number = base or "MotoProfil"
        invoice_date = datetime.now().strftime("%Y-%m-%d")

    # Parse CSV
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    items = []
    for row in reader:
        prefiks = (row.get("Prefiks") or "").strip()
        indeks = (row.get("Indeks") or "").strip()
        nazwa = (row.get("Nazwa") or "").strip()
        ilosc = (row.get("Ilosc") or "1").strip()
        cena = (row.get("Cena") or "0").strip().replace(",", ".")

        # MotoProfil: uporabi SAMO indeks (SKU) brez prefiksa (ABS, AMT, WP, ...)
        # Prefiks je dobaviteljeva interna oznaka, ne ga vključujemo v ProductCode.
        product_code = indeks if indeks else prefiks
        try:
            qty = float(ilosc)
            unit_price = float(cena)
        except ValueError:
            continue

        items.append({
            "product_number": product_code,
            "product_name": nazwa,
            "qty": int(qty) if qty == int(qty) else qty,
            "unit_price": unit_price,
            "value": round(qty * unit_price, 2),
        })

    return {
        "invoice_number": invoice_number,
        "invoice_date": invoice_date,
        "currency": "EUR",
        "vendor_name": "MotoProfil",
        "items": items,
    }


def _parse_intercars_xml(content: bytes) -> dict:
    """Parse Intercars eSLOG XML (DespatchAdvice).
    Imena izdelkov niso v dokumentu — pustimo prazno, user vpiše ročno."""
    import xml.etree.ElementTree as ET
    from datetime import datetime

    ns = {"e": "urn:eslog:2.00"}
    root = ET.fromstring(content)

    # Document number
    doc_num = ""
    unh = root.find(".//e:S_UNH/e:D_0062", ns)
    if unh is not None and unh.text:
        doc_num = unh.text.strip()
    else:
        bgm_n = root.find(".//e:S_BGM/e:C_C106/e:D_1004", ns)
        if bgm_n is not None and bgm_n.text:
            doc_num = bgm_n.text.strip()

    # Datum (first S_DTM s 2005=137 ali 11)
    invoice_date = ""
    for dtm in root.findall(".//e:S_DTM", ns):
        code = dtm.find("e:C_C507/e:D_2005", ns)
        date = dtm.find("e:C_C507/e:D_2380", ns)
        if code is not None and code.text == "137" and date is not None:
            invoice_date = date.text.strip()
            break
    if not invoice_date:
        invoice_date = datetime.now().strftime("%Y-%m-%d")
    # Format YYYY-MM-DD
    if len(invoice_date) == 8 and invoice_date.isdigit():
        invoice_date = f"{invoice_date[0:4]}-{invoice_date[4:6]}-{invoice_date[6:8]}"

    # Items: iz G_SG10 → G_SG17
    items = []
    for sg17 in root.findall(".//e:G_SG10/e:G_SG17", ns):
        code_el = sg17.find("e:S_LIN/e:C_C212/e:D_7140", ns)
        if code_el is None:
            continue
        product_code = (code_el.text or "").strip()

        # Qty: S_QTY/C_C186/D_6060
        qty = 1
        qty_el = sg17.find("e:S_QTY/e:C_C186/e:D_6060", ns)
        if qty_el is not None and qty_el.text:
            try:
                qty = float(qty_el.text)
            except ValueError:
                pass

        # Price: S_MOA/C_C516/D_5004 (kjer D_5025=203 = unit price)
        unit_price = 0.0
        for moa in sg17.findall("e:S_MOA", ns):
            code = moa.find("e:C_C516/e:D_5025", ns)
            val = moa.find("e:C_C516/e:D_5004", ns)
            if code is not None and code.text == "203" and val is not None and val.text:
                try:
                    unit_price = float(val.text)
                    break
                except ValueError:
                    continue

        items.append({
            "product_number": product_code,
            "product_name": "",  # Manjka v XML — user vpiše v UI
            "qty": int(qty) if qty == int(qty) else qty,
            "unit_price": unit_price,
            "value": round(qty * unit_price, 2),
        })

    return {
        "invoice_number": doc_num,
        "invoice_date": invoice_date,
        "currency": "EUR",
        "vendor_name": "Intercars",
        "items": items,
        "_missing_names": True,  # UI flag — kaže prompt za vpis imen
    }


def _parse_abakus_xlsx(content: bytes) -> dict:
    """Parse Abakus XLSX. Headers v vrstici 1, podatki od vrstice 3.
    Uporablja python-calamine (zaradi openpyxl 'xxid' bug-a)."""
    import io
    from datetime import datetime
    import re

    # Poskusi calamine
    try:
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_filelike(io.BytesIO(content))
        sheet_name = wb.sheet_names[0]
        ws = wb.get_sheet_by_name(sheet_name)
        data = ws.to_python()
    except ImportError:
        # Fallback: openpyxl ignoring styles (read_only mode)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
        except Exception as e:
            raise RuntimeError(f"Nemoremo prebrati Abakus XLSX. Manjka python-calamine v requirements? ({e})")
    except Exception as e:
        raise RuntimeError(f"Calamine read failed: {e}")

    # Row 1: meta (SPM/WDT/74/202605 je v stolpcu)
    invoice_number = ""
    if len(data) > 1:
        for cell in data[1]:
            cell_str = str(cell)
            if "SPM/WDT" in cell_str or "SPM\\WDT" in cell_str:
                invoice_number = cell_str.strip()
                break

    invoice_date = datetime.now().strftime("%Y-%m-%d")
    if "202" in sheet_name:
        m = re.search(r"_(\d{6})$", sheet_name)
        if m:
            ym = m.group(1)
            try:
                invoice_date = f"{ym[0:4]}-{ym[4:6]}-01"
            except Exception:
                pass

    items = []
    for row in data[2:]:
        if not row or len(row) < 5:
            continue
        try:
            lp = row[0]
            if lp is None or str(lp).strip() == "":
                continue
            float(lp)
        except (ValueError, TypeError):
            continue

        product_code = str(row[1] or "").strip()
        description = str(row[2] or "").strip()
        qty_raw = str(row[3] or "1").replace(",", ".").strip()
        net_price_raw = str(row[4] or "0").replace(",", ".").strip()

        try:
            qty = float(qty_raw)
            unit_price = float(net_price_raw)
        except ValueError:
            continue

        items.append({
            "product_number": product_code,
            "product_name": description,
            "qty": int(qty) if qty == int(qty) else qty,
            "unit_price": unit_price,
            "value": round(qty * unit_price, 2),
        })

    return {
        "invoice_number": invoice_number or sheet_name,
        "invoice_date": invoice_date,
        "currency": "EUR",
        "vendor_name": "Abakus",
        "items": items,
    }


def _parse_ikonka_pdf(content: bytes) -> dict:
    """Parse Ikonka/KIK PDF (FAKTURA VAT UE).
    Uporabimo Claude API ker je PDF struktura kompleksna."""
    import base64
    pdf_b64 = base64.standard_b64encode(content).decode('utf-8')

    msg = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=20000,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "document",
                    "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}
                },
                {
                    "type": "text",
                    "text": """Parse this Ikonka/KIK supplier invoice PDF. Extract structured data.

Return ONLY a valid JSON object — no markdown, no explanation.

CRITICAL: NET PRICE ONLY (cena netto), ignore brutto. We need supplier's net price without VAT.

JSON format:
{
  "invoice_number": "10941/HT/05/2026",
  "invoice_date": "2026-05-13",
  "items": [
    {
      "product_number": "KX3085",
      "product_name": "Kijki kije trekkingowe składane 2 sztuki czarne",
      "qty": 1,
      "unit_price": 3.98
    }
  ]
}

Rules:
- invoice_number: vzami iz "FAKTURA VAT UE XXXX" linije
- invoice_date: "Data wystawienia"
- product_number: kolona "Symbol" (npr. KX3085), NE EAN
- product_name: kolona "Nazwa"
- qty: kolona "Ilość szt."
- unit_price: kolona "Cena netto" (NE brutto)
- IGNORE "Przelew B2B_PAYMENT" in "kurier" rows — to so plačilo/dostava, ne produkti
- Return only product items
"""
                }
            ]
        }]
    )

    raw = msg.content[0].text.strip()
    # Strip markdown
    import re as _re
    raw = _re.sub(r'^```(?:json)?\s*', '', raw)
    raw = _re.sub(r'\s*```$', '', raw)
    parsed = json.loads(raw)

    # Doplnimo value
    for it in parsed.get("items", []):
        try:
            qty = float(it.get("qty", 1))
            up = float(it.get("unit_price", 0))
            it["value"] = round(qty * up, 2)
        except Exception:
            it["value"] = 0

    parsed["currency"] = "EUR"
    parsed["vendor_name"] = "Ikonka (KIK)"
    return parsed


# Supplier name lookup
SUPPLIER_NAMES = {
    "motoprofil": "MotoProfil",
    "intercars": "Intercars",
    "abakus": "Abakus",
    "ikonka": "Ikonka (KIK)",
    "amio": "AMiO",
    "unknown": "Unknown",
}

# Vendor ID-ji v računovodskem sistemu (siluxar/nextis)
# Default vrednosti — če /data/prevzemi/vendor_ids.json ne obstaja, uporabi te
SUPPLIER_VENDOR_IDS_DEFAULT = {
    "amio":       "000097",
    "abakus":     "D00003",
    "intercars":  "000177",
    "motoprofil": "P00010",
    "ikonka":     "000223",
}

VENDOR_IDS_FILE = PREVZEMI_DIR / "vendor_ids.json"


def _load_vendor_ids() -> dict:
    """Vrne vendor_id mapping iz JSON file-a (ali default če ni)."""
    try:
        if VENDOR_IDS_FILE.exists():
            data = json.loads(VENDOR_IDS_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
    except Exception as e:
        print(f"[vendor-ids] load error: {e}")
    # Fallback na default
    return dict(SUPPLIER_VENDOR_IDS_DEFAULT)


def _save_vendor_ids(mapping: dict) -> bool:
    """Shrani vendor_id mapping v JSON file."""
    try:
        VENDOR_IDS_FILE.parent.mkdir(parents=True, exist_ok=True)
        VENDOR_IDS_FILE.write_text(json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8")
        return True
    except Exception as e:
        print(f"[vendor-ids] save error: {e}")
        return False


# Globalni accessor za vse mesta ki uporabljajo SUPPLIER_VENDOR_IDS
# (vsakič preveri file — če je bil posodobljen brez restart-a, se to upošteva)
class _VendorIdsProxy:
    def __getitem__(self, key):
        return _load_vendor_ids().get(key, "")
    def get(self, key, default=""):
        return _load_vendor_ids().get(key, default)
    def items(self):
        return _load_vendor_ids().items()
    def keys(self):
        return _load_vendor_ids().keys()
    def values(self):
        return _load_vendor_ids().values()
    def __contains__(self, key):
        return key in _load_vendor_ids()
    def __iter__(self):
        return iter(_load_vendor_ids())

SUPPLIER_VENDOR_IDS = _VendorIdsProxy()


@app.get("/prevzemi-vendor-ids")
async def prevzemi_get_vendor_ids():
    """Vrne trenutni vendor_id mapping + seznam vseh znanih dobaviteljev."""
    current = _load_vendor_ids()
    # Vključi tudi default supplier-je ki morda še niso shranjeni
    all_suppliers = {}
    for sup_key, sup_name in SUPPLIER_NAMES.items():
        if sup_key == "unknown":
            continue
        all_suppliers[sup_key] = {
            "supplier_key": sup_key,
            "supplier_name": sup_name,
            "vendor_id": current.get(sup_key, ""),
            "is_default": current.get(sup_key) == SUPPLIER_VENDOR_IDS_DEFAULT.get(sup_key),
        }
    return {"ok": True, "suppliers": all_suppliers, "defaults": SUPPLIER_VENDOR_IDS_DEFAULT}


@app.post("/prevzemi-vendor-ids")
async def prevzemi_set_vendor_ids(data: dict):
    """Shrani vendor_id mapping. Body: {"mapping": {"amio": "000097", "abakus": "D00003", ...}}"""
    try:
        mapping = data.get("mapping", {})
        if not isinstance(mapping, dict):
            return {"ok": False, "error": "mapping mora biti slovar"}
        # Validacija — samo znani supplier_key-i
        valid = {}
        for k, v in mapping.items():
            if k not in SUPPLIER_NAMES:
                continue
            if k == "unknown":
                continue
            v_str = str(v or "").strip()
            valid[k] = v_str
        if not _save_vendor_ids(valid):
            return {"ok": False, "error": "Shranjevanje ni uspelo"}
        return {"ok": True, "saved": valid}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/prevzemi-normalize-invoice-numbers")
async def prevzemi_normalize_invoice_numbers():
    """Backfill: posodobi vse obstoječe prevzeme tako, da odstrani presledke iz invoice_number."""
    try:
        updated = 0
        skipped = 0
        details = []
        for d in PREVZEMI_DIR.iterdir():
            if not d.is_dir():
                continue
            # parsed.json
            parsed_path = d / "parsed.json"
            meta_path = d / "meta.json"
            if not parsed_path.exists():
                continue
            try:
                parsed = json.loads(parsed_path.read_text(encoding="utf-8"))
                old_num = parsed.get("invoice_number", "")
                new_num = _normalize_invoice_number(old_num)
                if old_num != new_num and new_num:
                    parsed["invoice_number"] = new_num
                    parsed_path.write_text(json.dumps(parsed, indent=2, ensure_ascii=False), encoding="utf-8")
                    # Posodobi tudi meta.json
                    if meta_path.exists():
                        meta = json.loads(meta_path.read_text(encoding="utf-8"))
                        if meta.get("invoice_number") != new_num:
                            meta["invoice_number"] = new_num
                            meta_path.write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")
                    updated += 1
                    details.append({"record_id": d.name, "old": old_num, "new": new_num})
                else:
                    skipped += 1
            except Exception as e:
                skipped += 1
                details.append({"record_id": d.name, "error": str(e)})
        return {"ok": True, "updated": updated, "skipped": skipped, "details": details[:30]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/prevzemi-update-supplier")
async def prevzemi_update_supplier(data: dict):
    """Posodobi supplier (in vendor_id) za posamezen prevzem.
    Body: {"record_id": "...", "supplier": "motoprofil"}
    """
    try:
        record_id = (data.get("record_id") or "").strip()
        new_supplier = (data.get("supplier") or "").strip().lower()
        if not record_id:
            return {"ok": False, "error": "record_id manjka"}
        if new_supplier not in SUPPLIER_NAMES:
            return {"ok": False, "error": f"Neznan dobavitelj: {new_supplier}"}

        rec_dir = PREVZEMI_DIR / record_id
        meta_path = rec_dir / "meta.json"
        if not meta_path.exists():
            return {"ok": False, "error": "Prevzem ne obstaja"}

        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        old_supplier = meta.get("supplier", "")
        old_vendor_id = meta.get("vendor_id", "")

        meta["supplier"] = new_supplier
        meta["supplier_name"] = SUPPLIER_NAMES.get(new_supplier, new_supplier)
        meta["vendor_id"] = _load_vendor_ids().get(new_supplier, "")

        meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[prevzemi-update] {record_id}: {old_supplier}({old_vendor_id}) -> {new_supplier}({meta['vendor_id']})")
        return {
            "ok": True,
            "supplier": new_supplier,
            "supplier_name": meta["supplier_name"],
            "vendor_id": meta["vendor_id"],
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/prevzemi-backfill-vendor-ids")
async def prevzemi_backfill_vendor_ids():
    """Posodobi vendor_id (in supplier če manjka) v vseh obstoječih meta.json datotekah.
    Če 'supplier' polje manjka, ga proba ugotoviti iz 'vendor_name' z heuristiko."""
    try:
        updated = 0
        skipped = 0
        details = []

        # Heuristika za ugibanje supplier-ja iz vendor_name (lowercase substring match)
        VENDOR_NAME_HINTS = {
            "amio":       ["amio", "suban"],
            "abakus":     ["abakus"],
            "intercars":  ["intercars", "inter cars", "inter-cars"],
            "motoprofil": ["motoprofil", "moto profil", "moto-profil"],
            "ikonka":     ["ikonka", "kik"],
        }

        for d in PREVZEMI_DIR.iterdir():
            if not d.is_dir():
                continue
            meta_path = d / "meta.json"
            if not meta_path.exists():
                continue
            try:
                meta = json.loads(meta_path.read_text(encoding="utf-8"))
                changed = False

                supplier_key = (meta.get("supplier") or "").lower()
                invoice_num = meta.get("invoice_number") or ""
                vendor_name = meta.get("vendor_name") or ""

                # Če supplier polje manjka, ga proba ugotoviti iz vendor_name
                if not supplier_key:
                    vname_l = vendor_name.lower()
                    for sup, hints in VENDOR_NAME_HINTS.items():
                        if any(h in vname_l for h in hints):
                            supplier_key = sup
                            meta["supplier"] = sup
                            meta["supplier_name"] = SUPPLIER_NAMES.get(sup, sup)
                            changed = True
                            break

                # OVERRIDE: če invoice ali vendor_name kažeta na drugega supplier-ja, popravi
                # (npr. "FA 162383/..." invoice = MotoProfil, ne AMiO)
                if supplier_key:
                    detected = _override_supplier_by_invoice(supplier_key, invoice_num, vendor_name)
                    if detected != supplier_key:
                        print(f"[backfill] Override {d.name}: {supplier_key} -> {detected} (invoice={invoice_num}, vendor={vendor_name})")
                        meta["supplier"] = detected
                        meta["supplier_name"] = SUPPLIER_NAMES.get(detected, detected)
                        meta["vendor_id"] = _load_vendor_ids().get(detected, "")
                        supplier_key = detected
                        changed = True

                # Posodobi vendor_id če manjka in supplier obstaja v mapping-u
                if supplier_key in SUPPLIER_VENDOR_IDS and not meta.get("vendor_id"):
                    meta["vendor_id"] = SUPPLIER_VENDOR_IDS[supplier_key]
                    changed = True

                if changed:
                    meta_path.write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")
                    updated += 1
                    details.append({"id": d.name, "supplier": supplier_key, "vendor_id": meta.get("vendor_id", "")})
                else:
                    skipped += 1
            except Exception as e:
                skipped += 1
                details.append({"id": d.name, "error": str(e)})

        return {
            "ok": True,
            "updated": updated,
            "skipped": skipped,
            "mapping": SUPPLIER_VENDOR_IDS,
            "details": details[:50],  # samo prvih 50 za debugging
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/prevzemi-parse-supplier")
async def prevzemi_parse_supplier(file: UploadFile = File(...)):
    """Auto-detect supplier + parse with appropriate parser. Returns same JSON as PDF endpoint."""
    try:
        content = await file.read()
        supplier = _detect_supplier(file.filename or "", content)

        if supplier == "unknown":
            return {"ok": False, "error": "Dobavitelj ni prepoznan. Podprti: AMiO, MotoProfil, Ikonka, Intercars, Abakus."}

        # Parse glede na dobavitelja
        if supplier == "amio_xml":
            # AMiO strukturiran XML — parse brez Claude (instant, $0)
            parsed = _parse_amio_xml(content)
            supplier = "amio"  # za vendor_id mapping je amio
        elif supplier == "motoprofil":
            parsed = _parse_motoprofil_csv(content, file.filename or "")
        elif supplier == "intercars":
            parsed = _parse_intercars_xml(content)
        elif supplier == "abakus":
            parsed = _parse_abakus_xlsx(content)
        elif supplier == "ikonka":
            # Ikonka rabi Claude (blocking) — v executor da ne blokira healthz
            _loop = asyncio.get_event_loop()
            parsed = await _loop.run_in_executor(None, lambda: _parse_ikonka_pdf(content))
        elif supplier == "amio":
            # Fallback na obstoječi PDF parser (Claude) — če je PDF, ne XML
            return {"ok": False, "error": "Za AMiO PDF uporabi /prevzemi-parse-pdf endpoint", "redirect": "amio"}
        else:
            return {"ok": False, "error": f"Parser za '{supplier}' še ni implementiran"}

        if not parsed.get("items"):
            return {"ok": False, "error": "Ni najdenih izdelkov v fajlu"}

        # Shrani na disk (enako kot /prevzemi-parse-pdf)
        from datetime import datetime
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        # Normaliziraj — odstrani presledke iz invoice_number (npr. 'FA 162383' → 'FA_162383')
        if parsed.get('invoice_number'):
            parsed['invoice_number'] = _normalize_invoice_number(parsed['invoice_number'])
        invoice_num_safe = _safe_filename(parsed.get('invoice_number', 'unknown'))
        record_id = f"{ts}_{invoice_num_safe}"
        target_dir = PREVZEMI_DIR / record_id
        target_dir.mkdir(parents=True, exist_ok=True)

        # Save source file
        safe_filename = _safe_filename(file.filename or "invoice")
        (target_dir / f'source_{safe_filename}').write_bytes(content)
        # Save parsed JSON
        (target_dir / 'parsed.json').write_text(json.dumps(parsed, indent=2, ensure_ascii=False), encoding='utf-8')
        # Save metadata
        # Override supplier če pattern invoice številke ali vendor_name kaže drugače
        detected_supplier = _override_supplier_by_invoice(
            supplier,
            parsed.get('invoice_number', ''),
            parsed.get('vendor_name', '')
        )
        if detected_supplier != supplier:
            print(f"[prevzemi-supplier] Override: {supplier} -> {detected_supplier} (invoice={parsed.get('invoice_number')}, vendor={parsed.get('vendor_name')})")
            supplier = detected_supplier
        meta = {
            'record_id': record_id,
            'original_filename': file.filename or 'invoice',
            'created_ts': ts,
            'supplier': supplier,
            'supplier_name': SUPPLIER_NAMES.get(supplier, supplier),
            'vendor_id': SUPPLIER_VENDOR_IDS.get(supplier, ''),
            'invoice_number': parsed.get('invoice_number', ''),
            'invoice_date': parsed.get('invoice_date', ''),
            'vendor_name': parsed.get('vendor_name', ''),
            'item_count': len(parsed.get('items', [])),
        }
        (target_dir / 'meta.json').write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding='utf-8')

        return {
            "ok": True,
            "record_id": record_id,
            "supplier": supplier,
            "supplier_name": SUPPLIER_NAMES.get(supplier, supplier),
            "missing_names": parsed.get("_missing_names", False),
            "parsed": parsed,
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


class UpdateItemsRequest(BaseModel):
    record_id: str
    items: list  # list of {product_number, product_name, qty, unit_price, value}


@app.post("/prevzemi-update-items")
async def prevzemi_update_items(req: UpdateItemsRequest):
    """Update items in a parsed record (npr. po Intercars name vpisu)."""
    try:
        rec_safe = _safe_filename(req.record_id)
        target = PREVZEMI_DIR / rec_safe / 'parsed.json'
        if not target.exists():
            return {"ok": False, "error": "Record not found"}
        parsed = json.loads(target.read_text(encoding='utf-8'))

        # Update items (replace whole array — frontend pošlje vse)
        parsed["items"] = req.items
        # Reset missing flag
        parsed.pop("_missing_names", None)

        target.write_text(json.dumps(parsed, indent=2, ensure_ascii=False), encoding='utf-8')
        return {"ok": True, "item_count": len(req.items)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


# ─── SHIPPING COSTS ANALYSIS ────────────────────────────────────────────────

SHIPPING_DIR = DATA_DIR / "shipping"
SHIPPING_DIR.mkdir(exist_ok=True, parents=True)


def _parse_shipping_xls(content: bytes) -> dict:
    """Parse shipping protocol/descr XLS file. Returns structured data with month auto-detected."""
    import io
    from collections import Counter
    from datetime import datetime as _dt

    try:
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_filelike(io.BytesIO(content))
    except ImportError:
        return {"ok": False, "error": "python-calamine ni instaliran v requirements.txt"}
    except Exception as e:
        return {"ok": False, "error": f"Cannot read XLS: {e}"}

    sheet_name = wb.sheet_names[0]
    ws = wb.get_sheet_by_name(sheet_name)
    data = ws.to_python()

    if len(data) < 4:
        return {"ok": False, "error": "Fajl je premajhen ali nima podatkov"}

    # Headers v vrstici 2
    headers = data[2]
    if not headers or "Date" not in str(headers[1]):
        return {"ok": False, "error": f"Headers ne najdejo v vrstici 2: {headers[:5]}"}

    # Indeksi stolpcev (od headers)
    col_map = {}
    for i, h in enumerate(headers):
        h_str = str(h).strip()
        col_map[h_str] = i

    def get(row, key, default=None):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    def num(v):
        if v is None or v == "":
            return 0.0
        try:
            return float(str(v).replace(",", ".").replace(" ", ""))
        except (ValueError, TypeError):
            return 0.0

    # Parse rows
    rows = []
    months_seen = Counter()
    for row in data[3:]:
        if not row or len(row) < 5:
            continue
        date_str = str(get(row, "Date", "") or "")
        # Skip "Total:" row
        if str(get(row, "ID", "") or "").lower().startswith("total"):
            continue
        if not date_str or "-" not in date_str:
            continue

        # Parse date
        try:
            d = _dt.strptime(date_str[:10], "%Y-%m-%d")
            month_key = d.strftime("%Y-%m")
            months_seen[month_key] += 1
        except Exception:
            continue

        rows.append({
            "date": date_str[:10],
            "waybill": str(get(row, "Waybill number", "") or ""),
            "order_number": str(get(row, "Order number", "") or ""),
            "country": str(get(row, "Country", "") or "").upper(),
            "service_type": str(get(row, "Service Type", "") or ""),
            "city": str(get(row, "City", "") or ""),
            "cod_eur": num(get(row, "COD in EUR", 0)),
            "weight_kg": num(get(row, "Weight in kg", 0)),
            "sum_before_disc": num(get(row, "Sum before disc EUR", 0)),
            "disc_pct": num(get(row, "Disc. %", 0)),
            "disc_sum": num(get(row, "Disc. sum EUR", 0)),
            "weight_price": num(get(row, "Weight price EUR", 0)),
            "cod_price": num(get(row, "COD Price EUR", 0)),
            "sms_price": num(get(row, "SMS Price EUR", 0)),
            "insurance_price": num(get(row, "Insurance price EUR", 0)),
            "fuel_tax_pct": num(get(row, "Fuel tax %", 0)),
            "fuel_tax": num(get(row, "Fuel tax EUR", 0)) or num(get(row, "Fuel tax price EUR", 0)),
            "extra_cost": num(get(row, "Extra cost EUR", 0)),
            "total_no_vat": num(get(row, "Total without VAT EUR", 0)),
            "courier": str(get(row, "Courier", "") or ""),
        })

    if not rows:
        return {"ok": False, "error": "Ni veljavnih vrstic s paketi"}

    # Determine month: take most common (>= 50%)
    if not months_seen:
        return {"ok": False, "error": "Ni najdenih datumov"}
    top_month, top_count = months_seen.most_common(1)[0]
    coverage = top_count / sum(months_seen.values())

    return {
        "ok": True,
        "month": top_month,
        "coverage": round(coverage, 3),
        "all_months": dict(months_seen),
        "rows": rows,
        "count": len(rows),
    }


def _aggregate_shipping_rows(rows: list, month: str = None, kind: str = "inventory") -> dict:
    """Sestavi vse agregate iz seznama rows. Vrne summary objekt.
    Vključuje fee fingerprints za anomaly detection."""
    from collections import defaultdict, Counter

    def safe_total(r):
        t = r.get("total_no_vat") or 0
        if t > 0:
            return float(t)
        return (r.get("weight_price") or 0) + (r.get("cod_price") or 0) + (r.get("sms_price") or 0) + \
               (r.get("insurance_price") or 0) + (r.get("fuel_tax") or 0) + (r.get("extra_cost") or 0) - \
               (r.get("disc_sum") or 0)

    def is_zero(r):
        return (r.get("total_no_vat") or 0) == 0

    def is_successful_cod(r):
        return (r.get("cod_eur") or 0) > 0 and (r.get("total_no_vat") or 0) > 0

    # Standard KPIs
    total_cost = 0.0
    total_weight = 0.0
    by_country = defaultdict(lambda: {"count": 0, "cost": 0.0, "weight": 0.0, "cod_count": 0,
                                       "cod_success": 0, "zero_count": 0, "cod_amount_sum": 0.0,
                                       "disc_sum": 0.0})
    by_courier = defaultdict(int)
    cost_breakdown = {"weight": 0.0, "cod": 0.0, "sms": 0.0, "insurance": 0.0,
                      "fuel": 0.0, "extra": 0.0, "disc": 0.0}
    weight_buckets = [
        {"label": "< 0.5 kg", "min": 0, "max": 0.5, "count": 0, "cost": 0.0},
        {"label": "0.5 - 1 kg", "min": 0.5, "max": 1, "count": 0, "cost": 0.0},
        {"label": "1 - 2 kg", "min": 1, "max": 2, "count": 0, "cost": 0.0},
        {"label": "2 - 5 kg", "min": 2, "max": 5, "count": 0, "cost": 0.0},
        {"label": "5 - 10 kg", "min": 5, "max": 10, "count": 0, "cost": 0.0},
        {"label": "> 10 kg", "min": 10, "max": 1e9, "count": 0, "cost": 0.0},
    ]
    heaviest = None
    priciest = None
    cod_rows_summary = []

    # === FEE FINGERPRINTS za anomaly detection (per country) ===
    fee_fingerprint = defaultdict(lambda: {
        "extra_cost_values": Counter(),       # {0.08: 1500, 0.12: 200, ...}
        "extra_cost_count": 0,
        "extra_cost_sum": 0.0,
        "fuel_tax_pct_values": Counter(),     # {5.5: 4000, 4.5: 100}
        "fuel_tax_count": 0,
        "fuel_tax_sum": 0.0,
        "sms_count": 0,
        "sms_sum": 0.0,
        "insurance_count": 0,
        "insurance_sum": 0.0,
        "cod_price_count": 0,
        "cod_price_sum": 0.0,
        "weight_price_sum": 0.0,
        "weight_price_count": 0,
    })

    for r in rows:
        cost = safe_total(r)
        weight = r.get("weight_kg") or 0
        country = r.get("country") or ""

        total_cost += cost
        total_weight += weight

        c = by_country[country]
        c["count"] += 1
        c["cost"] += cost
        c["weight"] += weight
        c["disc_sum"] += r.get("disc_sum") or 0
        if (r.get("cod_eur") or 0) > 0:
            c["cod_count"] += 1
            c["cod_amount_sum"] += r.get("cod_eur") or 0
        if is_successful_cod(r):
            c["cod_success"] += 1
        if is_zero(r):
            c["zero_count"] += 1

        by_courier[r.get("courier") or "Unknown"] += 1

        cost_breakdown["weight"] += r.get("weight_price") or 0
        cost_breakdown["cod"] += r.get("cod_price") or 0
        cost_breakdown["sms"] += r.get("sms_price") or 0
        cost_breakdown["insurance"] += r.get("insurance_price") or 0
        cost_breakdown["fuel"] += r.get("fuel_tax") or 0
        cost_breakdown["extra"] += r.get("extra_cost") or 0
        cost_breakdown["disc"] += r.get("disc_sum") or 0

        for b in weight_buckets:
            if b["min"] <= weight < b["max"]:
                b["count"] += 1
                b["cost"] += cost
                break

        if heaviest is None or weight > (heaviest.get("weight_kg") or 0):
            heaviest = {"waybill": r.get("waybill"), "country": country, "city": r.get("city"),
                        "weight_kg": weight, "cost": cost}
        if priciest is None or cost > (priciest.get("cost") or 0):
            priciest = {"waybill": r.get("waybill"), "country": country, "city": r.get("city"),
                        "weight_kg": weight, "cost": cost}

        if is_successful_cod(r):
            cod_rows_summary.append({
                "country": country,
                "month": month,
                "cost": cost,
                "weight": weight,
                "cod_amount": r.get("cod_eur") or 0,
            })

        # === Update fee fingerprint za to državo ===
        fp = fee_fingerprint[country]
        ec = round(r.get("extra_cost") or 0, 4)
        if ec > 0:
            fp["extra_cost_values"][ec] += 1
            fp["extra_cost_count"] += 1
            fp["extra_cost_sum"] += ec
        ftp = round(r.get("fuel_tax_pct") or 0, 2)
        if ftp > 0:
            fp["fuel_tax_pct_values"][ftp] += 1
        ft = r.get("fuel_tax") or 0
        if ft > 0:
            fp["fuel_tax_count"] += 1
            fp["fuel_tax_sum"] += ft
        sms = r.get("sms_price") or 0
        if sms > 0:
            fp["sms_count"] += 1
            fp["sms_sum"] += sms
        ins = r.get("insurance_price") or 0
        if ins > 0:
            fp["insurance_count"] += 1
            fp["insurance_sum"] += ins
        cod_p = r.get("cod_price") or 0
        if cod_p > 0:
            fp["cod_price_count"] += 1
            fp["cod_price_sum"] += cod_p
        wp = r.get("weight_price") or 0
        if wp > 0:
            fp["weight_price_sum"] += wp
            fp["weight_price_count"] += 1

    # Convert Counter to dict for JSON serialization
    fee_fp_clean = {}
    for c, fp in fee_fingerprint.items():
        fee_fp_clean[c] = {
            "extra_cost_values": dict(fp["extra_cost_values"]),
            "extra_cost_count": fp["extra_cost_count"],
            "extra_cost_sum": fp["extra_cost_sum"],
            "fuel_tax_pct_values": dict(fp["fuel_tax_pct_values"]),
            "fuel_tax_count": fp["fuel_tax_count"],
            "fuel_tax_sum": fp["fuel_tax_sum"],
            "sms_count": fp["sms_count"],
            "sms_sum": fp["sms_sum"],
            "insurance_count": fp["insurance_count"],
            "insurance_sum": fp["insurance_sum"],
            "cod_price_count": fp["cod_price_count"],
            "cod_price_sum": fp["cod_price_sum"],
            "weight_price_sum": fp["weight_price_sum"],
            "weight_price_count": fp["weight_price_count"],
        }

    cod_count_total = sum(c["cod_count"] for c in by_country.values())
    cod_success_total = sum(c["cod_success"] for c in by_country.values())
    zero_total = sum(c["zero_count"] for c in by_country.values())

    return {
        "month": month,
        "kind": kind,
        "row_count": len(rows),
        "total_cost": total_cost,
        "total_weight": total_weight,
        "by_country": dict(by_country),
        "by_courier": dict(by_courier),
        "cost_breakdown": cost_breakdown,
        "weight_buckets": weight_buckets,
        "heaviest": heaviest,
        "priciest": priciest,
        "cod_count_total": cod_count_total,
        "cod_success_total": cod_success_total,
        "zero_total": zero_total,
        "cod_rows": cod_rows_summary,
        "fee_fingerprint": fee_fp_clean,
    }


def _build_shipping_summary_for_month(month_dir):
    """Iz rows_<kind>.json gradi summary.json za en mesec."""
    for kind in ["inventory", "descr", "other"]:
        rf = month_dir / f"rows_{kind}.json"
        if rf.exists():
            try:
                rows = json.loads(rf.read_text(encoding='utf-8'))
                summary = _aggregate_shipping_rows(rows, month=month_dir.name, kind=kind)
                (month_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False), encoding='utf-8')
                return summary
            except Exception as e:
                print(f"[shipping-summary] Failed for {month_dir.name}: {e}")
    return None


@app.post("/shipping-upload")
async def shipping_upload(file: UploadFile = File(...)):
    """Naloži XLS fajl pošte. Auto-detect mesec, shrani v /data/shipping/YYYY-MM/."""
    try:
        content = await file.read()
        if len(content) > 100 * 1024 * 1024:
            return {"ok": False, "error": "Fajl prevelik (max 100MB)"}

        parsed = _parse_shipping_xls(content)
        if not parsed.get("ok"):
            return parsed

        month = parsed["month"]
        # Določi tip fajla iz imena
        fn = (file.filename or "").lower()
        if fn.startswith("descr"):
            kind = "descr"
        elif "inventory" in fn:
            kind = "inventory"
        else:
            kind = "other"

        # Shrani v /data/shipping/YYYY-MM/
        month_dir = SHIPPING_DIR / month
        month_dir.mkdir(parents=True, exist_ok=True)

        # Source file
        safe_name = _safe_filename(file.filename or f"shipping_{month}.xls")
        (month_dir / safe_name).write_bytes(content)

        # Parsed JSON (lahko se prepiše če uploadaš oba file-a)
        parsed_file = month_dir / f"parsed_{kind}.json"
        meta = {k: v for k, v in parsed.items() if k != "rows"}
        meta["kind"] = kind
        meta["original_filename"] = file.filename
        meta["uploaded_at"] = parsed.get("month", "")
        from datetime import datetime as _dt2
        meta["uploaded_ts"] = _dt2.now().isoformat()
        parsed_file.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
        (month_dir / f"rows_{kind}.json").write_text(json.dumps(parsed["rows"], ensure_ascii=False), encoding='utf-8')

        # Build summary cache TAKOJ ob uploadu
        summary = _aggregate_shipping_rows(parsed["rows"], month=month, kind=kind)
        (month_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False), encoding='utf-8')

        return {
            "ok": True,
            "month": month,
            "kind": kind,
            "count": parsed["count"],
            "coverage": parsed["coverage"],
            "all_months": parsed["all_months"],
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/shipping-summary")
async def shipping_summary():
    """Vrne agregirane summary podatke iz vseh mesecev. RES HITRO."""
    try:
        summaries = {}
        for month_dir in sorted(SHIPPING_DIR.iterdir()):
            if not month_dir.is_dir():
                continue
            sf = month_dir / "summary.json"
            # Če summary ne obstaja (stari upload), zgradi zdaj
            if not sf.exists():
                _build_shipping_summary_for_month(month_dir)
            if sf.exists():
                try:
                    summaries[month_dir.name] = json.loads(sf.read_text(encoding='utf-8'))
                except Exception as e:
                    print(f"[shipping-summary] Read failed for {month_dir.name}: {e}")
                    continue
        return {"ok": True, "summaries": summaries}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.post("/shipping-rebuild-cache")
async def shipping_rebuild_cache():
    """Force-rebuild summary.json za vse mesece (npr. po update-u logike agregacije)."""
    try:
        rebuilt = []
        for month_dir in sorted(SHIPPING_DIR.iterdir()):
            if not month_dir.is_dir():
                continue
            s = _build_shipping_summary_for_month(month_dir)
            if s:
                rebuilt.append({"month": month_dir.name, "rows": s.get("row_count", 0)})
        return {"ok": True, "rebuilt": rebuilt}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/shipping-anomalies")
async def shipping_anomalies():
    """Mesec-na-mesec anomaly detection. Vrne seznam zaznanih anomalij po državah."""
    try:
        # Naloži vse summarije
        summaries = {}
        for month_dir in sorted(SHIPPING_DIR.iterdir()):
            if not month_dir.is_dir():
                continue
            sf = month_dir / "summary.json"
            if not sf.exists():
                _build_shipping_summary_for_month(month_dir)
            if sf.exists():
                try:
                    summaries[month_dir.name] = json.loads(sf.read_text(encoding='utf-8'))
                except Exception:
                    continue

        months = sorted(summaries.keys())
        if len(months) < 2:
            return {"ok": False, "error": "Potrebno je vsaj 2 meseca za primerjavo"}

        anomalies = []

        # Za vsako državo primerjaj zadnji mesec z prejšnjim + povprečjem ostalih
        for i in range(1, len(months)):
            curr_m = months[i]
            prev_m = months[i-1]
            curr = summaries[curr_m].get("fee_fingerprint", {})
            prev = summaries[prev_m].get("fee_fingerprint", {})
            curr_by_c = summaries[curr_m].get("by_country", {})
            prev_by_c = summaries[prev_m].get("by_country", {})

            # Skupek vseh držav v obeh
            countries = set(list(curr.keys()) + list(prev.keys()))

            for country in countries:
                cfp = curr.get(country, {})
                pfp = prev.get(country, {})
                ccount = curr_by_c.get(country, {}).get("count", 0)
                pcount = prev_by_c.get(country, {}).get("count", 0)

                # Skip države z malo paketov (premalo signala)
                if ccount < 20 or pcount < 20:
                    continue

                # ── ANOMALY 1: Novi extra_cost vrednosti ──
                curr_ec_vals = set(float(v) for v in cfp.get("extra_cost_values", {}).keys())
                prev_ec_vals = set(float(v) for v in pfp.get("extra_cost_values", {}).keys())
                new_ec_vals = curr_ec_vals - prev_ec_vals
                if new_ec_vals:
                    for val in sorted(new_ec_vals):
                        affected = cfp.get("extra_cost_values", {}).get(str(val), 0) or cfp.get("extra_cost_values", {}).get(val, 0)
                        if affected > 5:  # min 5 paketov da signaliziramo
                            estimated_impact = val * affected
                            anomalies.append({
                                "type": "new_extra_cost",
                                "severity": "high" if affected > 100 else "medium",
                                "country": country,
                                "month": curr_m,
                                "prev_month": prev_m,
                                "title": f"Nov 'extra cost' {val:.2f}€ pri {country}",
                                "details": f"V {curr_m} se je pojavila nova vrednost 'extra cost' = {val:.2f}€ pri {affected} paketih. V {prev_m} te vrednosti ni bilo.",
                                "value": val,
                                "affected_packages": affected,
                                "estimated_monthly_impact": round(estimated_impact, 2),
                            })

                # ── ANOMALY 2: Extra cost frekvenca skok ──
                # Delež paketov z extra_cost
                pec_pct = pfp.get("extra_cost_count", 0) / pcount * 100 if pcount else 0
                cec_pct = cfp.get("extra_cost_count", 0) / ccount * 100 if ccount else 0
                # Če je skočil za 20%+ in v abs. več kot 100 paketov
                if cec_pct - pec_pct > 20 and cfp.get("extra_cost_count", 0) > 100:
                    new_packages = cfp.get("extra_cost_count", 0) - pfp.get("extra_cost_count", 0)
                    estimated_impact = cfp.get("extra_cost_sum", 0) - pfp.get("extra_cost_sum", 0)
                    anomalies.append({
                        "type": "extra_cost_frequency_jump",
                        "severity": "high",
                        "country": country,
                        "month": curr_m,
                        "prev_month": prev_m,
                        "title": f"Eksplozija 'extra cost' pri {country}",
                        "details": f"Delež paketov z extra cost je skočil iz {pec_pct:.1f}% ({pfp.get('extra_cost_count', 0)} paketov) na {cec_pct:.1f}% ({cfp.get('extra_cost_count', 0)} paketov)",
                        "value": cec_pct - pec_pct,
                        "affected_packages": cfp.get("extra_cost_count", 0),
                        "estimated_monthly_impact": round(estimated_impact, 2),
                    })

                # ── ANOMALY 3: Fuel tax % sprememba ──
                curr_ft_pcts = set(float(v) for v in cfp.get("fuel_tax_pct_values", {}).keys())
                prev_ft_pcts = set(float(v) for v in pfp.get("fuel_tax_pct_values", {}).keys())
                # Nove vrednosti %
                new_ft_pcts = curr_ft_pcts - prev_ft_pcts
                gone_ft_pcts = prev_ft_pcts - curr_ft_pcts
                # Če je sprememba
                if new_ft_pcts and prev_ft_pcts:
                    curr_max = max(curr_ft_pcts) if curr_ft_pcts else 0
                    prev_max = max(prev_ft_pcts) if prev_ft_pcts else 0
                    if curr_max > prev_max * 1.05:  # 5%+ povečanje
                        diff_pct = curr_max - prev_max
                        prev_fuel_sum = pfp.get("fuel_tax_sum", 0)
                        curr_fuel_sum = cfp.get("fuel_tax_sum", 0)
                        impact = curr_fuel_sum - prev_fuel_sum
                        anomalies.append({
                            "type": "fuel_tax_pct_increase",
                            "severity": "high" if (curr_max - prev_max) > 1 else "medium",
                            "country": country,
                            "month": curr_m,
                            "prev_month": prev_m,
                            "title": f"Povišan fuel tax pri {country}: {prev_max:.1f}% → {curr_max:.1f}%",
                            "details": f"Najvišji fuel tax procent se je povišal iz {prev_max:.1f}% na {curr_max:.1f}% (+{diff_pct:.1f} odstotne točke). Skupni fuel tax stroški: {prev_fuel_sum:.2f}€ → {curr_fuel_sum:.2f}€",
                            "value": diff_pct,
                            "affected_packages": cfp.get("fuel_tax_count", 0),
                            "estimated_monthly_impact": round(impact, 2),
                        })

                # ── ANOMALY 4: Povprečna teža-cena sprememba ──
                # Pov. weight price = total weight_price / count (uporabni za primerjavo)
                if ccount > 50 and pcount > 50:
                    p_avg_wp = pfp.get("weight_price_sum", 0) / pcount if pcount else 0
                    c_avg_wp = cfp.get("weight_price_sum", 0) / ccount if ccount else 0
                    if p_avg_wp > 0.1:
                        change_pct = (c_avg_wp - p_avg_wp) / p_avg_wp * 100
                        if abs(change_pct) > 8:  # 8%+ sprememba pov. cene/paket
                            impact = (c_avg_wp - p_avg_wp) * ccount
                            anomalies.append({
                                "type": "weight_price_change",
                                "severity": "high" if abs(change_pct) > 15 else "medium",
                                "country": country,
                                "month": curr_m,
                                "prev_month": prev_m,
                                "title": f"Povp. cena tehtanja {country}: {p_avg_wp:.2f}€ → {c_avg_wp:.2f}€ ({change_pct:+.1f}%)",
                                "details": f"Povp. weight price/paket se je {'povečal' if change_pct > 0 else 'znižal'} za {abs(change_pct):.1f}%. ({pcount} paketov v {prev_m}, {ccount} paketov v {curr_m})",
                                "value": change_pct,
                                "affected_packages": ccount,
                                "estimated_monthly_impact": round(impact, 2),
                            })

                # ── ANOMALY 5: SMS / Insurance fee novi ali porasel ──
                p_sms_pct = pfp.get("sms_count", 0) / pcount * 100 if pcount else 0
                c_sms_pct = cfp.get("sms_count", 0) / ccount * 100 if ccount else 0
                if c_sms_pct - p_sms_pct > 15 and cfp.get("sms_count", 0) > 50:
                    impact = cfp.get("sms_sum", 0) - pfp.get("sms_sum", 0)
                    anomalies.append({
                        "type": "sms_fee_increase",
                        "severity": "medium",
                        "country": country,
                        "month": curr_m,
                        "prev_month": prev_m,
                        "title": f"SMS fee povečan pri {country}: {p_sms_pct:.1f}% → {c_sms_pct:.1f}%",
                        "details": f"SMS fee se zdaj zaračuna {c_sms_pct:.1f}% paketov (prej {p_sms_pct:.1f}%)",
                        "value": c_sms_pct - p_sms_pct,
                        "affected_packages": cfp.get("sms_count", 0),
                        "estimated_monthly_impact": round(impact, 2),
                    })

                p_ins_pct = pfp.get("insurance_count", 0) / pcount * 100 if pcount else 0
                c_ins_pct = cfp.get("insurance_count", 0) / ccount * 100 if ccount else 0
                if c_ins_pct - p_ins_pct > 15 and cfp.get("insurance_count", 0) > 50:
                    impact = cfp.get("insurance_sum", 0) - pfp.get("insurance_sum", 0)
                    anomalies.append({
                        "type": "insurance_fee_increase",
                        "severity": "medium",
                        "country": country,
                        "month": curr_m,
                        "prev_month": prev_m,
                        "title": f"Insurance fee povečan pri {country}: {p_ins_pct:.1f}% → {c_ins_pct:.1f}%",
                        "details": f"Insurance se zdaj zaračuna {c_ins_pct:.1f}% paketov (prej {p_ins_pct:.1f}%)",
                        "value": c_ins_pct - p_ins_pct,
                        "affected_packages": cfp.get("insurance_count", 0),
                        "estimated_monthly_impact": round(impact, 2),
                    })

                # ── ANOMALY 6: Skupna cena/paket sprememba (samo ne-zero paketi) ──
                p_data = prev_by_c.get(country, {})
                c_data = curr_by_c.get(country, {})
                p_nonzero = p_data.get("count", 0) - p_data.get("zero_count", 0)
                c_nonzero = c_data.get("count", 0) - c_data.get("zero_count", 0)
                if p_nonzero > 30 and c_nonzero > 30:
                    p_avg = p_data.get("cost", 0) / p_nonzero if p_nonzero else 0
                    c_avg = c_data.get("cost", 0) / c_nonzero if c_nonzero else 0
                    if p_avg > 0.5:
                        chg_pct = (c_avg - p_avg) / p_avg * 100
                        if abs(chg_pct) > 7:
                            impact = (c_avg - p_avg) * c_nonzero
                            anomalies.append({
                                "type": "avg_total_change",
                                "severity": "high" if abs(chg_pct) > 15 else "medium",
                                "country": country,
                                "month": curr_m,
                                "prev_month": prev_m,
                                "title": f"Povp. skupna cena {country}: {p_avg:.2f}€ → {c_avg:.2f}€ ({chg_pct:+.1f}%)",
                                "details": f"Povprečna cena/paket (brez neprevzetih) se je spremenila za {chg_pct:+.1f}%",
                                "value": chg_pct,
                                "affected_packages": c_nonzero,
                                "estimated_monthly_impact": round(impact, 2),
                            })

        # Sortiraj: severity high → medium, znotraj po impact
        sev_order = {"high": 0, "medium": 1, "low": 2}
        anomalies.sort(key=lambda a: (sev_order.get(a["severity"], 9), -abs(a.get("estimated_monthly_impact", 0))))

        return {
            "ok": True,
            "anomalies": anomalies,
            "months_analyzed": months,
            "count": len(anomalies),
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/shipping-list")
async def shipping_list():
    """Vrne seznam vseh mesecev z naloženimi fajli."""
    try:
        months = []
        for month_dir in sorted(SHIPPING_DIR.iterdir(), reverse=True):
            if not month_dir.is_dir():
                continue
            month = month_dir.name
            # Najdi vse parsed_*.json
            kinds = []
            total_count = 0
            for kind in ["inventory", "descr", "other"]:
                pf = month_dir / f"parsed_{kind}.json"
                if pf.exists():
                    try:
                        meta = json.loads(pf.read_text(encoding='utf-8'))
                        kinds.append({
                            "kind": kind,
                            "count": meta.get("count", 0),
                            "filename": meta.get("original_filename", ""),
                            "uploaded_ts": meta.get("uploaded_ts", ""),
                        })
                        total_count += meta.get("count", 0)
                    except Exception:
                        continue
            if kinds:
                months.append({"month": month, "kinds": kinds, "total_count": total_count})
        return {"ok": True, "months": months}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/shipping-data")
async def shipping_data(month: str = "all"):
    """Vrne paketne podatke za en mesec ali vse. Če sta inventory + descr, preferira inventory (more columns)."""
    try:
        result_rows = []
        months_meta = {}

        target_dirs = []
        if month == "all":
            for d in SHIPPING_DIR.iterdir():
                if d.is_dir():
                    target_dirs.append(d)
        else:
            d = SHIPPING_DIR / month
            if d.exists():
                target_dirs.append(d)

        for month_dir in target_dirs:
            m = month_dir.name
            # Preferira inventory > descr > other (inventory ima Total without VAT + Courier)
            for kind in ["inventory", "descr", "other"]:
                rf = month_dir / f"rows_{kind}.json"
                if rf.exists():
                    try:
                        rows = json.loads(rf.read_text(encoding='utf-8'))
                        for r in rows:
                            r["_month"] = m
                            r["_kind"] = kind
                        result_rows.extend(rows)
                        months_meta[m] = {"kind": kind, "count": len(rows)}
                        break  # samo en kind per mesec
                    except Exception:
                        continue

        return {"ok": True, "rows": result_rows, "months": months_meta, "total": len(result_rows)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.delete("/shipping-month/{month}")
async def shipping_delete(month: str):
    """Briše vse fajle za en mesec."""
    try:
        month_safe = month.replace("/", "").replace("..", "")
        d = SHIPPING_DIR / month_safe
        if not d.exists():
            return {"ok": False, "error": "Mesec ne obstaja"}
        import shutil
        shutil.rmtree(d)
        return {"ok": True, "deleted": month_safe}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/shipping-debug")
async def shipping_debug():
    """Diagnostika: pokaže pot, ali obstaja, ali je persistent, koliko prostora, koliko fajlov."""
    import shutil
    from datetime import datetime as _dtdbg
    try:
        info = {
            "data_dir": str(DATA_DIR),
            "data_dir_exists": DATA_DIR.exists(),
            "data_dir_is_writable": os.access(str(DATA_DIR), os.W_OK) if DATA_DIR.exists() else False,
            "shipping_dir": str(SHIPPING_DIR),
            "shipping_dir_exists": SHIPPING_DIR.exists(),
            "env_DATA_DIR": os.environ.get("DATA_DIR", "(not set, default /data)"),
        }
        if DATA_DIR.exists():
            usage = shutil.disk_usage(str(DATA_DIR))
            info["disk_total_gb"] = round(usage.total / 1024**3, 2)
            info["disk_used_gb"] = round(usage.used / 1024**3, 2)
            info["disk_free_gb"] = round(usage.free / 1024**3, 2)
        # Files in shipping
        files = []
        if SHIPPING_DIR.exists():
            for month_dir in sorted(SHIPPING_DIR.iterdir()):
                if not month_dir.is_dir():
                    continue
                month_files = []
                for f in month_dir.iterdir():
                    month_files.append({
                        "name": f.name,
                        "size_kb": round(f.stat().st_size / 1024, 1),
                        "modified": _dtdbg.fromtimestamp(f.stat().st_mtime).isoformat(),
                    })
                files.append({"month": month_dir.name, "files": month_files})
        info["shipping_contents"] = files
        # Also count all top-level items in /data
        if DATA_DIR.exists():
            top = []
            for f in DATA_DIR.iterdir():
                top.append({
                    "name": f.name,
                    "is_dir": f.is_dir(),
                    "size_kb": round(f.stat().st_size / 1024, 1) if f.is_file() else None,
                })
            info["data_dir_top_level"] = top
        return info
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}


# ════════════════════════════════════════════════════════════════════
#  POTNI NALOGI (Knjigovodstvo → tab "Potni nalogi")
#  Nalog za službeno potovanje SUBAN d.o.o. — vnos, samodejni izračun
#  dnevnice (SLO uredba) in kilometrine (Google Maps), zgodovina,
#  izvoz v XLSX (predloga) in PDF.
# ════════════════════════════════════════════════════════════════════
PN_DIR = DATA_DIR / "potni_nalogi"
PN_DIR.mkdir(exist_ok=True, parents=True)
PN_STATE = PN_DIR / "nalogi.json"          # vsi nalogi (seznam dict)
PN_BOOK = PN_DIR / "imeniki.json"          # stalni imeniki: zaposleni, vozila, relacije
PN_TEMPLATE = Path("static") / "potni_nalog_template.xlsx"


def _pn_book_load():
    """Stalni imeniki za hitro štancanje (neodvisni od shranjenih nalogov)."""
    if PN_BOOK.exists():
        try:
            d = json.loads(PN_BOOK.read_text(encoding="utf-8"))
        except Exception:
            d = {}
    else:
        d = {}
    d.setdefault("zaposleni", [])   # [{oseba, dm, prebivalisce}]
    d.setdefault("vozila", [])      # ["KP FI 496", ...]
    d.setdefault("relacije", [])    # [{relacija, km}]
    return d


def _pn_book_save(book):
    tmp = PN_BOOK.with_suffix(".tmp")
    tmp.write_text(json.dumps(book, ensure_ascii=False, indent=2), encoding="utf-8")
    import os as _os
    _os.replace(str(tmp), str(PN_BOOK))


# Fiksni podatki podjetja
PN_COMPANY = "SUBAN d.o.o. Kazarje 3, 6230 Postojna ID za ddv: SI26391201"
PN_COMPANY_SHORT = "Suban d.o.o., Kazarje 3, 6230 Postojna"
PN_ODREDBA = "Irenej Suban, direktor"
PN_KM_RATE = 0.43                           # €/km (neobdavčeno 2026)

# Dnevnice SLO 2026 (Uredba o davčni obravnavi povračil — domače poti)
#   pod 6 ur → 0 ; 6–8 ur → 9,69 ; 8–12 ur → 13,88 ; nad 12 ur → 27,81
def _pn_dnevnica_za_ure(hours: float) -> float:
    """Vrne znesek dnevnice za podano trajanje v urah (en sam dan / ostanek)."""
    if hours <= 6:
        return 0.0
    if hours <= 8:
        return 9.69
    if hours <= 12:
        return 13.88
    return 27.81  # nad 12 do 24


def _pn_obracun_dnevnic(total_hours: float):
    """Večdnevni obračun: vsakih polnih 24 ur = polna dnevnica (27,81),
    za ostanek ur ustrezen razred. Vrne (st_dnevnic, znesek)."""
    if total_hours <= 0:
        return 0, 0.0
    full_days = int(total_hours // 24)
    rest = total_hours - full_days * 24
    amount = full_days * 27.81
    rest_amount = _pn_dnevnica_za_ure(rest)
    amount += rest_amount
    # število "dnevnic" za prikaz: polni dnevi + (1 če ostanek prinese kaj)
    count = full_days + (1 if rest_amount > 0 else 0)
    if count == 0 and amount > 0:
        count = 1
    return count, round(amount, 2)


def _pn_load():
    if PN_STATE.exists():
        try:
            return json.loads(PN_STATE.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []


def _pn_save(nalogi):
    tmp = PN_STATE.with_suffix(".tmp")
    tmp.write_text(json.dumps(nalogi, ensure_ascii=False, indent=2), encoding="utf-8")
    import os as _os
    _os.replace(str(tmp), str(PN_STATE))


def _pn_next_number(nalogi):
    """Naslednja zaporedna številka naloga."""
    mx = 0
    for n in nalogi:
        try:
            mx = max(mx, int(n.get("st", 0)))
        except (ValueError, TypeError):
            pass
    return mx + 1 if mx else 28  # nadaljuj od zadnje znane (izvirnik: 28–37)


def _pn_parse_hours(odhod: str, vrnitev: str, dni: int = 1):
    """Izračuna trajanje poti v urah iz ur odhoda/vrnitve.
    Za večdnevno (dni>1) prišteje (dni-1)*24 ur."""
    def _to_min(t):
        t = (t or "").strip()
        for fmt in ("%H:%M", "%H.%M", "%H:%M:%S"):
            try:
                from datetime import datetime as _d
                dt = _d.strptime(t, fmt)
                return dt.hour * 60 + dt.minute
            except Exception:
                continue
        return None
    o = _to_min(odhod); v = _to_min(vrnitev)
    if o is None or v is None:
        return None
    diff = v - o
    if diff < 0:
        diff += 24 * 60  # čez polnoč
    extra_days = max(0, int(dni) - 1)
    return diff / 60.0 + extra_days * 24.0


@app.get("/pn-list")
async def pn_list():
    """Vrne vse potne naloge (najnovejši prvi) + naslednjo številko + IMENIKE (stalni + izpeljani)."""
    try:
        nalogi = _pn_load()
        nalogi_sorted = sorted(nalogi, key=lambda n: int(n.get("st", 0)), reverse=True)
        book = _pn_book_load()

        # ZAPOSLENI: stalni imenik ima prednost; dopolni z izpeljanimi iz zgodovine
        zaposleni = {}
        for z in book.get("zaposleni", []):
            ime = (z.get("oseba") or "").strip()
            if ime:
                zaposleni[ime] = {"oseba": ime, "dm": z.get("dm", ""), "prebivalisce": z.get("prebivalisce", ""), "saved": True}
        for n in nalogi:
            ime = (n.get("oseba") or "").strip()
            if ime and ime not in zaposleni:
                zaposleni[ime] = {"oseba": ime, "dm": n.get("dm", ""), "prebivalisce": n.get("prebivalisce", ""), "saved": False}

        # VOZILA: stalni imenik + izpeljana
        vozila = set(v.strip() for v in book.get("vozila", []) if v and v.strip())
        for n in nalogi:
            if n.get("vozilo"):
                vozila.add(n["vozilo"].strip())

        # RELACIJE: stalni imenik (z znanimi km) + izpeljane iz zgodovine
        relacije = {}
        for r in book.get("relacije", []):
            rel = (r.get("relacija") or "").strip()
            if rel:
                relacije[rel] = {"relacija": rel, "km": r.get("km", 0), "saved": True}
        for n in nalogi:
            rel = (n.get("relacija") or "").strip()
            if rel and rel not in relacije:
                relacije[rel] = {"relacija": rel, "km": n.get("km", 0), "saved": False}

        return {
            "ok": True,
            "nalogi": nalogi_sorted,
            "next_st": _pn_next_number(nalogi),
            "zaposleni": list(zaposleni.values()),
            "vozila": sorted(vozila),
            "relacije": list(relacije.values()),
            "km_rate": PN_KM_RATE,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/pn-book")
async def pn_book_get():
    """Vrne samo stalne imenike (za urejanje)."""
    try:
        return {"ok": True, **_pn_book_load()}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/pn-book-save")
async def pn_book_save_ep(data: dict):
    """Doda/uredi/izbriše vnos v imeniku.
    body: { type: 'zaposleni'|'vozila'|'relacije', action: 'upsert'|'delete', item: {...} }"""
    try:
        typ = data.get("type")
        action = data.get("action", "upsert")
        item = data.get("item")
        if typ not in ("zaposleni", "vozila", "relacije"):
            return {"ok": False, "error": "Neveljaven tip imenika"}
        book = _pn_book_load()
        lst = book.get(typ, [])

        if typ == "vozila":
            val = (item or "").strip() if isinstance(item, str) else (item.get("vozilo", "").strip() if item else "")
            if action == "delete":
                lst = [v for v in lst if v.strip() != val]
            else:
                if val and val not in lst:
                    lst.append(val)
        elif typ == "zaposleni":
            ime = (item.get("oseba") or "").strip()
            if not ime:
                return {"ok": False, "error": "Manjka ime"}
            lst = [z for z in lst if (z.get("oseba") or "").strip() != ime]  # odstrani obstoječega
            if action != "delete":
                lst.append({"oseba": ime, "dm": (item.get("dm") or "").strip(), "prebivalisce": (item.get("prebivalisce") or "").strip()})
        elif typ == "relacije":
            rel = (item.get("relacija") or "").strip()
            if not rel:
                return {"ok": False, "error": "Manjka relacija"}
            lst = [r for r in lst if (r.get("relacija") or "").strip() != rel]
            if action != "delete":
                try:
                    km = float(item.get("km") or 0)
                except (ValueError, TypeError):
                    km = 0
                lst.append({"relacija": rel, "km": km})

        book[typ] = lst
        _pn_book_save(book)
        return {"ok": True, **book}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/pn-calc")
async def pn_calc(data: dict):
    """Samodejni izračun dnevnice + kilometrine. Frontend kliče ob spremembi vnosa.
    Vrne predlagane vrednosti (uporabnik jih lahko ročno popravi)."""
    try:
        odhod = data.get("odhod", "")
        vrnitev = data.get("vrnitev", "")
        dni = int(data.get("dni", 1) or 1)
        km = data.get("km")
        hours = _pn_parse_hours(odhod, vrnitev, dni)
        st_dnevnic, znesek_dnevnic = (0, 0.0)
        if hours is not None:
            st_dnevnic, znesek_dnevnic = _pn_obracun_dnevnic(hours)
        km_val = 0.0
        try:
            km_val = float(km) if km not in (None, "") else 0.0
        except (ValueError, TypeError):
            km_val = 0.0
        kilometrina = round(km_val * PN_KM_RATE, 2)
        skupaj = round(znesek_dnevnic + kilometrina, 2)
        return {
            "ok": True,
            "hours": round(hours, 2) if hours is not None else None,
            "st_dnevnic": st_dnevnic,
            "znesek_dnevnic": znesek_dnevnic,
            "km": km_val,
            "km_rate": PN_KM_RATE,
            "kilometrina": kilometrina,
            "skupaj": skupaj,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/pn-distance")
async def pn_distance(data: dict):
    """Google Maps Distance Matrix: km med kraji. Privzeto tja+nazaj (×2).
    Relacija je lahko niz krajev ločenih z '-' ali ','; sešteje segmente."""
    if not GOOGLE_MAPS_KEY:
        return {"ok": False, "error": "GOOGLE_MAPS_API_KEY ni nastavljen.", "manual": True}
    try:
        relacija = (data.get("relacija") or "").strip()
        origin = (data.get("origin") or "").strip()   # izhodišče (prebivališče)
        round_trip = data.get("round_trip", True)
        # razčleni kraje
        import re as _re
        kraji = [k.strip() for k in _re.split(r"[-,–]", relacija) if k.strip()]
        if not kraji:
            return {"ok": False, "error": "Ni krajev v relaciji.", "manual": True}
        # Pot se VEDNO začne pri izhodišču (prebivališču) — tako km ustrezajo izvirniku
        # (npr. Dobravlje → Ljubljana → Kranj → nazaj = 159 km).
        # Če prebivališče ni podano, uporabi sedež podjetja (Postojna) kot privzeto izhodišče.
        if not origin:
            origin = "Postojna"
        used_default_origin = (data.get("origin") or "").strip() == ""
        tocke = [origin] + kraji
        if len(tocke) < 2:
            return {"ok": False, "error": "Dodaj vsaj en kraj v relacijo.", "manual": True}

        total_m = 0
        segments = []
        async with httpx.AsyncClient(timeout=12.0) as hc:
            for i in range(len(tocke) - 1):
                a = tocke[i] + ", Slovenija"
                b = tocke[i + 1] + ", Slovenija"
                resp = await hc.get(
                    "https://maps.googleapis.com/maps/api/distancematrix/json",
                    params={"origins": a, "destinations": b,
                            "mode": "driving", "language": "sl", "key": GOOGLE_MAPS_KEY},
                )
                resp.raise_for_status()
                j = resp.json()
                try:
                    el = j["rows"][0]["elements"][0]
                    if el.get("status") != "OK":
                        return {"ok": False, "error": f"Maps: {el.get('status')} za {tocke[i]}→{tocke[i+1]}", "manual": True}
                    m = el["distance"]["value"]
                    total_m += m
                    segments.append({"from": tocke[i], "to": tocke[i + 1], "km": round(m / 1000, 1)})
                except (KeyError, IndexError):
                    return {"ok": False, "error": "Maps: neveljaven odgovor.", "manual": True}

        km_one = total_m / 1000.0
        km_total = km_one * (2 if round_trip else 1)
        return {
            "ok": True,
            "km_one_way": round(km_one, 1),
            "km": round(km_total),
            "round_trip": bool(round_trip),
            "origin": origin,
            "used_default_origin": used_default_origin,
            "segments": segments,
        }
    except httpx.HTTPError as e:
        return {"ok": False, "error": f"Maps napaka: {e}", "manual": True}
    except Exception as e:
        return {"ok": False, "error": str(e), "manual": True}


@app.post("/pn-save")
async def pn_save(data: dict):
    """Shrani (ali posodobi) potni nalog."""
    try:
        nalogi = _pn_load()
        st = data.get("st")
        if not st:
            st = _pn_next_number(nalogi)
        st = int(st)
        # izračunaj končne zneske (avtoriteta na strežniku, a spoštuj ročni override)
        hours = _pn_parse_hours(data.get("odhod", ""), data.get("vrnitev", ""), int(data.get("dni", 1) or 1))
        if data.get("znesek_dnevnic_manual") not in (None, ""):
            znesek_dnevnic = round(float(data["znesek_dnevnic_manual"]), 2)
            st_dnevnic = int(data.get("st_dnevnic", 1) or 1)
        else:
            st_dnevnic, znesek_dnevnic = _pn_obracun_dnevnic(hours or 0)
        try:
            km_val = float(data.get("km") or 0)
        except (ValueError, TypeError):
            km_val = 0.0
        kilometrina = round(km_val * PN_KM_RATE, 2)
        skupaj = round(znesek_dnevnic + kilometrina, 2)

        nalog = {
            "st": st,
            "datum": data.get("datum", ""),
            "oseba": (data.get("oseba") or "").strip(),
            "dm": (data.get("dm") or "").strip(),
            "prebivalisce": (data.get("prebivalisce") or "").strip(),
            "naloga": (data.get("naloga") or "").strip(),
            "vozilo": (data.get("vozilo") or "").strip(),
            "relacija": (data.get("relacija") or "").strip(),
            "origin": (data.get("origin") or "").strip(),
            "odhod": (data.get("odhod") or "").strip(),
            "vrnitev": (data.get("vrnitev") or "").strip(),
            "dni": int(data.get("dni", 1) or 1),
            "hours": round(hours, 2) if hours is not None else None,
            "st_dnevnic": st_dnevnic,
            "znesek_dnevnic": znesek_dnevnic,
            "km": km_val,
            "kilometrina": kilometrina,
            "skupaj": skupaj,
            "updated_at": _dt.now().isoformat(),
        }
        # posodobi obstoječega ali dodaj
        idx = next((i for i, n in enumerate(nalogi) if int(n.get("st", -1)) == st), None)
        if idx is not None:
            nalogi[idx] = nalog
        else:
            nalogi.append(nalog)
        _pn_save(nalogi)

        # samodejno dopolni stalni imenik (oseba, vozilo, relacija) — da se gradi sproti
        try:
            book = _pn_book_load()
            ime = nalog["oseba"]
            if ime:
                zlist = [z for z in book["zaposleni"] if (z.get("oseba") or "").strip() != ime]
                zlist.append({"oseba": ime, "dm": nalog["dm"], "prebivalisce": nalog["prebivalisce"]})
                book["zaposleni"] = zlist
            if nalog["vozilo"] and nalog["vozilo"] not in book["vozila"]:
                book["vozila"].append(nalog["vozilo"])
            rel = nalog["relacija"]
            if rel:
                rlist = [r for r in book["relacije"] if (r.get("relacija") or "").strip() != rel]
                rlist.append({"relacija": rel, "km": nalog["km"]})
                book["relacije"] = rlist
            _pn_book_save(book)
        except Exception:
            pass

        return {"ok": True, "nalog": nalog, "next_st": _pn_next_number(nalogi)}
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "tb": traceback.format_exc()}


@app.post("/pn-delete")
async def pn_delete(data: dict):
    try:
        st = int(data.get("st"))
        nalogi = _pn_load()
        nalogi = [n for n in nalogi if int(n.get("st", -1)) != st]
        _pn_save(nalogi)
        return {"ok": True, "next_st": _pn_next_number(nalogi)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _pn_find(st):
    for n in _pn_load():
        if int(n.get("st", -1)) == int(st):
            return n
    return None


def _pn_fmt_date(s):
    """ISO ali dd.mm.yyyy → dd.mm.yyyy"""
    if not s:
        return ""
    s = str(s)
    try:
        from datetime import datetime as _d
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%dT%H:%M:%S"):
            try:
                return _d.strptime(s[:19] if "T" in s else s, fmt).strftime("%d.%m.%Y")
            except Exception:
                continue
    except Exception:
        pass
    return s


@app.get("/pn-export-xlsx")
async def pn_export_xlsx(st: str = "", mesec: str = ""):
    """Izvozi nalog(e) v XLSX v formatu izvirnika (en list = en nalog).
    st=N za en nalog, mesec=YYYY-MM za vse naloge meseca."""
    try:
        nalogi = _pn_load()
        if st:
            sel = [n for n in nalogi if int(n.get("st", -1)) == int(st)]
            fname = f"Potni_nalog_{st}.xlsx"
        elif mesec:
            sel = [n for n in nalogi if (n.get("datum", "")[:7] == mesec)]
            fname = f"Potni_nalogi_{mesec}.xlsx"
        else:
            sel = nalogi
            fname = "Potni_nalogi.xlsx"
        if not sel:
            return JSONResponse({"ok": False, "error": "Ni nalogov za izvoz."}, status_code=404)
        sel = sorted(sel, key=lambda n: int(n.get("st", 0)))

        import io as _io
        # uporabi predlogo če obstaja, sicer zgradi list iz nič
        if PN_TEMPLATE.exists():
            wb = openpyxl.load_workbook(str(PN_TEMPLATE))
            tmpl_ws = wb[wb.sheetnames[0]]
            for i, n in enumerate(sel):
                ws = tmpl_ws if i == 0 else wb.copy_worksheet(tmpl_ws)
                ws.title = str(n.get("st", i + 1))
                _pn_fill_sheet(ws, n)
            # če predloga doda prazen prvi list, ga ohranimo kot prvi nalog (že napolnjen)
        else:
            wb = openpyxl.Workbook()
            first = True
            for n in sel:
                ws = wb.active if first else wb.create_sheet()
                ws.title = str(n.get("st"))
                _pn_build_sheet(ws, n)
                first = False

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except Exception as e:
        import traceback
        return JSONResponse({"ok": False, "error": str(e), "tb": traceback.format_exc()}, status_code=500)


def _pn_fill_sheet(ws, n):
    """Napolni list po koordinatah izvirne predloge (glej analizo strukture)."""
    od = _pn_fmt_date(n.get("datum"))
    ws["D3"] = n.get("st")
    ws["H3"] = od
    ws["C5"] = n.get("oseba", "")
    ws["C7"] = n.get("dm", "")
    ws["C9"] = n.get("prebivalisce", "")
    ws["C11"] = od
    ws["G11"] = n.get("odhod", "")
    ws["C13"] = PN_ODREDBA
    ws["C15"] = n.get("naloga", "")
    ws["C17"] = od
    ws["G17"] = n.get("dni", 1)
    ws["C19"] = n.get("vozilo", "")
    ws["G19"] = n.get("relacija", "")
    ws["C21"] = PN_COMPANY_SHORT
    ws["G31"] = "Irenej Suban"
    ws["F33"] = n.get("prebivalisce", "")
    ws["B36"] = od
    ws["C36"] = f"{n.get('odhod','')} uri"
    ws["B38"] = _pn_fmt_date(n.get("datum"))
    ws["C38"] = f"{n.get('vrnitev','')} uri"
    ws["E38"] = n.get("dni", 1)
    ws["G38"] = n.get("st_dnevnic", 1)
    ws["H38"] = n.get("znesek_dnevnic", 0)
    ws["H41"] = n.get("znesek_dnevnic", 0)
    ws["C43"] = n.get("km", 0)
    ws["E43"] = PN_KM_RATE
    ws["H43"] = n.get("kilometrina", 0)
    ws["H47"] = n.get("skupaj", 0)
    ws["H49"] = 0
    ws["H51"] = n.get("skupaj", 0)


def _pn_build_sheet(ws, n):
    """Fallback: zgradi list iz nič (če predloge ni)."""
    ws["A1"] = PN_COMPANY
    ws["A3"] = "Nalog za službeno potovanje"
    _pn_fill_sheet(ws, n)


@app.get("/pn-export-pdf")
async def pn_export_pdf(st: str = ""):
    """Izvozi en potni nalog v PDF (A4, postavitev po izvirniku)."""
    try:
        n = _pn_find(st) if st else None
        if not n:
            return JSONResponse({"ok": False, "error": "Nalog ne obstaja."}, status_code=404)
        import io as _io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as _canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        buf = _io.BytesIO()
        c = _canvas.Canvas(buf, pagesize=A4)
        W, H = A4
        # poskusi naložiti font s šumniki (DejaVu); fallback Helvetica
        font_name = "Helvetica"
        font_bold = "Helvetica-Bold"
        for fp in ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                   "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"):
            try:
                if "Bold" in fp:
                    pdfmetrics.registerFont(TTFont("DejaVuB", fp)); font_bold = "DejaVuB"
                else:
                    pdfmetrics.registerFont(TTFont("DejaVu", fp)); font_name = "DejaVu"
            except Exception:
                pass

        y = H - 20 * mm
        def line(txt, dy=7*mm, font=None, size=10, x=20*mm):
            nonlocal y
            c.setFont(font or font_name, size)
            c.drawString(x, y, txt)
            y -= dy

        od = _pn_fmt_date(n.get("datum"))
        c.setFont(font_bold, 9); c.drawString(20*mm, y, PN_COMPANY); y -= 10*mm
        c.setFont(font_bold, 13); c.drawCentredString(W/2, y, "Nalog za službeno potovanje"); 
        c.setFont(font_name, 10); c.drawRightString(W-20*mm, y, f"Št: {n.get('st')}   Datum: {od}"); y -= 12*mm

        rows = [
            ("Odrejam, da odpotuje:", n.get("oseba", "")),
            ("Na delovnem mestu:", n.get("dm", "")),
            ("Prebivališče:", n.get("prebivalisce", "")),
            ("Dne / ob uri:", f"{od}  ob {n.get('odhod','')}"),
            ("Po nalogu:", PN_ODREDBA),
            ("Z nalogo:", n.get("naloga", "")),
            ("Trajanje (dni):", str(n.get("dni", 1))),
            ("Prevozno sredstvo:", n.get("vozilo", "")),
            ("Potuje v kraj (relacija):", n.get("relacija", "")),
            ("Potne stroške plača:", PN_COMPANY_SHORT),
        ]
        for lbl, val in rows:
            c.setFont(font_bold, 10); c.drawString(20*mm, y, lbl)
            c.setFont(font_name, 10); c.drawString(75*mm, y, str(val)); y -= 8*mm

        y -= 4*mm
        c.setFont(font_bold, 11); c.drawString(20*mm, y, "RAČUN potnih stroškov"); y -= 9*mm
        calc = [
            ("Odsotnost:", f"{od} {n.get('odhod','')} → {n.get('vrnitev','')}  ({n.get('dni',1)} dni)"),
            (f"Dnevnice ({n.get('st_dnevnic',0)}×):", f"{n.get('znesek_dnevnic',0):.2f} EUR"),
            (f"Kilometrina ({n.get('km',0)} km × {PN_KM_RATE} EUR):", f"{n.get('kilometrina',0):.2f} EUR"),
        ]
        for lbl, val in calc:
            c.setFont(font_name, 10); c.drawString(20*mm, y, lbl)
            c.drawRightString(W-20*mm, y, val); y -= 8*mm
        y -= 2*mm
        c.setLineWidth(0.5); c.line(20*mm, y, W-20*mm, y); y -= 8*mm
        c.setFont(font_bold, 12); c.drawString(20*mm, y, "Skupaj za izplačilo:")
        c.drawRightString(W-20*mm, y, f"{n.get('skupaj',0):.2f} EUR"); y -= 20*mm

        c.setFont(font_name, 9)
        c.drawString(25*mm, y, "_______________________")
        c.drawString(120*mm, y, "_______________________"); y -= 5*mm
        c.drawString(30*mm, y, "(podpis prejemnika)")
        c.drawString(125*mm, y, "(podpis odredbodajalca)")

        c.showPage(); c.save()
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="Potni_nalog_{n.get("st")}.pdf"'},
        )
    except Exception as e:
        import traceback
        return JSONResponse({"ok": False, "error": str(e), "tb": traceback.format_exc()}, status_code=500)



# ════════════════════════════════════════════════════════════════════
#  SPREMEMBA POZICIJ (Zaloga → tab "Sprememba pozicij")
#  Prepozna SKU iz slike etikete (Claude vision) ALI ročni vnos z
#  autosuggest. SKU se VEDNO ujame proti bazi zaloge (sku/title) —
#  brez ujemanja = opozorilo. Druga slika = pozicija. Izpis SKU:Pozicija.
# ════════════════════════════════════════════════════════════════════
import unicodedata as _ud


def _poz_norm(s: str) -> str:
    """Normaliziraj za primerjavo: velike črke, brez šumnikov/ločil/presledkov."""
    if not s:
        return ""
    s = _ud.normalize("NFKD", str(s))
    s = "".join(c for c in s if not _ud.combining(c))
    return "".join(c for c in s.upper() if c.isalnum())


def _poz_load_stock():
    """Naloži seznam zaloge (sku, title, position) iz shranjenega CSV."""
    if not STOCK_CSV_FILE.exists():
        return []
    try:
        import csv as _csv
        from io import StringIO as _SIO
        text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        first_line = text.split("\n", 1)[0]
        sep = ";" if first_line.count(";") > first_line.count(",") else ","
        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        out = []
        for row in reader:
            sku = (row.get("product_sku") or row.get("sku") or "").strip()
            if not sku:
                continue
            out.append({
                "sku": sku,
                "title": (row.get("title") or "").strip(),
                "position": (row.get("position") or "").strip(),
                "product_id": (row.get("product_id") or "").strip(),
            })
        return out
    except Exception:
        return []


def _poz_match_sku(candidates, stock):
    """Ujemi kandidate (besede iz slike/vnosa) proti bazi.
    Vrne (match_dict | None, alternatives_list). Ujemanje: točno > prefiks > vsebuje."""
    cand_norm = [_poz_norm(c) for c in candidates if c and len(_poz_norm(c)) >= 2]
    if not cand_norm:
        return None, []
    exact, prefix, contains = [], [], []
    for it in stock:
        sku_n = _poz_norm(it["sku"])
        if not sku_n:
            continue
        for cn in cand_norm:
            if sku_n == cn:
                exact.append(it); break
            elif sku_n.startswith(cn) or cn.startswith(sku_n):
                prefix.append(it); break
            elif cn in sku_n or sku_n in cn:
                contains.append(it); break
    # dedup ohrani vrstni red
    seen = set(); ranked = []
    for it in exact + prefix + contains:
        if it["sku"] not in seen:
            seen.add(it["sku"]); ranked.append(it)
    if not ranked:
        return None, []
    return ranked[0], ranked[1:6]


@app.get("/pozicije-suggest")
async def pozicije_suggest(q: str = ""):
    """Autosuggest SKU iz baze (Elastic-style): prefiks SKU > vsebuje SKU > vsebuje naziv."""
    q = (q or "").strip()
    if len(q) < 1:
        return {"ok": True, "suggestions": []}
    stock = _poz_load_stock()
    qn = _poz_norm(q)
    ql = q.lower()
    pref, contains_sku, in_title = [], [], []
    for it in stock:
        skn = _poz_norm(it["sku"])
        if skn.startswith(qn):
            pref.append(it)
        elif qn in skn:
            contains_sku.append(it)
        elif ql in it["title"].lower():
            in_title.append(it)
    seen = set(); out = []
    for it in pref + contains_sku + in_title:
        if it["sku"] in seen:
            continue
        seen.add(it["sku"])
        out.append({"sku": it["sku"], "title": it["title"], "position": it["position"]})
        if len(out) >= 12:
            break
    return {"ok": True, "suggestions": out}


@app.post("/pozicije-recognize")
async def pozicije_recognize(data: dict):
    """Prepozna SKU iz slike etikete. Claude PREBERE vse berljive besede/kode,
    server jih UJAME proti bazi (sku/title). Vrne pravi SKU ali opozorilo."""
    image_b64 = data.get("image")
    media_type = data.get("media_type", "image/jpeg")
    if not image_b64:
        return {"ok": False, "error": "Ni slike."}
    stock = _poz_load_stock()
    if not stock:
        return {"ok": False, "error": "Baza zaloge ni naložena (naloži CSV v zavihku Zaloga)."}

    loop = asyncio.get_event_loop()
    _prompt = (
        "This is a warehouse product/carton label or a shelf-position note. "
        "Read ALL text you can see: product names, codes, SKUs, brand words. "
        "Return ONLY a JSON array of the distinct readable text tokens, most-prominent first. "
        "Include both full strings and meaningful sub-parts (e.g. a product name like 'SPINSTORE' AND any code). "
        "Do NOT guess or invent — only what is actually visible. "
        "Example: [\"SPINSTORE\", \"S-R4813-48-25-10000\", \"Home & Marker\"]"
    )

    def _call(model):
        return client.messages.create(
            model=model,
            max_tokens=400,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_b64}},
                {"type": "text", "text": _prompt},
            ]}],
        )

    msg = None
    for attempt in range(3):
        model = "claude-haiku-4-5-20251001" if attempt == 0 else "claude-sonnet-4-6"
        try:
            msg = await loop.run_in_executor(None, lambda m=model: _call(m))
            break
        except Exception as e:
            is_529 = "529" in str(e) or "overloaded" in str(e).lower()
            if attempt < 2:
                await asyncio.sleep(2)
            else:
                return {"ok": False, "error": f"Anthropic napaka: {e}"}
    if msg is None:
        return {"ok": False, "error": "Anthropic preobremenjen, poskusi znova."}

    raw = ""
    for block in msg.content:
        if getattr(block, "type", "") == "text":
            raw += block.text
    # razčleni JSON array iz odgovora
    candidates = []
    try:
        import re as _re
        m = _re.search(r"\[.*\]", raw, _re.S)
        if m:
            candidates = json.loads(m.group(0))
    except Exception:
        candidates = [w.strip() for w in raw.replace("\n", ",").split(",") if w.strip()]
    candidates = [str(c).strip() for c in candidates if str(c).strip()]

    match, alts = _poz_match_sku(candidates, stock)
    if match:
        return {
            "ok": True, "found": True,
            "sku": match["sku"], "title": match["title"],
            "current_position": match["position"],
            "read_tokens": candidates,
            "alternatives": alts,
        }
    return {
        "ok": True, "found": False,
        "read_tokens": candidates,
        "message": "SKU ni najden v bazi zaloge. Prebrano besedilo: " + ", ".join(candidates[:6]),
    }


@app.post("/pozicije-recognize-pos")
async def pozicije_recognize_pos(data: dict):
    """Prepozna POZICIJO iz slike (npr. ročno napisan 'P2-B'). Claude prebere,
    server ujame proti seznamu veljavnih pozicij in PREDLAGA (uporabnik potrdi)."""
    image_b64 = data.get("image")
    media_type = data.get("media_type", "image/jpeg")
    if not image_b64:
        return {"ok": False, "error": "Ni slike."}
    loop = asyncio.get_event_loop()
    _prompt = (
        "This image shows a warehouse shelf/rack position label, often HANDWRITTEN. "
        "Read the position code exactly as written (e.g. 'P2-B', 'P8-C', '02-3E'). "
        "Return ONLY a JSON array of the most likely readings, best first. "
        "If handwriting is ambiguous, include alternative interpretations. "
        "Do NOT invent. Example: [\"P2-B\", \"P22-B\"]"
    )

    def _call(model):
        return client.messages.create(
            model=model, max_tokens=200,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_b64}},
                {"type": "text", "text": _prompt},
            ]}],
        )

    msg = None
    for attempt in range(3):
        model = "claude-haiku-4-5-20251001" if attempt == 0 else "claude-sonnet-4-6"
        try:
            msg = await loop.run_in_executor(None, lambda m=model: _call(m))
            break
        except Exception as e:
            if attempt < 2:
                await asyncio.sleep(2)
            else:
                return {"ok": False, "error": f"Anthropic napaka: {e}"}
    if msg is None:
        return {"ok": False, "error": "Anthropic preobremenjen."}

    raw = ""
    for block in msg.content:
        if getattr(block, "type", "") == "text":
            raw += block.text
    readings = []
    try:
        import re as _re
        m = _re.search(r"\[.*\]", raw, _re.S)
        if m:
            readings = json.loads(m.group(0))
    except Exception:
        readings = [w.strip() for w in raw.replace("\n", ",").split(",") if w.strip()]
    readings = [str(r).strip() for r in readings if str(r).strip()]
    if not readings:
        return {"ok": True, "found": False, "read_tokens": [], "message": "Pozicije ni bilo mogoče prebrati."}

    # ujemi najboljše branje proti seznamu veljavnih
    primary = readings[0]
    result = _poz_match_position(primary)
    # če primarno ne najde, poskusi še druga branja za točno ujemanje
    if result["status"] in ("none", "suggest") and len(readings) > 1:
        for alt in readings[1:]:
            r2 = _poz_match_position(alt)
            if r2["status"] == "exact":
                result = r2
                primary = alt
                break
    return {
        "ok": True,
        "read_tokens": readings,
        "reading": primary,
        "status": result["status"],          # exact | suggest | none | no_list
        "value": result["value"],
        "suggestions": result["suggestions"],
    }


PN_POZ_PENDING = DATA_DIR / "pozicije_pending.json"
PN_POZ_VALID = DATA_DIR / "pozicije_valid.json"   # seznam veljavnih pozicij


def _poz_valid_load():
    if PN_POZ_VALID.exists():
        try:
            return json.loads(PN_POZ_VALID.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []


def _poz_levenshtein(a: str, b: str) -> int:
    """Razdalja med nizoma (število urejanj)."""
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def _poz_family(s: str) -> str:
    """Družina pozicije za boljše ujemanje: 'P' (P-regali), 'NUM' (številčni regali),
    'WORD' (Paleta, Omara, ...). Branje 'P22-B' tako ostane v P-družini → P2-B, ne 02-2B."""
    n = _poz_norm(s)
    if not n:
        return "WORD"
    if n[0] == "P" and len(n) > 1 and n[1].isdigit():
        return "P"
    if n[0].isdigit():
        return "NUM"
    return "WORD"


def _poz_match_position(reading: str):
    """Ujemi prebrano pozicijo proti seznamu veljavnih.
    Vrne dict: {status: 'exact'|'suggest'|'none', value, suggestions[]}.
    Razvrstitev: najprej razdalja, ob enaki razdalji ima prednost ista družina (P/številčna)."""
    valid = _poz_valid_load()
    rn = _poz_norm(reading)
    if not valid:
        # ni seznama → sprejmi kot je (brez validacije)
        return {"status": "no_list", "value": reading, "suggestions": []}
    # točno ujemanje (normalizirano)
    for v in valid:
        if _poz_norm(v) == rn:
            return {"status": "exact", "value": v, "suggestions": []}
    # fuzzy: razvrsti po (razdalja, družinski penal) — ista družina ima prednost
    rfam = _poz_family(reading)
    def _key(v):
        d = _poz_levenshtein(rn, _poz_norm(v))
        fam_penalty = 0 if _poz_family(v) == rfam else 1
        return (d, fam_penalty)
    scored = sorted(valid, key=_key)
    best = scored[:5]
    nearest_dist = _poz_levenshtein(rn, _poz_norm(best[0])) if best else 99
    if best and nearest_dist <= 3:
        return {"status": "suggest", "value": reading, "suggestions": best}
    return {"status": "none", "value": reading, "suggestions": best[:3]}


@app.get("/pozicije-valid")
async def pozicije_valid_get():
    """Vrne seznam veljavnih pozicij."""
    return {"ok": True, "positions": _poz_valid_load()}


@app.post("/pozicije-valid-save")
async def pozicije_valid_save(data: dict):
    """Nastavi/dopolni seznam veljavnih pozicij.
    body: { raw: '...', mode: 'replace'|'add' }"""
    try:
        import re as _re
        raw = data.get("raw", "")
        mode = data.get("mode", "replace")
        parsed = [p.strip() for p in _re.split(r"[\s,;\n\r\t]+", str(raw)) if p.strip()]
        # dedup ohrani vrstni red
        cur = [] if mode == "replace" else _poz_valid_load()
        seen = set(_poz_norm(x) for x in cur)
        for p in parsed:
            if _poz_norm(p) not in seen:
                cur.append(p)
                seen.add(_poz_norm(p))
        tmp = PN_POZ_VALID.with_suffix(".tmp")
        tmp.write_text(json.dumps(cur, ensure_ascii=False, indent=2), encoding="utf-8")
        import os as _os
        _os.replace(str(tmp), str(PN_POZ_VALID))
        return {"ok": True, "positions": cur, "count": len(cur)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/pozicije-pos-suggest")
async def pozicije_pos_suggest(q: str = ""):
    """Autosuggest veljavnih pozicij (za ročni vnos)."""
    q = (q or "").strip()
    valid = _poz_valid_load()
    if not valid:
        return {"ok": True, "suggestions": []}
    if not q:
        return {"ok": True, "suggestions": valid[:12]}
    qn = _poz_norm(q)
    pref, contains, fuzzy = [], [], []
    for v in valid:
        vn = _poz_norm(v)
        if vn.startswith(qn):
            pref.append(v)
        elif qn in vn:
            contains.append(v)
        elif _poz_levenshtein(qn, vn) <= 2:
            fuzzy.append(v)
    out = []
    for v in pref + contains + fuzzy:
        if v not in out:
            out.append(v)
        if len(out) >= 12:
            break
    return {"ok": True, "suggestions": out}


@app.get("/pozicije-pending")
async def pozicije_pending_get():
    """Vrne trenutni seznam pripravljenih sprememb (SKU:Pozicija) za to sejo."""
    try:
        if PN_POZ_PENDING.exists():
            return {"ok": True, "items": json.loads(PN_POZ_PENDING.read_text(encoding="utf-8"))}
        return {"ok": True, "items": []}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/pozicije-pending-save")
async def pozicije_pending_save(data: dict):
    """Doda/posodobi/izbriše vrstico v pripravljenem seznamu.
    action: 'add' | 'remove' | 'clear'. item: {sku, position, title}."""
    try:
        items = []
        if PN_POZ_PENDING.exists():
            items = json.loads(PN_POZ_PENDING.read_text(encoding="utf-8"))
        action = data.get("action", "add")
        if action == "clear":
            items = []
        elif action == "remove":
            sku = (data.get("sku") or "").strip()
            items = [x for x in items if x.get("sku") != sku]
        else:
            it = data.get("item") or {}
            sku = (it.get("sku") or "").strip()
            pos = (it.get("position") or "").strip()
            if not sku:
                return {"ok": False, "error": "Manjka SKU"}
            # preveri da SKU obstaja v bazi
            stock = _poz_load_stock()
            exists = any(s["sku"] == sku for s in stock)
            if not exists:
                return {"ok": False, "error": f"SKU '{sku}' ni v bazi zaloge"}
            items = [x for x in items if x.get("sku") != sku]  # zamenjaj obstoječega
            items.append({"sku": sku, "position": pos, "title": (it.get("title") or "").strip()})
        tmp = PN_POZ_PENDING.with_suffix(".tmp")
        tmp.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
        import os as _os
        _os.replace(str(tmp), str(PN_POZ_PENDING))
        return {"ok": True, "items": items}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/pozicije-apply")
async def pozicije_apply():
    """Zapiše pripravljene pozicije nazaj v CSV zaloge (stolpec position).
    Vrne število posodobljenih vrstic."""
    try:
        if not PN_POZ_PENDING.exists():
            return {"ok": False, "error": "Ni pripravljenih sprememb."}
        pending = json.loads(PN_POZ_PENDING.read_text(encoding="utf-8"))
        if not pending:
            return {"ok": False, "error": "Seznam je prazen."}
        if not STOCK_CSV_FILE.exists():
            return {"ok": False, "error": "Baza zaloge ni naložena."}

        import csv as _csv
        from io import StringIO as _SIO
        text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        first_line = text.split("\n", 1)[0]
        sep = ";" if first_line.count(";") > first_line.count(",") else ","
        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        fieldnames = reader.fieldnames or []
        if "position" not in fieldnames:
            fieldnames = fieldnames + ["position"]
        rows = list(reader)

        pos_map = {p["sku"]: p.get("position", "") for p in pending}
        sku_key = "product_sku" if "product_sku" in (fieldnames or []) else ("sku" if "sku" in (fieldnames or []) else None)
        updated = 0
        for row in rows:
            rsku = (row.get("product_sku") or row.get("sku") or "").strip()
            if rsku in pos_map:
                row["position"] = pos_map[rsku]
                updated += 1

        out = _SIO()
        writer = _csv.DictWriter(out, fieldnames=fieldnames, delimiter=sep, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        tmp = STOCK_CSV_FILE.with_suffix(".tmp")
        tmp.write_text(out.getvalue(), encoding="utf-8-sig")
        import os as _os
        _os.replace(str(tmp), str(STOCK_CSV_FILE))

        # počisti pending
        PN_POZ_PENDING.write_text("[]", encoding="utf-8")
        return {"ok": True, "updated": updated, "total_pending": len(pending)}
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "tb": traceback.format_exc()}


@app.post("/siluxar-push-positions")
async def siluxar_push_positions(data: dict):
    """Pošlje pozicije NAZAJ v siluxar (beta zapisovalni endpoint).
    Body: {"items": [{"sku": "1234", "position": "00-4B"}, ...]}
    Pošlje POST na www.siluxar.si/apistockexport z istim Authorization ključem."""
    try:
        import csv as _csv
        from io import StringIO as _SIO
        items = data.get("items") or []
        # podpri tudi enojni {sku, position?, stock?}
        # POMEMBNO: position vključi LE, če ga je frontend dejansko poslal.
        # Sicer siluxar prazni position interpretira kot reset in povozi position+stock z null.
        if not items and data.get("sku"):
            _single = {"sku": data.get("sku")}
            if "position" in data:
                _single["position"] = data.get("position")
            if "stock" in data:
                _single["stock"] = data.get("stock")
            if "warehouse" in data:
                _single["warehouse"] = data.get("warehouse")
            items = [_single]
        # mapping (SKU, warehouse) -> siluxar id iz lokalnega stock CSV.
        # Marko: apistockexport matcha po polju `id` (=naš siluxar_id), NE po product_id.
        # POMEMBNO: isti SKU ima lahko VEČ zapisov (silux + silux2), vsak svoj id.
        # Zato keyiramo po (sku, warehouse), da pošljemo id PRAVEGA skladišča.
        _id_by_sku_wh = {}
        _id_by_sku_first = {}
        try:
            if STOCK_CSV_FILE.exists():
                for _row in _csv.DictReader(_SIO(STOCK_CSV_FILE.read_text(encoding="utf-8"))):
                    _s = (_row.get("product_sku") or "").strip()
                    _sid = (_row.get("siluxar_id") or "").strip()
                    _wh = (_row.get("warehouse") or "").strip().lower()
                    if _s and _sid and _sid not in ("0", "0.0"):
                        _id_by_sku_wh.setdefault((_s, _wh), _sid)
                        _id_by_sku_first.setdefault(_s, _sid)
        except Exception:
            _id_by_sku_wh = {}; _id_by_sku_first = {}

        # očisti: sku + position (+ stock če podan), brez praznih
        payload = []
        for it in items:
            sku = (str(it.get("sku") or "")).strip()
            if not sku:
                continue
            entry = {"sku": sku}
            # position: vključi LE če je podan IN neprazen.
            # Prazen position siluxar tretira kot reset → povozi position+stock z null.
            if "position" in it:
                _pos = (str(it.get("position") or "")).strip()
                if _pos:
                    entry["position"] = _pos
            # stock: vključi SAMO če je podan (ne pošiljaj praznega/None).
            # Pošlji kot STRING — isto kot apistockexport izvaža/bere (Marko: konsistentno).
            stock_raw = it.get("stock")
            _has_stock = stock_raw is not None and str(stock_raw).strip() != ""
            if _has_stock:
                try:
                    entry["stock"] = str(int(float(str(stock_raw).replace(",", "."))))
                except Exception:
                    entry["stock"] = str(stock_raw).strip()
                # id = siluxar id (naš siluxar_id): match ključ za apistockexport (Marko).
                # SAMO pri stock pošiljanju. Pozicijski push ostane {sku, position} kot doslej.
                # Isti SKU ima lahko več skladišč (silux/silux2) z RAZLIČNIM id — zato po warehouse.
                _wh_it = (str(it.get("warehouse") or "")).strip().lower()
                _eid = (str(it.get("id") or "")).strip()
                if not _eid and _wh_it:
                    _eid = _id_by_sku_wh.get((sku, _wh_it), "")
                if not _eid:
                    _eid = _id_by_sku_first.get(sku, "")
                if _eid:
                    entry["id"] = _eid
            payload.append(entry)
        if not payload:
            return {"ok": False, "error": "Ni veljavnih postavk za pošiljanje."}

        key = os.environ.get("SILUXAR_STOCK_KEY", "")
        basic_user = os.environ.get("SILUXAR_BASIC_USER", "")
        basic_pass = os.environ.get("SILUXAR_BASIC_PASS", "")
        headers = {"Content-Type": "application/json"}
        _auth = None
        if key:
            headers["Authorization"] = key
        elif basic_user or basic_pass:
            _auth = httpx.BasicAuth(basic_user, basic_pass)

        # PRODUKCIJSKI zapisovalni endpoint
        url = "https://www.siluxar.si/apistockexport"

        def _zabelezi(status, ok, resp_text, err=None, exc=None):
            """Zabeleži pošiljanje v log (zadnjih 50) za debug."""
            try:
                log = []
                if SILUXAR_PUSH_LOG.exists():
                    log = json.loads(SILUXAR_PUSH_LOG.read_text(encoding="utf-8"))
                    if not isinstance(log, list):
                        log = []
                log.insert(0, {
                    "cas": _lj_now().strftime("%Y-%m-%d %H:%M:%S"),
                    "ts_utc": datetime.now(timezone.utc).isoformat(),
                    "poslano_stevilo": len(payload),
                    "primeri_poslanih": payload[:10],
                    "status": status,
                    "ok": ok,
                    "odgovor": (resp_text or "")[:500],
                    "napaka": err,
                    "exception": exc,
                    "url": url,
                })
                log = log[:50]
                SILUXAR_PUSH_LOG.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception:
                pass

        try:
            async with httpx.AsyncClient(timeout=60, auth=_auth) as cli:
                r = await cli.post(url, headers=headers, json=payload)
        except Exception as e:
            _zabelezi(None, False, None, err=f"Napaka pri klicu siluxar.si: {e}", exc=str(e))
            return {"ok": False, "error": f"Napaka pri klicu siluxar.si: {e}", "poslano": len(payload)}

        ok = 200 <= r.status_code < 300
        resp_text = (r.text or "")[:500]
        _zabelezi(r.status_code, ok, resp_text, err=(None if ok else f"siluxar vrnil status {r.status_code}"))
        return {
            "ok": ok,
            "status": r.status_code,
            "poslano": len(payload),
            "primeri_poslanih": payload[:5],
            "odgovor": resp_text,
            "error": None if ok else f"siluxar vrnil status {r.status_code}",
        }
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "tb": traceback.format_exc()}


@app.post("/siluxar-delete-alerts")
async def siluxar_delete_alerts(data: dict):
    """Pošlje array ps.id-jev za BRISANJE alarmov v siluxar (PRODUKCIJA).
    Body: {"ids": [12, 13, 14]} → pošlje [12,13,14] na apistockalertsexport (POST)."""
    try:
        ids_raw = data.get("ids") or []
        # očisti: samo ne-prazni, pretvori v int kjer gre
        ids = []
        for x in ids_raw:
            s = str(x).strip()
            if not s:
                continue
            try:
                ids.append(int(s))
            except Exception:
                ids.append(s)   # če ni številka, pusti kot string
        if not ids:
            return {"ok": False, "error": "Ni veljavnih ID-jev za brisanje."}

        key = os.environ.get("SILUXAR_STOCK_KEY", "")
        basic_user = os.environ.get("SILUXAR_BASIC_USER", "")
        basic_pass = os.environ.get("SILUXAR_BASIC_PASS", "")
        headers = {"Content-Type": "application/json"}
        _auth = None
        if key:
            headers["Authorization"] = key
        elif basic_user or basic_pass:
            _auth = httpx.BasicAuth(basic_user, basic_pass)

        url = "https://www.siluxar.si/apistockalertsexport"

        def _zabelezi(status, ok, resp_text, err=None, exc=None):
            try:
                log = []
                if SILUXAR_DELETE_LOG.exists():
                    log = json.loads(SILUXAR_DELETE_LOG.read_text(encoding="utf-8"))
                    if not isinstance(log, list):
                        log = []
                log.insert(0, {
                    "cas": _lj_now().strftime("%Y-%m-%d %H:%M:%S"),
                    "ts_utc": datetime.now(timezone.utc).isoformat(),
                    "brisanih_stevilo": len(ids),
                    "id_ji": ids[:50],
                    "status": status,
                    "ok": ok,
                    "odgovor": (resp_text or "")[:500],
                    "napaka": err,
                    "exception": exc,
                    "url": url,
                })
                log = log[:50]
                SILUXAR_DELETE_LOG.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception:
                pass

        try:
            async with httpx.AsyncClient(timeout=60, auth=_auth) as cli:
                r = await cli.post(url, headers=headers, json=ids)
        except Exception as e:
            _zabelezi(None, False, None, err=f"Napaka pri klicu siluxar.si: {e}", exc=str(e))
            return {"ok": False, "error": f"Napaka pri klicu siluxar.si: {e}", "poslano": len(ids)}

        ok = 200 <= r.status_code < 300
        resp_text = (r.text or "")[:500]
        _zabelezi(r.status_code, ok, resp_text, err=(None if ok else f"siluxar vrnil status {r.status_code}"))
        return {
            "ok": ok,
            "status": r.status_code,
            "poslano": len(ids),
            "id_ji": ids[:10],
            "odgovor": resp_text,
            "error": None if ok else f"siluxar vrnil status {r.status_code}",
        }
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "tb": traceback.format_exc()}


@app.get("/siluxar-delete-debug")
async def siluxar_delete_debug():
    """Debug: zadnjih 50 brisanj alarmov (kaj poslano, status, odgovor)."""
    if not SILUXAR_DELETE_LOG.exists():
        return {"sporocilo": "Še ni nobenega brisanja.", "zapisi": []}
    try:
        log = json.loads(SILUXAR_DELETE_LOG.read_text(encoding="utf-8"))
        return {"skupaj_zapisov": len(log) if isinstance(log, list) else 0,
                "zadnje_brisanje": log[0] if log else None, "zapisi": log}
    except Exception as e:
        return {"error": str(e)}


@app.get("/siluxar-push-authtest")
async def siluxar_push_authtest():
    """Diagnostika: preizkusi VEČ kombinacij avtentikacije na beta zapisovalnem endpointu.
    Pokaže, katera kombinacija NE vrne 401 — da vemo, kako beta pričakuje poverilnice."""
    key = os.environ.get("SILUXAR_STOCK_KEY", "")
    basic_user = os.environ.get("SILUXAR_BASIC_USER", "")
    basic_pass = os.environ.get("SILUXAR_BASIC_PASS", "")
    url = "https://www.siluxar.si/apistockexport"
    test_payload = [{"sku": "_TEST_", "position": "_TEST_"}]

    # katere poverilnice sploh imamo (brez razkrivanja gesel)
    env_status = {
        "SILUXAR_STOCK_KEY": "✓ nastavljen" if key else "✗ MANJKA",
        "SILUXAR_BASIC_USER": ("✓ " + basic_user) if basic_user else "✗ MANJKA",
        "SILUXAR_BASIC_PASS": "✓ nastavljen" if basic_pass else "✗ MANJKA",
    }

    rezultati = []

    async def _poskus(opis, headers=None, auth=None):
        try:
            async with httpx.AsyncClient(timeout=40, auth=auth) as cli:
                r = await cli.post(url, headers=(headers or {"Content-Type": "application/json"}), json=test_payload)
            body = (r.text or "")[:200]
            je_401 = r.status_code == 401
            rezultati.append({
                "nacin": opis,
                "status": r.status_code,
                "uspeh_ne_401": not je_401,
                "odgovor_zacetek": body,
            })
        except Exception as e:
            rezultati.append({"nacin": opis, "status": "EXCEPTION", "napaka": str(e)})

    # 1) samo aplikacijski ključ v Authorization (kot zdaj — pričakovano 401)
    if key:
        await _poskus("A) samo ključ v Authorization", headers={"Content-Type": "application/json", "Authorization": key})

    # 2) samo Basic Auth (user:pass) v Authorization header
    if basic_user or basic_pass:
        await _poskus("B) samo Basic Auth (Authorization: Basic)", auth=httpx.BasicAuth(basic_user, basic_pass))

    # 3) Basic Auth v Authorization + ključ v X-Api-Key header
    if (basic_user or basic_pass) and key:
        await _poskus("C) Basic Auth + ključ v X-Api-Key",
                      headers={"Content-Type": "application/json", "X-Api-Key": key},
                      auth=httpx.BasicAuth(basic_user, basic_pass))

    # 4) Basic Auth v Authorization + ključ v X-Authorization header
    if (basic_user or basic_pass) and key:
        await _poskus("D) Basic Auth + ključ v X-Authorization",
                      headers={"Content-Type": "application/json", "X-Authorization": key},
                      auth=httpx.BasicAuth(basic_user, basic_pass))

    # 5) Basic Auth v Authorization + ključ kot query param ?key=
    if (basic_user or basic_pass) and key:
        try:
            async with httpx.AsyncClient(timeout=40, auth=httpx.BasicAuth(basic_user, basic_pass)) as cli:
                r = await cli.post(url + "?key=" + key, headers={"Content-Type": "application/json"}, json=test_payload)
            rezultati.append({"nacin": "E) Basic Auth + ključ v ?key= query",
                              "status": r.status_code, "uspeh_ne_401": r.status_code != 401,
                              "odgovor_zacetek": (r.text or "")[:200]})
        except Exception as e:
            rezultati.append({"nacin": "E) Basic Auth + ključ v ?key= query", "status": "EXCEPTION", "napaka": str(e)})

    # najboljši rezultat
    uspesni = [r for r in rezultati if r.get("uspeh_ne_401")]
    return {
        "url": url,
        "okoljske_spremenljivke": env_status,
        "POVZETEK": (f"✓ Deluje: {uspesni[0]['nacin']} (status {uspesni[0]['status']})"
                     if uspesni else "✗ Nobena kombinacija ni uspela — vse vrnejo 401. Beta verjetno rabi druge poverilnice ali Render IP na whitelist."),
        "rezultati": rezultati,
        "namig": "Če B) (samo Basic Auth) ne vrne 401, je problem bil, da je ključ povozil Basic Auth. Pošlji ta izpis Marku.",
    }


@app.get("/siluxar-push-debug")
async def siluxar_push_debug():
    """Debug: pokaže zadnjih 50 pošiljanj pozicij v siluxar (kaj poslano, status, odgovor)."""
    if not SILUXAR_PUSH_LOG.exists():
        return {"sporocilo": "Še ni nobenega pošiljanja.", "zapisi": []}
    try:
        log = json.loads(SILUXAR_PUSH_LOG.read_text(encoding="utf-8"))
        return {
            "skupaj_zapisov": len(log) if isinstance(log, list) else 0,
            "zadnje_posiljanje": log[0] if log else None,
            "zapisi": log,
        }
    except Exception as e:
        return {"error": str(e)}


@app.post("/pozicije-update-one")
async def pozicije_update_one(data: dict):
    """Hitra sprememba pozicije IN/ALI količine ENEGA SKU-ja — zapiše takoj v CSV (pop-up pri zalogi).
    position: nova pozicija (opcijsko). stock: nova količina (opcijsko, samo če podana)."""
    try:
        sku = (data.get("sku") or "").strip()
        position = data.get("position")
        stock = data.get("stock")
        set_pos = position is not None
        set_stock = stock is not None and str(stock).strip() != ""
        if set_pos:
            position = str(position).strip()
        if not sku:
            return {"ok": False, "error": "Manjka SKU"}
        if not set_pos and not set_stock:
            return {"ok": False, "error": "Ni česa posodobiti (ne pozicije ne količine)."}
        if not STOCK_CSV_FILE.exists():
            return {"ok": False, "error": "Baza zaloge ni naložena."}

        # validiraj stock (cela števila)
        new_stock_val = None
        if set_stock:
            try:
                new_stock_val = str(int(float(str(stock).replace(",", "."))))
            except Exception:
                return {"ok": False, "error": "Količina mora biti število."}

        import csv as _csv
        from io import StringIO as _SIO
        text = STOCK_CSV_FILE.read_text(encoding="utf-8-sig", errors="replace")
        first_line = text.split("\n", 1)[0]
        sep = ";" if first_line.count(";") > first_line.count(",") else ","
        reader = _csv.DictReader(_SIO(text), delimiter=sep)
        fieldnames = reader.fieldnames or []
        if set_pos and "position" not in fieldnames:
            fieldnames = fieldnames + ["position"]
        if set_stock and "stock" not in fieldnames:
            fieldnames = fieldnames + ["stock"]
        rows = list(reader)

        # opcijsko: omeji na določeno skladišče (če podano), sicer vse vrstice tega SKU
        wh_filter = (data.get("warehouse") or "").strip().lower()

        updated = 0
        for row in rows:
            rsku = (row.get("product_sku") or row.get("sku") or "").strip()
            if rsku != sku:
                continue
            if wh_filter and (row.get("warehouse") or "").strip().lower() != wh_filter:
                continue
            if set_pos:
                row["position"] = position
            if set_stock:
                row["stock"] = new_stock_val
            updated += 1
        if updated == 0:
            return {"ok": False, "error": f"SKU '{sku}' ni v bazi zaloge"}

        out = _SIO()
        writer = _csv.DictWriter(out, fieldnames=fieldnames, delimiter=sep, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        tmp = STOCK_CSV_FILE.with_suffix(".tmp")
        tmp.write_text(out.getvalue(), encoding="utf-8-sig")
        import os as _os
        _os.replace(str(tmp), str(STOCK_CSV_FILE))
        result = {"ok": True, "updated": updated, "sku": sku}
        if set_pos:
            result["position"] = position
        if set_stock:
            result["stock"] = new_stock_val
        return result
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "tb": traceback.format_exc()}


@app.get("/zaloga-extra-positions")
async def zaloga_extra_positions_get():
    """Vrne vse dodatne (backup) lokacije: { sku: [pozicije...] }.
    Frontend to zlije s primarno pozicijo iz CSV za prikaz."""
    return {"ok": True, "extra": _zaloga_load_extra_pos()}

@app.post("/zaloga-extra-position-add")
async def zaloga_extra_position_add(data: dict):
    """Doda dodatno lokacijo SKU-ju (backup zaloga). Ne dotika se primarne pozicije."""
    try:
        sku = (data.get("sku") or "").strip()
        pos = (data.get("position") or "").strip()
        if not sku or not pos:
            return {"ok": False, "error": "Manjka SKU ali pozicija."}
        store = _zaloga_load_extra_pos()
        lst = store.get(sku, [])
        if not isinstance(lst, list):
            lst = []
        # ne podvajaj (case-insensitive)
        if any(p.strip().lower() == pos.lower() for p in lst):
            return {"ok": False, "error": f"Lokacija '{pos}' je že dodana."}
        lst.append(pos)
        store[sku] = lst
        _zaloga_save_extra_pos(store)
        return {"ok": True, "sku": sku, "positions": lst}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/zaloga-extra-position-remove")
async def zaloga_extra_position_remove(data: dict):
    """Odstrani eno dodatno lokacijo SKU-ja."""
    try:
        sku = (data.get("sku") or "").strip()
        pos = (data.get("position") or "").strip()
        if not sku or not pos:
            return {"ok": False, "error": "Manjka SKU ali pozicija."}
        store = _zaloga_load_extra_pos()
        lst = store.get(sku, [])
        if isinstance(lst, list):
            lst = [p for p in lst if p.strip().lower() != pos.lower()]
        if lst:
            store[sku] = lst
        else:
            store.pop(sku, None)
        _zaloga_save_extra_pos(store)
        return {"ok": True, "sku": sku, "positions": store.get(sku, [])}
    except Exception as e:
        return {"ok": False, "error": str(e)}



# ════════════════════════════════════════════════════════════════════
#  BADGE GENERATOR (skrita stran /badge-generator) — enkratna naloga
#  Zamenja title_suffix v dveh CSV-jih z AI-generiranimi značkami prek
#  Anthropic Message Batches API. Batch teče do 12h v ozadju; stanje
#  (batch ID, naloženi CSV) je na DATA_DIR (perzistenten disk).
# ════════════════════════════════════════════════════════════════════
import html as _html_mod
import csv      # csv ni globalno importiran v tem projektu (json in re sta — glej vrh datoteke)

BADGE_DIR = DATA_DIR / "badge_generator"
BADGE_DIR.mkdir(exist_ok=True, parents=True)
BADGE_STATE = BADGE_DIR / "state.json"
BADGE_MODEL = "claude-sonnet-4-6"
BADGE_TR_LOCALES = ["bg", "bs", "cs", "el", "hr", "hu", "it", "pl", "ro", "sk", "sr"]
BADGE_EXCLUDE = {"de", "de-AT"}
BADGE_DESC_MAX = 900


def _badge_kategorija(product_code):
    """Določi kategorijo iz product_code za boljši kontekst značke.
    - same številke (npr. 02187) → avto-moto
    - KX* → Ikonka, prosti čas / dom / hobi
    - drugo → brez specifičnega (mešano: elektronika ipd.)"""
    pc = str(product_code or "").strip()
    if not pc:
        return ""
    if pc[0].isdigit():
        return ("Kontekst: ta izdelek je AVTOMOBILSKI / MOTO dodatek. "
                "Besede kot 'vzglavnik' pomenijo naslon za glavo v avtu (ne postelje), "
                "'držalo' je za avto ipd.")
    if pc.upper().startswith("KX"):
        return "Kontekst: ta izdelek je za prosti čas, dom ali hobi."
    return ""


def _badge_state():
    if BADGE_STATE.exists():
        try:
            return json.loads(BADGE_STATE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def _badge_save_state(st):
    BADGE_STATE.write_text(json.dumps(st, ensure_ascii=False, indent=2), encoding="utf-8")


def _badge_clean_html(s):
    if not s:
        return ""
    s = re.sub(r"<[^>]+>", " ", str(s))
    s = _html_mod.unescape(s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _badge_desc(lead, content):
    d = (_badge_clean_html(lead) + " " + _badge_clean_html(content)).strip()
    return d[:BADGE_DESC_MAX]


def _badge_prompt_sl(title, desc, kategorija=""):
    kat = f"\n{kategorija}\n" if kategorija else ""
    return f"""Si copywriter za slovenski e-commerce. Napiši title_suffix značko za spodnji produkt.

PRAVILA:
- Format: 1 emoji + 1 beseda ali kratek izraz
- Maksimalno 19 znakov skupaj (emoji = 1 znak)
- Jezik: slovenščina
- Značka naj opiše ključno lastnost ali korist produkta
- NE uporabljaj generičnih fraz kot "Hitra dostava", "Preverjena kvaliteta"
- Upoštevaj kontekst izdelka (avto/dom/...) — ne prevajaj besed dobesedno če pomen ne ustreza
- Primeri dobrih značk: 🛞 Popravilo v sili, ⚡ 3× hitreje polni, 🧲 Drži brez truda, 📺 Zabava zadaj
{kat}
Produkt:
Naslov: {title}
Opis: {desc}

Odgovori SAMO z značko, brez razlage. Npr: ⚡ Hitro polnjenje"""


def _badge_prompt_tr(per_locale, kategorija=""):
    bloki = []
    for loc in BADGE_TR_LOCALES:
        d = per_locale.get(loc)
        if d:
            bloki.append(f"[{loc}] Naslov: {d['title']}\n     Opis: {d['desc']}")
    produkti = "\n".join(bloki)
    json_predloga = "{" + ", ".join(f'"{l}": "🔧 ..."' for l in BADGE_TR_LOCALES) + "}"
    kat = f"\n{kategorija}\n" if kategorija else ""
    return f"""Si copywriter za e-commerce v več jezikih. Napiši title_suffix značko za spodnji produkt v vseh spodaj navedenih jezikih.

PRAVILA:
- Format: 1 emoji + 1 beseda ali kratek izraz
- Maksimalno 15 znakov skupaj (emoji = 1 znak) — STROGO, raje krajše
- Vsaka značka v svojem jeziku
- sr (srbščina): OBVEZNO latinica, nikoli cirilica
- bg (bolgarščina): cirilica
- Značka naj opiše ključno lastnost ali korist produkta
- Upoštevaj kontekst izdelka (avto/dom/...) — ne prevajaj dobesedno če pomen ne ustreza
- NE uporabljaj generičnih fraz
{kat}
Produkt (isti produkt v 11 jezikih):
{produkti}

Odgovori SAMO v JSON formatu (brez markdown, brez razlage):
{json_predloga}"""


# srbska cirilica → latinica (varovalka)
_BADGE_CYR2LAT = {
    'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Ђ':'Đ','Е':'E','Ж':'Ž','З':'Z','И':'I','Ј':'J',
    'К':'K','Л':'L','Љ':'Lj','М':'M','Н':'N','Њ':'Nj','О':'O','П':'P','Р':'R','С':'S','Т':'T',
    'Ћ':'Ć','У':'U','Ф':'F','Х':'H','Ц':'C','Ч':'Č','Џ':'Dž','Ш':'Š',
    'а':'a','б':'b','в':'v','г':'g','д':'d','ђ':'đ','е':'e','ж':'ž','з':'z','и':'i','ј':'j',
    'к':'k','л':'l','љ':'lj','м':'m','н':'n','њ':'nj','о':'o','п':'p','р':'r','с':'s','т':'t',
    'ћ':'ć','у':'u','ф':'f','х':'h','ц':'c','ч':'č','џ':'dž','ш':'š',
}

def _badge_cyr2lat(s):
    return "".join(_BADGE_CYR2LAT.get(ch, ch) for ch in s)

def _badge_has_cyr(s):
    return any('\u0400' <= ch <= '\u04FF' for ch in s)

def _badge_clean(s, max_len=19):
    s = (s or "").strip().split("\n")[0].strip()
    s = s.strip('"').strip("'").strip()
    # uveljavi mejo znakov — če predolgo, poreži na meji besede (emoji + izraz ohranjen)
    MAX = max_len
    if len(s) > MAX:
        cut = s[:MAX]
        # če smo prerezali sredi besede, poreži do zadnjega presledka (a ohrani emoji na začetku)
        if " " in cut and not s[MAX:MAX+1].isspace():
            sp = cut.rfind(" ")
            if sp >= 3:   # ne poreži, če bi ostal samo emoji
                cut = cut[:sp]
        s = cut.strip()
    return s


@app.post("/badge-upload")
async def badge_upload(kind: str = Form(...), file: UploadFile = File(...)):
    """Naloži CSV (kind=sl ali kind=tr). Shrani na disk + zabeleži osnovne info."""
    try:
        if kind not in ("sl", "tr"):
            return {"ok": False, "error": "kind mora biti sl ali tr"}
        raw = await file.read()
        dest = BADGE_DIR / f"input_{kind}.csv"
        dest.write_bytes(raw)
        # preštej vrstice
        import io as _io
        rdr = csv.DictReader(_io.StringIO(raw.decode("utf-8")))
        rows = list(rdr)
        fields = rdr.fieldnames or []
        info = {"rows": len(rows), "fields": list(fields)}
        if kind == "tr":
            from collections import Counter
            locs = Counter(r.get("locale", "") for r in rows)
            info["locales"] = dict(locs)
            info["products"] = locs.get("bg", 0)  # 1 locale = št. produktov
        st = _badge_state()
        st[f"input_{kind}"] = {"filename": file.filename, **info}
        _badge_save_state(st)
        return {"ok": True, "kind": kind, **info}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.post("/badge-submit")
async def badge_submit(data: dict):
    """Pripravi batch zahtevke in odda na Anthropic Batches API.
    data: {kind: 'sl'|'tr', demo: int (0=vse)}"""
    try:
        kind = data.get("kind")
        demo = int(data.get("demo", 0) or 0)
        if kind not in ("sl", "tr"):
            return {"ok": False, "error": "kind mora biti sl ali tr"}
        src = BADGE_DIR / f"input_{kind}.csv"
        if not src.exists():
            return {"ok": False, "error": f"Najprej naloži {kind} CSV"}

        rows = list(csv.DictReader(src.open(encoding="utf-8")))
        requests = []

        if kind == "sl":
            limit = demo if demo > 0 else len(rows)
            for i in range(limit):
                r = rows[i]
                desc = _badge_desc(r.get("lead", ""), r.get("content", ""))
                kat = _badge_kategorija(r.get("product_code", ""))
                requests.append({
                    "custom_id": f"sl-{i}",
                    "params": {"model": BADGE_MODEL, "max_tokens": 60,
                               "messages": [{"role": "user", "content": _badge_prompt_sl(r.get("title",""), desc, kat)}]},
                })
        else:
            # TR CSV nima product_code → kategorijo dobimo iz SL CSV po istem indeksu
            sl_codes = []
            sl_src = BADGE_DIR / "input_sl.csv"
            if sl_src.exists():
                sl_codes = [r.get("product_code", "") for r in csv.DictReader(sl_src.open(encoding="utf-8"))]
            # grupiraj po indeksu prek locale
            from collections import defaultdict
            by = defaultdict(list)
            for r in rows:
                loc = r.get("locale", "")
                if loc in BADGE_EXCLUDE:
                    continue
                if loc in BADGE_TR_LOCALES:
                    by[loc].append(r)
            n = len(by[BADGE_TR_LOCALES[0]])
            limit = demo if demo > 0 else n
            for i in range(limit):
                per = {}
                for loc in BADGE_TR_LOCALES:
                    if i < len(by[loc]):
                        rr = by[loc][i]
                        per[loc] = {"title": rr.get("title",""), "desc": _badge_desc(rr.get("lead",""), rr.get("content",""))}
                kat = _badge_kategorija(sl_codes[i]) if i < len(sl_codes) else ""
                requests.append({
                    "custom_id": f"tr-{i}",
                    "params": {"model": BADGE_MODEL, "max_tokens": 300,
                               "messages": [{"role": "user", "content": _badge_prompt_tr(per, kat)}]},
                })

        if not requests:
            return {"ok": False, "error": "Ni zahtevkov"}

        # oddaj batch
        batch = client.messages.batches.create(requests=requests)
        st = _badge_state()
        st[f"batch_{kind}"] = {
            "id": batch.id, "status": batch.processing_status,
            "count": len(requests), "demo": demo,
            "submitted_at": datetime.now().isoformat(),
        }
        _badge_save_state(st)
        return {"ok": True, "kind": kind, "batch_id": batch.id, "count": len(requests),
                "status": batch.processing_status}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/badge-status")
async def badge_status(kind: str):
    """Preveri status batcha (ne blokira — samo poizve)."""
    try:
        st = _badge_state()
        b = st.get(f"batch_{kind}")
        if not b:
            return {"ok": False, "error": "Ni oddanega batcha"}
        batch = client.messages.batches.retrieve(b["id"])
        counts = batch.request_counts
        b["status"] = batch.processing_status
        b["counts"] = {"processing": counts.processing, "succeeded": counts.succeeded,
                       "errored": counts.errored, "canceled": counts.canceled, "expired": counts.expired}
        st[f"batch_{kind}"] = b
        _badge_save_state(st)
        return {"ok": True, "kind": kind, "status": batch.processing_status,
                "counts": b["counts"], "total": b["count"], "id": b["id"]}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.post("/badge-build")
async def badge_build(data: dict):
    """Prevzemi rezultate batcha + sestavi NEW CSV (ista struktura, LF, brez DE/de-AT za tr)."""
    try:
        kind = data.get("kind")
        st = _badge_state()
        b = st.get(f"batch_{kind}")
        if not b:
            return {"ok": False, "error": "Ni oddanega batcha"}

        # prevzemi rezultate prek custom_id
        results = {}
        errors = 0
        for r in client.messages.batches.results(b["id"]):
            if r.result.type == "succeeded":
                results[r.custom_id] = r.result.message.content[0].text
            elif r.result.type == "errored":
                errors += 1

        src = BADGE_DIR / f"input_{kind}.csv"
        rows = list(csv.DictReader(src.open(encoding="utf-8")))
        fields = list(rows[0].keys()) if rows else []
        miss = 0; cyr_fixed = 0; parse_fail = 0

        if kind == "sl":
            for i, row in enumerate(rows):
                cid = f"sl-{i}"
                if cid in results:
                    row["title_suffix"] = _badge_clean(results[cid])
                else:
                    miss += 1
            outp = BADGE_DIR / "maaarket-suffux-sl-NEW.csv"
            with outp.open("w", encoding="utf-8", newline="") as f:
                w = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
                w.writeheader(); w.writerows(rows)
            out_rows_count = len(rows)
        else:
            # razčleni JSON na produkt
            parsed = {}
            def get_badges(idx):
                nonlocal parse_fail
                if idx in parsed: return parsed[idx]
                raw = results.get(f"tr-{idx}")
                if not raw:
                    parsed[idx] = {}; return {}
                m = re.search(r"\{.*\}", raw, re.S)
                if not m:
                    parse_fail += 1; parsed[idx] = {}; return {}
                try:
                    parsed[idx] = json.loads(m.group(0))
                except Exception:
                    parse_fail += 1; parsed[idx] = {}
                return parsed[idx]

            from collections import defaultdict
            loc_counter = defaultdict(int)
            out_rows = []
            for row in rows:
                loc = row.get("locale", "")
                if loc in BADGE_EXCLUDE:
                    continue
                idx = loc_counter[loc]; loc_counter[loc] += 1
                badges = get_badges(idx)
                badge = _badge_clean(badges.get(loc, ""), max_len=15) if badges else ""  # tuji jeziki: max 15 znakov
                if badge:
                    if loc == "sr" and _badge_has_cyr(badge):
                        badge = _badge_cyr2lat(badge); cyr_fixed += 1
                    row["title_suffix"] = badge
                else:
                    miss += 1
                out_rows.append(row)
            outp = BADGE_DIR / "maaarket-suffux-translations-NEW.csv"
            with outp.open("w", encoding="utf-8", newline="") as f:
                w = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
                w.writeheader(); w.writerows(out_rows)
            out_rows_count = len(out_rows)

        st[f"build_{kind}"] = {"built_at": datetime.now().isoformat(), "rows": out_rows_count,
                               "missing": miss, "errors": errors, "cyr_fixed": cyr_fixed,
                               "parse_fail": parse_fail, "file": outp.name}
        _badge_save_state(st)
        return {"ok": True, "kind": kind, "rows": out_rows_count, "missing": miss,
                "errors": errors, "cyr_fixed": cyr_fixed, "parse_fail": parse_fail}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}


@app.get("/badge-download")
async def badge_download(kind: str):
    """Prenese sestavljen NEW CSV."""
    fn = "maaarket-suffux-sl-NEW.csv" if kind == "sl" else "maaarket-suffux-translations-NEW.csv"
    path = BADGE_DIR / fn
    if not path.exists():
        return JSONResponse({"error": "Datoteka še ni sestavljena"}, status_code=404)
    return FileResponse(str(path), media_type="text/csv", filename=fn)


@app.get("/badge-generator", response_class=HTMLResponse)
async def badge_generator_page():
    """Skrita stran (ni v meniju) za badge generator."""
    return HTMLResponse(_BADGE_HTML)


_BADGE_HTML = r"""<!DOCTYPE html>
<html lang="sl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="robots" content="noindex">
<title>Badge Generator (interno)</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'DM Sans', system-ui, sans-serif; background: #f1f5f9; color: #0f172a; padding: 24px; max-width: 760px; margin: 0 auto; }
  h1 { font-size: 22px; margin-bottom: 4px; }
  .sub { color: #64748b; font-size: 13px; margin-bottom: 20px; }
  .card { background: #fff; border: 1px solid #e2e8f0; border-radius: 14px; padding: 18px; margin-bottom: 16px; }
  .card h2 { font-size: 15px; margin-bottom: 12px; display: flex; align-items: center; gap: 8px; }
  .step-num { background: #3b82f6; color: #fff; width: 24px; height: 24px; border-radius: 50%; display: inline-flex; align-items: center; justify-content: center; font-size: 13px; font-weight: 700; }
  .row { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; margin-bottom: 10px; }
  label.lbl { font-weight: 600; font-size: 13px; min-width: 110px; }
  input[type=file] { font-size: 13px; }
  button { font-family: inherit; font-size: 14px; font-weight: 700; padding: 9px 16px; border-radius: 9px; border: none; cursor: pointer; background: #3b82f6; color: #fff; }
  button:hover { filter: brightness(1.08); }
  button.sec { background: #e2e8f0; color: #0f172a; }
  button:disabled { opacity: 0.5; cursor: not-allowed; }
  .info { font-size: 12px; color: #64748b; margin-top: 6px; white-space: pre-wrap; }
  .badge-pill { display: inline-block; background: #dbeafe; color: #1e40af; font-size: 11px; font-weight: 700; padding: 2px 8px; border-radius: 20px; }
  .warn { background: #fef3c7; border: 1px solid #fde68a; color: #92400e; font-size: 12px; padding: 8px 12px; border-radius: 8px; margin-bottom: 14px; }
  .ok { color: #16a34a; } .err { color: #dc2626; }
  progress { width: 100%; height: 10px; }
</style>
</head>
<body>
  <h1>🏷️ Badge Generator</h1>
  <div class="sub">Interno orodje — zamenjava title_suffix prek Anthropic Batch API. Batch teče do 12h v ozadju.</div>
  <div class="warn">⚠️ Enkratna naloga. Batch se procesira na Anthropicovih strežnikih — to stran lahko zapreš in se vrneš kasneje. Stanje se ohrani.</div>

  <div class="card">
    <h2><span class="step-num">1</span> Naloži CSV</h2>
    <div class="row">
      <label class="lbl">SL CSV:</label>
      <input type="file" id="fileSL" accept=".csv">
      <button onclick="upload('sl')">Naloži SL</button>
    </div>
    <div class="info" id="infoSL"></div>
    <div class="row" style="margin-top:10px">
      <label class="lbl">Prevodi CSV:</label>
      <input type="file" id="fileTR" accept=".csv">
      <button onclick="upload('tr')">Naloži prevode</button>
    </div>
    <div class="info" id="infoTR"></div>
  </div>

  <div class="card">
    <h2><span class="step-num">2</span> Oddaj batch</h2>
    <div class="row">
      <label class="lbl"><input type="checkbox" id="demoChk"> Samo demo</label>
      <input type="number" id="demoN" value="10" style="width:70px" min="1"> produktov
    </div>
    <div class="info">Brez kljukice = vsi produkti (~$26, do 12h). Z demo = test na N produktih.</div>
    <div class="row" style="margin-top:10px">
      <button onclick="submitB('sl')">▶ Oddaj SL batch</button>
      <button onclick="submitB('tr')">▶ Oddaj prevode batch</button>
    </div>
    <div class="info" id="infoSubmit"></div>
  </div>

  <div class="card">
    <h2><span class="step-num">3</span> Status & prenos</h2>
    <div class="row">
      <button class="sec" onclick="status('sl')">↻ Status SL</button>
      <button class="sec" onclick="status('tr')">↻ Status prevodi</button>
    </div>
    <div class="info" id="infoStatusSL"></div>
    <div class="info" id="infoStatusTR"></div>
    <div class="row" style="margin-top:12px">
      <button onclick="build('sl')">🔧 Sestavi SL NEW</button>
      <button onclick="build('tr')">🔧 Sestavi prevode NEW</button>
    </div>
    <div class="info" id="infoBuild"></div>
    <div class="row" style="margin-top:12px">
      <a href="/badge-download?kind=sl"><button class="sec">⬇ Prenesi SL NEW</button></a>
      <a href="/badge-download?kind=tr"><button class="sec">⬇ Prenesi prevode NEW</button></a>
    </div>
  </div>

<script>
async function upload(kind) {
  const f = document.getElementById('file'+kind.toUpperCase()).files[0];
  const el = document.getElementById('info'+kind.toUpperCase());
  if (!f) { el.textContent = 'Izberi datoteko.'; return; }
  el.textContent = 'Nalagam...';
  const fd = new FormData(); fd.append('kind', kind); fd.append('file', f);
  try {
    const r = await fetch('/badge-upload', { method:'POST', body: fd });
    const d = await r.json();
    if (d.ok) {
      let t = `✓ Naloženo: ${d.rows} vrstic`;
      if (d.products) t += ` · ${d.products} produktov · jeziki: ${Object.keys(d.locales).join(', ')}`;
      el.innerHTML = '<span class="ok">'+t+'</span>';
    } else el.innerHTML = '<span class="err">✗ '+d.error+'</span>';
  } catch(e) { el.innerHTML = '<span class="err">✗ napaka</span>'; }
}
async function submitB(kind) {
  const demo = document.getElementById('demoChk').checked ? parseInt(document.getElementById('demoN').value)||10 : 0;
  const el = document.getElementById('infoSubmit');
  el.textContent = 'Oddajam batch...';
  try {
    const r = await fetch('/badge-submit', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({kind, demo}) });
    const d = await r.json();
    if (d.ok) el.innerHTML = `<span class="ok">✓ ${kind.toUpperCase()} batch oddan: ${d.count} zahtevkov · ID ${d.batch_id} · status ${d.status}</span>`;
    else el.innerHTML = '<span class="err">✗ '+d.error+'</span>';
  } catch(e) { el.innerHTML = '<span class="err">✗ napaka</span>'; }
}
async function status(kind) {
  const el = document.getElementById('infoStatus'+kind.toUpperCase());
  el.textContent = 'Preverjam...';
  try {
    const r = await fetch('/badge-status?kind='+kind);
    const d = await r.json();
    if (d.ok) {
      const c = d.counts;
      el.innerHTML = `<b>${kind.toUpperCase()}</b>: ${d.status} — ✓${c.succeeded} ⏳${c.processing} ✗${c.errored} / ${d.total}`;
    } else el.innerHTML = '<span class="err">✗ '+d.error+'</span>';
  } catch(e) { el.innerHTML = '<span class="err">✗ napaka</span>'; }
}
async function build(kind) {
  const el = document.getElementById('infoBuild');
  el.textContent = 'Sestavljam NEW CSV...';
  try {
    const r = await fetch('/badge-build', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({kind}) });
    const d = await r.json();
    if (d.ok) el.innerHTML = `<span class="ok">✓ ${kind.toUpperCase()} sestavljen: ${d.rows} vrstic · manjka ${d.missing} · napak ${d.errors}${d.cyr_fixed?' · sr lat. popravkov '+d.cyr_fixed:''}${d.parse_fail?' · JSON fail '+d.parse_fail:''}</span>`;
    else el.innerHTML = '<span class="err">✗ '+d.error+'</span>';
  } catch(e) { el.innerHTML = '<span class="err">✗ napaka</span>'; }
}
</script>
</body>
</html>"""


# ═══════════════════════════════════════════════════════════════
#  POLCAR — SOAP integracija (dedal.polcar.com / Nemesis WebServices)
#  Faza 1: povezava + branje cen (GetCustomerPrices). Diagnostika.
# ═══════════════════════════════════════════════════════════════
POLCAR_WSDL_URL   = "https://dedal.polcar.com/Dystrybutorzy/Customers.asmx"
POLCAR_NS         = "http://dedal.polcar.com.pl/Nemesis/WebServices"
POLCAR_LOGIN      = os.environ.get("POLCAR_LOGIN", "silux")
POLCAR_PASSWORD   = os.environ.get("POLCAR_PASSWORD", "sk32H@sks!12lsMM")
POLCAR_CUSTOMER   = os.environ.get("POLCAR_CUSTOMER", "EX426")
POLCAR_DISTRIB    = os.environ.get("POLCAR_DISTRIBUTOR", "")


def _polcar_xml_escape(s: str) -> str:
    return (str(s or "")
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&apos;"))


@app.get("/polcar", response_class=HTMLResponse)
async def polcar_page():
    return FileResponse("static/polcar.html")


@app.get("/polcar-prices")
async def polcar_prices(limit: int = 50):
    """SOAP GetCustomerPrices — vrne surov odgovor + parsano strukturo (diagnostika faze 1)."""
    soap_action = f"{POLCAR_NS}/GetCustomerPrices"
    body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<soap:Envelope xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
        'xmlns:xsd="http://www.w3.org/2001/XMLSchema">'
        '<soap:Body>'
        f'<GetCustomerPrices xmlns="{POLCAR_NS}">'
        f'<DistributorCode>{_polcar_xml_escape(POLCAR_DISTRIB)}</DistributorCode>'
        f'<CustomerNumber>{_polcar_xml_escape(POLCAR_CUSTOMER)}</CustomerNumber>'
        f'<Login>{_polcar_xml_escape(POLCAR_LOGIN)}</Login>'
        f'<Password>{_polcar_xml_escape(POLCAR_PASSWORD)}</Password>'
        '</GetCustomerPrices>'
        '</soap:Body></soap:Envelope>'
    )
    headers = {
        "Content-Type": "text/xml; charset=utf-8",
        "SOAPAction": f'"{soap_action}"',
    }
    try:
        async with httpx.AsyncClient(timeout=60) as cli:
            r = await cli.post(POLCAR_WSDL_URL, content=body.encode("utf-8"), headers=headers)
        raw = r.text
        out = {
            "ok": r.status_code == 200,
            "status": r.status_code,
            "raw_len": len(raw),
            "raw_head": raw[:3000],
        }
        # poskusi izluščiti vsebino GetCustomerPricesResult
        try:
            root = ET.fromstring(raw)
            # poišči Result element ne glede na namespace prefix
            result_el = None
            for el in root.iter():
                if el.tag.endswith("GetCustomerPricesResult"):
                    result_el = el
                    break
            if result_el is not None:
                # vsebina je lahko nested XML elementi ALI besedilo
                children = list(result_el)
                if children:
                    # serializiraj prvih nekaj otrok za pregled strukture
                    sample = []
                    # poišči ponavljajoče se zapise (npr. tabela cen)
                    rows = []
                    for child in result_el.iter():
                        rows.append(child.tag.split("}")[-1])
                    from collections import Counter
                    tag_counts = Counter(rows)
                    out["result_tags"] = dict(tag_counts.most_common(20))
                    # vzorec: prvih `limit` "row" elementov (najpogostejši ponavljajoč tag globlje)
                    inner_xml = ET.tostring(result_el, encoding="unicode")
                    out["result_xml_head"] = inner_xml[:5000]
                    out["result_child_count"] = len(children)
                else:
                    txt = (result_el.text or "")
                    out["result_text_len"] = len(txt)
                    out["result_text_head"] = txt[:5000]
            else:
                out["note"] = "GetCustomerPricesResult element ni najden v odgovoru"
        except ET.ParseError as pe:
            out["parse_error"] = str(pe)
        return JSONResponse(out)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e), "type": type(e).__name__}, status_code=200)


# ═══════════════════════════════════════════════════════════════
#  OPTIMIZACIJA SLIK — regeneracija prek OpenAI gpt-image-2
#  Branje: maaarket API po SKU. Shramba: /data/regen (stabilni linki).
#  Write nazaj na maaarket: /regen-push (stub — programer doda POST).
# ═══════════════════════════════════════════════════════════════
MAAARKET_IMAGES_URL = os.environ.get("MAAARKET_IMAGES_URL", "https://api.maaarket.si/api/v1/images")
REGEN_DIR = DATA_DIR / "regen"
REGEN_MODEL = os.environ.get("REGEN_OPENAI_MODEL", "gpt-image-2-2026-04-21")


def _regen_ensure_dir():
    try:
        REGEN_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass


@app.get("/regen-read")
async def regen_read(sku: str):
    """Preberi slike izdelka po SKU iz maaarket (glavna + galerija)."""
    sku = (sku or "").strip()
    if not sku:
        return JSONResponse({"ok": False, "error": "Manjka SKU"}, status_code=200)
    try:
        async with httpx.AsyncClient(timeout=30) as cli:
            r = await cli.get(MAAARKET_IMAGES_URL, params={"sku": sku})
        if r.status_code != 200:
            return JSONResponse({"ok": False, "error": f"maaarket status {r.status_code}", "body": r.text[:300]}, status_code=200)
        data = r.json()
        return JSONResponse({"ok": True, "sku": sku, "data": data})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e), "type": type(e).__name__}, status_code=200)


@app.get("/regen-img/{fname}")
async def regen_img(fname: str):
    """Servira regenerirano sliko (stabilen link za maaarket prenos)."""
    # varnost: samo ime datoteke, brez poti
    safe = os.path.basename(fname)
    fpath = REGEN_DIR / safe
    if not fpath.exists():
        raise HTTPException(status_code=404, detail="ni najdena")
    media = "image/png" if safe.lower().endswith(".png") else "image/jpeg"
    return FileResponse(str(fpath), media_type=media)


def _pick_openai_size(w, h):
    if not w or not h:
        return "auto"
    ar = w / h
    if ar >= 1.25:
        return "1536x1024"   # landscape
    if ar <= 0.8:
        return "1024x1536"   # portrait
    return "1024x1024"       # ~kvadrat


async def _regen_generate_one(src_url: str, prompt: str):
    """Generira eno sliko prek OpenAI gpt-image-2 in shrani na /data/regen.
    Vrne (public_url, meta_dict, error_str). Ob napaki je public_url None."""
    openai_key = os.environ.get("OPENAI_API_KEY", "")
    if not openai_key:
        return None, {}, "OPENAI_API_KEY ni nastavljen."
    src_url = (src_url or "").strip()
    prompt = (prompt or "").strip()
    if not src_url:
        return None, {}, "Manjka image_url"
    if not prompt:
        return None, {}, "Manjka prompt"
    _regen_ensure_dir()
    # 1) prenesi izvorno sliko + izmeri dimenzije
    try:
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as cli:
            ir = await cli.get(src_url)
        if ir.status_code != 200:
            return None, {}, f"Vir slike status {ir.status_code}"
        src_bytes = ir.content
        src_mime = ir.headers.get("content-type", "image/jpeg").split(";")[0]
        src_ext = "png" if "png" in src_mime else "jpg"
        src_w = src_h = None
        try:
            from PIL import Image as _PILImage
            import io as _io2
            with _PILImage.open(_io2.BytesIO(src_bytes)) as _im:
                src_w, src_h = _im.size
        except Exception:
            src_w = src_h = None
    except Exception as e:
        return None, {}, f"Prenos vira: {e}"
    openai_size = _pick_openai_size(src_w, src_h)
    # 2) OpenAI gpt-image-2 edits (z retry ob 429 rate limit)
    try:
        files = {"image[]": (f"src.{src_ext}", src_bytes, src_mime)}
        form = {"model": REGEN_MODEL, "prompt": prompt, "size": openai_size, "n": "1", "output_format": "jpeg"}
        result = None
        last_status = None
        for _attempt in range(4):  # do 4 poskusov ob 429
            async with httpx.AsyncClient(timeout=240) as hc:
                resp = await hc.post(
                    "https://api.openai.com/v1/images/edits",
                    headers={"Authorization": f"Bearer {openai_key}"},
                    data=form, files=files,
                )
            last_status = resp.status_code
            if resp.status_code == 429:
                # spoštuj Retry-After, sicer eksponentni backoff
                try:
                    wait = float(resp.headers.get("retry-after", "")) or (2 ** _attempt) * 3
                except Exception:
                    wait = (2 ** _attempt) * 3
                await asyncio.sleep(min(wait, 30))
                continue
            result = resp.json()
            break
        if result is None:
            return None, {}, f"OpenAI rate limit (429) — preveč poskusov."
        if last_status != 200:
            return None, {}, result.get("error", {}).get("message", str(result))[:300]
        data_arr = result.get("data", [])
        if not (data_arr and data_arr[0].get("b64_json")):
            return None, {}, "OpenAI ni vrnil slike: " + str(result)[:200]
        out_b64 = data_arr[0]["b64_json"]
    except Exception as e:
        return None, {}, f"OpenAI: {e}"
    # 3) resize na TOČNO original + shrani JPG
    try:
        import base64 as _b64m
        out_bytes = _b64m.b64decode(out_b64)
        resized = False
        if src_w and src_h:
            try:
                from PIL import Image as _PILImage
                import io as _io3
                with _PILImage.open(_io3.BytesIO(out_bytes)) as _gen:
                    _gen = _gen.convert("RGB")
                    gw, gh = _gen.size
                    target_ar = src_w / src_h
                    gen_ar = gw / gh
                    if gen_ar > target_ar:
                        new_h = src_h
                        new_w = int(round(gw * (src_h / gh)))
                    else:
                        new_w = src_w
                        new_h = int(round(gh * (src_w / gw)))
                    _gen2 = _gen.resize((max(new_w, src_w), max(new_h, src_h)), _PILImage.LANCZOS)
                    gw2, gh2 = _gen2.size
                    left = (gw2 - src_w) // 2
                    top = (gh2 - src_h) // 2
                    _gen3 = _gen2.crop((left, top, left + src_w, top + src_h))
                    _buf = _io3.BytesIO()
                    _gen3.save(_buf, format="JPEG", quality=90)
                    out_bytes = _buf.getvalue()
                    resized = True
            except Exception:
                resized = False
        fname = f"{uuid.uuid4().hex}.jpg"
        (REGEN_DIR / fname).write_bytes(out_bytes)
        public_url = f"/regen-img/{fname}"
        meta = {
            "filename": fname,
            "src_size": (f"{src_w}x{src_h}" if src_w else None),
            "openai_size": openai_size,
            "resized": resized,
        }
        return public_url, meta, None
    except Exception as e:
        return None, {}, f"Shranjevanje: {e}"


class RegenImageReq(BaseModel):
    image_url: str
    prompt: str
    sku: Optional[str] = None
    image_id: Optional[str] = None
    kind: Optional[str] = None  # "main" ali "gallery"


@app.post("/regen-image")
async def regen_image(req: RegenImageReq):
    """Sinhrono generiranje ene slike (takojšen rezultat)."""
    url, meta, err = await _regen_generate_one(req.image_url, req.prompt)
    if err:
        return JSONResponse({"ok": False, "error": err}, status_code=200)
    return JSONResponse({
        "ok": True,
        "url": url,
        "filename": meta.get("filename"),
        "sku": req.sku,
        "image_id": req.image_id,
        "kind": req.kind,
        "src_size": meta.get("src_size"),
        "openai_size": meta.get("openai_size"),
        "resized": meta.get("resized"),
    })


class RegenPushReq(BaseModel):
    sku: str
    main: Optional[dict] = None       # {id, picture(link), picture_path}
    gallery: Optional[List[dict]] = None  # [{id, picture(link), picture_path}, ...] samo spremenjene


def _strip_domain(path: str) -> str:
    """picture_path brez domene (Marko zahteva). https://host/storage/... -> /storage/...
    Relativne poti (Igrace/...) pusti nedotaknjene."""
    p = (path or "").strip()
    if p.startswith("http://") or p.startswith("https://"):
        try:
            from urllib.parse import urlparse as _up
            parsed = _up(p)
            return parsed.path or p
        except Exception:
            return p
    return p


@app.post("/regen-push")
async def regen_push(req: RegenPushReq):
    """Pošlje regenerirane slike nazaj na maaarket (POST /api/v1/images).
    Payload: glavna na vrhu, picture_path BREZ domene, picture = public link."""
    # zgradi payload v strukturi, kot jo pričakuje maaarket (zrcalo branja)
    payload = {"sku": req.sku}
    if req.main:
        payload["picture"] = req.main.get("picture", "")
        payload["picture_path"] = _strip_domain(req.main.get("picture_path", ""))
        if req.main.get("id") is not None:
            payload["id"] = req.main.get("id")
    if req.gallery:
        payload["gallery"] = [
            {
                "id": g.get("id"),
                "picture": g.get("picture", ""),
                "picture_path": _strip_domain(g.get("picture_path", "")),
            }
            for g in req.gallery
        ]
    # POST na maaarket (isti URL kot branje, brez parametrov)
    try:
        async with httpx.AsyncClient(timeout=120) as cli:
            r = await cli.post(MAAARKET_IMAGES_URL, json=payload)
        body = r.text[:600]
        if r.status_code not in (200, 201):
            return JSONResponse({
                "ok": False,
                "error": f"maaarket status {r.status_code}",
                "body": body,
                "sent": payload,
            }, status_code=200)
        # poskusi prebrati JSON odgovor (če ga vrne)
        try:
            resp_json = r.json()
        except Exception:
            resp_json = None
        return JSONResponse({
            "ok": True,
            "status": r.status_code,
            "response": resp_json if resp_json is not None else body,
            "sent": payload,
        })
    except Exception as e:
        return JSONResponse({
            "ok": False,
            "error": str(e),
            "type": type(e).__name__,
            "sent": payload,
        }, status_code=200)



# ═══════════════════════════════════════════════════════════════
#  REGEN ČAKALNICA — backend job queue (generiranje teče na strežniku)
#  Push nazaj ostane ROČEN (po pregledu). Persistenca na /data.
# ═══════════════════════════════════════════════════════════════
REGEN_QUEUE_FILE = DATA_DIR / "regen_queue.json"
_regen_queue_lock = asyncio.Lock()


def _regen_queue_load():
    try:
        if REGEN_QUEUE_FILE.exists():
            return json.loads(REGEN_QUEUE_FILE.read_text(encoding="utf-8")) or []
    except Exception:
        pass
    return []


def _regen_queue_save(jobs):
    try:
        REGEN_QUEUE_FILE.write_text(json.dumps(jobs, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


class RegenEnqueueImage(BaseModel):
    image_id: Optional[str] = None
    kind: str = "gallery"          # "main" ali "gallery"
    src_url: str
    picture_path: Optional[str] = ""
    prompt: str


class RegenEnqueueReq(BaseModel):
    sku: str
    images: List[RegenEnqueueImage]


@app.post("/regen-enqueue")
async def regen_enqueue(req: RegenEnqueueReq):
    """Doda nov job v čakalnico. Worker ga obdela v ozadju."""
    if not req.images:
        return JSONResponse({"ok": False, "error": "Ni izbranih slik."}, status_code=200)
    job = {
        "id": uuid.uuid4().hex[:12],
        "sku": req.sku,
        "status": "pending",          # pending → processing → done/error
        "created": datetime.now(timezone.utc).isoformat(),
        "updated": datetime.now(timezone.utc).isoformat(),
        "images": [
            {
                "image_id": im.image_id,
                "kind": im.kind,
                "src_url": im.src_url,
                "picture_path": im.picture_path or "",
                "prompt": im.prompt,
                "status": "pending",      # pending → done/error
                "result_url": None,
                "error": None,
            }
            for im in req.images
        ],
    }
    async with _regen_queue_lock:
        jobs = _regen_queue_load()
        jobs.append(job)
        _regen_queue_save(jobs)
    return JSONResponse({"ok": True, "id": job["id"], "count": len(job["images"])})


@app.get("/regen-queue")
async def regen_queue_get():
    """Vrne celotno čakalnico (za polling v UI)."""
    return JSONResponse({"ok": True, "jobs": _regen_queue_load()})


@app.delete("/regen-queue/{job_id}")
async def regen_queue_delete(job_id: str):
    async with _regen_queue_lock:
        jobs = _regen_queue_load()
        jobs = [j for j in jobs if j.get("id") != job_id]
        _regen_queue_save(jobs)
    return JSONResponse({"ok": True})


@app.post("/regen-queue-clear-done")
async def regen_queue_clear_done():
    """Počisti dokončane/napake jobe iz vrste."""
    async with _regen_queue_lock:
        jobs = _regen_queue_load()
        jobs = [j for j in jobs if j.get("status") not in ("done", "error")]
        _regen_queue_save(jobs)
    return JSONResponse({"ok": True})


class RegenPushedReq(BaseModel):
    job_ids: Optional[List[str]] = None    # označi te jobe kot pushed
    pushed: bool = True                    # true=označi, false=odznači
    all_from_id: Optional[str] = None      # označi ta job IN vse STAREJŠE (po vrstnem redu v queue)


@app.post("/regen-queue-mark-pushed")
async def regen_queue_mark_pushed(req: RegenPushedReq):
    """Označi/odznači jobe kot 'poslane nazaj'. Persistentno na jobu."""
    async with _regen_queue_lock:
        jobs = _regen_queue_load()
        target_ids = set(req.job_ids or [])
        # "all_from_id": ta job + vsi starejši (joby so v vrsti dodani od najstarejšega → najnovejši,
        # torej index <= index(all_from_id))
        if req.all_from_id:
            idx = next((i for i, j in enumerate(jobs) if j.get("id") == req.all_from_id), None)
            if idx is not None:
                for j in jobs[:idx+1]:
                    target_ids.add(j.get("id"))
        n = 0
        for j in jobs:
            if j.get("id") in target_ids:
                j["pushed"] = bool(req.pushed)
                if req.pushed:
                    j["pushed_at"] = datetime.now(timezone.utc).isoformat()
                n += 1
        _regen_queue_save(jobs)
    return JSONResponse({"ok": True, "marked": n})


async def _regen_worker_loop():
    """Background worker: obdeluje pending jobe enega za drugim."""
    await asyncio.sleep(8)  # počakaj startup
    while True:
        try:
            # poišči prvi pending job
            async with _regen_queue_lock:
                jobs = _regen_queue_load()
                target = None
                for j in jobs:
                    if j.get("status") == "pending":
                        target = j
                        break
                if target:
                    target["status"] = "processing"
                    target["updated"] = datetime.now(timezone.utc).isoformat()
                    _regen_queue_save(jobs)
            if not target:
                await asyncio.sleep(5)
                continue
            # generiraj slike PARALELNO (Tier 2 = 20 IPM, 6 hkrati varno pod limitom)
            job_id = target["id"]
            _sema = asyncio.Semaphore(6)

            async def _process_image(idx, im):
                if im.get("status") == "done":
                    return
                async with _sema:
                    url, meta, err = await _regen_generate_one(im.get("src_url", ""), im.get("prompt", ""))
                # zapiši rezultat takoj (sproti viden napredek)
                async with _regen_queue_lock:
                    jobs2 = _regen_queue_load()
                    jj2 = next((x for x in jobs2 if x.get("id") == job_id), None)
                    if jj2 is None:
                        return
                    if idx < len(jj2["images"]):
                        if err:
                            jj2["images"][idx]["status"] = "error"
                            jj2["images"][idx]["error"] = err
                        else:
                            jj2["images"][idx]["status"] = "done"
                            jj2["images"][idx]["result_url"] = url
                            jj2["images"][idx]["error"] = None
                        jj2["updated"] = datetime.now(timezone.utc).isoformat()
                        _regen_queue_save(jobs2)

            await asyncio.gather(*[_process_image(i, im) for i, im in enumerate(target["images"])])
            # zaključi job: status glede na rezultate
            async with _regen_queue_lock:
                jobs = _regen_queue_load()
                jj = next((x for x in jobs if x.get("id") == job_id), None)
                if jj is not None:
                    states = [i.get("status") for i in jj["images"]]
                    if all(s == "done" for s in states):
                        jj["status"] = "done"
                    elif all(s == "error" for s in states):
                        jj["status"] = "error"
                    else:
                        jj["status"] = "done"  # delno — vsaj nekaj uspelo
                    jj["updated"] = datetime.now(timezone.utc).isoformat()
                    _regen_queue_save(jobs)
        except Exception as e:
            print(f"[regen-worker] napaka: {e}")
            await asyncio.sleep(5)


# ═══════════════════════════════════════════════════════════════
#  SCALE LOG — dnevnik dvigov bidcapa (Analiza beta, scale po državah)
#  Cooldown 7 POLNIH dni od datuma dviga. Match po imenu ad-seta.
#  Persistenca /data, deljeno med brskalniki.
# ═══════════════════════════════════════════════════════════════
SCALE_LOG_FILE = DATA_DIR / "scale_log.json"
SCALE_COOLDOWN_DAYS = 7
_scale_log_lock = asyncio.Lock()


def _scale_log_load():
    try:
        if SCALE_LOG_FILE.exists():
            return json.loads(SCALE_LOG_FILE.read_text(encoding="utf-8")) or []
    except Exception:
        pass
    return []


def _scale_log_save(entries):
    try:
        SCALE_LOG_FILE.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def _scale_cooldown_active(scaled_date_str: str):
    """Vrne (active: bool, days_left: int, end_date: str).
    Cooldown = 7 POLNIH dni od datuma dviga. Dvig 17.6. → aktiven do vključno 24.6., spet prost 25.6."""
    try:
        d = datetime.strptime(scaled_date_str[:10], "%Y-%m-%d").date()
    except Exception:
        return False, 0, ""
    from datetime import timedelta as _td
    end = d + _td(days=SCALE_COOLDOWN_DAYS)   # 17.6 + 7 = 24.6 (zadnji dan opazovanja); prost 25.6
    today = datetime.now(timezone.utc).date()
    days_left = (end - today).days
    active = today <= end
    return active, max(days_left, 0), end.isoformat()


@app.get("/scale-log")
async def scale_log_get():
    """Vrne vse zapise z izračunanim cooldown statusom (aktivni + potekli)."""
    entries = _scale_log_load()
    out = []
    for e in entries:
        active, days_left, end_date = _scale_cooldown_active(e.get("scaled_date", ""))
        out.append({**e, "cooldown_active": active, "days_left": days_left, "cooldown_end": end_date})
    return JSONResponse({"ok": True, "entries": out, "cooldown_days": SCALE_COOLDOWN_DAYS})


class ScaleLogReq(BaseModel):
    ad_set: str
    scaled_date: Optional[str] = None   # YYYY-MM-DD; če prazno → danes
    bid_before: Optional[str] = None
    bid_after: Optional[str] = None
    cpa_before: Optional[float] = None
    note: Optional[str] = None


@app.post("/scale-log")
async def scale_log_add(req: ScaleLogReq):
    ad_set = (req.ad_set or "").strip()
    if not ad_set:
        return JSONResponse({"ok": False, "error": "Manjka ad_set"}, status_code=200)
    sdate = (req.scaled_date or "").strip() or datetime.now(timezone.utc).date().isoformat()
    entry = {
        "id": uuid.uuid4().hex[:12],
        "ad_set": ad_set,
        "scaled_date": sdate,
        "bid_before": req.bid_before,
        "bid_after": req.bid_after,
        "cpa_before": req.cpa_before,
        "note": req.note,
        "created": datetime.now(timezone.utc).isoformat(),
    }
    async with _scale_log_lock:
        entries = _scale_log_load()
        # če isti ad-set že ima AKTIVEN cooldown, ga posodobi namesto podvajanja
        replaced = False
        for i, e in enumerate(entries):
            if e.get("ad_set") == ad_set:
                active, _, _ = _scale_cooldown_active(e.get("scaled_date", ""))
                if active:
                    entries[i] = entry
                    replaced = True
                    break
        if not replaced:
            entries.append(entry)
        _scale_log_save(entries)
    return JSONResponse({"ok": True, "id": entry["id"]})


@app.delete("/scale-log/{entry_id}")
async def scale_log_delete(entry_id: str):
    async with _scale_log_lock:
        entries = _scale_log_load()
        entries = [e for e in entries if e.get("id") != entry_id]
        _scale_log_save(entries)
    return JSONResponse({"ok": True})


@app.post("/scale-log-prune")
async def scale_log_prune():
    """Odstrani zapise, ki so potekli pred >30 dnevi (čiščenje starega dnevnika)."""
    from datetime import timedelta as _td
    cutoff = datetime.now(timezone.utc).date() - _td(days=30)
    async with _scale_log_lock:
        entries = _scale_log_load()
        kept = []
        for e in entries:
            try:
                d = datetime.strptime(e.get("scaled_date", "")[:10], "%Y-%m-%d").date()
                if d >= cutoff:
                    kept.append(e)
            except Exception:
                kept.append(e)
        _scale_log_save(kept)
    return JSONResponse({"ok": True, "kept": len(kept)})
